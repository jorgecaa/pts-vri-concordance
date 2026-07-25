#!/usr/bin/env python3
"""
validador_an_peyyala — barrido de las colas *peyyāla* de **A ii–A v**.

Los cuatro alineadores numerados (`validador_an2_catukka.py`, `_an3`, `_an4`, `_an5`) cierran el
cuerpo de cada volumen y dejan fuera, **por diseño**, dos clases de fila:

- las de los tramos donde una edición elide lo que la otra numera (`an_peyyala`, quinto régimen);
- las de **rango** (`4.277-279`), que ningún alineador miraba.

Aquí se barren las dos. El método no es el del cuerpo del volumen —allí hay un marcador por sutta y
el cotejo de texto decide— sino el de la **clase peyyāla de SN** (`STATUS.md`): con las dos
ediciones elidiendo, el REJECT del LLM es un falso negativo y lo que se puede exigir es una
**prueba mecánica**:

1. **Concordancia exacta de clave**: el `Sutta #` del Excel es el paranum del VRI (comprobado en
   las 1099 filas ya CONFIRMADO de A ii–A v, sin excepción), y la aritmética del rango cuadra con
   el `<p n="a-b">` del CST.
2. **Término distintivo LITERAL y EN ORDEN** en las dos ediciones: cada fila de la serie encuentra
   su término en el texto de PTS y en el del CST, y las posiciones crecen. Una sola fila fuera de
   sitio invalida la serie entera.
3. **Página**: el locus donde cae el término en PTS está en la página que declara el Excel (±1).

Lo que esta prueba SÍ detecta —y el LLM no podía— es la fila mal nombrada: en el *rāga-peyyāla* del
Catukka la serie de dieciséis kilesa falla en el último término porque el Excel escribe
`Macchariya` donde el impreso y el CST leen `pamādassa`, y `Pamāda` donde los dos leen `madassa`.
Las dieciséis casan en cuanto se corrigen los dos nombres (ver `--nombres`).

Las filas cuyo tramo NO está elidido (el *Sāmañña-vagga* de A v imprime cuatro suttas enteros, el
`7.85 Bhikkhu` de A iv otro) salen por la ruta normal: par PTS↔CST y **validador** (Gemini).

Uso: python3 validador_an_peyyala.py [--dry] [--live] [--vol ii,iii,iv,v]
"""
import json
import os
import re
import sqlite3
import sys
from collections import Counter, defaultdict

import an_peyyala as ap
import sutta_hash as sh
from openpyxl import load_workbook
from validador import validate_pair

DB = 'src/data/tipitaka.sqlite'
XLSX = 'PTS_Reference_Complete_Canon.xlsx'
OUT = 'validador_an_peyyala.json'

# Tramos, uno por nipāta. Cada uno declara su régimen de separación en el impreso: la columna
# «corte» documenta qué separa los suttas de PTS ahí, que NO es lo mismo en los cuatro volúmenes.
REGIONES = [
    # vol  nipāta        nº  stem       book  páginas     paranums     corte en el impreso
    ('ii',  'CATUKKA',   '4',  's0402m3', 18, (256, 257), (277, 783), 'un solo sutta numerado (271.)'),
    ('iii', 'PAÑCAKA',   '5',  's0403m1', 19, (271, 278), (251, 1151), 'ornamento U+E0E0'),
    ('iii', 'CHAKKA',    '6',  's0403m2', 19, (451, 452), (120, 649), 'sin corte: el peyyāla va seguido'),
    ('iv',  'SATTAKA',   '7',  's0403m3', 20, (144, 148), (85, 1124), 'raya de guiones'),
    ('iv',  'AṬṬHAKA',   '8',  's0404m1', 20, (347, 349), (91, 627),  'raya de guiones'),
    ('iv',  'NAVAKA',    '9',  's0404m2', 20, (460, 465), (71, 432),  'marcador de RANGO romano + uddāna'),
    ('v',   'DASAKA',    '10', 's0404m3', 21, (303, 310), (221, 746), 'numeral romano (CCX…) + raya'),
    ('v',   'EKĀDASAKA', '11', 's0404m4', 21, (359, 361), (22, 1151), 'un bloque para toda la serie'),
]

# ⚠️ Dos nombres del Excel que la prueba de la serie desmiente (A ii, *rāga-peyyāla* del Catukka).
# El impreso (A ii 257) y el CST (`s0402m3` n=304-783) enumeran los dieciséis kilesa en el mismo
# orden y los dos terminan `…mānassa atimānassa **madassa pamādassa**`. El Excel se saltó `mada` y
# repitió `macchariya` al final, de modo que sus dos últimas filas van corridas una posición.
# La aritmética confirma cuál es cuál: 30 paranums por kilesa desde 304 ⇒ 724-753 = el 15º = mada.
# El mismo error está en A iii, con la misma forma: el Pañcaka reparte 50 paranums por kilesa
# desde 353, así que `1053-1102` es el 15º (= mada) y `1103-1152` el 16º (= pamāda).
NOMBRES_CORREGIDOS = {
    ('ii', '4.724-753'): ('Pamāda peyyāla', 'Mada peyyāla'),
    ('ii', '4.754-783'): ('Macchariya peyyāla', 'Pamāda peyyāla'),
    ('iii', '5.1053-1102'): ('Pamāda peyyāla', 'Mada peyyāla'),
    ('iii', '5.1103-1152'): ('Macchariya peyyāla', 'Pamāda peyyāla'),
}

MIN_PROSA = 60          # tokens: por debajo de esto el texto está elidido y no hay qué cotejar


def excel_pendientes(solo_pendientes=True):
    """Filas de A ii–A v, en orden de lectura, con su paranum de arranque.

    ⚠️ Con `solo_pendientes=False` devuelve **todas**, y eso importa: la serie monótona necesita
    las filas ya CONFIRMADO como **anclas**. Corriéndola sólo sobre las pendientes, cada una queda
    suelta y la monotonía no puede apoyarse en nada — en el Pañcaka de A iii, `5.270` y `5.271` se
    situaban con el tramo entero y se perdían al quedarse solas. El análisis se hace sobre todas y
    se informa de las pendientes.
    """
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ci['Nikaya']] != 'AN':
            continue
        if solo_pendientes and row[ci['Estado']] == 'CONFIRMADO':
            continue
        vol = str(row[ci['PTS Roman']] or '').strip()
        if vol not in ('ii', 'iii', 'iv', 'v'):
            continue
        num = str(row[ci['Sutta #']])
        m = re.match(r'(\d+)\.(\d+)(?:-(\d+))?$', num)
        if not m:
            continue
        nombre = str(row[ci['Sutta Name']] or '')
        corr = NOMBRES_CORREGIDOS.get((vol, num))
        out.append({'vol': vol, 'num': num, 'nip': m.group(1),
                    'lo': int(m.group(2)), 'hi': int(m.group(3) or m.group(2)),
                    'name': nombre, 'name_ok': corr[1] if corr else nombre,
                    'page': row[ci['PTS Page']], 'legacy': str(row[ci['Validation']] or ''),
                    'pendiente': row[ci['Estado']] != 'CONFIRMADO'})
    return out


def region_de(fila):
    """Tramo al que pertenece la fila. El **prefijo de nipāta manda**: los paranums se repiten
    entre nipātas (el 140 existe en el Pañcaka y en el Chakka) y sin él `5.187` acababa en la
    región del Chakka."""
    for vol, nip, pref, stem, book, pgs, paras, _c in REGIONES:
        if fila['vol'] == vol and fila['nip'] == pref and paras[0] <= fila['lo'] <= paras[1]:
            return (vol, nip, stem, book, pgs, paras)
    return None


def texto_cst(unidades):
    """Concatena los textos DISTINTOS de las unidades (un grupo aporta el suyo una sola vez)."""
    partes, visto, tramos = [], set(), []
    off = 0
    for u in unidades:
        if u['grupo'] in visto:
            continue
        visto.add(u['grupo'])
        t = ap.fold(u['texto'])
        tramos.append((off, off + len(t), u['grupo']))
        partes.append(t)
        off += len(t) + 1
    return ' '.join(partes), tramos


def analiza(conn, region, filas):
    """Prueba mecánica de una región. Devuelve la lista de filas anotadas."""
    vol, nip, stem, book, pgs, paras = region
    T = ap.TextoPTS(conn, book, *pgs)
    unidades = ap.cst_unidades(stem, paras[0], paras[1])
    porpn = {u['pn']: u for u in unidades}
    cst_plano, _tramos = texto_cst([porpn[p] for p in sorted(porpn)])

    filas = sorted(filas, key=lambda f: f['lo'])
    series = [(f['num'], ap.sondas(f['name_ok'])) for f in filas]
    # sondas que aparecen en más de una fila del tramo: no distinguen y no pueden firmar sola
    _cuenta = Counter(s for _n, ss in series for s in set(ss))
    compartidas = {s for s, c in _cuenta.items() if c > 1}
    pos_pts, sin_pts = ap.serie_monotona(T.plano, series)
    pos_cst, sin_cst = ap.serie_monotona(cst_plano, series)

    for k, f in enumerate(filas):
        u = porpn.get(f['lo'])
        f['cst'] = u
        f['vri'] = f'{stem}:{f["lo"]}' if f['lo'] == f['hi'] else f'{stem}:{f["lo"]}-{f["hi"]}'
        f['pos_pts'], f['sonda_pts'] = (pos_pts or {}).get(f['num'], (None, None))
        f['pos_cst'], f['sonda_cst'] = (pos_cst or {}).get(f['num'], (None, None))
        f['locus'] = T.locus(f['pos_pts']) if f['pos_pts'] is not None else None
        # RUTA DEL APARATO: si el cuerpo no lo imprime, puede estar en las notas al pie de su
        # página, que son parte de la edición. En A iii 276 el cuerpo da `sāmaṇerā` y `upāsikā` y
        # el aparato `1 M. S. sāmaṇero sāmaṇerī` / `2 M. T. M6. M7 upāsako upāsikā`: los cuatro
        # suttas que el CST numera aparte. No entra en la serie monótona —una nota no tiene
        # posición en el texto corrido— pero atestigua la lectura en la página que toca.
        # ⚠️ La sonda del aparato tiene que ser EXCLUSIVA de la fila: si la comparten otras filas
        # del tramo no prueba nada. En el Navaka, `sammappadhāna` e `iddhipāda` son de las nueve
        # hermanas a la vez, y admitirlas habría firmado nueve filas con la misma palabra.
        f['aparato'] = None
        if f['pos_pts'] is None and isinstance(f['page'], int):
            for pg in (f['page'], f['page'] - 1, f['page'] + 1):
                ap_txt = T.aparato.get(pg, '')
                for sonda in ap.sondas(f['name_ok']):
                    if len(sonda) >= 6 and sonda not in compartidas and \
                            ap.buscar_palabra(ap_txt, sonda):
                        f['aparato'] = (pg, sonda)
                        break
                if f['aparato']:
                    break
        f['pagina_ok'] = (f['locus'] is not None and isinstance(f['page'], int)
                          and abs(f['locus'][0] - f['page']) <= 1)
        f['elidido'] = u is None or u['n'] > 1 or len(sh.tokens(u['texto'])) < MIN_PROSA
    # ventana de PTS: de su ancla a la de la fila siguiente (o al final del tramo)
    for k, f in enumerate(filas):
        if f['pos_pts'] is None:
            f['pts_txt'] = None
            continue
        sig = next((g['pos_pts'] for g in filas[k + 1:] if g['pos_pts'] is not None), None)
        a = T.idx[f['pos_pts']]
        b = T.idx[sig] if sig is not None else len(T.crudo)
        f['pts_txt'] = T.crudo[a:b]
    return filas, (sin_pts, sin_cst), T, porpn


def main():
    dry = '--dry' in sys.argv
    live = '--live' in sys.argv
    vols = (sys.argv[sys.argv.index('--vol') + 1].split(',')
            if '--vol' in sys.argv else ['ii', 'iii', 'iv', 'v'])

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    todas = [f for f in excel_pendientes(solo_pendientes=False) if f['vol'] in vols]
    porregion = defaultdict(list)
    sueltas = []
    for f in todas:
        r = region_de(f)
        (porregion[r].append(f) if r else sueltas.append(f))

    print('=' * 96)
    print('AN — colas peyyāla de A ii–A v: prueba mecánica (término literal y en orden)')
    print('=' * 96)
    tareas, resumen = [], Counter()
    for region in REGIONES:
        clave = (region[0], region[1], region[3], region[4], region[5], region[6])
        filas = porregion.get(clave)
        if not filas:
            continue
        filas, fallos, T, porpn = analiza(conn, clave, filas)
        filas = [f for f in filas if f['pendiente']]
        fallos = ([n for n in fallos[0] if n in {f['num'] for f in filas}],
                  [n for n in fallos[1] if n in {f['num'] for f in filas}])
        if not filas:
            continue
        vol, nip = region[0], region[1]
        ok_pts = sum(1 for f in filas if f['pos_pts'] is not None)
        ok_cst = sum(1 for f in filas if f['pos_cst'] is not None)
        ok_pg = sum(1 for f in filas if f['pagina_ok'])
        print(f'\n── A {vol} · {nip} · pp.{region[5][0]}-{region[5][1]} · '
              f'paranums {region[6][0]}-{region[6][1]} · {region[7]}')
        print(f'   filas pendientes: {len(filas)} | término hallado en PTS: {ok_pts}/{len(filas)}'
              f' | en CST: {ok_cst}/{len(filas)} | página ±1: {ok_pg}/{len(filas)}')
        if fallos[0]:
            print(f'   ⚠ sin situar en PTS ({len(fallos[0])}): {", ".join(fallos[0][:12])}')
        if fallos[1]:
            print(f'   ⚠ sin situar en CST ({len(fallos[1])}): {", ".join(fallos[1][:12])}')
        for f in filas:
            marca = ('prosa' if not f['elidido'] else 'elidido')
            print(f"   {f['num']:>12} {f['vri']:>22} p{str(f['page']):>4} "
                  f"{'PTS✓' if f['pos_pts'] is not None else 'PTS·'} "
                  f"{'CST✓' if f['pos_cst'] is not None else 'CST·'} "
                  f"{'pág✓' if f['pagina_ok'] else 'pág·'} {marca:<7} "
                  f"{str(f['locus']):>12} «{f['sonda_pts'] or '—'}»  {f['name_ok'][:38]}")
            resumen[(f['pos_pts'] is not None, f['pos_cst'] is not None, f['pagina_ok'],
                     f['elidido'])] += 1
            tareas.append((clave, f))
    if sueltas:
        print(f'\n── fuera de los tramos peyyāla: {len(sueltas)} filas '
              f'(régimen numerado; van por `an_rangos.py` o por arbitraje)')
        for f in sueltas:
            print(f"   A {f['vol']:>3} {f['num']:>12} p{str(f['page']):>4} {f['legacy']:<11} {f['name'][:44]}")
    print('\nresumen (PTS, CST, página, elidido) → nº de filas')
    for k, v in sorted(resumen.items(), key=str):
        print('  ', k, v)
    if dry:
        print('\n(dry-run: nada validado ni escrito)')
        return tareas
    return tareas


if __name__ == '__main__':
    main()
