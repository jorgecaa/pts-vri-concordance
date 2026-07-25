#!/usr/bin/env python3
"""Aligner + validador de **SN IV** (S iv): Excel(DPR) → massive.tsv → cst_paranum → XML VRI → CST.

Mismo pipeline que S iii y S v. El lado CST es una **concordancia exacta** (el `dpr_code` de
`massive.tsv` es literalmente el `Sutta #` del Excel), así que CONFIRMADO = concordancia ∧ Gemini.
El lado PTS lo da la **asignación global, inyectiva y monótona** (`align_rows.assign`) de las filas
del Excel a los marcadores de la BD (libro 15), nunca una búsqueda fila a fila.

⚠️ **El `Sutta #` del Excel NO es la numeración del CST en SN 35.** Es la cuenta **reducida de
Feer** (207), mientras `massive.tsv` usa la del CST (247 si se desglosa), así que emparejar por esa
clave mandaba 123 filas al sutta equivocado. Los desfases (+17 en p29; +27/+37/+47 en pp. 151-155)
caen exactamente donde Feer explica que comprimió: los 19 suttas de `Jātidhammādi…` +
`Aniccādi…` que el Excel lleva como 2 filas, y el **`satthi peyyala`**, donde él mismo «reduce» 60
suttantas a 20. Por eso el lado CST **no se resuelve por clave sino por POSICIÓN**: los bloques de
subhead del XML y las filas del Excel se corresponden **1:1 en los 10 saṃyuttas** (344 = 344), y
la comprobación por nombre da 344/344. `massive.tsv` se usa sólo para la página de cotejo.

**Verdad-terreno estructural** (`samyutta-vol-IV-info.txt`, front matter de Feer): 10 saṃyuttas
XXXV-XLIV con su página de arranque, 33 vaggos y 391 suttas. El contraste sale limpio: las 10
cabeceras de libro caen en la página exacta que declara Feer, 9 de los 10 recuentos son exactos y
la numeración de marcadores es contigua 1..N en todos. El único desajuste, **SN 39 Sāmaṇḍaka**
(PTS imprime 2, Feer cuenta 16), lo explica el propio Feer: *«All the MSS give only the beginning
and the end of the Samandaka»* — se imprimen el nº1 y el nº16 y se elide todo lo de en medio.

⚠️ **La numeración DPR no es la de PTS.** En SN 35 el DPR llega a 227 y PTS imprime 207; el Excel
tiene 201 filas. Por eso el emparejamiento va por **nombre + contenido + posición**, con el nº
corrido sólo como refuerzo, y se exige la comprobación cruzada independiente (nombre del marcador
≡ título CST, página del marcador ≡ `cst_p_page`) antes de gastar API.

Uso: python3 validador_sn4.py [--dry] [--all] [--n N] [--only 35.1,43.2]
"""
import json
import os
import re
import sqlite3
import sys
import csv
import xml.etree.ElementTree as ET
from collections import Counter

import sutta_hash as sh
from openpyxl import load_workbook

from align_rows import assign as ar_assign
from massive_reader import build_massive as _mr_build
from sn4_markers import find_markers_sn4
from validador import validate_pair
from validador_sn1 import _pts_page
import sn4_names

VRI = '/tmp/tipitaka-xml/romn/s0304m.mul.xml'
OUT = 'validador_sn4.json'
DB = 'src/data/tipitaka.sqlite'
SN4_BOOK = 15
XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# front matter de Feer: (saṃyutta, nº romano, título, vaggos, suttas, página de arranque)
FEER = [(35, 'XXXV', 'Saḷāyatana', 19, 207, 1), (36, 'XXXVI', 'Vedanā', 3, 29, 204),
        (37, 'XXXVII', 'Mātugāma', 3, 34, 238), (38, 'XXXVIII', 'Jambukhādaka', 1, 16, 251),
        (39, 'XXXIX', 'Sāmaṇḍaka', 1, 16, 261), (40, 'XL', 'Moggallāna', 1, 11, 262),
        (41, 'XLI', 'Citta', 1, 10, 281), (42, 'XLII', 'Gāmaṇi', 1, 13, 305),
        (43, 'XLIII', 'Asaṅkhata', 2, 44, 359), (44, 'XLIV', 'Avyākata', 1, 11, 374)]

_BOOK_RE = re.compile(r'^\s*BOOK\s+[IVXL]+\b.*?-?SA[ṂM]YUTTA', re.I)
_FOLD = str.maketrans({'ā': 'a', 'ī': 'i', 'ū': 'u', 'ṅ': 'n', 'ñ': 'n', 'ṇ': 'n',
                       'ṭ': 't', 'ḍ': 'd', 'ḷ': 'l', 'ṃ': 'm', 'ṁ': 'm'})


def _stem(t):
    t = re.sub(r'[^a-z]', '', re.sub(r'^[\d\s.,()-]*', '', (t or '').lower()).translate(_FOLD))
    t = re.sub(r'(suttantam|suttam|suttani|vaggo).*$', '', t)
    return re.sub(r'[aiueom]+$', '', re.sub(r'(.)\1+', r'\1', t))


def _name_score(a, b):
    x, y = _stem(a), _stem(b)
    if not x or not y or min(len(x), len(y)) < 3:
        return 0
    if x == y:
        return 100
    if x.startswith(y) or y.startswith(x):
        return 60
    if min(len(x), len(y)) >= 4 and (x in y or y in x):
        return 40
    return 0


# ── XML VRI → paranum → sutta CST ───────────────────────────────────────────
def build_vri_index(path=VRI):
    """`cst_paranum` → `{'title', 'text', 'block'}`, indexado **por párrafo**, no por bloque.

    Un subhead del CST puede cubrir un grupo («1-10. Jātidhammādisuttadasakaṃ») y llevar debajo el
    párrafo propio de cada miembro (`n=33` jātidhamma, `n=34` jarādhamma, …). Devolver el bloque
    entero para todos hacía que los diez se cotejaran contra el mismo texto — el defecto que ya
    costó caro en S iii. Indexar por párrafo lo arregla y **no fragmenta nada**: los 326 bloques
    individuales de este volumen tienen exactamente un paranum, así que para ellos párrafo y bloque
    coinciden. Los párrafos sin `n` se agregan al último paranum visto.
    """
    root = ET.parse(path).getroot()
    idx, title, last = {}, None, None
    for el in root.find('.//body').iter():
        if el.tag != 'p':
            continue
        rend, txt = el.get('rend'), ''.join(el.itertext())
        if rend == 'subhead':
            title, last = ' '.join(txt.split()), None
        elif rend == 'bodytext' and title is not None:
            n = el.get('n')
            m = re.match(r'(\d+)(?:-(\d+))?$', n or '')
            if m:
                keys = list(range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1))
                for k in keys:
                    idx.setdefault(k, {'title': title, 'text': ''})
                    idx[k]['text'] += ' ' + txt
                last = keys
            elif last:                              # párrafo sin `n`: sigue al anterior
                for k in last:
                    idx[k]['text'] += ' ' + txt
    return idx


def build_cst_blocks(path=VRI):
    """Bloques de subhead del XML VRI, **en orden de documento**, agrupados por saṃyutta.

    Ésta es la unidad correcta del lado CST: un bloque = un sutta o un grupo elidido, y su serie
    casa 1:1 con las filas del Excel (344 = 344, exacto en los 10 saṃyuttas). Resolver por
    `cst_paranum` a través del `Sutta #` NO vale en SN 35, donde las dos numeraciones divergen.
    """
    root = ET.parse(path).getroot()
    blocks, cur = [], None
    for el in root.find('.//body').iter():
        if el.tag != 'p':
            continue
        rend, txt = el.get('rend'), ''.join(el.itertext())
        if rend == 'subhead':
            cur = {'title': ' '.join(txt.split()), 'pn': [], 'text': ''}
            blocks.append(cur)
        elif rend == 'bodytext' and cur is not None:
            m = re.match(r'(\d+)(?:-(\d+))?$', el.get('n') or '')
            if m:
                cur['pn'] += list(range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1))
            cur['text'] += ' ' + txt
    para2sam, para2page = {}, {}
    for r in csv.DictReader(open('massive.tsv'), delimiter='\t'):
        if not (r['cst_code'] or '').startswith('sn4.'):
            continue
        pm = re.match(r'(\d+)', r['cst_paranum'] or '')
        sm = re.match(r'SN(\d+)\.', r['dpr_code'] or '')
        if pm and sm:
            para2sam.setdefault(int(pm.group(1)), int(sm.group(1)))
            para2page.setdefault(int(pm.group(1)), _pts_page(r['cst_p_page']))
    bysam = {}
    for b in blocks:
        b['sam'] = next((para2sam[p] for p in b['pn'] if p in para2sam), None)
        b['page'] = next((para2page[p] for p in b['pn'] if para2page.get(p)), None)
        bysam.setdefault(b['sam'], []).append(b)
    return bysam


def cst_for(entries, bysam):
    """`Sutta # → bloque CST` por **posición** dentro del saṃyutta (la serie es 1:1)."""
    out = {}
    for sam in {e['sam'] for e in entries}:
        rows = sorted([e for e in entries if e['sam'] == sam], key=lambda e: e['inner'])
        for e, b in zip(rows, bysam.get(sam, [])):
            out[e['num']] = b
    return out


# ── lado PTS: marcadores de la BD, segmentados por saṃyutta ─────────────────
def build_pts_suttas(cur):
    """Los marcadores de S iv con su `sam`, su texto acotado y su posición de lectura.

    La segmentación **no puede ir por página**: SN 39 y SN 40 comparten la página 262 (la cabecera
    de Moggallāna está en su línea 28), así que el `16 Dukkaram` del Sāmaṇḍaka caería en el
    saṃyutta siguiente. El corte lo dan las **cabeceras de libro**, que aparecen las 10 en la
    página exacta que declara Feer.
    """
    pages = {r['page_no']: (r['unitext'] or '').split('\n')
             for r in cur.execute('SELECT page_no,unitext FROM pages '
                                  'WHERE edition="mula" AND book_no=?', (SN4_BOOK,))}
    heads, flat, rows = [], [], []
    for pg in sorted(pages):
        for ln, line in enumerate(pages[pg], start=1):
            flat.append((pg, ln, line))
            if _BOOK_RE.match(line.replace('⁣', '')):
                heads.append((pg, ln))
        for ln, nums, names, _tag in find_markers_sn4('\n'.join(pages[pg])):
            rows.append((pg, ln, nums, names))

    def sam_of(pg, ln):
        s = FEER[0][0]
        for k, (hp, hl) in enumerate(heads):
            if (pg, ln) >= (hp, hl):
                s = FEER[k][0] if k < len(FEER) else s
        return s

    # CONTINUIDAD DE LA SECUENCIA. Una línea con número DESNUDO sólo es marcador si su número
    # continúa la serie del saṃyutta. Así se cae el falso «13 Pittaṃ semhaṃ ca vāto ca» (S iv 231)
    # —una gāthā que el alineador tomaba por sutta y que desplazaba 36.22 y siguientes— sin perder
    # el «5 Daṭṭhabbena» (S iv 207), que es un marcador legítimo sin posición en un saṃyutta que
    # normalmente sí la imprime. Exigir la posición por régimen se llevaba ese por delante: el
    # criterio no es la forma sino que el número encaje donde toca.
    rows.sort(key=lambda r: (r[0], r[1]))
    keep, expect = [], {}
    for r in rows:
        sam_, nums, names = sam_of(r[0], r[1]), r[2], r[3]
        has_pos = any(p is not None for p, _n, _rw in names)
        nxt = expect.get(sam_, 1)
        # `>=`, no `==`: una elisión salta HACIA DELANTE (S iv 262, «16 Dukkaram» tras el nº1 —
        # de los MSS de Sāmaṇḍaka sólo se imprimen principio y final), mientras que un nº de
        # párrafo o de verso reinicia dentro del sutta y por tanto queda POR DEBAJO del esperado.
        if not nums or has_pos or nums[0] >= nxt:
            keep.append(r)
            if nums:
                expect[sam_] = nums[-1] + 1
    rows = keep

    pos = {(pg, ln): i for i, (pg, ln, _t) in enumerate(flat)}
    ordered = sorted(rows, key=lambda r: pos[(r[0], r[1])])
    recs, last_num = [], 0
    for k, (pg, ln, nums, names) in enumerate(ordered):
        # régimen C (saṭṭhi-peyyāla): la línea trae sólo la posición y hereda el nº corrido
        if not nums:
            nums = [last_num]
        else:
            last_num = nums[-1]
        i0 = pos[(pg, ln)]
        i1 = pos[(ordered[k + 1][0], ordered[k + 1][1])] if k + 1 < len(ordered) else len(flat)
        text = ' '.join(sh.tokens(' '.join(t for _p, _l, t in flat[i0:i1]))[:350])
        nm = next((n for _p, n, _r in names if n), '')
        raw = next((r for _p, n, r in names if n), '')
        for n in nums:
            recs.append({'sam': sam_of(pg, ln), 'num': n, 'page': pg, 'line': ln,
                         'name': nm, 'raw': raw, 'ord': k, 'text': text,
                         'from_range': len(nums) > 1})
    return recs


def excel_entries():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['Nikaya']] or '') != 'SN' or str(row[ci['PTS Roman']] or '').strip().lower() != 'iv':
            continue
        num = str(row[ci['Sutta #']])
        m = re.match(r'(\d+)\.(\d+)$', num)
        if not m:
            continue
        name = str(row[ci['Sutta Name']] or '')
        out.append({'num': num, 'sam': int(m.group(1)), 'inner': int(m.group(2)),
                    'name': re.sub(r'\(SN[^)]*\)', '', name).strip(),
                    'page': row[ci['PTS Page']], 'legacy': str(row[ci['Validation']] or '')})
    return out


def assign_volume(entries, recs, cstmap):
    """Asignación **inyectiva y monótona** de las filas a los marcadores PTS, por saṃyutta.

    Las dos secuencias van ordenadas (filas en orden canónico, marcadores en orden de lectura), así
    que la asignación correcta es un alineamiento monótono con capacidad, no una búsqueda por fila:
    resolver cada fila por su cuenta permite que dos caigan en el mismo marcador y que una quede
    validada contra el locus de su vecina, sin que el validador pueda notarlo.
    """
    by_sam, marks_by_sam = {}, {}
    for e in entries:
        by_sam.setdefault(e['sam'], []).append(e)
    caps = Counter((q['sam'], q['page'], q['line']) for q in recs)
    for q in recs:
        marks_by_sam.setdefault(q['sam'], {}).setdefault((q['page'], q['line']), q)

    out, cov = {}, {}

    def content(e, q):
        """Solapamiento léxico CST↔marcador: la evidencia que no depende del nombre."""
        b = cstmap.get(e['num'])
        if not b:
            return 0.0
        key = (id(b), q['ord'])
        if key not in cov:
            a_ = set(sh.tokens(b['text'])[:350])
            cov[key] = len(a_ & set(q['text'].split())) / max(1, len(a_))
        return cov[key]

    def measured_caps(marks, rows):
        """Capacidad de cada marcador **medida en el texto**, no adivinada por el nombre.

        Feer junta a veces dos suttas del CST bajo un solo marcador, y el título no siempre lo
        dice: «Suddhikaṃ nirāmisam» y «Pubbeñāṇam» nombran los dos, pero «Agayha» (S iv 126)
        imprime los dos Rūpārāma sin nombrar ninguno. Y al revés, «Devadahakhaṇo» **parece** doble
        y no lo es: el Khaṇa del CST es el marcador siguiente. Contar raíces en el nombre acertaba
        en tres casos y fallaba en dos.

        Lo que sí decide es el contenido: la capacidad de un marcador es el número de filas que lo
        eligen a ÉL como el de mayor solapamiento léxico. Es autoconsistente y no depende de cómo
        titule Feer.
        """
        cap = Counter()
        for e in rows:
            cs = []
            for q in marks:
                base = e['page'] if isinstance(e['page'], int) else q['page']
                if abs(q['page'] - base) <= 2:
                    cs.append((content(e, q), (q['page'], q['line'])))
            if not cs:
                continue
            cs.sort(reverse=True)
            # Cuenta si el marcador es el MÁXIMO de la fila **y** la cobertura es alta. El umbral
            # alto es lo que filtra el ruido: este corpus es muy formulario (mediana 0.70) y un
            # argmax por 0.04 sobre coberturas mediocres no es evidencia — sin él, el nº27 de
            # SN 36 se llevaba dos filas y arrastraba 36.29-36.31 un puesto. Un margen sobre el
            # segundo mejor no sirve: dejaba fuera la PRIMERA mitad de los marcadores dobles
            # (36.24 de «Pubbeñāṇam»), y entonces la segunda se quedaba sin sitio.
            if cs[0][0] >= 0.60:
                cap[cs[0][1]] += 1
        return cap

    for sam, rows in by_sam.items():
        rows.sort(key=lambda e: e['inner'])
        mk = marks_by_sam.get(sam, {})
        marks = [mk[k] for k in sorted(mk, key=lambda k: mk[k]['ord'])]
        if not marks:
            continue

        def score(i, j, rows=rows, marks=marks):
            e, q = rows[i], marks[j]
            b = cstmap.get(e['num'])
            base = e['page'] if isinstance(e['page'], int) else (b['page'] if b else None)
            if not isinstance(base, int):
                base = q['page']
            d = abs(q['page'] - base)
            if d > 2:
                return None
            s = _name_score(e['name'], q['name']) + (30 if d == 0 else 15 if d == 1 else 0)
            s += 60 * content(e, q)
            if b:
                # el nombre va con el analizador PROPIO de S iv: Feer usa el término en
                # instrumental y el ordinal como dígito volado, y la regla genérica de raíces
                # falla en 96 filas bien alineadas
                s += sn4_names.score(b['title'], q['name'])
                ok = sn4_names.ordinal_ok(b['title'], q['name'])
                if ok is True:
                    s += 80                          # el ordinal desempata dentro de la serie
                elif ok is False:
                    s -= 120                         # «Sambodhena1» no puede ser «Dutiya…»
                if b['page'] is not None and abs(q['page'] - b['page']) <= 1:
                    s += 10
            return s

        mc = measured_caps(marks, rows)
        cc = [max(caps[(sam, q['page'], q['line'])], min(3, mc.get((q['page'], q['line']), 1)))
              for q in marks]
        idx = ar_assign(len(rows), len(marks), score, lambda j, cc=cc: cc[j],
                        skip_penalty=1000.0)
        for e, j in zip(rows, idx):
            out[e['num']] = marks[j] if j is not None else None
    return out


def measured_caps_for(entries, recs, cstmap):
    """Las capacidades medidas, expuestas para `audit_injectivity` (mismo camino que el driver)."""
    out = {}
    marks_by_sam = {}
    for q in recs:
        marks_by_sam.setdefault(q['sam'], {}).setdefault((q['page'], q['line']), q)
    for sam in {e['sam'] for e in entries}:
        rows = sorted([e for e in entries if e['sam'] == sam], key=lambda e: e['inner'])
        mk = marks_by_sam.get(sam, {})
        marks = [mk[k] for k in sorted(mk, key=lambda k: mk[k]['ord'])]
        cov = {}

        def content(e, q):
            b = cstmap.get(e['num'])
            if not b:
                return 0.0
            k = (e['num'], q['ord'])
            if k not in cov:
                a_ = set(sh.tokens(b['text'])[:350])
                cov[k] = len(a_ & set(q['text'].split())) / max(1, len(a_))
            return cov[k]

        for e in rows:
            cs = sorted(((content(e, q), (q['page'], q['line'])) for q in marks
                         if abs(q['page'] - (e['page'] if isinstance(e['page'], int) else q['page'])) <= 2),
                        reverse=True)
            if cs and cs[0][0] >= 0.60:
                k = (sam,) + cs[0][1]
                out[k] = min(3, out.get(k, 0) + 1)
    return out


def main():
    dry = '--dry' in sys.argv
    do_all = '--all' in sys.argv
    n = int(sys.argv[sys.argv.index('--n') + 1]) if '--n' in sys.argv else 30
    only = set(sys.argv[sys.argv.index('--only') + 1].split(',')) if '--only' in sys.argv else None

    bysam = build_cst_blocks()
    entries = excel_entries()
    cstmap = cst_for(entries, bysam)
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
    recs = build_pts_suttas(conn.cursor())
    assigned = assign_volume(entries, recs, cstmap)

    print('=' * 92)
    print('SN IV — pipeline VRI + estructura de Feer (samyutta-vol-IV-info.txt)')
    print('=' * 92)
    # el recuento comparable con Feer es el de NÚMEROS CORRIDOS distintos: las líneas del
    # saṭṭhi-peyyāla sin nº propio heredan el del marcador anterior y no son suttas nuevos
    nums = Counter()
    for sam_ in {q['sam'] for q in recs}:
        nums[sam_] = len({q['num'] for q in recs if q['sam'] == sam_})
    nb = sum(len(v) for v in bysam.values())
    print(f'PTS (BD libro 15): {len(recs)} marcadores | Excel «S iv»: {len(entries)} filas '
          f'| bloques CST: {nb}')
    print('correspondencia CST↔Excel por posición: ' +
          ', '.join(f'{s_}:{len(bysam.get(s_, []))}={sum(1 for e in entries if e["sam"] == s_)}'
                    for s_, *_r in FEER))
    bad = [f"{s}:{nums[s]}≠{f}" for s, _r, _t, _v, f, _p in FEER if nums[s] != f]
    print(f'contraste con Feer (suttas por saṃyutta): {10 - len(bad)}/10 exactos'
          + (f' — {", ".join(bad)}' if bad else ''))

    tasks, no_cst, no_pts, name_ok, page_ok, page_n, ord_ok, ord_n = [], [], [], 0, 0, 0, 0, 0
    for e in entries:
        if only is not None and e['num'] not in only:
            continue
        b = cstmap.get(e['num'])
        p = assigned.get(e['num'])
        if not b:
            no_cst.append(e); continue
        if not p:
            no_pts.append(e); continue
        # comprobaciones cruzadas INDEPENDIENTES de la clave de emparejamiento
        if sn4_names.score(b['title'], p['name']) >= 70 or _name_score(e['name'], p['name']):
            name_ok += 1
        oo = sn4_names.ordinal_ok(b['title'], p['name'])
        if oo is not None:
            ord_n += 1
            ord_ok += bool(oo)
        if b['page'] is not None:
            page_n += 1
            page_ok += abs(p['page'] - b['page']) <= 1
        cst = ' '.join(sh.tokens(b['text'])[:350])
        tasks.append((e, p['text'], cst, b['title'], p))

    tot = max(1, len(tasks))
    print(f'\nAlineadas con texto PTS+CST: {len(tasks)}/{len(entries)} '
          f'(sin CST: {len(no_cst)} | sin marcador PTS: {len(no_pts)})')
    print(f'  cruzada — raíz del nombre PTS ≡ título CST/Excel: {name_ok}/{tot} ({100*name_ok/tot:.0f}%)')
    print(f'  cruzada — página del marcador ≡ cst_p_page (±1): {page_ok}/{page_n} '
          f'({100*page_ok/max(1,page_n):.0f}%, sólo filas «exact»)')
    print(f'  cruzada — ordinal del marcador ≡ paṭhama/dutiya del CST: {ord_ok}/{ord_n} '
          f'({100*ord_ok/max(1,ord_n):.0f}%)')
    if no_pts:
        print('  sin marcador:', ', '.join(e['num'] for e in no_pts[:12]))
    if dry:
        print('\n(dry-run, nada validado ni escrito)')
        return

    run = tasks if (do_all or only) else tasks[:n]
    print(f'\nValidando {len(run)} (Gemini en vivo)...', flush=True)
    rows = []
    for k, (e, pts, cst, ctitle, p) in enumerate(run, 1):
        res = validate_pair(pts, cst, e['name'], ctitle, 'SN', concordant=True)
        rows.append({'num': e['num'], 'name': e['name'], 'cst_title': ctitle,
                     'legacy': e['legacy'], 'pts_name': p['name'], 'pts_page': p['page'],
                     'pts_line': p['line'], 'from_range': p['from_range'], **res})
        if k % 25 == 0:
            print(f'  ...{k}/{len(run)}', flush=True)
    prev = json.load(open(OUT)) if os.path.exists(OUT) else []
    merged = {r['num']: r for r in prev}
    merged.update({r['num']: r for r in rows})
    json.dump(list(merged.values()), open(OUT, 'w'), ensure_ascii=False, indent=1)
    print(f"\nEstado: {dict(Counter(r['estado'] for r in rows))} | "
          f"Validation: {dict(Counter(r['validation'] for r in rows))}")
    print(f'Resultados → {OUT} ({len(merged)} en total; nada escrito al Excel).')


if __name__ == '__main__':
    main()
