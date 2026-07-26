#!/usr/bin/env python3
"""
kn_vimanavatthu — el Vimānavatthu (85 filas) y el Petavatthu (51), una fila por vimāna/peta.

Las dos van juntas porque son **la misma obra con el signo cambiado** —el mérito y su reverso— y
comparten régimen editorial: cuatro/dos `div` de vaggas, título de colofón detrás del poema, ningún
rango de verso en el Excel y la biyección completa 85 = 85 y 51 = 51.

Va en fichero propio y no como un parámetro más de `kn_theragatha` (regla (6)) porque **el régimen
de prueba es otro**: aquí el contenido no decide, decide el nombre. Conviene explicar por qué,
porque la medición dice justo lo contrario de lo que uno esperaría.

### Por qué el contenido NO puede decidir

Lo que sigue está medido sobre el Vimānavatthu, que es el caso extremo; el Petavatthu se comporta
igual pero con menos gemelos.

La obra está construida a base de **pares deliberadamente repetidos**: `Paṭhamasuṇisā` /
`Dutiyasuṇisā`, `Paṭhamakaraṇīya` / `Dutiyakaraṇīya`, `Paṭhamasūci` / `Dutiyasūci`,
`Paṭhamanāga` / `Dutiyanāga`… El segundo miembro es el mismo poema con el donante cambiado, y en
tres casos PTS **ni siquiera lo reimprime**: en Vv 86 escribe `anantaraṃ pañcavimānaṃ yathā
kākatarasadāyakavimānaṃ tathā vitthāretabbaṃ` y remite. Medido sobre las 85 filas:

- con la ventana ajustada a la fila, la cobertura propia tiene mediana **0,94** — pero la de un
  sutta **ajeno** (±3, +10) tiene mediana **0,53**, y **106 de 239** pares ajenos pasan de 0,55. Un
  umbral fijo no separa nada: la lengua es formulaica, como en el Udāna.
- probando si el sutta propio es el que **más** puntúa de los 85, gana 75 de 85; y las diez que
  pierden, pierden **contra su propio gemelo** — `6.1.1.12` da 0,95 y su `Dutiyasuṇisā` 0,97.
- y abrir la ventana hacia atrás, que a primera vista arreglaba las ocho filas flojas (0,39 → 0,95),
  es **vacuo**: con esa ventana los vecinos puntúan igual o más (0,97 contra 0,95). Se descartó.

### El Petavatthu tiene además una evidencia mejor: el colofón impreso

PTS cierra cada peta con su nombre y su **ordinal dentro del vagga** —`nandāpetavatthu tatiyaṃ`,
`saṭṭhikūṭasahassapetavatthu soḷasamaṃ`— y eso es un **marcador del impreso**, del mismo tipo que
fija el lado PTS en SN y AN. Se leen **47 de los 51**, y con los vaggas delimitados por sus propios
encabezados (`URAGAVAGGO PAṬHAMO` p14, `III CŪḶAVAGGA` p47, `IV MAHĀVAGGA` p65 — nunca por inferir
dónde reinicia el ordinal, que se descuadra en cuanto falta uno) **la lista impresa coincide con la
del CST vagga por vagga y ordinal por ordinal**: 12, 13, 10 y 16.

Las discrepancias que quedan son legibles una a una y ninguna es de orden:

- el Ubbarivagga trae **dos colofones marcados `tatiyaṃ`** —`mattāpetavatthu` p22 y
  `nandāpetavatthu` p24— y **ninguno `catutthaṃ``**: el CST numera 3 = Mattā y 4 = Nandā, así que lo
  mal leído es el ordinal, no el orden;
- cuatro colofones no se leen en el texto de la BD (`Pañcaputtakhādaka` 1.6, `Kūṭavinicchayika` 3.9,
  `Seṭṭhiputta` 4.15 y el 1.6), y esas filas quedan **acotadas entre los colofones vecinos**;
- PTS abrevia el nombre —`Mātu-` por `Sāriputtattheramātu-`, `Maṭakuṇḍalī` por `Maṭṭhakuṇḍalī`,
  `Sāṇuvāsi` por `Sāṇavāsīthera`, `Amba-` por `Ambavana-`.

⚠️ Y **esto es lo que absolvió a las tres filas de cobertura baja** (`7.2.5` 0,48, `7.4.2` 0,50,
`7.4.4` 0,35), donde el sutta *siguiente* del CST puntuaba más alto que el propio sobre la ventana
de la fila — la firma exacta de un desplazamiento +1. No lo había: `revatīpetavatthu` está impreso
en p85 como **cuarto** del Mahāvagga, justo donde la fila dice. Lo que PTS no reimprime es el
**texto** (`revatīpetavatthu se vimānavatthu no …`, remite al Vimānavatthu, y Serīsaka **es** `Vv
84`), y los otros dos comparten página con su vecino. La cobertura no podía distinguir «la fila está
desplazada» de «PTS remite el texto a otra obra»; el colofón sí.

### Lo que sí decide

1. **El nombre, 85 de 85.** El `subhead` del CST casa con el nombre del Excel en las ochenta y
   cinco, y es la única señal que ve el ordinal `Paṭhama-`/`Dutiya-` que al contenido se le escapa.
2. **La biyección es completa y respeta el orden**: 85 suttas en el CST, 85 filas en el Excel, mismo
   orden, misma partición en los dos `div` (`Itthivimāna` 1-50, `Purisavimāna` 51-85) y serie de
   páginas monótona de la 1 a la 135.
3. **La cobertura, como corroboración** (nunca como discriminante): mediana 0,94, y donde baja está
   explicado — el poema arranca en la página anterior a la que declara la fila, o PTS lo elide.

### La notación (regla (5-ter))

`Vv <nº>` / `Pv <nº>` — el número corrido del vimāna o del peta, que el propio nombre de la fila trae (`Vv 84
Serīsakavimānavatthu`). El Excel no da rango de versos para esta obra (`PTS Alt/Verse` vacío en las
85), así que el rango del CST queda sólo en la `VRI Ref`.

Uso: python3 kn_vimanavatthu.py [--obra vv|pv] [--dry]
"""
import re
import shutil
import sqlite3
import sys
from datetime import datetime
from difflib import SequenceMatcher

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
# obra → (libro de la BD, fichero VRI, sigla, última página del libro, nº de filas)
OBRAS = {'vv': (27, 's0506m', 'Vv', 135, 85),
         'pv': (28, 's0507m', 'Pv', 95, 51)}
BOOK = STEM = SIGLA = ULTIMA = None
COV_COTA = 0.55          # sólo para etiquetar la corroboración, NO es condición para firmar

# Los vaggas del Petavatthu, delimitados por SUS PROPIOS ENCABEZADOS del impreso. No se infieren
# viendo dónde reinicia el ordinal del colofón: faltan cuatro colofones y la inferencia arrastra el
# desfase un vagga entero (le asignaba a `7.4.3 Nandaka` el colofón de `Rathakāra`).
VAGGAS_PV = [(1, 14), (15, 46), (47, 64), (65, 95)]
ORDS = ('pathamam dutiyam tatiyam catutham pancamam chatham satamam athamam navamam dasamam '
        'ekadasamam dvadasamam terasamam cudasamam pandarasamam solasamam').split()


def _limpia(t):
    """`(KN 6.84) Vv 84 Serīsakavimānavatthu` → `Serīsakavimānavatthu`."""
    t = re.sub(r'^\s*\(KN[^)]*\)\s*', '', t or '')
    return re.sub(r'^\s*(Vv|Pv)\s+[\d.]+\s*', '', t).strip()


def raiz(t):
    """Raíz comparable: sin ordinal de lista, sin `-vimānavatthu`, sin desinencia."""
    t = ap.fold(re.sub(r'^[\d.\s]+', '', t or '')).strip()
    t = re.sub(r'\s*(vimanavathu|peta(?:vathu|vathi)|petivathu|vimanam|petam|vathu|gatha)$',
               '', t).strip()
    return re.sub(r'[aiueom]+$', '', t)


def casa_nombre(excel, subhead):
    """¿Mismo vimāna/peta? Se prueban las variantes que el Excel escribe entre paréntesis.

    ⚠️ **El ordinal cuenta.** `raiz` lo conserva —`pathamasunisa` ≠ `dutiyasunisa`— y es lo único
    que distingue a los gemelos, que por texto son indistinguibles.
    """
    va = [raiz(x) for x in re.split(r'[()]', _limpia(excel)) if raiz(x)]
    vb = raiz(subhead)
    if len(vb) < 3:
        return any(a == vb for a in va)
    return any(len(a) >= 3 and (a == vb or a.startswith(vb) or vb.startswith(a)) for a in va)


def suttas_cst():
    """`[(subhead, primer verso, último verso, div, texto)]`, agrupando por `subhead`."""
    out = []
    for x in ap.cst_unidades(STEM):
        if not out or out[-1][0] != x['subhead']:
            out.append([x['subhead'], x['pn'], x['pn'], x['div'], ''])
        out[-1][2] = x['pn']
        out[-1][4] += ' ' + x['texto']
    return [(a, b, c, d, ap.fold(e)) for a, b, c, d, e in out]


def casa_colofon(subhead, colofon):
    """¿El colofón impreso nombra el mismo peta que el `subhead` del CST?

    PTS **abrevia**, y no siempre por el final: se come el primer elemento del compuesto
    (`Sāriputtattheramātu-` → `Mātu-`), simplifica la aspirada (`Maṭṭhakuṇḍalī` → `Maṭakuṇḍalī`) y
    usa otra forma del nombre (`Sāṇavāsīthera` → `Sāṇuvāsi`). Se compara sobre la raíz de los dos,
    sin aspiración, aceptando contención o parecido alto.
    """
    def nucleo(x):
        # sin el ordinal de lista, sin `-peta(vatthu|vatthi)` ni su femenino `-peti(vatthu)`, y sin
        # aspiración. NO se le quita la vocal final: `mātu` abreviado tiene cuatro letras y
        # recortarlo a `māt` lo dejaría por debajo del mínimo comparable.
        x = ap.fold(re.sub(r'^[\d.\s]+', '', x or '')).strip()
        x = re.sub(r'pet[ai](?:vathu|vathi)$|vathu$', '', x)
        # y sin la palabra de oficio con que el CST cierra el compuesto y PTS omite:
        # `sāṇavāsīthera` / `Sāṇuvāsi`, `uttaramātu` / `Uttarā-mātu`
        x = re.sub(r'(thera|theri|sethi)$', '', x)
        return re.sub('h', '', x)

    a, b = nucleo(subhead), nucleo(colofon)
    if len(a) < 4 or len(b) < 4:
        return a == b
    return a in b or b in a or SequenceMatcher(None, a, b).ratio() >= 0.8


def colofones_pv(conn):
    """`{(vagga, ordinal): (página, nombre)}` — el colofón que PTS imprime al cerrar cada peta.

    El vagga sale del rango de páginas del encabezado impreso; el ordinal, del propio colofón.
    """
    pgs = {r['page_no']: (r['unitext'] or '') for r in conn.execute(
        'SELECT page_no,unitext FROM pages WHERE edition="mula" AND book_no=28')}
    num = {o: k + 1 for k, o in enumerate(ORDS)}
    seq = []
    for pg in sorted(pgs):
        v = next((k + 1 for k, (a, b) in enumerate(VAGGAS_PV) if a <= pg <= b), None)
        for ln in pgs[pg].split('\n'):
            m = re.search(rf'([a-z\-]+pet[ai](?:vatthu|vathu))\s*({"|".join(ORDS)})', ap.fold(ln))
            if m and v:
                seq.append((v, num[m.group(2)], pg, m.group(1)))
    # ⚠️ El ordinal LEÍDO se corrige con el orden de impresión. El Ubbarivagga trae dos colofones
    # marcados `tatiyaṃ` —`mattā` p22 y `nandā` p24— y ninguno `catutthaṃ`: la transcripción leyó
    # mal el segundo. Los colofones aparecen en orden, así que un ordinal que no avanza es un
    # ordinal mal leído y vale `anterior + 1`. Los huecos SÍ se respetan (faltan cuatro colofones y
    # sus ordinales no deben reasignarse a nadie), por eso no se renumera de corrido.
    out, prev, vag = {}, 0, 0
    for v, o, pg, nm in seq:
        if v != vag:
            vag, prev = v, 0
        o = max(o, prev + 1)
        out[(v, o)] = (pg, nm)
        prev = o
    return out


def main():
    global BOOK, STEM, SIGLA, ULTIMA
    dry = '--dry' in sys.argv
    obra = sys.argv[sys.argv.index('--obra') + 1] if '--obra' in sys.argv else 'vv'
    BOOK, STEM, SIGLA, ULTIMA, ESPERADAS = OBRAS[obra]
    print(f'── {obra}: libro {BOOK} · {STEM} · «{SIGLA}»')
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    cst = suttas_cst()
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    fs = [{'num': str(r[ci['Sutta #']]), 'name': str(r[ci['Sutta Name']] or ''),
           'page': r[ci['PTS Page']], 'estado': r[ci['Estado']]}
          for r in ws.iter_rows(min_row=2, values_only=True)
          if r[ci['Nikaya']] == 'KN' and str(r[ci['PTS Vol']]) == SIGLA]
    print(f'CST {len(cst)} suttas · Excel {len(fs)} filas')
    if not (len(cst) == len(fs) == ESPERADAS):
        print('⚠ los recuentos no coinciden: no se escribe nada')
        return

    # (1) el nombre, en las 85; (2) la serie de páginas, monótona
    nombres = [casa_nombre(f['name'], cst[k][0]) for k, f in enumerate(fs)]
    pgs = [f['page'] for f in fs]
    mono = all(isinstance(b, int) and isinstance(a, int) and b >= a
               for a, b in zip(pgs, pgs[1:]))
    print(f'  nombre {sum(nombres)}/{len(fs)} · páginas monótonas {mono} '
          f'({pgs[0]}..{pgs[-1]} de {ULTIMA})')
    if not (all(nombres) and mono):
        print('⚠ la biyección por nombre no es completa o la serie retrocede: no se escribe nada')
        for k, f in enumerate(fs):
            if not nombres[k]:
                print(f"   ⚠ «{_limpia(f['name'])}» vs «{cst[k][0]}»")
        return

    # (3) sólo en el Petavatthu: el colofón impreso, que es la evidencia fuerte
    col, acuerdo = {}, [None] * len(fs)
    if obra == 'pv':
        col = colofones_pv(conn)
        for k, f in enumerate(fs):
            v, o = (int(x) for x in f['num'].split('.')[1:])
            if (v, o) not in col:
                continue                       # cuatro no se leen; la fila queda acotada por vecinos
            acuerdo[k] = casa_colofon(cst[k][0], col[(v, o)][1])
        vistos = sum(x is not None for x in acuerdo)
        print(f'  colofón impreso: {vistos}/{len(fs)} leídos · '
              f'nombre ≡ subhead del CST en {sum(1 for x in acuerdo if x)}/{vistos}')
        if any(x is False for x in acuerdo):
            print('⚠ un colofón nombra un peta distinto del que el CST pone en ese ordinal: '
                  'no se escribe nada')
            for k, f in enumerate(fs):
                if acuerdo[k] is False:
                    print(f'   ⚠ {f["num"]} «{cst[k][0]}» vs colofón «{col[tuple(int(x) for x in f["num"].split(".")[1:])][1]}»')
            return

    res, flojas = {}, 0
    for k, f in enumerate(fs):
        sub, a, b, div, tc = cst[k]
        pg = f['page']
        sig = fs[k + 1]['page'] if k + 1 < len(fs) else ULTIMA
        # ventana desde la página anterior: en esta obra el título va de colofón, detrás del poema,
        # así que la página que declara la fila es a veces la del cierre y el poema empieza antes
        txt = ' '.join(ap.fold(x['unitext'] or '') for x in conn.execute(
            'SELECT unitext FROM pages WHERE edition="mula" AND book_no=? AND page_no BETWEEN ? '
            'AND ?', (BOOK, max(1, pg - 1), max(sig, pg + 1 + len(tc) // 1200))))
        cov = R.cobertura(txt, tc)
        flojas += cov < COV_COTA
        m = re.search(rf'\b{SIGLA}\s+(\d+)', f['name'])
        if not m:
            print(f"   ✗ {f['num']}: el nombre no trae el nº de vimāna")
            continue
        res[f['num']] = (
            f'{STEM}:{a}-{b}' if b != a else f'{STEM}:{a}',
            f'{SIGLA} {m.group(1)}',
            f'{"Vimānavatthu: vimāna" if obra == "vv" else "Petavatthu: peta"} {k + 1} de '
            f'{ESPERADAS} → «{sub}» ({div}), versos {a}-{b} del CST. El nombre casa en las '
            f'{ESPERADAS} filas y la biyección respeta el orden y la partición en vaggas; '
            f'corroboración por contenido {cov:.2f}'
            + ('' if cov >= COV_COTA else
               ' (baja: el poema arranca en la página anterior o PTS lo elide remitiendo a su '
               'gemelo — en esta obra el contenido corrobora, no discrimina)'))
    print(f'\n{len(res)}/{len(fs)} firmadas · cobertura por debajo de {COV_COTA}: {flojas}')
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-{obra}.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, ci['Nikaya']).value != 'KN'
                or str(ws.cell(r, ci['PTS Vol']).value) != SIGLA):
            continue
        num = str(ws.cell(r, ci['Sutta #']).value)
        if num not in res or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, ref, det = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['PTS Ref']).value = ref
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
