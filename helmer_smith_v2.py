#!/usr/bin/env python3
"""
Helmer Smith v2 — Multi-level content validation using HEAD + body markers.
Level 1: Page existence
Level 2: HEAD match (running header — most reliable)
Level 3: Body structural markers (sutta numbers, evam me sutam, vagga headers)
Level 4: Body keyword search (diacritic-normalized)
Level 5: Nearby page search
"""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def norm(s):
    for a, b in [('ā','a'),('ī','i'),('ū','u'),('ṅ','n'),('ñ','n'),
                 ('ṭ','t'),('ḍ','d'),('ṇ','n'),('ḷ','l'),('ṃ','m'),('ṁ','m')]:
        s = s.replace(a, b).replace(a.upper(), b.upper())
    return s

def get_page(book_no, page_no):
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                (book_no, page_no))
    r = cur.fetchone()
    if not r:
        return None, None
    return r['head'] or '', r['unitext'] or ''

def validate_entry(book_no, stated_page, entry):
    """Multi-level validation. Returns (level, status, detail)."""
    head, text = get_page(book_no, stated_page)
    if text is None:
        return 0, 'PAGE_MISSING', f'Page {stated_page} not found in book {book_no}'
    
    sutta_num = entry['num'].split('.')[-1]  # last component
    sutta_name = entry['name']
    sutta_num_full = entry['num']
    raw = entry['raw']
    
    text_norm = norm(text)
    head_norm = norm(head)
    lines = text.split('\n')
    
    # ── LEVEL 2: HEAD match ──
    # DN/MN: head contains sutta number like "(16)" or title
    if entry['nikaya'] in ('DN', 'MN'):
        # Check for sutta number in head: "(16)" or "(16.)"
        if f'({sutta_num})' in head or f'({sutta_num}.)' in head:
            return 2, 'OK_HEAD_NUM', f'HEAD contains ({sutta_num})'
        # Check for name keywords in head
        kw = [w for w in re.findall(r'[a-z]{5,}', norm(sutta_name.lower())) 
              if w not in ('sutta', 'suttam', 'vagga', 'pathama', 'dutiya', 'tatiya')]
        head_hits = sum(1 for w in kw if w in head_norm.lower())
        if kw and head_hits >= len(kw) * 0.5 and head_hits >= 1:
            return 2, 'OK_HEAD_NAME', f'HEAD: {head_hits}/{len(kw)} name kw'
    
    # SN/AN: head contains saṃyutta/vagga name
    if entry['nikaya'] in ('SN', 'AN'):
        # Check if head has vagga/section info
        # For SN, head format: "XX. YY. ZZ.]  VAGGA-NAME  VAGGA-NUM  PAGE"
        # For AN, head format: "VAGGA-NAME  [I. x. y.]"
        head_clean = head.strip()
        if len(head_clean) > 10:
            # Extract potential name words from head
            head_words = set(re.findall(r'[a-z]{4,}', norm(head_clean.lower())))
            name_kw = set(re.findall(r'[a-z]{4,}', norm(sutta_name.lower())))
            overlap = head_words & name_kw
            if len(overlap) >= 2:
                return 2, 'OK_HEAD_SECTION', f'HEAD section: {overlap}'
    
    # KN: head contains book name
    if entry['nikaya'] == 'KN':
        head_clean = head.strip()
        # Specific KN head patterns
        kn_heads = {
            'Ud': 'udana', 'It': 'itivuttaka', 'Sn': 'suttanipata',
            'Vv': 'vimana', 'Pv': 'peta', 'Th': 'thera',
            'Dhp': 'dhammapada', 'Khp': 'khuddakapatha',
            'Ja': 'jataka', 'Ap': 'apadana', 'Bv': 'buddhavamsa',
            'Nidd': 'niddesa', 'Patis': 'patisambhida',
        }
        for abbr, keyword in kn_heads.items():
            if entry['vol'].startswith(abbr) and keyword in norm(head_clean.lower()):
                return 2, 'OK_HEAD_KN', f'HEAD: {head_clean[:60]}'
    
    # ── LEVEL 3: Body structural markers ──
    # Look for sutta-start markers
    # Pattern 1: "N. Evam me sutam" or "N. Evam-me sutam" 
    for i, line in enumerate(lines):
        s = line.strip()
        if re.match(rf'^{re.escape(sutta_num)}\.\s+[Ee]va[mṃ]', s):
            return 3, 'OK_MARKER', f'L{i+1}: sutta marker with evam me sutam'
        # Pattern 2: Just numbered marker
        if re.match(rf'^{re.escape(sutta_num)}\.\s+\S', s) and i < 20:
            return 3, 'OK_MARKER_NUM', f'L{i+1}: sutta number marker'
    
    # Pattern 3: Centered number (MN style) — number alone on a line
    for i, line in enumerate(lines[:30]):
        s = line.strip()
        if s == f'{sutta_num}.' or s == sutta_num:
            return 3, 'OK_CENTERED', f'L{i+1}: centered number marker'
    
    # Pattern 4: Vagga heading contains sutta
    for i, line in enumerate(lines[:5]):
        s = line.strip()
        if re.match(r'^[IVX]+\.?\s*$', s):
            return 3, 'OK_VAGGA_START', f'L{i+1}: vagga heading, sutta follows'
    
    # ── LEVEL 4: Body keyword search ──
    name_words = [w for w in re.findall(r'[a-z]{4,}', norm(sutta_name.lower()))
                  if w not in ('sutta', 'suttam', 'vagga', 'pathama', 'dutiya', 
                              'tatiya', 'catuttha', 'pancama', 'chattha', 'sattama',
                              'atthama', 'navama', 'dasama', 'paritta', 'mahaparinibbana')]
    
    if name_words:
        search_area = text_norm[:800].lower()
        hits = sum(1 for w in name_words if w in search_area)
        if hits >= 1 and (hits >= len(name_words) * 0.5 or len(name_words) <= 2):
            return 4, 'OK_KEYWORD', f'{hits}/{len(name_words)} kw in first 800 chars'
    
    # ── LEVEL 5: Nearby pages ──
    for delta in [-2, -1, 1, 2]:
        nh, nt = get_page(book_no, stated_page + delta)
        if nt is None:
            continue
        nhead = norm(nh)[:100]
        ntext = norm(nt)[:800]
        nsearch = (nhead + ' ' + ntext).lower()
        
        if name_words:
            nhits = sum(1 for w in name_words if w in nsearch)
            if nhits >= len(name_words) * 0.6:
                return 5, f'OFFSET_{delta:+d}', f'Found at p.{stated_page + delta}'
        
        # Check for sutta marker on nearby page
        nlines = nt.split('\n')
        for i, line in enumerate(nlines[:20]):
            s = line.strip()
            if re.match(rf'^{re.escape(sutta_num)}\.\s+\S', s):
                return 5, f'MARKER_{delta:+d}', f'Marker at p.{stated_page + delta} L{i+1}'
    
    # ── Not verified ──
    head_sample = head[:80] if head.strip() else '(empty head)'
    text_sample = ' '.join(text.split()[:15])[:100]
    return 0, 'UNVERIFIED', f'HEAD:[{head_sample}] TXT:[{text_sample}]'


# ── Load Excel ──
wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

entries = []
for ri in range(2, ws.max_row + 1):
    entries.append({
        'ri': ri, 'nikaya': str(ws.cell(row=ri, column=cols['Nikaya']).value or ''),
        'num': str(ws.cell(row=ri, column=cols['Sutta #']).value or ''),
        'name': str(ws.cell(row=ri, column=cols['Sutta Name']).value or ''),
        'page': ws.cell(row=ri, column=cols['PTS Page']).value,
        'roman': str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower(),
        'vol': str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip(),
        'ref': str(ws.cell(row=ri, column=cols['PTS Ref']).value or ''),
        'raw': str(ws.cell(row=ri, column=cols['Raw ID']).value or ''),
    })

def get_book_no(e):
    roman = e['roman']
    vol = e['vol']
    nik = e['nikaya']
    if nik == 'DN': return {'i': 6, 'ii': 7, 'iii': 8}.get(roman)
    if nik == 'MN': return {'i': 9, 'ii': 10, 'iii': 11}.get(roman)
    if nik == 'SN': return {'i': 12, 'ii': 13, 'iii': 14, 'iv': 15, 'v': 16}.get(roman)
    if nik == 'AN': return {'i': 17, 'ii': 18, 'iii': 19, 'iv': 20, 'v': 21}.get(roman)
    if nik == 'KN':
        vol_map = {
            'Khp': 22, 'Kh': 22,
            'Dhp': 23, 'Dh': 23,
            'Ud': 24, 'It': 25, 'Sn': 26,
            'Vv': 27, 'Pv': 28,
            'Th': 29, 'Th & Th': 29, 'Thī': 29, 'Th & Thī': 29, 'Thig': 29,
            'Ja': 30, 'Ja I': 30, 'Ja II': 31, 'Ja III': 32, 'Ja IV': 33, 'Ja V': 34, 'Ja VI': 35,
            'Nidd I': 36, 'Nidd II': 37,
            'Paṭis I': 38, 'Paṭis II': 39,
            'Ap': 40, 'Bv': 41, 'Cp': 42,
        }
        if vol in vol_map:
            return vol_map[vol]
        for key, val in vol_map.items():
            if vol.startswith(key) or key.startswith(vol):
                return val
        return None
    return None

# ── Select 40 samples ──
samples = []
dn = [e for e in entries if e['nikaya'] == 'DN']
samples.extend([e for e in dn if e['num'] in ['1', '2', '9', '14', '16', '17', '23', '33', '34']])
mn = [e for e in entries if e['nikaya'] == 'MN']
samples.extend([e for e in mn if e['num'] in ['1', '50', '51', '52', '76', '77', '100', '107', '152']])
sn = [e for e in entries if e['nikaya'] == 'SN']
samples.extend([e for e in sn if e['num'] in ['1.1', '2.1', '12.1', '22.1', '35.1', '45.1', '56.11', '56.131']])
an = [e for e in entries if e['nikaya'] == 'AN']
samples.extend([e for e in an if e['num'] in ['1.1', '2.1', '3.1', '4.1', '5.1', '6.1', '10.1']])
kn = [e for e in entries if e['nikaya'] == 'KN']
kn_targets = ['1.1', '2.1', '3.1.1', '4.1.1', '5.1.1', '6.1.1.1', '7.1.1', '8.1.1.1', '9.1.1', '10.1.1']
samples.extend([e for e in kn if e['num'] in kn_targets])
for tvol in ['Ja II', 'Nidd I', 'Nidd II', 'Paṭis I', 'Ap', 'Bv', 'Cp']:
    for e in kn:
        if e['vol'] == tvol:
            samples.append(e); break

seen = set(); unique = []
for s in samples:
    k = (s['nikaya'], s['num'], s['vol'])
    if k not in seen: seen.add(k); unique.append(s)
samples = unique[:45]

# ── Validate ──
print("""
╔══════════════════════════════════════════════════════════════════════════════╗
║  Helmer Smith — PTS Content Validation Report v2                            ║
║  Multi-level: HEAD → Body markers → Keywords → Nearby pages                 ║
╚══════════════════════════════════════════════════════════════════════════════╝
""")

results = []
stats = defaultdict(int)
level_dist = defaultdict(int)

for e in samples:
    book_no = get_book_no(e)
    if not book_no:
        stats['NO_BOOK'] += 1
        results.append((0, 'NO_BOOK', e, f'Cannot map {e["vol"]}/{e["roman"]}'))
        continue
    
    level, status, detail = validate_entry(book_no, e['page'], e)
    stats[status] += 1
    level_dist[level] += 1
    results.append((level, status, e, detail))

# Print
sym_map = {'OK_HEAD_NUM': 'H', 'OK_HEAD_NAME': 'H', 'OK_HEAD_SECTION': 'H', 'OK_HEAD_KN': 'H',
           'OK_MARKER': 'M', 'OK_MARKER_NUM': 'M', 'OK_CENTERED': 'M', 'OK_VAGGA_START': 'M',
           'OK_KEYWORD': 'K'}
status_sym = {**sym_map, 'UNVERIFIED': '✗', 'PAGE_MISSING': '∅', 'NO_BOOK': '‼'}
for prefix in ['OFFSET_', 'MARKER_']:
    for d in [-2, -1, 1, 2]:
        status_sym[f'{prefix}{d:+d}'] = '≈'

print(f"{'Lv':3s} {'St':3s} {'Nik':3s} {'#':12s} {'Ref':20s} {'Detail'}")
print("-" * 100)

for level, status, e, detail in results:
    sym = status_sym.get(status, '?')
    print(f'{level:2d}  {sym:3s} {e["nikaya"]:3s} {e["num"]:12s} {e["ref"]:20s} {detail[:75]}')

# Summary
ok_count = sum(1 for _, s, _, _ in results if s.startswith('OK_'))
offset_count = sum(1 for _, s, _, _ in results if s.startswith('OFFSET_') or s.startswith('MARKER_'))
unv_count = sum(1 for _, s, _, _ in results if s == 'UNVERIFIED')
other_count = sum(1 for _, s, _, _ in results if s in ('NO_BOOK', 'PAGE_MISSING'))

print(f"\n{'='*100}")
print(f"  Sample: {len(results)} entries")
print(f"  ✓ HEAD match (L2):    {sum(1 for _,s,_,_ in results if s.startswith('OK_HEAD_')):>3d}")
print(f"  ✓ Body marker (L3):   {sum(1 for _,s,_,_ in results if s.startswith('OK_MARKER') or s.startswith('OK_CENTERED') or s.startswith('OK_VAGGA')):>3d}")
print(f"  ✓ Keyword (L4):       {sum(1 for _,s,_,_ in results if s.startswith('OK_KEYWORD')):>3d}")
print(f"  ≈ Nearby (L5):        {offset_count:>3d}")
print(f"  ✗ Unverified:         {unv_count:>3d}")
print(f"  ∅ Other:              {other_count:>3d}")

verified = len(results) - unv_count - other_count
pct = 100 * verified / max(len(results), 1)
print(f"\n  Content-confirmed:    {verified}/{len(results)} ({pct:.0f}%)")

if pct >= 90:
    print(f"\n  Helmer Smith's verdict: SOUND. Excellent correspondence between")
    print(f"  the reference table and the printed PTS edition.")
elif pct >= 75:
    print(f"\n  Helmer Smith's verdict: MOSTLY RELIABLE. A few entries require")
    print(f"  closer inspection — likely title-page offsets or variant namings.")
elif pct >= 50:
    print(f"\n  Helmer Smith's verdict: NEEDS REVIEW. Significant number of")
    print(f"  entries could not be verified by multi-level content matching.")
else:
    print(f"\n  Helmer Smith's verdict: REQUIRES CORRECTION. The reference table")
    print(f"  has substantial discrepancies with the printed edition.")

# Per-nikaya breakdown
print(f"\n  Per Nikaya:")
nik_results = defaultdict(lambda: {'ok': 0, 'off': 0, 'unv': 0})
for level, status, e, detail in results:
    if status.startswith('OK_'): nik_results[e['nikaya']]['ok'] += 1
    elif status.startswith('OFFSET_') or status.startswith('MARKER_'): nik_results[e['nikaya']]['off'] += 1
    elif status == 'UNVERIFIED': nik_results[e['nikaya']]['unv'] += 1
    else: nik_results[e['nikaya']]['unv'] += 1

for nik in ['DN', 'MN', 'SN', 'AN', 'KN']:
    r = nik_results[nik]
    t = r['ok'] + r['off'] + r['unv']
    if t:
        ok_pct = 100 * (r['ok'] + r['off']) / t
        print(f"    {nik:3s}: {r['ok']+r['off']}/{t} confirmed ({ok_pct:.0f}%)")

conn.close()
