#!/usr/bin/env python3
"""
AN — Full parse + validation. Build complete sutta → (volume, page, line) map
from PTS text, then cross-reference against blog data in Excel.
"""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict, OrderedDict

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

AN_BOOKS = {1: 17, 2: 18, 3: 19, 4: 20, 5: 21}
ROMAN_MAP = {'i': 1, 'ii': 2, 'iii': 3, 'iv': 4, 'v': 5}
ROMAN_REV = {1: 'i', 2: 'ii', 3: 'iii', 4: 'iv', 5: 'v'}

def parse_an_volume(vol_no):
    """Parse an entire AN volume to get sequential sutta position map.
    Returns list of (nipata_no, sutta_no, page_no, line_no) for each sutta start."""
    book_no = AN_BOOKS[vol_no]
    suttas = []
    current_vagga = 0
    current_nipata = vol_no
    
    # Get max page for this book
    cur.execute('SELECT MAX(page_no) as mp FROM pages WHERE book_no=? AND edition="mula"', (book_no,))
    max_page = cur.fetchone()['mp'] or 500
    
    for page_no in range(1, max_page + 1):
        cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                   (book_no, page_no))
        r = cur.fetchone()
        if not r or not r['unitext']:
            continue
        
        head = r['head'] or ''
        text = r['unitext']
        lines = text.split('\n')
        
        for i, line in enumerate(lines):
            s = line.strip()
            
            # Vagga heading: Roman numeral alone
            if re.match(r'^[IVX]+\.?\s*$', s) and len(s) <= 8:
                # Convert Roman to int (simple)
                roman_val = 0
                roman_str = s.rstrip('.')
                roman_map = {'I':1, 'V':5, 'X':10, 'L':50}
                prev = 0
                for ch in reversed(roman_str):
                    v = roman_map.get(ch, 0)
                    if v >= prev:
                        roman_val += v
                    else:
                        roman_val -= v
                    prev = v
                if roman_val > 0:
                    current_vagga = roman_val
            
            # Nipata header (like "EKA-NIPĀTA.1", "DUKA-NIPĀTA.2")
            nip_match = re.search(r'NIPĀTA[.\s]*(\d+)', s, re.IGNORECASE)
            if nip_match:
                current_nipata = int(nip_match.group(1))
            
            # Sutta start: "N." followed by content
            m = re.match(r'^(\d+)\.\s+\S', s)
            if m:
                sutta_num = int(m.group(1))
                if sutta_num <= 1000:  # sanity check
                    suttas.append({
                        'nipata': current_nipata,
                        'vagga': current_vagga,
                        'sutta_in_vagga': sutta_num,
                        'page': page_no,
                        'line': i + 1
                    })
    
    return suttas


def assign_sequential_ids(suttas):
    """Assign composite sequential ID (like 1.1, 1.2...) to each sutta within each nipata."""
    by_nipata = defaultdict(list)
    for s in suttas:
        by_nipata[s['nipata']].append(s)
    
    for nipata, items in by_nipata.items():
        items.sort(key=lambda x: (x['page'], x['line']))
        for seq, item in enumerate(items, 1):
            item['seq_id'] = f'{nipata}.{seq}'
    
    return suttas


# ── Parse all AN volumes ──
print('Parsing AN volumes from PTS text...')
all_suttas = []
for vol in [1, 2, 3, 4, 5]:
    suttas = parse_an_volume(vol)
    all_suttas.extend(suttas)
    print(f'  Vol {vol} ({ROMAN_REV[vol]}): {len(suttas)} sutta starts found')

all_suttas = assign_sequential_ids(all_suttas)
print(f'\nTotal sutta starts parsed: {len(all_suttas)}')

# Build lookup: seq_id → page, line
pts_map = {}
for s in all_suttas:
    pts_map[s['seq_id']] = (s['page'], s['line'], s['vagga'], s['sutta_in_vagga'])

# ── Load Excel ──
wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(row=1, column=c).value
    if h:
        cols[h] = c

# Compare
print('\nComparing Excel vs PTS text...')
print('=' * 80)

correct = 0
wrong_page = 0
not_found = 0
fixed = 0

for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'AN':
        continue
    
    snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
    excel_page = ws.cell(row=ri, column=cols['PTS Page']).value
    excel_roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip()
    old_ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    
    if '-' in snum:
        # Range entry — skip for now, handled separately
        continue
    
    if snum in pts_map:
        pts_page, pts_line, pts_vagga, pts_svagga = pts_map[snum]
        pts_roman = ROMAN_REV.get(excel_roman and ROMAN_MAP.get(excel_roman.lower()), excel_roman)
        
        # Determine correct roman volume from the page
        for v in [1, 2, 3, 4, 5]:
            book = AN_BOOKS[v]
            if pts_page and book == AN_BOOKS.get(ROMAN_MAP.get(excel_roman.lower(), 0)):
                pass  # We trust the Excel volume assignment
        
        if pts_page == excel_page:
            correct += 1
        else:
            wrong_page += 1
            if wrong_page <= 20:
                print(f'  PAGE ERR: AN {snum:>8s} | Excel: A {excel_roman} {excel_page} | PTS: A {pts_roman} {pts_page} | {name[:40]}')
        
        # Fix reference
        new_vol_roman = ROMAN_REV.get(ROMAN_MAP.get(excel_roman.lower(), 0), excel_roman)
        new_ref = f'A {new_vol_roman} {pts_page}'
        if pts_line > 1:
            new_ref += f',{pts_line}'
        
        if new_ref != old_ref:
            ws.cell(row=ri, column=cols['PTS Page']).value = pts_page
            ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
            fixed += 1
    else:
        not_found += 1
        if not_found <= 10:
            print(f'  NOT FOUND: AN {snum} | {name[:50]}')

print(f'\n{"=" * 80}')
print(f'Correct page: {correct}')
print(f'Wrong page:   {wrong_page}')
print(f'Not found:    {not_found}')
print(f'Fixed refs:   {fixed}')
print(f'Total:        {correct + wrong_page + not_found}')

wb.save(XL)
print(f'Saved: {XL}')
conn.close()
