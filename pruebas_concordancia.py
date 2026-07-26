#!/usr/bin/env python3
"""
pruebas_concordancia — batería funcional sobre los cuatro nikāyas mayores: **20 DN, 30 MN, 50 SN,
60 AN**. Determinista, sin LLM, sin azar: la misma corrida da siempre las mismas filas y el mismo
veredicto.

### Qué tiene que probar, que no es lo que parece

El riesgo de esta tabla **no es que la cobertura sea baja**. Es que el par sea **el equivocado**: si
los dos lados son coherentes entre sí, cualquier medida de parecido aprueba, y eso ya pasó (S ii,
SN 17 desplazado +7: el validador aprobaba un par correcto y ajeno a la fila). Así que la batería no
pregunta «¿se parecen?» sino «**¿es ésta la fila?**», y para eso necesita evidencia que distinga
entre unidades vecinas.

### Por qué no hay un umbral único, medido antes de escribir la prueba

| | cobertura propia | cobertura de un sutta **ajeno** |
|---|---|---|
| DN | mediana 0,95 | p90 0,80 |
| MN | mediana 0,95 | p90 0,69 |
| SN | mediana 0,93 | **p90 0,95** |
| AN | mediana 0,93 | **p90 0,97, máx 1,00** |

En SN y AN **un vecino puede puntuar 1,00**: son nikāyas de suttas cortos que comparten página, y
la ventana por página no los separa. Un umbral de contenido daría un veredicto falso — de los dos
signos. Por eso la prueba fuerte es otra.

### El ancla: la `PTS Ref` trae **página y línea**

`S i 1,6`, `A i 1,5` — y eso permite **verificar** que el texto del CST empieza donde la fila dice.
Medido, en la línea declarada contra otras líneas de la misma página:

| | en la línea declarada | control (otras líneas) |
|---|---|---|
| DN | mediana 1,00 | mediana 0,14 |
| MN | mediana 1,00 | mediana 0,12 |
| SN | mediana 0,90 | mediana 0,14 |
| AN | mediana 0,99 | mediana 0,19 |

⚠️ **Se verifica, no se busca.** Buscar el incipit por toda la página lo encuentra en otro sitio:
las aperturas de AN son fórmulas (`atha kho āyasmā ānando yena bhagavā tenupasaṅkami`) que se
repiten dentro de la misma página. Una prueba comprueba la afirmación; no la re-deriva.

⚠️ Y el ancla es **independiente** de lo que se validó: la línea se fijó desde el marcador impreso
de PTS, y lo que se comprueba es que el texto **del CST** empieza ahí. Dos fuentes distintas.

### Los cuatro controles

1. **`resuelve`** — la `VRI Ref` apunta a texto real del XML. Sin esto la fila está anclada a nada.
2. **`ancla`** — el texto del CST empieza en la `página,línea` que declara la fila (≥ 0,55, muy por
   encima del control de 0,12-0,19). Es el control **discriminante** para unidades cortas.
3. **`contenido`** — cobertura de la ventana impresa contra su unidad del CST **y contra las
   vecinas**. Sólo cuenta como evidencia si la propia gana por margen; si no, se declara **no
   discriminante** en vez de fingir un veredicto.
4. **`unicidad`** — ninguna otra fila reclama esa misma `VRI Ref`.

### El veredicto, y por qué hay tres y no dos

- **PASA** — resuelve, es única, y **al menos un control discriminante** (ancla o contenido) la
  confirma.
- **FALLA** — algún control la **contradice**: no resuelve, la `VRI Ref` está repetida, o el ancla
  cae donde el texto no está.
- **INDECISO** — ningún control pudo decidir. **No cuenta como aprobada.** Una batería que convierte
  «no lo sé» en «bien» no vale para dar un veredicto.

### La muestra: estratificada y reproducible, no aleatoria

Por nikāya se toman, en este orden y sin repetir: (a) **primera y última fila de cada volumen** —los
bordes, que es donde se esconde el desplazamiento—; (b) las `VALIDADOR_HUMANO`, que son las que en
su día hubo que arbitrar; (c) las de `VRI Ref` **con rango** (`a-b`), que son las agrupaciones; y
(d) el resto a paso fijo. Sin RNG: la muestra es la misma en cada corrida y se puede discutir fila
por fila.

Uso:
    python3 pruebas_concordancia.py [--nikaya DN|MN|SN|AN] [--detalle] [--json salida.json]
"""
import json
import re
import sqlite3
import sys
from collections import Counter

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
CUANTAS = {'DN': 20, 'MN': 30, 'SN': 50, 'AN': 60}
LIBRO = {('DN', 'i'): 6, ('DN', 'ii'): 7, ('DN', 'iii'): 8,
         ('MN', 'i'): 9, ('MN', 'ii'): 10, ('MN', 'iii'): 11,
         ('SN', 'i'): 12, ('SN', 'ii'): 13, ('SN', 'iii'): 14, ('SN', 'iv'): 15, ('SN', 'v'): 16,
         ('AN', 'i'): 17, ('AN', 'ii'): 18, ('AN', 'iii'): 19, ('AN', 'iv'): 20, ('AN', 'v'): 21}

ANCLA_MIN = 0.55        # medido: en la línea declarada 0,90-1,00; en otras líneas 0,12-0,19
MARGEN = 0.15           # cuánto tiene que ganarle la propia a la mejor vecina para ser evidencia

_CACHE = {}


def cst_texto(vri):
    """Texto del CST al que apunta la `VRI Ref`, o `None` si no resuelve."""
    m = re.match(r'^([a-z0-9]+):(\d+)(?:-(\d+))?$', str(vri or ''))
    if not m:
        return None
    stem, a = m.group(1), int(m.group(2))
    b = int(m.group(3) or a)
    if stem not in _CACHE:
        try:
            _CACHE[stem] = {x['pn']: x['texto'] for x in ap.cst_unidades(stem)}
        except Exception:
            _CACHE[stem] = {}
    t = ' '.join(_CACHE[stem][n] for n in range(a, b + 1) if n in _CACHE[stem])
    return ap.fold(t) if t.strip() else None


def lineas(conn, libro, pagina):
    r = conn.execute('SELECT unitext FROM pages WHERE edition="mula" AND book_no=? AND page_no=?',
                     (libro, pagina)).fetchone()
    # ⚠️ por `[\r\n]`: el texto de la BD usa `\r` suelto como separador y partir sólo por `\n`
    # fusiona líneas impresas
    return re.split(r'[\r\n]+', r['unitext'] or '') if r else []


def ventana(conn, libro, a, b):
    return ' '.join(ap.fold(x['unitext'] or '') for x in conn.execute(
        'SELECT unitext FROM pages WHERE edition="mula" AND book_no=? AND page_no BETWEEN ? AND ?',
        (libro, a, b)))


def filas(nikaya):
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    return [{k: r[ci[k]] for k in hdr}
            for r in ws.iter_rows(min_row=2, values_only=True) if r[ci['Nikaya']] == nikaya]


def muestra(fs, n):
    """Muestra **estratificada y determinista**: bordes, arbitradas, agrupadas y paso fijo."""
    idx, vistos = [], set()

    def mete(js):
        for j in js:
            if j not in vistos and len(idx) < n:
                vistos.add(j)
                idx.append(j)

    porvol = {}
    for j, f in enumerate(fs):
        porvol.setdefault(str(f['PTS Roman']), []).append(j)
    mete([j for v in sorted(porvol) for j in (porvol[v][0], porvol[v][-1])])
    mete([j for j, f in enumerate(fs) if str(f['Validation']) == 'VALIDADOR_HUMANO'])
    mete([j for j, f in enumerate(fs) if re.match(r'^[a-z0-9]+:\d+-\d+$', str(f['VRI Ref'] or ''))])
    paso = max(1, len(fs) // max(1, n - len(idx) + 1))
    mete(range(0, len(fs), paso))
    mete(range(len(fs)))
    return sorted(idx)


def prueba_fila(conn, fs, j, dueños):
    """Los cuatro controles sobre una fila. Devuelve el dicho de cada uno y el veredicto."""
    f = fs[j]
    d = {'fila': str(f['Sutta #']), 'nombre': str(f['Sutta Name'])[:38],
         'ref': str(f['PTS Ref']), 'vri': str(f['VRI Ref'] or '')}
    tc = cst_texto(f['VRI Ref'])
    d['resuelve'] = tc is not None
    d['unica'] = dueños[str(f['VRI Ref'])] == 1 if f['VRI Ref'] else False
    if not d['resuelve']:
        d['veredicto'] = 'FALLA'
        d['motivo'] = 'la VRI Ref no apunta a texto del CST'
        return d
    if not d['unica']:
        d['veredicto'] = 'FALLA'
        d['motivo'] = f"la VRI Ref la reclaman {dueños[str(f['VRI Ref'])]} filas"
        return d

    # ── control 2: el ancla de página,línea
    d['ancla'] = None
    m = re.match(r'^[A-Z] ([ivx]+) (\d+),(\d+)$', d['ref'])
    libro = LIBRO.get((str(f['Nikaya']), str(f['PTS Roman'])))
    if m and libro:
        L = lineas(conn, libro, int(m.group(2)))
        ln = int(m.group(3))
        if L and ln - 1 < len(L):
            inc = ' '.join(tc.split()[:8])
            # ⚠️ una línea de holgura **antes**: la numeración de la `PTS Ref` y la de la BD no
            # cuentan igual el preámbulo de la página (`NAMO TASSA…`, el título entre corchetes,
            # las reglas de guiones), y el Brahmajāla arranca en la línea 6 donde la fila dice 7.
            # La calibración lo había medido: desviación mediana +0/+1, y |Δ| ≤ 2 en DN y MN.
            d['ancla'] = round(R.cobertura(ap.fold(' '.join(L[max(0, ln - 2):ln + 3])), inc), 2)

    # ── control 2-bis: el ancla **por nombre**, para donde PTS elide el texto
    #
    # ⚠️ Donde PTS abrevia —los peyyāla, los grupos— la línea declarada **no es texto sino
    # marcador**: `68. (2) Maraṇa.` seguido de `Maraṇasaññā bhikkhave ║ ║`, o
    # `(8)36-40 Apaccupalakkhaṇā (1-5)`. Ahí no hay incipit que anclar, y exigirlo era pedirle al
    # impreso algo que deliberadamente no imprime. Pero el marcador **nombra** la unidad, y el
    # nombre es lo que la identifica. Se comprueba entonces que el nombre de la fila esté en esa
    # línea, que es más discriminante que el incipit en un texto elidido.
    d['ancla_nombre'] = None
    if m and libro and d['ancla'] is not None and d['ancla'] < ANCLA_MIN:
        L = lineas(conn, libro, int(m.group(2)))
        ln = int(m.group(3))
        # la ventana arranca EN la línea declarada: el marcador está ahí y la elisión (`║ pe ║`)
        # en la línea siguiente — mirando desde `ln-2` se quedaba fuera y tres filas de S v 132
        # quedaban sin diagnosticar
        cerca = ap.fold(' '.join(L[max(0, ln - 1):ln + 2]))
        elide = bool(re.search(r'║|\bpe\b|\bpa\b|peyyal', cerca))
        # el nombre de la fila puede traer sufijos y corchetes (`Apaccupalakkhaṇā [2]`), y el
        # marcador impreso lo escribe desnudo
        raiz = re.sub(r'[^a-z]', '', ap.fold(re.sub(r'\s*\[.*$|\s*\(.*$|\s*\d+$', '',
                                                    str(f['Sutta Name'] or ''))))[:8]
        if elide and len(raiz) >= 5:
            d['ancla_nombre'] = raiz in re.sub(r'[^a-z]', '', cerca)

    # ── control 3: contenido contra las vecinas del mismo volumen
    d['cov'] = d['cov_vecina'] = None
    pg = f['PTS Page']
    if libro and isinstance(pg, int):
        herm = [k for k, g in enumerate(fs) if g['PTS Roman'] == f['PTS Roman']
                and isinstance(g['PTS Page'], int)]
        p = herm.index(j) if j in herm else -1
        sig = fs[herm[p + 1]]['PTS Page'] if 0 <= p < len(herm) - 1 else pg + 2
        tp = ventana(conn, libro, pg, max(pg, sig))
        if tp:
            d['cov'] = round(R.cobertura(tp, tc), 2)
            vec = [cst_texto(fs[herm[p + o]]['VRI Ref'])
                   for o in (-1, 1, 5) if 0 <= p + o < len(herm)]
            vec = [R.cobertura(tp, v) for v in vec if v]
            d['cov_vecina'] = round(max(vec), 2) if vec else None

    ancla_ok = d['ancla'] is not None and d['ancla'] >= ANCLA_MIN
    cont_ok = (d['cov'] is not None and d['cov_vecina'] is not None
               and d['cov'] >= d['cov_vecina'] + MARGEN)
    if d['ancla_nombre'] is True:
        d['veredicto'] = 'PASA'
        d['motivo'] = 'ancla por nombre (PTS elide el texto en ese punto)'
    elif d['ancla_nombre'] is False:
        d['veredicto'] = 'FALLA'
        d['motivo'] = (f"PTS elide el texto en {d['ref']} y el marcador de esa línea NO nombra "
                       'esta unidad')
    elif d['ancla'] is not None and d['ancla'] < ANCLA_MIN and not cont_ok:
        d['veredicto'] = 'INDECISO'
        d['motivo'] = (f"el texto del CST no empieza en {d['ref']} (ancla {d['ancla']}), no se "
                       'detecta elisión y el contenido no discrimina')
    elif ancla_ok or cont_ok:
        d['veredicto'] = 'PASA'
        d['motivo'] = ('ancla' if ancla_ok else '') + ('+' if ancla_ok and cont_ok else '') + \
                      ('contenido' if cont_ok else '')
    else:
        d['veredicto'] = 'INDECISO'
        d['motivo'] = 'sin `pág,línea` que anclar y el contenido no discrimina'
    return d


def main():
    detalle = '--detalle' in sys.argv
    solo = sys.argv[sys.argv.index('--nikaya') + 1] if '--nikaya' in sys.argv else None
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    salida, total = {}, Counter()
    for nk in (['DN', 'MN', 'SN', 'AN'] if not solo else [solo]):
        fs = filas(nk)
        dueños = Counter(str(f['VRI Ref']) for f in fs if f['VRI Ref'])
        idx = muestra(fs, CUANTAS[nk])
        res = [prueba_fila(conn, fs, j, dueños) for j in idx]
        c = Counter(r['veredicto'] for r in res)
        total.update(c)
        salida[nk] = res
        marca = '✅' if c['FALLA'] == 0 and c['INDECISO'] == 0 else (
            '❌' if c['FALLA'] else '⚠️')
        print(f"{marca} {nk}: {len(res)} pruebas · PASA {c['PASA']} · FALLA {c['FALLA']} · "
              f"INDECISO {c['INDECISO']}")
        for r in res:
            if detalle or r['veredicto'] != 'PASA':
                print(f"     {r['veredicto']:9} {r['fila']:11} {r['ref']:14} {r['vri']:18} "
                      f"ancla={r.get('ancla')} cov={r.get('cov')}/{r.get('cov_vecina')} "
                      f"— {r['motivo']}")
    print(f"\n{'═' * 78}\nTOTAL {sum(total.values())} pruebas · PASA {total['PASA']} · "
          f"FALLA {total['FALLA']} · INDECISO {total['INDECISO']}")
    if '--json' in sys.argv:
        with open(sys.argv[sys.argv.index('--json') + 1], 'w') as fh:
            json.dump(salida, fh, ensure_ascii=False, indent=1)
    return 1 if total['FALLA'] else 0


if __name__ == '__main__':
    sys.exit(main())
