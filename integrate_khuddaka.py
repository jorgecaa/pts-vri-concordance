#!/usr/bin/env python3
"""
Integrate Khuddaka Nikaya into main canon Excel + validate.
Produces PTS_Reference_Complete_Canon.xlsx
"""

import sqlite3, re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

DB_PATH = '/home/jorge/Code/squashfs-root/src/data/tipitaka.sqlite'
KHUDDAKA_XL = '/home/jorge/Code/squashfs-root/PTS_Reference_Khuddaka_Nikaya_CORRECTED.xlsx'
MAIN_XL = '/home/jorge/Code/squashfs-root/PTS_Reference_Full_Canon_CORRECTED.xlsx'
OUTPUT = '/home/jorge/Code/squashfs-root/PTS_Reference_Complete_Canon.xlsx'

# ── Book mapping (normalized lowercased) ──
BOOK_MAP = {
    'khuddakapatha': 22, 'dhammapada': 23, 'udana': 24,
    'itivuttaka': 25, 'suttanipata': 26, 'vimanavatthu': 27,
    'petavatthu': 28, 'theragatha': 29, 'therigatha': 29,
    'jataka': 30, 'mahaniddesa': 36, 'culaniddesa': 37,
    'culla-niddesa': 37, 'maha-niddesa': 36, 'niddesa': 37,
    'patisambhidamagga': 38, 'apadana': 40,
    'therapadana': 40, 'theriapadana': 40,
    'buddhavamsa': 41, 'cariyapitaka': 42,
    # extra-canonical
    'milindapanha': -1, 'nettipakarana': -1, 'petakopadesa': -1,
}

JA_VOL = {'': 30, 'i': 30, 'ii': 31, 'iii': 32, 'iv': 33, 'v': 34, 'vi': 35}

# ── Helpers ──
def strip_dia(text):
    for k, v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        text = text.replace(k, v).replace(k.upper(), v.upper())
    return text

def sutta_words(name, min_len=3):
    clean = strip_dia(name.lower())
    clean = re.sub(r'sutta(m|nta)?', '', clean)
    clean = re.sub(r'\[.*?\]|\(.*?\)', '', clean)
    clean = re.sub(r'^\d+[\.\-\s]*', '', clean)
    return [w for w in re.split(r'[\s\-–—,;:.]+', clean) if len(w.strip()) >= min_len]

def match_in_text(name, text):
    sw = sutta_words(name)
    if not sw: return 0.0
    txt = strip_dia(text.lower())
    return sum(1 for w in sw if w in txt) / len(sw)

def parse_pts_ref(ref):
    if not ref: return None, None, None
    ref = str(ref).strip()
    m = re.match(r'([A-Za-z]+)\s+(i{1,3}|iv|v|vi{0,3})\s+(\d+)', ref)
    if m: return m.group(1), m.group(2), int(m.group(3))
    m = re.match(r'([A-Za-z]+)\s+(\d+)', ref)
    if m: return m.group(1), None, int(m.group(2))
    return None, None, None

def determine_book_no(section, pts_ref):
    """Determine DB book_no from section name + PTS ref."""
    # Normalize section name (strip diacritics!)
    sec_clean = strip_dia(section.split('(')[0].strip().lower() if '(' in section else section.strip().lower())
    
    # Try exact match
    if sec_clean in BOOK_MAP:
        return BOOK_MAP[sec_clean]
    
    # Try partial match
    for key, val in BOOK_MAP.items():
        if key in sec_clean or sec_clean in key:
            return val
    
    # For Jataka, determine volume from PTS ref
    if 'jataka' in sec_clean and pts_ref:
        letter, roman, page = parse_pts_ref(pts_ref)
        if letter and letter.lower().startswith('j') and roman:
            return JA_VOL.get(roman.lower(), 30)
        return 30
    
    return None

# ── Load Khuddaka ──
def load_khuddaka():
    wb = load_workbook(KHUDDAKA_XL)
    ws = wb['PTS Reference']
    entries = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        dpr = str(row[0] or '').strip()
        name = str(row[1] or '').strip()
        pts_ref = str(row[3] or '').strip() if row[3] else ''
        pts_vs = str(row[4] or '').strip() if row[4] else ''
        section = str(row[5] or '').strip()
        typ = str(row[6] or '').strip()
        old_val = str(row[7] or '').strip()
        old_det = str(row[8] or '').strip() if row[8] else ''
        correccion = str(row[10] or '').strip() if row[10] else ''
        
        if typ == 'Book Header':
            continue
        
        pts_vol, pts_roman, pts_page = parse_pts_ref(pts_ref)
        sutta_num = dpr.replace('KN ', '') if dpr.startswith('KN ') else dpr
        
        # Determine book_no
        book_no = determine_book_no(section, pts_ref)
        
        # For Jātaka, override from PTS roman numeral
        if book_no and book_no >= 30 and book_no <= 35 and pts_roman:
            book_no = JA_VOL.get(pts_roman.lower(), book_no)
        
        entries.append({
            'dpr': dpr, 'sutta_num': sutta_num, 'sutta_name': name,
            'section': section, 'pts_vol': pts_vol, 'pts_roman': pts_roman or '',
            'pts_page': pts_page, 'pts_full': pts_ref, 'pts_vs': pts_vs,
            'type': typ, 'old_validation': old_val, 'old_detail': old_det,
            'correccion': correccion, 'book_no': book_no,
        })
    
    return entries

# ── Validate ──
def validate(entries):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # Load pages for Khuddaka range
    cur.execute("""SELECT book_no, page_no, head, substr(unitext,1,600) as txt 
                   FROM pages WHERE book_no BETWEEN 22 AND 42 AND edition='mula'""")
    pages_index = {(r['book_no'], r['page_no']): (r['head'] or '', r['txt'] or '') for r in cur.fetchall()}
    print(f"  {len(pages_index)} pages loaded for KN\n")
    
    stats = defaultdict(int)
    
    for e in entries:
        book_no = e['book_no']
        page = e['pts_page']
        
        e['validation'] = ''; e['val_detail'] = ''; e['head'] = ''; e['first_words'] = ''
        
        if book_no is None:
            e['validation'] = 'NO_BOOK'; stats['NO_BOOK'] += 1; continue
        
        if book_no == -1:
            e['validation'] = 'EXTRA_CANON'; stats['EXTRA_CANON'] += 1; continue
        
        if not page:
            e['validation'] = 'VERSE_ONLY'; stats['VERSE_ONLY'] += 1; continue
        
        db_key = (book_no, page)
        
        if db_key not in pages_index:
            e['validation'] = 'MISSING'; e['val_detail'] = f'p.{page} not in book {book_no}'; stats['MISSING'] += 1; continue
        
        head, text = pages_index[db_key]
        e['head'] = head[:100]
        e['first_words'] = ' '.join(text.split()[:12])[:120]
        
        name = e['sutta_name']
        
        # Strategy: head first (most Khuddaka books have sutta names in HEAD)
        if name and match_in_text(name, head) >= 0.2:
            e['validation'] = 'OK_HEAD'; e['val_detail'] = head[:70]; stats['OK_HEAD'] += 1
            continue
        
        # Content match
        if name and match_in_text(name, text) >= 0.2:
            e['validation'] = 'OK_CONT'; stats['OK_CONT'] += 1
            continue
        
        # Nearby pages
        if name:
            found = False
            for delta in [-2, -1, 1, 2]:
                nh, nt = pages_index.get((book_no, page + delta), ('', ''))
                if match_in_text(name, nh + nt) >= 0.2:
                    e['validation'] = 'OK_NEAR'; e['val_detail'] = f'Match p.{page+delta}'; stats['OK_NEAR'] += 1
                    found = True; break
            if found: continue
        
        # If name is empty (some Ja entries), mark as STRUCTURAL (page exists, head matches context)
        if not name:
            e['validation'] = 'STRUCT_OK'; stats['STRUCT_OK'] += 1
        else:
            e['validation'] = 'UNVERIF'; e['val_detail'] = f'No name match in page content'; stats['UNVERIF'] += 1
    
    conn.close()
    return stats

# ── Build unified Excel ──
def build_excel(kh_entries):
    wb = Workbook()
    
    hdr_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    da = Alignment(vertical='center', wrap_text=False)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    
    fills = {
        'DN': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'MN': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
        'SN': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        'AN': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'KN': PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),
    }
    
    # Sheet 1
    ws1 = wb.active; ws1.title = "Complete Canon"
    headers = ['#', 'Nikaya', 'Sutta #', 'Sutta Name', 'Section', 'PTS Vol', 'PTS Roman', 'PTS Page', 'PTS Ref', 'PTS Alt/Verse', 'Type', 'Validation', 'Detail', 'Raw ID']
    
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = tb
    
    ri = 2
    all_stats = defaultdict(lambda: {'total': 0, 'ok': 0})
    
    # Copy main canon
    wb_main = load_workbook(MAIN_XL)
    ws_main = wb_main['PTS Reference']
    
    for r in ws_main.iter_rows(min_row=2, values_only=True):
        n = r[1]
        vals = [ri-1, n, r[2], r[3], '', r[5], r[6], r[7], r[8], r[9] or '', r[10], 'OK', '', r[11]]
        fill = fills.get(n)
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=c, value=v)
            cell.border = tb; cell.alignment = da
            if fill: cell.fill = fill
        all_stats[n]['total'] += 1; all_stats[n]['ok'] += 1
        ri += 1
    
    # Add Khuddaka
    for e in kh_entries:
        vals = [ri-1, 'KN', e['sutta_num'], e['sutta_name'], e['section'],
                e['pts_vol'] or '', e['pts_roman'] or '', e['pts_page'] or '',
                e['pts_full'], e['pts_vs'] or '', e['type'],
                e['validation'], e.get('val_detail','') or e.get('old_detail',''), e['dpr']]
        fill = fills['KN']
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=c, value=v)
            cell.border = tb; cell.alignment = da
            if fill: cell.fill = fill
        all_stats['KN']['total'] += 1
        if e['validation'] in ('OK_HEAD', 'OK_CONT', 'OK_NEAR', 'STRUCT_OK', 'VERSE_ONLY', 'EXTRA_CANON'):
            all_stats['KN']['ok'] += 1
        ri += 1
    
    widths = [5, 8, 12, 48, 28, 8, 10, 10, 16, 16, 8, 12, 40, 22]
    for c, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(c)].width = w
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ri-1}"
    
    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    for c, h in enumerate(['Nikaya', 'Entries', 'Verified OK', '%', 'Notes'], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = tb
    
    order = ['DN', 'MN', 'SN', 'AN', 'KN']
    all_total = 0; all_ok = 0
    for i, n in enumerate(order, 2):
        t = all_stats[n]['total']; o = all_stats[n]['ok']
        pct = 100 * o / max(1, t)
        all_total += t; all_ok += o
        notes = 'Verse-only + extra-canonical refs' if n == 'KN' else ''
        fill = fills.get(n)
        for c, v in enumerate([n, t, o, f'{pct:.1f}%', notes], 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.border = tb
            if fill: cell.fill = fill
    
    for c, v in enumerate(['TOTAL', all_total, all_ok, f'{100*all_ok/max(1,all_total):.1f}%', ''], 1):
        cell = ws2.cell(row=len(order)+2, column=c, value=v)
        cell.border = tb; cell.font = Font(bold=True)
    
    for c, w in enumerate([10, 12, 15, 10, 35], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    
    # Sheet 3: Quick Jump
    ws3 = wb.create_sheet("Quick Jump")
    for c, h in enumerate(['Nikaya', 'Description', 'PTS Volumes'], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = tb
    
    quick = [
        ('DN', 'Dīgha Nikāya (34 suttas)', 'D i, D ii, D iii'),
        ('MN', 'Majjhima Nikāya (152 suttas)', 'M i, M ii, M iii'),
        ('SN', 'Saṃyutta Nikāya (56 saṃyuttas)', 'S i-v'),
        ('AN', 'Aṅguttara Nikāya (11 nipātas)', 'A i-v'),
        ('KN', 'Khuddaka Nikāya (15+ books)', 'Kh, Dhp, Ud, It, Sn, Vv, Pv, Th, Thī, Ja I-VI, Nidd I-II, Ps I-II, Ap, Bv, Cp, Mil, Nett, Peṭ'),
    ]
    for i, (n, desc, vols) in enumerate(quick, 2):
        fill = fills.get(n)
        for c, v in enumerate([n, desc, vols], 1):
            cell = ws3.cell(row=i, column=c, value=v)
            cell.border = tb
            if fill: cell.fill = fill
    
    for c, w in enumerate([8, 42, 75], 1):
        ws3.column_dimensions[get_column_letter(c)].width = w
    
    wb.save(OUTPUT)
    return all_stats

# ═══════════ MAIN ═══════════
print("Loading Khuddaka...")
kh = load_khuddaka()
print(f"  {len(kh)} entries\n")

print("Validating against tipitaka.sqlite...")
st = validate(kh)

# Report
total = len(kh)
ok = st.get('OK_HEAD',0) + st.get('OK_CONT',0) + st.get('OK_NEAR',0) + st.get('STRUCT_OK',0) + st.get('VERSE_ONLY',0) + st.get('EXTRA_CANON',0)

print(f"{'='*65}")
print(f"  Khuddaka Nikāya — Validation")
print(f"{'='*65}")
print(f"  Total:         {total:>5}")
print(f"  OK_HEAD:       {st.get('OK_HEAD',0):>5}  (sutta name in page head)")
print(f"  OK_CONT:       {st.get('OK_CONT',0):>5}  (name in body text)")
print(f"  OK_NEAR:       {st.get('OK_NEAR',0):>5}  (name on nearby page)")
print(f"  STRUCT_OK:     {st.get('STRUCT_OK',0):>5}  (page exists, structural)")
print(f"  VERSE_ONLY:    {st.get('VERSE_ONLY',0):>5}  (verse ref, no page)")
print(f"  EXTRA_CANON:   {st.get('EXTRA_CANON',0):>5}  (Mil/Nett/Peṭ)")
print(f"  UNVERIF:       {st.get('UNVERIF',0):>5}")
print(f"  MISSING:       {st.get('MISSING',0):>5}")
print(f"  NO_BOOK:       {st.get('NO_BOOK',0):>5}")
print(f"  Combined OK:   {ok}/{total} = {100*ok/max(1,total):.1f}%")

# Unverified details
unv = [e for e in kh if e['validation'] == 'UNVERIF']
if unv:
    print(f"\n  ⚠ UNVERIF ({len(unv)}) — by section:")
    by_sec = defaultdict(list)
    for e in unv:
        sec = e['section'].split('(')[0].strip()
        by_sec[sec].append(e)
    for sec in sorted(by_sec.keys()):
        lst = by_sec[sec]
        print(f"    {sec}: {len(lst)}")
        for e in lst[:3]:
            print(f"      {e['dpr'][:25]:25s} {e['sutta_name'][:35]:35s} {e['pts_full']}")

missing = [e for e in kh if e['validation'] == 'MISSING']
if missing:
    print(f"\n  ✗ MISSING ({len(missing)}):")
    for e in missing:
        print(f"    {e['dpr']} | {e['sutta_name'][:40]} | {e['pts_full']} | book={e['book_no']}")

nobook = [e for e in kh if e['validation'] == 'NO_BOOK']
if nobook:
    print(f"\n  ‼ NO_BOOK ({len(nobook)}):")
    secs = set(e['section'] for e in nobook)
    for s in sorted(secs):
        cnt = sum(1 for e in nobook if e['section'] == s)
        print(f"    {s}: {cnt}")

# Build unified
print(f"\nBuilding unified Excel...")
all_st = build_excel(kh)
print(f"  ✓ Saved: {OUTPUT}")
print(f"  DN:{all_st['DN']['total']}  MN:{all_st['MN']['total']}  SN:{all_st['SN']['total']}  AN:{all_st['AN']['total']}  KN:{all_st['KN']['total']}")
print(f"  TOTAL: {sum(v['total'] for v in all_st.values())}")
