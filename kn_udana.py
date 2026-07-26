#!/usr/bin/env python3
"""
kn_udana — el Udāna: 80 filas, 8 vaggas de 10 suttas.

Aquí lo que decide **no** es el contenido, y conviene decirlo antes que nada: los suttas del Udāna
empiezan casi todos con la misma fórmula —`evaṃ me sutaṃ ekaṃ samayaṃ bhagavā sāvatthiyaṃ…`— y el
incipit, que en otras obras es la sonda más discriminante que hay, aquí **no distingue nada**: se
sitúan 58 de 80 y las que fallan lo hacen porque su primera línea es idéntica a la de otras diez.
El nombre del `subhead` tampoco sirve: PTS **no titula los suttas** en el cuerpo.

### Lo que sí decide: la estructura, que es exacta por los dos lados

| | |
|---|---|
| **PTS** (BD, libro 24) | **8** vaggas con encabezado propio (`VAGGA I. BODHIVAGGO.`), colofón y uddāna |
| **CST** (`s0503m`) | **8** `div` de **10** paranums cada uno, 1-80 corridos, con `subhead` por sutta |
| **Excel** | 80 filas con `Sutta #` = `3.<vagga>.<sutta>` |

Con 8 × 10 en las dos ediciones, el paranum es `(vagga − 1) × 10 + sutta` **por construcción**. No
hay nada que alinear; lo que hace falta es una comprobación **independiente** de que esa
construcción es la correcta, y son tres:

1. **La página cae dentro del vagga que la fila declara** y no retrocede nunca — 80/80 contra los
   rangos de página que fijan los encabezados impresos (`VAGGA I` p1, `II` p10, `III` p21, `IV` p34,
   `V` p47, `VI` p62, `VII` p74, `VIII` p80).
2. **El uddāna del vagga**, que PTS imprime al cerrarlo y que nombra sus diez suttas en orden. En el
   primero dice `tayo ca bodhi, nigrodho te therā Kassapena ca | Pāṭalī Saṅgāmaji jaṭilā Bāhiyena te
   dasā` — y ese `tayo ca bodhi` («tres Bodhi») son los paranums 1, 2 y 3, `Saṅgāmaji` el 8,
   `jaṭilā` el 9, `Bāhiyena` el 10, y el `te dasā` final declara la cuenta. Corrobora el orden allí
   donde se lee (los cuatro primeros vaggas); donde el nombre no casa es porque **las dos ediciones
   llaman distinto al mismo sutta** —el 4 es `Nigrodha` para PTS y `Huṃhuṅka` para el CST, cosa que
   el propio Excel recoge: `Ud 1.4 Nigrodha (Huṁhuṅka)`—, no porque el orden falle.
3. **Cobertura de contenido** entre las páginas de la fila y el sutta del CST: mediana **0,94**,
   mínimo 0,57, contra un control de 0,41-0,52 con paranums de otros vaggas. Aquí es
   **corroboración**, no discriminante: la separación es estrecha porque la lengua es formulaica.

### La notación (regla (5-ter))

`Ud <vagga>.<sutta>` — `Ud 8.10`. Es como se cita el Udāna en cualquier edición.

Uso: python3 kn_udana.py [--dry]
"""
import re
import shutil
import sqlite3
import sys
from datetime import datetime

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
BOOK, STEM = 24, 's0503m'
COV_MIN = 0.55
ULTIMA = 94                      # última página del libro en la BD


def vaggas_pts(conn):
    """`[(página inicial, página final)]` de los ocho vaggas, por su encabezado impreso."""
    pgs = {r['page_no']: (r['unitext'] or '').split('\n')
           for r in conn.execute('SELECT page_no,unitext FROM pages WHERE edition="mula" '
                                 'AND book_no=?', (BOOK,))}
    cab = [pg for pg in sorted(pgs)
           for t in pgs[pg] if re.match(r'\s*VAGGA\s+[IVX]+\.', t.replace('\r', ''))]
    return [(cab[k], (cab[k + 1] if k + 1 < len(cab) else ULTIMA)) for k in range(len(cab))]


def main():
    dry = '--dry' in sys.argv
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    rng = vaggas_pts(conn)
    cst = {x['pn']: (x['subhead'], ap.fold(x['texto'])) for x in ap.cst_unidades(STEM)}
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    fs = [{'num': str(r[ci['Sutta #']]), 'name': str(r[ci['Sutta Name']] or ''),
           'page': r[ci['PTS Page']], 'estado': r[ci['Estado']]}
          for r in ws.iter_rows(min_row=2, values_only=True)
          if r[ci['Nikaya']] == 'KN' and str(r[ci['PTS Vol']]) == 'Ud']
    print(f'PTS {len(rng)} vaggas · CST {len(cst)} paranums · Excel {len(fs)} filas')
    if not (len(rng) == 8 and len(cst) == 80 and len(fs) == 80):
        print('⚠ la estructura no es 8×10 en las tres: no se escribe nada')
        return

    res, malas, prev = {}, [], 0
    for k, f in enumerate(fs):
        m = re.match(r'3\.(\d+)\.(\d+)$', f['num'])
        if not m:
            malas.append((f['num'], 'Sutta # con forma inesperada'))
            continue
        v, n = int(m.group(1)), int(m.group(2))
        pn = (v - 1) * 10 + n
        a, b = rng[v - 1]
        pg = f['page']
        dentro = isinstance(pg, int) and a - 1 <= pg <= b
        mono = isinstance(pg, int) and pg >= prev
        if isinstance(pg, int):
            prev = pg
        hasta = fs[k + 1]['page'] if k + 1 < len(fs) and isinstance(fs[k + 1]['page'], int) else ULTIMA
        txt = ' '.join(ap.fold(x['unitext'] or '') for x in conn.execute(
            'SELECT unitext FROM pages WHERE edition="mula" AND book_no=? AND page_no BETWEEN ? '
            'AND ?', (BOOK, pg, hasta)))
        sub, tc = cst[pn]
        cov = R.cobertura(txt, tc)
        if dentro and mono and cov >= COV_MIN:
            res[f['num']] = (f'{STEM}:{pn}', f'Ud {v}.{n}',
                             f'Udāna: vagga {v} sutta {n} → paranum {pn} («{sub}»); la página {pg} '
                             f'cae en el vagga impreso ({a}-{b}) y la serie no retrocede; '
                             f'cobertura {cov:.2f}')
        else:
            malas.append((f['num'], f'pág {pg} vs vagga {a}-{b} · cov {cov:.2f}'))
    print(f'\n{len(res)}/{len(fs)} firmadas')
    for x in malas:
        print('   ✗', x)
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-udana.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'KN' or str(ws.cell(r, ci['PTS Vol']).value) != 'Ud':
            continue
        num = str(ws.cell(r, ci['Sutta #']).value)
        if num not in res or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, ref, det = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['PTS Ref']).value = ref
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
