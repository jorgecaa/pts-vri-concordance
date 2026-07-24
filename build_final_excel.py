#!/usr/bin/env python3
"""
Final 1:1 Excel — DB-verified (PTS mula only). Fixed KN finders and DN content-line priority.

Sources: PTS canon DB (edition='mula') ONLY. No external editions — no RTE/BUDSIR, SuttaCentral,
or tipitaka.lk (project rules). A reference is OK if a marker is found on its stated page, else
UNVERIFIED. (CST/Helmer confirmation is applied separately; see helmer_*.py / STATUS.md.)
"""
import sqlite3
from openpyxl import load_workbook
from collections import defaultdict, Counter

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def norm(s):
    for a,b in [('ā','a'),('ī','i'),('ū','u'),('ṅ','n'),('ñ','n'),
                ('ṭ','t'),('ḍ','d'),('ṇ','n'),('ḷ','l'),('ṃ','m'),('ṁ','m')]:
        s = s.replace(a,b).replace(a.upper(),b.upper())
    return s

def get_page(book_no, page_no):
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None, None
    return r['head'] or '', r['unitext'] or ''

# ── Per-Nikaya marker finders ──
# Grammar lives in pts_markers.py (pyparsing; see docs/grammar.md). Behaviour is verified
# equivalent to the previous regex version over every real page (scratchpad/dev_markers.py).
from pts_markers import (
    find_markers_dn, find_markers_mn, find_markers_sn, find_markers_an, find_markers_kn,
)

# ── Load Excel ──
wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

entries = []
for ri in range(2, ws.max_row + 1):
    nik = str(ws.cell(row=ri, column=cols['Nikaya']).value or '')
    snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    if not snum or snum == 'None': continue
    entries.append({
        'ri': ri, 'nik': nik, 'num': snum,
        'name': str(ws.cell(row=ri, column=cols['Sutta Name']).value or ''),
        'roman': str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower(),
        'vol': str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip(),
        'page': ws.cell(row=ri, column=cols['PTS Page']).value,
        'ref': str(ws.cell(row=ri, column=cols['PTS Ref']).value or ''),
    })

def get_book_no(e):
    nik, roman, vol = e['nik'], e['roman'], e['vol']
    if nik == 'DN': return {'i':6,'ii':7,'iii':8}.get(roman)
    if nik == 'MN': return {'i':9,'ii':10,'iii':11}.get(roman)
    if nik == 'SN': return {'i':12,'ii':13,'iii':14,'iv':15,'v':16}.get(roman)
    if nik == 'AN': return {'i':17,'ii':18,'iii':19,'iv':20,'v':21}.get(roman)
    if nik == 'KN':
        m = {'Khp':22,'Kh':22,'Dhp':23,'Dh':23,'Ud':24,'It':25,'Sn':26,
             'Vv':27,'Pv':28,'Th':29,'Th & Th':29,'Thi':29,'Thī':29,
             'Ja':30,'Ja I':30,'Ja II':31,'Ja III':32,'Ja IV':33,'Ja V':34,'Ja VI':35,
             'Nidd':36,'Nidd I':36,'Nidd II':37,
             'Patis I':38,'Patis II':39,'Ap':40,'Bv':41,'Cp':42}
        if vol in m: return m[vol]
        for k,v in m.items():
            if vol.startswith(k) or k.startswith(vol): return v
    return None

# ── Group by page ──
page_groups = defaultdict(list)
for e in entries:
    book_no = get_book_no(e)
    if not book_no: continue
    page_groups[(e['nik'], book_no, e['page'])].append(e)

# ── Verify and update ──
print('\nVerifying and updating...')
stats = defaultdict(Counter)
updated = 0

for (nik, book_no, page), page_ents in sorted(page_groups.items()):
    head, text = get_page(book_no, page)
    
    if text is None:
        for e in page_ents:
            ws.cell(row=e['ri'], column=cols['Validation']).value = 'PAGE_MISSING'
        continue
    
    if nik == 'DN': markers = find_markers_dn(text)
    elif nik == 'MN': markers = find_markers_mn(text)
    elif nik == 'SN': markers = find_markers_sn(text)
    elif nik == 'AN': markers = find_markers_an(text)
    elif nik == 'KN': markers = find_markers_kn(text, page_ents[0]['vol'] if page_ents else '')
    else: markers = []
    
    vl = page_ents[0]['vol']; rm = page_ents[0]['roman']

    for ei, e in enumerate(page_ents):
        if ei < len(markers):
            line, mtype = markers[ei]
            new_ref = f'{vl} {rm} {page}'
            if line > 1: new_ref += f',{line}'

            if new_ref != e['ref']:
                ws.cell(row=e['ri'], column=cols['PTS Ref']).value = new_ref
                updated += 1

            ws.cell(row=e['ri'], column=cols['Validation']).value = 'OK'
            stats[nik]['verified'] += 1
        else:
            stats[nik]['unverified'] += 1
            ws.cell(row=e['ri'], column=cols['Validation']).value = 'UNVERIFIED'

wb.save(XL)

# ── Report ──
print()
for nik in ['DN', 'MN', 'SN', 'AN', 'KN']:
    s = stats[nik]; ver = s.get('verified',0); unv = s.get('unverified',0)
    total = ver + unv
    print(f'{nik:>4s}: {total:>5d} | verified={ver:>4d} unverified={unv:>4d} | {100*ver/max(total,1):.0f}%')

t_ver = sum(s.get('verified',0) for s in stats.values())
t_all = sum(s.get('verified',0)+s.get('unverified',0) for s in stats.values())
print(f'\nTOTAL: {t_all} | verified={t_ver} ({100*t_ver/t_all:.0f}%) | updated={updated}')
print(f'Saved: {XL}')
conn.close()
