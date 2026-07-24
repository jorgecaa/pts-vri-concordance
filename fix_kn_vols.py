#!/usr/bin/env python3
"""
Fix ALL KN entries with empty PTS Vol / PTS Roman / PTS Page.
Handles special formats: 'Bv & Cp X', 'Nidd ii', Petakopadesa page numbers.
"""
import re
from openpyxl import load_workbook

XL = 'PTS_Reference_Complete_Canon.xlsx'

# Map KN book number (from raw ID) → (vol_name, roman, db_book)
KN_BOOK_MAP = {
    '1': ('Khp', '', 22),
    '2': ('Dhp', '', 23),
    '3': ('Ud', '', 24),
    '4': ('It', '', 25),
    '5': ('Sn', '', 26),
    '6': ('Vv', '', 27),
    '7': ('Pv', '', 28),
    '8': ('Th', '', 29),
    '9': ('Thī', '', 29),
    '10': ('Ap', '', 40),
    '11': ('Ja', '', 30),
    '12': ('Bv', '', 41),   # Bv section of Bv&Cp
    '13': ('Bv', '', 41),   # Cp section of Bv&Cp (same physical vol, starts p.73)
    '14': ('Bv', '', 41),   # Also Bv
    '15': ('Cp', '', 42),   # Separate Cp volume
    '16': ('Nett', '', None),  # Extra-canonical
    '17': ('Pet', '', None),
    '18': ('Pet', '', None),
}

# Ja page ranges for volume assignment
JA_RANGES = [
    (1, 95, 'Ja', 'i', 30),
    (96, 266, 'Ja II', 'ii', 31),
    (267, 400, 'Ja III', 'iii', 32),
    (401, 510, 'Ja IV', 'iv', 33),
    (511, 600, 'Ja V', 'v', 34),
    (601, 700, 'Ja VI', 'vi', 35),
]


def parse_ref(ref_str):
    """Parse PTS ref into (vol_prefix, roman, page)."""
    ref_str = ref_str.strip()
    if not ref_str:
        return None, None, None
    
    # "Bv & Cp 73"
    m = re.match(r'^Bv\s*&?\s*Cp\s+(\d+)$', ref_str, re.IGNORECASE)
    if m:
        return 'Bv', '', int(m.group(1))
    
    # "Ja ii 96" or "Ja II 96"
    m = re.match(r'^Ja\s+([ivxlcdm]+)\s+(\d+)$', ref_str, re.IGNORECASE)
    if m:
        roman = m.group(1).lower()
        page = int(m.group(2))
        for lo, hi, jvol, jrom, jbook in JA_RANGES:
            if lo <= page <= hi:
                return jvol, jrom, page
        return 'Ja', roman, page
    
    # "Nidd i 1" or "Nidd I 1"
    m = re.match(r'^Nidd\s+([ivxlcdm]+)\s*(\d+)?$', ref_str, re.IGNORECASE)
    if m:
        roman = m.group(1).lower()
        page = int(m.group(2)) if m.group(2) else None
        vol = 'Nidd I' if roman == 'i' else 'Nidd II'
        return vol, '', page
    
    # "Nidd i" without page
    m = re.match(r'^Nidd\s+([ivxlcdm]+)$', ref_str, re.IGNORECASE)
    if m:
        roman = m.group(1).lower()
        vol = 'Nidd I' if roman == 'i' else 'Nidd II'
        return vol, '', None
    
    # "Ap i 3" — Apadana with roman and page
    m = re.match(r'^Ap\s+([ivxlcdm]+)\s+(\d+)$', ref_str, re.IGNORECASE)
    if m:
        return 'Ap', m.group(1).lower(), int(m.group(2))
    
    # "Thī 123" — single volume + page
    m = re.match(r'^(Thī|Thig|Th|Khp|Dhp|Dh|Ud|It|Sn|Vv|Pv)\s+(\d+)$', ref_str, re.IGNORECASE)
    if m:
        prefix_map = {'Thī': 'Th & Th', 'Thig': 'Th & Th', 'Th': 'Th & Th',
                      'Khp': 'Khp', 'Dh': 'Dhp', 'Dhp': 'Dhp',
                      'Ud': 'Ud', 'It': 'It', 'Sn': 'Sn', 'Vv': 'Vv', 'Pv': 'Pv'}
        vol = prefix_map.get(m.group(1), m.group(1))
        return vol, '', int(m.group(2))
    
    # Just a number (page only) — Petakopadesa, Netti style
    m = re.match(r'^(\d+)$', ref_str)
    if m:
        return None, '', int(m.group(1))
    
    return None, None, None


# ── Load Excel ──
wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

print('Fixing KN entries (pass 2)...')
fixed_vol = 0
fixed_page = 0
could_not = 0
issues = []

for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'KN':
        continue
    
    vol = str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip()
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip()
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '').strip()
    raw = str(ws.cell(row=ri, column=cols['Raw ID']).value or '')
    snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    
    # Extract KN book number from raw ID
    kn_match = re.match(r'KN\s+(\d+)', raw)
    kn_book = kn_match.group(1) if kn_match else None
    
    # Parse the ref
    ref_vol, ref_roman, ref_page = parse_ref(ref)
    
    # Determine vol from raw ID if ref parsing failed
    if not ref_vol and kn_book and kn_book in KN_BOOK_MAP:
        ref_vol, ref_roman, _ = KN_BOOK_MAP[kn_book]
    
    # Apply fixes
    changed = False
    
    if not vol and ref_vol:
        ws.cell(row=ri, column=cols['PTS Vol']).value = ref_vol
        fixed_vol += 1
        changed = True
    
    if not roman and ref_roman:
        ws.cell(row=ri, column=cols['PTS Roman']).value = ref_roman
        changed = True
    
    if not page and ref_page:
        ws.cell(row=ri, column=cols['PTS Page']).value = ref_page
        fixed_page += 1
        changed = True
    
    if changed and fixed_vol + fixed_page <= 20:
        old_v = vol or '(empty)'
        old_p = page or '(empty)'
        new_v = ref_vol or old_v
        new_p = ref_page or old_p
        print(f'  {snum:>10s} | {ref:>18s} → Vol={str(new_v):12s} Page={str(new_p):5s}')

    if not changed and (not vol or not page):
        could_not += 1
        if could_not <= 5:
            issues.append((snum, ref, raw[:60]))

wb.save(XL)

print(f'\nVol fixes: {fixed_vol}')
print(f'Page fixes: {fixed_page}')
print(f'Still missing: {could_not}')
if issues:
    print('Remaining issues:')
    for snum, ref, raw in issues:
        print(f'  {snum}: ref={ref} | raw={raw}')

# -- Verify final state --
empty_vol = 0
empty_page = 0
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'KN': continue
    if not str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip(): empty_vol += 1
    if not ws.cell(row=ri, column=cols['PTS Page']).value: empty_page += 1

print(f'\nFinal: {empty_vol} with empty vol, {empty_page} with empty page')
print(f'Saved: {XL}')
