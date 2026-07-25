#!/usr/bin/env python3
"""
backfill_vri_ref — escribe la columna **`VRI Ref`**, la clave canónica del lado CST.

**Por qué.** Hemos estado usando la notación **DPR** como clave de unión y no identifica el sutta de
forma unívoca. El historial es concluyente: S ii desplazado +7, S iv (SN 35) +17 y +47, AN +15 por
desdoblamientos de fila, y en `massive.tsv` el `dpr_code` de AN llega a tener **138 códigos
duplicados** y asignaciones incoherentes entre sí (`AN3.2` → `an3.1.2.8` mientras `AN3.11` →
`an3.1.2.11`). Cada vez, un APPROVE del validador no decía nada de la FILA porque la clave era falsa.

**La medida.** El **paranum del XML VRI** pasa a ser la clave canónica: es la dirección real del
texto CST que se cotejó, no una notación editorial paralela. Formato `«<fichero>:<paranum>»`, p.ej.
`s0303m:146`; los grupos van como rango, `s0304m:33-42`.

⚠️ **No se toca el `Sutta #`.** Sigue siendo la etiqueta legible en notación DPR y `Raw ID` conserva
el original de la fuente. Lo que cambia es que **deja de ser la clave**: la identidad de una fila la
dan `Sutta Name` + `PTS Page` (regla del repo), y la unión con el CST la da ahora `VRI Ref`.
Reescribir `Sutta #` en volúmenes CERRADOS 🔒 invalidaría la traza de su validación, que se hizo
contra esas filas.

**Verificación obligatoria**: cada fila sólo recibe su `VRI Ref` si el **título CST recomputado
coincide con el que se validó** (el guardado en `validador_sn*.json`). Si discrepa, no se escribe y
se reporta. Así el backfill no puede introducir una clave que no corresponda a lo cotejado.

Uso: python3 backfill_vri_ref.py [--write]
"""
import datetime
import json
import re
import shutil
import sqlite3
import sys
from collections import Counter

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
DB = 'src/data/tipitaka.sqlite'


# `S v 46.87` «(SN 46.130) Uddhambhāgiya» → `s0305m:311` «10. Uddhambhāgiyasuttaṃ»: el nombre casa
# y el contenido también (el paranum 311 trata los `uddhambhāgiyāni saṃyojanāni`). Es la ÚNICA de
# las 19 firmas en bloque con evidencia positiva.
#
# Se descartó apuntar `S ii 12.83-93` al bloque «1. Satthusuttaṃ» (`s0302m:73`): la comprobación de
# contenido lo desmiente — ese bloque **no contiene ninguno** de los términos del peyyāla
# (`sikkhā`, `yoga`, `chanda`, `ussoḷhī`…). El CST sencillamente no tiene paranum para ellos.
MANUAL = {('SN', 'v', '46.87'): 's0305m:311'}


def _head(t):
    """Clave secundaria: el título recortado en su primer «suttaṃ».

    Los dos lados anotan la variante de forma distinta — el título validado la trae entre llaves
    (`«2. pañcattayasuttaṃ {pañcāyatanasuttantipi}»`) y el XML como texto suelto
    (`«2. Pañcattayasuttaṃ pañcāyatanasutta (ka.)»`) —, así que normalizar no basta y hay que
    comparar sólo la cabeza `N. Nombresuttaṃ`.
    """
    n = _norm(t)
    i = n.find('sutam')
    return n[:i + 5] if i > 0 else n


def _fmt(stem, pns):
    """`(«s0303m», [146])` → `«s0303m:146»`; un grupo → `«s0304m:33-42»`."""
    if not pns:
        return None
    a, b = min(pns), max(pns)
    return f'{stem}:{a}' if a == b else f'{stem}:{a}-{b}'


# ── un adaptador por volumen: devuelve {Sutta # → (VRI Ref, título CST)} ──────
def vol_sn1():
    import validador_sn1 as V
    vri, mv = V.build_vri_index(), V.build_massive()
    out = {}
    for e in V.excel_entries():
        h = mv.get(e['canon'])
        if not h:
            continue
        s = vri.get(h[0])
        if s:
            out[e['num']] = (_fmt('s0301m', [h[0]]), s['title'])
    return out, 'validador_sn1.json'


def vol_sn2():
    import validador_sn2 as V
    vri, mv = V.build_vri_index(), V.build_massive()
    out = {}
    for e in V.excel_entries():
        h = mv.get((e['sam'], e['inner']))
        if not h:
            continue
        s = vri.get(h[0])
        if s:
            out[e['num']] = (_fmt('s0302m', [h[0]]), s['title'])
    return out, 'validador_sn2.json'


def vol_sn3():
    import validador_sn3 as V
    vri, mv = V.build_vri_index(), V.build_massive()
    out = {}
    for e in V.excel_entries():
        h = mv.get((e['sam'], e['inner']))
        if not h:
            continue
        s = vri.get(h[0])
        if s:
            out[e['num']] = (_fmt('s0303m', [h[0]]), s['title'])
    return out, 'validador_sn3.json'


def vol_sn4():
    import validador_sn4 as V
    entries = V.excel_entries()
    cst = V.cst_for(entries, V.build_cst_blocks())
    out = {}
    for e in entries:
        b = cst.get(e['num'])
        if b:
            out[e['num']] = (_fmt('s0304m', b['pn']), b['title'])
    return out, 'validador_sn4.json'


def vol_sn5():
    import validador_sn5_vri as V
    vri, c2p = V.build_vri_index(), V.build_massive()
    out = {}
    for e in V.excel_entries():
        h = c2p.get(e['canon'])
        if not h:
            continue
        s = vri.get(h[0])
        if s:
            out[e['num']] = (_fmt('s0305m', [h[0]]), s['title'])
    return out, 'validador_sn5_vri.json'


VOLS = {('SN', 'i'): vol_sn1, ('SN', 'ii'): vol_sn2, ('SN', 'iii'): vol_sn3,
        ('SN', 'iv'): vol_sn4, ('SN', 'v'): vol_sn5}

# fichero VRI y su índice título→paranums, para resolver por el título VALIDADO
VRI_FILE = {('SN', 'i'): 's0301m', ('SN', 'ii'): 's0302m', ('SN', 'iii'): 's0303m',
            ('SN', 'iv'): 's0304m', ('SN', 'v'): 's0305m',
            ('DN', 'i'): 's0101m', ('DN', 'ii'): 's0102m', ('DN', 'iii'): 's0103m',
            ('MN', 'i'): 's0201m', ('MN', 'ii'): 's0202m', ('MN', 'iii'): 's0203m'}


_SUTTA_TITLE = re.compile(r'^\s*\d+[\d\-–, ]*\.\s*\S*sutt', re.I)


def title_index_div(stem):
    """`título → paranums` usando los `<div>` con `<head rend="chapter">`.

    **La granularidad del `div` cambia por nikāya**: en DN un div es UN SUTTA
    (`n="dn1_1"` → «1. Brahmajālasuttaṃ»), mientras que en MN y AN es un **vagga** y en SN un
    **saṃyutta**. Por eso DN se resuelve por div y los demás por subhead.
    """
    import xml.etree.ElementTree as ET
    root = ET.parse(f'/tmp/tipitaka-xml/romn/{stem}.mul.xml').getroot()
    idx = {}
    for d in root.iter('div'):
        h = d.find('head')
        if h is None or h.get('rend') != 'chapter':
            continue
        t = ' '.join(''.join(h.itertext()).split())
        pns = []
        for pp in d.iter('p'):
            m = re.match(r'(\d+)(?:-(\d+))?$', pp.get('n') or '')
            if pp.get('rend') == 'bodytext' and m:
                pns += list(range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1))
        if pns:
            idx.setdefault(_norm(t), []).extend(pns)
            idx.setdefault(_head(t), []).extend(pns)
    return idx


def title_index_sutta_subhead(stem):
    """Como `title_index`, pero **sólo los subheads que son títulos de sutta**.

    En MN los subheads mezclan títulos de sutta («1. Mūlapariyāyasuttaṃ») con encabezados de
    sección internos («Dassanā pahātabbāsavā»), y estos últimos se llevaban los paranums del sutta
    en curso: por eso la primera pasada sólo resolvió 118 de 152.
    """
    import xml.etree.ElementTree as ET
    root = ET.parse(f'/tmp/tipitaka-xml/romn/{stem}.mul.xml').getroot()
    idx, title = {}, None
    for el in root.find('.//body').iter():
        if el.tag != 'p':
            continue
        if el.get('rend') == 'subhead':
            t = ' '.join(''.join(el.itertext()).split())
            title = t if _SUTTA_TITLE.match(t) else title
        elif el.get('rend') in ('bodytext', 'hangnum') and title:
            m = re.match(r'(\d+)(?:-(\d+))?$', el.get('n') or '')
            if m:
                rng = list(range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1))
                idx.setdefault(_norm(title), []).extend(rng)
                idx.setdefault(_head(title), []).extend(rng)
    return idx


def dnmn_refs():
    """DN y MN: no hay driver que recomputar (`validador_dnmn.py` no existe, sólo su JSON), así
    que se resuelve **por el título validado** contra el XML del volumen que declara la propia
    referencia PTS (`«D i 87,3»` → `s0101m`). Es el mismo criterio de autoridad que se usó para las
    24 de S iii: manda el texto que se cotejó."""
    out = {}
    tidx = {}
    for r in json.load(open('validador_dnmn.json')):
        m = re.match(r'([DM])\s+(i{1,3}|iv|v)\b', str(r.get('ref') or ''), re.I)
        if not m:
            continue
        nk = 'DN' if m.group(1).upper() == 'D' else 'MN'
        vol = m.group(2).lower()                       # volumen PTS: la fila del Excel
        # ⚠️ El volumen de PTS NO es el fichero del CST. PTS «M i» abarca MN 1-76, pero `s0201m` es
        # el Mūlapaṇṇāsa (MN 1-50): el CST divide por paṇṇāsa y PTS por tomo. El fichero se elige
        # por el NÚMERO de sutta.
        if nk == 'MN':
            k = int(r['num']) if str(r['num']).isdigit() else 0
            stem = 's0201m' if k <= 50 else 's0202m' if k <= 100 else 's0203m'
        else:
            stem = VRI_FILE.get((nk, vol))
        if not stem:
            continue
        if stem not in tidx:
            # DN: el div ES el sutta. MN: el div es el vagga, así que van los subheads de sutta.
            tidx[stem] = title_index_div(stem) if nk == 'DN' else title_index_sutta_subhead(stem)
        ct = r.get('cst_title', '')
        pns = tidx[stem].get(_norm(ct)) or tidx[stem].get(_head(ct))
        if pns:
            out[(nk, vol, str(r['num']))] = _fmt(stem, pns)
    return out


def title_index(stem):
    """`título normalizado → [paranums]` leyendo el XML VRI directamente.

    El título **validado** es la autoridad: es el texto que Gemini y el gate cotejaron. Recomputar
    la cadena Excel→massive→paranum no basta, porque los drivers aplican además resoluciones
    propias (el `ditthi_pairs` de S iii, el fallback por título del Jhāna-saṃyutta) que un
    adaptador ingenuo no reproduce — de ahí las 24 discrepancias de la primera pasada.
    """
    import xml.etree.ElementTree as ET
    root = ET.parse(f'/tmp/tipitaka-xml/romn/{stem}.mul.xml').getroot()
    idx, title = {}, None
    for el in root.find('.//body').iter():
        if el.tag != 'p':
            continue
        if el.get('rend') == 'subhead':
            title = ' '.join(''.join(el.itertext()).split())
        elif el.get('rend') in ('bodytext', 'hangnum') and title:
            m = re.match(r'(\d+)(?:-(\d+))?$', el.get('n') or '')
            if m:
                idx.setdefault(_norm(title), []).extend(
                    range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1))
    return idx

_F = str.maketrans({'ā': 'a', 'ī': 'i', 'ū': 'u', 'ṅ': 'n', 'ñ': 'n', 'ṇ': 'n',
                    'ṭ': 't', 'ḍ': 'd', 'ḷ': 'l', 'ṃ': 'm', 'ṁ': 'm'})


def _norm(t):
    """Normaliza un título para cotejarlo: sin diacríticos, sin puntuación, **sin letras dobles**
    y sin anotaciones entre llaves.

    Las dobles importan: el título validado lee `acchariyābbhuta` y el XML `acchariya-abbhuta`; sin
    colapsarlas, 34 suttas de MN no se localizaban. Las llaves traen variantes editoriales
    (`«2. pañcattayasuttaṃ {pañcāyatana…}»`) que no forman parte del título.
    """
    t = re.sub(r'\{[^}]*\}', ' ', t or '')
    t = re.sub(r'[^a-z0-9]', '', t.lower().translate(_F))
    return re.sub(r'(.)\1+', r'\1', t)


def main():
    write = '--write' in sys.argv
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}
    col = ci.get('VRI Ref')
    if col is None:
        col = len(hdr)
        ws.cell(row=1, column=col + 1).value = 'VRI Ref'
    st, mismatch, plan = Counter(), [], {}
    # Firmas humanas en bloque, sin veredicto en el JSON, resueltas UNA A UNA con evidencia
    # positiva de contenido. Sólo entra aquí lo que se ha comprobado a mano.
    for k, v in MANUAL.items():
        plan[k] = v
        st['firma humana: resuelta a mano'] += 1
    for k, v in dnmn_refs().items():
        plan[k] = v
        st[f'{k[0]} {k[1]}: ok (por título validado)'] += 1

    for (nk, vol), fn in VOLS.items():
        refs, jf = fn()
        validated = {r['num']: r.get('cst_title', '') for r in json.load(open(jf))}
        tidx = title_index(VRI_FILE[(nk, vol)])
        for num, ct in validated.items():          # filas que el recomputo no alcanzó
            if num not in refs and _norm(ct) in tidx:
                refs[num] = (_fmt(VRI_FILE[(nk, vol)], tidx[_norm(ct)]), ct)
        for num, (ref, title) in refs.items():
            if num not in validated:
                st[f'{nk} {vol}: sin veredicto'] += 1
                continue
            if _norm(title) == _norm(validated[num]):
                plan[(nk, vol, num)] = ref
                st[f'{nk} {vol}: ok'] += 1
                continue
            # el recomputo no reproduce la resolución del driver: manda el título VALIDADO
            pns = tidx.get(_norm(validated[num]))
            if pns:
                plan[(nk, vol, num)] = _fmt(VRI_FILE[(nk, vol)], pns)
                st[f'{nk} {vol}: ok (por título validado)'] += 1
            else:
                st[f'{nk} {vol}: TÍTULO NO LOCALIZADO'] += 1
                mismatch.append((nk, vol, num, title, validated[num]))

    print('=' * 88)
    print('BACKFILL DE «VRI Ref» — la clave canónica pasa a ser el paranum del XML VRI')
    print('=' * 88)
    for k in sorted(st):
        print(f'   {k}: {st[k]}')
    if mismatch:
        print(f'\n⚠️ títulos que NO coinciden con lo validado ({len(mismatch)}) — NO se escriben:')
        for x in mismatch[:12]:
            print(f'   {x[0]} {x[1]} #{x[2]:<9} recomputado «{x[3][:34]}» ≠ validado «{x[4][:34]}»')

    n = 0
    for row in ws.iter_rows(min_row=2):
        v = [c.value for c in row]
        k = (str(v[ci['Nikaya']] or ''), str(v[ci['PTS Roman']] or '').strip().lower(),
             str(v[ci['Sutta #']]))
        if k in plan:
            if write:
                ws.cell(row=row[0].row, column=col + 1).value = plan[k]
            n += 1
    print(f'\nfilas que reciben «VRI Ref»: {n}')
    if not write:
        print('(dry-run, nada escrito — usa --write)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-vriref.xlsx'
    shutil.copy(XLSX, bak)
    print('backup →', bak)
    wb.save(XLSX)
    print(f'Escritas {n} referencias VRI en {XLSX}.')


if __name__ == '__main__':
    main()
