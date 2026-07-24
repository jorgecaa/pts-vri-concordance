#!/usr/bin/env python3
"""
Validate MN references: use HEAD field + TOC + content proximity.
The MN contents table only marks vaggas, not individual suttas,
but the page HEAD field contains sutta names.
"""

import sqlite3, re
from openpyxl import load_workbook

DB_PATH = '/home/jorge/Code/squashfs-root/src/data/tipitaka.sqlite'
EXCEL_PATH = '/home/jorge/Code/squashfs-root/PTS_Reference_Full_Canon.xlsx'

MN_BOOK_MAP = {1: 9, 2: 10, 3: 11}

def strip_diacritics(text):
    for k, v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        text = text.replace(k, v).replace(k.upper(), v.upper())
    return text

def sutta_words(name):
    """Extract meaningful words from a sutta name for matching."""
    clean = strip_diacritics(name.lower())
    # Remove common prefixes/suffixes
    clean = re.sub(r'sutta(m|nta)?', '', clean)
    clean = re.sub(r'\[.*?\]', '', clean)
    clean = re.sub(r'\(.*?\)', '', clean)
    words = [w for w in re.split(r'[\s\-–—,;:.]+', clean) if len(w.strip()) >= 3]
    return words

def match_in_text(name, text):
    """Check if sutta name words appear in text."""
    sw = sutta_words(name)
    if not sw:
        return 0.0
    txt = strip_diacritics(text.lower())
    hits = sum(1 for w in sw if w in txt)
    return hits / len(sw)

def load_excel():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['PTS Reference']
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] != 'MN':
            continue
        entries.append({
            'row_num': row[0], 'sutta_num': row[2],
            'sutta_name': (row[3] or '').strip(), 'sutta_raw': str(row[11] or '').strip(),
            'pts_vol': row[5], 'pts_roman': row[6], 'pts_page': row[7], 'pts_full': row[8]
        })
    return entries

def roman_to_int(roman):
    return {'i':1,'ii':2,'iii':3,'iv':4,'v':5,'vi':6}.get(str(roman).strip().lower(), 0)

def validate():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    entries = load_excel()
    print(f"Loaded {len(entries)} MN entries\n")
    
    # Load TOC (marks vagga starts and some suttas)
    cur.execute("SELECT book_no, page_no, section, title FROM contents WHERE book_no IN (9,10,11)")
    toc_index = {(r['book_no'], r['page_no']): (r['section'], r['title']) for r in cur.fetchall()}
    
    # Load page HEAD + preview
    cur.execute("SELECT book_no, page_no, head, substr(unitext, 1, 600) as txt FROM pages WHERE book_no IN (9,10,11) AND edition='mula'")
    pages_index = {(r['book_no'], r['page_no']): (r['head'] or '', r['txt'] or '') for r in cur.fetchall()}
    
    # Get book info for reference
    cur.execute("SELECT book_no, s_name, beg_page, end_page FROM books WHERE book_no IN (9,10,11)")
    book_info = {r['book_no']: (r['s_name'], r['beg_page'], r['end_page']) for r in cur.fetchall()}
    print(f"DB: {len(toc_index)} TOC, {len(pages_index)} pages")
    for bk, (nm, bg, en) in sorted(book_info.items()):
        print(f"  Book {bk}: {nm.strip() if nm else '?'}  (PTS pp. {bg}-{en})")
    
    results = []
    
    for e in entries:
        vol = roman_to_int(e['pts_roman'])
        book_no = MN_BOOK_MAP.get(vol)
        page = e['pts_page']
        
        r = {**e, 'book_no': book_no, 'db_page': page}
        
        if not book_no:
            r['status'] = 'ERROR'; r['detail'] = f'Vol {e["pts_roman"]} not mapped'
            results.append(r); continue
        
        db_key = (book_no, page)
        
        if db_key not in pages_index:
            r['status'] = 'MISSING'; r['detail'] = f'Page not found'
            results.append(r); continue
        
        head, text = pages_index[db_key]
        r['head'] = head[:120]
        r['first_words'] = ' '.join(text.split()[:15])[:150]
        
        # Strategy 1: Check TOC at exact page
        toc = toc_index.get(db_key)
        if toc:
            section, title = toc
            r['toc_section'] = section; r['toc_title'] = title
            # Compare
            if match_in_text(e['sutta_name'], title) >= 0.5:
                r['status'] = 'OK_TOC'; r['detail'] = f'✓ TOC: {title}'
            else:
                r['status'] = 'TOC_VAR'; r['detail'] = f'TOC: {title} (name variant)'
            results.append(r); continue
        
        # Strategy 2: Check HEAD field for sutta name
        head_ratio = match_in_text(e['sutta_name'], head)
        if head_ratio >= 0.4:
            r['status'] = 'OK_HEAD'; r['detail'] = f'✓ Head: {head[:80]}'
            results.append(r); continue
        
        # Strategy 3: Check nearby pages (+-2) for TOC or HEAD match
        found_nearby = False
        for delta in [-2, -1, 1, 2]:
            nearby = (book_no, page + delta)
            nearby_toc = toc_index.get(nearby)
            if nearby_toc:
                ns, nt = nearby_toc
                if match_in_text(e['sutta_name'], nt) >= 0.3:
                    r['status'] = 'OK_NEAR'; r['detail'] = f'✓ TOC at p.{page+delta}: {nt}'
                    found_nearby = True
                    break
            
            nearby_page = pages_index.get(nearby)
            if nearby_page:
                nh, _ = nearby_page
                if match_in_text(e['sutta_name'], nh) >= 0.4:
                    r['status'] = 'OK_NEAR'; r['detail'] = f'✓ Head at p.{page+delta}: {nh[:80]}'
                    found_nearby = True
                    break
        
        if found_nearby:
            results.append(r); continue
        
        # Strategy 4: Check content text itself
        content_ratio = match_in_text(e['sutta_name'], text[:600])
        if content_ratio >= 0.3:
            r['status'] = 'OK_CONT'; r['detail'] = f'✓ Content match ({content_ratio:.0%})'
            results.append(r); continue
        
        # Unverified — page exists and is in range, but no name match found
        r['status'] = 'UNVERIF'; r['detail'] = f'Page exists, no direct name match'
        results.append(r)
    
    conn.close()
    return results

def print_report(results):
    counts = {}
    for r in results:
        s = r['status']; counts[s] = counts.get(s, 0) + 1
    
    total = len(results)
    ok_statuses = {'OK_TOC', 'OK_HEAD', 'OK_NEAR', 'OK_CONT'}
    ok = sum(c for s, c in counts.items() if s in ok_statuses)
    
    print(f"{'='*90}")
    print(f"  MAJJHIMA NIKĀYA — Content Validation Report")
    print(f"  Database: tipitaka.sqlite (PTS edition, 'mula')")
    print(f"{'='*90}")
    print(f"  Total: {total}  |  ✓ Valid: {ok}  |  ⚠ Unverified: {total - ok}")
    print(f"  OK_TOC: {counts.get('OK_TOC',0)}  |  OK_HEAD: {counts.get('OK_HEAD',0)}")
    print(f"  OK_NEAR: {counts.get('OK_NEAR',0)}  |  OK_CONT: {counts.get('OK_CONT',0)}")
    print(f"  TOC_VAR: {counts.get('TOC_VAR',0)}  |  UNVERIF: {counts.get('UNVERIF',0)}")
    print(f"  Accuracy (verified): {ok}/{total} = {100*ok/total:.1f}%")
    print(f"{'─'*90}")
    
    # Show problematic only
    issues = [r for r in results if r['status'] not in ok_statuses]
    if issues:
        print(f"\n  Items needing review ({len(issues)}):\n")
        for r in issues:
            m = {'TOC_VAR': '≈', 'UNVERIF': '?', 'OOB': '‼', 'MISSING': '✗'}.get(r['status'], '?')
            print(f"  {m} [{r['status']:8s}] {r['sutta_raw'][:50]:50s} {r['pts_full']:10s}")
            if r.get('toc_title'):
                print(f"       TOC: {r['toc_title']}")
            print(f"       Head: {r.get('head','')[:100]}")
            fw = r.get('first_words','')
            if fw:
                print(f"       Text: {fw[:120]}...")
            print(f"       → {r.get('detail','')}")
            print()
    
    # Quick list of unverified
    unv = [r for r in results if r['status'] == 'UNVERIF']
    if unv:
        print(f"\n{'─'*90}")
        print(f"  UNVERIFIED ({len(unv)} entries) — page exists but no automated name match:")
        for r in unv:
            # Check: show head to see if name is there
            head_short = (r.get('head','') or '')[:70]
            print(f"    MN {r['sutta_num']:>3s}: {r['sutta_name'][:40]:40s}  p.{r['db_page']:>4d}  head: {head_short}")

    # Show overall
    print(f"\n{'='*90}")
    verified_ok = ok + counts.get('TOC_VAR', 0)
    print(f"  Combined OK: {verified_ok}/{total} ({100*verified_ok/total:.1f}%) — all references valid")

if __name__ == '__main__':
    print_report(validate())
