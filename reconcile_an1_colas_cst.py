#!/usr/bin/env python3
"""
reconcile_an1_colas_cst — ancla las colas del Duka y del Tika **al tramo del CST que les
corresponde**, por decisión de Jorge (2026-07-25).

El problema era que la numeración del Excel **excede a la del CST**: su Duka llega a `2.479` y
`s0402m1` acaba en el paranum **246**; su Tika llega a `3.352` y `s0402m2` en **184**. Los paranums
`2.280`, `2.310` y los del tramo alto del Tika no existen, así que la `VRI Ref` aritmética era
imposible.

La decisión —anclar cada fila al tramo que le corresponde aunque el rango del Excel sea más ancho—
resulta **inmediata de verificar**, porque en estas colas el nombre de la fila **es el nombre del
`div`**:

| fila del Excel | `div` del CST | paranums |
|---|---|--:|
| `2.180-229` *Kodha [Kodhapeyyāla]* | `1. Kodhapeyyālaṃ` | 181-190 |
| `2.230-279` *Akusalapeyyāla* | `2. Akusalapeyyālaṃ` | 191-200 |
| `2.280-309` *Vinayapeyyāla* | `3. Vinayapeyyālaṃ` | 201-230 |
| `2.310-479` *Rāgapeyyāla* | `4. Rāgapeyyālaṃ` | 231-246 |
| `3.163-182` *Kammapathapeyyāla* | `7. Kammapathapeyyālaṃ` | 164-183 |
| `3.183-352` *Rāgapeyyāla* | `8. Rāgapeyyālaṃ` | 184 |

No hay que decidir *cuál* es la contraparte: la fila y el `div` se llaman igual. Y una de ellas
corrobora de propina: `3.163-182` declara **20** suttas y el `Kammapathapeyyālaṃ` tiene **20**
paranums.

### El Acelaka-vagga, que se resuelve de paso

`3.156` estaba en suspenso porque el cotejo empataba entre los paranums 162 y 163, y porque **Morris
declara ahí una conjetura** («The Acelaka-vagga **probably** included only suttas 151, 152…»). Pero
su duda es sobre **cuántos vaggas** forma ese material, no sobre qué texto es cada fila, y eso lo
fija la aritmética sin tocar su conjetura:

    `3.155` está cerrado en el paranum 156, el último del `Maṅgalavaggo`
    el `Acelakavaggo` es el `div` 157-163: **siete** paranums
    `3.156` (1) + `3.157-162` (6) = **siete** filas

Sólo hay un reparto: `3.156` → 157 y `3.157-162` → 158-163.

Uso: python3 reconcile_an1_colas_cst.py [--dry]
"""
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'

ANCLAS = {
    # el nombre de la fila ES el nombre del `div` del CST
    '2.180-229': ('s0402m1:181-190', 'el `div` «1. Kodhapeyyālaṃ» del CST: la fila y el div se '
                                     'llaman igual. El Excel numera 50 donde el CST tiene 10'),
    '2.230-279': ('s0402m1:191-200', 'el `div` «2. Akusalapeyyālaṃ»; el Excel numera 50 y el CST 10'),
    '2.280-309': ('s0402m1:201-230', 'el `div` «3. Vinayapeyyālaṃ»; el Excel numera 30 y el CST 30'),
    '2.310-479': ('s0402m1:231-246', 'el `div` «4. Rāgapeyyālaṃ», con el que acaba `s0402m1`; el '
                                     'Excel numera 170 y el CST 16'),
    '3.163-182': ('s0402m2:164-183', 'el `div` «7. Kammapathapeyyālaṃ»: la fila declara 20 suttas '
                                     'y el div tiene 20 paranums'),
    '3.183-352': ('s0402m2:184', 'el `div` «8. Rāgapeyyālaṃ», con el que acaba `s0402m2` (un solo '
                                 'paranum); el Excel numera 170'),
    # el Acelaka-vagga, forzado por la aritmética sin tocar la conjetura de Morris
    '3.156': ('s0402m2:157', '`3.155` cierra en el paranum 156 (último del Maṅgalavaggo) y el '
                             'Acelakavaggo es el div 157-163, siete paranums para las siete filas '
                             '`3.156` + `3.157-162`: sólo hay un reparto'),
    '3.157-162': ('s0402m2:158-163', 'ídem: las seis que siguen a `3.156` dentro del Acelakavaggo'),
}


def main():
    dry = '--dry' in sys.argv
    if not dry:
        cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-colascst.xlsx'
        shutil.copy(XLSX, cp)
        print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    filas = [r for r in range(2, ws.max_row + 1)
             if ws.cell(r, ci['Nikaya']).value == 'AN'
             and str(ws.cell(r, ci['PTS Roman']).value or '').strip() == 'i']
    for r in filas:
        s = str(ws.cell(r, ci['Sutta #']).value)
        if s not in ANCLAS or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, motivo = ANCLAS[s]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = ('anclada al tramo del CST (decisión de Jorge): ' + motivo)
        print(f'   {s:>12} → {vri}')
        n += 1
    pend = sum(1 for r in filas if ws.cell(r, ci['Estado']).value != 'CONFIRMADO')
    print(f'{n} filas ancladas · A i: {len(filas) - pend}/{len(filas)} CONFIRMADO')
    if dry:
        print('(dry-run: nada escrito)')
        return
    wb.save(XLSX)
    print(f'escrito → {XLSX}')


if __name__ == '__main__':
    main()
