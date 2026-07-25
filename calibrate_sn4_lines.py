#!/usr/bin/env python3
"""calibrate_sn4_lines.py — calibra el nº de LÍNEA de las referencias PTS de SN IV (S iv).

Mismo criterio que en S i, S iii y S v: la línea de `PTS Ref` (`S iv <pág>,<línea>`) es la del
**marcador del sutta** en la BD (libro 15), y se recalcula con **la misma asignación que usa el
validador** (`validador_sn4.assign_volume`, inyectiva y monótona, hueco 0). Repetir aquí una
cadena de heurísticas por fila no sería inyectivo: dos filas podrían apuntar al mismo marcador y
una recibiría la línea de su vecina.

Se reescribe **sólo la LÍNEA**, y sólo cuando el marcador está en la página que declara el Excel.
Los desacuerdos de página se listan aparte para arbitraje: reescribir la página es una afirmación
mayor que reescribir la línea.

Nota: varias filas comparten legítimamente marcador y, por tanto, página y línea — Feer imprime a
veces dos suttas del CST bajo un solo encabezado («Agayha» = los dos Rūpārāma, «Pubbeñāṇam»,
«Suddhikaṃ nirāmisam»). Se anotan aparte, no son un defecto.

Uso: python3 calibrate_sn4_lines.py [--write]
"""
import datetime
import re
import shutil
import sqlite3
import sys
from collections import Counter

from openpyxl import load_workbook

from validador_sn4 import (DB, XLSX, assign_volume, build_cst_blocks, build_pts_suttas,
                           cst_for, excel_entries)

REF_RE = re.compile(r'^(S\s+iv\s+)(\d+)(?:\s*,\s*(\d+))?\s*$')


def main():
    write = '--write' in sys.argv
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
    entries = excel_entries()
    cstmap = cst_for(entries, build_cst_blocks())
    recs = build_pts_suttas(conn.cursor())
    assigned = assign_volume(entries, recs, cstmap)

    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}

    shared = Counter()
    for num, q in assigned.items():
        if q:
            shared[(q['page'], q['line'])] += 1

    delta, fixes, otherpage, unresolved = Counter(), [], [], []
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'SN' or str(v[ci['PTS Roman']] or '').strip().lower() != 'iv':
            continue
        num = str(v[ci['Sutta #']])
        q = assigned.get(num)
        ref = str(v[ci['PTS Ref']] or '')
        if not q:
            unresolved.append(num)
            continue
        if q['page'] != v[ci['PTS Page']]:
            otherpage.append((num, ref, f"S iv {q['page']},{q['line']}"))
            continue
        m = REF_RE.match(ref)
        old = int(m.group(3)) if (m and m.group(3)) else None
        delta[None if old is None else old - q['line']] += 1
        new_ref = f"S iv {q['page']},{q['line']}"
        if ref != new_ref:
            fixes.append((ridx, num, ref, new_ref))

    print('=' * 84)
    print('CALIBRACIÓN DE LÍNEA — SN IV (marcador real en la BD, libro 15)')
    print('=' * 84)
    print('desfase (línea del Excel − línea del marcador), por frecuencia:')
    for d, n in delta.most_common(12):
        print(f'  {"sin línea" if d is None else f"{d:+d}":>10} : {n}')
    exact, tot = delta.get(0, 0), sum(delta.values())
    print(f'\ncoincidencia exacta: {exact}/{tot} ({100 * exact / max(1, tot):.0f}%)')
    print(f'a corregir: {len(fixes)} | página en disputa (NO se tocan): {len(otherpage)} '
          f'| sin marcador: {len(unresolved)}')
    for ridx, num, o, n in fixes[:20]:
        print(f'  SN {num:>7}: {o!r} → {n!r}')
    if otherpage:
        print('\npágina en disputa (Excel dice una, el marcador está en otra):')
        for num, o, n in otherpage:
            print(f'  SN {num:>7}: {o!r} vs marcador {n!r}')
    multi = {k: n for k, n in shared.items() if n > 1}
    if multi:
        print(f'\nmarcadores compartidos por varias filas (Feer imprime dos suttas juntos): '
              f'{len(multi)}')

    if not write:
        print('\n(dry-run, nada escrito — usa --write)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-sn4lineas.xlsx'
    shutil.copy(XLSX, bak)
    print('\nbackup →', bak)
    for ridx, _num, _o, new_ref in fixes:
        ws.cell(row=ridx, column=ci['PTS Ref'] + 1).value = new_ref
    wb.save(XLSX)
    print(f'Corregidas {len(fixes)} líneas en {XLSX}.')


if __name__ == '__main__':
    main()
