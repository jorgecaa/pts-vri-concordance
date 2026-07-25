#!/usr/bin/env python3
"""reconcile_sn5.py — vuelca validador_sn5_vri.json al Excel maestro (SN V).

Regla del proyecto: CONFIRMADO = concordancia (VRI) ∧ Gemini APPROVE → Validation=VALIDADOR.
Desacuerdo (Gemini REJECT) → queda para revisión humana (no se auto-firma). Este script
SOLO escribe las filas VALIDADOR salvo que se pase --humano MAP.json con dispensas
{canon: VALIDACION} firmadas por Jorge. Siempre hace backup antes de escribir.

Uso:
    python3 reconcile_sn5.py --dry               # resumen por disposición, no escribe
    python3 reconcile_sn5.py --write             # escribe solo VALIDADOR (auto)
    python3 reconcile_sn5.py --write --humano firmas.json  # + dispensas humanas
"""
import json, re, sys, shutil, datetime
from collections import Counter
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
RES = 'validador_sn5_vri.json'


def load_results():
    by_canon = {}
    for r in json.load(open(RES)):
        by_canon[r['canon']] = r
    return by_canon


def sn5_rows(ws, ci):
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'SN' or str(v[ci['PTS Roman']] or '').strip().lower() != 'v':
            continue
        name = str(v[ci['Sutta Name']] or '')
        m = re.search(r'\(SN\s+(\d+)\.(\d+)', name)
        canon = f'{m.group(1)}.{m.group(2)}' if m else str(v[ci['Sutta #']])
        yield ridx, canon, row


def main():
    dry = '--dry' in sys.argv or '--write' not in sys.argv
    humano = {}
    if '--humano' in sys.argv:
        humano = json.load(open(sys.argv[sys.argv.index('--humano') + 1]))
    res = load_results()
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}
    vcol, ecol = ci['Validation'] + 1, ci['Estado'] + 1

    plan = Counter()
    writes = []
    for ridx, canon, row in sn5_rows(ws, ci):
        if canon in humano:
            val = humano[canon]; est = 'CONFIRMADO' if val not in ('REF_ERROR_DPR', 'REVISAR') else 'PENDIENTE'
            plan[f'humano:{val}'] += 1; writes.append((ridx, val, est)); continue
        cur_val = str(ws.cell(row=ridx, column=vcol).value or '')
        if cur_val in ('VALIDADOR_HUMANO', 'PTS_CROSSREF_SN'):
            plan['protegido(humano/crossref)'] += 1; continue     # no degradar procedencia
        r = res.get(canon)
        if not r:
            plan['sin_resultado'] += 1; continue
        if r['gemini'] == 'APPROVE':          # concordancia (VRI) ∧ Gemini APPROVE
            plan['VALIDADOR(auto)'] += 1; writes.append((ridx, 'VALIDADOR', 'CONFIRMADO'))
        else:
            plan[f"arbitraje:{r.get('src')}"] += 1  # no auto-escribe (peyyāla elidido)

    print('Plan de escritura:', dict(plan))
    print('Filas a escribir:', len(writes))
    if dry:
        print('(dry-run, nada escrito)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}.xlsx'
    shutil.copy(XLSX, bak); print('backup →', bak)
    for ridx, val, est in writes:
        ws.cell(row=ridx, column=vcol).value = val
        ws.cell(row=ridx, column=ecol).value = est
    wb.save(XLSX)
    print(f'Escritas {len(writes)} filas en {XLSX}.')


if __name__ == '__main__':
    main()
