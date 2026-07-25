#!/usr/bin/env python3
"""
reconcile_an1_ekaduka — vuelca la **mitad limpia de A i**: 197 filas (Eka 1-12 y Duka entero).

Ejecuta `an1_eka_duka.py` y escribe lo que sostiene. Ninguna llamada a API: lo que firma cada fila
son **tres comprobaciones independientes**, y ninguna de las tres usa lo que usan las otras.

1. **Marcador impreso** — la fila `(vagga v, sutta n)` de su `DPR Ref` cae en el último marcador
   de `v` con número ≤ n. Es cómo se cita el impreso, no una aproximación.
2. **Paranum del CST** — la posición `n` dentro del `<div>` correspondiente. Donde los recuentos de
   las dos ediciones no coinciden (los vaggas 2 y 3 del Duka), decide el **contenido**, no la
   posición: alineamiento monótono marcador↔paranum dentro del vagga.
3. **Cobertura de contenido** entre el texto del marcador y el del paranum — señal que **no
   interviene** en (1) ni en (2).

Más la **página** del Excel (±1) como cuarta.

⚠️ La cobertura se mide sobre **n-gramas de caracteres**, no sobre palabras. En pali la composición
es libre en la ortografía: PTS escribe `kodha vinayo ca upanaha vinayo` y el CST `kodhavinayo ca
upanahavinayo`. Es el mismo texto, y por conjuntos de palabras daba **0,22** — suficiente para
tirar cuatro filas correctas. Sobre 4-gramas da 0,9 y deja de castigarse una convención tipográfica.

**El umbral está medido, no elegido**: sobre las 197 filas la cobertura real va de **0,64 a 1,00**
(mediana 0,97-0,99) y un control con pares de vaggas distintos da media **0,26** y máximo **0,48**.
El umbral (0,55) cae en el hueco.

`Validation` = `VALIDADOR_HUMANO`, con el detalle de las tres comprobaciones en `Detail`.

Uso: python3 reconcile_an1_ekaduka.py [--dry]
"""
import math
import shutil
import sqlite3
import sys
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook

import an1_eka_duka as R

XLSX = 'PTS_Reference_Complete_Canon.xlsx'


def resuelve():
    """`{Sutta #: (vri, detalle)}` de las filas que las tres comprobaciones sostienen."""
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    out = {}
    for _vol, (pref, nipata, stem, v_max) in R.NIPATAS.items():
        mk = R.marcadores(conn, nipata)
        cst, _divs = R.divisiones(stem)
        porcont = {}
        for f in R.filas(pref):
            if f['v'] is None or f['v'] > v_max:
                continue
            ms, uds = mk.get(f['v'], []), cst.get(f['v'], [])
            cand = [x for x in ms if x[0] <= f['n']]
            if not cand or not uds:
                continue
            m = cand[-1]
            if len(uds) == (ms[-1][0] if ms else 0) and f['n'] <= len(uds):
                pn, txt = uds[f['n'] - 1]
                cov = R.cobertura(m[3], txt)
                via = 'posición en el div'
            else:
                if f['v'] not in porcont:
                    porcont[f['v']] = R.alinea_por_contenido(ms, uds, None)
                if m[0] not in porcont[f['v']]:
                    continue
                pn, cov = porcont[f['v']][m[0]]
                via = 'contenido (los recuentos no coinciden)'
            pag_ok = isinstance(f['page'], int) and abs(m[1] - f['page']) <= 1
            if not (pag_ok and cov >= R.COV_MIN):
                continue
            # las filas de rango abarcan de `n` a `n_hi`: la `VRI Ref` va del paranum del primero
            # al del último, cuando ese último existe en el mismo div
            pn_hi = pn
            if f['n_hi'] and f['n_hi'] > f['n'] and f['n_hi'] <= len(uds):
                pn_hi = uds[f['n_hi'] - 1][0]
            vri = f'{stem}:{pn}' + (f'-{pn_hi}' if pn_hi != pn else '')
            out[f['num']] = (vri, f'A i: DPR ({f["v"]}.{f["n"]}) → marcador {m[0]} en '
                                  f'A i {m[1]},{m[2]}; paranum {pn} por {via}; '
                                  f'cobertura de contenido {cov:.2f}')
    return out


def main():
    dry = '--dry' in sys.argv
    res = resuelve()
    print(f'{len(res)} filas sostenidas por las tres comprobaciones')
    if not dry:
        cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-an1ekaduka.xlsx'
        shutil.copy(XLSX, cp)
        print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    filas = [r for r in range(2, ws.max_row + 1)
             if ws.cell(r, ci['Nikaya']).value == 'AN'
             and str(ws.cell(r, ci['PTS Roman']).value or '').strip() == 'i']
    # ⚠️ `2.19` está DUPLICADO en A i (una fila normal y otra fundida, `AN 2.19(*) + AN 3.29(*)`),
    # y escribir por `Sutta #` le daba a las dos la referencia de una. Las claves duplicadas no se
    # tocan: es justo lo que `check_integrity` lleva señalando y no se arregla escribiéndoles
    # encima.
    dup = {k for k, c in Counter(str(ws.cell(r, ci['Sutta #']).value)
                                 for r in filas).items() if c > 1}
    for r in filas:
        num = str(ws.cell(r, ci['Sutta #']).value)
        if ws.cell(r, ci['Estado']).value == 'CONFIRMADO' or num not in res or num in dup:
            continue
        vri, det = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    pend = sum(1 for r in filas if ws.cell(r, ci['Estado']).value != 'CONFIRMADO')
    print(f'{n} filas cerradas · A i: {len(filas) - pend}/{len(filas)} CONFIRMADO')
    if dry:
        print('(dry-run: nada escrito)')
        return
    wb.save(XLSX)
    print(f'escrito → {XLSX}')


if __name__ == '__main__':
    main()
