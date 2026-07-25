#!/usr/bin/env python3
"""
fix_kilesa_an3 — corrige los dos nombres de kilesa mal puestos en el *rāga-peyyāla* de **A iii**.

Mismo defecto que ya se corrigió en A ii (`reconcile_an2_cierre.py`) y con la misma prueba: el
impreso (A iii 277-278) y el CST (`s0403m1` n=308-1151) enumeran los **dieciséis kilesa en el mismo
orden**, terminando `…mānassa atimānassa **madassa pamādassa**`. El Excel se saltó `mada` y repitió
`macchariya` al final, de modo que sus dos últimas filas van corridas una posición.

La aritmética dice cuál es cuál sin margen: el Pañcaka reparte **50 paranums por kilesa** desde
353, así que `1053-1102` es el decimoquinto (*mada*) y `1103-1152` el decimosexto (*pamāda*).

Verificado antes de escribir con `an_peyyala.serie_monotona`: con los nombres corregidos la serie
se sitúa **16/16 en PTS y 16/16 en el CST**, en orden estricto. Con los del Excel se rompía en el
último término, en las dos ediciones a la vez — que es justo la señal de que el fallo estaba en el
nombre y no en el texto.

Sólo toca `Sutta Name`. Las filas siguen PENDIENTE: su cierre va con el barrido de A iii.
"""
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
CORRECCIONES = {                      # (Sutta #, nombre actual) → nombre correcto
    ('5.1053-1102', 'Pamāda peyyāla'): 'Mada peyyāla',
    ('5.1103-1152', 'Macchariya peyyāla'): 'Pamāda peyyāla',
    # Tercer nombre del mismo volumen desmentido por el texto: las DOS ediciones escriben
    # `Sāṭiya-` —PTS `sāṭiyaggāhāpako` (A iii 275), CST `sāṭiyagāhāpako` (grupo 273-285)— y sólo
    # el Excel escribe `Sāṭika-`. Con el nombre corregido la fila se sitúa y se cierra.
    ('5.282', 'Sāṭikagāhāpaka'): 'Sāṭiyagāhāpaka',
}


def main():
    dry = '--dry' in sys.argv
    if not dry:
        cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-kilesa-an3.xlsx'
        shutil.copy(XLSX, cp)
        print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    hechas = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'AN':
            continue
        if str(ws.cell(r, ci['PTS Roman']).value or '').strip() != 'iii':
            continue
        clave = (str(ws.cell(r, ci['Sutta #']).value), str(ws.cell(r, ci['Sutta Name']).value))
        if clave in CORRECCIONES:
            print(f'  {clave[0]:>12}  «{clave[1]}» → «{CORRECCIONES[clave]}»')
            ws.cell(r, ci['Sutta Name']).value = CORRECCIONES[clave]
            hechas += 1
    print(f'\n{hechas}/{len(CORRECCIONES)} corregidas')
    if dry:
        print('(dry-run: nada escrito)')
        return
    if hechas:
        wb.save(XLSX)
        print(f'escrito → {XLSX}')


if __name__ == '__main__':
    main()
