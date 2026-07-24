#!/usr/bin/env python3
"""SN all volumes — final line numbers with corrected patterns."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

SN_MAP = {1:12, 2:13, 3:14, 4:15, 5:16}
VOL_LETTER = {1:'i',2:'ii',3:'iii',4:'iv',5:'v'}

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1,column=c).value: c for c in range(1, ws.max_column+1)}

def find_marker(book_no, page_no, sutta_num_str):
    parts = sutta_num_str.split('.')
    if len(parts) < 2: return None, None, None
    sid = parts[1]
    
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None, None, None
    lines = (r['unitext'] or '').split(chr(10))
    
    esc = re.escape(sid)
    
    for i, line in enumerate(lines):
        s = line.strip()
        if not s: continue
        
        # "56. (4) Name" — with dot (SN V)
        m = re.match(r'^\s*' + esc + r'\.\s*\(\d+\)\s*(.*)', s)
        if m: return i+1, m.group(1).strip(), 'std_dot'
        
        # "56 (4) Name" — without dot (SN II-IV)
        m = re.match(r'^\s*' + esc + r'\s*\(\d+\)\s*(.*)', s)
        if m: return i+1, m.group(1).strip(), 'std'
        
        # Range "140-142 (5-7) Name"
        m = re.match(r'^\s*(\d+)\-(\d+)\s*\(\d+\-\d+\)\s*(.*)', s)
        if m:
            try:
                lo, hi = int(m.group(1)), int(m.group(2))
                if lo <= int(sid) <= hi:
                    return i+1, m.group(3).strip(), 'range'
            except: pass
        
        # "§ N. Name" — SN I section markers
        m = re.match(r'^\s*§\s*' + esc + r'[\.\s]+(.*)', s)
        if m: return i+1, m.group(1).strip(), 'section'
        
        # Simple "N Name" (peyyala)
        m = re.match(r'^\s*' + esc + r'\s+(\S.+)', s)
        if m:
            name = m.group(1).strip()
            skip_words = ['savatthi', 'evam', 'atha', 'tatra', 'bhagava', 'saddha']
            if not any(name.lower().startswith(w) for w in skip_words):
                if not re.match(r'^\d', name):
                    return i+1, name, 'simple'
    
    return None, None, None

stats = {}
total_updates = 0

for vol_num in [1, 2, 4, 5]:
    book_no = SN_MAP[vol_num]
    roman = VOL_LETTER[vol_num]
    
    vol_stats = {'std_dot':0, 'std':0, 'simple':0, 'range':0, 'section':0, 'not_found':0}
    count = 0
    
    for ri in range(2, ws.max_row + 1):
        if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
        if str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip() != roman: continue
        
        sutta_num = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
        page = ws.cell(row=ri, column=cols['PTS Page']).value
        ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
        if not page: continue
        
        count += 1
        
        # Skip if already has comma (except SN III which was done separately)
        if ',' in ref: continue
        
        line, mname, mtype = find_marker(book_no, page, sutta_num)
        
        if line:
            new_ref = 'S %s %d' % (roman, page)
            if line > 1: new_ref += ',%d' % line
            if new_ref != ref:
                ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
                total_updates += 1
            vol_stats[mtype] = vol_stats.get(mtype, 0) + 1
        else:
            vol_stats['not_found'] += 1
    
    stats[roman] = vol_stats

wb.save(XL)

# Report
for roman in ['i', 'ii', 'iv', 'v']:
    s = stats[roman]
    total = sum(s.values())
    found = total - s.get('not_found', 0)
    parts = []
    for k in ['std_dot','std','simple','range','section','not_found']:
        if s.get(k,0): parts.append('%s=%d' % (k, s.get(k,0)))
    print('SN %s: %d/%d  %s' % (roman.upper(), found, total, ' | '.join(parts)))

print()
print('Total PTS Ref updates: %d' % total_updates)
print('Saved: %s' % XL)
conn.close()
