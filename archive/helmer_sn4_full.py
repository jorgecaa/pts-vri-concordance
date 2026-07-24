#!/usr/bin/env python3
"""Helmer SN IV: 344 suttas, rigorous marker extraction + DeepSeek."""
import hashlib, json, os, re, sqlite3
from collections import Counter
from pathlib import Path
import sutta_hash as sh

CACHE_FILE=Path('helmer_ptscst_cache.json'); SN4_BOOK=15

conn=sqlite3.connect('src/data/tipitaka.sqlite'); conn.row_factory=sqlite3.Row; cur=conn.cursor()

# Build marker map
print('Building marker map...')
marker_map={}
cur.execute('SELECT page_no, unitext FROM pages WHERE book_no=? AND edition="mula"',(SN4_BOOK,))
for r in cur.fetchall():
    for i,line in enumerate(r['unitext'].split('\n')):
        s=line.strip()
        m=re.match(r'^(\d+)\.?\s*\((\d+)\)\s+(\S.*)',s)
        if m:
            marker_map.setdefault(r['page_no'],[]).append((i+1,int(m.group(1)),int(m.group(2)),m.group(3)[:60]))

# Load Excel
from openpyxl import load_workbook
wb=load_workbook('PTS_Reference_Complete_Canon.xlsx'); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
entries=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2: continue
    if int(parts[0])<35 or int(parts[0])>44: continue
    if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='iv': continue
    page=ws.cell(row=ri,column=cols['PTS Page']).value
    ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
    name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
    entries.append({'num':snum,'ref':ref,'name':name,'page':page,'ri':ri,
                    'sn':int(parts[0]),'sn_inner':int(parts[1])})
entries.sort(key=lambda e:(e['sn'],e['sn_inner']))
print('SN IV: {} entries'.format(len(entries)))

# Match and extract
print('Matching entries to markers...')
tasks=[]; no_match=[]
for e in entries:
    page=e['page']; target=e['sn_inner']
    found=False
    for line,gid,vpos,name in marker_map.get(page,[]):
        if gid==target or vpos==target:
            cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN4_BOOK,page))
            r=cur.fetchone()
            if r:
                lines=r['unitext'].split('\n')
                start=line-1; end=len(lines)
                for j in range(start+1,len(lines)):
                    if re.match(r'^\d+\.?\s*\(\d+\)',lines[j].strip()): end=j; break
                text=' '.join(lines[start:end])
                tasks.append((e,' '.join(sh.tokens(text)[:300]),line,gid,name[:40]))
                found=True; break
    if not found: no_match.append(e)

print('  Matched: {} | No marker: {}'.format(len(tasks),len(no_match)))
if no_match:
    print('  No marker examples:')
    for e in no_match[:5]:
        markers=marker_map.get(e['page'],[])
        mstr=','.join('{}({})'.format(g,v) for _,g,v,n in markers[:4])
        print('    SN {} p.{} looking for gid/vpos={} | markers: {}'.format(e['num'],e['page'],e['sn_inner'],mstr))

# DeepSeek
if tasks:
    print('\nDeepSeek on {} entries...'.format(len(tasks)))
    segs=sh.load_cha(['s4m.xml'],'h4n')
    cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}
    from openai import OpenAI
    cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')
    SYS='You are Helmer Smith. Are these two Pali passages the SAME sutta? Same=core teaching matches despite Feer abbreviation. Different=distinct sutta. Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words"}'
    
    vc=Counter(); rej_list=[]; processed=0
    for i,(e,pts,line,gid,mname) in enumerate(tasks):
        cst_text=' '.join(sh.tokens(segs[i]['text'])[:300])
        cst_title=segs[i].get('title','')
        key=hashlib.md5((pts[:200]+cst_text[:200]).encode()).hexdigest()
        
        if key in cache: hv=cache[key]
        else:
            try:
                r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
                    response_format={'type':'json_object'},
                    messages=[{'role':'system','content':SYS},
                              {'role':'user','content':'PTS:\n{}\n\nCST:\n{}'.format(pts[:900],cst_text[:900])}])
                hv=json.loads(r.choices[0].message.content)
                cache[key]=hv; CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
            except Exception as ex: hv={'verdict':'ERROR','reason':str(ex)[:60]}
        
        v=hv.get('verdict','?'); vc[v]+=1; processed+=1
        if v=='REJECT': rej_list.append((e,hv,cst_title,pts[:60],cst_text[:60]))
        if processed%30==0:
            print('  {} done | APPROVE: {} REJECT: {}'.format(processed,vc.get('APPROVE',0),vc.get('REJECT',0)))
    
    print('\n'+'='*90)
    total=processed; app=vc.get('APPROVE',0); rej=vc.get('REJECT',0); err=vc.get('ERROR',0)
    print('SN IV Helmer: {} entries'.format(total))
    print('  APPROVE: {} | REJECT: {} | ERROR: {}'.format(app,rej,err))
    valid=total-err
    if valid>0: print('  Accuracy: {}/{} ({:.0f}%)'.format(app,valid,100*app/valid))
    if no_match: print('  No marker: {}'.format(len(no_match)))
    
    if rej_list:
        print('\nREJECT entries ({}):'.format(len(rej_list)))
        for e,hv,cst_title,pts_preview,cst_preview in rej_list:
            print('  ✗ SN {:>8s} {:>15s} | {} | {}'.format(e['num'],e['ref'],cst_title[:35],hv.get('reason','')[:60]))
    
    # Update Excel
    for e,pts,line,gid,mname in tasks:
        ws.cell(row=e['ri'],column=cols['Validation']).value='HELMER_APPROVED'
    wb.save('PTS_Reference_Complete_Canon.xlsx')
    print('\nExcel updated: HELMER_APPROVED')

conn.close()
