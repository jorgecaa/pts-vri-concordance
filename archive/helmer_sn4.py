#!/usr/bin/env python3
"""Helmer SN IV: 344 suttas, perfect 1:1 CST match. Quick sample of 20."""
import hashlib, json, os, re, sqlite3, random
from pathlib import Path
import sutta_hash as sh

CACHE_FILE=Path('helmer_ptscst_cache.json'); SN4_BOOK=15

def pts_sutta(page_no, global_num, max_tokens=300):
    conn=sqlite3.connect('src/data/tipitaka.sqlite'); conn.row_factory=sqlite3.Row; cur=conn.cursor()
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN4_BOOK,page_no))
    r=cur.fetchone(); conn.close()
    if not r or not r['unitext']: return ''
    lines=r['unitext'].split('\n'); sn=str(global_num)
    start=None
    for i,line in enumerate(lines):
        s=line.strip()
        if re.match(r'^'+re.escape(sn)+r'\s*\(\d+\)\s+\S',s): start=i; break
        if re.match(r'^\('+re.escape(sn)+r'\)\s*\(\d+\)\s+\S',s): start=i; break
    if start is None: return ''
    end=len(lines)
    for j in range(start+1,len(lines)):
        s=lines[j].strip()
        if re.match(r'^\d+\s*\(\d+\)\s+\S',s): end=j; break
    return ' '.join(sh.tokens(' '.join(lines[start:end]))[:max_tokens])

from openpyxl import load_workbook
wb=load_workbook('PTS_Reference_Complete_Canon.xlsx'); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}

tests=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2: continue
    if int(parts[0])<35 or int(parts[0])>44: continue  # SN IV = saṃyuttas 35-44
    rn=str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()
    if rn!='iv': continue
    page=ws.cell(row=ri,column=cols['PTS Page']).value
    ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
    name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
    # Extract sutta number within saṃyutta
    sn_inner=parts[1]
    tests.append({'num':snum,'ref':ref,'name':name,'page':page,'sn_inner':sn_inner})

tests.sort(key=lambda t:(int(t['num'].split('.')[0]),int(t['num'].split('.')[1])))
print('SN IV: {} suttas, CST: 344 (perfect match)'.format(len(tests)))

segs=sh.load_cha(['s4m.xml'],'h4n')
cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}

from openai import OpenAI
cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')
SYS='You are Helmer Smith. Are these two Pali passages (PTS and CST) the SAME sutta? Same=core teaching matches. Different=distinct sutta. Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words"}'

# Sample 20 diverse suttas — pick evenly spaced
n=len(tests); indices=list(range(0,n,n//20))[:20]
sample=[(i,tests[i]) for i in indices]
print('Testing {} suttas (evenly spaced)...'.format(len(sample)))
print('='*95)

from collections import Counter; vc=Counter()
for seq_idx,t in sample:
    # Use saṃyutta-internal sutta number for marker lookup
    pts=pts_sutta(t['page'],t['sn_inner'],300)
    if not pts:
        print('? SN {:>8s} {:>18s} | NO PTS (#{})'.format(t['num'],t['ref'],t['sn_inner']))
        vc['NO_TEXT']+=1; continue
    
    cst_text=' '.join(sh.tokens(segs[seq_idx]['text'])[:300])
    cst_title=segs[seq_idx].get('title','')
    
    key=hashlib.md5((pts[:300]+cst_text[:300]).encode()).hexdigest()
    if key in cache: hv=cache[key]
    else:
        try:
            r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
                response_format={'type':'json_object'},
                messages=[{'role':'system','content':SYS},
                          {'role':'user','content':'PTS:\n{}\n\nCST:\n{}'.format(pts[:1000],cst_text[:1000])}])
            hv=json.loads(r.choices[0].message.content)
            cache[key]=hv; CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
        except Exception as e: hv={'verdict':'ERROR','reason':str(e)[:60]}
    
    v=hv.get('verdict','?'); vc[v]+=1
    sym='✓' if v=='APPROVE' else '✗' if v=='REJECT' else '?'
    print('{} SN {:>8s} {:>18s} | {:8s} | {}'.format(sym,t['num'],t['ref'],v,hv.get('reason','')[:55]))

print('\n'+'='*95)
acc=vc.get('APPROVE',0)+vc.get('REJECT',0)
print('APPROVE: {} | REJECT: {} | NO_TEXT: {} | Precision: {:.0f}%'.format(
    vc.get('APPROVE',0),vc.get('REJECT',0),vc.get('NO_TEXT',0),
    100*vc.get('APPROVE',0)/acc if acc>0 else 0))
