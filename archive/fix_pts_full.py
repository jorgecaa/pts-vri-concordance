#!/usr/bin/env python3
"""
Fix known errors in the PTS Reference Excel and save corrected version.
"""

from openpyxl import load_workbook
from copy import copy

EXCEL_PATH = '/home/jorge/Code/squashfs-root/PTS_Reference_Full_Canon.xlsx'
OUTPUT_PATH = '/home/jorge/Code/squashfs-root/PTS_Reference_Full_Canon_CORRECTED.xlsx'

# Known corrections: (nikaya, sutta_num, field, old_value, new_value, reason)
CORRECTIONS = [
    # MN typos from blog
    ('MN', '48', 'sutta_name', 'Kosambiyann', 'Kosambiya', 'Blog typo: extra n'),
    ('MN', '53', 'sutta_name', 'Sekhann', 'Sekha', 'Blog typo: extra n'),
    
    # Blog formatting glitches
    ('AN', '4.260', 'sutta_raw', '4.260 AN 4.257 Ājānīya 2', 'AN 4.260: Ājānīya 2', 'Blog formatting glitch'),
]

def apply_corrections():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['PTS Reference']
    
    # Find column indices from header
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[val] = col
    
    print(f"Columns: {headers}")
    
    # Map our field names to column letters
    field_to_col = {
        'sutta_name': headers.get('Sutta Name'),
        'sutta_raw': headers.get('Raw Sutta ID'),
        'sutta_num': headers.get('Sutta #'),
        'nikaya': headers.get('Nikaya'),
    }
    
    fixes_applied = 0
    
    for corr in CORRECTIONS:
        nikaya, sutta_num, field, old_val, new_val, reason = corr
        
        nikaya_col = field_to_col['nikaya']
        num_col = field_to_col['sutta_num']
        target_col = field_to_col[field]
        
        if not target_col:
            print(f"⚠ Column not found for field '{field}'")
            continue
        
        found = False
        for row in range(2, ws.max_row + 1):
            row_nikaya = ws.cell(row=row, column=nikaya_col).value
            row_num = str(ws.cell(row=row, column=num_col).value or '')
            
            if row_nikaya == nikaya and row_num == sutta_num:
                cell = ws.cell(row=row, column=target_col)
                current_val = str(cell.value or '')
                
                if old_val in current_val:
                    # Apply correction to Raw Sutta ID column too if fixing name
                    if field == 'sutta_name':
                        raw_col = field_to_col['sutta_raw']
                        raw_cell = ws.cell(row=row, column=raw_col)
                        raw_val = str(raw_cell.value or '')
                        raw_new = raw_val.replace(old_val, new_val)
                        raw_cell.value = raw_new
                        print(f"  Raw:      '{raw_val}' → '{raw_new}'")
                    
                    # Apply to sutta_name
                    new_cell_val = current_val.replace(old_val, new_val)
                    cell.value = new_cell_val
                    print(f"✓ Row {row}: [{nikaya} {sutta_num}] {field}: '{current_val}' → '{new_cell_val}'  ({reason})")
                    fixes_applied += 1
                    found = True
                    break
        
        if not found:
            print(f"⚠ Not found: [{nikaya} {sutta_num}] in {field}")
    
    # Also fix the specific AN entry that got mangled from blog formatting
    for row in range(2, ws.max_row + 1):
        raw_cell = ws.cell(row=row, column=field_to_col['sutta_raw'])
        raw_val = str(raw_cell.value or '')
        if raw_val.startswith('4.260'):
            raw_cell.value = 'AN 4.260: Ājānīya 2'
            # Fix nikaya
            ws.cell(row=row, column=field_to_col['nikaya']).value = 'AN'
            ws.cell(row=row, column=field_to_col['sutta_num']).value = '4.260'
            name_cell = ws.cell(row=row, column=field_to_col['sutta_name'])
            if str(name_cell.value or '') == 'AN 4.257 Ājānīya 2':
                name_cell.value = 'Ājānīya 2'
            print(f"✓ Row {row}: Fixed AN 4.260 formatting glitch")
            fixes_applied += 1
            break
    
    wb.save(OUTPUT_PATH)
    print(f"\n✓ Saved: {OUTPUT_PATH}")
    print(f"  {fixes_applied} corrections applied")

if __name__ == '__main__':
    apply_corrections()
