#!/usr/bin/env python3
"""Show first line of actual Pali text for every MN sutta — the real opening."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

MN_BOOKS = {1:9, 2:10, 3:11}
ROMAN = {'i':1,'ii':2,'iii':3}

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=1, column=c).value
    if h: cols[h] = c

entries = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] != 'MN': continue
    entries.append({
        'num': str(row[2] or ''),
        'name': str(row[3] or ''),
        'page': row[7],
        'ref': str(row[8] or ''),
        'roman': str(row[6] or '').strip(),
    })

def is_skip_line(s):
    """Is this line just a number marker, separator, or empty?"""
    s = s.strip()
    if not s: return True
    if re.match(r'^\d+\.?$', s): return True  # "82." or "82"
    if re.match(r'^\-{5,}$', s): return True  # "-----"
    if s.startswith('(') and ')' in s[:20]: return True  # "(The text of..."
    return False

def find_first_text_line(book_no, page_no, start_line):
    """Find the first line of actual Pali text starting from start_line."""
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
               (book_no, page_no))
    r = cur.fetchone()
    if not r: return None, 'PAGE MISSING', ''
    
    lines = (r['unitext'] or '').split('\n')
    head = (r['head'] or '')[:80]
    
    # Scan from start_line forward
    text_lines = []
    for i in range(start_line - 1, len(lines)):
        s = lines[i].strip()
        if is_skip_line(s):
            continue
        # Found actual text
        text = re.sub(r'\s+', ' ', s).strip()
        return i + 1, text[:130], head
    
    return None, '[empty page]', head

print('MN — First actual Pali line of every sutta')
print('=' * 90)

for e in entries:
    num = e['num']
    page = e['page']
    roman = e['roman']
    ref = e['ref']
    name = e['name']
    
    vol_num = ROMAN.get(roman.lower(), 1)
    book_no = MN_BOOKS.get(vol_num)
    if not book_no or not page: continue
    
    # Extract line number from ref (default 1)
    line_num = 1
    m = re.search(r',(\d+)$', ref)
    if m: line_num = int(m.group(1))
    
    text_line, first_text, head = find_first_text_line(book_no, page, line_num)
    
    if text_line is None:
        print(f'  ✗ MN {num:>3s}  {ref:16s}  {first_text}')
        continue
    
    # Classify the opening
    opening_type = ''
    tlower = first_text.lower()
    if tlower.startswith('evam'):
        opening_type = 'Evaṃ'
    elif tlower.startswith('atha'):
        opening_type = 'Atha'
    elif tlower.startswith('tena'):
        opening_type = 'Tena'
    elif tlower.startswith('idam'):
        opening_type = 'Idaṃ'
    
    # Show
    print(f'  {opening_type:6s} MN {num:>3s}  {ref:16s}  {first_text}')

conn.close()
