#!/usr/bin/env python3
"""SN IV Helmer: all 344 suttas. Match by nearest gid marker, DeepSeek compare."""
import hashlib, json, os, re, sqlite3
from collections import Counter, defaultdict
from pathlib import Path
import sutta_hash as sh

CACHE_FILE=Path('helmer_ptscst_cache.json'); SN4_BOOK=15
conn=sqlite3.connect('src/data/tipitaka.sqlite'); conn.row_factory=sqlite3.Row; cur=conn.cursor()

# Build markers
marker_map=defaultdict(list)
cur.execute('SELECT page_no, unitext FROM pages WHERE book_no=? AND edition="mula"',(SN4_BOOK,))
for r in cur.fetchall():
    for i,line in enumerate(r['unitext'].split('\n')):
        s=line.strip()
        m=re.match(r'^(\d+)\.?\s*\((\d+)\)\s+(\S.*)',s)
        if m: marker_map[r['page_no']].append((i+1,int(m.group(1)),int(m.group(2)),m.group(3)[:60]))
all_markers=sorted([(p,l,g,v,n) for p in marker_map for l,g,v,n in marker_map[p]],key=lambda x:x[2])

# Load Excel  
from openpyxl import load_workbook
wb=load_workbook('PTS_Reference_Complete_Canon.xlsx'); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
entries=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2: continue
    if int(parts[0])<35 or int(parts[0])>44: continue
    if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='iv': continue
    entries.append({'num':snum,'ref':str(ws.cell(row=ri,column=cols['PTS Ref']).value or ''),
        'name':str(ws.cell(row=ri,column=cols['Sutta Name']).value or ''),
        'page':ws.cell(row=ri,column=cols['PTS Page']).value,'ri':ri,
        'sn':int(parts[0]),'sn_inner':int(parts[1])})
entries.sort(key=lambda e:(e['sn'],e['sn_inner']))

# Match entries to nearest marker by gid
tasks=[]
for e in entries:
    target=e['sn_inner']
    best=None; best_dist=999
    for p,l,g,v,n in all_markers:
        d=abs(g-target)
        if d<best_dist: best_dist=d; best=(p,l,g,v,n)
    if best and best_dist<=8:  # within 8 suttas
        p,l,g,v,n=best
        cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN4_BOOK,p))
        r=cur.fetchone()
        if r:
            lines=r['unitext'].split('\n'); start=l-1; end=len(lines)
            for j in range(start+1,len(lines)):
                if re.match(r'^\d+\.?\s*\(\d+\)',lines[j].strip()): end=j; break
            text=' '.join(lines[start:end])
            tasks.append((e,' '.join(sh.tokens(text)[:300]),l,g,v,n[:30],best_dist))

print('{} entries, {} matched to markers (dist<=8)'.format(len(entries),len(tasks)))

# DeepSeek
segs=sh.load_cha(['s4m.xml'],'h4n')
cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}
from openai import OpenAI
cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')
SYS='You are Helmer Smith. Are these two Pali passages the SAME sutta? Same=core teaching matches despite Feer abbreviation. Different=distinct sutta. Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words"}'

vc=Counter(); rej_list=[]
for i,(e,pts,l,g,v,nm,dist) in enumerate(tasks):
    cst_idx=i  # same sequential order as entries
    if cst_idx>=len(segs): continue
    cst_text=' '.join(sh.tokens(segs[cst_idx]['text'])[:300])
    
    key=hashlib.md5((pts[:200]+cst_text[:200]).encode()).hexdigest()
    if key in cache: hv=cache[key]
    else:
        try:
            r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
                response_format={'type':'json_object'},
                messages=[{'role':'system','content':SYS},
                          {'role':'user','content':'PTS:\n{}\n\nCST:\n{}'.format(pts[:800],cst_text[:800])}])
            hv=json.loads(r.choices[0].message.content)
            cache[key]=hv; CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
        except Exception as ex: hv={'verdict':'ERROR','reason':str(ex)[:60]}
    
    v=hv.get('verdict','?'); vc[v]+=1
    if v=='REJECT': rej_list.append((e,hv,pts[:60],cst_text[:60]))
    if (i+1)%40==0: print('  {} done | APPROVE: {} REJECT: {}'.format(i+1,vc.get('APPROVE',0),vc.get('REJECT',0)))

print('\n'+'='*90)
total=len(tasks); app=vc.get('APPROVE',0); rej=vc.get('REJECT',0)
print('SN IV Helmer: {} entries'.format(total))
print('  APPROVE: {} | REJECT: {}'.format(app,rej))
if app+rej>0: print('  Accuracy: {:.0f}%'.format(100*app/(app+rej)))
if rej_list:
    print('\nREJECT:')
    for e,hv,pts,cst in rej_list:
        print('  ✗ SN {:>8s} {:>15s} | {}'.format(e['num'],e['ref'],hv.get('reason','')[:70]))
        print('    PTS: {:.60s}...'.format(pts))
        print('    CST: {:.60s}...'.format(cst))

conn.close()
