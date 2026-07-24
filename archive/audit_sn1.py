#!/usr/bin/env python3
"""SN1 audit: find page errors using sequential constraint + local FP search. REPORT ONLY."""
import re, sqlite3
from collections import defaultdict
import sutta_hash as sh
import sutta_fingerprint as fp

DB='src/data/tipitaka.sqlite'; XL='PTS_Reference_Complete_Canon.xlsx'
SN1_BOOK=12

def page_text(page_no):
    conn=sqlite3.connect(DB); conn.row_factory=sqlite3.Row; cur=conn.cursor()
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN1_BOOK,page_no))
    r=cur.fetchone(); conn.close()
    return r['unitext'] if r and r['unitext'] else ''

from openpyxl import load_workbook
wb=load_workbook(XL); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}

tests=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2 or int(parts[0])>11: continue
    rn=str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()
    if rn!='i': continue
    page=ws.cell(row=ri,column=cols['PTS Page']).value
    ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
    name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
    tests.append({'num':snum,'ref':ref,'name':name,'page':page,'ri':ri})

segs=sh.load_cha(['s1m.xml'],'h4n'); st=sh.Stemmer()
cst_pool=[]; 
for s in segs:
    seq=sh.stem_seq(s['text'],st) if s['text'] else []; cst_pool.append({'seq':seq})
cst_idf=fp.global_idf([s['seq'] for s in cst_pool])
for s in cst_pool: s['sig']=fp.prose_signature(s['seq'],cst_idf,32,6)

conn=sqlite3.connect(DB); cur=conn.cursor()
cur.execute('SELECT MAX(page_no) FROM pages WHERE book_no=? AND edition="mula"',(SN1_BOOK,))
max_page=cur.fetchone()[0]; conn.close()

page_fps={}
for p in range(1,max_page+1):
    txt=page_text(p)
    if txt: seq=sh.stem_seq(txt,st); page_fps[p]=seq
page_idf=fp.global_idf(list(page_fps.values()))
for p in page_fps: page_fps[p]=fp.prose_signature(page_fps[p],page_idf,32,6)

# Find sequential breaks
prev_page=0; prev_i=0; issues=[]
for i,t in enumerate(tests):
    page=t['page']
    if page and prev_page and page<prev_page-1:
        issues.append((prev_i,i,t,prev_page))
    if page: prev_page=page; prev_i=i

print('SN I: {} suttas, {} sequential breaks found'.format(len(tests),len(issues)))
print()

for start_i,end_i,t,prev_page in issues:
    # Search pages prev_page-2 to prev_page+20 for best match
    cst_sig=cst_pool[end_i]['sig']
    best_page,best_score=0,0
    for p in range(max(1,prev_page-2),min(max_page,prev_page+25)):
        if p not in page_fps: continue
        score=fp.prose_score(cst_sig,page_fps[p],page_idf)
        if score>best_score: best_score=score; best_page=p
    
    current_page=t['page']
    if best_page!=current_page:
        status='⚠ FIX' if best_score>0.15 else '? LOW CONF'
        print('{} SN {:>8s} | blog: S i {:>3d} | found: S i {:>3d} | score={:.3f} | {}'.format(
            status,t['num'],current_page,best_page,best_score,t['name'][:30]))
    else:
        print('✓ SN {:>8s} | blog: S i {:>3d} KEPT | score={:.3f}'.format(t['num'],current_page,best_score))

# Also check: for each sutta, does it actually appear on the stated page?
print('\n--- Manual verification of key entries ---')
check_list=[('3.1',68),('1.9',4),('1.10',5),('2.1',46)]
for snum,expected_page in check_list:
    for t in tests:
        if t['num']==snum:
            match='✓' if t['page']==expected_page else '✗ (blog says {})'.format(t['page'])
            print('  SN {} at S i {}: {}'.format(snum,expected_page,match))
