#!/usr/bin/env python3
"""Validación Thag/Thig: marcadores de pada + monotonicidad verso->pagina."""
import sqlite3, base64, re
from openpyxl import load_workbook

def decode(val):
    if not val: return ''
    try:
        raw = base64.b64decode(val.strip() + '==')
        if raw[:3] == b'\xef\xbb\xbf': raw = raw[3:]
        return raw.decode('utf-8', errors='replace')
    except: return str(val)[:80]

con = sqlite3.connect('src/data/tipitaka.sqlite')
wb = load_workbook('PTS_Reference_Khuddaka_Nikaya.xlsx')
ws = wb['PTS Reference']

VOL2BOOK = {"Th":29,"Thag":29,"Thi":29,"Thig":29}

entries = []
for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name = str(r[1]).strip() if r[1] else ''
    pts_vol = str(r[3]).strip() if r[3] else ''
    pts_vs = str(r[4]).strip() if r[4] else ''
    section = str(r[5]).strip() if len(r) > 5 and r[5] else ''
    rtype = str(r[6]).strip() if len(r) > 6 and r[6] else ''
    if rtype in ("Book Header", "Section Header"): continue
    if 'Theragatha' not in section.replace('ā','a').replace('ī','i') and \
       'Therigatha' not in section.replace('ā','a').replace('ī','i'): continue
    if not pts_vol: continue

    v = pts_vol.strip()
    if "," in v: v = v.split(",")[0].strip()
    v = re.sub(r"\s+"," ",v)
    m = re.match(r"^(.+?)\s+(\d+)$", v)
    if not m: continue
    vol_abbr, page = m.group(1).strip(), int(m.group(2))
    db_book = VOL2BOOK.get(vol_abbr)
    if not db_book: continue

    vs_m = re.search(r'v(\d+)(?:[–-](\d+))?', pts_vs) if pts_vs else None
    v_start = int(vs_m.group(1)) if vs_m else None
    v_end = int(vs_m.group(2)) if (vs_m and vs_m.group(2)) else v_start

    entries.append((name, page, v_start, v_end, db_book, pts_vs))

# ── Validar ──
verse_text = title_page = no_markers = non_mono = page_gap = 0
issues = []

for i, (name, page, v_start, v_end, db_book, pts_vs) in enumerate(entries):
    row = con.execute(
        "SELECT head, unitext FROM pages WHERE edition='mula' AND book_no=? AND page_no=?",
        (db_book, page)).fetchone()
    if not row:
        issues.append((name[:45], page, v_start, "PAGE NOT FOUND"))
        continue

    h = (decode(row[0]) or '')
    t = decode(row[1] or '')

    # 1. Marcadores de pada
    has_pada = bool(re.search(r'[|║]', t))
    is_title = bool(re.search(r'THER[AĪ]-GĀTHĀ', h.upper()))

    if has_pada:        verse_text += 1
    elif is_title:      title_page += 1
    else:
        no_markers += 1
        if len(issues) < 10:
            issues.append((name[:45], page, v_start, "sin | ni ║"))

    # 2. Monotonicidad
    if i > 0 and v_start is not None:
        prev_v = entries[i-1][3]
        prev_page = entries[i-1][1]
        if prev_v is not None and v_start < prev_v:
            non_mono += 1
            if len(issues) < 15:
                issues.append((name[:45], page, v_start,
                    f"v.{v_start} < v.{prev_v} anterior"))

    # 3. Salto de pagina
    if i > 0 and v_start is not None:
        prev_v = entries[i-1][3]
        prev_page = entries[i-1][1]
        if prev_v is not None:
            vj = v_start - prev_v
            pj = page - prev_page
            if pj > 5 and vj < 10:
                page_gap += 1
                if len(issues) < 15:
                    issues.append((name[:45], page, v_start,
                        f"salto p.+{pj} con solo +{vj} versos"))

total = len(entries)
print(f"Thag/Thig: {total} entradas\n")
print(f"  ✓ Texto con padas (| o ║)     : {verse_text} ({100*verse_text/total:.0f}%)")
print(f"  ~ Pagina de titulo             : {title_page}")
print(f"  ⚠ Sin marcadores               : {no_markers}")
print(f"  ⚠ No monotonico                : {non_mono}")
print(f"  ⚠ Salto de pagina anomalo      : {page_gap}")

if issues:
    print(f"\n  Anomalias ({len(issues)}):")
    for name, page, v_start, reason in issues[:15]:
        vs = f"v.{v_start}" if v_start else ""
        print(f"    {name:<47} p.{page} {vs}  {reason}")

con.close()
