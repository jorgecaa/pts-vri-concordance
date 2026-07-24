#!/usr/bin/env python3
"""Hunt SN III — split compound names into searchable chunks."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def sd(t):
    for k,v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        t=t.replace(k,v).replace(k.upper(),v.upper())
    return t

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=1, column=c).value
    if h: cols[h] = c

entries = []
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
    if str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip() != 'iii': continue
    entries.append({
        'ri': ri, 'num': str(ws.cell(row=ri, column=cols['Sutta #']).value or ''),
        'name': str(ws.cell(row=ri, column=cols['Sutta Name']).value or ''),
        'page': ws.cell(row=ri, column=cols['PTS Page']).value,
        'ref': str(ws.cell(row=ri, column=cols['PTS Ref']).value or ''),
    })

print('Loading book 14...')
cur.execute('SELECT page_no, head, unitext FROM pages WHERE book_no=14 AND edition="mula" ORDER BY page_no')
all_pages = {}
for r in cur.fetchall():
    all_pages[r['page_no']] = sd((r['head'] or '').lower()) + ' ' + sd((r['unitext'] or '').lower())[:2000]
print('%d pages' % len(all_pages))

def keywords(name):
    clean = sd(name.lower())
    clean = re.sub(r'\(.*?\)', '', clean)
    clean = re.sub(r'\[.*?\]', '', clean)
    skip = {'sutta','suttam','vagga','pathama','dutiya','tatiya','catuttha','pancaka','catukka','adisu','tika','duka','disu'}
    words = []
    for w in re.split(r'[\s\-,;:.]+', clean):
        w = w.strip()
        if len(w) < 3 or w in skip: continue
        if len(w) <= 8: words.append(w)
        else:
            for i in range(0, len(w)-2, 3):
                chunk = w[i:i+6]
                if len(chunk) >= 3 and chunk not in skip:
                    words.append(chunk)
    return list(set(words))[:8]

def hunt(name, stated_page):
    kw = keywords(name)
    if not kw: return None, 0
    best_page, best_score = None, 0
    for p in range(max(1,stated_page-8), min(597,stated_page+9)):
        if p not in all_pages: continue
        hits = sum(1 for w in kw if w in all_pages[p])
        if hits > best_score:
            best_score = hits; best_page = p
    if best_score >= 2: return best_page, best_score
    return None, 0

print()
correct = wrong = nf = 0
for e in entries:
    if ',' in e['ref']: correct += 1; continue  # already has line number
    pg, score = hunt(e['name'], e['page'])
    if pg is None: nf += 1; print('  XX SN %7s p.%-3d %s' % (e['num'], e['page'], e['name'][:50]))
    elif pg == e['page']: correct += 1
    else: wrong += 1; print('  !! SN %7s p.%-3d -> p.%-3d (score=%d) %s' % (e['num'], e['page'], pg, score, e['name'][:40]))

print()
print('  Verified: %d  |  Wrong: %d  |  Not found: %d  |  Total: %d' % (correct, wrong, nf, correct+wrong+nf))
conn.close()
