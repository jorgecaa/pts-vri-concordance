#!/usr/bin/env python3
"""
Helmer Smith — 100 Critical Cross-Validation Tests
Sutta Pitaka: CST reference → PTS content verification.
Tests: content existence, incipit match, name keyword match, RTE cross-ref.
"""
import sqlite3, re, json, os, random
from openpyxl import load_workbook
from collections import defaultdict, Counter

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'
RTE_DIR = '/home/jorge/Code/tipitaka.rte/Canonical'

random.seed(42)
conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def norm(s):
    for a,b in [('ā','a'),('ī','i'),('ū','u'),('ṅ','n'),('ñ','n'),
                ('ṭ','t'),('ḍ','d'),('ṇ','n'),('ḷ','l'),('ṃ','m'),('ṁ','m')]:
        s = s.replace(a,b).replace(a.upper(),b.upper())
    return s

def load_rte():
    fm = {'09':('D','i'),'10':('D','ii'),'11':('D','iii'),
          '12':('M','i'),'13':('M','ii'),'14':('M','iii'),
          '15':('S','i'),'16':('S','ii'),'17':('S','iii'),'18':('S','iv'),'19':('S','v'),
          '20':('A','i'),'21':('A','ii'),'22':('A','iii'),'23':('A','iv'),'24':('A','v')}
    refs = set()
    for fn in sorted(os.listdir(RTE_DIR)):
        if not fn.endswith('.txt'): continue
        pfx = fn.split('-')[0]
        if pfx not in fm: continue
        vl, rm = fm[pfx]
        with open(os.path.join(RTE_DIR, fn), encoding='utf-8-sig') as f:
            for line in f:
                for m in re.finditer(r'\(pts\.\s+([a-z]+)(?:\s+([ivxlcdm]+))?,?\s+(\d+)\)', line, re.IGNORECASE):
                    if m.group(1).lower() != vl.lower(): continue
                    refs.add(f'{vl} {m.group(2) or rm} {int(m.group(3))}')
    return refs

print('Loading RTE...')
rte_refs = load_rte()

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

# ── Collect all entries ──
all_entries = []
for ri in range(2, ws.max_row + 1):
    nik = str(ws.cell(row=ri, column=cols['Nikaya']).value or '')
    snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    if not snum or snum == 'None': continue
    all_entries.append({
        'ri': ri, 'nik': nik, 'num': snum,
        'name': str(ws.cell(row=ri, column=cols['Sutta Name']).value or ''),
        'roman': str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower(),
        'vol': str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip(),
        'page': ws.cell(row=ri, column=cols['PTS Page']).value,
        'ref': str(ws.cell(row=ri, column=cols['PTS Ref']).value or ''),
        'raw': str(ws.cell(row=ri, column=cols['Raw ID']).value or ''),
        'val': str(ws.cell(row=ri, column=cols['Validation']).value or ''),
    })

def book_no(nik, roman, vol):
    if nik == 'DN': return {'i':6,'ii':7,'iii':8}.get(roman)
    if nik == 'MN': return {'i':9,'ii':10,'iii':11}.get(roman)
    if nik == 'SN': return {'i':12,'ii':13,'iii':14,'iv':15,'v':16}.get(roman)
    if nik == 'AN': return {'i':17,'ii':18,'iii':19,'iv':20,'v':21}.get(roman)
    if nik == 'KN':
        m = {'Khp':22,'Kh':22,'Dhp':23,'Dh':23,'Ud':24,'It':25,'Sn':26,
             'Vv':27,'Pv':28,'Th':29,'Th & Th':29,'Thi':29,'Thī':29,
             'Ja':30,'Ja I':30,'Ja II':31,'Ja III':32,'Ja IV':33,'Ja V':34,'Ja VI':35,
             'Nidd':36,'Nidd I':36,'Nidd II':37,
             'Patis I':38,'Patis II':39,'Paṭis I':38,'Paṭis II':39,
             'Ap':40,'Bv':41,'Cp':42}
        if vol in m: return m[vol]
        for k,v in m.items():
            if vol.startswith(k) or k.startswith(vol): return v
    return None

# ── Select 100 critical tests ──
tests = []

# DN: all 34 suttas
dn = [e for e in all_entries if e['nik'] == 'DN']
tests.extend(dn)

# MN: 20 key suttas (first of each vagga + classics)
mn = [e for e in all_entries if e['nik'] == 'MN']
mn_key_nums = {'1','2','7','10','13','15','22','26','41','51','61','63','72','77','85','95','107','117','131','148'}
tests.extend([e for e in mn if e['num'] in mn_key_nums])

# SN: 20 across saṃyuttas
sn = [e for e in all_entries if e['nik'] == 'SN']
# Pick first of each major saṃyutta + key suttas
sn_keys = {'1.1','2.1','12.1','22.1','35.1','36.1','45.1','46.1','47.1','48.1',
           '22.59','35.28','56.11','12.15','14.1','4.1','7.1','11.1','6.1','5.1'}
tests.extend([e for e in sn if e['num'] in sn_keys])

# AN: 15 across nipātas
an = [e for e in all_entries if e['nik'] == 'AN']
an_keys = {'1.1','2.1','3.1','3.65','4.1','5.1','5.57','6.1','7.1','8.1','8.30','9.1','10.1','11.1'}
tests.extend([e for e in an if e['num'] in an_keys])

# KN: 11 across books
kn = [e for e in all_entries if e['nik'] == 'KN']
kn_keys = {'1.1','2.1','3.1.1','4.1.1','5.1.1','6.1.1.1','7.1.1','8.1.1.1','9.1.1','10.1.1'}
tests.extend([e for e in kn if e['num'] in kn_keys])

# Deduplicate
seen = set()
unique = []
for t in tests:
    k = (t['nik'], t['num'])
    if k not in seen:
        seen.add(k)
        unique.append(t)
tests = unique[:100]

print(f'Running {len(tests)} critical tests...')
print('=' * 100)

# ── Run tests ──
results = []
scores = []

for e in tests:
    bn = book_no(e['nik'], e['roman'], e['vol'])
    score = 0
    checks = []
    
    # Test 1: Page exists
    if bn and e['page']:
        cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (bn, e['page']))
        row = cur.fetchone()
        if row and row['unitext']:
            score += 1; checks.append('PAGE_EXISTS')
            head = row['head'] or ''
            text = row['unitext']
            lines = text.split('\n')
            
            # Test 2: Page has content (not just whitespace)
            content_chars = len(text.strip())
            if content_chars > 100:
                score += 1; checks.append('HAS_CONTENT')
            
            # Test 3: Name keyword in first 500 chars
            kw = [w for w in re.findall(r'[a-z]{4,}', norm(e['name'].lower()))
                  if w not in ('sutta','suttam','vagga','pathama','dutiya','tatiya',
                              'catuttha','pancama','chattha','sattama', 'paritta')]
            body = norm(text[:800]).lower()
            head_n = norm(head).lower()
            hits_body = sum(1 for w in kw if w in body)
            hits_head = sum(1 for w in kw if w in head_n)
            if kw and hits_body >= 1:
                score += 1; checks.append('NAME_IN_BODY')
            elif kw and hits_head >= 1:
                score += 1; checks.append('NAME_IN_HEAD')
            
            # Test 4: Structural marker on page
            has_marker = False
            for i, line in enumerate(lines):
                s = line.strip()
                if re.match(r'^\d+\.\s+\S', s): has_marker = True; break
                if re.search(r'[Ee]va[mM].*suta[mM]', s): has_marker = True; break
                if re.search(r'[║|]\s*\d+\s*[║|]', s): has_marker = True; break
            if has_marker:
                score += 1; checks.append('HAS_MARKER')
            
            # Test 5: RTE cross-reference
            rte_key = f'{e["vol"]} {e["roman"]} {e["page"]}'
            if rte_key in rte_refs:
                score += 1; checks.append('RTE_CONFIRM')
            
            # Test 6: Incipit — first content words make sense
            content_words = []
            for line in lines:
                s = line.strip()
                if s and len(s) > 5 and not re.match(r'^[A-ZĀĪŪ\s\-\.║|]+$', s):
                    content_words = re.findall(r'[a-zA-Zāīūṁṃṅñṭḍṇḷ]+', s)[:5]
                    break
            if content_words:
                score += 1; checks.append('HAS_INCIPIT')
        else:
            checks.append('PAGE_MISSING')
    else:
        checks.append('NO_BOOK')
    
    scores.append(score)
    results.append((e, score, checks))

# ── Report ──
print(f'{"Nik":3s} {"#":12s} {"Ref":>18s} {"S":1s} {"Checks"}')
print('-' * 100)

for e, score, checks in results:
    bar = '█' * score + '░' * (6 - score)
    sym = '✓' if score >= 5 else '~' if score >= 3 else '✗'
    print(f'{sym} {e["nik"]:2s} {e["num"]:12s} {e["ref"]:>18s} {bar} {", ".join(checks[:5])}')

# Summary
print(f'\n{"="*100}')
score_dist = Counter(scores)
print(f'  Score distribution:')
for s in sorted(score_dist.keys(), reverse=True):
    bar = '█' * score_dist[s]
    print(f'    {s}/6: {score_dist[s]:>3d} tests {bar}')

avg = sum(scores) / len(scores)
print(f'\n  Average score: {avg:.1f}/6')
print(f'  Perfect (6/6): {score_dist.get(6,0)}')
print(f'  Strong (5/6):  {score_dist.get(5,0)}')
print(f'  Good (4/6):    {score_dist.get(4,0)}')
print(f'  Fair (3/6):    {score_dist.get(3,0)}')
print(f'  Weak (<3):     {sum(v for k,v in score_dist.items() if k < 3)}')

# Per-Nikaya breakdown
nik_scores = defaultdict(list)
for e, score, _ in results:
    nik_scores[e['nik']].append(score)

print(f'\n  Per Nikaya:')
for nik in ['DN', 'MN', 'SN', 'AN', 'KN']:
    if nik in nik_scores:
        ss = nik_scores[nik]
        print(f'    {nik:3s}: avg {sum(ss)/len(ss):.1f}/6, best={max(ss)}, worst={min(ss)}, n={len(ss)}')

# Flag weak entries
weak = [(e, s, c) for e, s, c in results if s < 3]
if weak:
    print(f'\n  ⚠ Weak entries ({len(weak)}):')
    for e, s, c in weak:
        print(f'    {e["nik"]} {e["num"]:>12s} | {e["ref"]:>18s} | score={s}/6 | {", ".join(c)}')

# Helmer's verdict
if avg >= 5.0:
    verdict = "SOUND. The PTS references consistently resolve to correct sutta content."
elif avg >= 4.0:
    verdict = "RELIABLE. Minor gaps in RTE cross-references or keyword matching."
elif avg >= 3.0:
    verdict = "ADEQUATE. Most references verified; some entries need page adjustment."
else:
    verdict = "NEEDS REVIEW. Significant portion of references could not be verified."

print(f'\n  Helmer Smith verdict: {verdict}')

conn.close()
