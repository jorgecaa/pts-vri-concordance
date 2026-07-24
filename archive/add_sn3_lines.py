#!/usr/bin/env python3
"""SN III — fix remaining 49 entries by matching sutta NAMES to PTS markers."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=1, column=c).value
    if h: cols[h] = c

def strip_dia(t):
    for k,v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        t = t.replace(k,v).replace(k.upper(),v.upper())
    return t

def name_words(name):
    """Extract key words from sutta name for fuzzy matching."""
    clean = strip_dia(name.lower())
    clean = re.sub(r'\(.*?\)', '', clean)
    clean = re.sub(r'\[.*?\]', '', clean)
    words = [w for w in re.split(r'[\s\-,;:.]+', clean) if len(w) >= 3]
    return words

def find_by_name(book_no, page_no, excel_name):
    """Search page for ANY marker whose name matches the Excel sutta name."""
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None, None
    
    excel_words = name_words(excel_name)
    if len(excel_words) < 1: return None, None
    
    lines = (r['unitext'] or '').split('\n')
    
    # Find any "N (M) Name" marker — MUST have vagga position in parens
    marker_pat = re.compile(r'^\s*(\d+(?:\-\d+)?)\s*\(\d+(?:\-\d+)?\)\s*(.+)')
    
    best_match = None
    best_score = 0
    
    for i, line in enumerate(lines):
        m = marker_pat.match(line.strip())
        if not m: continue
        marker_name = m.group(2).strip()
        if not marker_name: continue
        
        marker_words = name_words(marker_name)
        if not marker_words: continue
        
        # Count matching words
        matches = sum(1 for w in excel_words if any(w in mw or mw in w for mw in marker_words))
        score = matches / max(len(excel_words), 1)
        
        if score > best_score and score >= 0.5:
            best_score = score
            best_match = (i+1, m.group(1), marker_name)
    
    return best_match

print('SN III — Name-based matching for remaining entries')
print('=' * 70)

fixed = 0
still_not = 0

for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip()
    if roman != 'iii': continue
    
    sutta_num = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
    
    if not page: continue
    
    # Skip if already has comma (line number)
    if ',' in ref: continue
    
    result = find_by_name(14, page, name)
    
    if result:
        line, marker_id, marker_name = result
        new_ref = 'S iii %d,%d' % (page, line)
        ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
        fixed += 1
        print('  SN %7s  %-16s  [%s] %s' % (sutta_num, new_ref, marker_id, marker_name[:60]))
    else:
        still_not += 1
        print('  ?? SN %7s  p.%-3d  %s' % (sutta_num, page, name[:50]))

wb.save(XL)
print()
print('  Fixed by name: %d  |  Still not found: %d' % (fixed, still_not))
print('  Saved: %s' % XL)
conn.close()
