#!/usr/bin/env python3
"""SN — universal fix: name-based matching when numeric ID fails (CST vs PTS numbering)."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

SN_MAP = {1:12, 2:13, 3:14, 4:15, 5:16}
VOL_LETTER = {1:'i',2:'ii',3:'iii',4:'iv',5:'v'}

def sd(t):
    for k,v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        t=t.replace(k,v).replace(k.upper(),v.upper())
    return t

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1,column=c).value: c for c in range(1, ws.max_column+1)}

def find_by_name(book_no, page_no, excel_name):
    """Search page for any marker whose name fuzzy-matches the Excel sutta name."""
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None
    lines = (r['unitext'] or '').split(chr(10))
    
    # Extract key words from Excel name
    clean = sd(excel_name.lower())
    clean = re.sub(r'\(.*?\)', '', clean)
    clean = re.sub(r'\[.*?\]', '', clean)
    name_words = [w for w in re.split(r'[\s\-,;:.]+', clean) if len(w) >= 3]
    skip = {'sutta','suttam','vagga','pathama','dutiya','tatiya','catuttha','pancaka','adisu','tika','duka'}
    name_words = [w for w in name_words if w not in skip]
    if not name_words: return None
    
    best_line = None
    best_score = 0
    
    for i, line in enumerate(lines):
        s = line.strip()
        if not s or len(s) > 100: continue
        if re.match(r'^\d+\s', s) and not re.search(r'\(\d', s): continue  # plain paragraph numbers
        
        marker_clean = sd(s.lower())
        hits = sum(1 for w in name_words if w in marker_clean)
        
        if hits > best_score:
            best_score = hits
            best_line = i + 1
    
    if best_score >= max(1, len(name_words) * 0.5):
        return best_line
    return None

total_fixed = 0
for vol_num in [1,2,3,4,5]:
    book_no = SN_MAP[vol_num]
    roman = VOL_LETTER[vol_num]
    vol_fixed = 0
    
    for ri in range(2, ws.max_row + 1):
        if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
        if str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip() != roman: continue
        
        sutta_num = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
        page = ws.cell(row=ri, column=cols['PTS Page']).value
        ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
        name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
        if not page: continue
        
        # Only fix entries at L1
        if not re.search(r',1$', ref): continue
        
        line = find_by_name(book_no, page, name)
        if line and line > 1:
            new_ref = 'S %s %d,%d' % (roman, page, line)
            ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
            vol_fixed += 1
    
    if vol_fixed:
        print('SN %s: fixed %d by name' % (roman.upper(), vol_fixed))
    total_fixed += vol_fixed

wb.save(XL)
print()
print('Total fixed: %d' % total_fixed)

# Final stats
from collections import defaultdict
stats = defaultdict(lambda: {'total':0,'l2':0})
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] != 'SN': continue
    roman = str(row[6] or '').strip()
    m = re.search(r',(\d+)$', str(row[8] or ''))
    line = int(m.group(1)) if m else 0
    stats[roman]['total'] += 1
    if line > 1: stats[roman]['l2'] += 1

for r in ['i','ii','iii','iv','v']:
    s = stats[r]
    print('SN %s: %d/%d (%.0f%%) L>1' % (r.upper(), s['l2'], s['total'], 100*s['l2']/s['total']))
total = sum(s['l2'] for s in stats.values())
print('TOTAL: %d/1806 (%.0f%%)' % (total, 100*total/1806))
conn.close()
