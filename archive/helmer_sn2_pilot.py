#!/usr/bin/env python3
"""Helmer SN II pilot v2: global sequential numbering."""
import hashlib, json, os, re, sqlite3, random
from pathlib import Path
import sutta_hash as sh

CACHE_FILE=Path('helmer_ptscst_cache.json'); SN2_BOOK=13

def pts_sutta(page_no, global_num, max_tokens=350):
    conn=sqlite3.connect('src/data/tipitaka.sqlite'); conn.row_factory=sqlite3.Row; cur=conn.cursor()
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN2_BOOK,page_no))
    r=cur.fetchone(); conn.close()
    if not r or not r['unitext']: return ''
    lines=r['unitext'].split('\n'); sn=str(global_num)
    start=None
    for i,line in enumerate(lines):
        s=line.strip()
        if re.match(r'^'+re.escape(sn)+r'\s*\(\d+\)\s+\S',s): start=i; break
        if re.match(r'^\('+re.escape(sn)+r'\)\s*\(\d+\)\s+\S',s): start=i; break
    if start is None: return ''
    end=len(lines)
    for j in range(start+1,len(lines)):
        s=lines[j].strip()
        if re.match(r'^\d+\s*\(\d+\)\s+\S',s): end=j; break
        if re.match(r'^\(\d+\)\s*\(\d+\)\s+\S',s): end=j; break
    return ' '.join(sh.tokens(' '.join(lines[start:end]))[:max_tokens])

from openpyxl import load_workbook
wb=load_workbook('PTS_Reference_Complete_Canon.xlsx'); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}

tests=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2: continue
    if int(parts[0])<12 or int(parts[0])>21: continue
    if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='ii': continue
    page=ws.cell(row=ri,column=cols['PTS Page']).value
    ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
    name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
    tests.append({'num':snum,'ref':ref,'name':name,'page':page})

# Sort sequentially
tests.sort(key=lambda t:(int(t['num'].split('.')[0]),int(t['num'].split('.')[1])))
print('SN II: {} suttas (sorted)'.format(len(tests)))

segs=sh.load_cha(['s2m.xml'],'h4n')
cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}

from openai import OpenAI
cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')
SYS='You are Helmer Smith. Are these two Pali passages (PTS and CST) the SAME sutta? Same=core teaching matches despite orthographic/expansion differences. Different=distinct sutta. Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words"}'

random.seed(42); indices=list(range(len(tests))); random.shuffle(indices)
sample=[(i,tests[i]) for i in indices[:15]]

print('Helmer SN II — 15 suttas (global sequential numbering)')
print('='*95)
from collections import Counter; vc=Counter()

for seq_idx,t in sample:
    global_num=seq_idx+1
    pts=pts_sutta(t['page'],global_num,350)
    if not pts:
        print('? SN {:>8s} {:>18s} | NO PTS (global #{})'.format(t['num'],t['ref'],global_num)); vc['NO_TEXT']+=1; continue
    
    cst_idx=seq_idx
    if cst_idx>=len(segs):
        print('? SN {:>8s} | CST idx {} OOB'.format(t['num'],cst_idx)); continue
    
    cst_text=' '.join(sh.tokens(segs[cst_idx]['text'])[:350])
    cst_title=segs[cst_idx].get('title','')
    
    key=hashlib.md5((pts[:300]+cst_text[:300]).encode()).hexdigest()
    if key in cache: hv=cache[key]
    else:
        try:
            r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
                response_format={'type':'json_object'},
                messages=[{'role':'system','content':SYS},
                          {'role':'user','content':'PTS:\n{}\n\nCST:\n{}'.format(pts[:1000],cst_text[:1000])}])
            hv=json.loads(r.choices[0].message.content)
            cache[key]=hv; CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
        except Exception as e: hv={'verdict':'ERROR','reason':str(e)[:60]}
    
    v=hv.get('verdict','?'); vc[v]+=1
    sym='✓' if v=='APPROVE' else '✗' if v=='REJECT' else '?'
    print('{} SN {:>8s} {:>18s} | {:8s} | CST[{}] {} | {}'.format(
        sym,t['num'],t['ref'],v,cst_idx,cst_title[:30],hv.get('reason','')[:50]))

print('\n'+'='*95)
print('APPROVE: {} | REJECT: {} | ERROR: {} | NO_TEXT: {}'.format(
    vc.get('APPROVE',0),vc.get('REJECT',0),vc.get('ERROR',0),vc.get('NO_TEXT',0)))
acc=vc.get('APPROVE',0)+vc.get('REJECT',0)
if acc>0: print('Precision: {:.0f}%'.format(100*vc.get('APPROVE',0)/acc))
