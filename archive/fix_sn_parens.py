#!/usr/bin/env python3
"""SN — fix regex to handle vagga position RANGES like (7-9), (10-12)."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

SN_MAP = {1:12, 2:13, 3:14, 4:15, 5:16}
VOL_LETTER = {1:'i',2:'ii',3:'iii',4:'iv',5:'v'}

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1,column=c).value: c for c in range(1, ws.max_column+1)}

def find_marker_v2(book_no, page_no, sutta_num_str):
    parts = sutta_num_str.split('.')
    if len(parts) < 2: return None
    sid = parts[1]
    
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None
    lines = (r['unitext'] or '').split(chr(10))
    esc = re.escape(sid)
    
    # PAREN_PAT: matches (4) or (7-9) or (10-12) or (19)
    PAREN = r'\(\d+(?:\-\d+)?\)'
    
    for i, line in enumerate(lines):
        s = line.strip()
        if not s: continue
        
        # "56. (4) Name" — with dot (SN V)
        m = re.match(r'^\s*' + esc + r'\.\s*' + PAREN + r'\s*(.*)', s)
        if m: return i+1
        
        # "56 (4) Name" — without dot
        m = re.match(r'^\s*' + esc + r'\s*' + PAREN + r'\s*(.*)', s)
        if m: return i+1
        
        # Range: "140-142 (5-7) Name"
        m = re.match(r'^\s*(\d+)\-(\d+)\s*' + PAREN + r'\s*(.*)', s)
        if m:
            try:
                lo, hi = int(m.group(1)), int(m.group(2))
                if lo <= int(sid) <= hi: return i+1
            except: pass
        
        # "§ N. Name" — SN I
        m = re.match(r'^\s*§\s*' + esc + r'[\.\s]+(.*)', s)
        if m: return i+1
    
    return None

total_fixed = 0
for vol_num in [1,2,3,4,5]:
    book_no = SN_MAP[vol_num]
    roman = VOL_LETTER[vol_num]
    fixed = 0
    
    for ri in range(2, ws.max_row + 1):
        if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
        if str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip() != roman: continue
        
        sutta_num = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
        page = ws.cell(row=ri, column=cols['PTS Page']).value
        ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
        if not page: continue
        
        # Only fix entries currently at L1
        if not re.search(r',1$', ref): continue
        
        line = find_marker_v2(book_no, page, sutta_num)
        if line and line > 1:
            new_ref = 'S %s %d,%d' % (roman, page, line)
            ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
            fixed += 1
    
    if fixed:
        print('SN %s: fixed %d entries' % (roman.upper(), fixed))
    total_fixed += fixed

wb.save(XL)
print()
print('Total fixed: %d' % total_fixed)
print('Saved: %s' % XL)
conn.close()
