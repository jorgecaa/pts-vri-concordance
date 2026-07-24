#!/usr/bin/env python3
"""Fix all SN volumes — name-based marker matching for shared-page suttas."""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

SN_MAP = {1:12, 2:13, 3:14, 4:15, 5:16}

def sd(t):
    for k,v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        t=t.replace(k,v).replace(k.upper(),v.upper())
    return t

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1,column=c).value: c for c in range(1, ws.max_column+1)}

# Find pages with >1 sutta where any are at L1
pages = defaultdict(list)
entry_map = {}
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
    sutta_num = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip()
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
    if not page: continue
    m = re.search(r',(\d+)$', ref)
    line = int(m.group(1)) if m else 0
    pages[(roman, page)].append((ri, sutta_num, line, name))
    entry_map[(roman, sutta_num)] = ri

def find_by_name(book_no, page_no, name):
    """Find any sutta marker on page whose name matches."""
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None
    lines = (r['unitext'] or '').split(chr(10))
    
    name_words = [w for w in re.split(r'[\s\-,;:.()]+', sd(name.lower())) if len(w) >= 3]
    if not name_words: return None
    
    for i, line in enumerate(lines):
        s = line.strip()
        # Try all marker patterns
        for pat in [r'^\s*§\s*\d+\.\s*(.+)', r'^\s*\d+\.?\s*\(\d+\)\s*(.+)', r'^\s*\d+\s+(\S.+)']:
            m = re.match(pat, s)
            if not m: continue
            marker = sd(m.group(1).strip().lower())
            hits = sum(1 for w in name_words if w in marker)
            if hits >= len(name_words) * 0.5 and hits >= 1:
                return i + 1
    return None

vol_map = {'i':1,'ii':2,'iii':3,'iv':4,'v':5}
updates = 0

for (roman, page), entries in pages.items():
    if len(entries) <= 1: continue
    # Only fix if some are at L1
    l1_entries = [(ri, sn, name) for ri, sn, line, name in entries if line <= 1]
    if not l1_entries: continue
    
    vol_num = vol_map.get(roman, 0)
    book_no = SN_MAP.get(vol_num)
    if not book_no: continue
    
    for ri, sn, name in l1_entries:
        line = find_by_name(book_no, page, name)
        if line and line > 1:
            new_ref = 'S %s %d,%d' % (roman, page, line)
            ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
            updates += 1

wb.save(XL)
print('Fixed: %d entries across all SN volumes' % updates)
conn.close()
