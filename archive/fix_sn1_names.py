#!/usr/bin/env python3
"""SN I — fix line numbers using name-matching for section markers."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def sd(t):
    for k,v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        t=t.replace(k,v).replace(k.upper(),v.upper())
    return t

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1,column=c).value: c for c in range(1, ws.max_column+1)}

def find_section_marker(book_no, page_no, excel_name):
    """Find '§ N. Name' marker matching the sutta name."""
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None
    lines = (r['unitext'] or '').split(chr(10))
    
    # Extract key name words
    name_words = [w for w in re.split(r'[\s\-,;:.()]+', sd(excel_name.lower())) if len(w) >= 3]
    if not name_words: return None
    
    for i, line in enumerate(lines):
        s = line.strip()
        m = re.match(r'^\s*§\s*\d+\.\s*(.+)', s)
        if not m: continue
        marker_name = sd(m.group(1).strip().lower())
        # Check how many name words appear in marker
        hits = sum(1 for w in name_words if w in marker_name)
        if hits >= len(name_words) * 0.5 and hits >= 1:
            return i + 1
    
    return None

updates = 0
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip()
    if roman != 'i': continue
    
    sutta_num = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
    if not page: continue
    
    # Only fix entries that are at L1 (likely wrong)
    if ',1' not in ref: continue
    
    line = find_section_marker(12, page, name)
    if line and line != 1:
        new_ref = 'S i %d,%d' % (page, line)
        ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
        updates += 1
        print('SN %7s  %s -> %s  %s' % (sutta_num, ref, new_ref, name[:40]))

wb.save(XL)
print()
print('Fixed: %d SN I entries' % updates)
conn.close()
