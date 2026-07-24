#!/usr/bin/env python3
"""DeepSeek validation: SN vol 1 (271 suttas). Sutta-bounded extraction."""
import hashlib, json, os, re, sqlite3
from collections import Counter
from pathlib import Path
import sutta_hash as sh

DB='src/data/tipitaka.sqlite'; XL='PTS_Reference_Complete_Canon.xlsx'
CACHE_FILE=Path('helmer_ptscst_cache.json'); SN1_BOOK=12

def pts_sutta(page_no, sutta_num, max_tokens=200):
    conn=sqlite3.connect(DB); conn.row_factory=sqlite3.Row; cur=conn.cursor()
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN1_BOOK,page_no))
    r=cur.fetchone(); conn.close()
    if not r or not r['unitext']: return ''
    lines=r['unitext'].split('\n'); sn=str(sutta_num)
    start=None
    for i,line in enumerate(lines):
        s=line.strip()
        if re.match(r'§\s*'+re.escape(sn)+r'\b',s) or re.match(r'^'+re.escape(sn)+r'\.\s+\S',s):
            start=i; break
    if start is None: return ''
    end=len(lines)
    for j in range(start+1,len(lines)):
        s=lines[j].strip()
        if re.match(r'§\s*\d+',s) or (re.match(r'^\d+\.\s+\S',s) and j>start+2):
            end=j; break
    return ' '.join(sh.tokens(' '.join(lines[start:end]))[:max_tokens])

from openpyxl import load_workbook
def get_sn1():
    wb=load_workbook(XL); ws=wb['Complete Canon']
    cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
    tests=[]
    for ri in range(2,ws.max_row+1):
        if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
        snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
        if not snum or snum=='None': continue
        parts=snum.split('.')
        if len(parts)<2 or int(parts[0])>11: continue
        if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='i': continue
        page=ws.cell(row=ri,column=cols['PTS Page']).value
        ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
        name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
        if page: tests.append({'num':snum,'ref':ref,'name':name,'page':page,'sn_inner':parts[1]})
    return tests

def main():
    from openai import OpenAI
    cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')
    cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}
    tests=get_sn1(); segs=sh.load_cha(['s1m.xml'],'h4n')
    
    SYS='You are Helmer Smith. Are these two Pali passages the SAME sutta? SN suttas are short verses. PTS (Feer) is more abbreviated than CST. Same=core verse/dialogue matches. Different=distinct content. Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words English"}'
    
    def ask(p,c):
        r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
            response_format={'type':'json_object'},
            messages=[{'role':'system','content':SYS},
                      {'role':'user','content':'PTS:\n{}\n\nCST:\n{}'.format(p[:800],c[:800])}])
        return json.loads(r.choices[0].message.content)
    
    # Pre-count cached
    cached_app=0
    to_run=[]
    for i,t in enumerate(tests):
        pts=pts_sutta(t['page'],t['sn_inner'])
        if not pts: continue
        cst_idx=i  # sequential 1:1 mapping
        if cst_idx>=len(segs): continue
        cst_text=' '.join(sh.tokens(segs[cst_idx]['text'])[:200])
        key=hashlib.md5((pts[:300]+cst_text[:300]).encode()).hexdigest()
        if key in cache and cache[key].get('verdict')=='APPROVE':
            cached_app+=1
        else:
            to_run.append((i,t))
    
    print('SN I: {} suttas | {} cached APPROVE | {} to validate'.format(len(tests),cached_app,len(to_run)))
    if not to_run: print('All done!'); return
    
    print('='*95)
    results=[]; rej=[]
    for i,t in to_run:
        pts=pts_sutta(t['page'],t['sn_inner'])
        cst_text=' '.join(sh.tokens(segs[i]['text'])[:200])
        cst_title=segs[i].get('title','')
        key=hashlib.md5((pts[:300]+cst_text[:300]).encode()).hexdigest()
        
        if key in cache: hv=cache[key]
        else:
            try: hv=ask(pts,cst_text); cache[key]=hv
            except Exception as e: hv={'verdict':'ERROR','reason':str(e)[:60]}
            CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
        
        results.append((t,hv,cst_title))
        v=hv.get('verdict','?')
        if v=='REJECT': rej.append((t,hv,cst_title))
        if len(results)%30==0:
            print('... {} validated | REJECT: {}'.format(len(results),len(rej)))
    
    # Print REJECTs
    if rej:
        print('\nREJECT entries:')
        for t,hv,cst_title in rej:
            print('  SN {:>8s} {:>18s} | {}'.format(t['num'],t['ref'],hv.get('reason','')[:70]))
    
    # Count ALL (including cached)
    all_app=cached_app
    for _,hv,_ in results:
        if hv.get('verdict')=='APPROVE': all_app+=1
    
    print('\n'+'='*95)
    print('SN I: {} suttas | APPROVE: {} | REJECT: {} | ERROR: {}'.format(
        len(tests),all_app,len(rej),sum(1 for _,hv,_ in results if hv.get('verdict')=='ERROR')))
    print('Precision: {}/{} ({:.0f}%)'.format(all_app,len(tests),100*all_app/max(len(tests),1)))

if __name__=='__main__': main()
