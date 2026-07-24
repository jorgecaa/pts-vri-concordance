#!/usr/bin/env python3
"""SN III (Khandha-vagga) — structure analysis + line numbers."""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=1, column=c).value
    if h: cols[h] = c

entries = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] != 'SN': continue
    if str(row[6] or '').strip() != 'iii': continue
    entries.append({
        'num': str(row[2] or ''),
        'name': str(row[3] or ''),
        'page': row[7],
        'ref': str(row[8] or ''),
        'raw': str(row[13] or ''),
        'row_idx': None,  # filled below
    })

# Map row indices
for ri in range(2, ws.max_row+1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
    if str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip() != 'iii': continue
    sn = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    for e in entries:
        if e['num'] == sn and e['row_idx'] is None:
            e['row_idx'] = ri
            break

print('SN III (Khandha-vagga) —', len(entries), 'entries')
print()

# ── Page sharing ──
pages = defaultdict(list)
for e in entries:
    pages[e['page']].append(e)

shared = {p: v for p, v in pages.items() if len(v) > 1}
print('Pages with >1 sutta:', len(shared), '/', len(pages))
for p in sorted(shared.keys()):
    suttas = shared[p]
    names = ', '.join('SN' + e['num'] for e in suttas)
    print('  p.%d: %d suttas — %s' % (p, len(suttas), names))

# ── First-line analysis (first 20) ──
print()
print('First-line sample:')
for e in entries[:20]:
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=14 AND page_no=? AND edition="mula"', (e['page'],))
    r = cur.fetchone()
    if not r: continue
    lines = (r['unitext'] or '').split('\n')
    head = (r['head'] or '')[:60]
    
    first = ''
    first_ln = 0
    for i, line in enumerate(lines):
        s = line.strip()
        if s and not s.lower().startswith('namo') and not re.match(r'^\-{3,}$', s) and not s.startswith('['):
            first = s[:90]
            first_ln = i + 1
            break
    
    print('  SN %7s  p.%3d L%d  %s' % (e['num'], e['page'], first_ln, first))

# ── Pattern discovery: check first line of all SN III suttas ──
print()
print('Opening pattern summary:')
evam_count = 0
other_count = 0
other_samples = []

for e in entries:
    cur.execute('SELECT unitext FROM pages WHERE book_no=14 AND page_no=? AND edition="mula"', (e['page'],))
    r = cur.fetchone()
    if not r: continue
    txt = (r['unitext'] or '').replace('\n',' ').lower()[:200]
    
    if txt.startswith('evam') or 'evam me suta' in txt[:100]:
        evam_count += 1
    else:
        other_count += 1
        if len(other_samples) < 5:
            other_samples.append((e['num'], e['page'], txt[:80]))

print('  Evaṃ me sutaṃ: %d' % evam_count)
print('  Other opening:  %d' % other_count)
for num, page, txt in other_samples:
    print('    SN %s p.%d: %s' % (num, page, txt))

conn.close()
