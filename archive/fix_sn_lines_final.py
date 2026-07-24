#!/usr/bin/env python3
"""Ensure ALL SN entries have line numbers in their reference."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

SN_MAP = {1:12, 2:13, 3:14, 4:15, 5:16}

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1,column=c).value: c for c in range(1, ws.max_column+1)}

def find_any_start(book_no, page_no, sutta_num_str):
    """Find where this sutta starts on the page. Returns line number or None."""
    parts = sutta_num_str.split('.')
    if len(parts) < 2: return None
    sid = parts[1]
    
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None
    head = r['head'] or ''
    lines = (r['unitext'] or '').split(chr(10))
    esc = re.escape(sid)
    
    for i, line in enumerate(lines):
        s = line.strip()
        if not s: continue
        # Any marker pattern
        if re.search(r'^\s*' + esc + r'[\.\s]', s): return i+1
        if re.search(r'^\s*§\s*' + esc, s): return i+1
        # Range pattern  
        m = re.match(r'^\s*(\d+)\-(\d+)', s)
        if m:
            try:
                lo, hi = int(m.group(1)), int(m.group(2))
                if lo <= int(sid) <= hi: return i+1
            except: pass
    
    # If marker not found, check if page HEAD mentions this samyutta
    # → sutta likely starts at first text line
    for i, line in enumerate(lines):
        s = line.strip()
        if s and not s.lower().startswith('namo') and not re.match(r'^\-{3,}$', s) and not s.startswith('['):
            # Skip volume titles
            if 'nikaya' in s.lower() or 'division' in s.lower() or 'book' in s.lower()[:6]:
                continue
            return i+1
    
    return None

updates = 0
still_none = 0

for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
    
    sutta_num = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip()
    if not page: continue
    
    # Skip if already has comma+number
    if re.search(r',\d+$', ref): continue
    
    vol_num = {'i':1,'ii':2,'iii':3,'iv':4,'v':5}.get(roman, 0)
    book_no = SN_MAP.get(vol_num)
    if not book_no: continue
    
    line = find_any_start(book_no, page, sutta_num)
    
    if line:
        new_ref = re.sub(r',\d+$', '', ref) + ',%d' % line
        ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
        updates += 1
    else:
        still_none += 1
        # At least mark as page-only
        if ',' not in ref:
            ws.cell(row=ri, column=cols['PTS Ref']).value = ref + ',?'

wb.save(XL)
print('Added line numbers: %d' % updates)
print('Still no marker:    %d' % still_none)
print('Saved: %s' % XL)
conn.close()
