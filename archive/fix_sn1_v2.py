#!/usr/bin/env python3
"""SN1 page fix v2: sequential constraint + local fingerprint search."""
import re, sqlite3, json
from collections import defaultdict
import sutta_hash as sh
import sutta_fingerprint as fp

DB='src/data/tipitaka.sqlite'; XL='PTS_Reference_Complete_Canon.xlsx'
SN1_BOOK=12

def page_text(page_no):
    conn=sqlite3.connect(DB); conn.row_factory=sqlite3.Row; cur=conn.cursor()
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN1_BOOK,page_no))
    r=cur.fetchone(); conn.close()
    return r['unitext'] if r and r['unitext'] else ''

from openpyxl import load_workbook
def get_sn1():
    wb=load_workbook(XL); ws=wb['Complete Canon']
    cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
    tests=[]
    for ri in range(2,ws.max_row+1):
        if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
        snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
        if not snum or snum=='None': continue
        parts=snum.split('.')
        if len(parts)<2 or int(parts[0])>11: continue
        if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='i': continue
        page=ws.cell(row=ri,column=cols['PTS Page']).value
        ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
        name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
        m=re.search(r',(\d+)$',ref); line=int(m.group(1)) if m else 1
        tests.append({'num':snum,'ref':ref,'name':name,'page':page,'line':line,'ri':ri})
    return tests

print('Loading...')
tests=get_sn1()
segs=sh.load_cha(['s1m.xml'],'h4n')
st=sh.Stemmer()

# Build CST fingerprints
cst_pool=[]
for s in segs:
    seq=sh.stem_seq(s['text'],st) if s['text'] else []
    cst_pool.append({'seq':seq})
cst_idf=fp.global_idf([s['seq'] for s in cst_pool])
for s in cst_pool: s['sig']=fp.prose_signature(s['seq'],cst_idf,32,6)

# Precompute page fingerprints
conn=sqlite3.connect(DB); cur=conn.cursor()
cur.execute('SELECT MAX(page_no) FROM pages WHERE book_no=? AND edition="mula"',(SN1_BOOK,))
max_page=cur.fetchone()[0]; conn.close()

page_fps={}
for p in range(1,max_page+1):
    txt=page_text(p)
    if txt:
        seq=sh.stem_seq(txt,st)
        page_fps[p]=seq

page_idf=fp.global_idf(list(page_fps.values()))
for p in page_fps: page_fps[p]=fp.prose_signature(page_fps[p],page_idf,32,6)

# Sequential constraint: pages must be non-decreasing
# Find entries that break monotonicity
issues=[]
prev_page=0
for i,t in enumerate(tests):
    page=t['page']
    if page and page<prev_page-1:  # allow -1 for shared pages going back 1
        issues.append((i,t,prev_page))
    if page: prev_page=page

print('Sequential issues (page decreases >1): {}'.format(len(issues)))
for i,t,prev in issues[:15]:
    print('  SN {:>8s} | p.{:>3d} (prev was p.{}) | {}'.format(t['num'],t['page'],prev,t['name'][:30]))

# For each issue, search nearby pages with fingerprint
print('\nResolving with local fingerprint search (±5 pages)...')
fixes=[]
for i,t,prev in issues:
    cst_sig=cst_pool[i]['sig']
    # Search pages from prev-2 to prev+10
    best_page,best_score=t['page'],0
    for p in range(max(1,prev-2), min(max_page,prev+15)):
        if p not in page_fps: continue
        score=fp.prose_score(cst_sig,page_fps[p],page_idf)
        if score>best_score: best_score=score; best_page=p
    
    if best_page!=t['page']:
        fixes.append((i,t,best_page,best_score))
        print('  SN {:>8s} | p.{:>3d} -> p.{:>3d} | score={:.3f} | {}'.format(
            t['num'],t['page'],best_page,best_score,t['name'][:30]))
    else:
        print('  SN {:>8s} | p.{:>3d} KEPT | {}'.format(t['num'],t['page'],t['name'][:30]))

if fixes:
    wb=load_workbook(XL); ws=wb['Complete Canon']
    cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
    for i,t,new_page,score in fixes:
        ri=t['ri']
        old_ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
        ws.cell(row=ri,column=cols['PTS Page']).value=new_page
        new_ref=re.sub(r'S i \d+','S i {}'.format(new_page),old_ref)
        new_ref=re.sub(r',\d+$','',new_ref)
        ws.cell(row=ri,column=cols['PTS Ref']).value=new_ref
        ws.cell(row=ri,column=cols['Validation']).value='SEQ_FIXED'
    wb.save(XL)
    print('\nApplied {} fixes'.format(len(fixes)))
