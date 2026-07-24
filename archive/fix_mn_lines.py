#!/usr/bin/env python3
"""Fix MN line numbers: for suttas where first line ≠ Evaṃ, find the centered number in body."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

MN_BOOKS = {1:9, 2:10, 3:11}
ROMAN = {'i':1,'ii':2,'iii':3}

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=1, column=c).value
    if h: cols[h] = c

def is_skip(s):
    s = s.strip()
    if not s: return True
    if re.match(r'^\d+\.?$', s): return True
    if re.match(r'^\-{5,}$', s): return True
    return False

def get_first_text(book, page, start_line):
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book, page))
    r = cur.fetchone()
    if not r: return None, ''
    lines = (r['unitext'] or '').split('\n')
    for i in range(start_line-1, len(lines)):
        s = lines[i].strip()
        if is_skip(s): continue
        return i+1, re.sub(r'\s+',' ',s).strip()[:130]
    return None, ''

def find_body_marker(book, page, num_str):
    """Find centered number marker in body (not HEAD)."""
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book, page))
    r = cur.fetchone()
    if not r: return None
    lines = (r['unitext'] or '').split('\n')
    for i, line in enumerate(lines):
        s = line.strip()
        if s == f'{num_str}.' or s == num_str:
            return i + 1
    return None

fixes = 0
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'MN': continue
    
    sn = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip()
    
    vol = ROMAN.get(roman.lower(), 1)
    book = MN_BOOKS.get(vol)
    if not book or not page: continue
    
    # Get current line
    cur_line = 1
    m = re.search(r',(\d+)$', ref)
    if m: cur_line = int(m.group(1))
    
    # Get first text
    txt_line, first_txt = get_first_text(book, page, cur_line)
    if not first_txt: continue
    
    # Check if it starts with Evaṃ
    tlow = first_txt.lower()
    if tlow.startswith('evam'):
        continue  # Already correct
    
    # Need to find the real start — search body for centered number
    marker_line = find_body_marker(book, page, sn)
    if marker_line and marker_line > cur_line:
        # The sutta starts after the marker
        new_line = marker_line + 1
        new_txt_line, new_txt = get_first_text(book, page, new_line)
        
        new_ref = f'M {roman} {page}'
        if new_line > 1:
            new_ref += f',{new_line}'
        
        old_ref = ws.cell(row=ri, column=cols['PTS Ref']).value
        ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
        
        print(f'  MN {sn:>3s}  {old_ref:16s} → {new_ref:16s}  {new_txt[:90] if new_txt else "?"}')
        fixes += 1
    elif marker_line:
        # Marker found but at same or earlier line — text starts right after
        new_line = marker_line + 1
        new_txt_line, new_txt = get_first_text(book, page, new_line)
        
        new_ref = f'M {roman} {page}'
        if new_line > 1:
            new_ref += f',{new_line}'
        
        old_ref = ws.cell(row=ri, column=cols['PTS Ref']).value
        if old_ref != new_ref:
            ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
            print(f'  MN {sn:>3s}  {old_ref:16s} → {new_ref:16s}  {new_txt[:90] if new_txt else "?"}')
            fixes += 1

wb.save(XL)
print(f'\n  Fixed: {fixes} entries')
print(f'  Saved: {XL}')
conn.close()
