#!/usr/bin/env python3
"""Fix SN1 pages using fingerprint matching against CST ground truth."""
import re, sqlite3, json
from collections import defaultdict
from pathlib import Path
import sutta_hash as sh
import sutta_fingerprint as fp

DB='src/data/tipitaka.sqlite'; XL='PTS_Reference_Complete_Canon.xlsx'
SN1_BOOK=12

def page_text(page_no):
    conn=sqlite3.connect(DB); conn.row_factory=sqlite3.Row; cur=conn.cursor()
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN1_BOOK,page_no))
    r=cur.fetchone(); conn.close()
    return r['unitext'] if r and r['unitext'] else ''

from openpyxl import load_workbook
def get_sn1():
    wb=load_workbook(XL); ws=wb['Complete Canon']
    cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
    tests=[]
    for ri in range(2,ws.max_row+1):
        if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
        snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
        if not snum or snum=='None': continue
        parts=snum.split('.')
        if len(parts)<2 or int(parts[0])>11: continue
        if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='i': continue
        page=ws.cell(row=ri,column=cols['PTS Page']).value
        ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
        name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
        m=re.search(r',(\d+)$',ref); line=int(m.group(1)) if m else 1
        tests.append({'num':snum,'ref':ref,'name':name,'page':page,'line':line,'ri':ri})
    return tests

print('Loading SN1 data...')
tests=get_sn1()
segs=sh.load_cha(['s1m.xml'],'h4n')
st=sh.Stemmer()

# Build CST fingerprint pool (ground truth)
print('Building CST fingerprints...')
cst_pool=[]
for s in segs:
    seq=sh.stem_seq(s['text'],st) if s['text'] else []
    cst_pool.append({'text':s['text'],'seq':seq,'title':s.get('title','')})

cst_idf=fp.global_idf([s['seq'] for s in cst_pool])
for s in cst_pool:
    s['sig']=fp.prose_signature(s['seq'],cst_idf,32,6)

# Get max page for SN vol 1
conn=sqlite3.connect(DB); cur=conn.cursor()
cur.execute('SELECT MAX(page_no) FROM pages WHERE book_no=? AND edition="mula"',(SN1_BOOK,))
max_page=cur.fetchone()[0]; conn.close()
print('SN I: {} suttas, pages 1-{}'.format(len(tests),max_page))

# Pre-compute page fingerprints for all pages
print('Building PTS page fingerprints...')
page_fps={}
for p in range(1,max_page+1):
    txt=page_text(p)
    if txt:
        seq=sh.stem_seq(txt,st)
        page_fps[p]=seq

page_idf=fp.global_idf(list(page_fps.values()))
for p in page_fps:
    page_fps[p]=fp.prose_signature(page_fps[p],page_idf,32,6)

# For each CST sutta, find best matching PTS page
print('Matching CST suttas to PTS pages...')
fixes=[]
no_match=[]
for i in range(min(len(tests),len(cst_pool))):
    cst_sig=cst_pool[i]['sig']
    best_page,best_score=0,-1
    for p,sig in page_fps.items():
        score=fp.prose_score(cst_sig,sig,page_idf)
        if score>best_score: best_score=score; best_page=p
    
    excel_page=tests[i]['page']
    if best_page!=excel_page and best_score>0.1:
        fixes.append((i,tests[i],best_page,best_score,cst_pool[i]['title']))
    elif best_score<=0.1:
        no_match.append((i,tests[i],best_score))

print('\nPage corrections needed: {}'.format(len(fixes)))
print('No match (score<=0.1): {}'.format(len(no_match)))

# Show top corrections
for i,t,new_page,score,title in fixes[:30]:
    print('  SN {:>8s} | Excel: p.{:>3d} -> Correct: p.{:>3d} | score={:.3f} | CST: {}'.format(
        t['num'],t['page'],new_page,score,title[:35]))

# Apply fixes
if fixes:
    wb=load_workbook(XL); ws=wb['Complete Canon']
    cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
    applied=0
    for i,t,new_page,score,title in fixes:
        ri=t['ri']
        old_page=ws.cell(row=ri,column=cols['PTS Page']).value
        old_ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
        if old_page!=new_page:
            ws.cell(row=ri,column=cols['PTS Page']).value=new_page
            # Update ref
            new_ref=re.sub(r'S i \d+','S i {}'.format(new_page),old_ref)
            new_ref=re.sub(r',\d+$','',new_ref)  # remove old line number (needs recalculation)
            ws.cell(row=ri,column=cols['PTS Ref']).value=new_ref
            ws.cell(row=ri,column=cols['Validation']).value='FP_CORRECTED'
            applied+=1
    wb.save(XL)
    print('\nApplied {} page corrections to Excel'.format(applied))
