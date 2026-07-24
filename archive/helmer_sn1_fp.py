#!/usr/bin/env python3
"""SN1 validation using sutta_hash fingerprints (no DeepSeek needed)."""
import re, sqlite3, json
from collections import defaultdict
from pathlib import Path
import sutta_hash as sh
import sutta_fingerprint as fp

DB='src/data/tipitaka.sqlite'; XL='PTS_Reference_Complete_Canon.xlsx'
SN1_BOOK=12

def pts_sutta(page_no, sutta_num):
    conn=sqlite3.connect(DB); conn.row_factory=sqlite3.Row; cur=conn.cursor()
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN1_BOOK,page_no))
    r=cur.fetchone(); conn.close()
    if not r or not r['unitext']: return ''
    lines=r['unitext'].split('\n'); sn=str(sutta_num)
    start=None
    for i,line in enumerate(lines):
        s=line.strip()
        if re.match(r'§\s*'+re.escape(sn)+r'\b',s) or re.match(r'^'+re.escape(sn)+r'\.\s+\S',s):
            start=i; break
    if start is None: return ''
    end=len(lines)
    for j in range(start+1,len(lines)):
        s=lines[j].strip()
        if re.match(r'§\s*\d+',s) or (re.match(r'^\d+\.\s+\S',s) and j>start+2):
            end=j; break
    return ' '.join(lines[start:end])

from openpyxl import load_workbook
def get_sn1():
    wb=load_workbook(XL); ws=wb['Complete Canon']
    cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
    tests=[]
    for ri in range(2,ws.max_row+1):
        if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
        snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
        if not snum or snum=='None': continue
        parts=snum.split('.')
        if len(parts)<2 or int(parts[0])>11: continue
        if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='i': continue
        page=ws.cell(row=ri,column=cols['PTS Page']).value
        ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
        name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
        if page: tests.append({'num':snum,'ref':ref,'name':name,'page':page,'sn_inner':parts[1]})
    return tests

print('Loading SN1 data...')
tests=get_sn1()
segs=sh.load_cha(['s1m.xml'],'h4n')
st=sh.Stemmer()

# Build PTS sutta pool
print('Computing PTS fingerprints...')
pts_texts=[]
for t in tests:
    txt=pts_sutta(t['page'],t['sn_inner'])
    pts_texts.append(txt if txt else '')

# Build CST sutta pool
print('Computing CST fingerprints...')
cst_texts=[s['text'] for s in segs]

# Stem and compute fingerprints
pool=[]
for txt in pts_texts+cst_texts:
    seq=sh.stem_seq(txt,st) if txt else []
    pool.append({'text':txt,'seq':seq})

idf=fp.global_idf([s['seq'] for s in pool])
for s in pool:
    s['sig']=fp.prose_signature(s['seq'],idf,32,6)

n=len(pts_texts)
print('Comparing {} PTS suttas against {} CST suttas...'.format(n,len(cst_texts)))
print('='*95)

# For each PTS sutta, find best CST match
scores=[]
matches=[]
for i in range(n):
    if not pts_texts[i]:
        matches.append((i,None,0,''))
        continue
    pts_sig=pool[i]['sig']
    best_score,best_j=-1,-1
    for j in range(len(cst_texts)):
        if not cst_texts[j]: continue
        cst_sig=pool[n+j]['sig']
        score=fp.prose_score(pts_sig,cst_sig,idf)
        if score>best_score: best_score=score; best_j=j
    matches.append((i,best_j,best_score,segs[best_j].get('title','') if best_j>=0 else ''))
    scores.append(best_score)

# Analyze
correct=sum(1 for i,j,s,_ in matches if j==i and s>0.1)
close=sum(1 for i,j,s,_ in matches if j!=i and s>0.1 and abs(j-i)<=2)
wrong=sum(1 for i,j,s,_ in matches if j!=i and s>0.1 and abs(j-i)>2)
no_match=sum(1 for i,j,s,_ in matches if s<=0.1)
empty=sum(1 for i,j,s,_ in matches if j is None)

print('Results (fingerprint matching):')
print('  Correct (j==i, same position): {}'.format(correct))
print('  Close (off by <=2):             {}'.format(close))
print('  Wrong (off by >2):              {}'.format(wrong))
print('  No match (score <= 0.1):        {}'.format(no_match))
print('  Empty PTS text:                 {}'.format(empty))
print('  Accuracy: {}/{} ({:.0f}%)'.format(correct+close,len(tests),100*(correct+close)/len(tests)))

# Show wrong matches
if wrong>0:
    print('\nWrong matches (>2 positions off):')
    for i,j,s,title in matches:
        if j is not None and j!=i and abs(j-i)>2 and s>0.1:
            print('  SN {:>8s} -> CST[{}] {:40s} score={:.3f}'.format(tests[i]['num'],j,title[:40],s))

# Score distribution
from collections import Counter
bins=Counter()
for s in scores:
    if s>0.5: bins['>0.5']+=1
    elif s>0.3: bins['0.3-0.5']+=1
    elif s>0.1: bins['0.1-0.3']+=1
    else: bins['<0.1']+=1
print('\nScore distribution:')
for k in ['>0.5','0.3-0.5','0.1-0.3','<0.1']:
    print('  {}: {}'.format(k,bins.get(k,0)))
