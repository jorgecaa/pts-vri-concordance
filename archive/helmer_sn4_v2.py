#!/usr/bin/env python3
"""SN IV Helmer v2: process by vagga. Markers at vagga starts only."""
import hashlib, json, os, re, sqlite3
from collections import Counter, defaultdict
from pathlib import Path
import sutta_hash as sh

CACHE_FILE=Path('helmer_ptscst_cache.json'); SN4_BOOK=15

conn=sqlite3.connect('src/data/tipitaka.sqlite'); conn.row_factory=sqlite3.Row; cur=conn.cursor()

# Build marker map
marker_map=defaultdict(list)
cur.execute('SELECT page_no, unitext FROM pages WHERE book_no=? AND edition="mula"',(SN4_BOOK,))
for r in cur.fetchall():
    for i,line in enumerate(r['unitext'].split('\n')):
        s=line.strip()
        m=re.match(r'^(\d+)\.?\s*\((\d+)\)\s+(\S.*)',s)
        if m:
            marker_map[r['page_no']].append((i+1,int(m.group(1)),int(m.group(2)),m.group(3)[:60]))

# Load Excel
from openpyxl import load_workbook
wb=load_workbook('PTS_Reference_Complete_Canon.xlsx'); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
entries=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2: continue
    if int(parts[0])<35 or int(parts[0])>44: continue
    if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='iv': continue
    page=ws.cell(row=ri,column=cols['PTS Page']).value
    ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
    name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
    entries.append({'num':snum,'ref':ref,'name':name,'page':page,'ri':ri,
                    'sn':int(parts[0]),'sn_inner':int(parts[1])})

# Group entries by saṃyutta, then by vagga (~10 suttas per vagga)
by_sn=defaultdict(list)
for e in entries: by_sn[e['sn']].append(e)

# For each saṃyutta, find all markers sorted by gid
all_markers=sorted([(p,l,g,v,n) for p in marker_map for l,g,v,n in marker_map[p]],key=lambda x:x[2])

print('SN IV: {} entries across {} saṃyuttas'.format(len(entries),len(by_sn)))
print('Total markers: {}'.format(len(all_markers)))

# For each entry, find the CLOSEST marker by gid (marker before or at this sutta)
tasks=[]
no_marker=[]
for e in entries:
    target=e['sn_inner']
    # Find the marker with largest gid <= target, or closest
    best_marker=None; best_dist=999
    for p,l,g,v,n in all_markers:
        # Only consider markers on the same page or nearby
        if abs(p-e['page'])<=3:
            if g<=target and target-g<best_dist:
                best_dist=target-g; best_marker=(p,l,g,v,n)
            elif g>target and g-target<best_dist:
                best_dist=g-target; best_marker=(p,l,g,v,n)
    
    if best_marker and best_dist<=5:  # within 5 suttas of a marker
        p,l,g,v,n=best_marker
        # Extract text from this marker
        cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN4_BOOK,p))
        r=cur.fetchone()
        if r:
            lines=r['unitext'].split('\n')
            start=l-1; end=len(lines)
            for j in range(start+1,len(lines)):
                if re.match(r'^\d+\.?\s*\(\d+\)',lines[j].strip()): end=j; break
            text=' '.join(lines[start:end])
            tasks.append((e,' '.join(sh.tokens(text)[:300]),l,g,v,n[:30],best_dist))
    else:
        no_marker.append(e)

print('Matched to nearest marker: {} | No nearby marker: {}'.format(len(tasks),len(no_marker)))

# Sample: first 3 entries of each saṃyutta for DeepSeek pilot
print('\n--- DeepSeek pilot: first 3 of each saṃyutta ---')
sample=[t for t in tasks if t[0]['sn_inner']<=3]
print('Sample size: {}'.format(len(sample)))

segs=sh.load_cha(['s4m.xml'],'h4n')
cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}
from openai import OpenAI
cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')
SYS='You are Helmer Smith. Are these two Pali passages the SAME sutta? Same=core teaching matches. Different=distinct sutta. Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words"}'

vc=Counter()
for i,(e,pts,l,g,v,nm,dist) in enumerate(sample):
    # CST: same sequential position
    cst_idx=sum(1 for ee in entries if ee['sn']<e['sn'] or (ee['sn']==e['sn'] and ee['sn_inner']<e['sn_inner']))
    if cst_idx>=len(segs): continue
    cst_text=' '.join(sh.tokens(segs[cst_idx]['text'])[:300])
    cst_title=segs[cst_idx].get('title','')
    
    key=hashlib.md5((pts[:200]+cst_text[:200]).encode()).hexdigest()
    if key in cache: hv=cache[key]
    else:
        try:
            r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
                response_format={'type':'json_object'},
                messages=[{'role':'system','content':SYS},
                          {'role':'user','content':'PTS (gid={}, dist={}):\n{}\n\nCST:\n{}'.format(g,dist,pts[:800],cst_text[:800])}])
            hv=json.loads(r.choices[0].message.content)
            cache[key]=hv; CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
        except Exception as ex: hv={'verdict':'ERROR','reason':str(ex)[:60]}
    
    v=hv.get('verdict','?'); vc[v]+=1
    sym='✓' if v=='APPROVE' else '✗' if v=='REJECT' else '?'
    print('{} SN {:>8s} {:>15s} | {:8s} | gid={} dist={} | CST[{}] {} | {}'.format(
        sym,e['num'],e['ref'],v,g,dist,cst_idx,cst_title[:25],hv.get('reason','')[:50]))

print('\nAPPROVE: {} | REJECT: {} | Precision: {:.0f}%'.format(
    vc.get('APPROVE',0),vc.get('REJECT',0),
    100*vc.get('APPROVE',0)/(vc.get('APPROVE',0)+vc.get('REJECT',0)) if vc.get('APPROVE',0)+vc.get('REJECT',0)>0 else 0))
conn.close()
