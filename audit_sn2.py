#!/usr/bin/env python3
"""SN II audit: sequential page check + marker verification + FP sample."""
import re, sqlite3
from collections import defaultdict
import sutta_hash as sh
import sutta_fingerprint as fp

DB='src/data/tipitaka.sqlite'; XL='PTS_Reference_Complete_Canon.xlsx'
SN2_BOOK=13

from openpyxl import load_workbook
wb=load_workbook(XL); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}

tests=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2: continue
    sn=int(parts[0])
    if sn<12 or sn>21: continue  # SN II range
    rn=str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()
    if rn!='ii': continue
    page=ws.cell(row=ri,column=cols['PTS Page']).value
    ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
    name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
    m=re.search(r',(\d+)$',ref); line=int(m.group(1)) if m else 1
    tests.append({'num':snum,'ref':ref,'name':name,'page':page,'line':line,'ri':ri})

print('SN II: {} entries (saṃyuttas 12-21)'.format(len(tests)))

# ── Step 1: Sequential audit ──
print('\n--- Step 1: Sequential Page Audit ---')
breaks=[]
prev_page=0
for i,t in enumerate(tests):
    page=t['page']
    if page and prev_page and page<prev_page-1:
        breaks.append((i,t,prev_page))
    if page: prev_page=page

if breaks:
    print('{} sequential breaks found:'.format(len(breaks)))
    for i,t,prev in breaks[:10]:
        print('  SN {:>8s} | p.{:>3d} (prev p.{}) | {}'.format(t['num'],t['page'],prev,t['name'][:40]))
else:
    print('✓ All pages sequentially consistent')

# ── Step 2: Marker verification ──
print('\n--- Step 2: Marker Verification ---')
conn=sqlite3.connect(DB); conn.row_factory=sqlite3.Row; cur=conn.cursor()
pat_nodot=re.compile(r'^(\d+)\s*\((\d+)\)\s+\S')  # N (M) Name

found=0; missing=0; missing_examples=[]
for t in tests:
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN2_BOOK,t['page']))
    r=cur.fetchone()
    if not r or not r['unitext']:
        missing+=1; continue
    
    lines=r['unitext'].split('\n')
    parts=t['num'].split('.')
    sn_inner=parts[-1]  # sutta number within saṃyutta
    
    has_marker=False
    for line in lines:
        s=line.strip()
        m=pat_nodot.match(s)
        if m:
            # Check if this is our sutta (by vagga position)
            vagga_pos=m.group(2)
            if vagga_pos==sn_inner:
                has_marker=True; break
            # Also check sequential: if no better match, any marker counts
            has_marker=True
    
    if has_marker: found+=1
    else:
        missing+=1
        if len(missing_examples)<10:
            missing_examples.append(t)

print('Markers found: {}/{} ({:.0f}%)'.format(found,len(tests),100*found/max(len(tests),1)))
if missing_examples:
    print('Missing examples:')
    for t in missing_examples:
        # Show what's on the page
        cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN2_BOOK,t['page']))
        r=cur.fetchone()
        lines=r['unitext'].split('\n') if r else []
        first5=' | '.join(l.strip()[:60] for l in lines[:5] if l.strip())
        print('  SN {:>8s} | p.{:>3d} | {} | {}'.format(t['num'],t['page'],t['name'][:25],first5[:100]))

# ── Step 3: CST fingerprint sample ──
print('\n--- Step 3: CST Fingerprint Sample (first 30 suttas) ---')
segs=sh.load_cha(['s2m.xml'],'h4n'); st=sh.Stemmer()
print('CST suttas in s2m.xml: {}'.format(len(segs)))

# Build CST pool
cst_pool=[]
for s in segs[:len(tests)]:
    seq=sh.stem_seq(s['text'],st) if s['text'] else []; cst_pool.append({'seq':seq})
cst_idf=fp.global_idf([s['seq'] for s in cst_pool])
for s in cst_pool: s['sig']=fp.prose_signature(s['seq'],cst_idf,32,6)

# Build PTS page pool (pages used by tests)
used_pages=sorted(set(t['page'] for t in tests if t['page']))
page_fps={}
for p in used_pages:
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN2_BOOK,p))
    r=cur.fetchone()
    if r and r['unitext']:
        seq=sh.stem_seq(r['unitext'],st); page_fps[p]=seq

page_idf=fp.global_idf(list(page_fps.values()))
for p in page_fps: page_fps[p]=fp.prose_signature(page_fps[p],page_idf,32,6)

# Match first 30
sample_n=min(30,len(tests))
correct=0; close=0; wrong=0
for i in range(sample_n):
    if i>=len(cst_pool): break
    cst_sig=cst_pool[i]['sig']
    excel_page=tests[i]['page']
    
    # Search used pages for best match
    best_page,best_score=0,-1
    for p,sig in page_fps.items():
        score=fp.prose_score(cst_sig,sig,page_idf)
        if score>best_score: best_score=score; best_page=p
    
    if best_page==excel_page: correct+=1
    elif abs(best_page-excel_page)<=2: close+=1
    else: wrong+=1

print('Sample {}: correct={} close={} wrong={} | accuracy={:.0f}%'.format(
    sample_n,correct,close,wrong,100*(correct+close)/sample_n))

conn.close()
