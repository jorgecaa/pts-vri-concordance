#!/usr/bin/env python3
"""
Corrige errores de paginación en la tabla PTS Khuddaka Nikaya.

Audita TODAS las secciones comparando el nombre del sutta contra el
contenido real en tipitaka.sqlite, usando fingerprints personalizados
por tipo de texto. Aplica solo correcciones verificadas (Δ=+1 o +2).

Uso:  python fix_pts_errors.py [input.xlsx] [output.xlsx]
"""

import sys, re, sqlite3, base64
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Paths ──
DB_PATH = Path(__file__).parent / "src" / "data" / "tipitaka.sqlite"
EXCEL_IN = Path(__file__).parent / "PTS_Reference_Khuddaka_Nikaya.xlsx"
EXCEL_OUT = Path(__file__).parent / "PTS_Reference_Khuddaka_Nikaya_CORRECTED.xlsx"
args = [a for a in sys.argv[1:] if not a.startswith("--")]
if args: EXCEL_IN = Path(args[0])
if len(args) > 1: EXCEL_OUT = Path(args[1])

# ── DB ──
def decode(val):
    if not val: return ''
    try:
        raw = base64.b64decode(val.strip() + '==')
        if raw[:3] == b'\xef\xbb\xbf': raw = raw[3:]
        return raw.decode('utf-8', errors='replace')
    except: return str(val)[:80]

def norm(s):
    s = s.lower()
    for a,b in [('ā','a'),('ī','i'),('ū','u'),('ṅ','n'),('ñ','n'),
                ('ṭ','t'),('ḍ','d'),('ṇ','n'),('ḷ','l'),('ṃ','m')]:
        s = s.replace(a,b)
    return re.sub(r'[^a-z0-9]','',s)

def clean_name(name):
    c = name
    c = re.sub(r'^\([^)]+\)\s*', '', c)
    c = re.sub(r'^\([^)]+\)\s*', '', c)
    c = re.sub(r'^[A-Za-zāīūṅñṭḍṇḷṃ]+\.?\s*[\d.]+\s+', '', c)
    c = re.sub(r'^[A-Za-zāīūṅñṭḍṇḷṃ]+\s+Ap\s+\d+\s+', '', c)
    c = re.sub(r'^Nd\s+\d+\s+\d+\s+', '', c)
    c = re.sub(r'^Ps\s+[\d.]+\s+', '', c)
    return c.strip()

# ── Mapa ──
VOL2BOOK = {
    "Kh":22,"Khp":22,"Dh":23,"Dhp":23,"Ud":24,"It":25,"Sn":26,"Snp":26,
    "Vv":27,"Pv":28,"Th":29,"Thag":29,"Thī":29,"Thig":29,
    "Ja i":30,"Ja ii":31,"Ja iii":32,"Ja iv":33,"Ja v":34,"Ja vi":35,
    "Nidd i":36,"Nidd ii":37,"Paṭis i":38,"Paṭis ii":39,
    "Ap i":40,"Ap ii":40,"Bv":41,"Bv & Cp":41,"Cp":42,
}

# ── Fingerprint strategies ──
def fp_words(nn, ml=4, mc=10):
    w = [x for x in nn.split() if len(x)>=ml]
    if w: w.sort(key=len, reverse=True); return w[0][:mc]
    return nn[:mc]

STRATEGIES = {
    'Khp': lambda c,n: (fp_words(n,4,10), False, None),
    'Dhp': lambda c,n: (fp_words(n.replace('vagga',''),5,10), False, None),
    'Ud':  lambda c,n: (fp_words(n,4,10), False, None),
    'It':  lambda c,n: (fp_words(n,4,10), False, None),
    'Sn':  lambda c,n: (fp_words(n,4,10), False, None),
    'Vv':  lambda c,n: (fp_words(n.replace('vimana','').replace('vatthu',''),5,10), False, None),
    'Pv':  lambda c,n: (fp_words(n.replace('peta','').replace('vatthu',''),5,10), False, None),
    'Thag':lambda c,n: (fp_words(n,4,10), True, None),
    'Thig':lambda c,n: (fp_words(n,4,10), True, None),
    'Ja':  lambda c,n: (n.replace('jataka','')[:12], False,
                        int(m.group(1)) if (m:=re.match(r'(\d+)',c)) else None),
    'Nd1': lambda c,n: (fp_words(n,4,10), True, None),
    'Nd2': lambda c,n: (fp_words(n,4,10), True, None),
    'Ps':  lambda c,n: (fp_words(n,4,10), True, None),
    'Ap':  lambda c,n: (fp_words(n.replace('theraapadana','').replace('theriapadana','').replace('apadana',''),5,12), False, None),
    'Bv':  lambda c,n: (fp_words(n.replace('buddhavamsa',''),5,10), False, None),
    'BvCp':lambda c,n: (fp_words(n.replace('cariya',''),5,10), False, None),
    'Cp':  lambda c,n: (fp_words(n.replace('cariya',''),5,10), False, None),
}

def find_page(con, db_book, ref_page, fp, head_only, num_hint):
    for delta in range(5):
        pg = ref_page + delta
        row = con.execute(
            "SELECT head, unitext FROM pages WHERE edition='mula' AND book_no=? AND page_no=?",
            (db_book, pg)).fetchone()
        if not row: break
        h = norm(decode(row[0] or '')[:120])
        t = norm(decode(row[1] or '')[:600])
        if head_only:
            if fp and fp in h: return pg, delta
        else:
            if fp and fp in (h + t): return pg, delta
        if num_hint:
            hr = decode(row[0] or '')
            if any(p in hr for p in [f"({num_hint})",f"({num_hint}.)",f" {num_hint}. "]):
                return pg, delta
    return None, None

# ── Styles ──
FIX_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

def main():
    print(f"Input:  {EXCEL_IN}")
    print(f"Output: {EXCEL_OUT}\n")

    con = sqlite3.connect(str(DB_PATH))
    wb = load_workbook(str(EXCEL_IN))
    ws = wb["PTS Reference"]

    col_k = 11
    k = ws.cell(row=1, column=col_k)
    k.value = "Corrección"
    k.font = Font(bold=True, size=11, color="FFFFFF")
    k.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    k.alignment = Alignment(horizontal="center", vertical="center")
    k.border = thin
    ws.column_dimensions["K"].width = 52

    fixed = 0
    for row_idx in range(2, ws.max_row + 1):
        name = str(ws.cell(row=row_idx, column=2).value or '').strip()
        pts_vol = str(ws.cell(row=row_idx, column=4).value or '').strip()
        rtype = str(ws.cell(row=row_idx, column=7).value or '').strip()
        section = str(ws.cell(row=row_idx, column=6).value or '').strip()
        if rtype in ("Book Header", "Section Header"): continue
        if not name or not pts_vol: continue

        v = pts_vol.strip()
        if "," in v: v = v.split(",")[0].strip()
        v = re.sub(r"\s+"," ",v)
        m = re.match(r"^(.+?)\s+(\d+)$", v)
        if not m: continue
        vol_abbr, ref_page = m.group(1).strip(), int(m.group(2))
        db_book = VOL2BOOK.get(vol_abbr)
        if db_book is None: continue

        clean = clean_name(name)
        name_norm = norm(clean)

        # Determinar sección desde el nombre del libro
        if 'Ja' in section or name.startswith('Ja '):
            sec = 'Ja'
        elif 'Khuddakap' in section:
            sec = 'Khp'
        elif 'Dhammapada' in section:
            sec = 'Dhp'
        elif 'Ud' in section and 'Udāna' in section:
            sec = 'Ud'
        elif 'Itivuttaka' in section:
            sec = 'It'
        elif 'Suttanipāta' in section:
            sec = 'Sn'
        elif 'Vimāna' in section:
            sec = 'Vv'
        elif 'Petavatthu' in section:
            sec = 'Pv'
        elif 'Theragāthā' == section.replace(' (Thag)',''):
            sec = 'Thag'
        elif 'Therīgāthā' == section.replace(' (Thig)',''):
            sec = 'Thig'
        elif 'Mahā-Niddesa' in section:
            sec = 'Nd1'
        elif 'Cūḷa-Niddesa' in section:
            sec = 'Nd2'
        elif 'Paṭisambhidāmagga' in section:
            sec = 'Ps'
        elif 'Therāpadāna' in section:
            sec = 'Ap'
        elif 'Therīapadāna' in section:
            sec = 'Ap'
        elif 'Buddhavaṃsa' in section:
            sec = 'Bv'
        elif 'Cariyāpiṭaka' in section:
            sec = 'BvCp'
        else:
            continue

        strat = STRATEGIES.get(sec, lambda c,n: (fp_words(n), False, None))
        fp, head_only, num_hint = strat(clean, name_norm)
        if not fp: continue

        found_pg, delta = find_page(con, db_book, ref_page, fp, head_only, num_hint)
        if found_pg is None or delta <= 0 or delta > 2:
            continue

        new_vol = f"{vol_abbr} {found_pg}"
        ws.cell(row=row_idx, column=4).value = new_vol
        c = ws.cell(row=row_idx, column=col_k)
        c.value = f"AUTO: p.{ref_page} → p.{found_pg}  Δ=+{delta}  [{sec}]"
        c.font = Font(color="C00000", bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = thin
        vc = ws.cell(row=row_idx, column=8); vc.value = "FIXED"; vc.fill = FIX_FILL
        fixed += 1

    wb.save(str(EXCEL_OUT))
    con.close()

    print(f"  Correcciones aplicadas: {fixed}")
    print(f"  Archivo: {EXCEL_OUT}")

if __name__ == "__main__":
    main()
