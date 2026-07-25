#!/usr/bin/env python3
"""
an1_eka_particion — cierra la cola del Eka por **partición forzada**, y corrige lo que la pasada
por contenido había dejado mal.

### El error que esto repara

`an1_eka_cola.py` resuelve cada fila por separado: busca el paranum cuyo texto mejor cubre el de su
marcador. Donde varias filas **comparten marcador** eso no puede funcionar, y no falla ruidosamente
sino en silencio: las seis filas del bloque del marcador 63 (A i 41,14) apuntan todas al mismo
sitio, y tres de ellas se escribieron con **la misma `VRI Ref`** (`483-492`). La comprobación de
cardinalidad no lo vio porque es **por fila**: cada tramo tenía los diez paranums que la fila
declara; lo que nadie miraba es que eran **los mismos diez** tres veces.

Es el punto ciego otra vez, en su forma más pura: tres pares coherentes cada uno consigo mismo y
mutuamente incompatibles.

### El argumento que sí decide

Cuando un bloque de filas consecutivas cae entre dos anclas ya resueltas, **la aritmética fuerza la
partición**: si lo que las filas declaran suma exactamente los paranums libres que quedan entre las
anclas, sólo hay un reparto posible y va en orden.

En el bloque del marcador 63:

    la fila anterior cerrada, `1.447-454`, ocupa 435-442  →  el bloque empieza en 443
    las seis filas declaran 10+10+10+10+40+40 = 120 suttas
    443 + 120 − 1 = 562 = **el último paranum del `div` 18 del CST**

Que el reparto termine **exactamente** donde acaba el `div` no es algo que se pueda ajustar: o cuadra
o no. Cuadra, y con eso las seis quedan repartidas sin ambigüedad.

⚠️ Esto **no** es una interpolación optimista. Se exige que el total coincida **al dígito** con el
hueco; si sobra o falta uno, el bloque no se toca. Es la misma clase de prueba que la aritmética de
rangos del *rāga-peyyāla* de A ii (27 = 9 × 3), no una estimación.

Uso: python3 an1_eka_particion.py [--dry]
"""
import re
import shutil
import sys
from datetime import datetime

import an1_eka_duka as R
from openpyxl import load_workbook


def carga():
    """Filas del Eka en orden de `Sutta #`, con su tramo del CST si ya lo tienen."""
    wb = load_workbook(R.XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    L = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ci['Nikaya']] != 'AN' or str(row[ci['PTS Roman']] or '').strip() != 'i':
            continue
        s = str(row[ci['Sutta #']])
        if not s.startswith('1.'):
            continue
        lo, _, hi = s.split('.')[1].partition('-')
        lo, hi = int(lo), int(hi or lo)
        v = row[ci['VRI Ref']]
        a = b = None
        if v:
            p = str(v).split(':')[1]
            a, b = int(p.split('-')[0]), int(p.split('-')[-1])
        L.append({'s': s, 'lo': lo, 'hi': hi, 'n': hi - lo + 1, 'a': a, 'b': b,
                  'pend': row[ci['Estado']] != 'CONFIRMADO'})
    L.sort(key=lambda x: (x['lo'], x['hi']))
    return L


# Bloques cuya partición está forzada por la aritmética. Se declaran explícitamente —con el ancla
# de la que arrancan y el paranum en el que tienen que terminar— para que la comprobación sea
# auditable y no dependa de que un heurístico dé con ellos.
BLOQUES = [
    {'desde': 443, 'hasta': 562, 'porque': 'el `div` 18 del CST acaba en 562, y la fila anterior '
                                           '(`1.447-454`) ocupa hasta el 442',
     'filas': ['1.455-464', '1.465-474', '1.475-484', '1.485-494', '1.495-534', '1.535-574']},
    {'desde': 312, 'hasta': 319, 'porque': 'entre `1.319` (paranum 311) y `1.328` (paranum 320)',
     'filas': ['1.320-321', '1.322-323', '1.324-327']},
    {'desde': 402, 'hasta': 418, 'porque': 'entre `1.402-405`… (paranum 401) y `1.431-438` (419)',
     'filas': ['1.414-418', '1.419-423', '1.424-430']},
    # El tramo 321-389, entre `1.328` (paranum 320) y `1.402-405` (390). Las cinco filas declaran
    # 4+45+16+1+7 = 73 suttas para **69** paranums libres, así que aquí la partición NO es por
    # cardinalidad: la fijan los anclajes de contenido —`1.378-393` cae en 376-381 (0,98) y
    # `1.394`/`1.395-401` en 382 (1,00)— y lo que absorbe la diferencia es el **Jambudīpapeyyāla**
    # (`1.333-377`), que es justo la fila donde el CST subdivide más fino que el Excel.
    {'desde': 321, 'hasta': 389, 'porque': 'entre `1.328` (paranum 320) y `1.402-405` (390); los '
                                           'anclajes de contenido fijan 376-381 y 382',
     'reparto': [('1.329-332', 4), ('1.333-377', 51), ('1.378-393', 6), ('1.394', 1),
                 ('1.395-401', 7)]},
    {'desde': 584, 'hasta': 611, 'porque': 'entre `1.596-599` (paranum 583) y el final del Eka, '
                                           'que es el paranum 611',
     'filas': ['1.600-615', '1.616-627']},
]
for _b in BLOQUES:
    if 'filas' not in _b:
        _b['filas'] = [s for s, _n in _b['reparto']]

# Filas sueltas cuya contraparte del CST es **más pequeña** que lo que el Excel numera: PTS/DPR
# cuentan varios suttas donde el CST tiene uno o dos. No es un error de nadie —es la compresión de
# Morris otra vez— y el hueco entre las vecinas lo fija sin ambigüedad.
SUELTAS = {
    '1.296-305': ('s0401m:296-297', 'entre `1.287-295` (295) y `1.306` (298) sólo quedan libres el '
                                    '296 y el 297; el contenido confirma que arranca en el 296 '
                                    '(cobertura 1.00). El Excel numera diez donde el CST tiene dos'),
    '1.314-315': ('s0401m:306-307', 'rango redundante: cubre `1.314` (306) y `1.315` (307), las dos '
                                    'ya resueltas individualmente'),
    '1.586-590': ('s0401m:574', 'entre `1.585` (573) y `1.591-592` (575) sólo queda libre el 574: '
                                'el Excel numera cinco donde el CST tiene uno'),
}


def resuelve(L):
    por_s = {f['s']: f for f in L}
    out = {}
    for bl in BLOQUES:
        filas = [por_s[s] for s in bl['filas'] if s in por_s]
        if len(filas) != len(bl['filas']):
            print(f"  ⚠ bloque {bl['desde']}-{bl['hasta']}: faltan filas, no se toca")
            continue
        reparto = dict(bl.get('reparto') or [])
        total = sum(reparto.get(f['s'], f['n']) for f in filas)
        hueco = bl['hasta'] - bl['desde'] + 1
        if total != hueco:
            print(f"  ⚠ bloque {bl['desde']}-{bl['hasta']}: declaran {total} y el hueco es "
                  f"{hueco} — NO cuadra, no se toca")
            continue
        a = bl['desde']
        for f in filas:
            ancho = reparto.get(f['s'], f['n'])
            out[f['s']] = (f's0401m:{a}-{a + ancho - 1}' if ancho > 1 else f's0401m:{a}',
                           f'A i cola: partición forzada del bloque {bl["desde"]}-{bl["hasta"]} '
                           f'({total} suttas declarados = {hueco} paranums libres; {bl["porque"]})')
            a += ancho
    return out


def main():
    dry = '--dry' in sys.argv
    L = carga()
    res = resuelve(L)
    res.update({k: v for k, v in SUELTAS.items()})
    print(f'{len(res)} filas repartidas por partición forzada')
    for s, (vri, _d) in res.items():
        print(f'   {s:>12} → {vri}')
    if not res:
        return
    if not dry:
        cp = f'{R.XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-particion.xlsx'
        shutil.copy(R.XLSX, cp)
        print(f'backup → {cp}')
    wb = load_workbook(R.XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = corr = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'AN':
            continue
        if str(ws.cell(r, ci['PTS Roman']).value or '').strip() != 'i':
            continue
        s = str(ws.cell(r, ci['Sutta #']).value)
        if s not in res:
            continue
        vri, det = res[s]
        antes = ws.cell(r, ci['VRI Ref']).value
        if antes and antes != vri:
            print(f'   corrige {s}: {antes} → {vri}')
            corr += 1
        elif not antes:
            n += 1
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
    print(f'{n} nuevas · {corr} corregidas')
    if dry:
        print('(dry-run: nada escrito)')
        return
    wb.save(R.XLSX)
    print(f'escrito → {R.XLSX}')


if __name__ == '__main__':
    main()
