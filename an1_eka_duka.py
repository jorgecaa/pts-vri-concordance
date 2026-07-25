#!/usr/bin/env python3
"""
an1_eka_duka — resolvedor de la **mitad limpia de A i**: el Duka entero y los 12 primeros vaggas
del Eka.

A i llevaba todo el proyecto aplazado con el argumento de que «el CST no da nombre por sutta». Es
verdad —el `s0401m.mul.xml` sólo titula vaggas— pero **no era el obstáculo**: aquí no hace falta
ningún nombre. Hacen falta dos coordenadas, y las dos existen.

### La coordenada de PTS: `(vagga, sutta)`, que ya está en el Excel

El Eka y el Duka usan el régimen `vagga_romano`: numeral romano centrado para el vagga, arábigo a
sangría 5 para el sutta, **reiniciando dentro de cada vagga**. Y la columna `DPR Ref` de A i trae
justo ese par —`AN 1.1.1`, `AN 2.14.1-12`— con tres niveles `nipāta.vagga.sutta`.

La regla que las une es simplemente **cómo se cita el impreso**: la fila `(v, n)` cae en **el
último marcador de `v` con número ≤ n**. No es una aproximación: en el Eka tardío PTS numera sus
suttas en ese sistema y sólo escribe el número donde arranca cada tirada elidida (el vagga 20 va
`1, 2, 10, 14, 18, 22, 32, 39, 47, 55, 63, 192`), así que «el último ≤ n» es exactamente el sutta
impreso que contiene a `n`. Resuelve **159/159** filas del Eka y **110/111** del Duka.

### La coordenada del CST: la posición dentro del `div`

En AN el `<div>` del CST **es** el vagga, y dentro de él los paranums van en orden. Donde las dos
ediciones cuentan lo mismo —los 12 primeros vaggas del Eka (`10,10,10,10,10,10,10,10,17,42,10,20`
en ambas) y 13 de los 15 del Duka— el sutta `n` del vagga `v` es el **n-ésimo paranum** de su
`div`, y eso da la `VRI Ref` sin conjeturar nada.

Donde no cuentan lo mismo (Duka 2 y 3: PTS 10, CST 11) hace falta decidir *dónde* está la fusión, y
eso ya no lo dice la estructura: lo dice el contenido. Esas filas salen marcadas y no se cierran a
ciegas.

### Las tres comprobaciones que se exigen

1. **`(v, n)` → marcador impreso** por la regla de arriba.
2. **`(v, n)` → paranum del CST** por posición en el `div`, sólo si los recuentos coinciden.
3. **cobertura de contenido** entre el texto del marcador y el del paranum, que es **independiente**
   de las dos anteriores: ninguna usa el texto.

Más la página del Excel (±1) como cuarta señal.

Uso: python3 an1_eka_duka.py [--vol eka|duka] [-v]
"""
import math
import re
import sqlite3
import sys
from collections import Counter, defaultdict

import an_peyyala as ap
from align_rows import assign as ar_assign
from an1_markers import collect_an1
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XLSX = 'PTS_Reference_Complete_Canon.xlsx'
BOOK = 17
# (prefijo del `Sutta #`, nipāta, fichero VRI, hasta qué vagga llega la «mitad limpia»)
NIPATAS = {'eka': ('1', 'EKA', 's0401m', 12), 'duka': ('2', 'DUKA', 's0402m1', 15)}
# Umbral con separación MEDIDA, no elegido a ojo: sobre las 197 filas de la mitad limpia la
# cobertura real va de 0,64 a 1,00 (mediana 0,97-0,99), y un control con pares de vaggas distintos
# da media 0,26 y máximo 0,48. El hueco entre 0,48 y 0,64 es donde cae el umbral.
COV_MIN = 0.55


def marcadores(conn, nipata):
    """`{vagga: [(nº, página, línea, texto)]}` en orden, con el texto hasta el marcador siguiente."""
    pgs = {r['page_no']: (r['unitext'] or '').split('\n')
           for r in conn.execute('SELECT page_no,unitext FROM pages WHERE edition="mula" '
                                 'AND book_no=?', (BOOK,))}
    ms = [m for m in collect_an1(pgs) if m['nipata'] == nipata and m['num'] and m.get('vagga')]
    ms.sort(key=lambda m: (m['page'], m['line']))
    plano = [(pg, ln, t) for pg in sorted(pgs) for ln, t in enumerate(pgs[pg], 1)]
    idx = {(pg, ln): i for i, (pg, ln, _t) in enumerate(plano)}
    out = defaultdict(list)
    for k, m in enumerate(ms):
        i0 = idx[(m['page'], m['line'])]
        i1 = idx[(ms[k + 1]['page'], ms[k + 1]['line'])] if k + 1 < len(ms) else len(plano)
        out[m['vagga']].append((m['num'], m['page'], m['line'],
                                ap.fold(' '.join(t for _p, _l, t in plano[i0:i1]))))
    for v in out:
        out[v].sort()
    return out


def divisiones(stem):
    """`{orden del div: [(paranum, texto plegado)]}` del fichero VRI."""
    out, divs = defaultdict(list), []
    for u in ap.cst_unidades(stem):
        if not divs or divs[-1] != u['div']:
            divs.append(u['div'])
        out[len(divs)].append((u['pn'], ap.fold(u['texto'])))
    return out, divs


def filas(pref):
    """Filas de A i del nipāta, con su `(vagga, sutta)` sacado del `DPR Ref` de tres niveles."""
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ci['Nikaya']] != 'AN' or str(row[ci['PTS Roman']] or '').strip() != 'i':
            continue
        num = str(row[ci['Sutta #']])
        if not num.startswith(pref + '.'):
            continue
        m = re.match(r'AN\s*\d+\.(\d+)\.(\d+)(?:-(\d+))?', str(row[ci['DPR Ref']] or ''))
        out.append({'num': num, 'name': str(row[ci['Sutta Name']] or ''),
                    'page': row[ci['PTS Page']], 'estado': row[ci['Estado']],
                    'v': int(m.group(1)) if m else None,
                    'n': int(m.group(2)) if m else None,
                    'n_hi': int(m.group(3)) if (m and m.group(3)) else
                            (int(m.group(2)) if m else None)})
    return out


def _gramas(t, n=4):
    """N-gramas de caracteres sobre el texto sin espacios."""
    x = t.replace(' ', '')
    return {x[i:i + n] for i in range(len(x) - n + 1)}


def cobertura(a, b, idf=None):
    """Cobertura por **n-gramas de caracteres**, no por palabras.

    En pali la composición es libre en la ortografía: PTS escribe `kodha vinayo ca upanaha vinayo`
    y el CST `kodhavinayo ca upanahavinayo`. Es **el mismo texto** y el solapamiento de conjuntos
    de palabras da 0,22 — suficiente para tirar una fila correcta. Sobre 4-gramas de caracteres, el
    compuesto partido y el compuesto junto comparten casi todo, y la medida deja de castigar una
    convención tipográfica.

    Asimétrica hacia el lado corto: una edición elide lo que la otra desarrolla (aquí el CST añade
    `etadaggaṃ bhikkhave imesaṃ dvinnaṃ…` donde PTS pone `pe`), y penalizar por eso castigaría
    justo a los pares correctos.
    """
    A, B = _gramas(a), _gramas(b)
    if not A or not B:
        return 0.0
    return len(A & B) / max(1, min(len(A), len(B)))


def alinea_por_contenido(ms, us, idf):
    """Marcadores ↔ paranums de un vagga cuyos recuentos NO coinciden, por contenido.

    Sólo se usa donde la estructura ya no decide. Es alineamiento **monótono**: las dos ediciones
    imprimen el vagga en el mismo orden, y lo que cambia es dónde una funde lo que la otra numera.
    En el vagga 2 del Duka sale limpio —el marcador 10 se lleva el paranum 21 y el **20 se queda
    sin marcador**: PTS imprime bajo un solo número el par `saddhammassa sammosāya` /
    `saddhammassa ṭhitiyā` que el CST numera aparte—.
    """
    M = [[cobertura(m[3], u[1], idf) for u in us] for m in ms]
    idx = ar_assign(len(ms), len(us), lambda i, j: max(M[i][j], 0.001), lambda j: 1,
                    skip_penalty=0.2)
    return {ms[i][0]: (us[j][0], M[i][j]) for i, j in enumerate(idx) if j is not None}


def main():
    vols = [sys.argv[sys.argv.index('--vol') + 1]] if '--vol' in sys.argv else ['eka', 'duka']
    verbose = '-v' in sys.argv
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    print('=' * 100)
    print('A i — mitad limpia: (vagga, sutta) del DPR contra el marcador impreso y el paranum del CST')
    print('=' * 100)
    total = Counter()
    for vol in vols:
        pref, nipata, stem, v_max = NIPATAS[vol]
        mk = marcadores(conn, nipata)
        cst, divs = divisiones(stem)
        df = Counter()
        for v in cst:
            for _pn, t in cst[v]:
                df.update(set(t.split()))
        N = max(1, sum(len(x) for x in cst.values()))
        idf = {w: math.log(N / c) for w, c in df.items()}

        print(f'\n── {nipata}: {len(mk)} vaggas impresos · {len(divs)} divs del CST · '
              f'mitad limpia hasta el vagga {v_max}')
        print(f"{'fila':>12} {'v.n':>9} {'marcador':>9} {'pág':>5} {'paranum':>8} {'cov':>5}  señal")
        porcont = {}
        for f in filas(pref):
            if f['v'] is None:
                total['sin DPR de 3 niveles'] += 1
                continue
            if f['v'] > v_max:
                total['cola (fase C)'] += 1
                continue
            ms = mk.get(f['v'], [])
            cand = [x for x in ms if x[0] <= f['n']]
            if not cand:
                total['sin marcador'] += 1
                print(f"  {f['num']:>12} {f['v']}.{f['n']:<7} {'—':>9}  sin marcador en el vagga")
                continue
            m = cand[-1]
            pag_ok = isinstance(f['page'], int) and abs(m[1] - f['page']) <= 1
            uds = cst.get(f['v'], [])
            cuadra = len(uds) == (ms[-1][0] if ms else 0)
            pn, cov = None, 0.0
            if cuadra and f['n'] <= len(uds):
                pn, txt = uds[f['n'] - 1]
                cov = cobertura(m[3], txt, idf)
            elif uds:
                # los recuentos no cuadran: decide el contenido, no la posición
                if f['v'] not in porcont:
                    porcont[f['v']] = alinea_por_contenido(ms, uds, idf)
                if m[0] in porcont[f['v']]:
                    pn, cov = porcont[f['v']][m[0]]
            señal = []
            if not cuadra:
                señal.append(f'por contenido (PTS {ms[-1][0]} · CST {len(uds)})')
            if not pag_ok:
                señal.append(f'página {m[1]} vs Excel {f["page"]}')
            if pn is not None and cov < COV_MIN:
                señal.append(f'cobertura baja')
            clave = ('cerrable' if (pn is not None and pag_ok and cov >= COV_MIN)
                     else 'a revisar')
            total[f'{nipata}: {clave}'] += 1
            if verbose or clave == 'a revisar':
                print(f"  {f['num']:>12} {f['v']}.{f['n']:<7} {m[0]:>9} {m[1]:>5} "
                      f"{str(pn):>8} {cov:>5.2f}  {' · '.join(señal)}")
    print('\n' + ' · '.join(f'{k}: {v}' for k, v in sorted(total.items())))


if __name__ == '__main__':
    main()
