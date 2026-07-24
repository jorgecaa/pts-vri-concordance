#!/usr/bin/env python3
"""
Restore AN entries from the original blog data (PTS_Reference_Full_Canon_CORRECTED.xlsx)
into PTS_Reference_Complete_Canon.xlsx, then add line numbers via page-level matching.
"""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict, Counter

DB = 'src/data/tipitaka.sqlite'
XL_CORRECTED = 'PTS_Reference_Full_Canon_CORRECTED.xlsx'
XL_COMPLETE = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

AN_BOOKS = {'i': 17, 'ii': 18, 'iii': 19, 'iv': 20, 'v': 21}

# ── Load original blog data for AN ──
wb_orig = load_workbook(XL_CORRECTED)
ws_orig = wb_orig['PTS Reference']
orig_headers = {ws_orig.cell(row=1, column=c).value: c for c in range(1, ws_orig.max_column + 1)}

orig_an = {}
for r in range(2, ws_orig.max_row + 1):
    if ws_orig.cell(row=r, column=orig_headers['Nikaya']).value != 'AN':
        continue
    snum = str(ws_orig.cell(row=r, column=orig_headers['Sutta #']).value or '')
    orig_an[snum] = {
        'name': str(ws_orig.cell(row=r, column=orig_headers['Sutta Name']).value or ''),
        'vol': str(ws_orig.cell(row=r, column=orig_headers['PTS Vol']).value or ''),
        'roman': str(ws_orig.cell(row=r, column=orig_headers['PTS Roman']).value or '').strip(),
        'page': ws_orig.cell(row=r, column=orig_headers['PTS Page']).value,
        'ref': str(ws_orig.cell(row=r, column=orig_headers['PTS Full Ref']).value or ''),
        'type': str(ws_orig.cell(row=r, column=orig_headers['Type']).value or ''),
        'raw': str(ws_orig.cell(row=r, column=orig_headers['Raw Sutta ID']).value or ''),
    }

print(f'Original AN entries: {len(orig_an)}')

# ── Restore into Complete Excel ──
wb = load_workbook(XL_COMPLETE)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(row=1, column=c).value
    if h:
        cols[h] = c

restored = 0
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'AN':
        continue
    snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    
    if snum in orig_an:
        o = orig_an[snum]
        ws.cell(row=ri, column=cols['PTS Vol']).value = o['vol']
        ws.cell(row=ri, column=cols['PTS Roman']).value = o['roman']
        ws.cell(row=ri, column=cols['PTS Page']).value = o['page']
        ws.cell(row=ri, column=cols['PTS Ref']).value = o['ref']
        restored += 1

print(f'Restored: {restored} AN entries to original blog pages')

# ── Now add line numbers via page-level sequential matching ──
print('\nAdding line numbers...')

def find_markers_on_page(book_no, page_no):
    """Find sutta start markers on page: 'N. text' pattern."""
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                (book_no, page_no))
    r = cur.fetchone()
    if not r or not r['unitext']:
        return []
    
    lines = r['unitext'].split('\n')
    markers = []
    
    for i, line in enumerate(lines):
        s = line.strip()
        # Vagga heading
        if re.match(r'^[IVX]+\.?\s*$', s) and len(s) <= 8:
            markers.append((i + 1, 'vagga', s))
        # Sutta start
        elif re.match(r'^\d+\.\s+\S', s):
            num = int(re.match(r'^(\d+)\.', s).group(1))
            if num <= 1000:
                markers.append((i + 1, 'sutta', num))
    
    return markers


# Group entries by (book_no, page)
page_entries = defaultdict(list)
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'AN':
        continue
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower()
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    book_no = AN_BOOKS.get(roman)
    if not book_no or not page:
        continue
    
    page_entries[(book_no, page)].append({
        'ri': ri, 'roman': roman, 'page': page
    })

results = Counter()
fixed = 0

for (book_no, page_no), entries in sorted(page_entries.items()):
    markers = find_markers_on_page(book_no, page_no)
    sutta_markers = [(l, c) for l, t, c in markers if t == 'sutta']
    vagga_markers = [(l, c) for l, t, c in markers if t == 'vagga']
    
    for ei, entry in enumerate(entries):
        line_num = None
        method = 'not_found'
        
        # Strategy 1: First entry after vagga heading
        if ei == 0 and vagga_markers and sutta_markers:
            vagga_line = vagga_markers[-1][0]
            for sl, sn in sutta_markers:
                if sl > vagga_line:
                    line_num = sl
                    method = 'vagga_next'
                    break
            if not line_num and sutta_markers:
                line_num = sutta_markers[0][0]
                method = 'vagga_next'
        
        # Strategy 2: Sequential match
        if not line_num:
            if ei < len(sutta_markers):
                line_num = sutta_markers[ei][0]
                method = 'sutta_marker'
        
        # Strategy 3: Find first content line
        if not line_num:
            cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                       (book_no, page_no))
            r = cur.fetchone()
            if r and r['unitext']:
                lines = r['unitext'].split('\n')
                for i, line in enumerate(lines):
                    s = line.strip()
                    if s and len(s) > 5 and not re.match(r'^[A-ZĀĪŪṄÑṬḌṆḶ\s\-\.\,\;\:]+$', s):
                        line_num = i + 1
                        method = 'page_top'
                        break
            if not line_num:
                line_num = 1
                method = 'page_top'
        
        results[method] += 1
        
        if line_num:
            old_ref = str(ws.cell(row=entry['ri'], column=cols['PTS Ref']).value or '')
            new_ref = 'A %s %d' % (entry['roman'], entry['page'])
            if line_num > 1:
                new_ref += ',%d' % line_num
            
            if new_ref != old_ref:
                ws.cell(row=entry['ri'], column=cols['PTS Ref']).value = new_ref
                fixed += 1

wb.save(XL_COMPLETE)

print(f'Results: {dict(results)}')
print(f'Fixed refs: {fixed}')
print(f'Saved: {XL_COMPLETE}')
conn.close()
