#!/usr/bin/env python3
"""Aligner DEFINITIVO SN V: Excel(DPR) → massive.tsv → cst_paranum → XML VRI → CST.

El Excel usa notación DPR. massive.tsv da, por código DPR, el `cst_paranum`. El XML
VRI (`romn/s0305m.mul.xml`, la fuente que massive referencia) da el texto CST exacto
por paranum (100% verificado). Así la alineación PTS↔CST es EXACTA, sin adivinar por
contenido ni reconciliar agrupaciones. Luego se valida con el validador (Modelo B).

Uso: python3 validador_sn5_vri.py [--dry] [--all] [--n N] [--only 45.112,55.40]
     --only  revalida SOLO esos `Sutta #` (los demás conservan su veredicto en el JSON)
"""
import csv, os, re, sqlite3, sys, json
import xml.etree.ElementTree as ET
from collections import Counter
import sutta_hash as sh
from openpyxl import load_workbook
from validador import validate_pair
from align_rows import assign as ar_assign
from massive_reader import build_massive as _mr_build
from validador_sn5 import (build_markers, pts_for, pts_by_name, pts_by_content, pts_page,
                           ordinal_of, _sutta_text as v5_text)

VRI = '/tmp/tipitaka-xml/romn/s0305m.mul.xml'
OUT = 'validador_sn5_vri.json'
DB = 'src/data/tipitaka.sqlite'
SN5_BOOK = 16


# ── XML VRI → paranum → texto CST del sutta ─────────────────────────────────
def build_vri_index():
    root = ET.parse(VRI).getroot()
    body = root.find('.//body')
    suttas, cur = [], None
    for el in body.iter():
        rend = el.get('rend')
        if el.tag == 'p' and rend == 'subhead':
            cur = {'title': ''.join(el.itertext()).strip(), 'pn': [], 'text': ''}
            suttas.append(cur)
        elif el.tag == 'p' and rend == 'bodytext' and cur is not None:
            n = el.get('n')
            if n:
                cur['pn'].append(n)
            cur['text'] += ' ' + ''.join(el.itertext())
    idx = {}
    for s in suttas:
        for n in s['pn']:
            m = re.match(r'(\d+)(?:-(\d+))?$', n)
            if m:
                a, b = int(m.group(1)), int(m.group(2) or m.group(1))
                for p in range(a, b + 1):
                    idx.setdefault(p, s)
    return idx


# ── massive.tsv → cada nº canónico DPR → cst_paranum ────────────────────────
def build_massive():
    """Cada nº canónico DPR → `(cst_paranum, título CST)`, con **expansión de rangos**.

    `massive.tsv` (que NO se modifica) da un único `cst_paranum` por grupo, el del primer miembro.
    Asignárselo a todos hace que los miembros se cotejen contra el primer sutta del grupo en vez de
    contra su propio párrafo. La expansión va con compuerta (`massive_reader`): solo si TODOS los
    paranum resultantes existen en el XML — normalmente porque caen en el **bloque elidido** del CST,
    que es el homólogo exacto del rango de PTS.
    """
    m = _mr_build('sn5.', build_vri_index())
    return {f'{k[0]}.{k[1]}': (v[0], v[1]) for k, v in m.items()}


def excel_entries():
    wb = load_workbook('PTS_Reference_Complete_Canon.xlsx', read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['Nikaya']] or '') != 'SN' or str(row[ci['PTS Roman']] or '').strip().lower() != 'v':
            continue
        num = str(row[ci['Sutta #']]); name = str(row[ci['Sutta Name']] or '')
        m = re.search(r'\(SN\s+(\d+)\.(\d+)', name)
        canon = f'{m.group(1)}.{m.group(2)}' if m else num
        # el nº corrido del marcador PTS = el canónico (no el Sutta# comprimido)
        inner = int(canon.split('.')[1]) if '.' in canon and canon.split('.')[1].isdigit() else None
        out.append({'num': num, 'canon': canon, 'inner': inner,
                    'name': re.sub(r'\(SN[^)]*\)', '', name).strip(),
                    'page': row[ci['PTS Page']], 'legacy': str(row[ci['Validation']] or '')})
    return out


def assign_volume(entries, cur, mm, canon2para, vri):
    """Asignación **inyectiva y monótona** de las filas de S v a sus marcadores, por saṃyutta.

    Igual que en S iii: resolver fila a fila permitía que dos filas cayeran en el mismo marcador
    (16 marcadores compartidos por 36 filas), con lo que una quedaba validada contra el locus de su
    vecino. Aquí las filas van en orden canónico y los marcadores en orden de lectura, y la
    asignación se resuelve globalmente con `align_rows.assign`.

    Devuelve `{Sutta # → (página, línea)}`.
    """
    marks, caps = [], Counter()
    for pg in sorted(mm):
        seen = set()
        for line, gid, _vpos, name in sorted(mm[pg]):
            caps[(pg, line)] += 1
            if line in seen:
                continue
            seen.add(line)
            marks.append({'page': pg, 'line': line, 'gid': gid, 'name': name})
    # La asignación va sobre TODO el volumen, no por saṃyutta: si se hace por partes, dos filas de
    # saṃyuttas contiguos pueden reclamar el mismo marcador en la frontera (S v 305: 52.21 y 53.1).
    rows = []
    for e in entries:
        m = re.match(r'(\d+)\.(\d+)', e['canon'])
        if m:
            rows.append((int(m.group(1)), int(m.group(2)), e))
    rows.sort(key=lambda t: (t[0], t[1]))
    rows = [(inner, e) for _sam, inner, e in rows]
    sub = marks

    if True:
        def score(i, j, rows=rows, sub=sub):
            inner, e = rows[i]
            q = sub[j]
            if not isinstance(e['page'], int):
                return None
            d = abs(q['page'] - e['page'])
            if d > 2:
                return None
            hit = canon2para.get(e['canon'])
            s = (30 if d == 0 else 15 if d == 1 else 0)
            if hit:
                t = vri.get(hit[0])
                if t:
                    s += _stem_score(t['title'], q['name'])
            s += _stem_score(e['name'], q['name'])
            if q['gid'] == inner:
                s += 25                              # el nº corrido que declara el canónico
            return s

        idx = ar_assign(len(rows), len(sub), score,
                        lambda j, sub=sub: caps[(sub[j]['page'], sub[j]['line'])],
                        skip_penalty=1000.0)
        out = {}
        for (_inner, e), j in zip(rows, idx):
            out[e['num']] = (sub[j]['page'], sub[j]['line']) if j is not None else None
    return out


_FOLD5 = str.maketrans({'ā': 'a', 'ī': 'i', 'ū': 'u', 'ṅ': 'n', 'ñ': 'n', 'ṇ': 'n',
                        'ṭ': 't', 'ḍ': 'd', 'ḷ': 'l', 'ṃ': 'm', 'ṁ': 'm'})


def _stem_score(a, b):
    """Afinidad de nombres, con la misma normalización que el resto del pipeline de S v."""
    def st(t):
        t = re.sub(r'[^a-z]', '', re.sub(r'^[\d\s.,()-]*', '', (t or '').lower()).translate(_FOLD5))
        t = re.sub(r'(suttantam|suttam|suttani|vaggo).*$', '', t)
        return re.sub(r'[aiueom]+$', '', re.sub(r'(.)\1+', r'\1', t))
    x, y = st(a), st(b)
    if not x or not y or min(len(x), len(y)) < 3:
        return 0
    if x == y:
        return 100
    if x.startswith(y) or y.startswith(x):
        return 60
    if min(len(x), len(y)) >= 4 and (x in y or y in x):
        return 40
    return 0


def main():
    dry = '--dry' in sys.argv
    do_all = '--all' in sys.argv
    n = int(sys.argv[sys.argv.index('--n') + 1]) if '--n' in sys.argv else 30
    only = set(sys.argv[sys.argv.index('--only') + 1].split(',')) if '--only' in sys.argv else None
    vri = build_vri_index()
    canon2para = build_massive()
    entries = excel_entries()
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    mm = build_markers(cur)
    assigned = assign_volume(entries, cur, mm, canon2para, vri)

    tasks, no_dpr, no_para, no_pts = [], [], [], []
    for e in entries:
        if only is not None and e['num'] not in only:
            continue
        hit = canon2para.get(e['canon'])
        if not hit:
            no_dpr.append(e); continue
        para, ctitle = hit
        s = vri.get(para)
        if not s:
            no_para.append(e); continue
        cst = ' '.join(sh.tokens(s['text'])[:350])
        # La localización del marcador la da ahora la ASIGNACIÓN GLOBAL (inyectiva y monótona):
        # la cadena anterior de heurísticas por fila permitía que dos filas cayeran en el mismo
        # marcador (16 marcadores compartidos por 36 filas).
        pts, src = None, None
        loc = assigned.get(e['num'])
        if loc:
            pts = v5_text(cur, loc[0], loc[1])
            src = 'dp'
        if not pts and isinstance(e['page'], int):
            # el marcador no aísla texto cotejable (grupos peyyāla): página completa, como antes
            pts = pts_page(cur, e['page'])
            if pts:
                src = 'page'
        if not pts:
            no_pts.append(e); continue
        tasks.append((e, pts, cst, s['title'], src))

    print('=' * 92)
    print('ALINEADOR VRI (Excel→DPR→massive→cst_paranum→XML VRI) — SN V')
    print('=' * 92)
    print(f'Excel SN V: {len(entries)}')
    print(f'  alineadas con texto CST+PTS: {len(tasks)} ({100*len(tasks)/len(entries):.0f}%)')
    print(f'  sin DPR en massive: {len(no_dpr)} | sin paranum en VRI: {len(no_para)} | sin texto PTS: {len(no_pts)}')
    if no_dpr:
        print('  muestra sin DPR:', [e['canon'] for e in no_dpr[:12]])
    if dry:
        return

    run = tasks if (do_all or only) else tasks[:n]
    print(f'\nValidando {len(run)} (Gemini en vivo)...', flush=True)
    rows = []
    for k, (e, pts, cst, ctitle, src) in enumerate(run, 1):
        res = validate_pair(pts, cst, e['name'], ctitle, 'SN', concordant=True)
        rows.append({'num': e['num'], 'canon': e['canon'], 'name': e['name'],
                     'cst_title': ctitle, 'legacy': e['legacy'], 'src': src, **res})
        if k % 25 == 0:
            print(f'  ...{k}/{len(run)}', flush=True)
    # fusión por `num`: una corrida parcial (--only/--n) no debe borrar lo ya validado
    prev = []
    if os.path.exists(OUT):
        prev = json.load(open(OUT))
    merged = {r['num']: r for r in prev}
    merged.update({r['num']: r for r in rows})
    json.dump(list(merged.values()), open(OUT, 'w'), ensure_ascii=False, indent=1)
    est = Counter(r['estado'] for r in rows); val = Counter(r['validation'] for r in rows)
    print(f'\nEstado: {dict(est)} | Validation: {dict(val)}')
    print(f'Resultados → {OUT} ({len(merged)} en total; nada escrito al Excel).')


if __name__ == '__main__':
    main()
