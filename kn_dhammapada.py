#!/usr/bin/env python3
"""
kn_dhammapada — el Dhammapada: 26 filas, **una por vagga**, referenciadas por ESTROFA.

### El cambio de notación, que es lo primero

En DN/MN/SN/AN la referencia es `<volumen> <página>,<línea>` porque lo que se cita es la página del
impreso. **En las obras en verso de KN eso no sirve**: lo que se cita —y lo que las tres ediciones
comparten— es el **número de estrofa**. `Dhp 21-32` identifica el Appamādavagga en cualquier
edición; `Dh 7` sólo dice en qué página de PTS empieza, y esa página es de PTS y de nadie más.

Así que aquí la `PTS Ref` pasa a ser **`Dhp <primera>-<última>`**. La página se conserva en su
columna, que sigue siendo cierta y útil, pero deja de ser la referencia.

### Las tres fuentes coinciden estrofa por estrofa

| | qué da |
|---|---|
| **PTS** (BD, libro 23) | numera **cada estrofa** en el margen (`1.`, `21.`, `383.`) y titula cada vagga (`2. Appamādavagga`) |
| **CST** (`s0502m`) | 26 `<div rend="chapter">`, y su `n` **es el número de estrofa**, corrido de 1 a 423 |
| **Excel** | 26 filas con el rango ya escrito en `PTS Alt/Verse` (`Dh v1-20`, `Dh v21-32`, …) |

Y coinciden **las 26**: `1-20, 21-32, 33-43, 44-59, … 383-423` en las tres. No hay nada que
alinear — hay que comprobarlo, que es distinto.

⚠️ Aquí el `n` del CST **sí** es corrido por fichero (a diferencia del Khuddakapāṭha, donde
reinicia por capítulo), así que la `VRI Ref` vuelve a la forma normal `s0502m:1-20`. La regla
(5-bis) no se aplica a todas las obras de KN: hay que mirar el fichero antes de elegir la forma.

### Las tres comprobaciones

1. el rango de estrofas del Excel ≡ el rango de `n` del `div` del CST en la misma posición;
2. la estrofa inicial **aparece numerada en el impreso** en la página que declara el Excel (±1);
3. cobertura de contenido entre el texto del vagga en las dos ediciones.

Uso: python3 kn_dhammapada.py [--dry]
"""
import re
import shutil
import sqlite3
import sys
import xml.etree.ElementTree as ET
from datetime import datetime

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
BOOK, STEM = 23, 's0502m'
COV_MIN = 0.55


def vaggas_pts(conn):
    """`{nº de estrofa: (página, línea)}` y el texto de cada vagga del impreso.

    El impreso numera la estrofa al margen (`21. appamādo amatapadaṃ…`), así que el ancla del lado
    PTS es el propio número de verso: no hace falta contar nada.
    """
    pgs = {r['page_no']: (r['unitext'] or '').split('\n')
           for r in conn.execute('SELECT page_no,unitext FROM pages WHERE edition="mula" '
                                 'AND book_no=?', (BOOK,))}
    versos, plano = {}, []
    for pg in sorted(pgs):
        for ln, t in enumerate(pgs[pg], 1):
            plano.append((pg, ln, t))
            m = re.match(r'\s*(\d+)\.\s+\S', t.replace('\r', ''))
            if m:
                versos.setdefault(int(m.group(1)), (pg, ln, len(plano) - 1))
    return versos, plano


def vaggas_cst():
    """`[(nombre, primera, última, texto)]` de los 26 capítulos del CST."""
    root = ET.parse(f'/tmp/tipitaka-xml/romn/{STEM}.mul.xml').getroot()
    out = []
    for d in root.iter('div'):
        h = d.find('head')
        if h is None or h.get('rend') != 'chapter':
            continue
        ns = [int(p.get('n')) for p in d.iter('p') if (p.get('n') or '').isdigit()]
        txt = ' '.join(''.join(p.itertext()) for p in d.iter('p'))
        if ns:
            out.append((' '.join(''.join(h.itertext()).split()), min(ns), max(ns), ap.fold(txt)))
    return out


def filas():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ci['Nikaya']] != 'KN' or str(r[ci['PTS Vol']]) != 'Dh':
            continue
        m = re.search(r'v(\d+)-(\d+)', str(r[ci['PTS Alt/Verse']] or ''))
        out.append({'num': str(r[ci['Sutta #']]), 'name': str(r[ci['Sutta Name']] or ''),
                    'page': r[ci['PTS Page']], 'estado': r[ci['Estado']],
                    'a': int(m.group(1)) if m else None, 'b': int(m.group(2)) if m else None})
    return out


def main():
    dry = '--dry' in sys.argv
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    versos, plano = vaggas_pts(conn)
    cst, fs = vaggas_cst(), filas()
    print(f'PTS: {len(versos)} estrofas numeradas · CST {len(cst)} capítulos · Excel {len(fs)} filas')
    if len(cst) != len(fs):
        print('⚠ los recuentos no coinciden: no se escribe nada')
        return
    res = {}
    print(f"\n{'fila':>6} {'estrofas Excel':>15} {'CST':>12} {'pág':>5} {'impreso':>9} {'cov':>5}  ok")
    for k, f in enumerate(fs):
        nom, a, b, tc = cst[k]
        # texto del vagga en el impreso: de su primera estrofa a la primera de la siguiente
        i0 = versos.get(f['a'], (None, None, None))[2]
        i1 = versos.get(cst[k + 1][1], (None, None, len(plano)))[2] if k + 1 < len(cst) else len(plano)
        tp = ap.fold(' '.join(t for _p, _l, t in plano[i0:i1])) if i0 is not None else ''
        cov = R.cobertura(tp, tc) if tp else 0.0
        pg_imp = versos.get(f['a'], (None,))[0]
        pag = isinstance(f['page'], int) and pg_imp is not None and abs(pg_imp - f['page']) <= 1
        rango_ok = (f['a'], f['b']) == (a, b)
        ok = rango_ok and pag and cov >= COV_MIN
        ex = f"{f['a']}-{f['b']}"
        print(f"  {f['num']:>6} {ex:>15} {f'{a}-{b}':>12} {str(f['page']):>5} "
              f"{str(pg_imp):>9} {cov:>5.2f}  {'✓' if ok else '✗'}")
        if ok:
            res[f['num']] = (f'{STEM}:{a}-{b}', f'Dhp {a}-{b}',
                             f'Dhammapada: el vagga «{nom}» va de la estrofa {a} a la {b} en las '
                             f'dos ediciones; la {a} está numerada en el impreso (Dh p{pg_imp}); '
                             f'cobertura de contenido {cov:.2f}')
    print(f'\n{len(res)}/{len(fs)} firmadas')
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-dhp.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'KN' or str(ws.cell(r, ci['PTS Vol']).value) != 'Dh':
            continue
        num = str(ws.cell(r, ci['Sutta #']).value)
        if num not in res or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, ref, det = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['PTS Ref']).value = ref          # ← por ESTROFA, no por página
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
