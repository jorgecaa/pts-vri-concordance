#!/usr/bin/env python3
"""
kn_buddhavamsa — el Buddhavaṃsa (29 filas) y el Cariyāpiṭaka (34), que el Excel mete bajo **una
sola etiqueta `Bv`** y son dos obras con su libro de la BD y su fichero VRI. Separarlas es el primer
paso, no un adorno: son `Sutta #` `12.x` y `13.x`.

### ⚠️ Lo primero: la paginación de la BD NO es la que cita el Excel

Y esto había que verificarlo antes de emparejar nada (regla del código de libro). Medido:

| | BD | Excel |
|---|--:|--:|
| Buddhavaṃsa | libro 41, **102** pp. | 1-68 |
| Cariyāpiṭaka | libro 42, **37** pp. | 73-101 |

No es un desfase constante que se pueda calibrar: las dos series **arrancan juntas y divergen**. El
Dīpaṅkara abre en la p9 de la BD y la fila dice 17; el Vessantara en la p7 de la BD (+72 = 79) y la
fila dice 78; al llegar al final el desfase es de seis páginas. Son **dos composiciones distintas
del mismo texto** — el Excel pagina el volumen único en que PTS imprimió las dos obras juntas (Bv
1-68, Cp 73-101, con el hueco del front matter entre medias), y la BD guarda otra tirada.

**Consecuencia: la columna `PTS Page` de estas 63 filas no se toca ni se usa como evidencia.** No
se «corrige» una paginación contra la otra: las dos son PTS y ninguna es errata de la otra. La
referencia de estas obras no la necesita — es el capítulo, no la página (regla (5-ter)).

### Lo que sí decide

**Buddhavaṃsa.** El CST tiene **29 capítulos** (`div`), el Excel **29 filas**, mismos nombres y
mismo orden. El impreso los encabeza con numeral romano —`II -- DĪPAṄKARABUDDHAVAṂSO`— y su romano
vale **el índice del CST menos uno**.

⚠️ El motivo es concreto y conviene decirlo bien: **PTS mete la Sumedhapatthanākathā DENTRO del
capítulo del Dīpaṅkara**, no la imprime suelta. El encabezado `II -- DĪPAṄKARABUDDHAVAṂSO` de la p9
abre con `Kappe ca satasahasse ca caturo ca asaṅkhiye | Amaraṃ nāma nagaraṃ…`, que es el arranque de
la **Sumedhapatthanā** del CST (cobertura 0,99); el Dīpaṅkara propiamente dicho no empieza hasta la
p21. El CST parte en dos lo que PTS imprime como un capítulo — y tiene sentido, porque Sumedha vive
en tiempos de Dīpaṅkara.

⚠️ Y de ahí sale la regla que gobierna esta tabla: **la posición de cada capítulo en la BD se
localiza por su INCIPIT, no por su encabezado**. El encabezado dice dónde empieza el capítulo
*impreso*; el incipit dice dónde empieza el texto *de la fila*. Con 29 capítulos el sondeo acierta
entre 0,91 y 1,00 y el segundo candidato nunca pasa de 0,60.

**Cariyāpiṭaka.** El impreso subtitula cada cariyā con su número dentro del vagga
(`5 Soṇapaṇḍitacariyaṃ`) y salen **35** en tres vaggas de **10 + 10 + 15**, exactamente los 35
`subhead` del CST y en el mismo orden.

⚠️ **Pero el Excel sólo tiene 34.** Le falta la última, `15. Mahālomahaṃsacariyā`, que **el impreso
sí trae** (p35 de la BD, cierra con `sabbattha samako homi esā me upekkhāpāramī ti` y el colofón de
la obra). No es una divergencia de ediciones: es una **fila que falta**, y se añade.

⚠️ Y el `n` del CST **reinicia en cada vagga** (regla (5-bis)): leer el fichero por paranum devuelve
las tres series entrelazadas —`Akitti 1`, `Mātuposaka 1`, `Yudhañjaya 1`, `Akitti 2`…— y da 337
«subheads» donde hay 35. Se lee en **orden de documento**, y la `VRI Ref` es `s0512m:c<vagga>.<n>`.

Uso: python3 kn_buddhavamsa.py [--obra bv|cp] [--dry]
"""
import re
import shutil
import sqlite3
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from difflib import SequenceMatcher

import an_peyyala as ap
import an1_eka_duka as R
import pali_norm
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
XML = '/tmp/tipitaka-xml/romn'
COV_MIN = 0.55
ROMANOS = ('I II III IV V VI VII VIII IX X XI XII XIII XIV XV XVI XVII XVIII XIX XX XXI XXII '
           'XXIII XXIV XXV XXVI XXVII XXVIII XXIX').split()
NUM_ROM = {r: k + 1 for k, r in enumerate(ROMANOS)}


def norm(t):
    """Raíz comparable de un nombre de capítulo/cariyā.

    ⚠️ El Excel escribe los nombres del Cariyāpiṭaka **con guiones blandos** (U+00AD):
    `Mahā\u00adsudas\u00adsa\u00adna\u00adcariya`. Invisibles al leerlos e imposibles de casar si
    no se quitan — 21 de las 34 filas fallaban por eso. Y usa `ṁ` donde el CST usa `ṃ`, que es la
    primera regla de `pali_norm`.
    """
    t = pali_norm.normaliza((t or '').replace('\u00ad', ''))
    t = re.sub(r'^\s*\(KN[^)]*\)\s*', '', t)          # `(KN 14.10) Bv 10 Paduma…`
    t = re.sub(r'^\s*(Bv|Cp)\s+\d+\s*', '', t)
    t = ap.fold(re.sub(r'^[\d.\sIVX\-]+', '', t)).strip()
    t = re.sub(r'(budhavams\w*|cariya\w*|kanda\w*|katha)$', '', t).strip()
    # ⚠️ una sola vocal final, no `[aiueom]+`: con el cuantificador, `paduma` se quedaba en `pad`
    # después de perder `-buddhavaṃso`, y tres capítulos no casaban con su propio nombre.
    return re.sub(r'[aiueo]m?$', '', t)


def casa(a, b):
    """⚠️ El mínimo es de **tres** letras, no cuatro. `Tissabuddhavaṃso` se queda en `tis` al
    perder el sufijo y la vocal final, y con el mínimo en cuatro el capítulo no casaba con su propio
    nombre. Tres basta porque lo que se compara son raíces ya despojadas del sufijo común.
    """
    a, b = norm(a), norm(b)
    return len(a) >= 3 and len(b) >= 3 and (a in b or b in a)


def mejor(nombre, cands, clave):
    """El candidato cuyo nombre más se parece a `nombre`, o `None`.

    ⚠️ **El primero que casa no vale.** Las raíces cortas se contienen unas a otras y
    `Padumuttarabuddhavaṃso` se quedaba con el tramo impreso del **Paduma** —`paduma` es prefijo de
    `padumuttara`—, con lo que la cobertura caía a 0,52 y la fila del Paduma y la del Padumuttara
    apuntaban al mismo sitio. Se elige por parecido máximo, no por orden.
    """
    a = norm(nombre)
    puntos = [(SequenceMatcher(None, a, norm(clave(c))).ratio(), c) for c in cands]
    r, c = max(puntos, key=lambda x: x[0]) if puntos else (0, None)
    return c if r >= 0.75 else None


def capitulos_cst(stem, top):
    """`[(head, texto)]` de los `div` en **orden de documento**.

    ⚠️ Nunca por paranum: en KN el `n` reinicia dentro de cada capítulo/vagga (regla (5-bis)) y
    ordenar por él entrelaza las series — en el Cariyāpiṭaka devuelve 337 unidades donde hay 35.
    """
    raiz = ET.parse(f'{XML}/{stem}.mul.xml').getroot()
    out = []
    for d in raiz.iter('div'):
        if d.get('n') == top:
            continue
        h = [''.join(x.itertext()).strip() for x in d.iter('head')]
        out.append((h[0] if h else '',
                    ap.fold(' '.join(''.join(p.itertext()) for p in d.iter('p')))))
    return out


def cariyas_cst(stem, top):
    """`[(vagga, nº, subhead, texto)]` — las 35 cariyās, en orden de documento."""
    raiz = ET.parse(f'{XML}/{stem}.mul.xml').getroot()
    out = []
    for v, d in enumerate((x for x in raiz.iter('div') if x.get('n') != top), 1):
        act = None
        for p in d.iter('p'):
            txt = ''.join(p.itertext()).strip()
            if p.get('rend') == 'subhead':
                m = re.match(r'\s*(\d+)\.', txt)
                act = [v, int(m.group(1)) if m else len(out) + 1, txt, '']
                out.append(act)
            elif act is not None:
                act[3] += ' ' + txt
    return [(a, b, c, ap.fold(d)) for a, b, c, d in out]


def pagina_bd(conn, libro, unidades):
    """`[(primera página, última página)]` de cada unidad **en la paginación de la BD**.

    Se localiza por el **incipit**: la página cuya cobertura con las primeras palabras de la unidad
    es máxima. No por el encabezado impreso, que puede abarcar más de una unidad del CST —el
    `II -- DĪPAṄKARABUDDHAVAṂSO` cubre la Sumedhapatthanā y el Dīpaṅkara— ni por la página del
    Excel, que en estas dos obras **es de otra composición** y no sirve.

    Sin esto, la tabla obliga a echar cuentas para encontrar el texto, que es tanto como no darlo.
    """
    pgs = {r['page_no']: ap.fold(r['unitext'] or '') for r in conn.execute(
        'SELECT page_no,unitext FROM pages WHERE edition="mula" AND book_no=?', (libro,))}
    ini = []
    for tc in unidades:
        inc = ' '.join(tc.split()[:14])
        ini.append(max((R.cobertura(pgs[p], inc), p) for p in pgs)[1] if inc else None)
    ultima = max(pgs)
    return [(a, (ini[k + 1] - 1 if k + 1 < len(ini) and ini[k + 1] and a and ini[k + 1] > a
                 else ultima)) for k, a in enumerate(ini)]


def plano(conn, libro):
    return [(r['page_no'], ln, l)
            for r in conn.execute('SELECT page_no,unitext FROM pages WHERE edition="mula" '
                                  'AND book_no=? ORDER BY page_no', (libro,))
            for ln, l in enumerate((r['unitext'] or '').split('\n'), 1)]


def tramos_bv(conn):
    """`{romano: (página, texto del tramo)}` — los capítulos que el impreso encabeza.

    El encabezado es `II -- DĪPAṄKARABUDDHAVAṂSO`. El romano se corrobora con el **nombre**, que es
    lo que de verdad identifica el capítulo: la transcripción se come alguna `I` (lee `VII` donde
    imprime `VIII`) y el nombre no deja lugar a duda.
    """
    pl = plano(conn, 41)
    cab = [(k, m.group(1), m.group(2), p) for k, (p, _ln, l) in enumerate(pl)
           if (m := re.match(r'\s*([IVX]+)\s*-{1,2}\s*([A-ZĀĪŪṂṆṬḌÑṄḶ\'\-]{5,})\s*$', l.strip()))]
    out = []
    for j, (k, rom, nom, pg) in enumerate(cab):
        fin = cab[j + 1][0] if j + 1 < len(cab) else len(pl)
        out.append((NUM_ROM.get(rom), nom, pg, ap.fold(' '.join(l for _p, _l, l in pl[k:fin]))))
    return out


def subtitulos_cp(conn):
    """`{(vagga, nº): (nombre, página, texto)}` — el impreso subtitula cada cariyā.

    El vagga sale de **dónde reinicia el número**, que aquí sí es fiable porque están los 35
    subtítulos y los tres vaggas son 10 + 10 + 15. La clave es el par (vagga, nº), y **el nombre es
    la corroboración, no la clave**: el impreso llama `Kurudhammacariyaṃ` a la que el CST titula
    `Kururājacariyā`, y buscar por nombre la deja sin pareja.
    """
    pl = plano(conn, 42)
    cab = [(k, int(m.group(1)), m.group(2), p) for k, (p, _ln, l) in enumerate(pl)
           # el nombre puede llevar espacio y apóstrofo dentro: `8 Saccasa' vhayapaṇḍitacariyaṃ`
           if (m := re.match(r"\s*(\d+)\s+([A-Za-zĀāĪīŪūṂṃṆṇṬṭḌḍÑñṄṅḶḷ' \-]+cariya[ṃm])\s*\d*\s*$",
                             l.strip()))]
    out, vag, prev = {}, 0, 99
    for j, (k, n, nom, pg) in enumerate(cab):
        fin = cab[j + 1][0] if j + 1 < len(cab) else len(pl)
        if n <= prev:
            vag += 1
        prev = n
        out[(vag, n)] = (nom, pg, ap.fold(' '.join(l for _p, _l, l in pl[k:fin])))
    return out


def filas(pref):
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    return [{'num': str(r[ci['Sutta #']]), 'name': str(r[ci['Sutta Name']] or ''),
             'page': r[ci['PTS Page']], 'estado': r[ci['Estado']]}
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[ci['Nikaya']] == 'KN' and str(r[ci['PTS Vol']]) in {'Bv', 'Cp'}
            and str(r[ci['Sutta #']]).startswith(pref)]


def resuelve_bv(conn):
    cst = capitulos_cst('s0511m', 'kn12')
    fs = filas('12.')
    tr = tramos_bv(conn)
    bd = pagina_bd(conn, 41, [tc for _h, tc in cst])
    print(f'CST {len(cst)} capítulos · Excel {len(fs)} filas · impreso {len(tr)} encabezados')
    res, avisos = {}, []
    for k, f in enumerate(fs):
        head, tc = cst[k]
        if not casa(f['name'], head):
            avisos.append(f'   ⚠ {f["num"]}: «{f["name"]}» vs «{head}»')
            continue
        m = mejor(head, tr, lambda t: t[1])
        cov = R.cobertura(m[3], tc) if m else None
        det = (f'Buddhavaṃsa: capítulo {k + 1} de 29 → «{head}». Excel y CST coinciden en nombre y '
               f'orden en los 29')
        if m:
            det += (f'; el impreso lo encabeza «{m[1]}» en la p{m[2]} de la BD, cobertura '
                    f'{cov:.2f}')
        else:
            det += '; el impreso no le pone encabezado romano propio'
        det += (f'. **En la BD el texto está en el libro 41, pp. {bd[k][0]}-{bd[k][1]}** (localizado '
                f'por incipit). ⚠ Esa paginación NO es la que cita el Excel (1-68): son dos '
                f'composiciones PTS distintas, así que la página del Excel no se verifica ni se '
                f'toca — pero la de la BD queda aquí para no tener que buscarla')
        if cov is None or cov >= COV_MIN:
            res[f['num']] = (f's0511m:c{k + 1}', f'Bv {k + 1}', det)
        else:
            avisos.append(f'   ✗ {f["num"]} «{head}» cobertura {cov:.2f}')
    return res, avisos, len(fs)


def resuelve_cp(conn):
    cst = cariyas_cst('s0512m', 'kn13')
    fs = filas('13.')
    sub = subtitulos_cp(conn)
    bd = pagina_bd(conn, 42, [x[3] for x in cst])
    print(f'CST {len(cst)} cariyās · Excel {len(fs)} filas · impreso {len(sub)} subtítulos')
    res, avisos = {}, []
    for k, f in enumerate(fs):
        v, n, head, tc = cst[k]
        # el nombre del Excel es **corroboración**: cuando difiere pero la cobertura del par
        # (vagga, nº) es decisiva, manda la cobertura. `Cp 16 Rurumigarāja` es la `Rururājacariyā`
        # del CST y la impresa `Rururājacariyaṃ` p18 — cobertura 0,99 contra 0,27-0,41 de las
        # vecinas. El nombre del Excel sólo añade «miga», el ciervo.
        if not casa(f['name'], head):
            avisos.append(f'   ⚠ nombre distinto en {f["num"]}: «{f["name"]}» vs «{head}» '
                          f'(decide la cobertura)')
        s = sub.get((v, n))
        cov = R.cobertura(s[2], tc) if s else 0.0
        if not s or cov < COV_MIN:
            avisos.append(f'   ✗ {f["num"]} «{head}» cobertura {cov:.2f}')
            continue
        nom_ok = casa(s[0], head)
        res[f['num']] = (
            f's0512m:c{v}.{n}', f'Cp {v}.{n}',
            f'Cariyāpiṭaka: cariyā {n} del vagga {v} → «{head}», la {k + 1} de 35. El impreso la '
            f'subtitula «{s[0]}» en la p{s[1]} de la BD, cobertura {cov:.2f}'
            + ('' if nom_ok else ' (el impreso usa otro nombre para la misma cariyā)')
            + f'. **En la BD el texto está en el libro 42, pp. {bd[k][0]}-{bd[k][1]}** (localizado '
              f'por incipit). ⚠ Esa paginación NO es la que cita el Excel (73-101): son dos '
              f'composiciones PTS distintas, así que la página del Excel no se verifica ni se '
              f'toca — pero la de la BD queda aquí para no tener que buscarla')
    if len(cst) == len(fs) + 1:
        avisos.append(f'   ⚠ el Excel no tiene fila para «{cst[-1][2]}», que el impreso SÍ trae '
                      f'(p{sub[max(sub)][1]} de la BD) — falta una fila')
    return res, avisos, len(fs)


def main():
    dry = '--dry' in sys.argv
    obra = sys.argv[sys.argv.index('--obra') + 1] if '--obra' in sys.argv else 'bv'
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    res, avisos, total = (resuelve_bv if obra == 'bv' else resuelve_cp)(conn)
    print(f'\n{len(res)}/{total} firmadas')
    for a in avisos:
        print(a)
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-{obra}.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'KN' or str(ws.cell(r, ci['PTS Vol']).value) not in {'Bv', 'Cp'}:
            continue
        num = str(ws.cell(r, ci['Sutta #']).value)
        if num not in res or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, ref, det = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['PTS Ref']).value = ref
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
