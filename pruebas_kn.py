#!/usr/bin/env python3
"""
pruebas_kn — batería funcional para el **Khuddaka**, sobre las 1.495 filas cotejables.

### Por qué no vale la batería de los cuatro nikāyas

`pruebas_concordancia.py` se apoya en que la `PTS Ref` de DN/MN/SN/AN trae **página y línea**
(`S i 1,6`), y verifica que el texto del CST empieza ahí. En KN esa referencia **no existe**: por la
regla (5-ter) la cita es la estrofa o la sección —`Dhp 21-32`, `Vv 12`, `Cp 3.15`, `It 5`—, que es lo
correcto para estas obras y lo inútil para aquel control. Y encima el contenido es donde menos
discrimina de todo el canon: los suttas del Udāna abren con la misma fórmula, los vimānas van en
pares deliberadamente repetidos, el Niddesa es prosa exegética que se repite a sí misma.

Así que hay que buscar evidencia de otra naturaleza. La idea es ésta:

> **Cuando no puedes verificar el punto, verifica el orden y la medida.**

Dos ordenaciones **independientes** —cómo PTS imprime y cómo el CST numera— tienen que coincidir, y
el tamaño de cada unidad tiene que ser proporcional al espacio que ocupa impresa. Una fila mal
asignada rompe una de las dos aunque su texto se parezca al del vecino; y las dos son ciegas a cómo
se hizo la alineación, que es lo que las hace prueba y no tautología.

### Los cuatro controles

1. **`orden`** — la `VRI Ref` no puede retroceder a lo largo del orden impreso. Es el control que
   caza la familia de defectos que más daño ha hecho en este proyecto: el **desplazamiento** (S ii
   con SN 17 a +7; `SN 46.36`/`46.37` a +1). Un solo par intercambiado produce una **inversión**, y
   la inversión se ve sin mirar una letra del texto.
2. **`biyección`** — cada unidad del CST dentro del tramo de la obra la reclama **exactamente una**
   fila, y cada fila reclama una. Caza el rango colapsado (varias filas sobre lo mismo) y el hueco
   (una unidad que nadie reclama).
3. **`medida`** — el texto del CST y el impreso tienen que **avanzar juntos**. No fila a fila: así
   se mide ruido, porque en KN muchas unidades cortas comparten página y la última de cada obra no
   tiene siguiente. Se mide **acumulado**: qué fracción del texto del CST lleva consumida la obra al
   llegar a la fila `k`, contra qué fracción de sus páginas impresas lleva recorridas. Las dos
   curvas van del 0 al 1 y, si la alineación es buena, **casi juntas**. Una fila mal asignada
   despega una de la otra y se ve como un salto en la diferencia.

   ⚠️ Y lo que se mide es **el salto, no el nivel**. Las dos curvas pueden separarse poco a poco sin
   que nada esté mal: el Cūḷaniddesa deriva de 0,00 a −0,28 a lo largo de dieciocho filas porque la
   densidad del impreso cambia, no porque haya un error. **Una fila mal asignada desplaza la curva
   de golpe**; una composición distinta la desplaza despacio. Así que se compara cada fila con la
   anterior y se marca la **discontinuidad**.

   Es el control que funciona donde el contenido no discrimina, porque no compara textos: compara
   **cuánto** texto hay y **dónde** va. Y es ciego a cómo se hizo la alineación.
4. **`contenido`** — cobertura de la ventana impresa contra su unidad del CST y contra las vecinas.
   Sólo cuenta si gana por margen; en KN muchas veces no puede, y entonces se dice.

### Los tres veredictos, como en la otra batería

**PASA** si algún control discriminante confirma; **FALLA** si alguno **contradice**; **INDECISO** si
ninguno pudo decidir. Un indeciso no cuenta como aprobado.

### Y la prueba de la prueba: mutación

Una batería que aprueba 1.495 de 1.495 puede ser una batería que no detecta nada, así que se la
somete a **mutaciones deliberadas**: intercambiar dos filas vecinas, desplazar una un puesto,
duplicar la referencia de la de al lado, apuntar a un sutta lejano, cambiar un rango. **Diez de
diez cazadas** (`--mutar`), y las dos que al principio se escapaban obligaron a corregir la regla de
biyección: había dado por legítimo compartir un rango sin más, y compartir explicado es un dato
mientras que compartir a secas es un defecto.

Uso: python3 pruebas_kn.py [--obra Vv|Th-ap|…] [--mutar] [--detalle] [--json salida.json]
"""
import json
import re
import sqlite3
import statistics as st
import sys
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
XMLDIR = '/tmp/tipitaka-xml/romn'
# sigla del CPD → libro de la BD. `Bv` y `Cp` van aparte: su paginación no es la de la BD y el
# tramo real lo trae el `Detail` de cada fila.
LIBRO = {'Khp': 22, 'Dhp': 23, 'Ud': 24, 'It': 25, 'Sn': 26, 'Vv': 27, 'Pv': 28, 'Th': 29,
         'Thī': 29, 'Nidd I': 36, 'Nidd II': 37, 'Paṭis': 38, 'Th-ap': 40, 'Thī-ap': 40,
         'Bv': 41, 'Cp': 42}
DESVIO = 0.15       # cuánto pueden separarse las dos progresiones acumuladas (0 = idénticas)
MARGEN = 0.15

_CST, _DIVS = {}, {}


def unidades(stem):
    if stem not in _CST:
        try:
            _CST[stem] = {x['pn']: x['texto'] for x in ap.cst_unidades(stem)}
        except Exception:
            _CST[stem] = {}
    return _CST[stem]


def divs(stem):
    """`[texto de cada div]` en orden de documento — para las `VRI Ref` con `c<capítulo>`."""
    if stem not in _DIVS:
        try:
            raiz = ET.parse(f'{XMLDIR}/{stem}.mul.xml').getroot()
            _DIVS[stem] = [d for d in raiz.iter('div') if '_' in (d.get('n') or '')]
        except Exception:
            _DIVS[stem] = []
    return _DIVS[stem]


def clave(vri):
    """`(fichero, capítulo, número, ítem)` — clave **ordenable** de una `VRI Ref`, en sus 5 formas.

    Es lo que permite preguntar si el orden del CST retrocede: sin una clave total no hay orden que
    comprobar. Las cinco formas del proyecto —`N`, `N-M`, `N.i`, `cN`, `cN.i`— caen todas aquí.
    """
    v = str(vri or '')
    m = re.match(r'^([a-z0-9]+):c(\d+)(?:\.(\d+)(?:-(\d+))?)?$', v)
    if m:
        return (m.group(1), int(m.group(2)), int(m.group(3) or 0), 0)
    m = re.match(r'^([a-z0-9]+):(\d+)(?:\.(\d+))?(?:-(\d+))?$', v)
    if m:
        return (m.group(1), 0, int(m.group(2)), int(m.group(3) or 0))
    return None


def comparte_declarado(f):
    """¿Declara la propia fila que su referencia se comparte a propósito?

    ⚠️ **La excepción se declara, no se adivina.** Intenté inferirla —«si el CST repite el mismo
    título a lo largo del rango, compartir vale»— y una prueba de mutación lo tumbó: dentro de una
    sola unidad que abarca once paranums el título tampoco cambia, así que la regla daba por
    legítima cualquier duplicación. Es el mismo principio que ya rige en `check_integrity` para el
    `deest in Be`: la ausencia registrada es un dato, la ausencia sin explicar es un error. Aquí
    igual — compartir explicado es un dato, compartir a secas es un defecto.
    """
    return 'comparte' in str(f['Detail'] or '').lower()


def texto_cst(vri):
    """Texto del CST de esa referencia, en cualquiera de las cinco formas."""
    v = str(vri or '')
    # ⚠️ **Cinco formas**, no cuatro: además de `N`, `N-M`, `N.i` y `cN`, el Paṭisambhidāmagga y el
    # Niddesa usan `c<capítulo>.<a>-<b>` —capítulo más **rango de paranums dentro de él**—, porque
    # ahí el `n` reinicia en cada vagga (regla (5-bis)). Sin contemplarla, sus 31 filas «no apuntan
    # a texto» y la obra entera falla: era la prueba, no el dato.
    m = re.match(r'^([a-z0-9]+):c(\d+)(?:\.(\d+)(?:-(\d+))?)?$', v)
    if m:                                        # capítulo, unidad o rango dentro del capítulo
        ds = divs(m.group(1))
        c = int(m.group(2))
        if c > len(ds):
            return None
        d = ds[c - 1]
        if not m.group(3):
            return ap.fold(' '.join(''.join(p.itertext()) for p in d.iter('p')))
        if m.group(4):                           # rango de paranums dentro del capítulo
            a, b = int(m.group(3)), int(m.group(4))
            out = [''.join(p.itertext()) for p in d.iter('p')
                   if p.get('n') and a <= int(p.get('n')) <= b]
            return ap.fold(' '.join(out)) or None
        # ⚠️ **`c<cap>.<n>` es ambiguo y hay que desambiguarlo por el dato.** En el Cariyāpiṭaka el
        # `n` es el **ítem** —la cariyā k-ésima del vagga, porque sus `<p>` no llevan `n` ninguno—;
        # en el Paṭisambhidāmagga es un **paranum dentro del capítulo**, coherente con la forma de
        # rango `c<cap>.<a>-<b>` que usa la misma obra. Es una arruga de la clave que esta prueba
        # destapó: se resuelve mirando si el capítulo numera sus párrafos y si alguno es ese número.
        # La regla: **el ítem manda, y el paranum es el recurso**. El Apadāna usa `c33.9` para la
        # novena unidad del vagga 33 —y su capítulo numera además los párrafos, así que leer el 9
        # como paranum devolvía otro texto y tumbaba 19 filas—. El Paṭisambhidāmagga usa `c1.236`,
        # donde 236 excede con mucho el número de unidades del capítulo: ahí sólo cabe el paranum.
        nn = int(m.group(3))
        subs = [q for q in d.iter('p') if q.get('rend') in ('subhead', 'title')]
        if nn > len(subs):
            porn = [q for q in d.iter('p') if q.get('n') and int(q.get('n')) == nn]
            if porn:
                return ap.fold(' '.join(''.join(q.itertext()) for q in porn))
        act, k, out = None, 0, []
        for p in d.iter('p'):
            if p.get('rend') in ('subhead', 'title'):
                k += 1
                act = k == nn
            elif act:
                out.append(''.join(p.itertext()))
        return ap.fold(' '.join(out)) or None
    m = re.match(r'^([a-z0-9]+):(\d+)(?:\.(\d+))?(?:-(\d+))?$', v)
    if m:
        u = unidades(m.group(1))
        a = int(m.group(2))
        b = int(m.group(4) or a)
        t = ' '.join(u[n] for n in range(a, b + 1) if n in u)
        if m.group(3) and t:                     # ítem dentro del paranum
            k = int(m.group(3))
            ini = t.find(f'({k})')
            fin = t.find(f'({k + 1})', ini + 1) if ini >= 0 else -1
            if ini >= 0:
                t = t[ini:fin if fin > ini else None]
        return ap.fold(t) if t.strip() else None
    return None


def filas():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = defaultdict(list)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ci['Nikaya']] != 'KN' or str(r[ci['Type']]) == 'Section Header':
            continue
        if not r[ci['VRI Ref']]:
            continue
        out[str(r[ci['PTS Vol']])].append({k: r[ci[k]] for k in hdr})
    return out


def tramo_bd(f):
    """`(libro, primera, última)` en la BD. En `Bv`/`Cp` lo trae el `Detail`, porque su paginación
    del Excel es de otra composición PTS y no sirve para buscar en la base."""
    m = re.search(r'En la BD el texto está en el libro (\d+), pp\. (\d+)-(\d+)',
                  str(f['Detail'] or ''))
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    lib = LIBRO.get(str(f['PTS Vol']))
    pg = f['PTS Page']
    return (lib, pg, pg) if lib and isinstance(pg, int) else (None, None, None)


def analiza(obra, fs, conn, dueños_kn=None):
    """Los cuatro controles sobre una obra entera."""
    res = [{'fila': str(f['Sutta #']), 'ref': str(f['PTS Ref']), 'vri': str(f['VRI Ref']),
            'nombre': str(f['Sutta Name'])[:34]} for f in fs]

    # ── 1 · orden: la clave del CST no puede retroceder a lo largo del orden impreso
    ks = [clave(f['VRI Ref']) for f in fs]
    for n, k in enumerate(ks):
        res[n]['orden'] = None if k is None else True
    for n in range(1, len(ks)):
        if ks[n] and ks[n - 1] and ks[n] < ks[n - 1]:
            res[n]['orden'] = res[n - 1]['orden'] = False

    # ── 2 · biyección: cada referencia, una fila
    #
    # ⚠️ Un rango **puede** compartirse, pero no porque sea un rango: sólo cuando el CST **repite el
    # mismo título** a lo largo de él —el `Navātasuttaṃ` de S iii ocupa cincuenta paranums seguidos
    # con el mismo nombre— y entonces el rango es de veras el locus de varias unidades impresas.
    # Permitirlo sin esa condición deja pasar una duplicación de verdad: una prueba de mutación lo
    # enseñó, duplicando a propósito la referencia de una fila vecina del Petavatthu y viendo que
    # la batería no la cazaba.
    # ⚠️ El recuento es **de todo KN, no de la obra**: dos obras pueden reclamar el mismo texto del
    # CST, y hay que verlo. Pasa de verdad — `Sn 5.18` y `Cnd 3` apuntan los dos a
    # `s0505m:1130-1136`, porque PTS imprime esas gāthās en las dos obras y el CST las tiene una
    # sola vez, en el Suttanipāta. Es legítimo y por eso va declarado; pero contarlo por obra lo
    # habría dejado invisible.
    dueños = dueños_kn if dueños_kn is not None else Counter(str(f['VRI Ref']) for f in fs)
    for n, f in enumerate(fs):
        v = str(f['VRI Ref'])
        res[n]['unica'] = dueños[v] == 1 or comparte_declarado(f)

    # ── 3 · medida: las dos progresiones acumuladas, la del CST y la del impreso
    #
    # ⚠️ **Por tramo continuo, no por obra.** Una obra puede cambiar de libro o de fichero a mitad
    # y entonces la progresión no significa nada: el Paṭisambhidāmagga **reinicia la paginación** en
    # su volumen II (pág. 1 otra vez, libro 39) y el Cūḷaniddesa tiene dos filas cuyo texto está en
    # el Suttanipāta, otro fichero. Medir eso de corrido daba 32 falsos positivos. Se corta el tramo
    # donde cambia el libro, cambia el fichero del CST, o la página retrocede.
    car, pag, lib_ = [], [], []
    for n, f in enumerate(fs):
        tc = texto_cst(f['VRI Ref'])
        res[n]['resuelve'] = tc is not None
        car.append(len(tc or ''))
        l, a, _b = tramo_bd(f)
        pag.append(a)
        k = clave(f['VRI Ref'])
        lib_.append((l, k[0] if k else None))
    cortes, ini = [], 0
    for n in range(1, len(fs)):
        if lib_[n] != lib_[n - 1] or (pag[n] and pag[n - 1] and pag[n] < pag[n - 1]):
            cortes.append((ini, n))
            ini = n
    cortes.append((ini, len(fs)))
    for d in res:
        d['medida'] = None
    for a_, b_ in cortes:
        seg = range(a_, b_)
        tot = sum(car[n] for n in seg)
        ps = [pag[n] for n in seg if pag[n]]
        if not tot or len(ps) < 2 or max(ps) == min(ps):
            continue
        p0, p1, acum = min(ps), max(ps), 0
        for n in seg:
            acum += car[n]
            if not pag[n]:
                continue
            # la página se toma **al final** de la fila —donde empieza la siguiente—, para que las
            # dos fracciones midan lo mismo: cuánto se lleva recorrido *después* de esta unidad
            sig = next((pag[m] for m in range(n + 1, b_) if pag[m]), p1)
            res[n]['_desv'] = acum / tot - (sig - p0) / (p1 - p0)
            res[n]['_peso'] = round(car[n] / tot, 2)
        # el salto respecto de la fila anterior del mismo tramo: la deriva lenta no cuenta
        prev = None
        for n in seg:
            if '_desv' not in res[n]:
                continue
            res[n]['medida'] = round(res[n]['_desv'] - (prev if prev is not None else
                                                       res[n]['_desv']), 2)
            prev = res[n]['_desv']

    # ── 4 · contenido: contra su unidad y contra las vecinas
    for n, f in enumerate(fs):
        res[n]['cov'] = res[n]['cov_vec'] = None
        lib, a, b = tramo_bd(f)
        tc = texto_cst(f['VRI Ref'])
        if not (lib and a and tc):
            continue
        sig = next((tramo_bd(g)[1] for g in fs[n + 1:] if tramo_bd(g)[1]), b)
        tp = ' '.join(ap.fold(x['unitext'] or '') for x in conn.execute(
            'SELECT unitext FROM pages WHERE edition="mula" AND book_no=? AND page_no BETWEEN ? '
            'AND ?', (lib, max(1, a), max(a, b, (sig or a)))))
        if not tp:
            continue
        res[n]['cov'] = round(R.cobertura(tp, tc), 2)
        vec = [texto_cst(fs[n + o]['VRI Ref']) for o in (-1, 1, 5) if 0 <= n + o < len(fs)]
        vec = [R.cobertura(tp, v) for v in vec if v]
        res[n]['cov_vec'] = round(max(vec), 2) if vec else None

    for d in res:
        cont_ok = (d['cov'] is not None and d['cov_vec'] is not None
                   and d['cov'] >= d['cov_vec'] + MARGEN)
        # ⚠️ El umbral **crece con lo que pesa la propia fila**: una unidad que es el 20 % del texto
        # de su obra desplaza la curva un 20 % ella sola, y eso no es un defecto sino su tamaño. Sin
        # esto, las obras cortas con una unidad dominante —el Cūḷaniddesa— salían todas marcadas.
        medida_mal = (d['medida'] is not None
                      and abs(d['medida']) > DESVIO + d.get('_peso', 0))
        if not d['resuelve']:
            d['veredicto'], d['motivo'] = 'FALLA', 'la VRI Ref no apunta a texto del CST'
        elif d['orden'] is False:
            d['veredicto'] = 'FALLA'
            d['motivo'] = 'la referencia del CST RETROCEDE respecto de la fila anterior'
        elif not d['unica']:
            d['veredicto'] = 'FALLA'
            d['motivo'] = f'la VRI Ref (unidad única) la reclaman {dueños[d["vri"]]} filas'
        elif medida_mal:
            d['veredicto'] = 'FALLA'
            d['motivo'] = (f'las dos progresiones se separan {d["medida"]:+}: el texto del CST y '
                           f'el impreso no avanzan juntos en este punto')
        elif cont_ok:
            d['veredicto'], d['motivo'] = 'PASA', 'orden + medida + contenido'
        elif d['orden'] and d['unica']:
            d['veredicto'] = 'PASA'
            d['motivo'] = 'orden + biyección + medida (el contenido no discrimina en esta obra)'
        else:
            d['veredicto'], d['motivo'] = 'INDECISO', 'ningún control pudo decidir'
    return res


MUTACIONES = [
    ('Pv', 'duplicar la referencia de la vecina',
     lambda fs: fs.__setitem__(20, {**fs[20], 'VRI Ref': fs[19]['VRI Ref']})),
    ('Sn', 'desplazar una fila un puesto',
     lambda fs: fs.__setitem__(30, {**fs[30], 'VRI Ref': fs[31]['VRI Ref']})),
    ('Th', 'duplicar la referencia de la vecina',
     lambda fs: fs.__setitem__(100, {**fs[100], 'VRI Ref': fs[99]['VRI Ref']})),
    ('Vv', 'intercambiar dos filas vecinas',
     lambda fs: (fs.__setitem__(11, {**fs[11], 'VRI Ref': fs[12]['VRI Ref']}),
                 fs.__setitem__(12, {**fs[12], 'VRI Ref': 's0506m:101-107'}))),
    ('It', 'desplazar una fila un puesto',
     lambda fs: fs.__setitem__(50, {**fs[50], 'VRI Ref': 's0504m:52'})),
    ('Ud', 'apuntar a un sutta lejano',
     lambda fs: fs.__setitem__(40, {**fs[40], 'VRI Ref': 's0503m:75'})),
    ('Th-ap', 'desplazar una fila un puesto',
     lambda fs: fs.__setitem__(300, {**fs[300], 'VRI Ref': 's0510m1:c31.2'})),
    ('Cp', 'desplazar una fila un puesto',
     lambda fs: fs.__setitem__(15, {**fs[15], 'VRI Ref': 's0512m:c2.7'})),
    ('Dhp', 'cambiar un rango entero',
     lambda fs: fs.__setitem__(10, {**fs[10], 'VRI Ref': 's0502m:200-220'})),
    ('Paṭis', 'desplazar una fila un puesto',
     lambda fs: fs.__setitem__(10, {**fs[10], 'VRI Ref': fs[11]['VRI Ref']})),
]


def mutar(conn):
    """¿Detecta la batería un defecto puesto a mano? Sin esto, un 100 % no dice nada."""
    todo, ok = filas(), 0
    for obra, desc, mut in MUTACIONES:
        base = Counter(d['veredicto'] for d in analiza(obra, todo[obra], conn))
        fs = [dict(x) for x in todo[obra]]
        mut(fs)
        tras = Counter(d['veredicto'] for d in analiza(obra, fs, conn))
        caza = tras['FALLA'] > base['FALLA']
        ok += caza
        print(f"   {obra:8} {desc:38} FALLA {base['FALLA']} → {tras['FALLA']}  "
              f"{'✓ cazada' if caza else '✗ NO LA VE'}")
    print(f'\n   mutaciones cazadas: {ok}/{len(MUTACIONES)}')
    return 0 if ok == len(MUTACIONES) else 1


def main():
    detalle = '--detalle' in sys.argv
    solo = sys.argv[sys.argv.index('--obra') + 1] if '--obra' in sys.argv else None
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    if '--mutar' in sys.argv:
        return mutar(conn)
    todo, total, salida = filas(), Counter(), {}
    dueños_kn = Counter(str(f['VRI Ref']) for fs in todo.values() for f in fs)
    for obra in sorted(todo, key=lambda o: -len(todo[o])):
        if solo and obra != solo:
            continue
        res = analiza(obra, todo[obra], conn, dueños_kn)
        c = Counter(r['veredicto'] for r in res)
        total.update(c)
        salida[obra] = res
        marca = '✅' if not c['FALLA'] and not c['INDECISO'] else ('❌' if c['FALLA'] else '⚠️')
        print(f"{marca} {obra:8} {len(res):4} · PASA {c['PASA']:4} · FALLA {c['FALLA']:3} · "
              f"INDECISO {c['INDECISO']:3}")
        for d in res:
            if detalle or d['veredicto'] != 'PASA':
                print(f"     {d['veredicto']:9} {d['fila']:11} {d['ref']:14} {d['vri']:18} "
                      f"medida={d['medida']} cov={d['cov']}/{d['cov_vec']} — {d['motivo']}")
    print(f"\n{'═' * 78}\nTOTAL {sum(total.values())} pruebas · PASA {total['PASA']} · "
          f"FALLA {total['FALLA']} · INDECISO {total['INDECISO']}")
    if '--json' in sys.argv:
        with open(sys.argv[sys.argv.index('--json') + 1], 'w') as fh:
            json.dump(salida, fh, ensure_ascii=False, indent=1)
    return 1 if total['FALLA'] else 0


if __name__ == '__main__':
    sys.exit(main())
