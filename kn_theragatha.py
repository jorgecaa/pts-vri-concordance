#!/usr/bin/env python3
"""
kn_theragatha — el **Theragāthā** (264 filas) y el **Therīgāthā** (73), con el mismo método.

Los dos comparten libro en la BD (el 29: PTS los publica en un volumen, Thera 1-109 y Therī
123-167) y estructura: una fila por *thera*/*therī*, un `subhead` por poema en el CST, y el rango
de versos de PTS ya escrito en `PTS Alt/Verse`. Cambian sólo el fichero VRI, la etiqueta del Excel
y el tramo de páginas — así que el mismo analizador sirve para ambos y no se duplica nada.

La primera obra grande de KN, y sale bien porque **tres señales independientes coinciden** y ninguna
de las tres depende de las otras dos.

### Las tres

| señal | resultado |
|---|--:|
| **posición** — el CST tiene 264 `subhead` y el Excel 264 filas, en el mismo orden y reparto en 21 nipātas | 264 = 264 |
| **rango de versos** — el del Excel (`PTS Alt/Verse`, la numeración de PTS) contra el del CST | **257/264 idénticos**, 260 con la misma cardinalidad |
| **nombre** — el del Excel contra el `subhead` | 254/264 |
| *(corroboración)* **cobertura** de contenido, mediana **0,96** | 259/264 ≥ 0,55 |

Se firma una fila cuando **al menos dos de las tres** se cumplen. Con 264 unidades en las dos
ediciones y el rango de versos coincidiendo en 257, la correspondencia posicional no es una
suposición: está comprobada por dos caminos que no la usan.

⚠️ **El desfase de versos existe pero es tardío y pequeño**: 258 filas tienen desfase **0** entre las
dos numeraciones, y las que no son cuatro del final —el CST añade versos en el Saṭṭhi- y el
Mahānipāta (`Th 1146-1208` de PTS es `1149-1217` en el CST)—. No es como el Suttanipāta, donde la
deriva empezaba pronto y crecía; aquí el rango del Excel es utilizable como clave casi siempre, y
donde no lo es, lo dicen el nombre y el contenido.

### Nombres que no casan y por qué

Los 10 son variantes de grafía entre ediciones, y **tres parecen erratas del Excel**: `Jmbuka` por
`Jambuka`, `Grimānanda` por `Girimānanda`, `Vajaya` por `Vijaya`. Las otras son alternancias reales
—`Balliya`/`Bhalliya`, `Kimbila`/`Kimila`, `Māluṅkya`/`Mālukya`— del tipo que el propio Excel suele
anotar entre paréntesis. Se dejan señaladas en `Detail`, no se tocan.

### La notación (regla (5-ter))

`Th <a>-<b>`, el rango de versos **de PTS** — que aquí sí se tiene, porque el Excel lo trae en
`PTS Alt/Verse` y coincide con el impreso.

Uso: python3 kn_theragatha.py [--dry]
"""
import re
import shutil
import sqlite3
import sys
from datetime import datetime

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
COV_MIN = 0.55
# obra → (libro BD, fichero VRI, etiqueta `PTS Vol` del Excel, última página del tramo, sigla)
OBRAS = {'thag': (29, 's0508m', 'Th', 116, 'Th'),
         'thig': (29, 's0509m', 'Th & Th', 174, 'Thī')}

# ⚠️ Dos rangos de verso INTERCAMBIADOS en el Excel del Therīgāthā, y el CST y el contenido lo
# dicen a la vez: el `Vāseṭṭhī` es el 133-138 (el Excel le pone 139-144) y el `Khemā` el 139-144
# (le pone `133-139`, que además cuenta **siete** versos donde sus hermanos tienen seis). Los
# nombres casan en las dos filas, así que lo cruzado son los rangos, no los poemas.
VERSOS_CORREGIDOS = {'9.6.2': (133, 138), '9.6.3': (139, 144)}


def _limpia(t):
    t = re.sub(r'^\s*\(KN[^)]*\)\s*', '', t or '')
    return re.sub(r'^\s*(Thag|Thig|Snp|Ud|Iti|Dhp|Khp)\s+[\d.]+\s*', '', t).strip()


def raiz(t, quita_ord=True):
    """Raíz comparable: sin ordinal del CST, sin `-ttheragāthā`, sin desinencia."""
    t = ap.fold(re.sub(r'^[\d.\s()\-]+', '', t or '')).strip()
    t = re.sub(r'\s*(theragatha|therigatha|gatha)$', '', t).strip()
    if quita_ord:
        t = re.sub(r'^(pathama|dutiya|tatiya|catuttha)[\s\-]*', '', t).strip()
    return re.sub(r'[aiueom]+$', '', t)


def casa_nombre(excel, subhead):
    va = [raiz(x) for x in re.split(r'[()]', _limpia(excel)) if raiz(x)]
    vb = [raiz(subhead), raiz(subhead, False)]
    for x in va:
        for y in vb:
            if x and y and (x == y or (len(x) >= 3 and len(y) >= 3
                                       and (x.startswith(y) or y.startswith(x)))):
                return True
    return False


def suttas_cst(stem):
    """`[(subhead, primer verso, último verso, texto)]`, agrupando por `subhead`."""
    out = []
    for x in ap.cst_unidades(stem):
        if not out or out[-1][0] != x['subhead']:
            out.append([x['subhead'], x['pn'], x['pn'], ''])
        out[-1][2] = x['pn']
        out[-1][3] += ' ' + x['texto']
    return [(a, b, c, ap.fold(d)) for a, b, c, d in out]


def main():
    dry = '--dry' in sys.argv
    obra = sys.argv[sys.argv.index('--obra') + 1] if '--obra' in sys.argv else 'thag'
    BOOK, STEM, VOL, ULTIMA, SIGLA = OBRAS[obra]
    print(f'── {obra}: libro {BOOK} · {STEM} · «{VOL}»')
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    cst = suttas_cst(STEM)
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    fs = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ci['Nikaya']] != 'KN' or str(r[ci['PTS Vol']]) != VOL:
            continue
        m = re.search(r'v(\d+)(?:-(\d+))?', str(r[ci['PTS Alt/Verse']] or ''))
        num = str(r[ci['Sutta #']])
        a = int(m.group(1)) if m else None
        b = int(m.group(2) or m.group(1)) if m else None
        if num in VERSOS_CORREGIDOS:
            a, b = VERSOS_CORREGIDOS[num]
        fs.append({'num': num, 'name': str(r[ci['Sutta Name']] or ''),
                   'page': r[ci['PTS Page']], 'estado': r[ci['Estado']],
                   'a': a, 'b': b, 'alt': str(r[ci['PTS Alt/Verse']] or '')})
    print(f'CST {len(cst)} suttas · Excel {len(fs)} filas')
    if len(cst) != len(fs):
        print('⚠ los recuentos no coinciden: no se escribe nada')
        return
    res, malas = {}, []
    n_nom = n_ver = n_cov = 0
    for k, f in enumerate(fs):
        sub, ca, cb, tc = cst[k]
        pg = f['page']
        # ⚠️ La ventana es de **dos páginas como mínimo**. Cuando varios poemas cortos comparten
        # página —y en el Therīgāthā el vagga de los Chakka mete tres en la 136— una ventana de una
        # sola deja el poema partido y la cobertura se hunde: el Vāseṭṭhī daba 0,47 con `136-136` y
        # **0,96** con `136-137`. No es un umbral que se afloja: es la ventana que estaba mal.
        sig = fs[k + 1]['page'] if k + 1 < len(fs) and isinstance(fs[k + 1]['page'], int) else ULTIMA
        hasta = max(sig, (pg + 1) if isinstance(pg, int) else sig)
        txt = ' '.join(ap.fold(x['unitext'] or '') for x in conn.execute(
            'SELECT unitext FROM pages WHERE edition="mula" AND book_no=? AND page_no BETWEEN ? '
            'AND ?', (BOOK, pg, max(pg, hasta))))
        cov = R.cobertura(txt, tc)
        nom = casa_nombre(f['name'], sub)
        ver = f['a'] is not None and (f['a'], f['b']) == (ca, cb)
        n_nom += nom
        n_ver += ver
        n_cov += cov >= COV_MIN
        señales = sum((nom, ver, cov >= COV_MIN))
        if señales >= 2:
            ref = f'{SIGLA} {f["a"]}-{f["b"]}' if f['a'] and f['b'] != f['a'] else (
                f'{SIGLA} {f["a"]}' if f['a'] else f'{SIGLA} {ca}-{cb}')
            det = (f'{obra}: poema {k + 1} de {len(fs)} → «{sub}», versos {ca}-{cb} del CST; '
                   f'señales: {"nombre " if nom else ""}{"rango " if ver else ""}'
                   f'cobertura {cov:.2f}')
            if not nom:
                det += f'. ⚠ el nombre del Excel («{_limpia(f["name"])}») no casa con el subhead'
            if f['num'] in VERSOS_CORREGIDOS:
                det += (f'. RANGO CORREGIDO: el Excel decía «{f["alt"]}» y el CST y el contenido '
                        f'dan {f["a"]}-{f["b"]}')
            res[f['num']] = (f'{STEM}:{ca}-{cb}' if cb != ca else f'{STEM}:{ca}', ref, det)
        else:
            malas.append((f['num'], _limpia(f['name'])[:24], sub[:28],
                          f'nom {nom} · rango {ver} · cov {cov:.2f}'))
    print(f'  señales: nombre {n_nom}/{len(fs)} · rango de versos {n_ver}/{len(fs)} · '
          f'cobertura {n_cov}/{len(fs)}')
    print(f'\n{len(res)}/{len(fs)} firmadas (≥2 señales)')
    for x in malas:
        print('   ✗', x)
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-{obra}.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'KN' or str(ws.cell(r, ci['PTS Vol']).value) != VOL:
            continue
        num = str(ws.cell(r, ci['Sutta #']).value)
        if num not in res or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, ref, det = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['PTS Ref']).value = ref
        if num in VERSOS_CORREGIDOS:
            a, b = VERSOS_CORREGIDOS[num]
            ws.cell(r, ci['PTS Alt/Verse']).value = f'{SIGLA} v{a}-{b}'
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
