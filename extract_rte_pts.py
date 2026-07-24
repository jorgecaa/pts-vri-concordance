#!/usr/bin/env python3
"""
Extract PTS cross-references from Royal Thai (BUDSIR) edition.
Format: (pts. X r, NNN) embedded in the text at PTS page boundaries.
Maps PTS ref → Royal Thai position for orthogonal validation.
"""
import re, json, os
from collections import defaultdict

RTE_DIR = '/home/jorge/Code/tipitaka.rte/Canonical'

# Royal Thai file → Nikaya mapping
# Files are: 09-11=D, 12-14=M, 15-19=S, 20-24=A, 25-33=KN
FILE_MAP = {
    '09': ('DN', 'D', 'i'), '10': ('DN', 'D', 'ii'), '11': ('DN', 'D', 'iii'),
    '12': ('MN', 'M', 'i'), '13': ('MN', 'M', 'ii'), '14': ('MN', 'M', 'iii'),
    '15': ('SN', 'S', 'i'), '16': ('SN', 'S', 'ii'), '17': ('SN', 'S', 'iii'),
    '18': ('SN', 'S', 'iv'), '19': ('SN', 'S', 'v'),
    '20': ('AN', 'A', 'i'), '21': ('AN', 'A', 'ii'), '22': ('AN', 'A', 'iii'),
    '23': ('AN', 'A', 'iv'), '24': ('AN', 'A', 'v'),
}

def parse_pts_refs(filepath, nikaya, vol_letter, roman):
    """Extract all (pts. X r, NNN) references with their page positions."""
    refs = []
    current_page = 0
    
    with open(filepath, encoding='utf-8-sig') as f:
        for line in f:
            line = line.rstrip('\n')
            
            # Track Royal Thai page numbers
            m = re.match(r'^page number: (\d{3})$', line)
            if m:
                current_page = int(m.group(1))
                continue
            
            # Find PTS references: (pts. X r, NNN) or (pts. X, NNN)
            for m in re.finditer(r'\(pts\.\s+([a-z]+)(?:\s+([ivxlcdm]+))?,?\s+(\d+)\)', line, re.IGNORECASE):
                pts_letter = m.group(1).lower()
                pts_roman = m.group(2)
                pts_page = int(m.group(3))
                
                # Verify the letter matches expected
                expected_letter = vol_letter.lower()
                if pts_letter != expected_letter:
                    continue
                
                refs.append({
                    'pts_ref': f'{vol_letter} {roman} {pts_page}',
                    'pts_letter': pts_letter,
                    'pts_roman': roman,
                    'pts_page': pts_page,
                    'rte_page': current_page,
                    'rte_file': os.path.basename(filepath),
                })
    
    return refs

# ── Extract ──
print('Extracting PTS references from Royal Thai edition...')
all_refs = []

for filename in sorted(os.listdir(RTE_DIR)):
    if not filename.endswith('.txt'): continue
    prefix = filename.split('-')[0]
    if prefix not in FILE_MAP: continue
    
    nikaya, vol_letter, roman = FILE_MAP[prefix]
    filepath = os.path.join(RTE_DIR, filename)
    
    refs = parse_pts_refs(filepath, nikaya, vol_letter, roman)
    all_refs.extend(refs)
    print(f'  {filename}: {len(refs)} PTS refs')

print(f'\nTotal PTS refs from Royal Thai: {len(all_refs)}')

# ── Build lookup: PTS ref → [RTE pages] ──
from collections import defaultdict
pts_to_rte = defaultdict(list)
for r in all_refs:
    pts_to_rte[r['pts_ref']].append(r['rte_page'])

# ── Compare with our Excel ──
from openpyxl import load_workbook
wb = load_workbook('PTS_Reference_Complete_Canon.xlsx')
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

print('\nCross-referencing with Excel...')
print('=' * 80)

matches = 0
mismatches = 0
no_rte_ref = 0
mismatch_examples = []

for ri in range(2, ws.max_row + 1):
    nik = str(ws.cell(row=ri, column=cols['Nikaya']).value or '')
    if nik not in ('DN', 'MN', 'SN', 'AN'): continue
    snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    if not snum or snum == 'None': continue
    
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower()
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    vol = str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip()
    
    # Build PTS ref key: "D i 47", "M i 6", etc.
    pts_key = f'{vol} {roman} {page}'
    
    if pts_key in pts_to_rte:
        matches += 1
    else:
        no_rte_ref += 1
        if no_rte_ref <= 10:
            name = ws.cell(row=ri, column=cols['Sutta Name']).value
            ref = ws.cell(row=ri, column=cols['PTS Ref']).value
            mismatch_examples.append((snum, ref, name))

print(f'Excel entries with matching RTE PTS ref: {matches}')
print(f'Excel entries without RTE PTS ref:     {no_rte_ref}')
print(f'\nSample without RTE ref:')
for snum, ref, name in mismatch_examples:
    print(f'  {snum:>8s} | {ref:>15s} | {str(name)[:50]}')

# ── Show RTE ref distribution ──
print(f'\nUnique PTS refs in RTE: {len(pts_to_rte)}')
# Show sample
for pts_ref in list(pts_to_rte.keys())[:10]:
    print(f'  {pts_ref} → RTE pages {pts_to_rte[pts_ref][:5]}')

# Save
with open('rte_pts_refs.json', 'w') as f:
    json.dump({k: v for k, v in pts_to_rte.items()}, f, indent=2)
print(f'\nSaved rte_pts_refs.json')
