#!/usr/bin/env python3
"""
Helmer Smith — KN content validation: 30 strategic points across all KN books.
Verifies page content actually matches the expected sutta/section.
"""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def norm(s):
    for a,b in [('ā','a'),('ī','i'),('ū','u'),('ṅ','n'),('ñ','n'),
                ('ṭ','t'),('ḍ','d'),('ṇ','n'),('ḷ','l'),('ṃ','m'),('ṁ','m')]:
        s = s.replace(a,b).replace(a.upper(),b.upper())
    return s

def get_page(book_no, page_no):
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None, None
    return r['head'] or '', r['unitext'] or ''

def get_book_no(vol_str):
    m = {
        'Khp':22,'Kh':22,'Dhp':23,'Dh':23,'Ud':24,'It':25,'Sn':26,
        'Vv':27,'Pv':28,'Th':29,'Th & Th':29,'Thi':29,'Thī':29,
        'Ja':30,'Ja I':30,'Ja II':31,'Ja III':32,'Ja IV':33,'Ja V':34,'Ja VI':35,
        'Nidd':36,'Nidd I':36,'Nidd II':37,
        'Patis I':38,'Patis II':39,'Paṭis I':38,'Paṭis II':39,
        'Ap':40,'Bv':41,'Cp':42,
    }
    if vol_str in m: return m[vol_str]
    for k,v in m.items():
        if vol_str.startswith(k) or k.startswith(vol_str): return v
    return None

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

# Collect KN entries
kn_entries = []
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'KN': continue
    snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    if not snum or snum == 'None': continue
    kn_entries.append({
        'ri': ri, 'num': snum,
        'name': str(ws.cell(row=ri, column=cols['Sutta Name']).value or ''),
        'vol': str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip(),
        'page': ws.cell(row=ri, column=cols['PTS Page']).value,
        'ref': str(ws.cell(row=ri, column=cols['PTS Ref']).value or ''),
        'val': str(ws.cell(row=ri, column=cols['Validation']).value or ''),
    })

# Group by volume
by_vol = defaultdict(list)
for e in kn_entries:
    by_vol[e['vol']].append(e)

# Select 30 strategic samples: first + last of each book, plus mid-points
import random
random.seed(42)

samples = []
for vol, entries in sorted(by_vol.items()):
    if vol in ('Nett', 'Pet'): continue
    if len(entries) == 0: continue
    entries_sorted = sorted(entries, key=lambda e: (e['page'] or 0, e['num']))
    
    # First, last, and 1-2 mid-points
    to_sample = [entries_sorted[0], entries_sorted[-1]]
    if len(entries) > 4:
        to_sample.append(entries_sorted[len(entries)//3])
    if len(entries) > 10:
        to_sample.append(entries_sorted[2*len(entries)//3])
    
    for e in to_sample:
        if len(samples) < 35:
            samples.append(e)

# Also add some UNVERIFIED entries
unverified = [e for e in kn_entries if e['val'] in ('UNVERIFIED', 'RTE_ONLY', '')]
random.shuffle(unverified)
for e in unverified[:5]:
    if e not in samples:
        samples.append(e)

print('Helmer Smith — KN Content Validation')
print('=' * 90)
print(f'Sample: {len(samples)} entries across {len(by_vol)} books')
print()

results = []
stats = defaultdict(int)

for e in samples:
    book_no = get_book_no(e['vol'])
    if not book_no:
        results.append((e, 'NO_BOOK', f'Cannot map vol="{e["vol"]}"'))
        stats['NO_BOOK'] += 1
        continue
    
    head, text = get_page(book_no, e['page'])
    if text is None:
        results.append((e, 'PAGE_MISSING', f'Page {e["page"]} not in book {book_no}'))
        stats['PAGE_MISSING'] += 1
        continue
    
    lines = text.split('\n')
    name_kw = [w for w in re.findall(r'[a-z]{4,}', norm(e['name'].lower()))
               if w not in ('sutta','suttam','vagga','pathama','dutiya','tatiya',
                           'catuttha','pancama','chattha','sattama','atthama','navama','dasama')]
    
    # Level 1: HEAD match
    head_match = False
    head_detail = ''
    if name_kw:
        head_hits = sum(1 for w in name_kw if w in norm(head).lower())
        if head_hits >= 1:
            head_match = True
            head_detail = f'HEAD: {head_hits}/{len(name_kw)} kw'
    
    # Level 2: Body keyword match (first 600 chars)
    body = norm(text[:600]).lower()
    body_hits = sum(1 for w in name_kw if w in body) if name_kw else 0
    
    # Level 3: Structural markers on page
    markers = []
    for i, line in enumerate(lines):
        s = line.strip()
        if re.match(r'^\d+\.\s+\S', s):
            markers.append(f'L{i+1}:num')
        elif re.search(r'[Ee]va[mM].*suta[mM]', s):
            markers.append(f'L{i+1}:evam')
        elif re.search(r'[║]\s*\d+\s*[║]', s):
            markers.append(f'L{i+1}:verse_end')
    
    # Level 4: Name match in first 8 lines
    first8 = ' '.join(lines[:8])
    name_in_first8 = False
    if name_kw:
        name_in_first8 = sum(1 for w in name_kw if w in norm(first8).lower()) >= 1
    
    # Determine status
    status = 'UNVERIFIED'
    detail = ''
    
    if head_match:
        status = 'HEAD'
        detail = head_detail
    elif body_hits >= max(1, len(name_kw)*0.5):
        status = 'BODY_KW'
        detail = f'{body_hits}/{len(name_kw)} kw in body'
    elif name_in_first8:
        status = 'FIRST8'
        detail = f'name in first 8 lines'
    elif markers:
        status = 'MARKERS'
        detail = f'{len(markers)} markers: {markers[:3]}'
    
    # Check nearby pages if unverified
    if status == 'UNVERIFIED':
        for delta in [-2, -1, 1, 2]:
            nh, nt = get_page(book_no, e['page'] + delta)
            if nt is None: continue
            nbody = norm(nt[:600]).lower()
            nhits = sum(1 for w in name_kw if w in nbody) if name_kw else 0
            if nhits >= 1:
                status = f'NEAR_{delta:+d}'
                detail = f'Found at p.{e["page"]+delta}, {nhits}/{len(name_kw)} kw'
                break
    
    stats[status] += 1
    results.append((e, status, detail))

# Print results
print(f'{"Vol":>12s} {"#":12s} {"Ref":>18s} {"Status":12s} {"Detail"}')
print('-' * 90)

for e, status, detail in results:
    sym = '✓' if status in ('HEAD','BODY_KW','FIRST8','MARKERS') else '≈' if status.startswith('NEAR') else '✗'
    print(f'{sym} {e["vol"]:>11s} {e["num"]:12s} {e["ref"]:>18s} {status:12s} {detail[:55]}')

print(f'\n{"="*90}')
verified = sum(1 for _, s, _ in results if s in ('HEAD','BODY_KW','FIRST8','MARKERS'))
nearby = sum(1 for _, s, _ in results if s.startswith('NEAR'))
unv = sum(1 for _, s, _ in results if s == 'UNVERIFIED')
other = sum(1 for _, s, _ in results if s in ('NO_BOOK','PAGE_MISSING'))

print(f'  ✓ Verified:   {verified}/{len(results)} ({100*verified/len(results):.0f}%)')
print(f'  ≈ Nearby:     {nearby}')
print(f'  ✗ Unverified: {unv}')
print(f'  ∅ Other:      {other}')

# Show unverified details
unv_entries = [(e, d) for e, s, d in results if s == 'UNVERIFIED']
if unv_entries:
    print(f'\n  Unverified entries (check manually):')
    for e, detail in unv_entries:
        book_no = get_book_no(e['vol'])
        head, text = get_page(book_no, e['page'])
        if text:
            first_line = text.split('\n')[0].strip()[:80]
            print(f'    {e["num"]:>12s} | {e["ref"]:>18s} | {e["name"][:40]}')
            print(f'      Page {e["page"]} first line: {first_line}')

# Assessment
if unv <= 2:
    print(f'\n  Helmer Smith verdict: SOUND. KN references confirmed against content.')
elif unv <= len(results) * 0.15:
    print(f'\n  Helmer Smith verdict: MOSTLY RELIABLE. {unv} entries need manual inspection.')
else:
    print(f'\n  Helmer Smith verdict: NEEDS REVIEW. {unv} entries could not be content-verified.')

conn.close()
