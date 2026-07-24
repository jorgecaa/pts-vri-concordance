#!/usr/bin/env python3
"""Genera tabla Excel completa: todos los nikayas desde tipitaka.sqlite."""
import sqlite3, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DB = 'src/data/tipitaka.sqlite'
OUT = 'PTS_Reference_Complete.xlsx'

con = sqlite3.connect(DB)

# ── Mapa book_no → PTS abbrev ──
BOOK_ABBR = {
    1:"Vin I",2:"Vin II",3:"Vin III",4:"Vin IV",5:"Vin V",
    6:"DN I",7:"DN II",8:"DN III",
    9:"MN I",10:"MN II",11:"MN III",
    12:"SN I",13:"SN II",14:"SN III",15:"SN IV",16:"SN V",
    17:"AN I",18:"AN II",19:"AN III",20:"AN IV",21:"AN V",
    22:"Khp",23:"Dhp",24:"Ud",25:"It",26:"Sn",
    27:"Vv",28:"Pv",29:"Th & Th",30:"Th & Th",
    31:"Ja II",32:"Ja III",33:"Ja IV",34:"Ja V",35:"Ja VI",
    36:"Nidd I",37:"Nidd II",38:"Paṭis I",39:"Paṭis II",
    40:"Ap",41:"Bv",42:"Cp",
    43:"Dhs",44:"Vibh",45:"Dhātuk",46:"Pp",47:"Kv",
    48:"Yam I",49:"Yam II",50:"Paṭṭh I",51:"Paṭṭh II",52:"Paṭṭh II",53:"Paṭṭh III",
}

PITAKA = {}
for b in range(1,6): PITAKA[b] = "Vinaya"
for b in range(6,9): PITAKA[b] = "Dīgha Nikāya"
for b in range(9,12): PITAKA[b] = "Majjhima Nikāya"
for b in range(12,17): PITAKA[b] = "Saṃyutta Nikāya"
for b in range(17,22): PITAKA[b] = "Aṅguttara Nikāya"
for b in range(22,43): PITAKA[b] = "Khuddaka Nikāya"
for b in range(43,54): PITAKA[b] = "Abhidhamma"

# ── Extraer contenidos ──
rows = []
for book_no, seq, page_no, section, title in con.execute(
    "SELECT book_no, seq, page_no, section, title FROM contents ORDER BY book_no, seq"
):
    abbr = BOOK_ABBR.get(book_no, f"Book {book_no}")
    pitaka = PITAKA.get(book_no, "")
    # El título está en Thai corrupto — intentar extraer lo que se pueda
    title_clean = str(title or '').strip()
    section_clean = str(section or '').strip()
    rows.append((book_no, abbr, pitaka, page_no, section_clean, title_clean))

con.close()
print(f"Entradas extraídas: {len(rows)}")

# ── Excel ──
wb = Workbook()
ws = wb.active
ws.title = "PTS Reference"

hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(bold=True, size=11, color="FFFFFF")
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

headers = ["Book #", "PTS Vol", "Piṭaka", "Page", "Section", "Title (raw)"]
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin

current_book = None
for ri, (book_no, abbr, pitaka, page_no, section, title) in enumerate(rows, 2):
    is_new_book = (book_no != current_book)
    current_book = book_no
    
    for ci, val in enumerate([book_no, abbr, pitaka, page_no, section, title], 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = thin
        if is_new_book:
            c.fill = section_fill
            c.font = Font(bold=True)

ws.column_dimensions["A"].width = 9
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 8
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 40
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{len(rows)+1}"

wb.save(OUT)
print(f"Guardado: {OUT}")
