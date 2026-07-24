#!/usr/bin/env python3
"""Re-audit ALL 152 MN suttas using centered number marker. Correct the Excel."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

MN_BOOKS = {1:9, 2:10, 3:11}
ROMAN = {'i':1,'ii':2,'iii':3}

ORIGINAL = [
    (1,'i',1),(2,'i',6),(3,'i',12),(4,'i',16),(5,'i',24),(6,'i',33),(7,'i',36),
    (8,'i',40),(9,'i',46),(10,'i',55),(11,'i',63),(12,'i',68),(13,'i',83),
    (14,'i',91),(15,'i',95),(16,'i',101),(17,'i',104),(18,'i',108),(19,'i',114),
    (20,'i',118),(21,'i',122),(22,'i',130),(23,'i',142),(24,'i',145),(25,'i',151),
    (26,'i',160),(27,'i',175),(28,'i',184),(29,'i',192),(30,'i',198),(31,'i',205),
    (32,'i',212),(33,'i',220),(34,'i',225),(35,'i',227),(36,'i',237),(37,'i',251),
    (38,'i',256),(39,'i',271),(40,'i',281),(41,'i',285),(42,'i',290),(43,'i',292),
    (44,'i',299),(45,'i',305),(46,'i',309),(47,'i',317),(48,'i',320),(49,'i',326),
    (50,'i',332),(51,'i',339),(52,'i',349),(53,'i',353),(54,'i',359),(55,'i',368),
    (56,'i',371),(57,'i',387),(58,'i',392),(59,'i',396),(60,'i',400),(61,'i',414),
    (62,'i',420),(63,'i',426),(64,'i',432),(65,'i',437),(66,'i',447),(67,'i',456),
    (68,'i',462),(69,'i',469),(70,'i',473),(71,'i',481),(72,'i',483),(73,'i',489),
    (74,'i',497),(75,'i',501),(76,'i',513),
    (77,'ii',1),(78,'ii',22),(79,'ii',29),(80,'ii',40),(81,'ii',45),(82,'ii',54),
    (83,'ii',74),(84,'ii',83),(85,'ii',91),(86,'ii',97),(87,'ii',106),(88,'ii',112),
    (89,'ii',118),(90,'ii',125),(91,'ii',133),(92,'ii',146),(93,'ii',147),
    (94,'ii',157),(95,'ii',164),(96,'ii',177),(97,'ii',184),(98,'ii',196),
    (99,'ii',196),(100,'ii',209),(101,'ii',214),(102,'ii',228),(103,'ii',238),
    (104,'ii',243),(105,'ii',252),(106,'ii',261),
    (107,'iii',1),(108,'iii',7),(109,'iii',15),(110,'iii',20),(111,'iii',25),
    (112,'iii',29),(113,'iii',37),(114,'iii',45),(115,'iii',61),(116,'iii',68),
    (117,'iii',71),(118,'iii',78),(119,'iii',88),(120,'iii',99),(121,'iii',104),
    (122,'iii',109),(123,'iii',118),(124,'iii',124),(125,'iii',128),(126,'iii',138),
    (127,'iii',144),(128,'iii',152),(129,'iii',163),(130,'iii',178),(131,'iii',187),
    (132,'iii',189),(133,'iii',192),(134,'iii',199),(135,'iii',202),(136,'iii',207),
    (137,'iii',215),(138,'iii',223),(139,'iii',230),(140,'iii',237),(141,'iii',248),
    (142,'iii',253),(143,'iii',258),(144,'iii',263),(145,'iii',267),(146,'iii',270),
    (147,'iii',277),(148,'iii',280),(149,'iii',287),(150,'iii',290),(151,'iii',293),
    (152,'iii',298),
]

def find_marker(book_no, sutta_num):
    """Find page+line where centered number marker appears."""
    ns = str(sutta_num)
    for page in range(1, 600):
        cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page))
        r = cur.fetchone()
        if not r: continue
        head = r['head'] or ''
        lines = (r['unitext'] or '').split('\n')
        
        # 1. HEAD has sutta number: "(14)"
        if f'({ns})' in head or f'({ns}.)' in head:
            for i, line in enumerate(lines):
                if line.strip() and len(line.strip()) > 3:
                    return page, i+1, 'head'
        
        # 2. Body: centered "82." alone
        for i, line in enumerate(lines):
            s = line.strip()
            if s == f'{ns}.' or s == ns:
                return page, i+1, 'marker'
        
        # 3. Body: "82. Evam me..."
        for i, line in enumerate(lines):
            s = line.strip()
            if re.match(rf'^{re.escape(ns)}\.\s+\S', s):
                return page, i+1, 'marker+txt'
    return None, None, 'not_found'

print('MN Re-audit: centered number markers')
print('=' * 60)

results = []
page_ok = page_fix = not_found = line_count = 0

for num, roman, stated in ORIGINAL:
    book = MN_BOOKS[ROMAN[roman]]
    actual_pg, actual_ln, method = find_marker(book, num)
    
    if actual_pg is None:
        print(f'  XX MN {num:>3d} — NOT FOUND')
        not_found += 1
        continue
    
    delta = actual_pg - stated
    if delta == 0: page_ok += 1
    else: page_fix += 1
    if actual_ln > 1: line_count += 1
    
    ref = f'M {roman} {actual_pg}'
    if actual_ln > 1: ref += f',{actual_ln}'
    
    marker = '✓' if delta == 0 else f'Δ={delta:+d}'
    ln_str = f'L{actual_ln}' if actual_ln > 1 else ''
    print(f'  {marker:5s} MN {num:>3d}  {ref:16s} {ln_str:5s}  ({method})')
    
    results.append({'num': num, 'roman': roman, 'page': actual_pg, 'line': actual_ln, 'ref': ref})

print(f'\n{"="*60}')
print(f'  Blog correct:  {page_ok}')
print(f'  Blog wrong:    {page_fix}')
print(f'  Not found:     {not_found}')
print(f'  With line #:   {line_count}')
print(f'  Total:         {len(results)}/152')

# Apply to Excel
wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=1, column=c).value
    if h: cols[h] = c

updates = 0
for res in results:
    for ri in range(2, ws.max_row+1):
        if ws.cell(row=ri, column=cols['Nikaya']).value != 'MN': continue
        if str(ws.cell(row=ri, column=cols['Sutta #']).value or '') != str(res['num']): continue
        
        old_pg = ws.cell(row=ri, column=cols['PTS Page']).value
        old_ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
        
        if old_pg != res['page'] or old_ref != res['ref']:
            ws.cell(row=ri, column=cols['PTS Page']).value = res['page']
            ws.cell(row=ri, column=cols['PTS Ref']).value = res['ref']
            updates += 1
        break

wb.save(XL)
print(f'  Excel updates: {updates}')
print(f'  Saved: {XL}')
conn.close()
