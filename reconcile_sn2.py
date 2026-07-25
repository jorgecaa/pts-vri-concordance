#!/usr/bin/env python3
"""reconcile_sn2.py — vuelca `validador_sn2.json` al Excel maestro (SN II, S ii).

Regla del proyecto: CONFIRMADO = concordancia (VRI) ∧ Gemini APPROVE → `Validation=VALIDADOR`.
Gemini REJECT → NO se auto-firma: queda para arbitraje humano (`--humano firmas.json`).

En SN II el emparejamiento es por **clave `(saṃyutta, nº corrido)`** — la notación que Feer mismo
propone («XII. 25. 4») y la que usa el `Sutta #` del Excel — con la estructura del volumen cotejada
contra su front matter (10 saṃyuttas, 27 vaggas, 286 suttas) y `cst_p_page` coincidiendo en las 248.
Un REJECT sería, por tanto, un problema de cotejo de texto, no de alineación.

NOTA: el Excel solo cubre 248 de los 286 suttas de S ii; las 38 filas que faltan (SN 12.76–93,
17.32–43, 18.15–22) no las crea este script.

Uso:
    python3 reconcile_sn2.py --dry
    python3 reconcile_sn2.py --write [--humano firmas.json]
"""
import json, re, sys, shutil, datetime
from collections import Counter
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
RES = 'validador_sn2.json'


def sn2_rows(ws, ci):
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'SN' or str(v[ci['PTS Roman']] or '').strip().lower() != 'ii':
            continue
        yield ridx, str(v[ci['Sutta #']]), v


def main():
    dry = '--write' not in sys.argv
    humano = {}
    if '--humano' in sys.argv:
        humano = json.load(open(sys.argv[sys.argv.index('--humano') + 1]))
    res = {r['num']: r for r in json.load(open(RES))}
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}
    vcol, ecol = ci['Validation'] + 1, ci['Estado'] + 1

    plan, writes, pend = Counter(), [], []
    for ridx, num, v in sn2_rows(ws, ci):
        if num in humano:
            val = humano[num]
            est = 'PENDIENTE' if val in ('REVISAR', 'REF_ERROR_DPR') else 'CONFIRMADO'
            plan[f'humano:{val}'] += 1
            writes.append((ridx, val, est))
            continue
        cur_val = str(ws.cell(row=ridx, column=vcol).value or '')
        if cur_val in ('VALIDADOR_HUMANO', 'PTS_CROSSREF_SN'):
            plan['protegido(humano/crossref)'] += 1
            continue
        r = res.get(num)
        if not r:
            plan['sin_resultado'] += 1
            continue
        if r['gemini'] == 'APPROVE':
            plan['VALIDADOR(auto)'] += 1
            writes.append((ridx, 'VALIDADOR', 'CONFIRMADO'))
        else:
            plan['arbitraje(gemini REJECT)'] += 1
            pend.append(r)

    print('Plan de escritura:', dict(plan))
    print('Filas a escribir:', len(writes))
    if pend:
        print(f'\nA arbitraje ({len(pend)}) — alineación segura, discrepa el cotejo de texto:')
        for r in pend:
            print(f'  SN {r["num"]:>7} PTS "{r.get("pts_name","")[:20]:20}" vs CST '
                  f'"{r["cst_title"][:26]:26}" p{r.get("pts_page")} | {str(r["reason"])[:64]}')
    if dry:
        print('\n(dry-run, nada escrito)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-sn2.xlsx'
    shutil.copy(XLSX, bak)
    print('backup →', bak)
    for ridx, val, est in writes:
        ws.cell(row=ridx, column=vcol).value = val
        ws.cell(row=ridx, column=ecol).value = est
    wb.save(XLSX)
    print(f'Escritas {len(writes)} filas en {XLSX}.')


if __name__ == '__main__':
    main()
