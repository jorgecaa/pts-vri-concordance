#!/usr/bin/env python3
"""SN1 pilot v2: sutta-bounded text extraction using § markers."""
import hashlib, json, os, re, sqlite3
from pathlib import Path
import sutta_hash as sh

CACHE_FILE = Path('helmer_ptscst_cache.json')
SN1_BOOK = 12

def pts_sutta_text(page_no, sutta_num, max_tokens=200):
    """Extract PTS text for a specific sutta on a page. Uses § markers for SN I."""
    conn=sqlite3.connect('src/data/tipitaka.sqlite'); conn.row_factory=sqlite3.Row; cur=conn.cursor()
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN1_BOOK,page_no))
    r=cur.fetchone(); conn.close()
    if not r or not r['unitext']: return ''
    lines=r['unitext'].split('\n')
    
    # Find the sutta marker: § N. or N. Name
    start_idx = None
    sn = str(sutta_num)
    for i, line in enumerate(lines):
        s = line.strip()
        # § N. Name
        if re.match(r'§\s*' + re.escape(sn) + r'\b', s):
            start_idx = i; break
        # N. Name (without §)
        if re.match(r'^' + re.escape(sn) + r'\.\s+\S', s):
            start_idx = i; break
    
    if start_idx is None:
        # Fallback: start from the line number in the ref
        return ''
    
    # Extract from start_idx until next § or end of page
    end_idx = len(lines)
    for j in range(start_idx + 1, len(lines)):
        s = lines[j].strip()
        if re.match(r'§\s*\d+', s):
            end_idx = j; break
        # Also stop at next sutta number marker
        if re.match(r'^\d+\.\s+\S', s) and j > start_idx + 2:
            end_idx = j; break
    
    sutta_lines = lines[start_idx:end_idx]
    text = ' '.join(sutta_lines)
    tokens = sh.tokens(text)
    return ' '.join(tokens[:max_tokens])

# Test extraction
print('Testing sutta-bounded extraction:')
for snum, page in [('4',3),('5',3),('6',3)]:
    txt = pts_sutta_text(page, snum, 200)
    print('SN 1.{} | page {}: {:.80s}...'.format(snum, page, txt))
print()

# Now run DeepSeek on first 10
from openpyxl import load_workbook
wb=load_workbook('PTS_Reference_Complete_Canon.xlsx')
ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}

tests=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2 or int(parts[0])!=1: continue  # SN 1 only
    if int(parts[1])>10: continue  # first 10
    page=ws.cell(row=ri,column=cols['PTS Page']).value
    ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
    name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
    if page: tests.append((snum,name,page,parts[1],ref))

segs=sh.load_cha(['s1m.xml'],'h4n')
cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}

from openai import OpenAI
cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')

SYS='You are Helmer Smith. Are these two Pali passages the SAME sutta? SN suttas are short verses. Same=core verse/dialogue matches. Different=distinct content. Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words English"}'

def ask(p,c,cli):
    r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
        response_format={'type':'json_object'},
        messages=[{'role':'system','content':SYS},
                  {'role':'user','content':'PTS:\n{}\n\nCST:\n{}'.format(p[:800],c[:800])}])
    return json.loads(r.choices[0].message.content)

print('DeepSeek SN1 pilot (sutta-bounded):')
print('='*95)
from collections import Counter
vc=Counter()

for snum,name,page,sn_inner,ref in tests:
    pts=pts_sutta_text(page,sn_inner,200)
    if not pts:
        print('? SN {} | {} | PTS extraction failed'.format(snum,ref)); continue
    
    cst_idx=int(sn_inner)-1
    cst_text=' '.join(sh.tokens(segs[cst_idx]['text'])[:200])
    cst_title=segs[cst_idx].get('title','')
    
    print('SN {:>5s} {:>15s} | PTS: {:.60s}...'.format(snum,ref,pts[:60]))
    
    key=hashlib.md5((pts[:300]+cst_text[:300]).encode()).hexdigest()
    if key in cache and cache[key].get('verdict')=='APPROVE':
        hv=cache[key]; vc['CACHED']+=1
        print('  -> CACHED APPROVE: {}'.format(hv.get('reason','')))
    else:
        try: hv=ask(pts,cst_text,cli); cache[key]=hv; CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
        except Exception as e: hv={'verdict':'ERROR','reason':str(e)[:60]}
        v=hv.get('verdict','?'); vc[v]+=1
        sym='✓' if v=='APPROVE' else '✗' if v=='REJECT' else '?'
        print('  -> {} {}: {}'.format(sym,v,hv.get('reason','')))
    print()

print('='*95)
print('APPROVE: {} | REJECT: {} | ERROR: {} | CACHED: {}'.format(
    vc.get('APPROVE',0),vc.get('REJECT',0),vc.get('ERROR',0),vc.get('CACHED',0)))
