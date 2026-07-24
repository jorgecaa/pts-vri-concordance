#!/usr/bin/env python3
"""
Fix MN page references in PTS_Reference_Complete_Canon.xlsx.
Pattern: blog alternates between vagga-header page (even) and sutta-header page (odd).
"""
import sqlite3
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'
OUT = 'PTS_Reference_Complete_Canon.xlsx'  # overwrite

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

# Corrections: (nikaya, sutta_num, old_page, new_page, reason)
CORRECTIONS = [
    # MN I — off-by-one (sutta header is on odd page, blog gave even vagga page)
    ('MN', '1', 1, 3, 'Title page vs sutta header'),
    ('MN', '2', 6, 7, 'Vagga page vs sutta header'),
    ('MN', '3', 12, 13, 'Vagga page vs sutta header'),
    ('MN', '4', 16, 17, 'Vagga page vs sutta header'),
    ('MN', '5', 24, 25, 'Vagga page vs sutta header'),
    ('MN', '7', 36, 37, 'Vagga page vs sutta header'),
    ('MN', '8', 40, 41, 'Vagga page vs sutta header'),
    ('MN', '9', 46, 47, 'Vagga page vs sutta header'),
    ('MN', '12', 68, 69, 'Vagga page vs sutta header'),
    ('MN', '17', 104, 105, 'Vagga page vs sutta header'),
    ('MN', '18', 108, 109, 'Vagga page vs sutta header'),
    ('MN', '19', 114, 115, 'Vagga page vs sutta header'),
    ('MN', '20', 118, 119, 'Vagga page vs sutta header'),
    ('MN', '21', 122, 123, 'Vagga page vs sutta header'),
    ('MN', '22', 130, 131, 'Vagga page vs sutta header'),
    ('MN', '23', 142, 143, 'Vagga page vs sutta header'),
    ('MN', '26', 160, 161, 'Vagga page vs sutta header'),
    ('MN', '28', 184, 185, 'Vagga page vs sutta header'),
    ('MN', '29', 192, 193, 'Vagga page vs sutta header'),
    ('MN', '30', 198, 199, 'Vagga page vs sutta header'),
    ('MN', '32', 212, 213, 'Vagga page vs sutta header'),
    ('MN', '33', 220, 221, 'Vagga page vs sutta header'),
    ('MN', '38', 256, 257, 'Vagga page vs sutta header'),
    ('MN', '42', 290, 291, 'Vagga page vs sutta header'),
    ('MN', '43', 292, 293, 'Vagga page vs sutta header'),
    ('MN', '48', 320, 321, 'Vagga page vs sutta header'),
    ('MN', '49', 326, 327, 'Vagga page vs sutta header'),
    ('MN', '50', 332, 333, 'Vagga page vs sutta header'),
    ('MN', '55', 368, 369, 'Vagga page vs sutta header'),
    ('MN', '58', 392, 393, 'Vagga page vs sutta header'),
    ('MN', '59', 396, 397, 'Vagga page vs sutta header'),
    ('MN', '60', 400, 401, 'Vagga page vs sutta header'),
    ('MN', '61', 414, 415, 'Vagga page vs sutta header'),
    ('MN', '62', 420, 421, 'Vagga page vs sutta header'),
    ('MN', '63', 426, 427, 'Vagga page vs sutta header'),
    ('MN', '64', 432, 433, 'Vagga page vs sutta header'),
    ('MN', '67', 456, 457, 'Vagga page vs sutta header'),
    ('MN', '68', 462, 463, 'Vagga page vs sutta header'),
    
    # MN II — off-by-one
    ('MN', '77', 1, 3, 'Title page vs sutta header (MN II)'),
    ('MN', '78', 22, 23, 'Vagga page vs sutta header'),
    ('MN', '80', 40, 41, 'Vagga page vs sutta header'),
    ('MN', '83', 74, 75, 'Vagga page vs sutta header'),
    ('MN', '87', 106, 107, 'Vagga page vs sutta header'),
    ('MN', '88', 112, 113, 'Vagga page vs sutta header'),
    ('MN', '89', 118, 119, 'Vagga page vs sutta header'),
    ('MN', '95', 164, 165, 'Vagga page vs sutta header'),
    ('MN', '97', 184, 185, 'Vagga page vs sutta header'),
    ('MN', '99', 196, 197, 'Vagga page vs sutta header'),
    ('MN', '101', 214, 215, 'Vagga page vs sutta header'),
    ('MN', '102', 228, 229, 'Vagga page vs sutta header'),
    ('MN', '103', 238, 239, 'Vagga page vs sutta header'),
    ('MN', '105', 252, 253, 'Vagga page vs sutta header'),
    
    # MN III — off-by-one (different pattern!)
    ('MN', '107', 1, 3, 'Title page vs sutta header (MN III)'),
    ('MN', '110', 20, 21, 'Vagga page vs sutta header'),
    ('MN', '116', 68, 69, 'Vagga page vs sutta header'),
    ('MN', '118', 78, 79, 'Vagga page vs sutta header'),
    ('MN', '119', 88, 89, 'Vagga page vs sutta header'),
    ('MN', '121', 104, 105, 'Vagga page vs sutta header'),
    ('MN', '123', 118, 119, 'Vagga page vs sutta header'),
    ('MN', '124', 124, 125, 'Vagga page vs sutta header'),
    ('MN', '125', 128, 129, 'Vagga page vs sutta header'),
    ('MN', '126', 138, 139, 'Vagga page vs sutta header'),
    ('MN', '127', 144, 145, 'Vagga page vs sutta header'),
    ('MN', '128', 152, 153, 'Vagga page vs sutta header'),
    ('MN', '130', 178, 179, 'Vagga page vs sutta header'),
    ('MN', '133', 192, 193, 'Vagga page vs sutta header'),
    ('MN', '135', 202, 203, 'Vagga page vs sutta header'),
    ('MN', '139', 230, 229, 'Vagga page vs sutta header (reverse!)'),
    ('MN', '140', 237, 239, 'Vagga page vs sutta header (Δ=2)'),
    ('MN', '141', 248, 249, 'Vagga page vs sutta header'),
    ('MN', '143', 258, 259, 'Vagga page vs sutta header'),
    ('MN', '144', 263, 265, 'Vagga page vs sutta header (Δ=2)'),
    ('MN', '146', 270, 271, 'Vagga page vs sutta header'),
    ('MN', '148', 280, 281, 'Vagga page vs sutta header'),
    ('MN', '150', 290, 291, 'Vagga page vs sutta header'),
    ('MN', '152', 298, 299, 'Vagga page vs sutta header'),
    
    # MN II — unverified, manually resolved
    ('MN', '92', 146, 147, 'Sela: on same page as Assalayana (MN 93) — sutta header at p.147'),
    ('MN', '98', 196, 196, 'Vāseṭṭha: shares p.196 with Subha (MN 99), sutta starts mid-page — correct'),
]

# Apply corrections
wb = load_workbook(XL)
ws = wb['Complete Canon']

# Find column indices
col_map = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(row=1, column=c).value
    if h:
        col_map[h] = c

nikaya_col = col_map['Nikaya']
num_col = col_map['Sutta #']
page_col = col_map['PTS Page']
ref_col = col_map['PTS Ref']
roman_col = col_map['PTS Roman']

applied = 0
for nikaya, sutta_num, old_page, new_page, reason in CORRECTIONS:
    for row in range(2, ws.max_row + 1):
        n = ws.cell(row=row, column=nikaya_col).value
        sn = str(ws.cell(row=row, column=num_col).value or '')
        pg = ws.cell(row=row, column=page_col).value
        
        if n == nikaya and sn == sutta_num and pg == old_page:
            # Update page
            ws.cell(row=row, column=page_col).value = new_page
            
            # Update PTS Full Ref (e.g., "M i 6" → "M i 7")
            old_ref = str(ws.cell(row=row, column=ref_col).value or '')
            new_ref = old_ref.replace(str(old_page), str(new_page))
            ws.cell(row=row, column=ref_col).value = new_ref
            
            applied += 1
            break

wb.save(OUT)
print(f'Applied {applied} corrections to {OUT}')
print(f'Total corrections: {len(CORRECTIONS)}')

# Summary
print(f'\nMN I: 39 corrections (1 title page + 38 vagga→sutta)')
print(f'MN II: 14 corrections')
print(f'MN III: 24 corrections')
print(f'Total: {len(CORRECTIONS)} corrections')
conn.close()
