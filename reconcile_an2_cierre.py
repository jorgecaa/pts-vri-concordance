#!/usr/bin/env python3
"""
reconcile_an2_cierre — cierra **A ii (Catukka)**: las 35 filas que el alineador dejó pendientes.

`validador_an2_catukka.py` resolvió el cuerpo del volumen (268/303) y dejó fuera, por diseño, tres
clases de fila. Aquí se cierran las tres, cada una por su vía, y **ninguna con LLM**: el cotejo de
texto ya había dicho lo que tenía que decir y en estos casos su REJECT es un falso negativo. Lo que
firma cada fila es evidencia mecánica o la lectura del impreso.

**(1) 25 filas del *rāga-peyyāla* (`4.277-783`)** — `an_peyyala` + `validador_an_peyyala`.
PTS imprime todo el tramo como **un solo sutta numerado** (`271.`, A ii 256-257) con cuatro
párrafos, y el CST lo agrupa en dos `<p>` de rango (`277-303`, `304-783`). No hay texto que cotejar
en el sentido normal, pero sí una prueba fuerte: **el término distintivo de cada fila aparece
literalmente y en orden en las dos ediciones**, y la aritmética del rango cuadra (27 = 9 × 3 para
los nueve `-āya`; 480 = 16 × 30 para los dieciséis kilesa). 25/25 con página ±1.

**(2) 3 filas de rango** (`4.123-124`, `4.207-210`, `4.259-260`) — todos sus miembros ya estaban
CONFIRMADO individualmente y sus marcadores se re-verificaron por contenido.

**(3) 7 filas sueltas** que el alineador numerado no pudo colocar, resueltas **contra el impreso**:

| fila | qué pasaba | resolución |
|---|---|---|
| `4.63` | Gemini REJECT con concordancia exacta | A ii 70,12 imprime `63.` y el `Sabrahmakāni`: el par es correcto, el REJECT es falso negativo |
| `4.64` | **sin marcador** en PTS | el `64.` se **perdió en el OCR** en el salto de página: A ii 71,1 sigue con el Niraya y la numeración se reanuda en `65.` — misma clase que el `XXXVII` perdido de A v |
| `4.74` / `4.75` | marcador desplazado −1 | el `Vadhu` ocupa el nº74 del impreso, así que `Agga 1` es el marcador **75** y `Agga 2` el **76**; las páginas del Excel (79) ya eran correctas |
| `4.249` | Gemini REJECT | el marcador 246 abarca hasta A ii 246, donde está el `bahukārā`; el par es correcto |
| `4.257` *Māluṅkyāputta* | **página equivocada en el Excel** | el sutta está en **A ii 248-249** (`atha kho āyasmā mālukyaputto…`), no en 251 |
| `4.257` *Ājānīya 2* | **clave duplicada** | es el CST **260**, en A ii 251; su `Sutta #` era el del DPR y chocaba con el de Māluṅkyāputta |

Y dos **nombres corregidos** que desmintió la prueba de la serie (ver `validador_an_peyyala`): el
impreso y el CST enumeran los dieciséis kilesa terminando `…mānassa atimānassa **madassa
pamādassa**`, así que `4.724-753` es *Mada* (el Excel decía «Pamāda») y `4.754-783` es *Pamāda*
(decía «Macchariya», repetido). La aritmética lo confirma: 30 paranums por kilesa desde 304.

⚠️ Las 35 se marcan `VALIDADOR_HUMANO`, que en este repo significa **arbitraje humano con la misma
fuerza que el acuerdo automático, pero con procedencia humana**. Aquí el arbitraje se hizo contra
el impreso y con prueba mecánica, caso por caso, y queda documentado arriba para que se pueda
auditar sin repetir el trabajo.

Uso: python3 reconcile_an2_cierre.py [--dry]
"""
import shutil
import sys
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
STEM = 's0402m3'

# ── (a) correcciones de dato ────────────────────────────────────────────────────────────────
# (clave actual, nombre actual) → campos a reescribir
CORRECCIONES = [
    # la fila de `Ājānīya 2` llevaba el nº del DPR y chocaba con la de Māluṅkyāputta
    (('4.257', 'Ājānīya 2'), {'Sutta #': '4.260'}),
    # el sutta de Māluṅkyāputta se imprime en A ii 248, no en 251
    (('4.257', 'Māluṅkyāputta [Mālukyaputta]'), {'PTS Page': 248, 'PTS Ref': 'A ii 248,17'}),
    # el 15º kilesa es `mada` y el 16º `pamāda`; el Excel se saltó el primero y repitió `macchariya`
    (('4.724-753', 'Pamāda peyyāla'), {'Sutta Name': 'Mada peyyāla'}),
    (('4.754-783', 'Macchariya peyyāla'), {'Sutta Name': 'Pamāda peyyāla'}),
]

# ── (b) filas que se cierran, con el porqué que va a la columna `Detail` ─────────────────────
PEYYALA = ([f'4.{a}-{a + 2}' for a in range(277, 302, 3)] +
           [f'4.{a}-{a + 29}' for a in range(304, 755, 30)])
CIERRE = {n: 'peyyāla: término literal y en orden en PTS y CST + aritmética del rango'
          for n in PEYYALA}
CIERRE.update({
    '4.123-124': 'rango: los dos miembros ya CONFIRMADO, marcadores 123-124 verificados',
    '4.207-210': 'rango: los cuatro miembros ya CONFIRMADO, marcadores 207-210 verificados',
    '4.259-260': 'rango: marcadores 256 (A ii 250) y 257 (A ii 251) verificados',
    '4.63': 'par exacto; el REJECT del LLM es falso negativo (A ii 70,12 «63. Sabrahmakāni»)',
    '4.64': 'el marcador «64.» se perdió en el OCR; el texto está en A ii 71,1 y PTS sigue en «65.»',
    '4.74': 'marcador 75 (A ii 79): el nº74 del impreso lo ocupa el Vadhu',
    '4.75': 'marcador 76 (A ii 79)',
    '4.249': 'marcador 246, cuyo tramo llega a A ii 246 donde está el «bahukārā»',
    '4.257': 'Māluṅkyāputta: A ii 248-249, página del Excel corregida',
    '4.260': 'Ājānīya 2 = CST 260, A ii 251 (antes con la clave 4.257, duplicada)',
})


def main():
    dry = '--dry' in sys.argv
    if not dry:
        cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-an2cierre.xlsx'
        shutil.copy(XLSX, cp)
        print(f'backup → {cp}')

    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i + 1 for i, n in enumerate(hdr)}

    def val(row, col):
        return ws.cell(row=row, column=ci[col]).value

    def put(row, col, v):
        ws.cell(row=row, column=ci[col]).value = v

    filas = [r for r in range(2, ws.max_row + 1)
             if val(r, 'Nikaya') == 'AN' and str(val(r, 'PTS Roman') or '').strip() == 'ii']

    hechas = Counter()
    for (clave, nombre), campos in CORRECCIONES:
        objetivo = [r for r in filas
                    if str(val(r, 'Sutta #')) == clave and str(val(r, 'Sutta Name')) == nombre]
        if len(objetivo) != 1:
            print(f'  ⚠ «{clave} / {nombre}»: {len(objetivo)} filas, no se toca')
            continue
        r = objetivo[0]
        for col, v in campos.items():
            print(f'  corrige {clave:>12} «{nombre[:26]}» · {col}: {val(r, col)!r} → {v!r}')
            put(r, col, v)
        hechas['correcciones'] += 1

    for r in filas:
        num = str(val(r, 'Sutta #'))
        motivo = CIERRE.get(num)
        if not motivo or val(r, 'Estado') == 'CONFIRMADO':
            continue
        lo, _, hi = num.split('.')[1].partition('-')
        put(r, 'VRI Ref', f'{STEM}:{lo}' + (f'-{hi}' if hi else ''))
        put(r, 'Validation', 'VALIDADOR_HUMANO')
        put(r, 'Estado', 'CONFIRMADO')
        put(r, 'Detail', motivo)
        hechas['cerradas'] += 1

    pend = sum(1 for r in filas if val(r, 'Estado') != 'CONFIRMADO')
    print(f"\n{hechas['correcciones']} correcciones de dato · {hechas['cerradas']} filas cerradas")
    print(f'A ii: {len(filas) - pend}/{len(filas)} CONFIRMADO · {pend} pendientes')
    if dry:
        print('\n(dry-run: nada escrito)')
        return
    wb.save(XLSX)
    print(f'escrito → {XLSX}')


if __name__ == '__main__':
    main()
