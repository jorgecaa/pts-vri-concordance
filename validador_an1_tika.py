#!/usr/bin/env python3
"""Aligner + validador del **Tika-nipāta de AN I** (A i 101-300, `book_no=17`).

Es el primero de los tres nipātas de AN I que se aborda, y por una razón concreta: **es el único
donde el CST trae nombre por sutta** (10 subheads por vagga) y donde las tres cuentas casi
coinciden — PTS 163 marcadores, Excel 156 filas individuales, CST 155 subheads. En Eka y Duka el
CST sólo da párrafos numerados bajo el vagga, así que allí no hay cotejo por nombre que sirva de
red y hace falta otra estrategia.

**Verdad-terreno del lado PTS** (`anguttara-vol-I-info.txt`, índice de Morris 1885): 16 vaggas con
su página de arranque. El contraste sale **16/16 exacto** — el marcador nº 10k+1 cae siempre en la
página que declara Morris — y los números corridos van **1..163 sin un solo hueco**. Es el lado PTS
más limpio del proyecto.

**Claves.** El lado CST se identifica con el **paranum del XML VRI** (`s0402m2:N`), que es la clave
canónica del proyecto (regla (5)); el `Sutta #` es sólo etiqueta. Ojo: **el `Sutta #` NO es el nº
corrido de PTS** — coincide en 103 de 156 y diverge donde el Excel intercala filas de rango o
desdobla, así que el emparejamiento va por **alineamiento monótono**, nunca por clave.

⚠️ **9 filas de rango** (`3.2-9 «Lakkhaṇa - Khata»`, `3.163-182 «Kammapathapeyyāla»`…) son
**redundantes** con filas individuales que ya existen. Quedan fuera de esta alineación y a
arbitraje, como en su día `12.74` (S ii) y `22.148` (S iii).

Uso: python3 validador_an1_tika.py [--dry] [--all] [--n N] [--only 3.5,3.9]
"""
import json
import os
import re
import sqlite3
import sys
import xml.etree.ElementTree as ET
from collections import Counter

import an_names
import sutta_hash as sh
from align_rows import assign as ar_assign
from an1_markers import collect_an1
from openpyxl import load_workbook
from validador import validate_pair

VRI = '/tmp/tipitaka-xml/romn/s0402m2.mul.xml'
VRI_STEM = 's0402m2'
OUT = 'validador_an1_tika.json'
DB = 'src/data/tipitaka.sqlite'
AN1_BOOK = 17
XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# índice de Morris: (vagga, nombre, página de arranque). Verdad-terreno estructural del lado PTS.
MORRIS = [(1, 'Bāla', 101), (2, 'Rathakāra', 106), (3, 'Puggala', 118), (4, 'Devadūta', 132),
          (5, 'Cūla', 150), (6, 'Brāhmaṇa', 155), (7, 'Mahā', 173), (8, 'Ānanda', 215),
          (9, 'Samaṇa', 229), (10, 'Loṇaphala', 239), (11, 'Sambodhi', 258), (12, 'Āpāyika', 265),
          (13, 'Kusināra', 274), (14, 'Yodhājīva', 284), (15, 'Maṅgala', 292), (16, 'Acelaka', 295)]


def build_cst_units(path=VRI):
    """Suttas del CST en orden de documento: `[{title, pn, vagga, text}]`.

    En AN el `<div>` es un **vagga** (no un sutta como en DN, ni un saṃyutta como en SN), así que
    el sutta es el `subhead` de dentro y sus paranums son los `bodytext` que le siguen.
    """
    root = ET.parse(path).getroot()
    out = []
    for d in root.iter('div'):
        h = d.find('head')
        if h is None or h.get('rend') != 'chapter':
            continue
        vagga = ' '.join(''.join(h.itertext()).split())
        cur = None
        for p in d.iter('p'):
            rend, txt = p.get('rend'), ''.join(p.itertext())
            if rend == 'subhead':
                cur = {'title': ' '.join(txt.split()), 'pn': [], 'text': '', 'vagga': vagga}
                out.append(cur)
            elif rend in ('bodytext', 'gatha1', 'gatha2', 'gatha3', 'gathalast') and cur is not None:
                m = re.match(r'(\d+)(?:-(\d+))?$', p.get('n') or '')
                if m:
                    cur['pn'] += list(range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1))
                cur['text'] += ' ' + txt
    return out


def excel_entries():
    """Filas del Tika, separando las **individuales** de las de **rango** (redundantes)."""
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    ind, rng = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['Nikaya']] or '') != 'AN' or str(row[ci['PTS Roman']] or '').strip().lower() != 'i':
            continue
        num = str(row[ci['Sutta #']])
        if not num.startswith('3.'):
            continue
        e = {'num': num, 'name': str(row[ci['Sutta Name']] or '').strip(),
             'page': row[ci['PTS Page']], 'legacy': str(row[ci['Validation']] or ''),
             'dpr': str(row[ci['DPR Ref']] or '').strip()}
        (rng if '-' in num else ind).append(e)
    ind.sort(key=lambda e: int(e['num'].split('.')[1]))
    return ind, rng


def pts_markers(cur):
    """Marcadores del Tika: nº corrido → `(página, línea, texto acotado)`."""
    pages = {r['page_no']: (r['unitext'] or '').split('\n')
             for r in cur.execute('SELECT page_no,unitext FROM pages '
                                  'WHERE edition="mula" AND book_no=?', (AN1_BOOK,))}
    marks = [x for x in collect_an1(pages) if x['nipata'] == 'TIKA' and x['num'] is not None]
    marks.sort(key=lambda x: (x['page'], x['line']))
    flat = [(pg, ln, t) for pg in sorted(pages) for ln, t in enumerate(pages[pg], 1)]
    pos = {(pg, ln): i for i, (pg, ln, _t) in enumerate(flat)}
    out = []
    for k, m in enumerate(marks):
        i0 = pos[(m['page'], m['line'])]
        i1 = pos[(marks[k + 1]['page'], marks[k + 1]['line'])] if k + 1 < len(marks) else len(flat)
        out.append({'num': m['num'], 'page': m['page'], 'line': m['line'], 'ord': k,
                    'vagga': (m['num'] - 1) // 10 + 1,
                    'text': ' '.join(sh.tokens(' '.join(t for _p, _l, t in flat[i0:i1]))[:350])})
    return out


def align(entries, units):
    """Alineamiento **monótono** filas↔suttas del CST (`align_rows`), por nombre y ordinal.

    No se empareja por clave: el `Sutta #` sólo coincide con el nº corrido de PTS en 103 de 156.
    """
    def score(i, j):
        e, u = entries[i], units[j]
        s = an_names.score(u['title'], e['name'])
        ok = an_names.ordinal_ok(u['title'], e['name'])
        if ok is True:
            s += 60                       # el ordinal desempata dentro de la serie
        elif ok is False:
            s -= 120                      # «Pāpaṇika 1» no puede ser «Dutiya…»
        return s if s > 0 else 0.1        # nunca None: la serie debe poder avanzar

    idx = ar_assign(len(entries), len(units), score, lambda j: 1, skip_penalty=200.0)
    return {entries[i]['num']: (units[j] if j is not None else None) for i, j in enumerate(idx)}


def pts_anchor(e):
    """Nº corrido de PTS que declara la fila: la **`DPR Ref` si existe**, si no el `Sutta #`.

    Es el ancla determinista del lado PTS, y sale del re-parseo de `Raw ID` (`an_rawid.py`): la
    `DPR Ref` describe exactamente los desdoblamientos y fusiones del Excel —`3.32 Ānanda` →
    `AN 3.32a` y `3.33 Sāriputta` → `AN 3.32b` (una del DPR partida en dos), `3.39 Sukhumāla` →
    `AN 3.38-9` (dos fundidas en una)—, que es justo lo que hace que haya **163 marcadores para
    156 filas**. Sin ella el alineamiento acumulaba deriva: `3.111 «Nidāna 1»` acababa en el
    marcador 110 cuando su `Raw ID` dice `[AN 3.107]-08`, o sea los marcadores **107-108**.
    """
    src = e.get('dpr') or ('AN ' + e['num'])
    m = re.search(r'(\d+)\.(\d+)', src)
    if not m:
        # el punto se perdió en la fuente: «AN 3102b» por «AN 3.102b»
        m = re.search(r'AN\s*(\d)(\d+)', src)
    return int(m.group(2)) if m else None


def align_pts(entries, marks, cst):
    """Alineamiento **monótono** filas↔marcadores de PTS, anclado en la PÁGINA del Excel.

    No vale la posición: hay **163 marcadores para 156 filas** —PTS numera por separado suttas que
    el Excel lleva agrupados— y emparejar la fila k con el marcador k+1 acumula deriva (desde
    `3.34` la página se va +2 y sigue creciendo). Sólo cuadraba en 104 de 156.
    """
    # Cuántas filas declaran cada nº corrido: los sufijos `a`/`b` de la `DPR Ref` marcan un sutta
    # de PTS que el Excel **parte en dos filas** (`3.32 Ānanda` → `AN 3.32a`, `3.33 Sāriputta` →
    # `AN 3.32b`), así que ese marcador tiene capacidad 2, no 1.
    anchors = Counter(a for a in (pts_anchor(x) for x in entries) if a is not None)
    reserved = set(anchors)

    def score(i, j):
        e, m = entries[i], marks[j]
        # El nº corrido que declara la fila es evidencia DIRECTA, no una preferencia: va como
        # restricción dura y además se le reserva el marcador, para que ninguna otra fila lo ocupe.
        a = pts_anchor(e)
        if a is not None:
            return 1000.0 if m['num'] == a else None
        if m['num'] in reserved:
            return None
        if not isinstance(e['page'], int):
            return 0.1
        d = abs(m['page'] - e['page'])
        if d > 3:
            return None
        s = 40 - 10 * d
        u = cst.get(e['num'])
        if u:                                   # el título CST refuerza, no decide
            s += an_names.score(u['title'], e['name']) / 10

        return max(s, 0.1)

    idx = ar_assign(len(entries), len(marks), score,
                    lambda j: max(1, anchors.get(marks[j]['num'], 1)), skip_penalty=500.0)
    return {entries[i]['num']: (marks[j] if j is not None else None) for i, j in enumerate(idx)}


def main():
    dry = '--dry' in sys.argv
    do_all = '--all' in sys.argv
    n = int(sys.argv[sys.argv.index('--n') + 1]) if '--n' in sys.argv else 30
    only = set(sys.argv[sys.argv.index('--only') + 1].split(',')) if '--only' in sys.argv else None

    units = build_cst_units()
    ind, rng = excel_entries()
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
    marks = pts_markers(conn.cursor())
    by_num = {m['num']: m for m in marks}
    cst = align(ind, units)

    print('=' * 92)
    print('AN I — TIKA-NIPĀTA: pipeline VRI + índice de Morris (anguttara-vol-I-info.txt)')
    print('=' * 92)
    ok16 = sum(1 for v, _nm, pg in MORRIS if by_num.get(v * 10 - 9, {}).get('page') == pg)
    print(f'PTS (BD libro 17): {len(marks)} marcadores, nº corridos '
          f'{min(by_num)}..{max(by_num)} | Excel: {len(ind)} individuales + {len(rng)} de rango '
          f'| CST: {len(units)} suttas')
    print(f'estructura vs Morris: {ok16}/16 páginas de arranque de vagga exactas')

    pts = align_pts(ind, marks, cst)
    tasks, no_cst, no_pts, name_ok, ord_ok, ord_n, page_ok = [], [], [], 0, 0, 0, 0
    for k, e in enumerate(ind):
        if only is not None and e['num'] not in only:
            continue
        u = cst.get(e['num'])
        if not u:
            no_cst.append(e); continue
        if an_names.score(u['title'], e['name']) >= 60:
            name_ok += 1
        oo = an_names.ordinal_ok(u['title'], e['name'])
        if oo is not None:
            ord_n += 1; ord_ok += bool(oo)
        m = pts.get(e['num'])
        if not m:
            no_pts.append(e); continue
        if isinstance(e['page'], int) and abs(m['page'] - e['page']) <= 1:
            page_ok += 1
        tasks.append((e, m, u))

    tot = max(1, len(ind))
    print(f'\nalineadas con sutta del CST: {len(ind) - len(no_cst)}/{len(ind)}')
    print(f'  cruzada — nombre PTS ≡ título CST: {name_ok}/{tot} ({100 * name_ok / tot:.0f}%)')
    print(f'  cruzada — ordinal (dígito PTS ≡ paṭhama/dutiya del CST): {ord_ok}/{ord_n}')
    print(f'  cruzada — página del marcador ≡ página del Excel (±1): {page_ok}/{tot} '
          f'({100 * page_ok / tot:.0f}%)  | sin marcador: {len(no_pts)}')
    if dry:
        print('\n(dry-run, nada validado ni escrito)')
        return

    run = tasks if (do_all or only) else tasks[:n]
    print(f'\nValidando {len(run)} (Gemini en vivo)...', flush=True)
    rows = []
    for k, (e, m, u) in enumerate(run, 1):
        cstx = ' '.join(sh.tokens(u['text'])[:350])
        res = validate_pair(m['text'], cstx, e['name'], u['title'], 'AN', concordant=True)
        rows.append({'num': e['num'], 'name': e['name'], 'cst_title': u['title'],
                     'vri_ref': f"{VRI_STEM}:{min(u['pn'])}" if u['pn'] else None,
                     'pts_num': m['num'], 'pts_page': m['page'], 'pts_line': m['line'],
                     'legacy': e['legacy'], **res})
        if k % 25 == 0:
            print(f'  ...{k}/{len(run)}', flush=True)
    prev = json.load(open(OUT)) if os.path.exists(OUT) else []
    merged = {r['num']: r for r in prev}
    merged.update({r['num']: r for r in rows})
    json.dump(list(merged.values()), open(OUT, 'w'), ensure_ascii=False, indent=1)
    print(f"\nEstado: {dict(Counter(r['estado'] for r in rows))} | "
          f"Validation: {dict(Counter(r['validation'] for r in rows))}")
    print(f'Resultados → {OUT} ({len(merged)} en total; nada escrito al Excel).')


if __name__ == '__main__':
    main()
