#!/usr/bin/env python3
"""calibrate_sn3_lines.py — calibra el nº de LÍNEA de las referencias PTS de SN III (S iii).

Igual criterio que en S i y S v: la línea de `PTS Ref` (`S ii <pág>,<línea>`) es la del **marcador
del sutta**, aquí el `N (M) Nombre` que localiza `sn2_markers.py`. La posición no se adivina: el
emparejamiento va por clave `(saṃyutta, nº corrido)` y la estructura del volumen está cotejada
contra el front matter de Feer, con `cst_p_page` coincidiendo en las 248 filas.

OJO con los miembros de un rango peyyāla: varios suttas comparten el marcador de cabecera (S ii
243: `38-43 (8) Pitā (9) Bhātā …`), así que comparten página y línea. Es correcto, y se anota.

Solo se reescribe la LÍNEA, y solo cuando el marcador está en la página que declara el Excel; los
desacuerdos de página se listan aparte (reescribir la página es una afirmación mayor).

`--fix-pages` corrige la página SOLO donde hay doble evidencia independiente: el marcador en la BD
**y** el ancla `cst_p_page` del concordance apuntan a la misma página, distinta de la del Excel.
Excluye por diseño los miembros de un marcador de rango.

Uso: python3 calibrate_sn2_lines.py [--write] [--fix-pages]
"""
import datetime, re, shutil, sqlite3, sys
from collections import Counter

from openpyxl import load_workbook

from validador_sn3 import (DB, XLSX, DITTHI, build_massive, build_pts_suttas,
                           calibrate_offsets, ditthi_pairs, excel_entries, resolve_pts)

REF_RE = re.compile(r'^(S\s+iii\s+)(\d+)(?:\s*,\s*(\d+))?\s*$')


def main():
    write = '--write' in sys.argv
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
    pts_by_key = build_pts_suttas(conn.cursor())
    canon2para0 = build_massive()
    by_page = {}
    for (sam, num), q in pts_by_key.items():
        by_page.setdefault((sam, q['page']), []).append(q)
    # misma resolución que el validador: por nombre sobre la página del ancla (la clave no sirve)
    # misma resolución que el validador, incluidos los offsets calibrados y el Diṭṭhi por posición
    offsets = calibrate_offsets(pts_by_key, canon2para0,
                                {k: v[1] for k, v in canon2para0.items()})
    ditthi = ditthi_pairs(pts_by_key)
    entries, pts = [], []
    for e in excel_entries():
        if e['sam'] == DITTHI and 1 <= e['inner'] <= len(ditthi):
            q = ditthi[e['inner'] - 1][1]
        else:
            q = resolve_pts(e, canon2para0.get((e['sam'], e['inner'])), pts_by_key, by_page, offsets)
        if q:
            entries.append(e); pts.append(q)
    print(f'PTS: {len(pts_by_key)} suttas | filas del Excel resueltas: {len(entries)}')

    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}

    # fila del Excel por `Sutta #`, para escribir sin depender del orden de iteración
    ridx_of = {}
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') == 'SN' and str(v[ci['PTS Roman']] or '').strip().lower() == 'iii':
            ridx_of[str(v[ci['Sutta #']])] = ridx

    fix_pages = '--fix-pages' in sys.argv
    canon2para = build_massive()

    delta, fixes, otherpage, ranges, pagefix = Counter(), [], [], [], []
    for e, p in zip(entries, pts):
        ref = e['ref']
        if p['page'] != e['page']:
            hit = canon2para.get((e['sam'], e['inner']))
            anchor = hit[2] if hit else None
            # doble evidencia: marcador y concordance coinciden, y no es miembro de un rango
            corroborated = (anchor == p['page']) and not p['from_range']
            otherpage.append((e['num'], ref, f"S iii {p['page']},{p['line']}", p['name'] or '',
                              anchor, corroborated))
            if corroborated:
                pagefix.append((ridx_of[e['num']], e['num'], e['page'], p['page'],
                                f"S iii {p['page']},{p['line']}"))
            continue
        rm = REF_RE.match(ref)
        old = int(rm.group(3)) if (rm and rm.group(3)) else None
        delta[None if old is None else old - p['line']] += 1
        new_ref = f"S iii {p['page']},{p['line']}"
        if p['from_range']:
            # miembros de un rango: comparten encabezado → misma página y línea (es correcto)
            ranges.append((e['num'], new_ref))
        if ref != new_ref:
            fixes.append((ridx_of[e['num']], e['num'], ref, new_ref))

    print('=' * 84)
    print('CALIBRACIÓN DE LÍNEA — SN III (marcador real en la BD, libro 14)')
    print('=' * 84)
    print('desfase (línea del Excel − línea del marcador):')
    for d, n in delta.most_common(12):
        print(f'  {"sin línea" if d is None else f"{d:+d}":>10} : {n}')
    exact, tot = delta.get(0, 0), sum(delta.values())
    print(f'\ncoincidencia exacta: {exact}/{tot} ({100 * exact / max(1, tot):.0f}%)')
    print(f'a corregir: {len(fixes)} | página en disputa (NO se tocan): {len(otherpage)} | '
          f'en marcador de rango §§: {len(ranges)}')
    for _r, num, o, n in fixes[:15]:
        print(f'  SN {num:>7}: {o!r} → {n!r}')
    for num, o, n, nm, anchor, corr in otherpage:
        tag = f'  ← concordance p{anchor} CORROBORA el marcador' if corr else \
              (f'  (concordance p{anchor}; miembro de rango §§)' if anchor else '')
        print(f'  PÁGINA SN {num:>7}: {o!r} vs marcador {n!r} ("{nm}"){tag}')
    if ranges:
        print('  rango §§ (dos suttas, un encabezado):', ', '.join(f'{a}→{b}' for a, b in ranges))

    if not write:
        print('\n(dry-run, nada escrito — usa --write)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-sn3lineas.xlsx'
    shutil.copy(XLSX, bak); print('\nbackup →', bak)
    for ridx, _num, _o, new_ref in fixes:
        ws.cell(row=ridx, column=ci['PTS Ref'] + 1).value = new_ref
    npg = 0
    if fix_pages:
        for ridx, num, oldpg, newpg, new_ref in pagefix:
            ws.cell(row=ridx, column=ci['PTS Page'] + 1).value = newpg
            ws.cell(row=ridx, column=ci['PTS Ref'] + 1).value = new_ref
            print(f'  PÁGINA corregida SN {num}: p{oldpg} → p{newpg}')
            npg += 1
    wb.save(XLSX)
    print(f'Corregidas {len(fixes)} líneas' + (f' y {npg} páginas' if npg else '') + f' en {XLSX}.')


if __name__ == '__main__':
    main()
