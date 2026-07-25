#!/usr/bin/env python3
"""Aligner + validador del **Catukka-nipāta de AN II** (A ii, `book_no=18`).

Segundo tramo de AN. Mismo régimen de marcadores que el Tika de AN I —nº corrido **centrado**, el
vagga no se imprime— así que reutiliza `an1_markers` y `an_names`; lo que cambia es la estructura,
y hubo que fijarla primero (ver `anguttara-vol-II-info.txt`).

**Verdad-terreno del lado PTS.** No sirve la aritmética `10k+1` contra el índice de Morris: su
índice tiene **6 entradas defectuosas de 26** —dos son el **número de sutta** impreso en la columna
de páginas (`Cara` 11 → p13, `Sañcetanika` 171 → p157, la que hacía imposible el índice porque el
vagga 18 salía *después* del 19) y otras cuatro discrepan por poco—. La estructura la fijan los
**26 colofones de vagga** del propio texto, y con ellos sale regular: **26 vaggas de 10 suttas, 259
en total** (el vagga 7 tiene 9 porque el nº **64** falta en la BD).

⚠️ Los colofones vienen en **cuatro formas** y ceñirse a la canónica pierde 8 de los 26: en
versales y sin nombre (`VAGGO PAṬHAMO`), sin ordinal (`Macalavaggo.`), con la llamada de nota
pegada antes de «vaggo» (`[Indriya-]1vaggo`) y con el punto dentro del corchete
(`Caravaggo [dutiyo.]`). Lo resuelve `an1_markers._COLOPHON`.

⚠️ **El CST no cuenta igual**: 28 vaggas y 278 subheads, con los vaggas 23-25 de **once** suttas
—lo que Hardy documenta para AN IV/V («the Commentary counts from one to ten **or eleven**, if
there are more Suttas than ten in a Vagga»)— y un `Rāgapeyyālaṃ` final de 5 subheads sobre **510**
paranums. Por eso el emparejamiento va por **alineamiento monótono**, nunca por clave.

**Claves**: el lado CST se identifica con el paranum del XML VRI (`s0402m3:N`), la clave canónica
(regla (5)). El ancla del lado PTS es la **`DPR Ref`** donde existe (56 de 275 filas), que describe
los desdoblamientos (`4.173` → `AN 4.174a`, `4.174` → `AN 4.174b`) y las fusiones (`4.104` →
`AN 4.104,105`).

Uso: python3 validador_an2_catukka.py [--dry] [--all] [--n N] [--only 4.5,4.9]
"""
import json
import os
import re
import sqlite3
import sys
import xml.etree.ElementTree as ET
from collections import Counter

import an_names
import sutta_hash as sh
from align_rows import assign as ar_assign
from an1_markers import collect_an1
from openpyxl import load_workbook
from validador import validate_pair

VRI = '/tmp/tipitaka-xml/romn/s0402m3.mul.xml'
VRI_STEM = 's0402m3'
OUT = 'validador_an2_catukka.json'
DB = 'src/data/tipitaka.sqlite'
AN2_BOOK = 18
XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# índice de Morris: (vagga, nombre, página de arranque). Verdad-terreno estructural del lado PTS.
# Índice de Morris **corregido contra el texto** (ver `anguttara-vol-II-info.txt`): 20 de sus 26
# entradas son la página correcta; `Cara` y `Sañcetanika` traen el nº de sutta en su lugar, y otras
# cuatro discrepan por poco. Aquí van las páginas REALES del primer sutta de cada vagga.
MORRIS = [(1, 'Bhaṇḍagāma', 1), (2, 'Cara', 13), (3, 'Uruvela', 20), (4, 'Cakka', 32),
          (5, 'Rohitassa', 44), (6, 'Puññābhisanda', 54), (7, 'Pattakamma', 65),
          (8, 'Apaṇṇaka', 76), (9, 'Macala', 83), (10, 'Asura', 91), (11, 'Valāhaka', 102),
          (12, 'Kesi', 112), (13, 'Bhaya', 121), (14, 'Puggala', 133), (15, 'Ābhā', 139),
          (16, 'Indriya', 141), (17, 'Paṭipadā', 149), (18, 'Sañcetanika', 157),
          (19, 'Yodhājīva', 170), (20, 'Mahā', 185), (21, 'Sappurisa', 217),
          (22, 'Sobhana', 225), (23, 'Sucarita', 228), (24, 'Kamma', 230),
          (25, 'Āpatti', 239), (26, 'Abhiññā', 246)]


def build_cst_units(path=VRI):
    """Suttas del CST en orden de documento: `[{title, pn, vagga, text}]`.

    En AN el `<div>` es un **vagga** (no un sutta como en DN, ni un saṃyutta como en SN), así que
    el sutta es el `subhead` de dentro y sus paranums son los `bodytext` que le siguen.
    """
    root = ET.parse(path).getroot()
    out = []
    for d in root.iter('div'):
        h = d.find('head')
        if h is None or h.get('rend') != 'chapter':
            continue
        vagga = ' '.join(''.join(h.itertext()).split())
        cur = None
        for p in d.iter('p'):
            rend, txt = p.get('rend'), ''.join(p.itertext())
            if rend == 'subhead':
                cur = {'title': ' '.join(txt.split()), 'pn': [], 'text': '', 'vagga': vagga}
                out.append(cur)
            elif rend in ('bodytext', 'gatha1', 'gatha2', 'gatha3', 'gathalast') and cur is not None:
                m = re.match(r'(\d+)(?:-(\d+))?$', p.get('n') or '')
                if m:
                    cur['pn'] += list(range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1))
                cur['text'] += ' ' + txt
    return out


def excel_entries():
    """Filas del Tika, separando las **individuales** de las de **rango** (redundantes)."""
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    ind, rng = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['Nikaya']] or '') != 'AN' or str(row[ci['PTS Roman']] or '').strip().lower() != 'ii':
            continue
        num = str(row[ci['Sutta #']])
        if not num.startswith('4.'):
            continue
        e = {'num': num, 'name': str(row[ci['Sutta Name']] or '').strip(),
             'page': row[ci['PTS Page']], 'legacy': str(row[ci['Validation']] or ''),
             'dpr': str(row[ci['DPR Ref']] or '').strip()}
        (rng if '-' in num else ind).append(e)
    ind.sort(key=lambda e: int(e['num'].split('.')[1]))
    return ind, rng


def pts_markers(cur):
    """Marcadores del Tika: nº corrido → `(página, línea, texto acotado)`."""
    pages = {r['page_no']: (r['unitext'] or '').split('\n')
             for r in cur.execute('SELECT page_no,unitext FROM pages '
                                  'WHERE edition="mula" AND book_no=?', (AN2_BOOK,))}
    marks = [x for x in collect_an1(pages) if x['nipata'] == 'CATUKKA' and x['num'] is not None]
    marks.sort(key=lambda x: (x['page'], x['line']))
    flat = [(pg, ln, t) for pg in sorted(pages) for ln, t in enumerate(pages[pg], 1)]
    pos = {(pg, ln): i for i, (pg, ln, _t) in enumerate(flat)}
    out = []
    for k, m in enumerate(marks):
        i0 = pos[(m['page'], m['line'])]
        i1 = pos[(marks[k + 1]['page'], marks[k + 1]['line'])] if k + 1 < len(marks) else len(flat)
        out.append({'num': m['num'], 'page': m['page'], 'line': m['line'], 'ord': k,
                    'vagga': (m['num'] - 1) // 10 + 1,
                    # el texto va SIN recortar: cuando el CST parte en varios lo que PTS numera
                    # como uno, el párrafo homólogo puede caer muy por dentro del sutta (el
                    # `Sarada` es el §4 del nº92, dos páginas más allá) y un tope de 350 tokens lo
                    # dejaba fuera del alcance de `pts_window`
                    'text': ' '.join(sh.tokens(' '.join(t for _p, _l, t in flat[i0:i1])))})
    return out


def align(entries, units):
    """Alineamiento **monótono** filas↔suttas del CST (`align_rows`), por nombre y ordinal.

    No se empareja por clave: el `Sutta #` sólo coincide con el nº corrido de PTS en 103 de 156.
    """
    def score(i, j):
        e, u = entries[i], units[j]
        s = an_names.score(u['title'], e['name'])
        ok = an_names.ordinal_ok(u['title'], e['name'])
        if ok is True:
            s += 60                       # el ordinal desempata dentro de la serie
        elif ok is False:
            s -= 120                      # «Pāpaṇika 1» no puede ser «Dutiya…»
        return s if s > 0 else 0.1        # nunca None: la serie debe poder avanzar

    idx = ar_assign(len(entries), len(units), score, lambda j: 1, skip_penalty=200.0)
    return {entries[i]['num']: (units[j] if j is not None else None) for i, j in enumerate(idx)}


def pts_anchor(e):
    """Nº corrido de PTS que declara la fila: la **`DPR Ref` si existe**, si no el `Sutta #`.

    Es el ancla determinista del lado PTS, y sale del re-parseo de `Raw ID` (`an_rawid.py`): la
    `DPR Ref` describe exactamente los desdoblamientos y fusiones del Excel —`3.32 Ānanda` →
    `AN 3.32a` y `3.33 Sāriputta` → `AN 3.32b` (una del DPR partida en dos), `3.39 Sukhumāla` →
    `AN 3.38-9` (dos fundidas en una)—, que es justo lo que hace que haya **163 marcadores para
    156 filas**. Sin ella el alineamiento acumulaba deriva: `3.111 «Nidāna 1»` acababa en el
    marcador 110 cuando su `Raw ID` dice `[AN 3.107]-08`, o sea los marcadores **107-108**.
    """
    src = e.get('dpr') or ('AN ' + e['num'])
    m = re.search(r'(\d+)\.(\d+)', src)
    if not m:
        # el punto se perdió en la fuente: «AN 3102b» por «AN 3.102b»
        m = re.search(r'AN\s*(\d)(\d+)', src)
    return int(m.group(2)) if m else None


def align_pts(entries, marks, cst):
    """Alineamiento **monótono** filas↔marcadores de PTS, anclado en la PÁGINA del Excel.

    No vale la posición: hay **163 marcadores para 156 filas** —PTS numera por separado suttas que
    el Excel lleva agrupados— y emparejar la fila k con el marcador k+1 acumula deriva (desde
    `3.34` la página se va +2 y sigue creciendo). Sólo cuadraba en 104 de 156.
    """
    # Cuántas filas declaran cada nº corrido: los sufijos `a`/`b` de la `DPR Ref` marcan un sutta
    # de PTS que el Excel **parte en dos filas** (`3.32 Ānanda` → `AN 3.32a`, `3.33 Sāriputta` →
    # `AN 3.32b`), así que ese marcador tiene capacidad 2, no 1.
    anchors = Counter(a for a in (pts_anchor(x) for x in entries) if a is not None)
    reserved = set(anchors)

    def score(i, j):
        e, m = entries[i], marks[j]
        # El nº corrido que declara la fila es evidencia DIRECTA, no una preferencia: va como
        # restricción dura y además se le reserva el marcador, para que ninguna otra fila lo ocupe.
        a = pts_anchor(e)
        if a is not None:
            return 1000.0 if m['num'] == a else None
        if m['num'] in reserved:
            return None
        if not isinstance(e['page'], int):
            return 0.1
        d = abs(m['page'] - e['page'])
        if d > 3:
            return None
        s = 40 - 10 * d
        u = cst.get(e['num'])
        if u:                                   # el título CST refuerza, no decide
            s += an_names.score(u['title'], e['name']) / 10

        return max(s, 0.1)

    idx = ar_assign(len(entries), len(marks), score,
                    lambda j: max(1, anchors.get(marks[j]['num'], 1)), skip_penalty=500.0)
    return {entries[i]['num']: (marks[j] if j is not None else None) for i, j in enumerate(idx)}


def pts_window(pts_text, cst_text, width=350):
    """Acota el texto de PTS al **párrafo** que corresponde al sutta del CST.

    En el Tika, PTS numera como UN sutta lo que el CST parte en varios: el `Saradasuttaṃ` del CST
    es el **§4 del nº92** de PTS (A i 242,17), el `Nimittasuttaṃ` vive dentro del nº100 y el
    `Samaṇabrāhmaṇasuttaṃ` dentro del nº102. La alineación al sutta contenedor es correcta, pero
    cotejar el sutta breve del CST contra el contenedor entero hunde la cobertura (0.06) y produce
    un REJECT que no dice nada de la fila.

    Se localiza el incipit del CST dentro del texto de PTS y se devuelve una ventana desde ahí. Si
    no se encuentra, se deja el texto tal cual: nunca se inventa un locus.
    """
    pt, ct = pts_text.split(), cst_text.split()
    if len(pt) <= width or len(ct) < 6:
        return pts_text
    probe = ct[:6]
    for i in range(len(pt) - len(probe)):
        if pt[i:i + len(probe)] == probe:
            return ' '.join(pt[i:i + width])
    # sin coincidencia exacta: se busca la mejor ventana por solapamiento léxico
    head = set(ct[:60])
    best, bi = -1, None
    for i in range(0, max(1, len(pt) - 60), 20):
        ov = len(head & set(pt[i:i + 60]))
        if ov > best:
            best, bi = ov, i
    return ' '.join(pt[bi:bi + width]) if bi is not None and best >= 12 else pts_text


def main():
    dry = '--dry' in sys.argv
    do_all = '--all' in sys.argv
    n = int(sys.argv[sys.argv.index('--n') + 1]) if '--n' in sys.argv else 30
    only = set(sys.argv[sys.argv.index('--only') + 1].split(',')) if '--only' in sys.argv else None

    units = build_cst_units()
    ind, rng = excel_entries()
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
    marks = pts_markers(conn.cursor())
    by_num = {m['num']: m for m in marks}
    cst = align(ind, units)

    print('=' * 92)
    print('AN II — CATUKKA-NIPĀTA: pipeline VRI + estructura por colofones (anguttara-vol-II-info.txt)')
    print('=' * 92)
    ok16 = sum(1 for v, _nm, pg in MORRIS if by_num.get(v * 10 - 9, {}).get('page') == pg)
    print(f'PTS (BD libro 18): {len(marks)} marcadores, nº corridos '
          f'{min(by_num)}..{max(by_num)} | Excel: {len(ind)} individuales + {len(rng)} de rango '
          f'| CST: {len(units)} suttas')
    print(f'estructura vs colofones: {ok16}/26 páginas de arranque de vagga exactas')

    pts = align_pts(ind, marks, cst)
    tasks, no_cst, no_pts, name_ok, ord_ok, ord_n, page_ok = [], [], [], 0, 0, 0, 0
    for k, e in enumerate(ind):
        if only is not None and e['num'] not in only:
            continue
        u = cst.get(e['num'])
        if not u:
            no_cst.append(e); continue
        if an_names.score(u['title'], e['name']) >= 60:
            name_ok += 1
        oo = an_names.ordinal_ok(u['title'], e['name'])
        if oo is not None:
            ord_n += 1; ord_ok += bool(oo)
        m = pts.get(e['num'])
        if not m:
            no_pts.append(e); continue
        if isinstance(e['page'], int) and abs(m['page'] - e['page']) <= 1:
            page_ok += 1
        tasks.append((e, m, u))

    tot = max(1, len(ind))
    print(f'\nalineadas con sutta del CST: {len(ind) - len(no_cst)}/{len(ind)}')
    print(f'  cruzada — nombre PTS ≡ título CST: {name_ok}/{tot} ({100 * name_ok / tot:.0f}%)')
    print(f'  cruzada — ordinal (dígito PTS ≡ paṭhama/dutiya del CST): {ord_ok}/{ord_n}')
    print(f'  cruzada — página del marcador ≡ página del Excel (±1): {page_ok}/{tot} '
          f'({100 * page_ok / tot:.0f}%)  | sin marcador: {len(no_pts)}')
    if dry:
        print('\n(dry-run, nada validado ni escrito)')
        return

    run = tasks if (do_all or only) else tasks[:n]
    print(f'\nValidando {len(run)} (Gemini en vivo)...', flush=True)
    rows = []
    for k, (e, m, u) in enumerate(run, 1):
        cstx = ' '.join(sh.tokens(u['text'])[:350])
        ptsx = pts_window(m['text'], cstx)
        res = validate_pair(ptsx, cstx, e['name'], u['title'], 'AN', concordant=True)
        rows.append({'num': e['num'], 'name': e['name'], 'cst_title': u['title'],
                     'vri_ref': f"{VRI_STEM}:{min(u['pn'])}" if u['pn'] else None,
                     'pts_num': m['num'], 'pts_page': m['page'], 'pts_line': m['line'],
                     'legacy': e['legacy'], **res})
        if k % 25 == 0:
            print(f'  ...{k}/{len(run)}', flush=True)
    prev = json.load(open(OUT)) if os.path.exists(OUT) else []
    merged = {r['num']: r for r in prev}
    merged.update({r['num']: r for r in rows})
    json.dump(list(merged.values()), open(OUT, 'w'), ensure_ascii=False, indent=1)
    print(f"\nEstado: {dict(Counter(r['estado'] for r in rows))} | "
          f"Validation: {dict(Counter(r['validation'] for r in rows))}")
    print(f'Resultados → {OUT} ({len(merged)} en total; nada escrito al Excel).')


if __name__ == '__main__':
    main()
