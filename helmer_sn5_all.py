#!/usr/bin/env python3
"""SN V Helmer v2: saṃyutta-aware CST mapping."""
import hashlib, json, os, re, sqlite3
from collections import defaultdict, Counter
from pathlib import Path
import sutta_hash as sh

CACHE_FILE=Path('helmer_ptscst_cache.json'); SN5_BOOK=16; BATCH=50
conn=sqlite3.connect('src/data/tipitaka.sqlite'); conn.row_factory=sqlite3.Row; cur=conn.cursor()

marker_map=defaultdict(list)
cur.execute('SELECT page_no, unitext FROM pages WHERE book_no=? AND edition="mula"',(SN5_BOOK,))
for r in cur.fetchall():
    for i,line in enumerate(r['unitext'].split('\n')):
        s=line.strip()
        m=re.match(r'^(\d+)\.?\s*\((\d+)\)\s+(\S.*)',s)
        if m: marker_map[r['page_no']].append((i+1,int(m.group(1)),int(m.group(2)),m.group(3)[:60]))

from openpyxl import load_workbook
wb=load_workbook('PTS_Reference_Complete_Canon.xlsx'); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
entries=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2: continue
    if int(parts[0])<45 or int(parts[0])>56: continue
    if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='v': continue
    entries.append({'num':snum,'ref':str(ws.cell(row=ri,column=cols['PTS Ref']).value or ''),
        'name':str(ws.cell(row=ri,column=cols['Sutta Name']).value or ''),
        'page':ws.cell(row=ri,column=cols['PTS Page']).value,
        'sn':int(parts[0]),'sn_inner':int(parts[1])})
entries.sort(key=lambda e:(e['sn'],e['sn_inner']))

# Get CST saṃyutta boundaries
segs=sh.load_cha(['s5m.xml'],'h4n')
cst_boundaries={}
current_sn=45
for i,s in enumerate(segs):
    m=re.match(r'^1\.\s+\S',s.get('title',''))
    if m:
        cst_boundaries[current_sn]=i
        current_sn+=1
cst_boundaries[current_sn]=len(segs)  # end marker

# Get exact-match tasks
all_tasks=[]
for e in entries:
    page=e['page']; target=e['sn_inner']
    for line,gid,vpos,name in marker_map.get(page,[]):
        if gid==target:
            cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN5_BOOK,page))
            r=cur.fetchone()
            if r:
                lines=r['unitext'].split('\n'); start=line-1; end=len(lines)
                for j in range(start+1,len(lines)):
                    if re.match(r'^\d+\.?\s*\(\d+\)',lines[j].strip()): end=j; break
                text=' '.join(lines[start:end])
                all_tasks.append((e,' '.join(sh.tokens(text)[:300]),line,gid,name[:40]))
            break

print('SN V: {} entries, {} exact matches'.format(len(entries),len(all_tasks)))

cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}
from openai import OpenAI
cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')
SYS='You are Helmer Smith. Same sutta? Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words"}'

total_vc=Counter()
for batch_start in range(0,len(all_tasks),BATCH):
    batch=all_tasks[batch_start:batch_start+BATCH]
    vc=Counter()
    for i,(e,pts,l,g,n) in enumerate(batch):
        # Saṃyutta-aware CST index
        sn=e['sn']
        if sn in cst_boundaries:
            cst_idx=cst_boundaries[sn]+(e['sn_inner']-1)
        else:
            continue
        if cst_idx>=len(segs): continue
        cst_text=' '.join(sh.tokens(segs[cst_idx]['text'])[:300])
        key=hashlib.md5((pts[:200]+cst_text[:200]).encode()).hexdigest()
        if key in cache: hv=cache[key]
        else:
            try:
                r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
                    response_format={'type':'json_object'},
                    messages=[{'role':'system','content':SYS},
                              {'role':'user','content':'PTS(gid={}):\n{}\n\nCST:\n{}'.format(g,pts[:800],cst_text[:800])}])
                hv=json.loads(r.choices[0].message.content)
            except Exception as ex:
                hv={'verdict':'ERROR','reason':str(ex)[:60]}
            cache[key]=hv; CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
        v=hv.get('verdict','?'); vc[v]+=1; total_vc[v]+=1
    
    acc=vc.get('APPROVE',0)+vc.get('REJECT',0)
    if acc>0:
        done=batch_start+len(batch)
        print('  {}/{} | APPROVE: {} REJECT: {} = {:.0f}%'.format(done,len(all_tasks),vc.get('APPROVE',0),vc.get('REJECT',0),100*vc.get('APPROVE',0)/acc))

acc=total_vc.get('APPROVE',0)+total_vc.get('REJECT',0)
print('\n' + '='*70)
print('SN V: APPROVE {} REJECT {} = {:.0f}%'.format(total_vc.get('APPROVE',0),total_vc.get('REJECT',0),100*total_vc.get('APPROVE',0)/acc if acc>0 else 0))
conn.close()
