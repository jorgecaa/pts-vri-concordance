#!/usr/bin/env python3
"""SN V Helmer pilot: first 10 exact gid matches."""
import hashlib, json, os, re, sqlite3
from collections import defaultdict, Counter
from pathlib import Path
import sutta_hash as sh

CACHE_FILE=Path('helmer_ptscst_cache.json'); SN5_BOOK=16
conn=sqlite3.connect('src/data/tipitaka.sqlite'); conn.row_factory=sqlite3.Row; cur=conn.cursor()

# Build markers for SN V (book 16) — format: N. (M) Name (with dot)
marker_map=defaultdict(list)
cur.execute('SELECT page_no, unitext FROM pages WHERE book_no=? AND edition="mula"',(SN5_BOOK,))
for r in cur.fetchall():
    for i,line in enumerate(r['unitext'].split('\n')):
        s=line.strip()
        m=re.match(r'^(\d+)\.?\s*\((\d+)\)\s+(\S.*)',s)
        if m: marker_map[r['page_no']].append((i+1,int(m.group(1)),int(m.group(2)),m.group(3)[:60]))

from openpyxl import load_workbook
wb=load_workbook('PTS_Reference_Complete_Canon.xlsx'); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
entries=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2: continue
    if int(parts[0])<45 or int(parts[0])>56: continue
    if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='v': continue
    entries.append({'num':snum,'ref':str(ws.cell(row=ri,column=cols['PTS Ref']).value or ''),
        'name':str(ws.cell(row=ri,column=cols['Sutta Name']).value or ''),
        'page':ws.cell(row=ri,column=cols['PTS Page']).value,
        'sn':int(parts[0]),'sn_inner':int(parts[1])})
entries.sort(key=lambda e:(e['sn'],e['sn_inner']))
print('SN V: {} entries'.format(len(entries)))

# First 10 with exact gid match
tasks=[]
for e in entries:
    if len(tasks)>=10: break
    page=e['page']; target=e['sn_inner']
    for line,gid,vpos,name in marker_map.get(page,[]):
        if gid==target:
            cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN5_BOOK,page))
            r=cur.fetchone()
            if r:
                lines=r['unitext'].split('\n'); start=line-1; end=len(lines)
                for j in range(start+1,len(lines)):
                    if re.match(r'^\d+\.?\s*\(\d+\)',lines[j].strip()): end=j; break
                text=' '.join(lines[start:end])
                tasks.append((e,' '.join(sh.tokens(text)[:300]),line,gid,name[:40]))
            break

print('Pilot: {} with exact match'.format(len(tasks)))
for e,pts,l,g,n in tasks:
    print('  SN {:>8s} p.{:>3d} gid={} | {}'.format(e['num'],e['page'],g,e['name'][:30]))

# DeepSeek
segs=sh.load_cha(['s5m.xml'],'h4n')
cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}
from openai import OpenAI
cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')
SYS='You are Helmer Smith. Same sutta? Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words"}'

vc=Counter()
for i,(e,pts,l,g,n) in enumerate(tasks):
    cst_idx=sum(1 for ee in entries if ee['sn']<e['sn'] or (ee['sn']==e['sn'] and ee['sn_inner']<e['sn_inner']))
    cst_text=' '.join(sh.tokens(segs[cst_idx]['text'])[:300])
    key=hashlib.md5((pts[:200]+cst_text[:200]).encode()).hexdigest()
    if key in cache: hv=cache[key]
    else:
        r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
            response_format={'type':'json_object'},
            messages=[{'role':'system','content':SYS},
                      {'role':'user','content':'PTS(gid={}):\n{}\n\nCST:\n{}'.format(g,pts[:800],cst_text[:800])}])
        hv=json.loads(r.choices[0].message.content)
        cache[key]=hv; CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
    v=hv.get('verdict','?'); vc[v]+=1
    sym='✓' if v=='APPROVE' else '✗'
    print('{} SN {:>8s} {:>15s} | {:8s} | {}'.format(sym,e['num'],e['ref'],v,hv.get('reason','')[:50]))

acc=vc.get('APPROVE',0)+vc.get('REJECT',0)
print('\nSN V Pilot: APPROVE {} REJECT {} = {:.0f}%'.format(vc.get('APPROVE',0),vc.get('REJECT',0),100*vc.get('APPROVE',0)/acc if acc>0 else 0))
conn.close()
