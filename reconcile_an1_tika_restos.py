#!/usr/bin/env python3
"""
reconcile_an1_tika_restos — fase D: los restos del **Tika** de A i.

Diez filas se cierran y once quedan anotadas. Las once no son deuda de método: son sitios donde la
numeración del Excel no tiene contraparte en el CST, o donde **el propio Morris declara que está
conjeturando**.

### Lo que se cierra

**Seis filas de rango, por redundancia probada** (`3.2-9`, `3.37-38`, `3.51-52`, `3.58-9`,
`3.96-98`, `3.143-145`): **todos** sus miembros ya están CONFIRMADO individualmente y con su
`VRI Ref`, así que el rango se escribe con los paranums de sus extremos. Ojo a que **no** coinciden
con el `Sutta #`: en el Tika hay un desfase de +1 desde el paranum 49 (documentado), y por eso
`3.51-52` es `52-53` y `3.58-9` es `59-60`.

**Los tres `REVISAR`** (`3.33 Sāriputta`, `3.138 Sampadā`, `3.139 Vuddhi`). Los tres rechazaban por
la misma razón —**PTS imprime dos suttas del CST bajo un solo número**: el marcador 32 cubre el
paranum 33, y el 136 cubre el 139 y el 140— y el cotejo lo confirma: la cobertura del marcador
contra cada paranum es **0,90** en los tres. Es el caso de A ii 174 otra vez (Mahākoṭṭhika +
Ānanda), y se arbitra igual.

**`3.62 Bhaya`**, que el alineador había dejado fuera: el marcador 62 (A i 178) da el paranum **63**
con cobertura 0,96 contra 0,60 del segundo.

### Lo que NO se cierra, y por qué

- **`3.156 Paṭipadā [Acelaka] 1`** — el marcador 151 empata entre los paranums 162 y 163 (0,97 los
  dos). Y es justo donde **Morris avisa de que conjetura**: «The Acelaka-vagga **probably** included
  only suttas 151, 152; so that the ten suttas 153-162 made a second vagga…». Con el editor
  declarando que no lo sabe, resolverlo por cobertura sería inventar precisión.
- **`3.157-162`, `3.163-182`, `3.183-352`** y las cuatro colas del Duka (`2.180-229`, `2.230-279`,
  `2.280-309`, `2.310-479`) — la numeración del Excel **excede la del CST**: su Tika llega a 352 y
  el `s0402m2` acaba en **184**; su Duka llega a 479 y el `s0402m1` acaba en **246**. Los paranums
  `2.280`, `2.310` y los del tramo alto del Tika **no existen**. Es el mismo caso que dejó fuera
  `11.502-981` en A v, y la salida correcta es la misma: decisión editorial, no aritmética.
- **`2.19` (×2)** — la clave duplicada de siempre (`AN 2.19(*) + AN 3.29(*)`), que junta un sutta
  del Duka con uno del Tika. No se toca: es lo que `check_integrity` señala y no se arregla
  escribiéndole encima.

Uso: python3 reconcile_an1_tika_restos.py [--dry]
"""
import shutil
import sys
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'

CIERRE = {
    # (a) rangos redundantes: todos sus miembros ya CONFIRMADO, con su VRI Ref
    '3.2-9': ('s0402m2:2-9', 'rango: los 8 miembros ya CONFIRMADO individualmente'),
    '3.37-38': ('s0402m2:37-38', 'rango: los 2 miembros ya CONFIRMADO'),
    '3.51-52': ('s0402m2:52-53', 'rango: los 2 miembros ya CONFIRMADO (el Tika va +1 desde el 49)'),
    '3.58-9': ('s0402m2:59-60', 'rango «58-9» = 58-59: los 2 miembros ya CONFIRMADO'),
    '3.96-98': ('s0402m2:97-99', 'rango: los 3 miembros ya CONFIRMADO'),
    '3.143-145': ('s0402m2:144-146', 'rango: los 3 miembros ya CONFIRMADO'),
    # (b) los tres REVISAR: PTS imprime dos suttas del CST bajo un número
    '3.33': ('s0402m2:33', 'el marcador 32 (A i 132) cubre también este sutta: cobertura 0.90. '
                           'Mismo caso que A ii 174 (Mahākoṭṭhika + Ānanda)'),
    '3.138': ('s0402m2:139', 'el marcador 136 (A i 287) cubre los paranums 139 y 140: cobertura 0.90'),
    '3.139': ('s0402m2:140', 'el marcador 136 (A i 287) cubre los paranums 139 y 140: cobertura 0.90'),
    # (c) la que el alineador dejó fuera
    '3.62': ('s0402m2:63', 'marcador 62 (A i 178) → paranum 63 por contenido: 0.96 contra 0.60 del '
                           'segundo'),
}

NO_CERRAR = {
    '3.156': 'el marcador 151 empata entre los paranums 162 y 163 (0.97 los dos), y es justo donde '
             'MORRIS DECLARA QUE CONJETURA: «The Acelaka-vagga probably included only suttas 151, '
             '152…». Resolverlo por cobertura sería inventar precisión que el editor no tiene',
    '3.157-162': 'la numeración del Excel excede la del CST: su Tika llega a 352 y `s0402m2` acaba '
                 'en 184',
    '3.163-182': 'ídem; además el paranum 163 del CST está en el Acelakavagga, no en el '
                 'Kammapathapeyyāla que nombra la fila',
    '3.183-352': 'ídem: el tramo alto no existe en el CST',
    '2.180-229': 'la numeración del Excel excede la del CST: su Duka llega a 479 y `s0402m1` acaba '
                 'en 246',
    '2.230-279': 'ídem',
    '2.280-309': 'ídem: el paranum 280 no existe en `s0402m1`',
    '2.310-479': 'ídem: el paranum 310 no existe en `s0402m1`',
}


def main():
    dry = '--dry' in sys.argv
    if not dry:
        cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-an1tika.xlsx'
        shutil.copy(XLSX, cp)
        print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    filas = [r for r in range(2, ws.max_row + 1)
             if ws.cell(r, ci['Nikaya']).value == 'AN'
             and str(ws.cell(r, ci['PTS Roman']).value or '').strip() == 'i']
    dup = {k for k, c in Counter(str(ws.cell(r, ci['Sutta #']).value)
                                 for r in filas).items() if c > 1}
    n, anot = 0, 0
    for r in filas:
        num = str(ws.cell(r, ci['Sutta #']).value)
        if ws.cell(r, ci['Estado']).value == 'CONFIRMADO' or num in dup:
            continue
        if num in NO_CERRAR:
            ws.cell(r, ci['Detail']).value = 'sin resolver: ' + NO_CERRAR[num]
            anot += 1
            continue
        if num not in CIERRE:
            continue
        vri, motivo = CIERRE[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = motivo
        n += 1
    pend = sum(1 for r in filas if ws.cell(r, ci['Estado']).value != 'CONFIRMADO')
    print(f'{n} filas cerradas · {anot} anotadas sin resolver · '
          f'A i: {len(filas) - pend}/{len(filas)} CONFIRMADO')
    if dry:
        print('(dry-run: nada escrito)')
        return
    wb.save(XLSX)
    print(f'escrito → {XLSX}')


if __name__ == '__main__':
    main()
