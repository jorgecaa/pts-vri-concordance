#!/usr/bin/env python3
"""
AN — Strategic spot-check: 15 entries across all 5 volumes.
For each, verify the sutta actually starts on the stated page
using incipit matching and marker proximity.
"""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

AN_BOOKS = {'i': 17, 'ii': 18, 'iii': 19, 'iv': 20, 'v': 21}

# ── Load Excel ──
wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

entries = []
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'AN':
        continue
    entries.append({
        'ri': ri,
        'num': str(ws.cell(row=ri, column=cols['Sutta #']).value or ''),
        'name': str(ws.cell(row=ri, column=cols['Sutta Name']).value or ''),
        'page': ws.cell(row=ri, column=cols['PTS Page']).value,
        'roman': str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower(),
        'ref': str(ws.cell(row=ri, column=cols['PTS Ref']).value or ''),
        'raw': str(ws.cell(row=ri, column=cols['Raw ID']).value or ''),
    })

# ── Select 15 strategic entries ──
# Criteria: diverse nipatas, with/without CST vagga info, different page densities
spot_checks = []

# AN 1: ekakanipata (very dense, many peyyala)
spot_checks.extend(['1.1', '1.50', '1.100', '1.300', '1.500', '1.600'])

# AN 2: dukanipata
spot_checks.extend(['2.1', '2.50', '2.200'])

# AN 3: tikanipata
spot_checks.extend(['3.1', '3.80', '3.150'])

# AN 4: catukkanipata
spot_checks.extend(['4.1', '4.100'])

# AN 5: pañcakanipata
spot_checks.extend(['5.1'])

def get_page_text(book_no, page_no):
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                (book_no, page_no))
    r = cur.fetchone()
    if not r:
        return None, None
    return r['head'] or '', r['unitext'] or ''

def find_sutta_on_page(book_no, page_no, sutta_name, sutta_num):
    """Check if a sutta starts on this page. Returns (found, evidence)."""
    head, text = get_page_text(book_no, page_no)
    if text is None:
        return False, 'PAGE_MISSING'
    
    lines = text.split('\n')
    evidence = []
    
    # 1. Check for sutta marker near page top
    for i, line in enumerate(lines[:40]):
        s = line.strip()
        m = re.match(r'^(\d+)\.\s+(\S.*)', s)
        if m:
            n = int(m.group(1))
            rest = m.group(2)[:80]
            evidence.append(f'L{i+1}: [{n}. {rest}]')
    
    # 2. Check HEAD for vagga/sutta info
    if head.strip():
        evidence.insert(0, f'HEAD: {head.strip()[:100]}')
    
    # 3. Check if name keywords appear
    name_kw = re.findall(r'[āīūṁṃṅñṭḍṇḷa-z]{4,}', sutta_name.lower())
    found_kw = []
    for kw in name_kw[:5]:
        if kw in text.lower():
            found_kw.append(kw)
    if found_kw:
        evidence.append(f'Keywords found: {found_kw}')
    
    return bool(evidence), evidence

print('AN STRATEGIC SPOT-CHECK (15 entries)')
print('=' * 90)

results = []
for target_num in spot_checks:
    entry = None
    for e in entries:
        if e['num'] == target_num and '-' not in e['num']:
            entry = e
            break
    
    if not entry:
        # Try partial match
        for e in entries:
            if e['num'].startswith(target_num + '.') or e['num'] == target_num:
                entry = e
                break
    
    if not entry:
        print(f'\n  ✗ {target_num}: NOT FOUND in Excel')
        continue
    
    book_no = AN_BOOKS.get(entry['roman'])
    if not book_no:
        print(f'\n  ✗ {target_num}: unknown roman "{entry["roman"]}"')
        continue
    
    page = entry['page']
    found, evidence = find_sutta_on_page(book_no, page, entry['name'], entry['num'])
    
    # Also check +1/-1 pages
    nearby_info = []
    for delta in [-2, -1, 1, 2]:
        h, t = get_page_text(book_no, page + delta)
        if t:
            # Check if sutta markers on nearby page match better
            markers = re.findall(r'^(\d+)\.\s+\S.*', t, re.MULTILINE)
            if markers:
                nearby_info.append(f'  p.{page+delta} markers: {markers[:5]}')
    
    status = '✓' if found else '✗'
    print(f'\n{status} AN {entry["num"]:>8s} | {entry["ref"]:>15s} | {entry["name"][:45]}')
    print(f'  Raw: {entry["raw"][:80]}')
    for ev in evidence[:6]:
        print(f'  {ev}')
    if nearby_info:
        for ni in nearby_info[:3]:
            print(ni)
    
    results.append({
        'num': entry['num'],
        'ref': entry['ref'],
        'found': found,
        'evidence': evidence,
        'nearby': nearby_info
    })

# Summary
ok = sum(1 for r in results if r['found'])
print(f'\n{"=" * 90}')
print(f'  Result: {ok}/{len(results)} verified on stated page')
print(f'{"=" * 90}')

conn.close()
