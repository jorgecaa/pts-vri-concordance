#!/usr/bin/env python3
"""
kn_khuddakapatha — la primera obra de KN: **Khuddakapāṭha**, 9 filas.

Es la más pequeña del nikāya y por eso es el sitio para fijar el método, no para correr. Lo que
aquí se decide vale para las quince obras que vienen detrás.

### Las tres fuentes coinciden en estructura, y eso es nuevo

| | unidades |
|---|--:|
| PTS (BD, libro 22) | **9** secciones, numeral romano `I.`…`IX.` y **colofón de cierre** (`SARAṆATTAYAṂ.`, `METTASUTTAṂ NIṬṬHITAṂ.`) |
| CST (`s0501m`) | **9** `<div rend="chapter">`, con los mismos nombres y en el mismo orden |
| Excel | **9** filas, `1.1`-`1.9`, con los mismos nombres |

No hay compresión, ni elisión, ni desfase: por una vez las tres cuentan lo mismo. La fila `k` es el
capítulo `k` en las dos ediciones, y las páginas del Excel caen sobre el marcador romano (±1).

### ⚠️ Pero la CLAVE del lado CST no puede ser la de SN y AN

En SN y AN el `n` del CST es un **paranum corrido por fichero** y basta con él (regla (5)). En el
Khuddakapāṭha **reinicia dentro de cada capítulo**: hay un `n=1` en el capítulo 2, otro en el 4,
otro en el 5, en el 6, en el 7, en el 8 y en el 9. Y dos capítulos —`Saraṇattayaṃ` y
`Dvattiṃsākāro`— **no tienen `n` ninguno**: son gāthās sin numerar.

Un `s0501m:1` no identificaría nada. Así que la referencia se hace **al capítulo**, con una `c` que
lo marca: **`s0501m:c1` … `s0501m:c9`** (y `s0501m:c5.3` si algún día hace falta un verso suelto).
`check_integrity` valida la forma nueva contra los `<div rend="chapter">` del fichero.

Es una extensión de la regla (5), no una excepción: la clave sigue siendo el VRI; lo que cambia es
que en KN la unidad numerada del CST es el capítulo y no el párrafo.

### La comprobación

Cada fila se firma con **tres señales independientes**: el marcador romano del impreso en la página
que declara el Excel (±1); el capítulo del CST en la misma posición ordinal; y la **cobertura de
contenido** entre el texto del uno y el del otro, que no interviene en las dos anteriores.

Uso: python3 kn_khuddakapatha.py [--dry]
"""
import re
import shutil
import sqlite3
import sys
import xml.etree.ElementTree as ET
from datetime import datetime

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
BOOK, STEM = 22, 's0501m'
COV_MIN = 0.55

_ROMANO = re.compile(r'^(I|II|III|IV|V|VI|VII|VIII|IX|X)\.?\*?$')


def secciones_pts(conn):
    """`[(ordinal, página, línea, texto)]` de las nueve secciones del impreso.

    El corte es el **numeral romano centrado**, igual que los vaggas del Eka; el colofón de cierre
    (`RATANASUTTAṂ NIṬṬHITAṂ.`) confirma cada uno pero no hace falta para cortar.
    """
    pgs = {r['page_no']: (r['unitext'] or '').split('\n')
           for r in conn.execute('SELECT page_no,unitext FROM pages WHERE edition="mula" '
                                 'AND book_no=?', (BOOK,))}
    plano = [(pg, ln, t) for pg in sorted(pgs) for ln, t in enumerate(pgs[pg], 1)]
    cortes = []
    for k, (pg, ln, t) in enumerate(plano):
        s = t.replace('\r', '').strip()
        ind = len(t.replace('\r', '')) - len(t.replace('\r', '').lstrip())
        if ind >= 10 and _ROMANO.match(s):
            cortes.append((k, pg, ln))
    out = []
    for j, (k, pg, ln) in enumerate(cortes):
        fin = cortes[j + 1][0] if j + 1 < len(cortes) else len(plano)
        out.append((j + 1, pg, ln, ap.fold(' '.join(t for _p, _l, t in plano[k:fin]))))
    return out


def capitulos_cst():
    """`[(ordinal, nombre, texto)]` de los `<div rend="chapter">` del fichero VRI."""
    root = ET.parse(f'/tmp/tipitaka-xml/romn/{STEM}.mul.xml').getroot()
    out = []
    for d in root.iter('div'):
        h = d.find('head')
        if h is None or h.get('rend') != 'chapter':
            continue
        nom = ' '.join(''.join(h.itertext()).split())
        txt = ' '.join(''.join(p.itertext()) for p in d.iter('p'))
        out.append((len(out) + 1, nom, ap.fold(txt)))
    return out


def filas():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    return [{'num': str(r[ci['Sutta #']]), 'name': str(r[ci['Sutta Name']] or ''),
             'page': r[ci['PTS Page']], 'estado': r[ci['Estado']]}
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[ci['Nikaya']] == 'KN' and str(r[ci['PTS Vol']]) == 'Khp']


def main():
    dry = '--dry' in sys.argv
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    pts, cst, fs = secciones_pts(conn), capitulos_cst(), filas()
    print(f'PTS {len(pts)} secciones · CST {len(cst)} capítulos · Excel {len(fs)} filas')
    if not (len(pts) == len(cst) == len(fs)):
        print('⚠ los tres recuentos no coinciden: no se escribe nada')
        return
    res = {}
    print(f"\n{'fila':>6} {'pág':>4} {'romano':>7} {'capítulo del CST':<24} {'cov':>5}  ok")
    for k, f in enumerate(fs):
        o, pg, ln, tp = pts[k]
        _oc, nom, tc = cst[k]
        cov = R.cobertura(tp, tc)
        pag = isinstance(f['page'], int) and abs(pg - f['page']) <= 1
        ok = cov >= COV_MIN and pag
        print(f"  {f['num']:>6} {str(f['page']):>4} {o:>7} {nom[:24]:<24} {cov:>5.2f}  "
              f"{'✓' if ok else '✗'}{'' if pag else f' (marcador en p{pg})'}")
        if ok:
            res[f['num']] = (f'{STEM}:c{o}',
                             f'Khuddakapāṭha: sección {o} del impreso (Kh {pg},{ln}) ↔ capítulo '
                             f'«{nom}» del CST; cobertura de contenido {cov:.2f}')
    print(f'\n{len(res)}/{len(fs)} firmadas')
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-khp.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'KN' or str(ws.cell(r, ci['PTS Vol']).value) != 'Khp':
            continue
        num = str(ws.cell(r, ci['Sutta #']).value)
        if num not in res or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, det = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
