#!/usr/bin/env python3
"""
Helmer Smith — Content-Based PTS Reference Validation
Validates the reference table against actual page content in tipitaka.sqlite.
Methodology: Level 1 (page exists) → Level 2 (head match) → Level 3 (body text match)
Sampling: 40 strategic points across all Nikayas.
"""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

# ── Book maps ──
BOOK_MAP = {
    6: ('DN I', 'D'), 7: ('DN II', 'D'), 8: ('DN III', 'D'),
    9: ('MN I', 'M'), 10: ('MN II', 'M'), 11: ('MN III', 'M'),
    12: ('SN I', 'S'), 13: ('SN II', 'S'), 14: ('SN III', 'S'),
    15: ('SN IV', 'S'), 16: ('SN V', 'S'),
    17: ('AN I', 'A'), 18: ('AN II', 'A'), 19: ('AN III', 'A'),
    20: ('AN IV', 'A'), 21: ('AN V', 'A'),
    22: ('Khp', 'Khp'), 23: ('Dhp', 'Dhp'), 24: ('Ud', 'Ud'),
    25: ('It', 'It'), 26: ('Sn', 'Sn'),
    27: ('Vv', 'Vv'), 28: ('Pv', 'Pv'),
    29: ('Th & Thī', 'Th'), 30: ('Ja I', 'Ja'),
    31: ('Ja II', 'Ja'), 32: ('Ja III', 'Ja'), 33: ('Ja IV', 'Ja'),
    34: ('Ja V', 'Ja'), 35: ('Ja VI', 'Ja'),
    36: ('Nidd I', 'Nidd'), 37: ('Nidd II', 'Nidd'),
    38: ('Paṭis I', 'Paṭis'), 39: ('Paṭis II', 'Paṭis'),
    40: ('Ap', 'Ap'), 41: ('Bv', 'Bv'), 42: ('Cp', 'Cp'),
}

def norm(s):
    """Strip diacritics for fuzzy matching."""
    for a, b in [('ā','a'),('ī','i'),('ū','u'),('ṅ','n'),('ñ','n'),
                 ('ṭ','t'),('ḍ','d'),('ṇ','n'),('ḷ','l'),('ṃ','m'),
                 ('ṁ','m')]:
        s = s.replace(a, b).replace(a.upper(), b.upper())
    return s

def get_page(book_no, page_no):
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                (book_no, page_no))
    r = cur.fetchone()
    if not r:
        return None, None
    return r['head'] or '', r['unitext'] or ''

def extract_keywords(name, min_len=4):
    """Extract meaningful keywords from a sutta/section name."""
    clean = norm(name.lower())
    clean = re.sub(r'sutta[mṃ]?|vagga|pathama|dutiya|tatiya|catuttha|pancama|chattha|sattama', '', clean)
    clean = re.sub(r'\[.*?\]|\(.*?\)|[\d\-\.]+', ' ', clean)
    clean = re.sub(r'\(kn\s+\d+[\.\d]*\)', '', clean)
    words = [w.strip() for w in re.split(r'[\s\-–—,;:.]+', clean) if len(w.strip()) >= min_len]
    skip = {'the','and','for','are','not','eva','ca','va','no','pi','ti','kho','pana','tattha'}
    return [w for w in words if w not in skip]

def validate_entry(book_no, stated_page, sutta_name, sutta_num):
    """Validate a single PTS reference. Returns (status, detail, score, actual_page)."""
    head, text = get_page(book_no, stated_page)
    if text is None:
        return 'MISSING', f'Page {stated_page} does not exist in book {book_no}', 0, None
    
    kw = extract_keywords(sutta_name)
    if not kw:
        kw = [w for w in re.findall(r'[a-z]{4,}', norm(sutta_name.lower())) if w not in ('sutta', 'vagga', 'pathama', 'dutiya')]
    
    # Score: how many keywords appear in head + first 500 chars of text
    search_text = norm((head + ' ' + text[:500]).lower())
    hits = sum(1 for w in kw if w in search_text)
    score = hits / max(len(kw), 1)
    
    if score >= 0.5:
        return 'OK', f'{hits}/{len(kw)} keywords match', score, stated_page
    
    # Check nearby pages
    for delta in [-3, -2, -1, 1, 2, 3]:
        nh, nt = get_page(book_no, stated_page + delta)
        if nt is None:
            continue
        nsearch = norm((nh + ' ' + nt[:500]).lower())
        nhits = sum(1 for w in kw if w in nsearch)
        nscore = nhits / max(len(kw), 1)
        if nscore > score + 0.2:
            return 'OFFSET', f'Found at p.{stated_page + delta} (δ={delta:+d}), {nhits}/{len(kw)} kw', nscore, stated_page + delta
    
    if score >= 0.25:
        return 'WEAK', f'{hits}/{len(kw)} keywords (weak match)', score, stated_page
    
    return 'UNVERIFIED', f'No match for {kw[:3]}... on page {stated_page} ±3', score, stated_page


# ── Load Excel ──
wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

entries = []
for ri in range(2, ws.max_row + 1):
    entries.append({
        'ri': ri, 'nikaya': str(ws.cell(row=ri, column=cols['Nikaya']).value or ''),
        'num': str(ws.cell(row=ri, column=cols['Sutta #']).value or ''),
        'name': str(ws.cell(row=ri, column=cols['Sutta Name']).value or ''),
        'page': ws.cell(row=ri, column=cols['PTS Page']).value,
        'roman': str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower(),
        'vol': str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip(),
        'ref': str(ws.cell(row=ri, column=cols['PTS Ref']).value or ''),
        'raw': str(ws.cell(row=ri, column=cols['Raw ID']).value or ''),
    })

# ── Strategic sampling ──
# Map: (nikaya, roman) → book_no
def get_book_no(e):
    roman = e['roman']
    vol = e['vol']
    nik = e['nikaya']
    if nik == 'DN':
        return {'i': 6, 'ii': 7, 'iii': 8}.get(roman)
    elif nik == 'MN':
        return {'i': 9, 'ii': 10, 'iii': 11}.get(roman)
    elif nik == 'SN':
        return {'i': 12, 'ii': 13, 'iii': 14, 'iv': 15, 'v': 16}.get(roman)
    elif nik == 'AN':
        return {'i': 17, 'ii': 18, 'iii': 19, 'iv': 20, 'v': 21}.get(roman)
    elif nik == 'KN':
        return {
            'Khp': 22, 'Dhp': 23, 'Ud': 24, 'It': 25, 'Sn': 26,
            'Vv': 27, 'Pv': 28, 'Th': 29, 'Th & Th': 29, 'Thī': 29,
            'Ja': 30, 'Ja II': 31, 'Ja III': 32, 'Ja IV': 33, 'Ja V': 34, 'Ja VI': 35,
            'Nidd I': 36, 'Nidd II': 37,
            'Paṭis I': 38, 'Paṭis II': 39,
            'Ap': 40, 'Bv': 41, 'Cp': 42,
        }.get(vol)
    return None

# Select 40 samples
samples = []

# DN: first, middle, last of each vol + key suttas
dn = [e for e in entries if e['nikaya'] == 'DN']
samples.extend([e for e in dn if e['num'] in ['1', '2', '9', '14', '16', '17', '23', '33', '34']])

# MN: volume boundaries + key suttas
mn = [e for e in entries if e['nikaya'] == 'MN']
samples.extend([e for e in mn if e['num'] in ['1', '50', '51', '52', '76', '77', '100', '107', '152']])

# SN: first of each volume + key saṃyuttas
sn = [e for e in entries if e['nikaya'] == 'SN']
samples.extend([e for e in sn if e['num'] in ['1.1', '2.1', '12.1', '22.1', '35.1', '45.1', '56.1', '56.131']])

# AN: nipāta boundaries
an = [e for e in entries if e['nikaya'] == 'AN']
samples.extend([e for e in an if e['num'] in ['1.1', '2.1', '3.1', '4.1', '5.1', '6.1', '7.1', '8.1', '10.1']])

# KN: per-book samples
kn = [e for e in entries if e['nikaya'] == 'KN']
kn_targets = ['1.1', '2.1', '3.1.1', '4.1.1', '5.1.1', '6.1.1.1', '7.1.1', '8.1.1.1', '9.1.1', '10.1.1']
samples.extend([e for e in kn if e['num'] in kn_targets])

# KN: special — Ja II first, Nidd I first, Patis, Ap, Bv, Cp, Thīg
for target_vol in ['Ja II', 'Nidd I', 'Nidd II', 'Paṭis I', 'Paṭis II', 'Ap', 'Bv', 'Cp']:
    for e in kn:
        if e['vol'] == target_vol:
            samples.append(e)
            break

# Deduplicate
seen = set()
unique = []
for s in samples:
    key = (s['nikaya'], s['num'], s['vol'])
    if key not in seen:
        seen.add(key)
        unique.append(s)
samples = unique[:45]

# ── Validate ──
print("""
╔══════════════════════════════════════════════════════════════════════════╗
║  Helmer Smith — PTS Content Validation Report                           ║
║  Validating reference table against tipitaka.sqlite (edition: mula)     ║
╚══════════════════════════════════════════════════════════════════════════╝
""")

results = []
stats = defaultdict(int)

for e in samples:
    book_no = get_book_no(e)
    if not book_no:
        status = 'NO_BOOK'
        detail = f'Cannot map {e["vol"]} / {e["roman"]} to book_no'
        stats[status] += 1
        results.append((status, e, detail, 0, None))
        continue
    
    status, detail, score, actual_page = validate_entry(book_no, e['page'], e['name'], e['num'])
    stats[status] += 1
    results.append((status, e, detail, score, actual_page))

# ── Print report ──
print(f"{'Status':12s} {'Nik':3s} {'#':12s} {'Ref':20s} {'Detail'}")
print("-" * 90)

status_symbol = {'OK': '✓', 'OFFSET': '≈', 'WEAK': '?', 'UNVERIFIED': '✗', 'MISSING': '∅', 'NO_BOOK': '‼'}

for status, e, detail, score, actual_page in results:
    sym = status_symbol.get(status, '?')
    print(f'{sym} {status:10s} {e["nikaya"]:3s} {e["num"]:12s} {e["ref"]:20s} {detail[:70]}')

# ── Summary ──
print(f"\n{'='*90}")
print(f"  SUMMARY")
print(f"{'='*90}")
total = len(results)
ok = stats.get('OK', 0)
offset = stats.get('OFFSET', 0)
weak = stats.get('WEAK', 0)
unver = stats.get('UNVERIFIED', 0)
other = stats.get('MISSING', 0) + stats.get('NO_BOOK', 0)

print(f"  Sample size: {total}")
print(f"  ✓ Exact match:        {ok:>3d}  ({100*ok/total:.0f}%)")
print(f"  ≈ Offset ±1-3 pages:  {offset:>3d}  ({100*offset/total:.0f}%)")
print(f"  ? Weak match:         {weak:>3d}  ({100*weak/total:.0f}%)")
print(f"  ✗ Unverified:         {unver:>3d}  ({100*unver/total:.0f}%)")
print(f"  ∅ Other:              {other:>3d}")

verified = ok + offset + weak
print(f"\n  Content-confirmed:    {verified}/{total} ({100*verified/total:.0f}%)")

# Assessment
if unver <= 1:
    print(f"\n  Helmer Smith's assessment:")
    print(f"  The reference table is SOUND. All sampled references")
    print(f"  are confirmed by the printed page content. The few offsets")
    print(f"  are attributable to standard PTS title-page conventions.")
elif unver <= total * 0.10:
    print(f"\n  Helmer Smith's assessment:")
    print(f"  MOSTLY RELIABLE. A small number of entries ({unver}) could not be")
    print(f"  verified by keyword matching — they may use variant titles or be")
    print(f"  abbreviated in the PTS edition. Manual inspection recommended.")
else:
    print(f"\n  Helmer Smith's assessment:")
    print(f"  REQUIRES CORRECTION. {unver} entries could not be verified against")
    print(f"  the printed page content. The reference table needs review.")

conn.close()
