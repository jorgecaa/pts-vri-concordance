#!/usr/bin/env python3
"""
AN — add line numbers using sequential position matching within pages.
Sutta markers: "N. text" pattern, numbers reset per vagga.
Vagga markers: "I.", "II.", "III." etc. centered on a line.
"""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

AN_BOOKS = {1: 17, 2: 18, 3: 19, 4: 20, 5: 21}
ROMAN_MAP = {'i': 1, 'ii': 2, 'iii': 3, 'iv': 4, 'v': 5}

def find_markers_on_page(book_no, page_no):
    """Find all sutta start markers and vagga headings on a page.
    Returns list of (line, type, content) where type is 'vagga' or 'sutta'."""
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                (book_no, page_no))
    r = cur.fetchone()
    if not r or not r['unitext']:
        return []
    
    text = r['unitext']
    lines = text.split('\n')
    markers = []
    
    for i, line in enumerate(lines):
        s = line.strip()
        # Vagga heading: Roman numeral alone or with period
        # e.g., "I." "II." "III." "IV." but not "I" in middle of text
        if re.match(r'^[IVX]+\.?\s*$', s) and len(s) <= 8:
            markers.append((i + 1, 'vagga', s))
        # Sutta start: "N." followed by text (not "N." alone)
        elif re.match(r'^\d+\.\s+\S', s):
            num = int(re.match(r'^(\d+)\.', s).group(1))
            # Ignore false positives (year numbers like 2025, etc.)
            if num <= 1000:
                markers.append((i + 1, 'sutta', num))
    
    return markers


# ── Load Excel ──
wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(row=1, column=c).value
    if h:
        cols[h] = c

# Group AN entries by (book_no, page_no)
page_entries = defaultdict(list)
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'AN':
        continue
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower()
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    vol_no = ROMAN_MAP.get(roman, 0)
    book_no = AN_BOOKS.get(vol_no, 0)
    if not book_no or not page:
        continue
    
    sutta_num = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    
    page_entries[(book_no, page)].append({
        'ri': ri, 'num': sutta_num, 'name': name, 'ref': ref, 'roman': roman,
        'page': page, 'book_no': book_no
    })

print(f'AN pages with entries: {len(page_entries)}')
print(f'Total AN entries: {sum(len(v) for v in page_entries.values())}')

# ── Process each page ──
results = {'sutta_marker': 0, 'vagga_next': 0, 'page_top': 0, 'not_found': 0}
fixed = 0

for (book_no, page_no), entries in sorted(page_entries.items()):
    markers = find_markers_on_page(book_no, page_no)
    sutta_markers = [(l, c) for l, t, c in markers if t == 'sutta']
    vagga_markers = [(l, c) for l, t, c in markers if t == 'vagga']
    
    # Assign line numbers to entries
    for ei, entry in enumerate(entries):
        line_num = None
        method = 'not_found'
        
        # Strategy 1: If entry is first on page and vagga heading exists,
        # the sutta starts at the first sutta marker after the vagga
        if ei == 0 and vagga_markers and sutta_markers:
            # Find first sutta marker after the last vagga marker
            vagga_line = vagga_markers[-1][0]
            for sl, sn in sutta_markers:
                if sl > vagga_line:
                    line_num = sl
                    method = 'vagga_next'
                    break
            if not line_num and sutta_markers:
                line_num = sutta_markers[0][0]
                method = 'vagga_next'
        
        # Strategy 2: Sequential match with sutta markers
        if not line_num:
            if ei < len(sutta_markers):
                line_num = sutta_markers[ei][0]
                method = 'sutta_marker'
        
        # Strategy 3: Page top (first non-empty content line)
        if not line_num:
            # Get page text and find first content line
            cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                       (book_no, page_no))
            r = cur.fetchone()
            if r and r['unitext']:
                lines = r['unitext'].split('\n')
                for i, line in enumerate(lines):
                    s = line.strip()
                    if s and len(s) > 5 and not re.match(r'^[A-ZĀĪŪṄÑṬḌṆḶ\s\-\.]+$', s):
                        line_num = i + 1
                        method = 'page_top'
                        break
            if not line_num:
                line_num = 1
                method = 'page_top'
        
        results[method] = results.get(method, 0) + 1
        
        # Update Excel
        if line_num:
            new_ref = 'A %s %d' % (entry['roman'], entry['page'])
            if line_num > 1:
                new_ref += ',%d' % line_num
            
            if new_ref != entry['ref']:
                ws.cell(row=entry['ri'], column=cols['PTS Ref']).value = new_ref
                fixed += 1

wb.save(XL)

print(f'\nResults:')
for k, v in sorted(results.items()):
    print(f'  {k}: {v}')
print(f'Fixed: {fixed} refs updated')
print(f'Saved: {XL}')
conn.close()
