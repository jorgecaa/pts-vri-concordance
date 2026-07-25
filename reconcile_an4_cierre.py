#!/usr/bin/env python3
"""
reconcile_an4_cierre — cierra **A iv (Sattaka + Aṭṭhaka + Navaka)**: las 50 filas pendientes.

Mismo criterio que A ii y A v: donde las dos ediciones eliden, el REJECT del LLM es un falso
negativo y firma la **prueba mecánica**. Cero API. Pero A iv trae dos cosas que no habían salido
antes, y las dos obligan a **no** aplicar la regla general a ciegas.

### ⚠️ 1. La cola del Aṭṭhaka va **+1** respecto al CST

La premisa que sostiene todo el pipeline de AN —«el `Sutta #` del Excel ES el paranum del VRI»,
comprobada sin excepción en las 1099 filas CONFIRMADO de A ii–A v— **deja de cumplirse en la cola
del Aṭṭhaka**. El grupo *Upāsikā* del CST es `<p n="91-116">` (26 miembros) y el Excel lo numera
`8.91-117` (27), y de ahí en adelante todo va corrido uno. Se ve sin ambigüedad en tres filas
consecutivas, cotejando **nombre contra contenido**:

| Excel | dice | el texto es | CST |
|---|---|---|---|
| `8.118` | Aṭṭhamaggaṅga | `sammādiṭṭhi sammāsaṅkappo…` (el óctuple sendero) | **117** |
| `8.119` | Aṭṭhābhibhāyatana | `ajjhattaṃ rūpasaññī bahiddhā rūpāni…` | **118** |
| `8.120` | Aṭṭhavimokkha | `rūpī rūpāni passati…` | **119** |

Así que a esas filas se les escribe la `VRI Ref` **corregida**, no la aritmética. Anotarlo importa:
`check_integrity` no lo habría cazado (los paranums desplazados también existen), y es justo el
punto ciego —un par coherente consigo mismo pero ajeno a la fila.

### ⚠️ 2. El Navaka elide **simétricamente**, y ninguna edición imprime los suttas

`9.74-81` y `9.84-91` no tienen texto en ninguna de las dos ediciones: las dos remiten al vagga
anterior. El CST escribe `yathā satipaṭṭhānavagge tathā sammappadhānavasena vitthāretabbā` y PTS
`(Yāva vaggā sammappadhānavasena vitthārenti.)`. No hay nada que cotejar, y lo que identifica cada
fila es la **coincidencia exacta de los límites del rango en las tres fuentes**: el `<p n="74-81">`
del CST, el marcador de rango `LXXIII-LXXXI.` de PTS (A iv 462) y las ocho filas del Excel.

De paso, esa estructura desmiente dos nombres: el vagga de referencia (CST 63-72, y las propias
filas `9.63-72` del Excel) va `sikkhādubbalya, **nīvaraṇa, kāmaguṇa**, upādānakkhandha…`, y la serie
del *iddhipāda* (`9.84-91`) lo respeta — pero la del *sammappadhāna* trae `9.74 Kāmaguṇa` y
`9.75 Nīvaraṇa` **invertidas**. Se corrigen.

### El resto

- **Sattaka (15)**: 14 con término literal y en orden en las dos ediciones; `7.103-614` cae dentro
  del mismo `<p n="96-622">` del CST y del mismo bloque del impreso, y se firma por aritmética.
- **7 filas de rango** del régimen numerado (`7.11-12`, `7.22-27`, `7.73-74`, `8.21-22`, `8.59-60`,
  `8.68-70`, `9.71-72`): **todos** sus miembros ya estaban CONFIRMADO individualmente.
- **`8.18`** (REVISAR): el par es correcto —`aṭṭhahi ākārehi puriso itthiṃ bandhati`, A iv 197— y el
  rechazo venía de que el marcador abarca también el sutta gemelo.

Uso: python3 reconcile_an4_cierre.py [--dry]
"""
import shutil
import sys
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
STEM = {'7': 's0403m3', '8': 's0404m1', '9': 's0404m2'}

NOMBRES = {                       # (Sutta #, nombre actual) → nombre correcto
    ('9.74', 'Kāmaguṇa sammappadhāna'): 'Nīvaraṇa sammappadhāna',
    ('9.75', 'Nīvaraṇa sammappadhāna'): 'Kāmaguṇa sammappadhāna',
}

# `Sutta #` → `VRI Ref` cuando NO coincide con la aritmética (la cola del Aṭṭhaka, +1)
VRI_CORREGIDA = {
    '8.91-117': 's0404m1:91-116',
    '8.118': 's0404m1:117',
    '8.119': 's0404m1:118',
    '8.120': 's0404m1:119',
    '8.121-147': 's0404m1:120-146',
    '8.148-627': 's0404m1:147-626',
}

_ELIDE_SIM = ('elisión SIMÉTRICA: las dos ediciones remiten al vagga anterior '
              '(CST «yathā satipaṭṭhānavagge…», PTS «Yāva vaggā…»); identifica la fila la '
              'coincidencia exacta de los límites del rango en las tres fuentes')
_LEXICA = 'peyyāla: término distintivo literal y en orden en PTS y en el CST, con la página'

CIERRE = {
    # Sattaka
    **{f'7.{n}': _LEXICA for n in range(85, 96)},
    '7.96-102': _LEXICA,
    '7.103-614': 'mismo `<p n="96-622">` del CST y mismo bloque del impreso; aritmética del rango',
    '7.615-644': _LEXICA,
    '7.645-1124': _LEXICA,
    # Aṭṭhaka (con la VRI Ref corregida, ver arriba)
    '8.91-117': 'grupo Upāsikā = `<p n="91-116">`; el Excel numera 27 donde el CST numera 26',
    '8.118': 'nombre ≡ contenido: el óctuple sendero es el paranum 117, no el 118',
    '8.119': 'nombre ≡ contenido: los ocho abhibhāyatana son el paranum 118',
    '8.120': 'nombre ≡ contenido: los ocho vimokkha son el paranum 119',
    '8.121-147': 'rāga pariññāya = `<p n="120-146">` del CST (la cola del Aṭṭhaka va +1)',
    '8.148-627': 'dosādi peyyāla = `<p n="147-626">` del CST',
    # Navaka
    '9.73': _LEXICA,
    **{f'9.{n}': _ELIDE_SIM for n in list(range(74, 82)) + list(range(84, 92))},
    '9.93': _LEXICA,
    '9.94': _LEXICA,
    '9.95-112': _LEXICA,
    '9.113-432': _LEXICA,
    # rangos del régimen numerado
    **{n: 'rango: todos sus miembros ya CONFIRMADO individualmente'
       for n in ('7.11-12', '7.22-27', '7.73-74', '8.21-22', '8.59-60', '8.68-70', '9.71-72')},
    # arbitraje
    '8.18': 'par correcto (A iv 197, «aṭṭhahi ākārehi puriso itthiṃ bandhati»); el REJECT venía de '
            'que el marcador abarca también el sutta gemelo',
}


def main():
    dry = '--dry' in sys.argv
    if not dry:
        cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-an4cierre.xlsx'
        shutil.copy(XLSX, cp)
        print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    filas = [r for r in range(2, ws.max_row + 1)
             if ws.cell(r, ci['Nikaya']).value == 'AN'
             and str(ws.cell(r, ci['PTS Roman']).value or '').strip() == 'iv']

    for r in filas:
        clave = (str(ws.cell(r, ci['Sutta #']).value), str(ws.cell(r, ci['Sutta Name']).value))
        if clave in NOMBRES:
            print(f'  nombre {clave[0]:>10}: «{clave[1]}» → «{NOMBRES[clave]}»')
            ws.cell(r, ci['Sutta Name']).value = NOMBRES[clave]

    hechas, faltan = Counter(), []
    for r in filas:
        num = str(ws.cell(r, ci['Sutta #']).value)
        if ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        motivo = CIERRE.get(num)
        if not motivo:
            faltan.append(num)
            continue
        if num in VRI_CORREGIDA:
            vri = VRI_CORREGIDA[num]
        else:
            lo, _, hi = num.split('.')[1].partition('-')
            vri = f"{STEM[num.split('.')[0]]}:{lo}" + (f'-{hi}' if hi else '')
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = motivo
        hechas['cerradas'] += 1
    pend = sum(1 for r in filas if ws.cell(r, ci['Estado']).value != 'CONFIRMADO')
    print(f"\n{hechas['cerradas']} filas cerradas · A iv: {len(filas) - pend}/{len(filas)} CONFIRMADO")
    if faltan:
        print(f'  sin regla: {faltan}')
    if dry:
        print('(dry-run: nada escrito)')
        return
    wb.save(XLSX)
    print(f'escrito → {XLSX}')


if __name__ == '__main__':
    main()
