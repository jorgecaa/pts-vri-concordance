#!/usr/bin/env python3
"""Driver SN III (S iii, book_no=14) — pipeline VRI por concordancia + validador (Modelo B).

Diferencias propias de S ii frente a S i / S v:

1. **Marcadores `N (M) Nombre` sin puntos** (ver `sn3_markers.py`, pyparsing), con `N` = nº corrido
   dentro del saṃyutta y `M` = posición en el vagga. La estructura se coteja contra el front matter
   de Feer (`samyutta-vol-III-info.txt`): **13 saṃyuttas (XXII–XXXIV)** y la página de arranque de
   cada uno — cuadra al 100% —, y el nº de suttantas cuadra en **12 de 13** (158/46/10/10/10/10/50/
   46/112/57/55/55); el 13º, Diṭṭhi, lo explica Feer mismo (ver la tabla FEER más abajo).
2. **La clave `(saṃyutta, nº corrido)` NO sirve en este volumen.** PTS imprime 158 suttas en el
   Khandha-saṃyutta (Feer lo dice) mientras el CST/DPR cuenta 159, así que la numeración se desfasa
   +1 en la cola. El lado PTS se resuelve por **nombre del marcador en la página que ancla el
   concordance** (`cst_p_page`; ver `resolve_pts`). El Excel y el concordance SÍ concuerdan entre
   sí — al revés que en S ii, donde el corrupto era el `Sutta #` del Excel.
3. **Los grupos peyyāla van en UNA fila de nombre colectivo** (`Jātisuttādidasakaṃ` = PTS 72–81,
   `Suvaṇṇanikkhasuttādiaṭṭhakaṃ` = 17.13–20, `Pitusuttādichakkaṃ` = 17.38–43), misma convención
   que en SN V: 257 filas cubren los 286 suttas. Por eso `--faltantes` lista suttas sin fila PROPIA
   y eso NO es incompletitud — son los miembros de esos grupos.
4. **Cuando el CST agrupa lo que PTS numera** (antara-peyyālaṃ, SN 12.83–93), el ítem `(N)` del
   grupo CST se casa con la posición `(M)` del marcador PTS: `build_cst_group_items`.

Uso: python3 validador_sn2.py [--dry] [--all] [--n N] [--only 12.1,17.5] [--faltantes]
"""
import csv, os, re, sqlite3, sys, json
import xml.etree.ElementTree as ET
from collections import Counter

import sutta_hash as sh
from openpyxl import load_workbook

from validador import validate_pair
from massive_reader import build_massive as _mr_build
from validador_sn1 import _pts_page, _TEXT_RENDS, build_vri_index as _build_vri
from sn3_markers import find_markers_sn3, collect_suttas

VRI = '/tmp/tipitaka-xml/romn/s0303m.mul.xml'
OUT = 'validador_sn3.json'
DB = 'src/data/tipitaka.sqlite'
SN3_BOOK = 14
XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# Estructura de S iii según el front matter de Feer: (saṃyutta, nombre, nº de suttantas, página).
# Feer distingue TÍTULOS de SUTTANTAS: Nāga son 14 títulos = 50 suttantas, Gandhabba 10 = 112. Aquí
# van los suttantas, que es lo que el marcador PTS numera (los rangos expanden). Su total, 733,
# cuenta Diṭṭhi como 114; el volumen IMPRIME 96 (él mismo confiesa haber omitido 18 del texto) y de
# esos 96 solo 72 llevan marcador: los 72–95 de la 4ª gamana van elididos (PTS salta de 71 a 96).
FEER = [(22, 'Khandha', 158, 1), (23, 'Rādha', 46, 188), (24, 'Diṭṭhi', 96, 202),
        (25, 'Okkantika', 10, 225), (26, 'Uppāda', 10, 228), (27, 'Kilesa', 10, 232),
        (28, 'Sāriputta', 10, 235), (29, 'Nāga', 50, 240), (30, 'Supaṇṇa', 46, 246),
        (31, 'Gandhabba', 112, 249), (32, 'Valāha', 57, 254), (33, 'Vacchagotta', 55, 257),
        (34, 'Jhāna', 55, 263)]


# Cuando el CST agrupa lo que PTS numera, el ítem `(N)` del grupo se casa con la posición `(M)` del
# marcador PTS y el grupo se fija POR TÍTULO (ver `validador_sn2.py`, antara-peyyālaṃ).
CST_GROUP = {}          # sin correspondencias fijadas a mano en este volumen (por ahora)


def build_vri_index():
    return _build_vri(VRI)


def build_cst_by_title(path=VRI):
    """`título del subhead` → texto completo del bloque, para casar por TÍTULO cuando el
    `cst_paranum` del concordance apunta a otro sitio.

    Hace falta en el Jhāna-saṃyutta: sus 10 series van colapsadas (`n=696-701`
    «Kallitamūlakaārammaṇasuttādichakkaṃ», `n=707-710` «Gocaramūlakaabhinīhāra…») y el concordance
    manda varias filas al bloque de la serie **ṭhiti-**. Pero el título del subhead **nombra la
    serie** y es casi idéntico al `Sutta Name` del Excel, así que el título resuelve lo que el
    paranum no.
    """
    body = ET.parse(path).getroot().find('.//body')
    out, title = {}, None
    for el in body.iter():
        if el.tag != 'p':
            continue
        rend = el.get('rend')
        if rend == 'subhead':
            title = ''.join(el.itertext()).strip()
        elif rend in _TEXT_RENDS and title:
            out.setdefault(title, '')
            out[title] += ' ' + ''.join(el.itertext())
    return out


def build_cst_group_items(path=VRI):
    """`(título del grupo, nº de ítem)` → texto, para los grupos CST cuyos miembros van como
    ítems `(N)` dentro de un solo `subhead`.

    Hace falta en el *antara-peyyālaṃ*: PTS imprime `83 (2) Sikkhā` … `93 (12) Appamādo` como 11
    suttas numerados, mientras el CST los mete en `2-12. Sikkhāsuttādipeyyālaekādasakaṃ` con los
    ítems `(2)`…`(12)`. La **posición `(M)` del marcador PTS es el nº de ítem del CST**, así que el
    emparejamiento es exacto y mecánico (no hay que adivinar nada).
    """
    body = ET.parse(path).getroot().find('.//body')
    out, title = {}, None
    for el in body.iter():
        if el.tag != 'p':
            continue
        rend = el.get('rend')
        if rend == 'subhead':
            title = ''.join(el.itertext()).strip()
        elif rend in _TEXT_RENDS and title:
            t = ''.join(el.itertext()).strip()
            m = re.match(r'[‘"\(]*\((\d+)\)', t)
            if m:
                out.setdefault((title, int(m.group(1))), '')
                out[(title, int(m.group(1)))] += ' ' + t
    return out


def build_massive(prefix='sn3.'):
    """`(saṃyutta, nº DPR)` → `(cst_paranum, título CST, página PTS)` con **expansión de rangos**.

    `massive.tsv` da un solo `cst_paranum` por grupo (el del primer miembro), así que sin expandir
    los 17 miembros de `SN24.19-35` se cotejaban todos contra el sutta 1 del saṃyutta. La expansión
    va con compuerta (ver `massive_reader`): solo si TODOS los paranum resultantes existen en el XML
    — que es lo que ocurre cuando el destino cae en el bloque elidido del CST.
    """
    m = _mr_build(prefix, build_vri_index(), pts_page=_pts_page)
    return {k: (v[0], v[1], v[2]) for k, v in m.items()}


def build_pts_suttas(cur):
    """Los 286 suttas de S ii, con `sam` (nº de saṃyutta XII–XXI) y el texto acotado."""
    pages = {r['page_no']: (r['unitext'] or '').split('\n')
             for r in cur.execute('SELECT page_no,unitext FROM pages '
                                  'WHERE edition="mula" AND book_no=?', (SN3_BOOK,))}
    flat, rows = [], []
    for pg in sorted(pages):
        for ln, line in enumerate(pages[pg], start=1):
            flat.append((pg, ln, line))
        for ln, nums, names, tag in find_markers_sn3('\n'.join(pages[pg])):
            rows.append((pg, ln, nums, names, tag))
    suttas, _sams = collect_suttas(rows)
    pos = {(pg, ln): i for i, (pg, ln, _t) in enumerate(flat)}

    # el texto de cada sutta va de su marcador al del siguiente (en orden de lectura)
    ordered = sorted(suttas, key=lambda s: (pos[(s['page'], s['line'])], s['num']))
    out = {}
    for k, s in enumerate(ordered):
        i0 = pos[(s['page'], s['line'])]
        i1 = len(flat)
        for nxt in ordered[k + 1:]:
            j = pos[(nxt['page'], nxt['line'])]
            if j > i0:
                i1 = j
                break
        sam = FEER[s['sam_idx'] - 1][0]
        rec = dict(s)
        rec['sam'] = sam
        rec['text'] = ' '.join(sh.tokens(' '.join(t for _p, _l, t in flat[i0:i1]))[:350])
        out[(sam, s['num'])] = rec
    return out


_ORDINAL = {'pathama': 1, 'paṭhama': 1, 'dutiya': 2, 'tatiya': 3, 'catuttha': 4, 'pancama': 5,
            'pañcama': 5}
_ORD_RE = re.compile(r'^(pa[ṭt]hama|dutiya|tatiya|catuttha|pa[ñn]cama)', re.I)
# nombre compuesto del Jhāna-saṃyutta: «Gocara-mūlaka-abhinīhāra-suttādi-catukkaṃ»
_MULAKA = re.compile(r'^(.+?)m[ūu]lak[aā](.+?)(?:sutt|$)', re.I)


def _ordinal_prefix(name):
    m = _ORD_RE.match((name or '').strip())
    return m.group(1).lower().replace('ṭ', 't').replace('ñ', 'n') if m else ''


def _name_score(a, b):
    """Prefijo común de las raíces, con bonus si una contiene a la otra. Feer abrevia y declina
    («Assādo» vs «Assāda») y marca los repetidos con «(2)» donde el Excel escribe «Dutiya-»."""
    x0, y0 = _ORD_RE.sub('', _stem(a)), _stem(b)
    x, y = re.sub(r'[aiueom]+$', '', x0), re.sub(r'[aiueom]+$', '', y0)
    if not x or not y:
        return 0
    # la contención se mide SIN recortar la terminación de caso: al recortarla «Saññā» queda en
    # «san», por debajo del mínimo, y se perdía su compuesto «Rūpasaññā» (S iii 25.6)
    if min(len(x0), len(y0)) >= 4 and (x0 in y0 or y0 in x0) and x0 != y0:
        return 40 + min(len(x0), len(y0))
    lcp = 0
    for c, d in zip(x, y):
        if c != d:
            break
        lcp += 1
    if x == y:
        return 100
    if x.startswith(y) or y.startswith(x):
        return 50 + lcp
    # Contención, no solo prefijo: Feer antepone palabras («Eso attā» ⊃ CST «Soattā») y al quitar el
    # ordinal el sandhi se come la vocal inicial («Dutiyaabhinivesa» → «bhinivesa» ⊂ «Abhinivesa»).
    if min(len(x), len(y)) >= 4 and (x in y or y in x):
        return 40 + min(len(x), len(y))
    return lcp if lcp >= 4 else 0


DITTHI = 24


def ditthi_pairs(pts, path=VRI):
    """Alineación del **Diṭṭhi-saṃyutta** (SN 24), que no sigue ninguna de las numeraciones.

    Sus 32 filas del Excel están numeradas `24.1`…`24.32`, un índice corrido que **no es DPR**
    (`24.21 «Rūpīattā»` es DPR `SN24.37`) ni PTS. Pero las tres fuentes listan exactamente los
    mismos **32 suttas EXPLÍCITOS** — los que no caen en un bloque elidido — y en el mismo orden:

        Excel 24.1–18   ↔ CST 206–223 (Sotāpattivagga)      ↔ PTS 1–18
        Excel 24.19     ↔ CST 224  (2ª gamana, 1. Vāta)      ↔ PTS `19 (1) Vātā`
        Excel 24.20     ↔ CST 241  (2ª gamana, 18)           ↔ PTS `36 (18)`
        Excel 24.21–28  ↔ CST 242–249 (2ª gamana, 19–26)     ↔ PTS `37 (19)`…`44 (26)`
        Excel 24.29/30  ↔ CST 250 / 275 (3ª gamana, 1 / 26)  ↔ PTS `45 (1)` / `70 (26)`
        Excel 24.31/32  ↔ CST 276 / 301 (4ª gamana, 1 / 26)  ↔ PTS `71 (1)` / `96 (26)`

    Los bloques elididos quedan fuera **en las dos ediciones** y lo dicen igual: el CST con
    `(Purimavagge viya aṭṭhārasa veyyākaraṇāni vitthāretabbānī)` / `(Dutiyavagge viya catuvīsati
    suttāni pūretabbāni)` y PTS con los marcadores de rango `20--35 (2--17)` / `46--69 (2--25)` y el
    salto de `71 (1)` a `96 (26)`.

    Devuelve `[(paranum CST, título CST, sutta PTS)]` en ese orden, para indexar por `nº − 1`.
    """
    body = ET.parse(path).getroot().find('.//body')
    title, cst = None, []
    for el in body.iter():
        if el.tag != 'p':
            continue
        rend = el.get('rend')
        if rend == 'subhead':
            title = ''.join(el.itertext()).strip()
        elif rend in _TEXT_RENDS and el.get('n') and title:
            n = el.get('n')
            if '-' in n:                     # bloque elidido: no es un sutta explícito
                continue
            p0 = int(re.match(r'(\d+)', n).group(1))
            if 206 <= p0 <= 301:
                cst.append((p0, title))
    # lado PTS: los marcadores del saṃyutta que NO son de rango, en orden de lectura
    pt = sorted((q for (sam, _n), q in pts.items() if sam == DITTHI and not q['from_range']),
                key=lambda q: (q['page'], q['line']))
    return list(zip(cst, pt))


def calibrate_offsets(pts, massive, cst_titles):
    """Offset `PTS − CST` por tramos, para los saṃyuttas donde PTS **agrupa o fusiona**.

    En S iii 179 PTS imprime TRES suttas (`146-148 Kulaputtena dukkhā (1)(2)(3)`) donde el CST tiene
    CUATRO (146 Nibbidābahula, 147 Aniccānupassī, 148 Dukkhānupassī, 149 Anattānupassī), así que a
    partir de ahí el último vagga corre con **PTS = CST − 1** (PTS 158 «Ānandena» = CST 159 «Ānanda»).

    El offset no se adivina: se prueba cada desplazamiento pequeño y se acepta el tramo solo si con
    él **casan TODOS los nombres** desde ese punto hasta el final del saṃyutta. Ese acuerdo en bloque
    es la prueba (en el Khandha son 10/10) y hace innecesario el LLM para estas filas.

    Devuelve `{(sam, nº CST): nº PTS}` solo para las claves donde el offset no es 0.
    """
    fix = {}
    for sam in sorted({s for s, _n in pts}):
        nums = sorted(n for s, n in pts if s == sam)
        if not nums:
            continue
        for off in (-1, -2, -3):
            # busca el punto desde el que un offset constante hace casar todos los nombres
            for start in nums:
                run = [k for k in nums if k >= start and (sam, k) in cst_titles]
                if len(run) < 4:
                    continue
                if all(_name_score(cst_titles[(sam, k)], (pts.get((sam, k + off)) or {}).get('name'))
                       >= 40 for k in run):
                    # y con offset 0 el tramo NO casa (si no, no hay nada que corregir)
                    if all(_name_score(cst_titles[(sam, k)], (pts.get((sam, k)) or {}).get('name'))
                           >= 40 for k in run):
                        break
                    for k in run:
                        fix[(sam, k)] = k + off
                    break
            if fix:
                break
    return fix


def resolve_pts(e, h, pts, by_page, offsets=None):
    """Localiza el sutta PTS de una fila de S iii.

    **No se puede usar la clave `(saṃyutta, nº)`**: PTS imprime 158 suttas en el Khandha-saṃyutta
    (así lo dice Feer) mientras el CST/DPR cuenta 159, de modo que la numeración se desfasa +1 en la
    cola y arrastra el resto. Así que se resuelve por **nombre del marcador en la página que ancla
    el concordance** (`cst_p_page`, ±1) — el ancla es independiente del Excel y coincidió en el 98%
    de las filas — y solo si eso falla se cae a la clave o al marcador único de la página.
    """
    # tramo con offset calibrado (PTS agrupa/fusiona): manda el nº, no el nombre
    if offsets and (e['sam'], e['inner']) in offsets:
        q = pts.get((e['sam'], offsets[(e['sam'], e['inner'])]))
        if q:
            return q
    anchor = h[2] if h else None
    base = anchor if anchor is not None else e['page']
    # VÍA 0 — si el marcador del nº que declara el Excel está en la página que declara el Excel y su
    # nombre tiene alguna afinidad, ese manda. Va primero porque el Excel usa compuestos al estilo
    # CST («Rūpasaññā», «Rūpataṇhā») donde Feer imprime solo la cola («Saññā», «Taṇhā»), y buscar
    # por nombre premia el PREFIJO: «Rūpasaññā» casaba con «Rūpa» de la página vecina en vez de con
    # «Saññā», que es el sutta correcto (S iii 25.6).
    # VÍA -1 — nombre COMPUESTO `X-mūlaka-Y` (Jhāna-saṃyutta): el marcador es el compuesto de PTS
    # («Kallita -- ārammaṇa», «Gocara-Abhinīhāra»), no el simple. Buscar por afinidad suelta elige
    # «Kallita» o «Ārammaṇa», que son otros suttas.
    mm = _MULAKA.match(e['name'] or '')
    if mm:
        pa, pb = _stem(mm.group(1)), _stem(mm.group(2))
        for d in (0, 1, -1):
            for q in by_page.get((e['sam'], (e['page'] or 0) + d), []):
                st = _stem(q['name'])
                if len(pa) >= 4 and len(pb) >= 4 and pa[:5] in st and pb[:5] in st:
                    return q

    q0 = pts.get((e['sam'], e['inner']))
    if q0 and isinstance(e['page'], int) and abs(q0['page'] - e['page']) <= 1 \
            and _name_score(e['name'], q0['name']):
        return q0
    # Feer no escribe «Dutiya-»: imprime dos veces el mismo nombre (o le añade «(2)»). Si el
    # nombre del Excel lleva ordinal, hay que tomar el k-ésimo homónimo en orden de lectura, o se
    # coteja el sutta equivocado — es la causa de la mayoría de los REJECT del primer pase.
    ord_k = _ORDINAL.get(_ordinal_prefix(e['name']), 1)
    cands = []
    for d in (0, 1, -1):
        for q in by_page.get((e['sam'], (base or 0) + d), []):
            sc = _name_score(e['name'], q['name'])
            if sc:
                cands.append((abs(d), -sc, q['num'], q))
    if not cands and isinstance(e['page'], int) and e['page'] != base:
        # el ancla y el Excel discrepan: si el nombre casa en la página que declara el Excel, esa
        # manda (S iii 34.3 «Samādhimūlakavuṭṭhāna» está en p265, no en la p273 del ancla)
        for d in (0, 1, -1):
            for q in by_page.get((e['sam'], e['page'] + d), []):
                sc = _name_score(e['name'], q['name'])
                if sc:
                    cands.append((abs(d), -sc, q['num'], q))
    if cands:
        cands.sort(key=lambda c: (c[0], c[1], c[2]))
        top = [c for c in cands if (c[0], c[1]) == (cands[0][0], cands[0][1])]
        if ord_k > 1:
            # los homónimos van en orden de lectura; se amplía la ventana si el k-ésimo cae fuera
            same = sorted({(q['num'], q['page'], q['line']): q for _d, _s, _n, q in cands
                           if _name_score(e['name'], q['name']) >= 50}.values(),
                          key=lambda q: (q['page'], q['line']))
            if len(same) >= ord_k:
                return same[ord_k - 1]
        return top[0][3]
    q = pts.get((e['sam'], e['inner']))
    if q and base and abs(q['page'] - base) <= 1:
        return q
    same = by_page.get((e['sam'], base), [])
    return same[0] if len(same) == 1 else q


def excel_entries(roman='iii'):
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['Nikaya']] or '') != 'SN' or str(row[ci['PTS Roman']] or '').strip().lower() != roman:
            continue
        num = str(row[ci['Sutta #']]); name = str(row[ci['Sutta Name']] or '')
        m = re.match(r'(\d+)\.(\d+)', num)
        sam, inner = (int(m.group(1)), int(m.group(2))) if m else (None, None)
        out.append({'num': num, 'sam': sam, 'inner': inner,
                    'name': re.sub(r'\(SN[^)]*\)', '', name).strip(),
                    'page': row[ci['PTS Page']], 'ref': str(row[ci['PTS Ref']] or ''),
                    'legacy': str(row[ci['Validation']] or '')})
    return out


def main():
    dry = '--dry' in sys.argv
    do_all = '--all' in sys.argv
    n = int(sys.argv[sys.argv.index('--n') + 1]) if '--n' in sys.argv else 30
    only = set(sys.argv[sys.argv.index('--only') + 1].split(',')) if '--only' in sys.argv else None

    vri = build_vri_index()
    massive = build_massive()
    entries = excel_entries()
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    pts = build_pts_suttas(cur)

    print('=' * 92)
    print('SN III — pipeline VRI + estructura de Feer (samyutta-vol-III-info.txt)')
    print('=' * 92)
    print(f'PTS (BD libro 14): {len(pts)} suttas | Feer: {sum(f[2] for f in FEER)}')
    print(f'Excel marcado «S iii»: {len(entries)} filas | paranums en el XML VRI: {len(vri)}')

    mislabel = [e for e in entries if e['sam'] not in dict((f[0], 1) for f in FEER)]
    if mislabel:
        print(f'\n!! {len(mislabel)} filas con saṃyutta ajeno a S iii (volumen mal puesto): '
              f'{", ".join(e["num"] for e in mislabel)}')
    faltan = [(sam, k) for sam, _nm, exp, _pg in FEER for k in range(1, exp + 1)
              if not any(e['sam'] == sam and e['inner'] == k for e in entries)]
    if faltan:
        print(f'!! {len(faltan)} suttas de S iii SIN fila en el Excel: '
              + ', '.join(f'{a}.{b}' for a, b in faltan[:6]) + ' …')
    if '--faltantes' in sys.argv:
        print('\nDetalle de los que faltan (marcador PTS + mapeo CST):')
        for sam, k in faltan:
            p = pts.get((sam, k)); h = massive.get((sam, k))
            print(f'  SN {sam}.{k:<3} PTS: '
                  f'{"p%-3d L%-2d %-26s" % (p["page"], p["line"], (p["name"] or "(sin nombre)")[:26]) if p else "NO HALLADO":38}'
                  f' CST: {h[1][:34] if h else "(no en massive)"}')
        return

    groups = build_cst_group_items()
    by_title = build_cst_by_title()
    by_page = {}
    for (sam, num), q in pts.items():
        by_page.setdefault((sam, q['page']), []).append(q)
    cst_titles = {k: v[1] for k, v in massive.items()}
    ditthi = ditthi_pairs(pts)
    offsets = calibrate_offsets(pts, massive, cst_titles)
    if offsets:
        byc = Counter(k[0] for k in offsets)
        print(f'offsets calibrados (PTS agrupa/fusiona): {sum(byc.values())} claves '
              f'en saṃyuttas {dict(sorted(byc.items()))}')
    tasks, no_cst, name_ok, page_ok = [], [], 0, 0
    for e in entries:
        h = massive.get((e['sam'], e['inner']))
        p = resolve_pts(e, h, pts, by_page, offsets)
        ditthi_hit = None
        if e['sam'] == DITTHI and 1 <= e['inner'] <= len(ditthi):
            # el Diṭṭhi va por POSICIÓN sobre los suttas explícitos (ver `ditthi_pairs`)
            (para, title), q = ditthi[e['inner'] - 1]
            p, ditthi_hit = q, (para, title)
        s = vri.get(h[0]) if h else None
        if ditthi_hit:
            s = vri.get(ditthi_hit[0])
            h = (ditthi_hit[0], ditthi_hit[1], h[2] if h else None)
        elif s is not None and _name_score(e['name'], h[1]) == 0:
            # El paranum apunta a un bloque cuyo título NO tiene NADA que ver con el de la fila
            # (Jhāna-saṃyutta: manda varias series a la de «ṭhiti-»). Solo entonces se busca el
            # subhead que sí la nombra, y se exige coincidencia CASI EXACTA: con un umbral laxo la
            # regla se dispara en decenas de filas cuyo paranum era correcto.
            best = max(((t, _name_score(e['name'], t)) for t in by_title), key=lambda x: x[1],
                       default=(None, 0))
            if best[1] >= 100:
                s = {'title': best[0], 'text': by_title[best[0]]}
        if p and not s and (e['sam'], e['inner']) in CST_GROUP:
            # el CST agrupa lo que PTS numera: el ítem (N) del grupo es la posición (M) del
            # marcador PTS. El grupo se FIJA por título (no se busca): hay varios grupos con los
            # mismos números de ítem y buscar por parecido de texto elige el equivocado.
            gt = CST_GROUP[(e['sam'], e['inner'])]
            txt = groups.get((gt, p['pos']))
            if txt:
                s = {'title': f'{gt} — ítem ({p["pos"]})', 'text': txt}
        if not (p and s):
            no_cst.append(e); continue
        if h and h[2] and abs(h[2] - p['page']) <= 1:
            page_ok += 1
        if _name_score(p['name'], s['title']) or _name_score(p['name'], e['name']):
            name_ok += 1
        tasks.append((e, p, ' '.join(sh.tokens(s['text'])[:350]), s['title']))

    print(f'\nAlineadas con texto PTS+CST: {len(tasks)}/{len(entries)} (sin par: {len(no_cst)})')
    if tasks:
        print(f'  cruzada — página del marcador ≡ cst_p_page (±1): {page_ok}/{len(tasks)} '
              f'({100*page_ok/len(tasks):.0f}%)')
        print(f'  cruzada — raíz del nombre PTS ⊂ título CST: {name_ok}/{len(tasks)} '
              f'({100*name_ok/len(tasks):.0f}%)')
    if no_cst:
        print('  sin par:', ', '.join(e['num'] for e in no_cst[:14]))
    if dry:
        print('\n(dry-run: nada validado, nada escrito)')
        return

    run = [t for t in tasks if t[0]['num'] in only] if only else (tasks if do_all else tasks[:n])
    print(f'\nValidando {len(run)} (Gemini en vivo)...', flush=True)
    rows = []
    for k, (e, p, cst, ctitle) in enumerate(run, 1):
        res = validate_pair(p['text'], cst, e['name'], ctitle, 'SN', concordant=True)
        rows.append({'num': e['num'], 'name': e['name'], 'cst_title': ctitle,
                     'legacy': e['legacy'], 'pts_name': p['name'], 'pts_page': p['page'],
                     'pts_line': p['line'], 'from_range': p['from_range'], **res})
        if k % 25 == 0:
            print(f'  ...{k}/{len(run)}', flush=True)
    prev = json.load(open(OUT)) if os.path.exists(OUT) else []
    merged = {r['num']: r for r in prev}
    merged.update({r['num']: r for r in rows})
    json.dump(list(merged.values()), open(OUT, 'w'), ensure_ascii=False, indent=1)
    print(f'\nEstado: {dict(Counter(r["estado"] for r in rows))} | '
          f'Validation: {dict(Counter(r["validation"] for r in rows))}')
    print(f'Resultados → {OUT} ({len(merged)} en total; nada escrito al Excel).')


_FOLD = str.maketrans({'ā': 'a', 'ī': 'i', 'ū': 'u', 'ṅ': 'n', 'ñ': 'n', 'ṇ': 'n',
                       'ṭ': 't', 'ḍ': 'd', 'ḷ': 'l', 'ṃ': 'm', 'ṁ': 'm'})
def _stem(t):
    n = re.sub(r'[^a-z]', '', re.sub(r'^[\d\s.,()-]*', '', (t or '').lower()).translate(_FOLD))
    n = re.sub(r'(suttantam|suttam|suttani|vaggo).*$', '', n)
    return re.sub(r'(.)\1+', r'\1', n)


if __name__ == '__main__':
    main()
