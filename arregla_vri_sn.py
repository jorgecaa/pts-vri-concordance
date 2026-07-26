#!/usr/bin/env python3
"""
arregla_vri_sn — resuelve las filas de SN/AN que la batería destapó, según criterio filológico.

Todas eran la misma cosa vista por varias caras: **el Excel numera más fino que el CST**, y la clave
no tenía cómo decirlo. Lo que sigue es lo que dicta la práctica lexicográfica: **cada fila apunta al
locus de su texto**, y cuando ese locus es una subdivisión dentro de un paranum, se la nombra.

### La notación nueva: `<fichero>:<paranum>.<ítem>`

No es una invención nuestra: **es la numeración que las dos fuentes ya traen**. El peyyāla de
`s0302m:73` numera sus miembros `(2)`, `(3)` … `(12)` dentro del propio párrafo, y su uddāna los
cuenta —`Satthā sikkhā ca yogo ca, chando ussoḷhipañcamī; … appamādena **dvādasā**`—: doce, que son
`12.82` más las once que estaban sin clave. Es la tercera extensión de la clave VRI, hermana de las
dos de KN (`:c<capítulo>` de la regla (5-bis) y `:c<capítulo>.<ítem>`), y por el mismo motivo: **la
unidad numerada no siempre es el paranum**.

### Lo que se arregla, con su evidencia

| filas | qué eran | a qué pasan |
|---|---|---|
| `SN 12.83`-`12.93` (11) | el peyyāla `Sikkhā`…`Appamāda`, que el CST mete entero en un paranum | `s0302m:73.2` … `:73.12` |
| `SN 46.58`-`46.62` (5) | los **frutos** de la Aṭṭhikasaññā (`aññataraphala`, `mahattha`, `yogakkhema`, `saṃvega`, `phāsuvihāra`), que las **dos** ediciones imprimen dentro de un solo sutta —PTS lo numera `57. (1) Aṭṭhika` y ocupa S v 129-131; el CST lo cierra con `Paṭhamaṃ`— | `s0305m:238.2` … `:238.6` |
| `SN 46.36`, `46.37` (2) | **desplazadas +1**: `Yonisomanasikāra` apuntaba al `Ayoniso` (216) y `Buddhi` al `Yoniso` (217), dejando el 218 sin dueño | 217 y 218 |
| `SN 23.24`, `23.26` (2) | el CST tiene **tres** `Nirodhadhammasuttaṃ` (181, 193, 205) y `23.22` ya tenía el primero; estas dos compartían el rango entero | 193 y 205 |
| `SN 24.28`, `24.30`, `24.32` (3) | tres filas y **tres** `Adukkhamasukhīsuttaṃ` en el CST (249, 275, 301), todas con el rango colapsado | 249, 275, 301 |
| `AN 8.32` (1) | `Dāna 2` sin clave, entre `8.31`→`:31` (`Paṭhamadāna`) y `8.33`→`:33` (`Dānavatthu`) | `s0404m1:32` (`Dutiyadāna`) |

### Lo que NO se toca, y por qué

- **`SN 24.29` y `24.31`** (y sus pares) comparten `s0303m:250-300` y **está bien**: ahí el CST
  repite el título `Navātasuttaṃ` en **cincuenta** paranums seguidos —los vaggas abreviados—, así
  que el rango *es* el locus. Un rango compartido no es un defecto por sí mismo.
- **`SN 24.29` / `24.31`** ya dicho: el rango **es** el locus.
- **`49.1`, `49.3`-`49.5`, `50.1`, `50.3`-`50.6`** apuntaban a un solo paranum dentro de su bloque
  —unas al primero, otras al último, `50.5` a uno de en medio—. Igualadas al **rango que el propio
  CST escribe** en el atributo `n` del bloque.
- ⚠️ Queda un hueco sin fila: el CST salta de `770` a `792` (veintiún paranums) y ninguna fila del
  Excel lo reclama. Es anterior a este arreglo y no se ha tocado.

Uso: python3 arregla_vri_sn.py [--dry]
"""
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# (nikāya, Sutta #) → (VRI Ref nueva, por qué)
ARREGLOS = {}
for k in range(2, 13):                       # el peyyāla de la Nidānasaṃyutta
    ARREGLOS[('SN', f'12.{81 + k}')] = (
        f's0302m:73.{k}',
        f'ítem {k} de los doce que el CST mete en s0302m:73 y que su propio uddāna cuenta '
        f'(«Satthā sikkhā ca yogo ca… appamādena dvādasā»); el texto los numera ({k})')
for k, num in enumerate(('46.58', '46.59', '46.60', '46.61', '46.62'), 2):
    ARREGLOS[('SN', num)] = (
        f's0305m:238.{k}',
        f'ítem {k} del sutta que las dos ediciones imprimen entero: PTS lo numera «57. (1) '
        f'Aṭṭhika» y ocupa S v 129-131, y el CST lo cierra con «Paṭhamaṃ». Esta fila es uno de '
        f'sus frutos, no un sutta aparte')
# ── los cinco bloques de repetición del Bojjhaṅga y los dos huecos reservados ──────────────
#
# ⚠️ Aquí **la cuenta manda sobre el nombre**, y hay que decir por qué. Las dos ediciones parten la
# serie de repeticiones en bloques del mismo tamaño —12, 10, 12, 10, 10 en el Excel contra
# 12, 10, 12, 11, 10 en el CST— pero **atan el nombre a bloques distintos**: el CST llama
# `Punagaṅgānadīādisuttaṃ` al que empieza en 324 y el Excel al que empieza en 312. En una serie de
# repeticiones los nombres son notoriamente resbaladizos —lo dice el propio material, que es el
# mismo sutta repetido con otro término— mientras que la posición y el tamaño no se mueven.
for _n, _r in (('46.88', '312-323'), ('46.89', '324-333'), ('46.90', '334-345'),
               ('46.91', '346-356'), ('46.92', '357-366')):
    ARREGLOS[('SN', _n)] = (
        f's0305m:{_r}',
        f'bloque de repetición del Bojjhaṅga. La secuencia de tamaños casa posición a posición '
        f'(12-10-12-10-10 en el Excel contra 12-10-12-11-10 en el CST); el nombre NO, porque cada '
        f'edición se lo ata a un bloque distinto — el CST llama Punagaṅgānadī al que abre en 324 y '
        f'el Excel al que abre en 312. En una serie de repeticiones el tamaño y la posición son '
        f'firmes y el nombre no')

# Los dos huecos que el CST **reserva y no imprime**: su numeración salta de `651-662` a `673-684`
# y de `705-716` a `749-758`, dejando exactamente los diez y los treinta y dos que PTS sí imprime.
# Es el `deest` del aparato, pero del lado del CST: la numeración es común, el texto falta en ese
# testigo. La referencia es correcta como **locus**; lo que no hay es texto que cotejar.
ARREGLOS[('SN', '49.2')] = (
    's0305m:663-672',
    'SN 49.13-22 (Appamādavagga). El CST **reserva estos números y no imprime el texto**: sus '
    'grupos saltan de 651-662 a 673-684. La referencia es el locus; el texto es «deest in Be»')
ARREGLOS[('SN', '50.2')] = (
    's0305m:717-748',
    'SN 50.13-22. El CST **reserva estos números y no imprime el texto**: sus grupos saltan de '
    '705-716 a 749-758. La referencia es el locus; el texto es «deest in Be»')

# ── rangos consistentes en el Sammappadhāna y el Bala ─────────────────────────────────────
#
# Estas siete apuntaban a **un solo paranum dentro de su bloque** —unas al primero, otras al
# último, `50.5` a uno de en medio—. No estaban mal: el paranum existe y cae dentro. Pero sus
# hermanas llevan rango y **el locus de un grupo es el grupo**, así que se igualan. El rango no se
# calcula: es el que el propio CST escribe en el atributo `n` de cada bloque (`n="673-684"`), y el
# nombre del bloque casa con el de la fila uno a uno.
for _n, _r in (('49.1', '651-662'), ('49.3', '673-684'), ('49.4', '685-694'), ('49.5', '695-704'),
               ('50.1', '705-716'), ('50.3', '749-758'), ('50.4', '759-770'),
               ('50.5', '792-802'), ('50.6', '803-812')):
    ARREGLOS[('SN', _n)] = (
        f's0305m:{_r}',
        f'rango del bloque, tal como el CST lo escribe en el atributo `n` de ese grupo; el nombre '
        f'del bloque casa con el de la fila. Antes apuntaba a un solo paranum de dentro, que no '
        f'estaba mal pero no era el locus del grupo')

ARREGLOS.update({
    ('SN', '46.36'): ('s0305m:217', 'estaba desplazada +1: apuntaba al Ayonisomanasikāra (216) '
                                    'siendo el Yonisomanasikāra, que es el 217'),
    ('SN', '46.37'): ('s0305m:218', 'estaba desplazada +1: apuntaba al Yonisomanasikāra (217) '
                                    'siendo el Buddhi, que es el 218'),
    ('SN', '23.24'): ('s0303m:193', 'el CST tiene tres Nirodhadhammasuttaṃ (181, 193, 205) y '
                                    '23.22 ya tenía el primero; ésta es la segunda'),
    ('SN', '23.26'): ('s0303m:205', 'el CST tiene tres Nirodhadhammasuttaṃ (181, 193, 205); '
                                    'ésta es la tercera'),
    ('SN', '24.28'): ('s0303m:249', 'tres filas y tres Adukkhamasukhīsuttaṃ en el CST '
                                    '(249, 275, 301); ésta es la primera'),
    ('SN', '24.30'): ('s0303m:275', 'ídem, la segunda'),
    ('SN', '24.32'): ('s0303m:301', 'ídem, la tercera'),
    ('AN', '8.32'): ('s0404m1:32', 'entre 8.31 → :31 (Paṭhamadānasuttaṃ) y 8.33 → :33 '
                                   '(Dānavatthusuttaṃ) está el :32, Dutiyadānasuttaṃ'),
})


def main():
    dry = '--dry' in sys.argv
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    hechos, faltan = [], dict(ARREGLOS)
    for r in range(2, ws.max_row + 1):
        clave = (str(ws.cell(r, ci['Nikaya']).value), str(ws.cell(r, ci['Sutta #']).value))
        if clave not in ARREGLOS:
            continue
        vri, porque = ARREGLOS[clave]
        antes = str(ws.cell(r, ci['VRI Ref']).value or '—')
        hechos.append((clave, antes, vri))
        faltan.pop(clave, None)
        if not dry:
            ws.cell(r, ci['VRI Ref']).value = vri
            ws.cell(r, ci['Detail']).value = f'VRI Ref corregida ({antes} → {vri}): {porque}'
    for (nk, n), antes, vri in hechos:
        print(f'   {nk} {n:9} {antes:18} → {vri}')
    print(f'\n{len(hechos)}/{len(ARREGLOS)} filas')
    if faltan:
        print(f'⚠ no encontradas: {sorted(faltan)}')
        return
    if dry:
        print('(dry-run: nada escrito)')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-vrisn.xlsx'
    shutil.copy(XLSX, cp)
    wb.save(XLSX)
    print(f'backup → {cp}\nescrito → {XLSX}')


if __name__ == '__main__':
    main()
