#!/usr/bin/env python3
"""
Parse the PTS Reference blog HTML using BeautifulSoup and save to Excel.

Source: https://palistudies.blogspot.com/2020/02/sutta-number-to-pts-reference-converter.html
"""

import re
import html as html_mod
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Read HTML
with open('/home/jorge/Code/squashfs-root/pts_full_table.html', 'r', encoding='utf-8') as f:
    html_content = f.read()

soup = BeautifulSoup(html_content, 'html.parser')

# Find all table rows in the main content
# The data is in <tr> elements with 3 <td> children
rows = []

# Find all tables and their rows
for table in soup.find_all('table'):
    for tr in table.find_all('tr'):
        tds = tr.find_all('td')
        if len(tds) == 3:
            # Get text from each td
            sutta_text = tds[0].get_text(strip=True)
            type_text = tds[1].get_text(strip=True)
            ref_text = tds[2].get_text(strip=True)
            
            # Skip header/navigation rows
            if not sutta_text or 'Sutta Number' in sutta_text:
                continue
            if 'PTS Vol' in sutta_text:
                continue
            if 'Go to' in sutta_text or 'New!' in sutta_text or 'Quick Jump' in sutta_text:
                continue
            if 'Dīgha Nikāya' in sutta_text or 'Majjhima Nikāya' in sutta_text:
                continue
            if 'Saṁyutta Nikāya' in sutta_text or 'Aṅguttara Nikāya' in sutta_text:
                continue
            if 'Use the' in sutta_text or 'search boxes' in sutta_text:
                continue
            if sutta_text.startswith('Search'):
                continue
            if 'Labels' in sutta_text or 'Share' in sutta_text:
                continue
            if '©' in sutta_text:
                continue
            if len(sutta_text) < 3:
                continue
                
            # Clean up Unicode
            sutta_text = html_mod.unescape(sutta_text)
            type_text = html_mod.unescape(type_text)
            ref_text = html_mod.unescape(ref_text)
            
            # Normalize whitespace
            sutta_text = re.sub(r'\s+', ' ', sutta_text).strip()
            type_text = re.sub(r'\s+', ' ', type_text).strip()
            ref_text = re.sub(r'\s+', ' ', ref_text).strip()
            
            rows.append((sutta_text, type_text, ref_text))

print(f"Found {len(rows)} rows from BeautifulSoup parsing")

# Parse each row
parsed_rows = []

for sutta_raw, type_raw, ref_raw in rows:
    # Extract alternate PTS ref if present
    # Format: "D i 1" or "S i 1 [S i 1]" or "S i 1 {S i 1}" or "S i 1 (alt)"
    alt_pts = ''
    ref_main = ref_raw
    
    # Check for brackets/parens with alternate ref
    alt_match = re.search(r'[\[\(\{]([SDA]\s+[ivxlcdm]+\s+\d+)[\]\)\}]', ref_raw)
    if alt_match:
        alt_pts = alt_match.group(1).strip()
        ref_main = ref_raw[:alt_match.start()].strip()
    
    # Parse PTS volume/page from main ref
    pts_vol_letter = ''
    pts_vol_roman = ''
    pts_page = ''
    
    vol_match = re.match(r'([DMAS])\s+(i{1,3}|iv|v|vi{0,3})\s+(\d+)', ref_main)
    if vol_match:
        pts_vol_letter = vol_match.group(1)
        pts_vol_roman = vol_match.group(2)
        pts_page = int(vol_match.group(3))
    
    # Determine nikaya and parse sutta number
    nikaya = ''
    sutta_num = ''
    sutta_name = ''
    cst_num = ''
    
    # Clean the sutta text
    sutta_clean = sutta_raw.strip()
    
    # Fix known typos: "N 4.260 AN 4.257" → "AN 4.260"
    if sutta_clean.startswith('N '):
        sutta_clean = re.sub(r'^N\s+', '', sutta_clean)
    # Fix "4.260 AN 4.257" → "AN 4.260"
    if re.match(r'^\d', sutta_clean):
        sutta_clean = re.sub(r'^[\d.]+\s*', '', sutta_clean).strip()
    
    if sutta_clean.startswith('DN '):
        nikaya = 'DN'
        rest = sutta_clean[3:]
    elif sutta_clean.startswith('MN '):
        nikaya = 'MN'
        rest = sutta_clean[3:]
    elif sutta_clean.startswith('SN '):
        nikaya = 'SN'
        rest = sutta_clean[3:]
    elif sutta_clean.startswith('AN '):
        nikaya = 'AN'
        rest = sutta_clean[3:]
    else:
        nikaya = '??'
        rest = sutta_clean
    
    # Parse number and name based on nikaya format
    if nikaya in ('DN', 'MN'):
        # DN 1: Brahmajāla
        m = re.match(r'(\d+):?\s*(.*)', rest)
        if m:
            sutta_num = m.group(1)
            sutta_name = m.group(2).strip()
    elif nikaya == 'SN':
        # SN 1.1 (SN 1) Oghataraṇa  OR  SN 12.1 Paṭiccasamuppāda
        m = re.match(r'([\d.]+)\s*(?:\(SN\s*\d+\)\s*)?(.*)', rest)
        if m:
            sutta_num = m.group(1)
            sutta_name = m.group(2).strip()
        # Extract CST number
        cst_m = re.search(r'\(SN\s*(\d+)\)', sutta_clean)
        if cst_m:
            cst_num = cst_m.group(1)
    elif nikaya == 'AN':
        # AN 1.1 [AN 1.1.1]: Cittapariyādāna [Rūpādi] 1
        m = re.match(r'([\d.]+)\s*(?:\[AN\s*[\d.]+\][:\]]?\s*)?(.*)', rest)
        if m:
            sutta_num = m.group(1)
            sutta_name = m.group(2).strip()
        # Extract CST number
        cst_m = re.search(r'\[AN\s*([\d.]+)\]', sutta_clean)
        if cst_m:
            cst_num = cst_m.group(1)
    
    parsed_rows.append({
        'nikaya': nikaya,
        'sutta_num': sutta_num,
        'sutta_name': sutta_name,
        'cst_num': cst_num,
        'sutta_raw': sutta_clean,
        'pts_vol_letter': pts_vol_letter,
        'pts_vol_roman': pts_vol_roman,
        'pts_page': pts_page,
        'pts_full': ref_main,
        'pts_alt': alt_pts,
        'type': type_raw if type_raw else 'Sutta'
    })

print(f"Parsed {len(parsed_rows)} rows")

# Deduplicate (same sutta_raw and pts_full)
seen = set()
deduped = []
for row in parsed_rows:
    key = (row['sutta_raw'], row['pts_full'])
    if key not in seen:
        seen.add(key)
        deduped.append(row)

print(f"After dedup: {len(deduped)} rows (removed {len(parsed_rows) - len(deduped)})")

parsed_rows = deduped

# =============================================
# Create Excel
# =============================================
wb = Workbook()

# Styles
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_align = Alignment(vertical='center', wrap_text=False)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
nikaya_fills = {
    'DN': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'MN': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
    'SN': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'AN': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
}

# ---- Sheet 1: Full Reference Table ----
ws1 = wb.active
ws1.title = "PTS Reference"

headers = ['#', 'Nikaya', 'Sutta #', 'Sutta Name', 'CST #', 
           'PTS Vol', 'PTS Roman', 'PTS Page', 'PTS Full Ref', 
           'PTS Alt (Somaratne)', 'Type', 'Raw Sutta ID']

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for i, row in enumerate(parsed_rows, 2):
    values = [
        i - 1,
        row['nikaya'],
        row['sutta_num'],
        row['sutta_name'],
        row['cst_num'],
        row['pts_vol_letter'],
        row['pts_vol_roman'],
        row['pts_page'],
        row['pts_full'],
        row['pts_alt'],
        row['type'],
        row['sutta_raw']
    ]
    
    fill = nikaya_fills.get(row['nikaya'])
    
    for col, val in enumerate(values, 1):
        cell = ws1.cell(row=i, column=col, value=val)
        cell.border = thin_border
        cell.alignment = data_align
        if fill:
            cell.fill = fill

# Column widths
col_widths = [5, 8, 12, 48, 8, 8, 12, 10, 18, 22, 7, 55]
for col, width in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(col)].width = width

ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(parsed_rows)+1}"

# ---- Sheet 2: Summary ----
ws2 = wb.create_sheet("Summary")
ws2_headers = ['Nikaya', 'Entries', 'PTS Volume Range', 'Page Range', 'Sutta # Range']
for col, h in enumerate(ws2_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

nikaya_stats = defaultdict(lambda: {'count': 0, 'vols': set(), 'pages': [], 'nums': []})

for row in parsed_rows:
    n = row['nikaya']
    nikaya_stats[n]['count'] += 1
    if row['pts_vol_letter'] and row['pts_vol_roman']:
        nikaya_stats[n]['vols'].add(f"{row['pts_vol_letter']} {row['pts_vol_roman']}")
    if row['pts_page']:
        nikaya_stats[n]['pages'].append(row['pts_page'])
    if row['sutta_num']:
        try:
            nikaya_stats[n]['nums'].append(float(row['sutta_num']))
        except ValueError:
            pass

for i, (nikaya, stats) in enumerate(sorted(nikaya_stats.items()), 2):
    vols_str = ', '.join(sorted(stats['vols']))
    if stats['pages']:
        page_range = f"{min(stats['pages'])} – {max(stats['pages'])}"
    else:
        page_range = ''
    if stats['nums']:
        num_range = f"{min(stats['nums'])} – {max(stats['nums'])}"
    else:
        num_range = ''
    
    values = [nikaya, stats['count'], vols_str, page_range, num_range]
    fill = nikaya_fills.get(nikaya)
    for col, val in enumerate(values, 1):
        cell = ws2.cell(row=i, column=col, value=val)
        cell.border = thin_border
        if fill:
            cell.fill = fill

for col, w in enumerate([8, 12, 35, 15, 15], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# ---- Sheet 3: Quick Jump ----
ws3 = wb.create_sheet("Quick Jump")
ws3_headers = ['Nikaya', 'Description', 'PTS Volumes']
for col, h in enumerate(ws3_headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

quick = [
    ('DN', 'Dīgha Nikāya (34 suttas)', 'D i, D ii, D iii'),
    ('MN', 'Majjhima Nikāya (152 suttas)', 'M i, M ii, M iii'),
    ('SN', 'Saṁyutta Nikāya (56 saṁyuttas)', 'S i, S ii, S iii, S iv, S v'),
    ('AN', 'Aṅguttara Nikāya (11 nipātas)', 'A i, A ii, A iii, A iv, A v'),
]
for i, (n, desc, vols) in enumerate(quick, 2):
    fill = nikaya_fills.get(n)
    for col, val in enumerate([n, desc, vols], 1):
        cell = ws3.cell(row=i, column=col, value=val)
        cell.border = thin_border
        if fill:
            cell.fill = fill

for col, w in enumerate([8, 38, 38], 1):
    ws3.column_dimensions[get_column_letter(col)].width = w

# ---- Save ----
output = '/home/jorge/Code/squashfs-root/PTS_Reference_Full_Canon.xlsx'
wb.save(output)
print(f"\nSaved: {output}")

# ---- Report ----
print("\n=== Nikaya Stats ===")
for n in ['DN', 'MN', 'SN', 'AN', '??']:
    subset = [r for r in parsed_rows if r['nikaya'] == n]
    if subset:
        first = subset[0]
        last = subset[-1]
        print(f"\n{n}: {len(subset)} entries")
        print(f"  First: {first['sutta_raw'][:70]}")
        print(f"  Last:  {last['sutta_raw'][:70]}")
        print(f"  PTS range: {first['pts_full']} – {last['pts_full']}")

# Check for issues
missing_pts = [r for r in parsed_rows if not r['pts_vol_letter']]
if missing_pts:
    print(f"\n⚠ Missing PTS refs ({len(missing_pts)}):")
    for r in missing_pts[:10]:
        print(f"  {r['sutta_raw'][:60]} -> '{r['pts_full']}'")

unparsed = [r for r in parsed_rows if r['nikaya'] == '??']
if unparsed:
    print(f"\n⚠ Unparsed ({len(unparsed)}):")
    for r in unparsed[:10]:
        print(f"  {r['sutta_raw'][:80]} -> {r['pts_full']}")

# Count with PTS alt
with_alt = [r for r in parsed_rows if r['pts_alt']]
print(f"\nEntries with alternate PTS ref (Somaratne): {len(with_alt)}")
print(f"Total entries: {len(parsed_rows)}")

print("\nDone!")
