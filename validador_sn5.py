#!/usr/bin/env python3
"""Driver SN V para el validador (Modelo B).

Resolvedor CST específico de SN V (book 16, saṃyuttas 45–56, s5m.xml), portado de
helmer_sn5_all.py: parsea los marcadores PTS `N. (gid) nombre`, extrae el texto del
sutta hasta el siguiente marcador, y mapea al CST por fronteras de saṃyutta
(cst_idx = frontera[saṃyutta] + (inner-1)).  OJO: CST≠PTS numbering en SN — si el
mapeo desalinea, el Modelo B lo marca como desacuerdo/REVISAR, no lo auto-confirma.

Uso: python3 validador_sn5.py [--n 30] [--all]
No escribe el Excel; vuelca a validador_sn5_batch.json.
"""
import re, sqlite3, sys, json
from collections import defaultdict, Counter
import sutta_hash as sh
from openpyxl import load_workbook
from validador import validate_pair, _nfold

SN5_BOOK = 16
DB = 'src/data/tipitaka.sqlite'
ALIGN_WINDOW = 30     # segmentos CST hacia delante a considerar por sutta PTS
ALIGN_MIN = 0.30      # cobertura mínima para aceptar una alineación


def _nset(s, n=150):
    return {_nfold(x) for x in sh.tokens(s)[:n]}

# ordinales pāli → distinguir paṭhama/dutiya/tatiya… en series repetitivas
_ORD = {'paṭhama': 1, 'dutiya': 2, 'tatiya': 3, 'catuttha': 4, 'pañcama': 5, 'chaṭṭha': 6,
        'sattama': 7, 'aṭṭhama': 8, 'navama': 9, 'dasama': 10, 'ekādasama': 11, 'dvādasama': 12}
def ordinal_of(s):
    t = re.sub(r'^\s*[\d\-]+\.\s*', '', (s or '').lower())   # quita "4. " / "2-6. "
    for w, k in _ORD.items():
        if t.startswith(w):
            return k
    return None
def is_range(title):
    return bool(re.match(r'^\s*\d+\s*-\s*\d+\.', title or ''))


def align(pts_items, cst_items):
    """Alineación monótona por cobertura con consistencia de ordinal y avance
    rango-aware. pts_items: [(key, pts_text, name)]; cst_items: [(text, title)].
    Devuelve [(key, cst_idx|None, cov)]."""
    cst_sets = [_nset(t) for t, _ in cst_items]
    cst_ord = [ordinal_of(title) for _, title in cst_items]
    cst_rng = [is_range(title) for _, title in cst_items]
    out, ptr = [], 0
    for key, pts, name in pts_items:
        A = _nset(pts)
        pord = ordinal_of(name)
        best_cov, best_idx = -1.0, None
        for j in range(ptr, min(ptr + ALIGN_WINDOW, len(cst_items))):
            if pord and cst_ord[j] and pord != cst_ord[j]:
                continue                          # ordinal en conflicto → no es este miembro
            cov = len(A & cst_sets[j]) / max(1, len(A))
            if cov > best_cov:
                best_cov, best_idx = cov, j
        if best_idx is not None and best_cov >= ALIGN_MIN:
            out.append((key, best_idx, round(best_cov, 2)))
            ptr = best_idx                        # re-ancla (reuso permitido para grupos)
        else:
            out.append((key, None, round(max(best_cov, 0), 2)))  # no mover el puntero
    return out


def load_entries():
    wb = load_workbook('PTS_Reference_Complete_Canon.xlsx', read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['Nikaya']] or '') != 'SN':
            continue
        snum = str(row[ci['Sutta #']] or '')
        parts = snum.split('.')
        if len(parts) < 2 or not parts[0].isdigit():
            continue
        if not (45 <= int(parts[0]) <= 56):
            continue
        if str(row[ci['PTS Roman']] or '').strip().lower() != 'v':
            continue
        out.append({'num': snum, 'ref': str(row[ci['PTS Ref']] or ''),
                    'name': str(row[ci['Sutta Name']] or ''), 'page': row[ci['PTS Page']],
                    'sn': int(parts[0]), 'sn_inner': int(parts[1]),
                    'legacy': str(row[ci['Validation']] or '')})
    out.sort(key=lambda e: (e['sn'], e['sn_inner']))
    return out


_OCR_D = str.maketrans({'O': '0', 'o': '0', 'l': '1', 'I': '1'})
def _fix_ocr_num(s):
    """Repara el nº del marcador cuando el OCR leyó letras por dígitos ("4O. (10)
    Nandiya." en S v 397). Solo toca el prefijo numérico de una línea que ya tiene
    forma de marcador `NN. (NN) Nombre`, nunca el texto."""
    m = re.match(r'^([0-9OolI]{1,4})(\s*-+\s*[0-9OolI]{1,4})?(\.?\s*\()([0-9OolI]{1,3})', s)
    if not m or (m.group(1).isdigit() and m.group(4).isdigit()):
        return s
    head = m.group(0).translate(_OCR_D)
    return head + s[m.end():]


def build_markers(cur):
    mm = defaultdict(list)
    cur.execute('SELECT page_no,unitext FROM pages WHERE book_no=? AND edition="mula"', (SN5_BOOK,))
    for r in cur.fetchall():
        for i, line in enumerate((r['unitext'] or '').split('\n')):
            s = _fix_ocr_num(line.strip())
            lead = len(line) - len(line.lstrip())
            # marcador "48. (8) Nombre", rango "103--108. (1--6) Pācīna", grupo sin
            # nombre "63--72. (1--10)."
            m = re.match(r'^(\d+)(?:\s*-+\s*(\d+))?\.?\s*\((\d+)(?:\s*-+\s*(\d+))?\)\.?\s*(\S.*)?$', s)
            if not m:
                # grupo profundo sin paréntesis: "89--98.1--10."
                m = re.match(r'^(\d+)\s*-+\s*(\d+)\.\s*(\d+)(?:\s*-+\s*\d+)?\.\s*$', s)
            if m:
                a = int(m.group(1)); b = int(m.group(2)) if m.group(2) else a
                nm = (m.group(5) if m.lastindex >= 5 else '') or ''
                for run in range(a, b + 1):        # registra cada nº corrido del rango
                    mm[r['page_no']].append((i + 1, run, int(m.group(3)), nm[:60]))
                continue
            # marcador CENTRADO simple sin paréntesis: "5. Bhikkhu.", "1. Rājā.",
            # "7. Kūṭāgāra." — discriminado por la sangría (>=8) frente a los nº de
            # sección del texto (margen izquierdo). El nombre no cruza separadores ║.
            if lead >= 8:
                m2 = re.match(r'^(\d+)(?:\s*-+\s*(\d+))?\.\s*([A-ZĀĪŪṄÑṆṬḌḶ][^\d║]*?)\d?\s*\.?\s*$', s)
                if m2 and m2.group(3) and len(m2.group(3).strip()) >= 3:
                    a = int(m2.group(1)); b = int(m2.group(2)) if m2.group(2) else a
                    nm = m2.group(3).strip()
                    for run in range(a, b + 1):
                        mm[r['page_no']].append((i + 1, run, a, nm[:60]))
    return mm


# Los suttas peyyāla de PTS son legítimamente cortísimos ("virāgasaññā bhikkhave pe"),
# así que el umbral solo descarta el marcador SIN cuerpo (su propio título y nada más).
MIN_PTS_TOKENS = 3


def _sutta_text(cur, page, line):
    """Texto del sutta que arranca en `line` de `page`, cortado en el marcador
    siguiente. Devuelve None si sale un fragmento demasiado corto para cotejar
    (marcadores de rango consecutivos: el "sutta" sale reducido a su propio título)."""
    r = cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                    (SN5_BOOK, page)).fetchone()
    if not r:
        return None
    lines = (r['unitext'] or '').split('\n')
    start, end = line - 1, len(lines)
    for j in range(start + 1, len(lines)):
        if re.match(r'^\d+(?:\s*-+\s*\d+)?\.?\s*\(', _fix_ocr_num(lines[j].strip())):
            end = j
            break
    toks = sh.tokens(' '.join(lines[start:end]))
    return ' '.join(toks[:350]) if len(toks) >= MIN_PTS_TOKENS else None


def pts_for(cur, page, target, mm, span=1):
    """Texto PTS del nº corrido `target`. Busca en la página declarada y, si no está,
    en las adyacentes (el off-by-one de página documentado en PTS vol. V)."""
    for pg in [page] + [page + d for d in range(1, span + 1)] + [page - d for d in range(1, span + 1)]:
        for line, gid, vpos, name in mm.get(pg, []):
            if gid == target:
                t = _sutta_text(cur, pg, line)
                if t:
                    return t
    return None


_FOLD_N = str.maketrans({'ā': 'a', 'ī': 'i', 'ū': 'u', 'ṅ': 'n', 'ñ': 'n', 'ṇ': 'n',
                         'ṭ': 't', 'ḍ': 'd', 'ḷ': 'l', 'ṃ': 'm', 'ṁ': 'm'})
def _norm_name(t):
    t = re.sub(r'^[\d\-\s]*\.?\s*', '', (t or '').lower()).translate(_FOLD_N)
    return re.sub(r'[^a-z]', '', t)

def _stem(t):
    """Raíz del nombre: quita el sufijo sutta/vagga y la terminación de caso (PTS
    "Ogho" vs CST "Oghasuttaṃ"). El sufijo se quita ANTES de colapsar consonantes
    dobles — al revés, "suttaṃ" ya se ha vuelto "sutam" y no lo reconoce."""
    n = re.sub(r'(suttantam|suttam|suttani|vaggo).*$', '', _norm_name(t))
    return re.sub(r'[aiueom]+$', '', re.sub(r'(.)\1+', r'\1', n))


def _name_score(tgt, ns):
    """Afinidad entre la raíz del título CST y la del nombre del marcador PTS.
    4 = idéntico; 3 = uno es prefijo del otro ("Ogho" ⊂ "Oghasuttaṃ"); 2 = uno está
    contenido en el otro ("Khandā" ⊂ "Upādānakkhandha"); 1 = prefijo común largo
    (flexión divergente: "Nivaraṇāni" vs "Nīvaraṇasuttaṃ"). 0 = no casan."""
    if not ns or len(ns) < 3:
        return 0
    if tgt == ns:
        return 4
    if tgt.startswith(ns) or ns.startswith(tgt):
        return 3
    if len(ns) >= 5 and (ns in tgt or tgt in ns):
        return 2
    lcp = 0
    for a, b in zip(tgt, ns):
        if a != b:
            break
        lcp += 1
    return 1 if lcp >= 7 and lcp >= 0.7 * min(len(tgt), len(ns)) else 0


def pts_by_name(cur, page, cst_title, mm, span=1, min_score=1, require_unique=False):
    """Resuelve el texto PTS casando el NOMBRE del título CST contra el nombre del
    marcador PTS (robusto al off-by-one de numeración de las vaggas peyyāla).
    Elige el MEJOR candidato, no el primero: la página declarada gana a las vecinas
    y la coincidencia más fuerte gana a la más laxa.

    `min_score` fija la fuerza exigida (ver `_name_score`). Los nombres LAXOS (2, 1)
    no distinguen los pares paṭhama/dutiya de las series peyyāla — el marcador PTS
    los llama igual ("Pathavī1." / "Pathavī2.") — así que el nombre solo debe
    anteponerse al nº corrido cuando la coincidencia es fuerte (>=3), o cuando es
    INEQUÍVOCA: con `require_unique` se exige un único marcador candidato en la
    ventana, y entonces no hay homónimo que confundir."""
    tgt = _stem(cst_title)
    if not tgt or len(tgt) < 3:
        return None
    pages = [page] + [page + d for d in range(1, span + 1)] + [page - d for d in range(1, span + 1)]
    cands = []                                    # (score, dist_pagina, pg, line)
    for dist, pg in enumerate(pages):
        seen = set()
        for line, gid, vpos, name in mm.get(pg, []):
            if line in seen:
                continue
            seen.add(line)
            sc = _name_score(tgt, _stem(name))
            if sc >= min_score:
                cands.append((sc, dist, pg, line))
    if require_unique and len(cands) != 1:
        return None
    # score fuerte primero; a igual score, la página declarada antes que las vecinas y,
    # dentro de la página, el marcador de MÁS ARRIBA (los homónimos de una serie
    # peyyāla van en orden: "Duggatibhaya" antes que "Duggativinipātabhaya").
    for sc, _d, pg, line in sorted(cands, key=lambda c: (-c[0], c[1], c[3])):
        t = _sutta_text(cur, pg, line)
        if t:
            return t
    return None


def pts_page(cur, page, span=0):
    """Fallback final: texto de la(s) página(s) PTS completas. Para los grupos
    peyyāla donde el marcador no aísla el sutta pero la concordancia SÍ garantiza
    la alineación (la página PTS es la que afirma el propio Excel)."""
    if not isinstance(page, int):
        return None
    parts = []
    for pg in [page] + [page + d for d in range(1, span + 1)]:
        r = cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (SN5_BOOK, pg)).fetchone()
        if r and r['unitext']:
            parts.append(r['unitext'])
    return ' '.join(sh.tokens(' '.join(parts))[:350]) if parts else None


def _ctoks(t, n=40):
    return {re.sub(r'(.)\1+', r'\1', x.translate(_FOLD_N)) for x in sh.tokens(t)[:n]}

def pts_by_content(cur, page, cst_text, mm, span=1, thr=0.30):
    """Último fallback: casa por CONTENIDO — el marcador de la página cuyo texto
    comparte más tokens con la incipit CST (nombres divergentes: manussacuti↔pañcagati)."""
    A = _ctoks(cst_text)
    if not A:
        return None
    best_ov, best_txt = 0.0, None
    for pg in [page] + [page + d for d in range(1, span + 1)] + [page - d for d in range(1, span + 1)]:
        r = cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (SN5_BOOK, pg)).fetchone()
        if not r:
            continue
        lines = (r['unitext'] or '').split('\n')
        seen = set()
        for line, gid, vpos, name in mm.get(pg, []):
            if line in seen:
                continue
            seen.add(line)
            start, end = line - 1, len(lines)
            for j in range(start + 1, len(lines)):
                if re.match(r'^\d+(?:\s*-+\s*\d+)?\.', lines[j].strip()):
                    end = j
                    break
            ov = len(A & _ctoks(' '.join(lines[start:end]))) / max(1, len(A))
            if ov > best_ov:
                best_ov, best_txt = ov, ' '.join(sh.tokens(' '.join(lines[start:end]))[:350])
    return best_txt if best_ov >= thr else None


def cst_boundaries(segs):
    b, cur = {}, 45
    for i, s in enumerate(segs):
        if re.match(r'^1\.\s+\S', s.get('title', '')):
            b[cur] = i
            cur += 1
    b[cur] = len(segs)
    return b


def main():
    import statistics as st
    n = int(sys.argv[sys.argv.index('--n') + 1]) if '--n' in sys.argv else 30
    do_all = '--all' in sys.argv
    dry = '--dry' in sys.argv
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    entries = load_entries()
    mm = build_markers(cur)
    segs = sh.load_cha(['s5m.xml'], 'h4n')
    cst_items = [(' '.join(sh.tokens(s['text'])[:350]), s.get('title', '')) for s in segs]

    pts_items, meta = [], {}
    for e in entries:
        pts = pts_for(cur, e['page'], e['sn_inner'], mm)
        if not pts:
            continue
        pts_items.append((e['num'], pts, e['name']))
        meta[e['num']] = (e, pts)

    aligned = align(pts_items, cst_items)
    if not do_all and not dry:
        aligned = aligned[:n]
    na = sum(1 for _, i, _ in aligned if i is None)
    print(f'SN V: {len(entries)} entradas, {len(pts_items)} PTS resolubles; '
          f'alineador → {len(aligned) - na}/{len(aligned)} alineadas')

    if dry:
        covs = [c for _, i, c in aligned if i is not None]
        # consistencia de título: ¿la raíz del nombre PTS aparece en el título CST?
        ok = tot = 0
        for num, idx, _c in aligned:
            if idx is None:
                continue
            tot += 1
            base = _nfold(re.sub(r'^(paṭhama|dutiya|tatiya|catuttha|pañcama|chaṭṭha|sattama|aṭṭhama|navama|dasama)', '',
                                 meta[num][0]['name'].lower()))[:6]
            if base and base in _nfold(cst_items[idx][1].lower()):
                ok += 1
        print(f'[DRY] cobertura alineada: min={min(covs):.2f} mediana={st.median(covs):.2f} '
              f'max={max(covs):.2f} | sin alinear: {na}')
        print(f'[DRY] consistencia de título (raíz PTS ⊂ título CST): {ok}/{tot} ({100*ok/max(1,tot):.0f}%)')
        for thr in (0.45, 0.55, 0.65):
            print(f'  cov>={thr}: {sum(1 for c in covs if c >= thr)}/{len(aligned)}')
        return

    rows = []
    n_align = sum(1 for _, i, _ in aligned if i is not None)
    done = 0
    for num, idx, cov in aligned:
        e, pts = meta[num]
        if idx is not None:
            done += 1
            if done % 25 == 0:
                print(f'  ...{done}/{n_align} alineadas validadas (Gemini)', flush=True)
        if idx is None:
            rows.append({'num': num, 'name': e['name'], 'cst_title': '(sin alinear)', 'legacy': e['legacy'],
                         'estado': 'PENDIENTE', 'validation': 'REVISAR', 'gate': {'cov': cov},
                         'gemini': None, 'reason': f'sin alineación CST confiable (cov={cov})'})
            continue
        cst, ct = cst_items[idx]
        res = validate_pair(pts, cst, e['name'], ct, 'SN')
        rows.append({'num': num, 'name': e['name'], 'cst_title': ct, 'legacy': e['legacy'],
                     'aligncov': cov, **res})
    json.dump(rows, open('validador_sn5_batch.json', 'w'), ensure_ascii=False, indent=1)
    report(rows)


def report(rows):
    est = Counter(r['estado'] for r in rows)
    val = Counter(r['validation'] for r in rows)
    revisar = [r for r in rows if r['validation'] == 'REVISAR']
    print('=' * 92)
    print(f'VALIDADOR (Modelo B) — SN V, primera tanda ({len(rows)})')
    print('=' * 92)
    print(f'Estado: {dict(est)} | Validation: {dict(val)}')
    print(f'\nA revisión humana ({len(revisar)}) — gate y Gemini discrepan:')
    for r in revisar[:60]:
        g = r['gate'] or {}
        gj = f"cov={g['cov']}" if 'cov' in g else (f"jac={g.get('jac')} tit={g.get('title')}" if g else '')
        print(f'  SN {r["num"]:>7} | {r["name"][:20]:20} -> {r["cst_title"][:22]:22} | {gj} | {r["reason"][:40]}')
    print('\nResultados en validador_sn5_batch.json (nada escrito al Excel).')


if __name__ == '__main__':
    main()
