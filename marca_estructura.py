#!/usr/bin/env python3
"""
marca_estructura — pone en los metadatos lo que la convención filológica pide, sin tocar ni una
referencia. Dos cosas distintas que hoy se confunden en `PENDIENTE`:

### 1. Las cabeceras de sección no son entradas (`Type = Section Header`)

Un rótulo que el impreso trae —`Therāpadāna`, `2. Pucchā.`— es un dato del testimonio pero no una
unidad citable. **No se borra** (no se destruye estructura atestiguada): se tipa, para que quede
fuera del recuento de entradas sin desaparecer del fichero. La columna `Type` ya tiene el valor y
seis de las siete lo llevaban; faltaban `Cnd 2 Pucchā` y la cabecera del Milindapañha.

Su `Validation` decía `VERSE_ONLY`, que es un **veredicto de cotejo** y no describe nada: una
cabecera no se coteja. Pasa a `SECTION_HEADER`, que es un hecho estructural.

### 2. La ausencia en un testimonio es un dato positivo (`Validation = DEEST_PTS`)

En crítica textual, que un testimonio no traiga un pasaje **se registra**: `om.` (*omittit*) para
una lección omitida, **`deest`** («falta») para un pasaje ausente. Es la práctica de los aparatos de
PTS y del *Critical Pāli Dictionary*. La ausencia es una lectura del testimonio, no un hueco.

Catorce filas están **verificadamente ausentes de PTS**, y dejarlas como «no verificadas» confunde
*evidencia negativa* con *falta de evidencia* — la distinción sobre la que está construida la regla
del `Estado`:

- **el vagga 56 del Therāpadāna** (11): el impreso cierra en `[547. Cūlasugandha]` (p510) y en la
  p511 va el colofón de toda la obra. El propio Excel lo delata dando la 511 a las once;
- **tres del vagga 34**: el uddāna impreso nombra **diez** —`Gandhodaka-Pūjanī ca
  Punnāga-Ekadussikā | Phusito ca Pabhaṅkāro Kuṭido Uttarīyako | Savanī Ekapadumī`— y la edición
  imprime **siete**: entre `[331. Gandhathūpiya]` (p267) y `[332. Phussitakammiya]` (p268) no hay
  nada. Dos de ellas llevan además un **renvoi**: su texto coincide al 0,96 y al 0,93 con apadānas
  que PTS **sí** imprime (`Ap 106`, `Ap 419`) y que ya tienen fila propia.

⚠️ **`Estado` sigue siendo binario.** No se inventa un tercer valor: la regla de dos es buena y el
fichero ya tiene el precedente —las 297 invalidables llevan `Validation = EXTRA_CANON`, que también
es un hecho editorial y no un veredicto—. Estas catorce siguen en `PENDIENTE`, porque CONFIRMADO
exige un par cotejado y aquí no hay lado PTS. Lo que cambia es que ahora **se sabe por qué**, y el
recuento honrado sale del denominador.

⚠️ **Y un límite que hay que dejar escrito**: lo verificado es *esta transcripción* de PTS, no el
volumen impreso. La evidencia interna es fuerte —el uddāna que anuncia diez, el colofón que cierra
la obra— y apunta a la edición misma, pero la convención distingue `deest in Ee` de «no está en
nuestra fuente». Queda anotado en el `Detail` de cada fila.

⚠️ El `PTS_CROSSREF_SN` de MN 92 y 98 **no es precedente aplicable**: allí PTS *remite
explícitamente* al Suttanipāta y por eso es CONFIRMADO. Aquí PTS **omite**; es una laguna, no un
reenvío.

Uso: python3 marca_estructura.py [--dry]
"""
import re
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# ── cabeceras: (fragmento del nombre, sección esperada o None si no se toca)
#
# ⚠️ `Cnd 2 Pucchā` **sí tiene página** (la 6, donde PTS imprime el rótulo) y aun así es una
# cabecera: encabeza toda la sección de preguntas del Cūḷaniddesa y su contenido son las filas
# `Cnd 5-22`. Tener página no la convierte en unidad citable, así que la lista va explícita y no
# por «las que no tienen página».
CABECERAS = [
    ('Nd 2 2 Pucchā', None),
    ('Milindapa', 'Milindapañha (Mil)'),
]

_DEEST_56 = (
    'DEEST: PTS no imprime este apadāna. Su edición del Therāpadāna cierra en '
    '«[547. Cūlasugandha]» (p510) y la p511 es el colofón de toda la obra; las once filas del vagga '
    '56 del CST (Yasavaggo) llevan por eso esa misma p511, que es la del colofón. No hay lado PTS '
    'que cotejar, así que no puede ser CONFIRMADO — pero la ausencia está verificada, no es una '
    'fila sin mirar. ⚠ Comprobado sobre la transcripción de la BD, no sobre el volumen impreso.')

_DEEST_34 = (
    'DEEST: PTS no imprime este apadāna, aunque su propio uddāna lo nombra. El uddāna del vagga 34 '
    '(«Gandhodaka-Pūjanī ca Punnāga-Ekadussikā | Phusito ca Pabhaṅkāro Kuṭido Uttarīyako | Savanī '
    'Ekapadumī») anuncia diez y la edición imprime siete: entre «[331. Gandhathūpiya]» (p267) y '
    '«[332. Phussitakammiya]» (p268) no hay nada. No hay lado PTS que cotejar. '
    '⚠ Comprobado sobre la transcripción de la BD, no sobre el volumen impreso.')

# renvoi: el texto está impreso en otro lugar, bajo su propia entrada, que ya tiene fila
_CF = {
    '10.34.2': ' cf. Ap 106 «Udakapūjaka» (p142), homónimo cuyo texto coincide al 0,96 con éste en '
               'el CST y que PTS sí imprime, con fila propia. La referencia NO se comparte: dos '
               'filas sobre un mismo texto es lo que impide `audit_injectivity`.',
    '10.34.4': ' cf. Ap 419 «Ekadussadāyaka» (p379), homónimo cuyo texto coincide al 0,93 con éste '
               'en el CST y que PTS sí imprime, con fila propia. La referencia NO se comparte.',
}

AUSENTES = ({f'10.34.{k}': _DEEST_34 + _CF.get(f'10.34.{k}', '') for k in (2, 3, 4)}
            | {f'10.56.{k}': _DEEST_56 for k in range(1, 12)})


def main():
    dry = '--dry' in sys.argv
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    cab, aus, val = [], [], []
    for r in range(2, ws.max_row + 1):
        nom = str(ws.cell(r, ci['Sutta Name']).value or '')
        num = str(ws.cell(r, ci['Sutta #']).value or '')
        tipo = str(ws.cell(r, ci['Type']).value or '')
        if tipo == 'Section Header':
            val.append(r)
        for frag, sec in CABECERAS:
            if frag in nom and tipo != 'Section Header':
                cab.append((r, nom, sec))
        if num in AUSENTES and str(ws.cell(r, ci['PTS Vol']).value) in {'Th-ap', 'Thī-ap'}:
            aus.append((r, num))
    print(f'cabeceras a tipar: {len(cab)} · ya tipadas: {len(val)} · ausentes a marcar: {len(aus)}')
    for _r, nom, sec in cab:
        print(f'   → Section Header: «{nom}»' + (f'  ·  Section = {sec}' if sec else ''))
    if len(aus) != 14:
        print(f'⚠ esperaba 14 filas ausentes y encuentro {len(aus)}: no se escribe nada')
        return
    if dry:
        print('(dry-run: nada escrito)')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-estructura.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    for r, nom, sec in cab:
        ws.cell(r, ci['Type']).value = 'Section Header'
        if sec:
            ws.cell(r, ci['Section']).value = sec
            # `Milindapa ñha`: el espacio dentro de la palabra es un artefacto de la transcripción
            ws.cell(r, ci['Sutta Name']).value = re.sub(r'\s+ñ', 'ñ', nom)
        val.append(r)
    for r in set(val):
        ws.cell(r, ci['Validation']).value = 'SECTION_HEADER'
        ws.cell(r, ci['Estado']).value = 'PENDIENTE'
        ws.cell(r, ci['Detail']).value = (
            'Rótulo de sección del impreso, no una unidad citable: no se coteja y no cuenta como '
            'entrada. Se conserva porque es estructura atestiguada del testimonio.')
    for r, num in aus:
        ws.cell(r, ci['Validation']).value = 'DEEST_PTS'
        ws.cell(r, ci['Estado']).value = 'PENDIENTE'
        ws.cell(r, ci['Detail']).value = AUSENTES[num]
    wb.save(XLSX)
    print(f'{len(set(val))} cabeceras + {len(aus)} ausencias marcadas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
