#!/usr/bin/env python3
"""fix_sn2_antarapeyyala.py — ajusta el *antara-peyyālaṃ* de SN 12 a la estructura de PTS.

**Qué pasaba.** Dos filas de S ii venían con la forma del CST, no la de PTS:

1. `12.74 «Dutiyasuttādidasakaṃ»` — **no es un sutta en PTS**. El CST parte el sutta PTS 82
   («Satthā») en 1 + 10: `1. Satthusuttaṃ` (el ítem jarāmaraṇa) y
   `2-11. Dutiyasatthusuttādidasakaṃ` (los ítems jāti…saṅkhāra). PTS imprime todo eso como UN
   sutta y lo dice literalmente: tras el §1 escribe **`Suttanto eko`** («un solo sutta», S ii 130
   L28) y luego `Sabbesam evam peyyālo` con los §§2–11 (S ii 131 L1–28). Su propio uddāna remata:
   «Satthā Sikkhā … **Appamādena dvādasāti**» = **doce** suttas, y «Pare te dvādasa honti, suttā
   dvattiṃsasatāni» = doce, o 132 si se cuentan por ítem. Así que esa fila no tiene referencia PTS
   propia → **se borra**.
   NB: no es una elisión. El texto está en PTS; lo que no existe es la división en suttas. (Distinto
   del caso de SN V, donde PTS SÍ numera los suttas — `122--132. (2--12)` — y elide solo el texto
   con `vitthāretabbo`.)

2. `12.75 «Sikkhāsuttādipeyyālaekādasakaṃ»` — el CST agrupa, pero **PTS imprime los 11 con
   marcador y número propios**: `83 (2) Sikkhā`, `84 (3) Yogo`, … `93 (12) Appamādo`. Aquí PTS no
   agrupa (a diferencia de 12.72, 17.13 o 17.38, donde sí imprime rango), así que la fila se
   **expande a 11**, con la página y la línea reales de cada marcador.

Decisión de Jorge, 2026-07-25. Uso: python3 fix_sn2_antarapeyyala.py [--write]
"""
import datetime, re, shutil, sqlite3, sys

from openpyxl import load_workbook

from validador_sn2 import DB, XLSX, build_massive, build_pts_suttas

DELETE = '12.74'          # división solo-CST, sin sutta en PTS
EXPAND = '12.75'          # grupo CST → los 11 suttas que PTS numera
EXPAND_RANGE = range(83, 94)

# el nombre del marcador arrastra el dígito de la llamada a nota al pie: «Ussoḷhī1»
_FOOTNOTE = re.compile(r'\d+$')


def clean_name(s):
    return _FOOTNOTE.sub('', (s or '').strip()).strip()


def main():
    write = '--write' in sys.argv
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
    pts = build_pts_suttas(conn.cursor())
    massive = build_massive()

    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}

    del_row = exp_row = None
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'SN' or str(v[ci['PTS Roman']] or '').strip().lower() != 'ii':
            continue
        if str(v[ci['Sutta #']]) == DELETE:
            del_row = (ridx, v)
        elif str(v[ci['Sutta #']]) == EXPAND:
            exp_row = (ridx, v)
    if not (del_row and exp_row):
        print(f'No encuentro las filas {DELETE} / {EXPAND} — ¿ya aplicado?')
        return

    print(f'A BORRAR   fila {del_row[0]}: {DELETE} «{del_row[1][ci["Sutta Name"]]}» '
          f'p{del_row[1][ci["PTS Page"]]}')
    print(f'A EXPANDIR fila {exp_row[0]}: {EXPAND} «{exp_row[1][ci["Sutta Name"]]}» → 11 filas:')
    plan = []
    for k in EXPAND_RANGE:
        p = pts.get((12, k))
        h = massive.get((12, k))
        if not p:
            print(f'   !! SN 12.{k} sin marcador PTS — abortando'); return
        plan.append((k, p, h))
        print(f'   12.{k:<3} «{clean_name(p["name"])[:18]:18}» pos ({p["pos"]}) '
              f'S ii {p["page"]},{p["line"]:<3} CST={(h[1][:26] if h else "grupo 2-12 ítem (%s)" % p["pos"])}')

    if not write:
        print('\n(dry-run, nada escrito — usa --write)')
        return

    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-sn2antara.xlsx'
    shutil.copy(XLSX, bak); print('\nbackup →', bak)

    # 1) la fila a expandir se reutiliza para el primero (12.83) y se insertan 10 debajo
    ridx, v = exp_row
    ws.insert_rows(ridx + 1, amount=len(plan) - 1)
    for j, (k, p, h) in enumerate(plan):
        r = ridx + j
        if j:                                   # las filas nuevas heredan las columnas fijas
            for col in range(1, len(hdr) + 1):
                ws.cell(row=r, column=col).value = v[col - 1]
        ws.cell(row=r, column=ci['Sutta #'] + 1).value = f'12.{k}'
        ws.cell(row=r, column=ci['Sutta Name'] + 1).value = clean_name(p['name']) or f'(sin nombre {k})'
        ws.cell(row=r, column=ci['PTS Page'] + 1).value = p['page']
        ws.cell(row=r, column=ci['PTS Ref'] + 1).value = f"S ii {p['page']},{p['line']}"
        ws.cell(row=r, column=ci['Validation'] + 1).value = 'DB_VERIFIED'
        ws.cell(row=r, column=ci['Estado'] + 1).value = 'PENDIENTE'
        ws.cell(row=r, column=ci['Detail'] + 1).value = (
            f'Antara-peyyālaṃ: PTS imprime este sutta con marcador y nº propios ({k} ({p["pos"]})); '
            f'el CST lo agrupa en «2-12. Sikkhāsuttādipeyyālaekādasakaṃ», ítem ({p["pos"]}).')
        ws.cell(row=r, column=ci['Raw ID'] + 1).value = f"SN 12.{k} {clean_name(p['name'])}"
    print(f'Expandida {EXPAND} en {len(plan)} filas (12.83–12.93).')

    # 2) borrar la fila solo-CST (su índice se ha desplazado por la inserción)
    off = len(plan) - 1
    drow = del_row[0] + (off if del_row[0] > ridx else 0)
    assert str(ws.cell(row=drow, column=ci['Sutta #'] + 1).value) == DELETE, 'fila movida'
    ws.delete_rows(drow)
    print(f'Borrada la fila {DELETE} (división solo-CST, sin sutta en PTS).')

    # 3) renumerar la columna `#` del canon completo
    n = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=ci['Nikaya'] + 1).value:
            n += 1
            ws.cell(row=r, column=ci['#'] + 1).value = n
    wb.save(XLSX)
    print(f'Renumerada la columna `#` ({n} filas). Guardado {XLSX}.')


if __name__ == '__main__':
    main()
