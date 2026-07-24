#!/usr/bin/env python3
"""
Validate Aṅguttara Nikāya (1,738 entries) against tipitaka.sqlite.
"""

import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict

DB_PATH = '/home/jorge/Code/squashfs-root/src/data/tipitaka.sqlite'
EXCEL_PATH = '/home/jorge/Code/squashfs-root/PTS_Reference_Full_Canon_CORRECTED.xlsx'

AN_BOOK_MAP = {1: 17, 2: 18, 3: 19, 4: 20, 5: 21}
AN_BOOK_RANGE = {17: (1, 304), 18: (1, 257), 19: (1, 452), 20: (1, 466), 21: (1, 361)}

def strip_diacritics(text):
    for k, v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        text = text.replace(k, v).replace(k.upper(), v.upper())
    return text

def sutta_words(name, min_len=3):
    """Extract meaningful words from a sutta name."""
    clean = strip_diacritics(name.lower())
    clean = re.sub(r'sutta(m|nta)?', '', clean)
    clean = re.sub(r'\[.*?\]|\(.*?\)', '', clean)
    clean = re.sub(r'^\d+[\.\-\s]*', '', clean)
    words = [w for w in re.split(r'[\s\-–—,;:.]+', clean) if len(w.strip()) >= min_len]
    return words

def match_in_text(name, text):
    sw = sutta_words(name)
    if not sw:
        return 0.0
    txt = strip_diacritics(text.lower())
    return sum(1 for w in sw if w in txt) / len(sw)

def load_excel():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['PTS Reference']
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] != 'AN':
            continue
        entries.append({
            'row_num': row[0], 'sutta_num': str(row[2] or ''),
            'sutta_name': (row[3] or '').strip(), 'sutta_raw': str(row[11] or '').strip(),
            'pts_vol': row[5], 'pts_roman': row[6], 'pts_page': row[7], 'pts_full': row[8]
        })
    return entries

def roman_to_int(roman):
    return {'i':1,'ii':2,'iii':3,'iv':4,'v':5,'vi':6}.get(str(roman).strip().lower(), 0)

def validate():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    entries = load_excel()
    print(f"Loaded {len(entries)} AN entries\n")
    
    # Load TOC
    cur.execute("SELECT book_no, page_no, section, title FROM contents WHERE book_no BETWEEN 17 AND 21")
    toc_index = {(r['book_no'], r['page_no']): (r['section'], r['title']) for r in cur.fetchall()}
    
    # Load pages — only HEAD + first 400 chars for speed
    cur.execute("SELECT book_no, page_no, head, substr(unitext, 1, 400) as txt FROM pages WHERE book_no BETWEEN 17 AND 21 AND edition='mula'")
    pages_index = {(r['book_no'], r['page_no']): (r['head'] or '', r['txt'] or '') for r in cur.fetchall()}
    
    print(f"DB: {len(toc_index)} TOC entries, {len(pages_index)} pages for AN\n")
    
    results = []
    stats = defaultdict(int)
    
    for e in entries:
        vol = roman_to_int(e['pts_roman'])
        book_no = AN_BOOK_MAP.get(vol)
        page = e['pts_page']
        
        r = {**e, 'book_no': book_no, 'db_page': page, 'head': '', 'first_words': ''}
        
        if not book_no:
            r['status'] = 'ERROR'; results.append(r); stats['ERROR'] += 1; continue
        
        # Page exists?
        db_key = (book_no, page)
        if db_key not in pages_index:
            r['status'] = 'MISSING'; r['detail'] = f'Page {page} not in book {book_no}'; results.append(r); stats['MISSING'] += 1; continue
        
        head, text = pages_index[db_key]
        r['head'] = head[:100]
        r['first_words'] = ' '.join(text.split()[:12])[:120]
        
        # Strategy 1: TOC at exact page
        toc = toc_index.get(db_key)
        if toc:
            section, title = toc
            r['toc_section'] = section; r['toc_title'] = title
            ratio = match_in_text(e['sutta_name'], title)
            if ratio >= 0.4:
                r['status'] = 'OK_TOC'; r['detail'] = title; stats['OK_TOC'] += 1
            else:
                r['status'] = 'TOC_NEAR'; r['detail'] = f'TOC: {title}'; stats['TOC_NEAR'] += 1
            results.append(r); continue
        
        # Strategy 2: HEAD match
        if match_in_text(e['sutta_name'], head) >= 0.3:
            r['status'] = 'OK_HEAD'; r['detail'] = head[:70]; stats['OK_HEAD'] += 1
            results.append(r); continue
        
        # Strategy 3: Content text match
        if match_in_text(e['sutta_name'], text) >= 0.3:
            r['status'] = 'OK_CONT'; stats['OK_CONT'] += 1
            results.append(r); continue
        
        # Strategy 4: Check pages +-1
        found = False
        for delta in [-1, 1]:
            nearby = (book_no, page + delta)
            nh, nt = pages_index.get(nearby, ('', ''))
            if match_in_text(e['sutta_name'], nh + nt) >= 0.3:
                r['status'] = 'OK_NEAR'; r['detail'] = f'Match at p.{page+delta}'; stats['OK_NEAR'] += 1
                found = True; break
        if found:
            results.append(r); continue
        
        r['status'] = 'UNVERIF'; r['detail'] = 'No match in page content'; stats['UNVERIF'] += 1
        results.append(r)
    
    conn.close()
    return results, stats

def print_report(results, stats):
    total = len(results)
    verified = total - stats.get('UNVERIF', 0) - stats.get('ERROR', 0) - stats.get('MISSING', 0)
    
    print(f"{'='*90}")
    print(f"  AṄGUTTARA NIKĀYA — Content Validation")
    print(f"  Database: tipitaka.sqlite (PTS edition, 'mula')")
    print(f"{'='*90}")
    print(f"  Total: {total}")
    print(f"  ✓ OK_TOC: {stats.get('OK_TOC',0):>5}  |  OK_HEAD: {stats.get('OK_HEAD',0):>5}  |  OK_CONT: {stats.get('OK_CONT',0):>5}")
    print(f"  ≈ TOC_NEAR:{stats.get('TOC_NEAR',0):>5}  |  OK_NEAR: {stats.get('OK_NEAR',0):>5}")
    print(f"  ⚠ UNVERIF: {stats.get('UNVERIF',0):>5}  |  ERROR:   {stats.get('ERROR',0):>5}  |  MISSING: {stats.get('MISSING',0):>5}")
    print(f"  Accuracy: {verified}/{total} = {100*verified/total:.1f}%")
    print(f"{'─'*90}")
    
    # Show problematic entries
    issues = [r for r in results if r['status'] in ('UNVERIF', 'ERROR', 'MISSING')]
    if issues:
        print(f"\n  ⚠ Items needing review: {len(issues)}\n")
        for r in issues[:50]:  # Show first 50
            m = {'UNVERIF': '?', 'ERROR': '‼', 'MISSING': '✗'}.get(r['status'], '?')
            print(f"  {m} [{r['status']:7s}] {r['sutta_raw'][:55]:55s} {r['pts_full']:10s}")
            if r.get('head'):
                print(f"       Head: {r['head'][:90]}")
            fw = r.get('first_words','')
            if fw:
                print(f"       Text: {fw[:110]}")
        if len(issues) > 50:
            print(f"\n  ... and {len(issues)-50} more")
    
    # Quick analysis of UNVERIF by nipata/volume
    unv_by_vol = defaultdict(list)
    for r in results:
        if r['status'] == 'UNVERIF':
            unv_by_vol[r['pts_roman']].append(r)
    
    if unv_by_vol:
        print(f"\n  UNVERIF by volume:")
        for vol in sorted(unv_by_vol.keys()):
            entries = unv_by_vol[vol]
            # Show first 3 examples
            print(f"    Vol {vol}: {len(entries)} entries")
            for e in entries[:3]:
                print(f"      {e['sutta_raw'][:60]}")
    
    # Summary
    print(f"\n{'='*90}")
    print(f"  Combined verified: {verified}/{total} ({100*verified/total:.1f}%)")

if __name__ == '__main__':
    results, stats = validate()
    print_report(results, stats)
