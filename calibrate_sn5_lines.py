#!/usr/bin/env python3
"""calibrate_sn5_lines.py — calibra el nº de LÍNEA de las referencias PTS de SN V.

La referencia del Excel es `S v <página>,<línea>`; la línea no es fiable (viene del
raspado, no del texto). Aquí se recalcula contra el texto real de la BD (libro 16):
se reusa la MISMA asignación del pipeline VRI —global, inyectiva y monótona
(`assign_volume`)—, que devuelve la posición exacta del marcador de cada sutta, y se
compara. Antes se repetía aquí la cadena de heurísticas por fila, que no es inyectiva:
dos filas podían apuntar al mismo marcador y una recibía la línea de su vecina.

Hallazgo (2026-07-25): las líneas del Excel NO son posiciones de línea reales — se
agolpan en 2–8 sobre páginas de ~29 líneas (distribución imposible), y donde
discrepan, el marcador de la BD es el correcto (verificado a mano: p7 L21, p3 L12,
p4 L16, p11 L16, p13 L8 tienen UN solo marcador centrado y es el del sutta). Las 75
que ya coincidían sí eran líneas buenas: el dato venía mezclado.

`--dry` (por defecto) solo mide el desfase; `--write` corrige el `PTS Ref` del Excel
(con backup) SOLO en las filas cuyo marcador está en la PÁGINA DECLARADA, dejando
intactas las demás — reescribir la página es una afirmación mayor que la línea, así
que los desacuerdos de página se reportan para arbitraje, no se tocan.

Uso: python3 calibrate_sn5_lines.py [--write]
"""
import datetime, json, re, shutil, sqlite3, sys
from collections import Counter

import sutta_hash as sh
from openpyxl import load_workbook

from validador_sn5 import build_markers, locate_by_name, locate_for, ordinal_of
from validador_sn5_vri import (DB, SN5_BOOK, build_massive, build_vri_index, assign_volume,
                               excel_entries as vri_entries)

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
REF_RE = re.compile(r'^(S\s+v\s+)(\d+)(?:\s*,\s*(\d+))?\s*$')


def locate_all(cur, mm, canon2para, vri):
    """`Sutta # → (página, línea)` con la MISMA asignación global que el validador.

    Antes esto repetía la cadena de heurísticas por fila, que no es inyectiva: dos filas podían
    apuntar al mismo marcador y una recibía la línea de su vecina.
    """
    return assign_volume(vri_entries(), cur, mm, canon2para, vri)


def locate(cur, e, ctitle, mm):
    """Posición del marcador PTS del sutta. Misma prioridad que `validador_sn5_vri`."""
    if not (isinstance(e['page'], int) and e['inner']):
        return None, None
    at = locate_by_name(cur, e['page'], ctitle, mm, min_score=3)
    if at:
        return at, 'name'
    if ordinal_of(ctitle) is None:
        at = locate_by_name(cur, e['page'], ctitle, mm, min_score=1, require_unique=True)
        if at:
            return at, 'name!'
    at = locate_for(cur, e['page'], e['inner'], mm)
    if at:
        return at, 'num'
    at = locate_by_name(cur, e['page'], ctitle, mm, min_score=1)
    return (at, 'name~') if at else (None, None)


def main():
    write = '--write' in sys.argv
    vri, canon2para = build_vri_index(), build_massive()
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    mm = build_markers(cur)
    assigned = locate_all(cur, mm, canon2para, vri)

    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}

    delta, src_c, fixes, unresolved, otherpage = Counter(), Counter(), [], [], []
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'SN' or str(v[ci['PTS Roman']] or '').strip().lower() != 'v':
            continue
        name = str(v[ci['Sutta Name']] or '')
        m = re.search(r'\(SN\s+(\d+)\.(\d+)', name)
        canon = f"{m.group(1)}.{m.group(2)}" if m else str(v[ci['Sutta #']])
        inner = int(canon.split('.')[1]) if '.' in canon and canon.split('.')[1].isdigit() else None
        e = {'page': v[ci['PTS Page']], 'inner': inner}

        at, src = assigned.get(str(v[ci['Sutta #']])), 'dp'
        src_c[src if at else 'sin-marcador'] += 1
        if not at:
            unresolved.append((str(v[ci['Sutta #']]), canon))
            continue

        pg, line = at
        ref = str(v[ci['PTS Ref']] or '')
        if pg != e['page']:
            # el marcador cayó en una página vecina: eso cuestiona la PÁGINA, no la
            # línea. Fuera del alcance de esta calibración → a arbitraje.
            otherpage.append((str(v[ci['Sutta #']]), ref, f'S v {pg},{line}', src))
            continue
        rm = REF_RE.match(ref)
        old_line = int(rm.group(3)) if (rm and rm.group(3)) else None
        delta[(None if old_line is None else old_line - line)] += 1
        new_ref = f'S v {pg},{line}'
        if ref != new_ref:
            fixes.append((ridx, ref, new_ref, src))

    print('=' * 84)
    print('CALIBRACIÓN DE LÍNEA — SN V (marcador real en la BD, libro 16)')
    print('=' * 84)
    print('resolución del marcador:', dict(src_c))
    print('\ndesfase (línea del Excel − línea del marcador), por frecuencia:')
    for d, n in delta.most_common(15):
        print(f'  {"sin línea" if d is None else f"{d:+d}":>10} : {n}')
    exact = delta.get(0, 0)
    tot = sum(delta.values())
    print(f'\ncoincidencia exacta: {exact}/{tot} ({100 * exact / max(1, tot):.0f}%)')
    print(f'referencias a corregir: {len(fixes)} | sin marcador: {len(unresolved)} | '
          f'marcador en otra página (a arbitraje, NO se tocan): {len(otherpage)}')
    for ridx, o, n, src in fixes[:20]:
        print(f'  fila {ridx}: {o!r} → {n!r}  ({src})')
    if otherpage:
        print('\npágina en disputa (Excel dice una, el marcador está en otra):')
        for num, o, n, src in otherpage:
            print(f'  SN {num:>8}: {o!r} vs marcador {n!r} ({src})')
    if unresolved:
        print('\nsin marcador resoluble (línea intacta):',
              ', '.join(f'{a}' for a, _b in unresolved))

    if not write:
        print('\n(dry-run, nada escrito — usa --write)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-lineas.xlsx'
    shutil.copy(XLSX, bak); print('\nbackup →', bak)
    for ridx, _o, new_ref, _s in fixes:
        ws.cell(row=ridx, column=ci['PTS Ref'] + 1).value = new_ref
    wb.save(XLSX)
    print(f'Corregidas {len(fixes)} referencias en {XLSX}.')


if __name__ == '__main__':
    main()
