#!/usr/bin/env python3
"""Driver SN I (S i, book_no=12) — pipeline VRI por concordancia + validador (Modelo B).

Mismo esqueleto que `validador_sn5_vri.py`, con tres diferencias propias de SN I:

1. **Marcadores PTS `§ N. Nombre.`** reiniciados por vagga (ver `sn1_markers.py`, pyparsing).
   La estructura PTS se valida contra el **front matter de Feer** (`samyutta-vol-I-info.txt`):
   271 suttas, 28 vaggas, y la página de arranque de cada vagga. Cuadra al 100%, así que el lado
   PTS queda fijado **sin LLM y sin conjeturas**.
2. **Alineación por POSICIÓN, no difusa.** Excel (271 filas en orden canónico) y marcadores PTS
   (271 en orden de lectura) son la misma secuencia: el recuento por saṃyutta del Excel
   (81/30/25/25/10/15/22/12/14/12/25) coincide exacto con la tabla de Feer. El CST entra por
   concordancia (`massive.tsv` → `cst_paranum` → XML VRI), independiente de la posición, así que
   el acuerdo de nombre PTS↔CST es una comprobación *cruzada*, no la fuente de la alineación.
3. **XML VRI del Sagāthāvagga**: el nº de párrafo lo llevan `bodytext` **y `hangnum`** (223+48=271
   — los suttas solo-verso usan `hangnum`), y el texto vive en los `gatha1/2/3/gathalast`. El
   índice de SN V (solo `bodytext`) se queda en 223 y hay que ampliarlo.

Uso: python3 validador_sn1.py [--dry] [--all] [--n N] [--only 1.1,2.3]
"""
import csv, os, re, sqlite3, sys, json
import xml.etree.ElementTree as ET
from collections import Counter

import sutta_hash as sh
from openpyxl import load_workbook

from validador import validate_pair
from sn1_markers import find_markers_sn1, find_submarkers_sn1, build_vaggas

VRI = '/tmp/tipitaka-xml/romn/s0301m.mul.xml'
OUT = 'validador_sn1.json'
DB = 'src/data/tipitaka.sqlite'
SN1_BOOK = 12
XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# Estructura de S i según el front matter de Feer (samyutta-vol-I-info.txt):
# (saṃyutta, nombre del vagga, página de arranque, nº de suttas)
FEER = [(1, 'Naḷa', 1, 10), (1, 'Nandana', 5, 10), (1, 'Satti', 13, 10),
        (1, 'Satullapakāyika', 16, 10), (1, 'Āditta', 31, 10), (1, 'Jarā', 36, 10),
        (1, 'Addha', 39, 10), (1, 'Chetvā', 41, 11),
        (2, 'I', 46, 10), (2, 'Anāthapiṇḍika', 51, 10), (2, 'Nānātitthiya', 56, 10),
        (3, 'I', 68, 10), (3, 'II', 77, 10), (3, 'Pañcaka', 93, 5),
        (4, 'I', 103, 10), (4, 'II', 109, 10), (4, 'Uparipañca', 117, 5),
        (5, 'Bhikkhunī', 128, 10), (6, 'I', 136, 10), (6, 'Pañcaka', 153, 5),
        (7, 'Arahanta', 160, 10), (7, 'Upāsaka', 172, 12),
        (8, 'Vaṅgīsa', 185, 12), (9, 'Vana', 197, 14), (10, 'Yakkha', 206, 12),
        (11, 'I', 216, 10), (11, 'II', 228, 10), (11, 'Pañcaka', 237, 5)]

# rends del XML VRI que aportan texto del sutta (el Sagāthāvagga es casi todo verso)
_TEXT_RENDS = {'bodytext', 'gatha1', 'gatha2', 'gatha3', 'gathalast', 'hangnum', 'unindented'}


def build_vri_index(path=VRI):
    """paranum → sutta CST (`title`, `text`). Amplía el índice de SN V: el nº lo llevan
    `bodytext` y `hangnum`, y el texto está repartido por los rends de gāthā."""
    body = ET.parse(path).getroot().find('.//body')
    suttas, cur = [], None
    for el in body.iter():
        if el.tag != 'p':
            continue
        rend = el.get('rend')
        if rend == 'subhead':
            cur = {'title': ''.join(el.itertext()).strip(), 'pn': [], 'text': ''}
            suttas.append(cur)
        elif rend in _TEXT_RENDS and cur is not None:
            n = el.get('n')
            if n:
                cur['pn'].append(n)
            cur['text'] += ' ' + ''.join(el.itertext())
    idx = {}
    for s in suttas:
        for n in s['pn']:
            m = re.match(r'(\d+)(?:-(\d+))?$', n)
            if m:
                for p in range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1):
                    idx.setdefault(p, s)
    return idx


def _pts_page(v):
    """`cst_p_page` → nº de página PTS. El campo es `vol.pppp` (`1.0001` = vol I p. 1) pero
    viene guardado como DECIMAL, así que los ceros de cola se perdieron: `1.004` es la p. 40 y
    `1.01` la p. 100. Hay que RELLENAR la parte fraccionaria a 4 dígitos, no leerla tal cual."""
    m = re.match(r'(\d+)\.(\d+)$', (v or '').strip())
    if not m:
        return None
    frac = m.group(2)[:4].ljust(4, '0')
    return int(frac) or None


def build_massive(prefix='sn1.'):
    """Cada nº canónico DPR → (cst_paranum, título CST, página PTS de la concordancia)."""
    canon2para = {}
    for r in csv.DictReader(open('massive.tsv'), delimiter='\t'):
        if not (r['cst_code'] or '').startswith(prefix):
            continue
        pm = re.match(r'(\d+)', r['cst_paranum'] or '')
        m = re.match(r'SN(\d+)\.(\d+)(?:-(\d+))?', r['dpr_code'] or '')
        if not (pm and m):
            continue
        sam, a = m.group(1), int(m.group(2))
        b = int(m.group(3)) if m.group(3) else a
        pp_ = _pts_page(r['cst_p_page'])
        for k in range(a, b + 1):
            canon2para.setdefault(f'{sam}.{k}', (int(pm.group(1)), r['cst_sutta'], pp_))
    return canon2para


# ── lado PTS: marcadores § y texto del sutta (puede cruzar páginas) ─────────────
def build_pts_suttas(cur):
    """Los 271 suttas de S i en orden de lectura: `(page, line, nums, name, text)`.

    El texto va del marcador al siguiente (cruzando páginas: los suttas de S i se reparten en
    varias). En un marcador de rango (`§§ 4,5`) el bloque se PARTE por los submarcadores internos
    (`4.` / `5.` centrados a solas): si no, los dos suttas comparten todo el texto y el cotejo se
    cae — en S i 82 son dos batallas con desenlaces opuestos.
    """
    pages = {r['page_no']: (r['unitext'] or '').split('\n')
             for r in cur.execute('SELECT page_no,unitext FROM pages '
                                  'WHERE edition="mula" AND book_no=?', (SN1_BOOK,))}
    flat, marks, subs = [], [], []      # flat: [(page, line, texto)] en orden de lectura
    for pg in sorted(pages):
        for ln, line in enumerate(pages[pg], start=1):
            flat.append((pg, ln, line))
        page_text = '\n'.join(pages[pg])
        for ln, nums, name, tag in find_markers_sn1(page_text):
            marks.append((pg, ln, nums, name, tag))
        for ln, n in find_submarkers_sn1(page_text):
            subs.append((pg, ln, n))
    marks.sort(key=lambda m: (m[0], m[1]))
    subs.sort()
    pos = {(pg, ln): i for i, (pg, ln, _t) in enumerate(flat)}

    def text_of(i0, i1):
        return ' '.join(sh.tokens(' '.join(t for _p, _l, t in flat[i0:i1]))[:350])

    out = []
    for k, (pg, ln, nums, name, tag) in enumerate(marks):
        i0 = pos[(pg, ln)]
        i1 = pos[(marks[k + 1][0], marks[k + 1][1])] if k + 1 < len(marks) else len(flat)
        if len(nums) == 1:
            out.append({'page': pg, 'line': ln, 'num': nums[0], 'name': name,
                        'tag': tag, 'text': text_of(i0, i1), 'range': False})
            continue
        # bloque de rango: parte por el submarcador de cada nº (si está)
        inner = {n: (spg, sln) for spg, sln, n in subs if i0 < pos[(spg, sln)] < i1 and n in nums}
        cuts = [(n, pos[inner[n]] if n in inner else i0) for n in nums]
        for j, (n, start) in enumerate(cuts):
            end = cuts[j + 1][1] if j + 1 < len(cuts) else i1
            spg, sln = inner.get(n, (pg, ln))
            out.append({'page': spg, 'line': sln, 'num': n, 'name': name, 'tag': tag,
                        'text': text_of(start, end), 'range': True,
                        'submark': n in inner})
    return out, marks


def excel_entries():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['Nikaya']] or '') != 'SN' or str(row[ci['PTS Roman']] or '').strip().lower() != 'i':
            continue
        num = str(row[ci['Sutta #']]); name = str(row[ci['Sutta Name']] or '')
        m = re.search(r'\(SN\s+(\d+)\.(\d+)', name)
        canon = f'{m.group(1)}.{m.group(2)}' if m else num
        out.append({'num': num, 'canon': canon,
                    'name': re.sub(r'\(SN[^)]*\)', '', name).strip(),
                    'page': row[ci['PTS Page']], 'ref': str(row[ci['PTS Ref']] or ''),
                    'legacy': str(row[ci['Validation']] or '')})
    return out


_FOLD = str.maketrans({'ā': 'a', 'ī': 'i', 'ū': 'u', 'ṅ': 'n', 'ñ': 'n', 'ṇ': 'n',
                       'ṭ': 't', 'ḍ': 'd', 'ḷ': 'l', 'ṃ': 'm', 'ṁ': 'm'})
def _stem(t):
    n = re.sub(r'[^a-z]', '', re.sub(r'^[\d\s.,-]*', '', (t or '').lower()).translate(_FOLD))
    n = re.sub(r'(suttantam|suttam|suttani|vaggo).*$', '', n)
    return re.sub(r'[aiueom]+$', '', re.sub(r'(.)\1+', r'\1', n))


def name_agrees(a, b):
    x, y = _stem(a), _stem(b)
    if not x or not y or min(len(x), len(y)) < 3:
        return False
    return x == y or x.startswith(y) or y.startswith(x) or x in y or y in x


def check_structure(vaggas):
    """Coteja la estructura PTS hallada contra el front matter de Feer. Devuelve incidencias."""
    bad = []
    if len(vaggas) != len(FEER):
        bad.append(f'vaggas: {len(vaggas)} != {len(FEER)} (Feer)')
        return bad
    for k, v in enumerate(vaggas):
        sam, nm, fpg, exp = FEER[k]
        cnt = sum(len(x[2]) for x in v)
        nums = sorted(n for x in v for n in x[2])
        if v[0][0] != fpg:
            bad.append(f'vagga {k+1} (SN {sam} {nm}): arranca en p{v[0][0]}, Feer dice p{fpg}')
        if cnt != exp or [x for x in range(1, exp + 1) if x not in nums]:
            bad.append(f'vagga {k+1} (SN {sam} {nm}): {cnt} suttas, Feer dice {exp}')
    return bad


def main():
    dry = '--dry' in sys.argv
    do_all = '--all' in sys.argv
    n = int(sys.argv[sys.argv.index('--n') + 1]) if '--n' in sys.argv else 30
    only = set(sys.argv[sys.argv.index('--only') + 1].split(',')) if '--only' in sys.argv else None

    vri = build_vri_index()
    canon2para = build_massive()
    entries = excel_entries()
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    pts, marks = build_pts_suttas(cur)

    print('=' * 92)
    print('SN I — pipeline VRI (Excel→DPR→massive→cst_paranum→XML VRI) + estructura de Feer')
    print('=' * 92)
    bad = check_structure(build_vaggas(marks))
    print(f'Estructura PTS: {len(marks)} marcadores → {len(pts)} suttas; '
          f'{"CUADRA con Feer (28 vaggas, 271 suttas, páginas de arranque)" if not bad else "INCIDENCIAS:"}')
    for b in bad:
        print('   ', b)
    print(f'Excel S i: {len(entries)} filas | paranums en el XML VRI: {len(vri)}')
    if len(pts) != len(entries):
        print(f'!! desajuste PTS({len(pts)}) vs Excel({len(entries)}) — alineación por posición no válida')
        return

    tasks, no_cst, name_ok, page_ok = [], [], 0, 0
    for k, e in enumerate(entries):
        p = pts[k]                             # alineación por POSICIÓN (secuencia canónica)
        hit = canon2para.get(e['canon'])
        s = vri.get(hit[0]) if hit else None
        if not s:
            no_cst.append(e); continue
        if name_agrees(p['name'], s['title']):
            name_ok += 1
        if hit[2] and abs(hit[2] - p['page']) <= 1:
            page_ok += 1
        cst = ' '.join(sh.tokens(s['text'])[:350])
        tasks.append((e, p, cst, s['title']))

    print(f'\nAlineadas con texto PTS+CST: {len(tasks)}/{len(entries)} (sin CST: {len(no_cst)})')
    print(f'  comprobación cruzada — nombre marcador PTS ≡ título CST: {name_ok}/{len(tasks)} '
          f'({100*name_ok/max(1,len(tasks)):.0f}%)')
    print(f'  comprobación cruzada — página del marcador ≡ cst_p_page (±1): {page_ok}/{len(tasks)} '
          f'({100*page_ok/max(1,len(tasks)):.0f}%)')
    if dry:
        print('\n(dry-run: nada validado, nada escrito)')
        return

    run = [t for t in tasks if t[0]['num'] in only] if only else (tasks if do_all else tasks[:n])
    print(f'\nValidando {len(run)} (Gemini en vivo)...', flush=True)
    rows = []
    for k, (e, p, cst, ctitle) in enumerate(run, 1):
        res = validate_pair(p['text'], cst, e['name'], ctitle, 'SN', concordant=True)
        rows.append({'num': e['num'], 'canon': e['canon'], 'name': e['name'],
                     'cst_title': ctitle, 'legacy': e['legacy'], 'pts_name': p['name'],
                     'pts_page': p['page'], 'pts_line': p['line'], 'tag': p['tag'], **res})
        if k % 25 == 0:
            print(f'  ...{k}/{len(run)}', flush=True)
    prev = json.load(open(OUT)) if os.path.exists(OUT) else []
    merged = {r['num']: r for r in prev}
    merged.update({r['num']: r for r in rows})
    json.dump(list(merged.values()), open(OUT, 'w'), ensure_ascii=False, indent=1)
    est = Counter(r['estado'] for r in rows); val = Counter(r['validation'] for r in rows)
    print(f'\nEstado: {dict(est)} | Validation: {dict(val)}')
    print(f'Resultados → {OUT} ({len(merged)} en total; nada escrito al Excel).')


if __name__ == '__main__':
    main()
