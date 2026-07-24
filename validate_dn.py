#!/usr/bin/env python3
"""
Validate all Dīgha Nikāya entries against tipitaka.sqlite.

Uses:
- contents table for sutta title matching
- pages table for direct content verification
"""

import sqlite3
import re
from openpyxl import load_workbook

DB_PATH = '/home/jorge/Code/squashfs-root/src/data/tipitaka.sqlite'
EXCEL_PATH = '/home/jorge/Code/squashfs-root/PTS_Reference_Full_Canon.xlsx'

# DN book mapping (volume -> book_no in DB)
DN_BOOK_MAP = {1: 6, 2: 7, 3: 8}
DN_BOOK_NAMES = {6: 'DN I', 7: 'DN II', 8: 'DN III'}

def normalize(text):
    """Normalize Pali text for comparison."""
    if not text:
        return ""
    text = text.lower().strip()
    # Remove diacritics for fuzzy matching
    return text

def strip_diacritics(text):
    """Remove Pali diacritics for comparison."""
    replacements = {
        'ā': 'a', 'ī': 'i', 'ū': 'u', 'ṁ': 'm', 'ṃ': 'm',
        'ṅ': 'n', 'ñ': 'n', 'ṭ': 't', 'ḍ': 'd', 'ṇ': 'n',
        'ḷ': 'l', 'ṝ': 'r', 'ṝ': 'r', 'ś': 's', 'ṣ': 's',
        'Ā': 'A', 'Ī': 'I', 'Ū': 'U', 'Ṅ': 'N', 'Ñ': 'N',
        'Ṭ': 'T', 'Ḍ': 'D', 'Ṇ': 'N', 'Ḷ': 'L',
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text

def word_match_ratio(name, content, min_word_len=3):
    """Calculate what fraction of sutta name words appear in content."""
    name_clean = strip_diacritics(name.lower())
    content_clean = strip_diacritics(content.lower())
    
    # Extract words from name
    name_words = set()
    for w in re.split(r'[\s\-–—,;:.\[\]()]+', name_clean):
        w = w.strip()
        if len(w) >= min_word_len:
            name_words.add(w)
    
    if not name_words:
        return 0, set()
    
    matched = set()
    for w in name_words:
        if w in content_clean:
            matched.add(w)
    
    return len(matched) / len(name_words), matched

def load_excel_dn():
    """Load DN entries from Excel."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb['PTS Reference']
    
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nikaya = row[1]
        if nikaya != 'DN':
            continue
        
        entry = {
            'row_num': row[0],
            'nikaya': nikaya,
            'sutta_num': row[2],
            'sutta_name': row[3].strip() if row[3] else '',
            'pts_vol': row[5],
            'pts_roman': row[6],
            'pts_page': row[7],
            'pts_full': row[8],
            'sutta_raw': str(row[11]).strip() if row[11] else ''
        }
        entries.append(entry)
    
    return entries

def roman_to_int(roman):
    """Roman numeral to int."""
    mapping = {'i': 1, 'ii': 2, 'iii': 3, 'iv': 4, 'v': 5, 'vi': 6}
    return mapping.get(str(roman).strip().lower(), 0)

def validate_dn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    entries = load_excel_dn()
    print(f"Loaded {len(entries)} DN entries from Excel\n")
    
    # Load contents table for DN books (6, 7, 8)
    cur.execute("""
        SELECT book_no, page_no, section, title 
        FROM contents 
        WHERE book_no IN (6, 7, 8)
        ORDER BY book_no, page_no
    """)
    contents_rows = cur.fetchall()
    
    # Build index: (book_no, page_no) -> (section, title)
    contents_index = {}
    for row in contents_rows:
        key = (row['book_no'], row['page_no'])
        contents_index[key] = (row['section'], row['title'])
    
    # Also load actual page content for verification
    # Get first 200 chars of each page
    cur.execute("""
        SELECT book_no, page_no, head, substr(unitext, 1, 500) as text_preview
        FROM pages
        WHERE book_no IN (6, 7, 8) AND edition = 'mula'
    """)
    pages_index = {}
    for row in cur.fetchall():
        key = (row['book_no'], row['page_no'])
        pages_index[key] = (row['head'], row['text_preview'])
    
    print(f"DB contents: {len(contents_rows)} TOC entries for DN")
    print(f"DB pages: {len(pages_index)} pages for DN\n")
    
    results = []
    ok = warn = fail = 0
    
    for entry in entries:
        vol_num = roman_to_int(entry['pts_roman'])
        book_no = DN_BOOK_MAP.get(vol_num)
        page_no = entry['pts_page']
        sutta_name = entry['sutta_name']
        
        result = {**entry, 'book_no': book_no, 'db_page': page_no}
        
        if not book_no:
            result['status'] = 'ERROR'
            result['detail'] = f'Volume {entry["pts_roman"]} not in map'
            fail += 1
            results.append(result)
            continue
        
        # Check page exists in DB
        db_key = (book_no, page_no)
        
        if db_key not in pages_index:
            result['status'] = 'MISSING'
            result['detail'] = f'Page {page_no} not found in book {book_no}'
            fail += 1
            results.append(result)
            continue
        
        head, text_preview = pages_index[db_key]
        result['head'] = (head or '')[:100]
        result['text_preview'] = (text_preview or '')[:200]
        
        # Get first meaningful words from page content
        # Skip title page markers
        text_clean = text_preview or ''
        first_words = ' '.join(text_clean.split()[:15]) if text_clean else ''
        result['first_words'] = first_words[:150]
        
        # Check contents table for this page
        toc_entry = contents_index.get(db_key)
        if toc_entry:
            toc_section, toc_title = toc_entry
            result['toc_section'] = toc_section
            result['toc_title'] = toc_title
            
            # Compare sutta name with TOC title
            if toc_title:
                ratio, matched = word_match_ratio(sutta_name, toc_title)
                result['match_ratio_toc'] = ratio
                if ratio >= 0.5:
                    result['status'] = 'OK'
                    result['detail'] = f'TOC: {toc_title} ({ratio:.0%} match)'
                    ok += 1
                elif ratio >= 0.25:
                    result['status'] = 'PARTIAL'
                    result['detail'] = f'TOC: {toc_title} ({ratio:.0%} match - partial)'
                    warn += 1
                else:
                    result['status'] = 'MISMATCH'
                    result['detail'] = f'TOC: {toc_title} ({ratio:.0%} match - check)'
                    warn += 1
            else:
                result['status'] = 'NO_TOC_TITLE'
                result['detail'] = f'Section: {toc_section}, no title'
                warn += 1
        else:
            # No TOC entry for this page - check if this is expected
            # For DN, almost every sutta start should have a TOC entry
            result['status'] = 'NO_TOC'
            result['detail'] = f'No TOC entry for book={book_no} page={page_no}'
            result['toc_section'] = ''
            result['toc_title'] = ''
            warn += 1
        
        results.append(result)
    
    conn.close()
    return results, ok, warn, fail

def print_report(results, ok, warn, fail):
    """Print formatted validation report."""
    total = len(results)
    
    print(f"{'='*90}")
    print(f"  DĪGHA NIKĀYA — Content Validation Report")
    print(f"  Database: tipitaka.sqlite (PTS edition, 'mula')")
    print(f"{'='*90}")
    print(f"  Total: {total}  |  ✓ OK: {ok}  |  ⚠ Review: {warn}  |  ✗ Fail: {fail}")
    print(f"  Accuracy: {ok}/{total} = {100*ok/total:.1f}%")
    print(f"{'─'*90}")
    
    for r in results:
        s = r['status']
        marker = {
            'OK': '✓', 'PARTIAL': '⚡', 'MISMATCH': '⚠', 
            'NO_TOC': '?', 'NO_TOC_TITLE': '?',
            'MISSING': '✗', 'ERROR': '‼'
        }.get(s, '?')
        
        sutta_id = r['sutta_raw']
        pts = r['pts_full']
        
        print(f"\n  {marker} [{s}] {sutta_id}")
        print(f"     PTS: {pts}  |  Book {r['book_no']} (p.{r['db_page']})")
        
        if r.get('toc_title'):
            print(f"     TOC: [{r.get('toc_section','')}] {r['toc_title']}")
        
        if r.get('head'):
            h = r['head'].replace('\n', ' ')[:100]
            print(f"     Head: {h}")
        
        fw = r.get('first_words', '')
        if fw:
            fw_clean = fw.replace('\n', ' ')[:130]
            print(f"     Content: {fw_clean}...")
        
        if r.get('detail'):
            print(f"     → {r['detail']}")
    
    # Summary of mismatches
    mismatches = [r for r in results if r['status'] not in ('OK',)]
    if mismatches:
        print(f"\n{'='*90}")
        print(f"  Items requiring review: {len(mismatches)}")
        print(f"{'─'*90}")
        for r in mismatches:
            print(f"  [{r['status']}] {r['sutta_raw']:40s}  PTS: {r['pts_full']:10s}  → {r.get('detail','')[:100]}")

if __name__ == '__main__':
    results, ok, warn, fail = validate_dn()
    print_report(results, ok, warn, fail)
