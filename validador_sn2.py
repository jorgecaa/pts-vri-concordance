#!/usr/bin/env python3
"""Driver SN II (S ii, book_no=13) — pipeline VRI por concordancia + validador (Modelo B).

Diferencias propias de S ii frente a S i / S v:

1. **Marcadores `N (M) Nombre` sin puntos** (ver `sn2_markers.py`, pyparsing), con `N` = nº corrido
   dentro del saṃyutta y `M` = posición en el vagga. La estructura se coteja contra el front matter
   de Feer (`samyutta-vol-II-info.txt`): **10 saṃyuttas (XII–XXI), 27 vaggas, 286 suttas**
   (93/11/39/20/13/43/22/21/12/12) y la página de arranque de cada uno. Cuadra al 100%.
2. **La clave de emparejamiento es `(saṃyutta, nº corrido)`** — que es justo la notación que Feer
   propone en su introducción («XII. 25. 4» = saṃyutta, sutta, párrafo) y la que usa el `Sutta #`
   del Excel. No hace falta alineación posicional ni difusa.
3. **Los grupos peyyāla van en UNA fila de nombre colectivo** (`Jātisuttādidasakaṃ` = PTS 72–81,
   `Suvaṇṇanikkhasuttādiaṭṭhakaṃ` = 17.13–20, `Pitusuttādichakkaṃ` = 17.38–43), misma convención
   que en SN V: 257 filas cubren los 286 suttas. Por eso `--faltantes` lista suttas sin fila PROPIA
   y eso NO es incompletitud — son los miembros de esos grupos.
4. **Cuando el CST agrupa lo que PTS numera** (antara-peyyālaṃ, SN 12.83–93), el ítem `(N)` del
   grupo CST se casa con la posición `(M)` del marcador PTS: `build_cst_group_items`.

Uso: python3 validador_sn2.py [--dry] [--all] [--n N] [--only 12.1,17.5] [--faltantes]
"""
import csv, os, re, sqlite3, sys, json
import xml.etree.ElementTree as ET
from collections import Counter

import sutta_hash as sh
from openpyxl import load_workbook

from validador import validate_pair
from validador_sn1 import _pts_page, _TEXT_RENDS, build_vri_index as _build_vri
from sn2_markers import find_markers_sn2, collect_suttas

VRI = '/tmp/tipitaka-xml/romn/s0302m.mul.xml'
OUT = 'validador_sn2.json'
DB = 'src/data/tipitaka.sqlite'
SN2_BOOK = 13
XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# Estructura de S ii según el front matter de Feer: (saṃyutta, nombre, nº de suttas, página)
FEER = [(12, 'Nidāna', 93, 1), (13, 'Abhisamaya', 11, 133), (14, 'Dhātu', 39, 140),
        (15, 'Anamatagga', 20, 178), (16, 'Kassapa', 13, 194), (17, 'Lābhasakkāra', 43, 225),
        (18, 'Rāhula', 22, 244), (19, 'Lakkhaṇa', 21, 254), (20, 'Opamma', 12, 262),
        (21, 'Bhikkhu', 12, 273)]


# Correspondencia estructural verificada a mano (S ii 130–133 y el XML VRI): el *antara-peyyālaṃ*
# de PTS son 12 suttas numerados (82 Satthā, 83 Sikkhā … 93 Appamādo — su propio uddāna dice
# «Appamādena dvādasāti»), mientras el CST reúne del 83 al 93 en un solo bloque
# `2-12. Sikkhāsuttādipeyyālaekādasakaṃ` cuyos ítems (2)…(12) son exactamente la posición `(M)`
# del marcador PTS. Se fija por título porque hay otros grupos con los mismos nºs de ítem.
CST_GROUP = {(12, k): '2-12. Sikkhāsuttādipeyyālaekādasakaṃ' for k in range(83, 94)}


def build_vri_index():
    return _build_vri(VRI)


def build_cst_group_items(path=VRI):
    """`(título del grupo, nº de ítem)` → texto, para los grupos CST cuyos miembros van como
    ítems `(N)` dentro de un solo `subhead`.

    Hace falta en el *antara-peyyālaṃ*: PTS imprime `83 (2) Sikkhā` … `93 (12) Appamādo` como 11
    suttas numerados, mientras el CST los mete en `2-12. Sikkhāsuttādipeyyālaekādasakaṃ` con los
    ítems `(2)`…`(12)`. La **posición `(M)` del marcador PTS es el nº de ítem del CST**, así que el
    emparejamiento es exacto y mecánico (no hay que adivinar nada).
    """
    body = ET.parse(path).getroot().find('.//body')
    out, title = {}, None
    for el in body.iter():
        if el.tag != 'p':
            continue
        rend = el.get('rend')
        if rend == 'subhead':
            title = ''.join(el.itertext()).strip()
        elif rend in _TEXT_RENDS and title:
            t = ''.join(el.itertext()).strip()
            m = re.match(r'[‘"\(]*\((\d+)\)', t)
            if m:
                out.setdefault((title, int(m.group(1))), '')
                out[(title, int(m.group(1)))] += ' ' + t
    return out


def build_massive(prefix='sn2.'):
    """`(saṃyutta, nº corrido)` → `(cst_paranum, título CST, página PTS del concordance)`."""
    out = {}
    for r in csv.DictReader(open('massive.tsv'), delimiter='\t'):
        if not (r['cst_code'] or '').startswith(prefix):
            continue
        pm = re.match(r'(\d+)', r['cst_paranum'] or '')
        m = re.match(r'SN(\d+)\.(\d+)(?:-(\d+))?', r['dpr_code'] or '')
        if not (pm and m):
            continue
        sam, a = int(m.group(1)), int(m.group(2))
        b = int(m.group(3)) if m.group(3) else a
        for k in range(a, b + 1):
            out.setdefault((sam, k), (int(pm.group(1)), r['cst_sutta'], _pts_page(r['cst_p_page'])))
    return out


def build_pts_suttas(cur):
    """Los 286 suttas de S ii, con `sam` (nº de saṃyutta XII–XXI) y el texto acotado."""
    pages = {r['page_no']: (r['unitext'] or '').split('\n')
             for r in cur.execute('SELECT page_no,unitext FROM pages '
                                  'WHERE edition="mula" AND book_no=?', (SN2_BOOK,))}
    flat, rows = [], []
    for pg in sorted(pages):
        for ln, line in enumerate(pages[pg], start=1):
            flat.append((pg, ln, line))
        for ln, nums, names, tag in find_markers_sn2('\n'.join(pages[pg])):
            rows.append((pg, ln, nums, names, tag))
    suttas, _sams = collect_suttas(rows)
    pos = {(pg, ln): i for i, (pg, ln, _t) in enumerate(flat)}

    # el texto de cada sutta va de su marcador al del siguiente (en orden de lectura)
    ordered = sorted(suttas, key=lambda s: (pos[(s['page'], s['line'])], s['num']))
    out = {}
    for k, s in enumerate(ordered):
        i0 = pos[(s['page'], s['line'])]
        i1 = len(flat)
        for nxt in ordered[k + 1:]:
            j = pos[(nxt['page'], nxt['line'])]
            if j > i0:
                i1 = j
                break
        sam = FEER[s['sam_idx'] - 1][0]
        rec = dict(s)
        rec['sam'] = sam
        rec['text'] = ' '.join(sh.tokens(' '.join(t for _p, _l, t in flat[i0:i1]))[:350])
        out[(sam, s['num'])] = rec
    return out


def excel_entries(roman='ii'):
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['Nikaya']] or '') != 'SN' or str(row[ci['PTS Roman']] or '').strip().lower() != roman:
            continue
        num = str(row[ci['Sutta #']]); name = str(row[ci['Sutta Name']] or '')
        m = re.match(r'(\d+)\.(\d+)', num)
        sam, inner = (int(m.group(1)), int(m.group(2))) if m else (None, None)
        out.append({'num': num, 'sam': sam, 'inner': inner,
                    'name': re.sub(r'\(SN[^)]*\)', '', name).strip(),
                    'page': row[ci['PTS Page']], 'ref': str(row[ci['PTS Ref']] or ''),
                    'legacy': str(row[ci['Validation']] or '')})
    return out


def main():
    dry = '--dry' in sys.argv
    do_all = '--all' in sys.argv
    n = int(sys.argv[sys.argv.index('--n') + 1]) if '--n' in sys.argv else 30
    only = set(sys.argv[sys.argv.index('--only') + 1].split(',')) if '--only' in sys.argv else None

    vri = build_vri_index()
    massive = build_massive()
    entries = excel_entries()
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    pts = build_pts_suttas(cur)

    print('=' * 92)
    print('SN II — pipeline VRI + estructura de Feer (samyutta-vol-II-info.txt)')
    print('=' * 92)
    print(f'PTS (BD libro 13): {len(pts)} suttas | Feer: {sum(f[2] for f in FEER)}')
    print(f'Excel marcado «S ii»: {len(entries)} filas | paranums en el XML VRI: {len(vri)}')

    mislabel = [e for e in entries if e['sam'] not in dict((f[0], 1) for f in FEER)]
    if mislabel:
        print(f'\n!! {len(mislabel)} filas con saṃyutta ajeno a S ii (volumen mal puesto): '
              f'{", ".join(e["num"] for e in mislabel)}')
    faltan = [(sam, k) for sam, _nm, exp, _pg in FEER for k in range(1, exp + 1)
              if not any(e['sam'] == sam and e['inner'] == k for e in entries)]
    if faltan:
        print(f'!! {len(faltan)} suttas de S ii SIN fila en el Excel: '
              + ', '.join(f'{a}.{b}' for a, b in faltan[:6]) + ' …')
    if '--faltantes' in sys.argv:
        print('\nDetalle de los que faltan (marcador PTS + mapeo CST):')
        for sam, k in faltan:
            p = pts.get((sam, k)); h = massive.get((sam, k))
            print(f'  SN {sam}.{k:<3} PTS: '
                  f'{"p%-3d L%-2d %-26s" % (p["page"], p["line"], (p["name"] or "(sin nombre)")[:26]) if p else "NO HALLADO":38}'
                  f' CST: {h[1][:34] if h else "(no en massive)"}')
        return

    groups = build_cst_group_items()
    tasks, no_cst, name_ok, page_ok = [], [], 0, 0
    for e in entries:
        p = pts.get((e['sam'], e['inner']))
        h = massive.get((e['sam'], e['inner']))
        s = vri.get(h[0]) if h else None
        if p and not s and (e['sam'], e['inner']) in CST_GROUP:
            # el CST agrupa lo que PTS numera: el ítem (N) del grupo es la posición (M) del
            # marcador PTS. El grupo se FIJA por título (no se busca): hay varios grupos con los
            # mismos números de ítem y buscar por parecido de texto elige el equivocado.
            gt = CST_GROUP[(e['sam'], e['inner'])]
            txt = groups.get((gt, p['pos']))
            if txt:
                s = {'title': f'{gt} — ítem ({p["pos"]})', 'text': txt}
        if not (p and s):
            no_cst.append(e); continue
        if h and h[2] and abs(h[2] - p['page']) <= 1:
            page_ok += 1
        if _stem(p['name']) and _stem(p['name'])[:4] in _stem(s['title']):
            name_ok += 1
        tasks.append((e, p, ' '.join(sh.tokens(s['text'])[:350]), s['title']))

    print(f'\nAlineadas con texto PTS+CST: {len(tasks)}/{len(entries)} (sin par: {len(no_cst)})')
    if tasks:
        print(f'  cruzada — página del marcador ≡ cst_p_page (±1): {page_ok}/{len(tasks)} '
              f'({100*page_ok/len(tasks):.0f}%)')
        print(f'  cruzada — raíz del nombre PTS ⊂ título CST: {name_ok}/{len(tasks)} '
              f'({100*name_ok/len(tasks):.0f}%)')
    if no_cst:
        print('  sin par:', ', '.join(e['num'] for e in no_cst[:14]))
    if dry:
        print('\n(dry-run: nada validado, nada escrito)')
        return

    run = [t for t in tasks if t[0]['num'] in only] if only else (tasks if do_all else tasks[:n])
    print(f'\nValidando {len(run)} (Gemini en vivo)...', flush=True)
    rows = []
    for k, (e, p, cst, ctitle) in enumerate(run, 1):
        res = validate_pair(p['text'], cst, e['name'], ctitle, 'SN', concordant=True)
        rows.append({'num': e['num'], 'name': e['name'], 'cst_title': ctitle,
                     'legacy': e['legacy'], 'pts_name': p['name'], 'pts_page': p['page'],
                     'pts_line': p['line'], 'from_range': p['from_range'], **res})
        if k % 25 == 0:
            print(f'  ...{k}/{len(run)}', flush=True)
    prev = json.load(open(OUT)) if os.path.exists(OUT) else []
    merged = {r['num']: r for r in prev}
    merged.update({r['num']: r for r in rows})
    json.dump(list(merged.values()), open(OUT, 'w'), ensure_ascii=False, indent=1)
    print(f'\nEstado: {dict(Counter(r["estado"] for r in rows))} | '
          f'Validation: {dict(Counter(r["validation"] for r in rows))}')
    print(f'Resultados → {OUT} ({len(merged)} en total; nada escrito al Excel).')


_FOLD = str.maketrans({'ā': 'a', 'ī': 'i', 'ū': 'u', 'ṅ': 'n', 'ñ': 'n', 'ṇ': 'n',
                       'ṭ': 't', 'ḍ': 'd', 'ḷ': 'l', 'ṃ': 'm', 'ṁ': 'm'})
def _stem(t):
    n = re.sub(r'[^a-z]', '', re.sub(r'^[\d\s.,()-]*', '', (t or '').lower()).translate(_FOLD))
    n = re.sub(r'(suttantam|suttam|suttani|vaggo).*$', '', n)
    return re.sub(r'(.)\1+', r'\1', n)


if __name__ == '__main__':
    main()
