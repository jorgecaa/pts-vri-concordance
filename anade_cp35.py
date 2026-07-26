#!/usr/bin/env python3
"""
anade_cp35 — añade al Excel la fila que le falta al Cariyāpiṭaka: la **35ª cariyā**,
`Mahālomahaṃsacariyā`.

No es una divergencia entre ediciones ni una conjetura: las dos fuentes la traen y el Excel no.

- **CST** (`s0512m`): el Yudhañjayavaggo tiene **quince** `subhead` y el decimoquinto es
  `15. Mahālomahaṃsacariyā`. Los tres vaggas son 10 + 10 + 15 = **35**.
- **El impreso** (BD, libro 42): subtitula `15 Mahālomahaṃsacariyaṃ` en su p35 y cierra con
  `sabbattha samako homi esā me upekkhāpāramī ti` y el colofón de la obra
  (`Cariyāpiṭakaṃ niṭṭhitaṃ`). Los subtítulos impresos son **35**, con la misma partición.
- **El Excel** tenía **34**, terminando en `13.34 Ekarājacariya`.

La página que se le asigna es la **101**, la misma que el Excel da a `13.34 Ekarāja`: en el impreso
las dos comparten página (la BD las pone las dos en su p35) y el Cariyāpiṭaka acaba ahí. Es el único
valor compatible con la serie, y de todos modos **la referencia de esta obra no es la página** sino
la cariyā (regla (5-ter)): `Cp 3.15`.

Es la operación inversa de los dos borrados que ya hizo este proyecto —`12.74` en S ii y
`2.19 «Adhikaraṇa 9 + Andha»` en A i—, y como aquéllos **renumera la columna `#`**. El total pasa de
6.097 a **6.098** filas.

Uso: python3 anade_cp35.py [--dry]
"""
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
ANCLA = '13.34'
NUEVA = {
    'Nikaya': 'KN', 'Sutta #': '13.35', 'Sutta Name': '(KN 15.35) Cp 35 Mahālomahaṃsacariya',
    'Section': 'Cariyāpiṭaka (Cp)', 'PTS Vol': 'Cp', 'PTS Page': 101, 'PTS Ref': 'Cp 3.15',
    'Type': 'Sutta', 'Validation': 'UNVERIFIED', 'Estado': 'PENDIENTE', 'Raw ID': 'KN 13.35',
    'Detail': 'Fila añadida: el CST y el impreso traen 35 cariyās (10+10+15) y el Excel tenía 34',
}


def main():
    dry = '--dry' in sys.argv
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    fila = next((r for r in range(2, ws.max_row + 1)
                 if str(ws.cell(r, ci['Sutta #']).value) == ANCLA), None)
    if fila is None:
        print(f'⚠ no encuentro la fila {ANCLA}: no se toca nada')
        return
    if any(str(ws.cell(r, ci['Sutta #']).value) == NUEVA['Sutta #']
           for r in range(2, ws.max_row + 1)):
        print(f'⚠ la fila {NUEVA["Sutta #"]} ya existe: no se toca nada')
        return
    print(f'insertar tras la fila {fila} (#{ws.cell(fila, ci["#"]).value} = {ANCLA}) · '
          f'{ws.max_row - 1} filas → {ws.max_row}')
    if dry:
        print('(dry-run: nada escrito)')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-cp35.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    ws.insert_rows(fila + 1)
    for col, val in NUEVA.items():
        ws.cell(fila + 1, ci[col]).value = val
    for r in range(2, ws.max_row + 1):          # `#` corrido, como en los dos borrados anteriores
        ws.cell(r, ci['#']).value = r - 1
    wb.save(XLSX)
    print(f'añadida {NUEVA["Sutta #"]} · {ws.max_row - 1} filas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
