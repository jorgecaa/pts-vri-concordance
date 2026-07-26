#!/usr/bin/env python3
"""
kn_suttanipata — el Suttanipāta: 73 filas. **La transcripción trae muchos errores** y eso decide
el método.

### Lo que la transcripción tiene mal, medido

- **214 de 7.731 palabras distintas (2,8 %) no pasan la fonotáctica** ni después de normalizar, y
  son **truncamientos**: `adhivāsetv`, `anagāriy`, `anuttar`, `appicch`. Son palabras partidas al
  final de línea **sin marcar el guion**, así que no se pueden recomponer como en A ii-A v.
- **87 de los 1.149 números de verso no están** en el texto (`5, 15, 35, 41, 55, 62, 81…`).
- Y una diferencia que no es error de nadie: PTS escribe la niggahīta final como **`ñ` de sandhi**
  —`ahañ ca`, `ajjhattañ ca`, `kiñ ca`— **194 veces**, y el CST la lleva siempre a `ṃ` (0 `ñ`
  finales contra 4.604 `ṃ`). Se añadió a `pali_norm` como regla medida: sin ella, 194 palabras
  serían discrepancias inventadas. Bajó el residuo fonotáctico del **4,9 % al 2,8 %**.

### Por qué el número de verso NO puede ser la clave

El CST numera **1.155** versos y PTS **1.149**, y el desfase **se acumula**: el verso 553 del CST
está impreso en Sn 109 y la fila que le corresponde declara la página 102. Buscar «el verso `a` del
CST» en la numeración de PTS empieza bien y se descuadra siete páginas al llegar al Mahāvagga. Son
**dos numeraciones distintas**, no una con erratas.

### Lo que sí decide

**El CST tiene exactamente 73 suttas** (contados por `subhead`) y el Excel exactamente 73 filas, en
el mismo orden y con el mismo reparto en cinco vaggas. La fila `k` es el sutta `k`, y eso se
comprueba con dos señales independientes:

1. **El nombre** — el del Excel contra el `subhead` del CST. 72 de 73.
2. **La cobertura de contenido** entre las páginas de la fila y el texto del sutta en el CST, que
   resiste bien los truncamientos porque va por **n-gramas de caracteres**.

⚠️ La comparación de nombres **no usa `an_names`**: ese módulo está afinado para AN, donde los
nombres son largos, y en KN se rompe con los cortos — `cst_parts('Kāmasuttaṃ')` devuelve `'k'` y
`pts_variants('Kāma')` devuelve **lista vacía**. No se toca (lo usan cinco volúmenes ya cerrados):
la comparación de KN va aparte.

### La notación (regla (5-ter))

`Sn <vagga>.<sutta>` — `Sn 4.1`. **No** se da rango de versos de PTS: con 87 números ausentes del
texto y las dos numeraciones divergiendo, sería inventarlo. El rango del CST queda en la `VRI Ref`.

Uso: python3 kn_suttanipata.py [--dry]
"""
import re
import shutil
import sqlite3
import sys
from datetime import datetime

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
BOOK, STEM = 26, 's0505m'
COV_MIN = 0.55
ULTIMA = 223

# ⚠️ Dos páginas INTERCAMBIADAS en el Excel, y el impreso lo dice sin lugar a duda: Sn **151** abre
# con `IV. AṬṬHAKAVAGGA. / 1. Kāmasutta.` y el verso `766. Kāmaṃ kāmayamānassa…`, mientras que el
# Guhaṭṭhaka empieza en la **152**. El Excel las tenía al revés (`5.4.1`→152, `5.4.2`→151), lo que
# además hacía retroceder la serie de páginas.
PAGINA_CORREGIDA = {'5.4.1': 151, '5.4.2': 152}


def _limpia(t):
    """Quita del nombre del Excel los prefijos de catálogo: `(KN 5.4) Snp 1.4 Kasibhāradvāja`."""
    t = re.sub(r'^\s*\(KN[^)]*\)\s*', '', t or '')
    return re.sub(r'^\s*(Snp|Ud|Iti|Dhp|Khp|Thag|Thig|Vv|Pv)\s+[\d.]+\s*', '', t).strip()


def raiz(t):
    """Raíz comparable de un nombre de KN: sin ordinal, sin `-suttaṃ`/`-gāthā`, sin desinencia."""
    t = ap.fold(re.sub(r'^[\d.\s]+', '', t or '')).strip()
    t = re.sub(r'\s*(sutantam|sutani|sutam|suta|manavapucha|pucha|gatha)$', '', t).strip()
    return re.sub(r'[aiueom]+$', '', t)


def casa_nombre(excel, subhead):
    """¿El nombre del Excel y el `subhead` del CST son el mismo? Se prueban **las variantes**.

    El Excel escribe la del CST entre paréntesis cuando difieren —`Kapila (Dhammacariya)`— y a
    veces la buena es ésa.
    """
    va = [raiz(x) for x in re.split(r'[()]', _limpia(excel)) if raiz(x)]
    vb = raiz(subhead)
    if len(vb) < 3:
        return any(a == vb for a in va)
    return any(len(a) >= 3 and (a == vb or a.startswith(vb) or vb.startswith(a)) for a in va)


def suttas_cst():
    """`[(subhead, primer verso, último verso, texto)]` — los 73 suttas, agrupando por `subhead`."""
    out = []
    for x in ap.cst_unidades(STEM):
        if not out or out[-1][0] != x['subhead']:
            out.append([x['subhead'], x['pn'], x['pn'], ''])
        out[-1][2] = x['pn']
        out[-1][3] += ' ' + x['texto']
    return [(a, b, c, ap.fold(d)) for a, b, c, d in out]


def main():
    dry = '--dry' in sys.argv
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    cst = suttas_cst()
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    fs = [{'num': str(r[ci['Sutta #']]), 'name': str(r[ci['Sutta Name']] or ''),
           'page': r[ci['PTS Page']], 'estado': r[ci['Estado']]}
          for r in ws.iter_rows(min_row=2, values_only=True)
          if r[ci['Nikaya']] == 'KN' and str(r[ci['PTS Vol']]) == 'Sn']
    print(f'CST {len(cst)} suttas · Excel {len(fs)} filas')
    if len(cst) != len(fs):
        print('⚠ los recuentos no coinciden: no se escribe nada')
        return
    res, malas, prev = {}, [], 0
    for k, f in enumerate(fs):
        sub, a, b, tc = cst[k]
        pg = PAGINA_CORREGIDA.get(f['num'], f['page'])
        sig = fs[k + 1] if k + 1 < len(fs) else None
        hasta = (PAGINA_CORREGIDA.get(sig['num'], sig['page'])
                 if sig and isinstance(sig['page'], int) else ULTIMA)
        txt = ' '.join(ap.fold(x['unitext'] or '') for x in conn.execute(
            'SELECT unitext FROM pages WHERE edition="mula" AND book_no=? AND page_no BETWEEN ? '
            'AND ?', (BOOK, pg, max(pg, hasta))))
        cov = R.cobertura(txt, tc)
        nom = casa_nombre(f['name'], sub)
        mono = isinstance(pg, int) and pg >= prev
        if isinstance(pg, int):
            prev = pg
        m = re.match(r'5\.(\d+)\.(\d+)$', f['num'])
        if cov >= COV_MIN and mono and m:
            v, n = m.group(1), m.group(2)
            res[f['num']] = (f'{STEM}:{a}-{b}', f'Sn {v}.{n}',
                             f'Suttanipāta: sutta {k + 1} de 73 → «{sub}», versos {a}-{b} del CST; '
                             f'cobertura {cov:.2f}'
                             + ('' if nom else '. ⚠ el nombre del Excel no casa con el subhead')
                             + (f'. PÁGINA CORREGIDA: el Excel decía {f["page"]} y el impreso '
                                f'abre en la {pg}' if f['num'] in PAGINA_CORREGIDA else ''),
                             PAGINA_CORREGIDA.get(f['num']))
        else:
            malas.append((f['num'], _limpia(f['name'])[:26], sub, f'cov {cov:.2f}'))
        if not nom:
            print(f"   ⚠ nombre: «{_limpia(f['name'])}» vs «{sub}»")
    print(f'\n{len(res)}/{len(fs)} firmadas')
    for x in malas:
        print('   ✗', x)
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-sn.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'KN' or str(ws.cell(r, ci['PTS Vol']).value) != 'Sn':
            continue
        num = str(ws.cell(r, ci['Sutta #']).value)
        if num not in res or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, ref, det, pag_new = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['PTS Ref']).value = ref
        if pag_new:
            ws.cell(r, ci['PTS Page']).value = pag_new
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
