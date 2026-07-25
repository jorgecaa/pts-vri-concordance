#!/usr/bin/env python3
"""reid_sn2.py — re-identifica cada fila de S ii por NOMBRE+PÁGINA y detecta `Sutta #` erróneos.

**Por qué hace falta.** El primer pase de `validador_sn2.py` supuso que el `Sutta #` del Excel era
el nº corrido del sutta dentro del saṃyutta (la notación que Feer propone). Es cierto en la mayor
parte del volumen, pero **falso en SN 17 y SN 18**: ahí el Excel salta los grupos peyyāla y su
numeración queda comprimida y desplazada (la fila que llama `17.21` se titula *Chavi* y está en la
p. 237, que es el sutta **17.28** tanto en PTS como en DPR/CST). Emparejar por ese número compara
un par PTS↔CST correcto entre sí pero **ajeno a la fila**, y el APPROVE queda mal adjudicado.

**Criterio de re-identificación.** Dentro de una fila, `Sutta Name` y `PTS Page` concuerdan entre
sí y con las dos fuentes independientes; el `Sutta #` es el dato que se corrompió. Así que la
identidad se resuelve por **nombre del marcador en la página declarada** (±1), y se exige que el
**ancla `cst_p_page`** del concordance confirme el nº resultante. Solo con esos dos testigos de
acuerdo se propone cambiar el `Sutta #`.

Uso:
    python3 reid_sn2.py            # informe
    python3 reid_sn2.py --write    # corrige el `Sutta #` de las filas con doble evidencia
"""
import datetime, re, shutil, sqlite3, sys
from collections import Counter

from openpyxl import load_workbook

from validador_sn2 import (DB, XLSX, FEER, build_massive, build_pts_suttas, excel_entries,
                           _stem)

# Nombres de GRUPO al estilo CST: «Pitusuttādichakkaṃ» = «los seis, Pitu-sutta etc.» (el rango
# 38–43), «Sikkhāsuttādipeyyālaekādasakaṃ» = «los once, Sikkhā-sutta etc.». Para casarlos con el
# marcador PTS hay que quedarse con la CABEZA del nombre (Pitu-, Sikkhā-).
_GROUP_TAIL = re.compile(r'sutt[aā]di.*$|peyy[aā]l.*$', re.I)
_CASE_END = re.compile(r'[aiueom]+$')


def head_of(name):
    return _GROUP_TAIL.sub('', name or '') or (name or '')


def name_score(a, b):
    """Afinidad de nombres: longitud del prefijo común, con bonus si uno contiene al otro.
    Un umbral de prefijo fijo no sirve — «Kusalamūla» y «Kusaladhamma» comparten 6 letras y son
    suttas distintos —, así que se puntúa y se exige que el mejor candidato sea único."""
    # se recorta también la terminación de caso: Feer declina («Pitā») donde el CST cita el tema
    # («Pitu-suttādi-chakkaṃ»), y sin esto «pitu» y «pita» no se reconocen
    x, y = (_CASE_END.sub('', _stem(head_of(t))) for t in (a, b))
    if not x or not y:
        return 0
    lcp = 0
    for c, d in zip(x, y):
        if c != d:
            break
        lcp += 1
    if x == y:
        return 100
    if x.startswith(y) or y.startswith(x):
        return 50 + lcp
    return lcp if lcp >= 5 else 0


def _best(cands):
    """`[(dist, score, num, payload)]` → el mejor si es ÚNICO en su (dist, score)."""
    if not cands:
        return None
    cands = sorted(cands, key=lambda c: (c[0], -c[1]))
    top = cands[0]
    ties = [c for c in cands if c[0] == top[0] and c[1] == top[1]]
    return top if len(ties) == 1 else None


def reidentify():
    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
    pts = build_pts_suttas(conn.cursor())
    massive = build_massive()
    entries = excel_entries()

    by_page = {}
    for (sam, num), p in pts.items():
        by_page.setdefault((sam, p['page']), []).append((num, p))
    # índice del concordance por página PTS: los `Sutta Name` del Excel son de estilo CST
    # ("Ekaputtaka" ↔ "Ekaputtakasuttaṃ"), así que el título CST resuelve lo que el nombre
    # abreviado de Feer no alcanza.
    cst_by_page = {}
    for (sam, num), h in massive.items():
        if h[2] is not None:
            cst_by_page.setdefault((sam, h[2]), []).append((num, h[1]))

    rows = []
    for e in entries:
        if e['sam'] not in [f[0] for f in FEER]:
            continue                                    # filas de otro volumen
        true_num = p = None
        # VÍA 1 — dar por bueno el nº que declara el Excel si su marcador o su título CST casan
        # con el nombre de la fila. Va PRIMERO a propósito: buscar por página antes de esto
        # confunde los pares «X (1)» / «X (2)» (S ii 87–88 Mahārukkho, 135–136 Pathavī), donde
        # el nº era correcto y lo desfasado era la PÁGINA.
        q = pts.get((e['sam'], e['inner']))
        h0 = massive.get((e['sam'], e['inner']))
        if q and (name_score(e['name'], q['name']) or (h0 and name_score(e['name'], h0[1]))):
            true_num, p = e['inner'], q
        if true_num is None:
            # VÍA 2 — nombre del marcador en la página declarada (±1)
            cands = [(abs(d), sc, num, qq)
                     for d in (0, 1, -1)
                     for num, qq in by_page.get((e['sam'], (e['page'] or 0) + d), [])
                     for sc in [name_score(e['name'], qq['name'])] if sc]
            best = _best(cands)
            true_num, p = (best[2], best[3]) if best else (None, None)
        if true_num is None:
            # VÍA 3 — el nº declarado, si su marcador y el ancla caen en la página declarada
            # (Feer abrevia: «Loko» vs «Loka», «Samaṇa-brāhmaṇā (2)» vs «Dutiyasamaṇabrāhmaṇa»)
            if q and isinstance(e['page'], int) and abs(q['page'] - e['page']) <= 1 \
                    and h0 and h0[2] is not None and abs(h0[2] - q['page']) <= 1:
                true_num, p = e['inner'], q
        if true_num is None:
            # VÍA 4 — el título CST en la página declarada (los `Sutta Name` son de estilo CST)
            hits = [(abs(d), sc, num, title)
                    for d in (0, 1, -1)
                    for num, title in cst_by_page.get((e['sam'], (e['page'] or 0) + d), [])
                    for sc in [name_score(e['name'], title)] if sc]
            best = _best(hits)
            if best:
                true_num = best[2]
                p = pts.get((e['sam'], true_num))
        anchor = None
        if true_num is not None:
            h = massive.get((e['sam'], true_num))
            anchor = h[2] if h else None
        rows.append({'num': e['num'], 'sam': e['sam'], 'inner': e['inner'], 'name': e['name'],
                     'page': e['page'], 'true_num': true_num,
                     'pts_name': p['name'] if p else None,
                     'pts_page': p['page'] if p else None, 'anchor': anchor,
                     'ok': true_num == e['inner'],
                     # el ancla del concordance debe respaldar el nº re-identificado
                     'corroborated': (true_num is not None and anchor is not None
                                      and abs(anchor - (p['page'] if p else 0)) <= 1)})
    return rows


def main():
    write = '--write' in sys.argv
    rows = reidentify()
    ok = [r for r in rows if r['ok']]
    unresolved = [r for r in rows if r['true_num'] is None]
    wrong = [r for r in rows if r['true_num'] is not None and not r['ok']]
    fixable = [r for r in wrong if r['corroborated']]

    print('=' * 96)
    print('RE-IDENTIFICACIÓN DE S ii POR NOMBRE+PÁGINA')
    print('=' * 96)
    print(f'filas: {len(rows)} | `Sutta #` correcto: {len(ok)} | erróneo: {len(wrong)} '
          f'| sin resolver por nombre: {len(unresolved)}')
    print(f'de los erróneos, con el ancla del concordance a favor: {len(fixable)}')
    if wrong:
        print('\n`Sutta #` a corregir:')
        for r in sorted(wrong, key=lambda x: (x['sam'], x['inner'])):
            flag = '' if r['corroborated'] else '   (SIN respaldo del concordance — no se toca)'
            print(f"  {r['num']:>7} «{r['name'][:22]:22}» p{r['page']:>3} → "
                  f"{r['sam']}.{r['true_num']:<3} (marcador «{(r['pts_name'] or '')[:20]:20}» "
                  f"p{r['pts_page']}, ancla p{r['anchor']}){flag}")
    if unresolved:
        print(f'\nsin resolver por nombre ({len(unresolved)}) — se dejan como están:')
        for r in unresolved[:20]:
            print(f"  {r['num']:>7} «{r['name'][:26]:26}» p{r['page']}")
    print('\npor saṃyutta:', dict(Counter(r['sam'] for r in wrong)))

    if not write:
        print('\n(informe, nada escrito — usa --write)')
        return
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}
    target = {r['num']: r for r in fixable}
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-sn2reid.xlsx'
    shutil.copy(XLSX, bak); print('backup →', bak)
    n = 0
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'SN' or str(v[ci['PTS Roman']] or '').strip().lower() != 'ii':
            continue
        r = target.get(str(v[ci['Sutta #']]))
        if not r:
            continue
        ws.cell(row=ridx, column=ci['Sutta #'] + 1).value = f"{r['sam']}.{r['true_num']}"
        n += 1
    wb.save(XLSX)
    print(f'Corregidos {n} `Sutta #` en {XLSX}.')


if __name__ == '__main__':
    main()
