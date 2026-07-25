#!/usr/bin/env python3
"""
an1_eka_cola — resuelve la **cola del Eka** (vaggas 13-21 de PTS) por CONTENIDO.

En la mitad limpia (`an1_eka_duka.py`) las dos coordenadas del Excel bastan: el `DPR Ref` de tres
niveles da el marcador impreso y el `Sutta #` **es** el paranum del CST (comprobado en las 91 filas
cerradas: la diferencia `Sutta # − paranum` es **0** en las 91). En la cola eso se rompe, y de una
manera que conviene ver antes de decidir nada:

    fila      DPR   marcador   el CST que dice el `Sutta #`   el CST que dice el CONTENIDO
    1.306    17.1        1              306                            298
    1.318    18.3        3              318                            310
    1.394    20.2        2              394                            382
    1.616-627 21.47     47              616                            600

El `Sutta #` **deriva** del paranum, y la deriva crece: −8 en el vagga 17, −12 en el 20, −16 al
final. No es un desfase constante que se pueda corregir con una resta: se abre y se cierra según
qué agrupa cada edición.

Así que en la cola **el `Sutta #` no sirve como clave del lado CST** y hay que resolverlo por
contenido. Lo que sí sigue valiendo es el lado PTS: el `DPR Ref` da el marcador y la página del
Excel lo confirma.

### El método

Para cada fila: se toma el texto de **su marcador impreso** (el último de su vagga con número ≤ n)
y se busca el paranum del Eka **cuya cobertura de n-gramas sea máxima**, sobre los 611. Se acepta
si:

- la cobertura del ganador es **≥ 0,85**;
- la **página** del marcador es la del Excel (±1);
- y, en las filas individuales, hay **separación** con el segundo (≥ 0,08). En las de rango no se
  exige, porque un rango cubre una tirada de suttas de fórmula idéntica y el empate es lo esperado
  —ahí lo que se escribe es el **tramo**, no un paranum suelto—.

63 de las 68 filas de la cola pasan. Las 5 que no quedan anotadas: `1.394` empata con `1.395-401`
(las dos apuntan al 382, y una de las dos sobra), `1.414-418`/`1.419-423` caen en el mismo tramo
del vagga 20, `1.424-430` no separa, y `1.600-615` discrepa de página (marcador en A i 43, Excel
dice 45).

⚠️ **Esto NO es adivinar por contenido.** El lado PTS está fijado por dos claves independientes
—el `DPR Ref` y la página— antes de mirar una sola palabra; el contenido sólo decide **cuál de los
611 paranums** es la contraparte, y con separación medida. Es la misma disciplina del piloto de
A ii: bloquear primero por estructura, decidir después por contenido.

Uso: python3 an1_eka_cola.py [-v]
"""
import re
import sqlite3
import sys

import an1_eka_duka as R
from openpyxl import load_workbook

COV_MIN = 0.85
GAP_MIN = 0.08


def resuelve(conn):
    """`{Sutta #: (vri, detalle, ok)}` para las filas pendientes de la cola del Eka."""
    mk = R.marcadores(conn, 'EKA')
    cst, _divs = R.divisiones('s0401m')
    todos = [(pn, t) for v in sorted(cst) for pn, t in cst[v]]
    wb = load_workbook(R.XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ci['Nikaya']] != 'AN' or str(row[ci['PTS Roman']] or '').strip() != 'i':
            continue
        if row[ci['Estado']] == 'CONFIRMADO':
            continue
        num = str(row[ci['Sutta #']])
        if not num.startswith('1.'):
            continue
        m = re.match(r'AN\s*\d+\.(\d+)\.(\d+)', str(row[ci['DPR Ref']] or ''))
        if not m:
            continue
        v, n = int(m.group(1)), int(m.group(2))
        cand = [x for x in mk.get(v, []) if x[0] <= n]
        if not cand:
            continue
        mar = cand[-1]
        punt = sorted(((R.cobertura(mar[3], t), pn) for pn, t in todos), reverse=True)
        cov, pn = punt[0]
        gap = cov - punt[1][0]
        pag = isinstance(row[ci['PTS Page']], int) and abs(mar[1] - row[ci['PTS Page']]) <= 1
        rango = '-' in num.split('.')[1]
        lo_n, _, hi_n = num.split('.')[1].partition('-')
        n_fila = int(hi_n) - int(lo_n) + 1 if hi_n else 1
        # en las filas de rango la `VRI Ref` es el TRAMO: la tirada contigua de paranums que
        # comparte la cobertura del ganador (fórmula idéntica repetida), no un paranum suelto
        if rango:
            altos = {p for c, p in punt if c >= cov - 0.05}
            lo = hi = pn
            while lo - 1 in altos:
                lo -= 1
            while hi + 1 in altos:
                hi += 1
            vri = f's0401m:{lo}-{hi}' if hi > lo else f's0401m:{lo}'
            n_tramo = hi - lo + 1
        else:
            vri = f's0401m:{pn}'
            n_tramo = 1
        # ⚠️ CUARTA comprobación, y es la que hace rigurosas las filas de rango: **el tramo hallado
        # por contenido tiene que tener tantos paranums como suttas declara la fila**. `1.431-438`
        # son ocho y el contenido da `419-426`, ocho. Es una coincidencia numérica que el contenido
        # no puede fabricar: si el tramo saliera corrido o mal cortado, la cardinalidad lo delata.
        card_ok = (n_tramo == n_fila)
        # …o bien las DOS coordenadas se confirman entre sí: el paranum que elige el contenido cae
        # DENTRO del rango que declara el `Sutta #`. Eso pasa en los vaggas 13-16, donde el
        # `Sutta #` todavía es el paranum (el Etadagga va 188-267 en las dos ediciones) y el
        # contenido no puede delimitar la tirada porque ahí los suttas no son de fórmula: cada uno
        # nombra a un discípulo distinto. Cuando se confirman, la `VRI Ref` es el rango del Excel.
        lo_x, hi_x = int(lo_n), int(hi_n) if hi_n else int(lo_n)
        confirman = lo_x <= pn <= hi_x
        if confirman:
            vri = f's0401m:{lo_x}' + (f'-{hi_x}' if hi_x > lo_x else '')
        ok = cov >= COV_MIN and pag and (confirman or card_ok) and (rango or gap >= GAP_MIN)
        out[num] = (vri, f'A i cola: DPR ({v}.{n}) → marcador {mar[0]} en A i {mar[1]},{mar[2]}; '
                         f'paranum por contenido (cobertura {cov:.2f}'
                         + (f', separación {gap:.2f}' if not rango
                            else f', el paranum cae dentro del rango del Sutta #'
                            if confirman
                            else f', tramo de {n_tramo} = las {n_fila} de la fila'
                            if card_ok else f', tramo de {n_tramo} ≠ {n_fila} de la fila') + ')', ok)
    return out


def main():
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    res = resuelve(conn)
    ok = sum(1 for v in res.values() if v[2])
    print(f'cola del Eka: {len(res)} filas · resueltas {ok} · sin resolver {len(res) - ok}')
    for num, (vri, det, bien) in res.items():
        if bien and '-v' not in sys.argv:
            continue
        print(f"  {'✓' if bien else '✗'} {num:>12} → {vri:<20} {det}")


if __name__ == '__main__':
    main()
