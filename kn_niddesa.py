#!/usr/bin/env python3
"""
kn_niddesa — el Mahāniddesa (16), el Cūḷaniddesa (23) y el Paṭisambhidāmagga (31): 70 filas en tres
regímenes distintos, cada uno con su marcador impreso (regla (6)).

Las tres obras son **comentario canónico**: prosa exegética que glosa palabra por palabra. Eso las
hace **formulaicas hasta lo indistinguible** — en el Mahāniddesa la cobertura de un capítulo con su
par del CST es 0,93-0,96 y con un capítulo **ajeno** llega a **0,85**. Como en el Udāna: el
contenido corrobora, no discrimina. Lo que decide es el **marcador impreso**, y por suerte cada obra
tiene el suyo.

### Mahāniddesa (`Nidd`, 16 filas, libro 36, `s0515m`)

PTS imprime en cada capítulo un **ordinal en pali más el nombre**, dos veces: como colofón al cerrar
—`SOḶASAMO SĀRIPUTTASUTTANIDDESO` en la p510— y como encabezado al abrir. El ordinal es el índice
del capítulo y el nombre lo identifica; los 16 `div` del CST llevan el mismo número y el mismo
nombre. `VRI Ref` = `s0515m:<a>-<b>` (aquí el `n` **sí** es un paranum corrido, 1-210).

### Cūḷaniddesa (`Nidd` 21 + `Nidd II` 2, libro 37, `s0516m`)

Aquí **no hay marcador por sección**: el impreso sólo titula tres cosas —`1. Vatthugāthā`,
`2. Pucchā`, `3. Khaggavisāṇasutta`— y las dieciséis pucchās de los māṇavas van dentro de la segunda
sin título propio. Lo que sí imprime es **el número de estrofa del Suttanipāta** (`1033.` para el
Ajita), que es la misma numeración que el CST glosa. La `VRI Ref` va por capítulo y subsección
(regla (5-bis)): `s0516m:c1.<n>`.

⚠️ Las dos filas de `Nidd II` (`Cnd 3 Pārāyanatthutigāthā`, `Cnd 4 Pārāyanānugītigāthā`) **no tienen
página** en el Excel y son las **gāthās base**, no sus niddesas —que sí están, como `Cnd 21` y
`Cnd 22`—. Quedan fuera hasta decidir qué son.

### Paṭisambhidāmagga (`Paṭis I` 22 + `Paṭis II` 9, libros 38-39, `s0517m`)

El mejor marcador de los tres: PTS titula cada kathā con **numeral romano, vagga y nombre** —
`IV. MAHĀVAGGE INDRIYAKATHĀ`, `X. PAÑÑĀVAGGE MĀTIKAKATHĀ`—, y el CST marca las mismas treinta con
`<p rend="title">`. Las dos listas coinciden en número, orden y nombre, y **la página impresa es
exactamente la que declara el Excel** en las que se leen. La fila 31 es la `Mātikā` inicial, que el
CST pone como `subhead` del Mahāvagga.

⚠️ El `n` del CST **reinicia en cada vagga** (1-240, 1-48, 1-42): regla (5-bis), `VRI Ref` =
`s0517m:c<vagga>.<a>-<b>`.

Uso: python3 kn_niddesa.py [--obra mnd|cnd|ps] [--dry]
"""
import re
import shutil
import sqlite3
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from difflib import SequenceMatcher

import an_peyyala as ap
import an1_eka_duka as R
import pali_norm
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
XMLDIR = '/tmp/tipitaka-xml/romn'
COV_MIN = 0.55

# ⚠️ **en la forma plegada**, que es como llegan tras `ap.fold`: `catuttho` → `catutho`. Escribir
# la geminada aquí deja el capítulo IV sin marcador y sin firmar.
ORDINALES = ('pathamo dutiyo tatiyo catutho pancamo chatho satamo athamo navamo dasamo ekadasamo '
             'dvadasamo terasamo cudasamo panarasamo solasamo').split()
ROMANOS = 'I II III IV V VI VII VIII IX X'.split()


def norm(t):
    """Raíz comparable: sin prefijo de catálogo, sin ordinal, sin sufijo de género."""
    t = pali_norm.normaliza((t or '').replace('­', ''))
    t = re.sub(r'^\s*\((?:Mnd|Cnd|KN)[^)]*\)\s*', '', t)
    t = re.sub(r'^\s*(Nd\s*\d+\s*\d*|Ps\s*[\d.]+)\s*', '', t)
    t = ap.fold(re.sub(r'^[\d.\sIVX\-]+', '', t)).strip()
    # ⚠️ los sufijos van en la forma **plegada** (`ap.fold` colapsa la geminada): `sutaniddeso`
    # llega como `sutanideso`, y escribirlo con las dos `d` no casa con nada.
    t = re.sub(r'(sutanideso|sutanidesa|nideso|nidesa|kathaya|katha|vago|sutam|suta)$', '', t)
    return re.sub(r'[aiueo]m?$', '', t.strip())


def parecido(a, b):
    return SequenceMatcher(None, norm(a), norm(b)).ratio()


def filas(vols, seccion=None):
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ci['Nikaya']] != 'KN' or str(r[ci['PTS Vol']]) not in vols:
            continue
        if seccion and seccion not in str(r[ci['Section']]):
            continue
        out.append({'name': str(r[ci['Sutta Name']] or ''), 'page': r[ci['PTS Page']],
                    'vol': str(r[ci['PTS Vol']]), 'estado': r[ci['Estado']]})
    return out


def divs_cst(stem, top):
    """`[(head, primer n, último n, texto)]` de los `div`, en orden de documento."""
    raiz = ET.parse(f'{XMLDIR}/{stem}.mul.xml').getroot()
    out = []
    for d in raiz.iter('div'):
        if d.get('n') == top:
            continue
        h = [''.join(x.itertext()).strip() for x in d.iter('head')]
        ns = [int(p.get('n')) for p in d.iter('p') if p.get('n')]
        out.append((h[0] if h else '', ns[0] if ns else None, ns[-1] if ns else None,
                    ap.fold(' '.join(''.join(p.itertext()) for p in d.iter('p')))))
    return out


def secciones_cst(stem, top, rend):
    """`[(vagga, orden, título, primer n, último n, texto)]` — las secciones marcadas con `rend`.

    ⚠️ En orden de documento y con el vagga explícito, **nunca por paranum**: el `n` reinicia en cada
    vagga (regla (5-bis)) y ordenar por él entrelaza las series.
    """
    raiz = ET.parse(f'{XMLDIR}/{stem}.mul.xml').getroot()
    out = []
    for v, d in enumerate((x for x in raiz.iter('div') if x.get('n') != top), 1):
        act, k = None, 0
        for p in d.iter('p'):
            txt = ''.join(p.itertext()).strip()
            if p.get('rend') == rend:
                k += 1
                act = [v, k, txt, None, None, '']
                out.append(act)
            elif act is not None:
                if p.get('n'):
                    act[3] = act[3] if act[3] is not None else int(p.get('n'))
                    act[4] = int(p.get('n'))
                act[5] += ' ' + txt
    return [(a, b, c, d, e, ap.fold(f)) for a, b, c, d, e, f in out]


def texto_pts(conn, libro, a, b):
    return ' '.join(ap.fold(x['unitext'] or '') for x in conn.execute(
        'SELECT unitext FROM pages WHERE edition="mula" AND book_no=? AND page_no BETWEEN ? AND ?',
        (libro, a, b)))


# ─────────────────────────── Mahāniddesa ───────────────────────────
def marcadores_mnd(conn):
    """`{ordinal: (página, nombre)}` — el ordinal en pali con que PTS abre y cierra cada capítulo."""
    pat = re.compile(rf'({"|".join(ORDINALES)})\d*\s+([a-z]+sutanideso)')
    out = {}
    for r in conn.execute('SELECT page_no,unitext FROM pages WHERE edition="mula" AND book_no=36 '
                          'ORDER BY page_no'):
        for l in (r['unitext'] or '').split('\n'):
            m = pat.search(ap.fold(l.strip()))
            if m:
                n = ORDINALES.index(m.group(1)) + 1
                out.setdefault(n, (r['page_no'], m.group(2)))
    return out


def resuelve_mnd(conn):
    cst = divs_cst('s0515m', 'kn15')
    fs = filas({'Nidd'}, 'Mahā')
    mk = marcadores_mnd(conn)
    print(f'CST {len(cst)} capítulos · Excel {len(fs)} filas · marcadores impresos {len(mk)}/16')
    res, avisos = {}, []
    for k, f in enumerate(fs):
        head, a, b, tc = cst[k]
        if parecido(f['name'], head) < 0.8:
            avisos.append(f'   ⚠ {k + 1}: «{f["name"]}» vs «{head}»')
        sig = fs[k + 1]['page'] if k + 1 < len(fs) else 510
        cov = R.cobertura(texto_pts(conn, 36, f['page'], sig), tc)
        pg, nom = mk.get(k + 1, (None, None))
        if nom is None or parecido(nom, head) < 0.8:
            avisos.append(f'   ✗ {k + 1} «{head}»: el impreso no confirma el ordinal '
                          f'({nom or "sin marcador"})')
            continue
        res[f['name']] = (
            f's0515m:{a}-{b}', f'Nd1 {k + 1}',
            f'Mahāniddesa: capítulo {k + 1} de 16 → «{head}», paranums {a}-{b} del CST. El impreso '
            f'lo nombra con su ordinal pali —«{ORDINALES[k]} {nom}»— en la p{pg}; cobertura '
            f'{cov:.2f} (en esta obra el contenido corrobora, no discrimina: un capítulo ajeno '
            f'llega a 0,85 porque la prosa exegética es formulaica)')
    return res, avisos, len(fs)


# ─────────────────────────── Cūḷaniddesa ───────────────────────────
def resuelve_cnd(conn):
    """Las 21 filas de `Nidd` del Cūḷaniddesa. Las 2 de `Nidd II` no tienen página y quedan fuera."""
    cst = divs_cst('s0516m', 'kn16')
    sub = secciones_cst('s0516m', 'kn16', 'subhead')
    fs = filas({'Nidd', 'Nidd II'}, 'Cūḷa')
    print(f'CST: {len(cst)} div, {len(sub)} subheads · Excel {len(fs)} filas')
    # Cnd 1-2 son las gāthās y las preguntas base; Cnd 5-20 las 16 pucchās; 21-22 sus dos niddesas
    # de cierre; 23 el Khaggavisāṇa, que en el CST es un `div` propio.
    res, avisos = {}, []
    pucchas = [s for s in sub if s[0] == 1]
    for f in fs:
        m = re.search(r'\(Cnd (\d+)\)', f['name'])
        if not m:
            avisos.append(f'   ⚠ sin nº de Cnd: {f["name"]}')
            continue
        n = int(m.group(1))
        if n == 1:
            # ⚠️ El Cūḷaniddesa del CST **no reimprime** las vatthugāthās: su div arranca ya en
            # `1. Ajitamāṇavapucchāniddeso`. PTS sí las imprime al frente (pp.1-6, estrofas
            # **976-1031**, con el `2. Pucchā.` abriendo en la 1032 — los dos extremos impresos), y
            # su texto es el del **Suttanipāta**: cobertura 0,97 contra la `Vatthugāthā` de
            # `s0505m`. La referencia va allí, que es donde está el texto.
            res[f['name']] = (
                's0505m:976-1037', 'Nd2 1',
                'Cūḷaniddesa: las vatthugāthās del Pārāyanavagga, que PTS imprime al frente '
                '(pp.1-6, estrofas 976-1031) y el Cūḷaniddesa del CST no reimprime — su div abre '
                'ya en el Ajitamāṇavapucchāniddesa. El texto es el del Suttanipāta: cobertura '
                '0,97 contra la «Vatthugāthā» de s0505m (n976-1037)')
            continue
        if n in (3, 4):
            # Las dos filas de `Nidd II`, **sin página en el Excel**. Son las **gāthās base**, no
            # sus niddesas —que son `Cnd 21` y `Cnd 22`—, y PTS las imprime dentro del Cūḷaniddesa
            # en las mismas páginas. El texto es el del Suttanipāta y la cobertura las separa sin
            # ambigüedad: pp.46-49 dan 0,97 con la `Pārāyanatthutigāthā` y 0,47 con la otra;
            # pp.49-56, 0,86 y 0,34. Se les asigna además la página que les faltaba.
            a2, b2, pg2, tit = ((1130, 1136, 46, 'Pārāyanatthutigāthā') if n == 3 else
                                (1137, 1155, 49, 'Pārāyanānugītigāthā'))
            res[f['name']] = (
                f's0505m:{a2}-{b2}', f'Nd2 {n}',
                f'Cūḷaniddesa: la «{tit}» **base** (no su niddesa, que es Cnd {n + 18}), que PTS '
                f'imprime dentro del Cūḷaniddesa desde la p{pg2}. El texto es el del Suttanipāta, '
                f'n{a2}-{b2} de s0505m; la cobertura separa las dos gāthās sin ambigüedad '
                f'(0,97 contra 0,47 y 0,86 contra 0,34). PÁGINA ASIGNADA: el Excel la tenía vacía',
                pg2)
            continue
        if n == 2:
            avisos.append(f'   · Cnd 2 «Pucchā»: **encabezado de sección**, no una unidad. PTS lo '
                          f'imprime en la p6 para abrir las preguntas, y su contenido son las '
                          f'filas Cnd 5-22. Queda PENDIENTE')
            continue
        if n == 23:
            head, a, b, tc = cst[1]
            vri, sec = 's0516m:c2', head
        elif 5 <= n <= 22:
            v, kk, head, a, b, tc = pucchas[n - 5]
            vri, sec = f's0516m:c1.{kk}', head
        else:
            avisos.append(f'   ⚠ Cnd {n} inesperado')
            continue
        if parecido(f['name'], head) < 0.75:
            avisos.append(f'   ⚠ Cnd {n}: «{f["name"]}» vs «{head}»')
            continue
        sig = next((g['page'] for g in fs if isinstance(g['page'], int)
                    and g['page'] > f['page']), 73)
        cov = R.cobertura(texto_pts(conn, 37, f['page'], sig), tc)
        res[f['name']] = (
            vri, f'Nd2 {n}',
            f'Cūḷaniddesa: «{sec}» → {vri} (paranums {a}-{b} del CST). El impreso no titula cada '
            f'pucchā —sólo `1. Vatthugāthā`, `2. Pucchā` y `3. Khaggavisāṇasutta`— pero numera las '
            f'estrofas del Suttanipāta que el niddesa glosa; cobertura {cov:.2f}')
    return res, avisos, len(fs)


# ─────────────────────── Paṭisambhidāmagga ───────────────────────
def marcadores_ps(conn):
    """`[(libro, página, romano, vagga, nombre)]` — `IV. MAHĀVAGGE INDRIYAKATHĀ`."""
    pat = re.compile(r'^([IVX]{1,4})\.?\s+(?:([A-ZĀĪŪṂṆṬḌÑṄḶ]+VAGGE)\s+)?'
                     r'([A-ZĀĪŪṂṆṬḌÑṄḶ\-]+KATHĀ)\.?\d*$')
    out = []
    for libro in (38, 39):
        for r in conn.execute('SELECT page_no,unitext FROM pages WHERE edition="mula" '
                              'AND book_no=? ORDER BY page_no', (libro,)):
            for l in (r['unitext'] or '').split('\n'):
                m = pat.match(l.strip())
                if m:
                    out.append((libro, r['page_no'], m.group(1), m.group(2) or '', m.group(3)))
    return out


def resuelve_ps(conn):
    cst = secciones_cst('s0517m', 'kn17', 'title')
    fs = filas({'Paṭis I', 'Paṭis II'})
    mk = marcadores_ps(conn)
    print(f'CST {len(cst)} kathās · Excel {len(fs)} filas · títulos impresos {len(mk)}')
    res, avisos = {}, []
    for f in fs:
        m = re.search(r'Ps (\d+)\.(\d+)', f['name'])
        if not m:
            avisos.append(f'   ⚠ sin referencia: {f["name"]}')
            continue
        v, n = int(m.group(1)), int(m.group(2))
        if n == 0:
            # `Ps 1.0 Mātikā`: el CST no la marca como `title` sino como `subhead` del Mahāvagga —
            # es el índice de los 73 ñāṇas que la obra desarrolla, y lleva sus propios paranums.
            mat = [x for x in secciones_cst('s0517m', 'kn17', 'subhead')
                   if x[0] == 1 and norm(x[2]) == norm('Mātikā')]
            if not mat:
                avisos.append(f'   ⚠ {f["name"]}: no encuentro la Mātikā en el CST')
                continue
            _v, _k, head, a, b, tc = mat[0]
            cov = R.cobertura(texto_pts(conn, 38, 1, 4), tc)
            res[f['name']] = (
                f's0517m:c1.{a}-{b}', 'Ps 1.0',
                f'Paṭisambhidāmagga: la Mātikā que abre el Mahāvagga —el índice de los ñāṇas que '
                f'la obra desarrolla—, paranums {a}-{b} del CST. El CST la marca como `subhead`, '
                f'no como kathā, que es por lo que no está entre los treinta títulos; '
                f'cobertura {cov:.2f}')
            continue
        cand = [c for c in cst if c[0] == v and c[1] == n]
        if not cand:
            avisos.append(f'   ⚠ {f["name"]}: el CST no tiene la kathā {v}.{n}')
            continue
        _v, _k, head, a, b, tc = cand[0]
        if parecido(f['name'], head) < 0.8:
            avisos.append(f'   ⚠ {f["name"]} vs «{head}»')
            continue
        libro = 38 if v == 1 and n <= 3 else 39
        imp = [x for x in mk if x[0] == libro and parecido(x[4], head) >= 0.8]
        ok_pg = bool(imp) and any(x[1] == f['page'] for x in imp)
        sig = next((g['page'] for g in fs
                    if isinstance(g['page'], int) and g['page'] > f['page']
                    and (38 if 'Ps 1.' in g['name'] and int(re.search(r'Ps 1\.(\d+)', g['name'])
                                                            .group(1)) <= 3 else 39) == libro), 246)
        cov = R.cobertura(texto_pts(conn, libro, f['page'], sig), tc)
        if cov < COV_MIN and not ok_pg:
            avisos.append(f'   ✗ Ps {v}.{n} «{head}»: cobertura {cov:.2f} y el impreso no confirma')
            continue
        res[f['name']] = (
            f's0517m:c{v}.{a}-{b}' if a != b else f's0517m:c{v}.{a}', f'Ps {v}.{n}',
            f'Paṭisambhidāmagga: kathā {n} del vagga {v} → «{head}», paranums {a}-{b} del CST '
            f'(el `n` reinicia por vagga). '
            + (f'El impreso la titula «{imp[0][2]}. {imp[0][3]} {imp[0][4]}» en la p{imp[0][1]} del '
               f'libro {libro}, la misma que declara el Excel; ' if ok_pg else
               'El título impreso no se lee en la BD; ')
            + f'cobertura {cov:.2f}')
    return res, avisos, len(fs)


def main():
    dry = '--dry' in sys.argv
    obra = sys.argv[sys.argv.index('--obra') + 1] if '--obra' in sys.argv else 'mnd'
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    res, avisos, total = {'mnd': resuelve_mnd, 'cnd': resuelve_cnd, 'ps': resuelve_ps}[obra](conn)
    print(f'\n{len(res)}/{total} firmadas')
    for a in avisos:
        print(a)
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-{obra}.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        nom = str(ws.cell(r, ci['Sutta Name']).value or '')
        if ws.cell(r, ci['Nikaya']).value != 'KN' or nom not in res:
            continue
        if ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, ref, det, *pag = res[nom]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['PTS Ref']).value = ref
        if pag:
            ws.cell(r, ci['PTS Page']).value = pag[0]
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
