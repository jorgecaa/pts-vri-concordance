#!/usr/bin/env python3
"""
kn_itivuttaka — el Itivuttaka: 112 filas, una por sutta.

La obra más cómoda de KN hasta ahora, y por una razón concreta: **PTS numera sus suttas de 1 a 112
de corrido y el CST usa exactamente la misma numeración**. No hay que alinear nada; hay que
comprobarlo, y comprobarlo por tres caminos que no se apoyan entre sí.

### Las tres fuentes

| | qué da |
|---|---|
| **PTS** (BD, libro 25) | `112. (Cat.13) Vuttaṃ hetaṃ bhagavatā…` — el **nº corrido** y, entre paréntesis, la referencia relativa al nipāta |
| **CST** (`s0504m`) | `n` = **1..112**, corrido, en cuatro `div` (Ekaka 1-27, Duka 28-49, Tika 50-99, Catukka 100-112) |
| **Excel** | 112 filas; el nº corrido está en el **nombre** (`Iti 112 Lokavabodha`) y el `Sutta #` es la referencia de tres niveles (`4.4.13` = Catukka, vagga 4, sutta 13) |

Y las dos referencias del Excel se confirman entre sí: su `Sutta #` `4.4.13` es el `(Cat.13)` que
el impreso escribe al lado del `112.`.

### La notación (regla (5-ter))

La referencia es **`It <nº de sutta>`**, no la página. Es la citación estándar de la obra y la única
que sirve fuera de PTS.

### Las comprobaciones

1. el nº del nombre del Excel existe como **paranum del CST**;
2. ese mismo nº está **impreso como marcador** en la página que declara el Excel (±1);
3. **cobertura de contenido** entre el sutta del impreso y el del CST.

Uso: python3 kn_itivuttaka.py [--dry]
"""
import re
import shutil
import sqlite3
import sys
import xml.etree.ElementTree as ET
from datetime import datetime

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
BOOK, STEM = 25, 's0504m'
COV_MIN = 0.55

# ⚠️ Dos páginas del Excel que el impreso desmiente. En los dos casos la cobertura del par es
# 0,97-0,98 —o sea el sutta es el que dice la fila— y lo que falla es la página:
#   · `Iti 6 Māna` se imprime en **It 3** (`6. (Ek. I.6)`), no en la 5, donde los marcadores son
#     el 9 y el 10;
#   · `Iti 38 Vitakka` se imprime en **It 31** (`38. (Duk. II.1)`), no en la 21, que **no tiene
#     ningún marcador** — parece una transposición de dígitos, 31 → 21.
PAGINA_CORREGIDA = {6: 3, 38: 31}


def suttas_pts(conn):
    """`{nº: (página, línea, texto)}` — el impreso numera cada sutta al margen (`112. (Cat.13)…`)."""
    pgs = {r['page_no']: (r['unitext'] or '').split('\n')
           for r in conn.execute('SELECT page_no,unitext FROM pages WHERE edition="mula" '
                                 'AND book_no=?', (BOOK,))}
    plano = [(pg, ln, t) for pg in sorted(pgs) for ln, t in enumerate(pgs[pg], 1)]
    marcas = []
    for k, (pg, ln, t) in enumerate(plano):
        # El marcador es el nº del sutta seguido del paréntesis con su referencia de nipāta. Entre
        # los dos cabe casi cualquier cosa, y el impreso lo demuestra: `90.* (Tik. V.1)`,
        # `98 (Tik. V.9)` **sin punto**, `110. ** (Cat.11)` con dos asteriscos. Exigir la forma
        # canónica `N. (Nip.…)` perdía cinco de los 112 — y los cinco existen, sólo que con la
        # llamada de nota metida en medio.
        m = re.match(r'\s{0,8}(\d+)\.?[\s*\d]*\((?:Ek|Duk|Tik|Cat|Pañ)',
                     t.replace('\r', ''))
        if m:
            marcas.append((int(m.group(1)), pg, ln, k))
    out = {}
    for j, (n, pg, ln, k) in enumerate(marcas):
        fin = marcas[j + 1][3] if j + 1 < len(marcas) else len(plano)
        out[n] = (pg, ln, ap.fold(' '.join(t for _p, _l, t in plano[k:fin])))
    return out


def suttas_cst():
    """`{paranum: (div, texto)}` del fichero VRI.

    ⚠️ Se usa `an_peyyala.cst_unidades`, **no un lector propio**. El CST sólo pone el atributo `n`
    en el PRIMER `<p>` de cada sutta y los de continuación —la prosa que sigue y las gāthās— van sin
    él. Leer sólo los numerados deja en cada sutta del Itivuttaka su primer párrafo, que es la
    fórmula `Vuttaṃ hetaṃ bhagavatā vuttamarahatā…` **idéntica en los 112**: el cotejo pasa a
    comparar la misma frase consigo misma y todos los suttas empatan. Escribí aquí un lector nuevo y
    reintroduje exactamente ese defecto, ya corregido en `an_peyyala`.
    """
    return {u['pn']: (u['div'], ap.fold(u['texto'])) for u in ap.cst_unidades(STEM)}


def filas():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ci['Nikaya']] != 'KN' or str(r[ci['PTS Vol']]) != 'It':
            continue
        nom = str(r[ci['Sutta Name']] or '')
        m = re.search(r'Iti\s+(\d+)', nom)
        out.append({'num': str(r[ci['Sutta #']]), 'name': nom, 'page': r[ci['PTS Page']],
                    'estado': r[ci['Estado']], 'n': int(m.group(1)) if m else None})
    return out


def main():
    dry = '--dry' in sys.argv
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    pts, cst, fs = suttas_pts(conn), suttas_cst(), filas()
    print(f'PTS {len(pts)} suttas numerados · CST {len(cst)} paranums · Excel {len(fs)} filas')
    res, malas = {}, []
    for f in fs:
        if f['n'] is None or f['n'] not in cst:
            malas.append((f['num'], f['name'][:34], 'sin nº en el nombre o sin paranum'))
            continue
        div, tc = cst[f['n']]
        if f['n'] not in pts:
            malas.append((f['num'], f['name'][:34], f'el impreso no numera el {f["n"]}'))
            continue
        pg, ln, tp = pts[f['n']]
        cov = R.cobertura(tp, tc)
        esperada = PAGINA_CORREGIDA.get(f['n'], f['page'])
        pag = isinstance(esperada, int) and abs(pg - esperada) <= 1
        if cov >= COV_MIN and pag:
            det = (f'Itivuttaka: el sutta {f["n"]} lo numeran las dos ediciones igual; '
                   f'impreso en It {pg},{ln} ({div}); cobertura {cov:.2f}')
            if f['n'] in PAGINA_CORREGIDA:
                det += (f'. PÁGINA CORREGIDA: el Excel decía {f["page"]} y el marcador está '
                        f'en la {pg}')
            res[f['num']] = (f'{STEM}:{f["n"]}', f'It {f["n"]}', det, PAGINA_CORREGIDA.get(f['n']))
        else:
            malas.append((f['num'], f['name'][:34],
                          f'cov {cov:.2f}' + ('' if pag else f' · pág {pg} vs {f["page"]}')))
    print(f'\n{len(res)}/{len(fs)} firmadas')
    for x in malas:
        print('   ✗', x)
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-iti.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'KN' or str(ws.cell(r, ci['PTS Vol']).value) != 'It':
            continue
        num = str(ws.cell(r, ci['Sutta #']).value)
        if num not in res or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, ref, det, pag_new = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['PTS Ref']).value = ref
        if pag_new:
            ws.cell(r, ci['PTS Page']).value = pag_new
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
