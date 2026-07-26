#!/usr/bin/env python3
"""
normaliza_siglas_cpd — pone la columna `PTS Vol` en las **siglas del *Critical Pāli Dictionary***.

La norma lexicográfica del pali la fija el CPD y sus *Epilegomena* (Copenhague, 1948), que el
*Dictionary of Pāli* de Cone continúa. De ahí salen tres cosas, y la tercera es la que arregla esta
columna:

1. **siglas de edición**: `Ee` (*editio Europaea*, o sea PTS), `Be` (birmana, que es el CST),
   `Ce` (cingalesa), `Se` (siamesa) — la nomenclatura en que se escribe `deest in Ee`;
2. **la unidad de cita depende del género**: prosa por volumen, página y línea (`M I 1,5`); verso
   por **número de estrofa**, porque la estrofa es estable entre ediciones y la página no. Es la
   regla (5-ter) de este proyecto, adoptada antes de saber que coincidía;
3. **una sigla por obra**, con el volumen dentro de la referencia y no en la sigla.

El punto 3 es el que faltaba, y explica un defecto que este repo llevaba documentado como rareza del
fichero: **«tres etiquetas mezclan obras»** no era una rareza, era la consecuencia de no usar siglas
de obra. `Pet` rotulaba el Milindapañha (248 filas) **y** el Peṭakopadesa (9); `Bv`, el Buddhavaṃsa
y el Cariyāpiṭaka; `Ap`, los dos Apadānas; `Nidd`, el Mahā- y el Cūḷaniddesa. Y `Paṭis I`/`Paṭis II`
no eran obras sino **volúmenes**.

DN, MN, SN y AN ya venían con `D`, `M`, `S`, `A`, que son las del CPD: no se tocan.

Uso: python3 normaliza_siglas_cpd.py [--dry]
"""
import re
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# `Section` del Excel → sigla del CPD. La `Section` es la clave porque es la que **no** mezcla
# obras; la vieja `PTS Vol` no puede serlo, que es justo el defecto que se corrige.
SIGLA = {
    'Khuddakapāṭha': 'Khp', 'Dhammapada': 'Dhp', 'Udāna': 'Ud', 'Itivuttaka': 'It',
    'Suttanipāta': 'Sn', 'Vimānavatthu': 'Vv', 'Petavatthu': 'Pv', 'Theragāthā': 'Th',
    'Therīgāthā': 'Thī', 'Jātaka': 'Ja', 'Therāpadāna': 'Th-ap', 'Therīapadāna': 'Thī-ap',
    'Buddhavaṃsa': 'Bv', 'Cariyāpiṭaka': 'Cp', 'Mahā-Niddesa': 'Nidd I',
    'Cūḷa-Niddesa': 'Nidd II', 'Paṭisambhidāmagga': 'Paṭis', 'Nettipakaraṇa': 'Nett',
    'Peṭakopadesa': 'Peṭ', 'Milindapañha': 'Mil', 'Niddesa': 'Nidd',
}

# prefijos de `PTS Ref` que yo escribí y que no eran los del CPD
REF = [(r'^Nd1\b', 'Nidd I'), (r'^Nd2\b', 'Nidd II'),
       (r'^Th Ap\b', 'Th-ap'), (r'^Thī Ap\b', 'Thī-ap'), (r'^Thi Ap\b', 'Thī-ap')]


def sigla_de(seccion):
    for nombre, sg in SIGLA.items():
        if seccion.startswith(nombre):
            return sg
    return None


def main():
    dry = '--dry' in sys.argv
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    cambios, refs, sin = {}, 0, []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ci['Nikaya']).value != 'KN':
            continue
        sec = str(ws.cell(r, ci['Section']).value or '')
        sg = sigla_de(sec)
        if sg is None:
            sin.append(sec)
            continue
        viejo = str(ws.cell(r, ci['PTS Vol']).value)
        if viejo != sg:
            cambios[(viejo, sg)] = cambios.get((viejo, sg), 0) + 1
        ref = str(ws.cell(r, ci['PTS Ref']).value or '')
        nueva = ref
        for pat, rep in REF:
            nueva = re.sub(pat, rep, nueva)
        if nueva != ref:
            refs += 1
        if not dry:
            ws.cell(r, ci['PTS Vol']).value = sg
            if nueva != ref:
                ws.cell(r, ci['PTS Ref']).value = nueva
    print(f'{sum(cambios.values())} filas cambian de sigla · {refs} `PTS Ref` reescritas')
    for (a, b), n in sorted(cambios.items(), key=lambda x: -x[1]):
        print(f'   {a:10} → {b:8} {n:5}')
    if sin:
        print(f'⚠ {len(sin)} filas con Section sin sigla: {sorted(set(sin))}')
    if dry:
        print('(dry-run: nada escrito)')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-cpd.xlsx'
    shutil.copy(XLSX, cp)
    wb.save(XLSX)
    print(f'backup → {cp}\nescrito → {XLSX}')


if __name__ == '__main__':
    main()
