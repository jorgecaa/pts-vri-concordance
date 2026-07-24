#!/usr/bin/env python3
"""DeepSeek validation for ALL 34 DN suttas."""
import hashlib, json, os, re, sqlite3
from collections import Counter
from pathlib import Path
import sutta_hash as sh

DB_PTS = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'
CACHE_FILE = Path('helmer_ptscst_cache.json')

_dpd = None
def get_dpd():
    global _dpd
    if _dpd is None: _dpd = sqlite3.connect('file:/dev/shm/dpd.db?mode=ro', uri=True)
    return _dpd

DN_BOOKS = {'i':6, 'ii':7, 'iii':8}

def cst_for_dn(snum):
    sn = int(snum)
    if sn <= 13: return ['d1m.xml'], 'h2n', sn-1
    elif sn <= 23: return ['d2m.xml'], 'h2n', sn-14
    else: return ['d3m.xml'], 'h2n', sn-24

def pts_text(book_no, page_no, line_no=1, n_tokens=400):
    conn = sqlite3.connect(DB_PTS); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone(); conn.close()
    if not r or not r['unitext']: return ''
    lines = r['unitext'].split('\n'); start = max(0, (line_no or 1)-1)
    tokens = sh.tokens(' '.join(lines[start:])); return ' '.join(tokens[:n_tokens])

def collate(pts, cst, maxw=200):
    from collatex import Collation, collate
    ta = sh.tokens(pts)[:maxw]; tb = sh.tokens(cst)[:maxw]
    if not ta or not tb: return []
    a, b = ' '.join(ta), ' '.join(tb)
    c = Collation(); c.add_plain_witness('PTS', a); c.add_plain_witness('CST', b)
    table = collate(c, output='table', segmentation=True)
    diffs = []
    for col in table.columns:
        d = col.tokens_per_witness
        sa = ' '.join(str(t) for t in d.get('PTS', [])).strip()
        sb = ' '.join(str(t) for t in d.get('CST', [])).strip()
        if sa != sb: diffs.append((sa, sb))
    return diffs

def dpd_known(w1, w2):
    for a, b in ((w1,w2),(w2,w1)):
        if not a: continue
        r = get_dpd().execute("SELECT variant FROM lookup WHERE lookup_key=?", (a,)).fetchone()
        if r and r[0] not in ('', '[]') and b and b in r[0]: return True
    return False

def classify(diffs):
    out = []
    for sa, sb in diffs[:25]:
        wa, wb = sa.split(), sb.split()
        if not sa or not sb: tag = 'ins/om'
        elif len(wa)==1 and len(wb)==1 and dpd_known(wa[0], wb[0]): tag = 'DPD'
        elif len(wa)==1 and len(wb)==1 and re.sub(r'[ṃṁ]','m',wa[0])==re.sub(r'[ṃṁ]','m',wb[0]): tag = 'ortho'
        else: tag = 'diverg'
        out.append((sa, sb, tag))
    return out

SYS = '''You are Helmer Smith, Pali philologist. You audit whether two Pali passages (PTS and CST editions) are the SAME sutta. Same=core narrative matches; editions may differ in orthography, elisions (peyyala), or one may have expanded passages. Different=completely distinct sutta/topic. Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words English"}'''

def ask(p, c, v, cli):
    body = ['PTS:', p[:1200], '', 'CST:', c[:1200], '', 'CollateX (PTS|CST|DPD):']
    for sa, sb, tag in v: body.append('  "{}" | "{}" | {}'.format(sa, sb, tag))
    r = cli.chat.completions.create(model='deepseek-v4-flash', temperature=0,
        response_format={'type': 'json_object'},
        messages=[{'role': 'system', 'content': SYS}, {'role': 'user', 'content': '\n'.join(body)}])
    return json.loads(r.choices[0].message.content)

from openpyxl import load_workbook
def get_dn_tests():
    wb = load_workbook(XL); ws = wb['Complete Canon']
    cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    tests = []
    for ri in range(2, ws.max_row+1):
        nik = str(ws.cell(row=ri, column=cols['Nikaya']).value or '')
        if nik != 'DN': continue
        snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
        if not snum or snum == 'None': continue
        roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower()
        page = ws.cell(row=ri, column=cols['PTS Page']).value
        ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
        name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
        m = re.search(r',(\d+)$', ref); line = int(m.group(1)) if m else 1
        bn = DN_BOOKS.get(roman)
        if not bn or not page: continue
        tests.append({'num': snum, 'ref': ref, 'name': name, 'bn': bn, 'page': page, 'line': line})
    return sorted(tests, key=lambda t: int(t['num']))

def main():
    from openai import OpenAI
    cli = OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'], base_url='https://api.deepseek.com')
    cache = json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}
    tests = get_dn_tests()
    
    print('Preloading CST for DN...')
    cst_cache = {}
    for xf in [['d1m.xml'], ['d2m.xml'], ['d3m.xml']]:
        cst_cache[tuple(xf)] = sh.load_cha(list(xf), 'h2n')
    
    print('Running ALL {} DN suttas through DeepSeek...'.format(len(tests)))
    print('=' * 95)
    results = []
    
    for t in tests:
        xf, lv, cst_idx = cst_for_dn(t['num'])
        pts = pts_text(t['bn'], t['page'], t['line'], 400)
        if not pts:
            print('✗ DN {:>3s} {:>18s} | ERROR | PTS page missing'.format(t['num'], t['ref']))
            results.append((t, {'verdict': 'ERROR', 'reason': 'PTS page missing', 'nvar': 0}, ''))
            continue
        
        segs = cst_cache[tuple(xf)]
        if cst_idx >= len(segs):
            print('✗ DN {:>3s} {:>18s} | ERROR | CST idx out of range'.format(t['num'], t['ref']))
            results.append((t, {'verdict': 'ERROR', 'reason': 'CST idx out of range', 'nvar': 0}, ''))
            continue
        
        cst_sutta = segs[cst_idx]; cst_text = ' '.join(sh.tokens(cst_sutta['text'])[:400])
        cst_title = cst_sutta.get('title', '')
        
        key = hashlib.md5((pts[:300] + cst_text[:300]).encode()).hexdigest()
        if key in cache:
            hv = cache[key]
        else:
            try:
                diffs = collate(pts, cst_text); variants = classify(diffs)
                hv = ask(pts, cst_text, variants, cli); hv['nvar'] = len(variants)
            except Exception as e:
                hv = {'verdict': 'ERROR', 'reason': str(e)[:60], 'nvar': 0}
            cache[key] = hv
            CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2))
        
        results.append((t, hv, cst_title))
        v = hv.get('verdict', '?')
        sym = '✓' if v == 'APPROVE' else '✗' if v == 'REJECT' else '?'
        print('{} DN {:>3s} {:>18s} | {:8s} | CST: {:40s} | {}'.format(
            sym, t['num'], t['ref'], v, cst_title[:40], hv.get('reason', '')[:55]))
    
    vc = Counter(hv.get('verdict', '?') for _, hv, _ in results)
    app = vc.get('APPROVE', 0); rej_count = vc.get('REJECT', 0); err = vc.get('ERROR', 0)
    print('\n' + '=' * 95)
    print('  DN: {} suttas | APPROVE: {} | REJECT: {} | ERROR: {}'.format(len(tests), app, rej_count, err))
    valid = len(tests) - err
    if valid > 0:
        print('  Precision: {}/{} ({}%)'.format(app, valid, int(100*app/valid)))
    
    if rej_count > 0:
        print('\n  REJECT -- revisar referencia PTS:')
        for t, hv, cst_title in results:
            if hv.get('verdict') == 'REJECT':
                print('    DN {:>3s} | {} | {}'.format(t['num'], t['ref'], hv.get('reason', '')[:80]))

if __name__ == '__main__':
    main()
