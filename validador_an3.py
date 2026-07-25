#!/usr/bin/env python3
"""Aligner + validador de **AN III** (A iii, `book_no=19`): Pañcaka y Chakka nipāta.

Tercer volumen de AN. Cambia el editor —muere Morris y sigue **Hardy**— y cambia el régimen de
marcador: A iii **no** usa el nº corrido centrado del Tika y el Catukka, sino el **numeral ROMANO
centrado**, y ahí el romano es el **SUTTA** (corrido, hasta `CCL`), no el vagga; los arábigos a
sangría 5 son sus **párrafos**. Es el cuarto régimen distinto dentro del mismo nikāya, y se elige
en `an1_markers.REGIMEN`.

**Verdad-terreno del lado PTS** (`anguttara-vol-III-info.txt`, índice de Hardy 1895/1897): 26 vaggas
en el Pañcaka y 12 en el Chakka. El contraste da **25/26** arranques exactos en el Pañcaka; el 26º
es el vagga **Upasampadā**, que Hardy titula entre corchetes por ser conjetura suya y que **no
numera sus suttas**: los separa una fila de ornamentos (`U+E0E0`), única en todo el Aṅguttara.

Recuentos: **Pañcaka 271 suttas** (250 numerados + 21 tras los ornamentos) —la cuenta exacta de
Hardy— y **Chakka 124**, ambos sin un solo hueco.

⚠️ **Los ficheros VRI no se reparten como los volúmenes de PTS**: `s0403m1` es el Pañcaka y
`s0403m2` el Chakka, pero `s0403m3` ya es el **Sattaka**, o sea AN IV. Elegir el XML por volumen
impreso es el mismo error que en MN, donde `s0201m` es el Mūlapaṇṇāsa (MN 1-50) y no el tomo I
(MN 1-76).

Uso: python3 validador_an3.py [--dry] [--all] [--n N] [--only 5.7,6.2]
"""
import json
import os
import re
import sqlite3
import sys
import xml.etree.ElementTree as ET
from collections import Counter

import an_names
import sutta_hash as sh
from align_rows import assign as ar_assign
from an1_markers import collect_an1
from openpyxl import load_workbook
from validador import validate_pair

# (nipāta, fichero VRI, prefijo del «Sutta #» en el Excel)
SECCIONES = [('PAÑCAKA', 's0403m1', '5.'), ('CHAKKA', 's0403m2', '6.')]
OUT = 'validador_an3.json'
DB = 'src/data/tipitaka.sqlite'
AN3_BOOK = 19
XLSX = 'PTS_Reference_Complete_Canon.xlsx'

# índice de Morris: (vagga, nombre, página de arranque). Verdad-terreno estructural del lado PTS.
# Índice de Morris **corregido contra el texto** (ver `anguttara-vol-II-info.txt`): 20 de sus 26
# entradas son la página correcta; `Cara` y `Sañcetanika` traen el nº de sutta en su lugar, y otras
# cuatro discrepan por poco. Aquí van las páginas REALES del primer sutta de cada vagga.
# Índice de Hardy, por sección. Contrastado contra la BD: 25/26 exactos en el Pañcaka (el 26º es
# el peyyāla Upasampadā, sin numerar) y los 12 del Chakka.
HARDY = {
    'PAÑCAKA': [(1, 'Sekhabala', 1), (2, 'Bala', 9), (3, 'Pañcaṅgika', 14), (4, 'Sumana', 32),
                (5, 'Muṇḍarāja', 45), (6, 'Nīvaraṇa', 63), (7, 'Saññā', 79), (8, 'Yodhājīva', 84),
                (9, 'Thera', 110), (10, 'Kakudha', 118), (11, 'Phāsuvihāra', 127),
                (12, 'Andhakavinda', 136), (13, 'Gilāna', 142), (14, 'Rāja', 147),
                (15, 'Tikaṇḍakī', 164), (16, 'Saddhamma', 174), (17, 'Āghāta', 185),
                (18, 'Upāsaka', 203), (19, 'Arañña', 219), (20, 'Brāhmaṇa', 221),
                (21, 'Kimbila', 247), (22, 'Akkosaka', 252), (23, 'Dīghacārika', 257),
                (24, 'Āvāsika', 261), (25, 'Duccarita', 267), (26, 'Upasampadā', 271)],
    'CHAKKA': [(1, 'Āhuneyya', 279), (2, 'Sāraṇīya', 288), (3, 'Anuttariya', 309),
               (4, 'Devatā', 329), (5, 'Dhammika', 344), (6, 'Mahā', 374), (7, 'Devatā', 421),
               (8, 'Arahatta', 429), (9, 'Sīti', 435), (10, 'Ānisaṃsa', 441), (11, 'Tika', 445),
               (12, 'Vaggāsaṅgahitā', 449)],
}


def build_cst_units(stem):
    """Suttas del CST en orden de documento: `[{title, pn, vagga, text}]`.

    En AN el `<div>` es un **vagga** (no un sutta como en DN, ni un saṃyutta como en SN), así que
    el sutta es el `subhead` de dentro y sus paranums son los `bodytext` que le siguen.
    """
    root = ET.parse(f'/tmp/tipitaka-xml/romn/{stem}.mul.xml').getroot()
    out = []
    for d in root.iter('div'):
        h = d.find('head')
        if h is None or h.get('rend') != 'chapter':
            continue
        vagga = ' '.join(''.join(h.itertext()).split())
        cur = None
        for p in d.iter('p'):
            rend, txt = p.get('rend'), ''.join(p.itertext())
            if rend == 'subhead':
                cur = {'title': ' '.join(txt.split()), 'pn': [], 'text': '', 'vagga': vagga}
                out.append(cur)
            elif rend in ('bodytext', 'gatha1', 'gatha2', 'gatha3', 'gathalast') and cur is not None:
                m = re.match(r'(\d+)(?:-(\d+))?$', p.get('n') or '')
                if m:
                    cur['pn'] += list(range(int(m.group(1)), int(m.group(2) or m.group(1)) + 1))
                cur['text'] += ' ' + txt
    return out


def excel_entries(pref):
    """Filas del Tika, separando las **individuales** de las de **rango** (redundantes)."""
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    ind, rng = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[ci['Nikaya']] or '') != 'AN' or str(row[ci['PTS Roman']] or '').strip().lower() != 'iii':
            continue
        num = str(row[ci['Sutta #']])
        if not num.startswith(pref):
            continue
        e = {'num': num, 'name': str(row[ci['Sutta Name']] or '').strip(),
             'page': row[ci['PTS Page']], 'legacy': str(row[ci['Validation']] or ''),
             'dpr': str(row[ci['DPR Ref']] or '').strip()}
        (rng if '-' in num else ind).append(e)
    ind.sort(key=lambda e: int(e['num'].split('.')[1]))
    return ind, rng


def colofones(cur, nipata):
    """Colofones de vagga del nipāta, en orden de lectura. Son la estructura fiable del lado PTS."""
    pages = {r['page_no']: (r['unitext'] or '').split('\n')
             for r in cur.execute('SELECT page_no,unitext FROM pages '
                                  'WHERE edition="mula" AND book_no=?', (AN3_BOOK,))}
    return [x for x in collect_an1(pages)
            if x.get('colofon') is not None and x['nipata'] == nipata]


def vagga_starts(marks, colofones):
    """Página del primer sutta de cada vagga, delimitando por los **colofones** del impreso."""
    out, prev = [], 0
    pos = lambda x: x['page'] * 1000 + x['line']
    for c in colofones:
        ns = [m for m in marks if prev < pos(m) <= pos(c)]
        if ns:
            out.append(ns[0]['page'])
        prev = pos(c)
    return out


def pts_markers(cur, nipata):
    """Marcadores del Tika: nº corrido → `(página, línea, texto acotado)`."""
    pages = {r['page_no']: (r['unitext'] or '').split('\n')
             for r in cur.execute('SELECT page_no,unitext FROM pages '
                                  'WHERE edition="mula" AND book_no=?', (AN3_BOOK,))}
    marks = [x for x in collect_an1(pages) if x['nipata'] == nipata and x['num'] is not None]
    marks.sort(key=lambda x: (x['page'], x['line']))
    flat = [(pg, ln, t) for pg in sorted(pages) for ln, t in enumerate(pages[pg], 1)]
    pos = {(pg, ln): i for i, (pg, ln, _t) in enumerate(flat)}
    out = []
    for k, m in enumerate(marks):
        i0 = pos[(m['page'], m['line'])]
        i1 = pos[(marks[k + 1]['page'], marks[k + 1]['line'])] if k + 1 < len(marks) else len(flat)
        out.append({'num': m['num'], 'page': m['page'], 'line': m['line'], 'ord': k,
                    'vagga': (m['num'] - 1) // 10 + 1,
                    # el texto va SIN recortar: cuando el CST parte en varios lo que PTS numera
                    # como uno, el párrafo homólogo puede caer muy por dentro del sutta (el
                    # `Sarada` es el §4 del nº92, dos páginas más allá) y un tope de 350 tokens lo
                    # dejaba fuera del alcance de `pts_window`
                    'text': ' '.join(sh.tokens(' '.join(t for _p, _l, t in flat[i0:i1])))})
    return out


def align(entries, units):
    """Alineamiento **monótono** filas↔suttas del CST (`align_rows`), por nombre y ordinal.

    No se empareja por clave: el `Sutta #` sólo coincide con el nº corrido de PTS en 103 de 156.
    """
    def score(i, j):
        e, u = entries[i], units[j]
        s = an_names.score(u['title'], e['name'])
        ok = an_names.ordinal_ok(u['title'], e['name'])
        if ok is True:
            s += 60                       # el ordinal desempata dentro de la serie
        elif ok is False:
            s -= 120                      # «Pāpaṇika 1» no puede ser «Dutiya…»
        return s if s > 0 else 0.1        # nunca None: la serie debe poder avanzar

    idx = ar_assign(len(entries), len(units), score, lambda j: 1, skip_penalty=200.0)
    return {entries[i]['num']: (units[j] if j is not None else None) for i, j in enumerate(idx)}


def pts_anchor(e):
    """Nº corrido de PTS que declara la fila: la **`DPR Ref` si existe**, si no el `Sutta #`.

    Es el ancla determinista del lado PTS, y sale del re-parseo de `Raw ID` (`an_rawid.py`): la
    `DPR Ref` describe exactamente los desdoblamientos y fusiones del Excel —`3.32 Ānanda` →
    `AN 3.32a` y `3.33 Sāriputta` → `AN 3.32b` (una del DPR partida en dos), `3.39 Sukhumāla` →
    `AN 3.38-9` (dos fundidas en una)—, que es justo lo que hace que haya **163 marcadores para
    156 filas**. Sin ella el alineamiento acumulaba deriva: `3.111 «Nidāna 1»` acababa en el
    marcador 110 cuando su `Raw ID` dice `[AN 3.107]-08`, o sea los marcadores **107-108**.
    """
    src = e.get('dpr') or ('AN ' + e['num'])
    m = re.search(r'(\d+)\.(\d+)', src)
    if not m:
        # el punto se perdió en la fuente: «AN 3102b» por «AN 3.102b»
        m = re.search(r'AN\s*(\d)(\d+)', src)
    return int(m.group(2)) if m else None


def align_pts(entries, marks, cst):
    """Alineamiento **monótono** filas↔marcadores de PTS, anclado en la PÁGINA del Excel.

    No vale la posición: hay **163 marcadores para 156 filas** —PTS numera por separado suttas que
    el Excel lleva agrupados— y emparejar la fila k con el marcador k+1 acumula deriva (desde
    `3.34` la página se va +2 y sigue creciendo). Sólo cuadraba en 104 de 156.
    """
    # Cuántas filas declaran cada nº corrido: los sufijos `a`/`b` de la `DPR Ref` marcan un sutta
    # de PTS que el Excel **parte en dos filas** (`3.32 Ānanda` → `AN 3.32a`, `3.33 Sāriputta` →
    # `AN 3.32b`), así que ese marcador tiene capacidad 2, no 1.
    anchors = Counter(a for a in (pts_anchor(x) for x in entries) if a is not None)
    reserved = set(anchors)

    def score(i, j):
        e, m = entries[i], marks[j]
        # El nº corrido que declara la fila es evidencia DIRECTA, no una preferencia: va como
        # restricción dura y además se le reserva el marcador, para que ninguna otra fila lo ocupe.
        a = pts_anchor(e)
        if a is not None:
            return 1000.0 if m['num'] == a else None
        if m['num'] in reserved:
            return None
        if not isinstance(e['page'], int):
            return 0.1
        d = abs(m['page'] - e['page'])
        if d > 3:
            return None
        s = 40 - 10 * d
        u = cst.get(e['num'])
        if u:                                   # el título CST refuerza, no decide
            s += an_names.score(u['title'], e['name']) / 10

        return max(s, 0.1)

    idx = ar_assign(len(entries), len(marks), score,
                    lambda j: max(1, anchors.get(marks[j]['num'], 1)), skip_penalty=500.0)
    return {entries[i]['num']: (marks[j] if j is not None else None) for i, j in enumerate(idx)}


def pts_window(pts_text, cst_text, width=350):
    """Acota el texto de PTS al **párrafo** que corresponde al sutta del CST.

    En el Tika, PTS numera como UN sutta lo que el CST parte en varios: el `Saradasuttaṃ` del CST
    es el **§4 del nº92** de PTS (A i 242,17), el `Nimittasuttaṃ` vive dentro del nº100 y el
    `Samaṇabrāhmaṇasuttaṃ` dentro del nº102. La alineación al sutta contenedor es correcta, pero
    cotejar el sutta breve del CST contra el contenedor entero hunde la cobertura (0.06) y produce
    un REJECT que no dice nada de la fila.

    Se localiza el incipit del CST dentro del texto de PTS y se devuelve una ventana desde ahí. Si
    no se encuentra, se deja el texto tal cual: nunca se inventa un locus.
    """
    pt, ct = pts_text.split(), cst_text.split()
    if len(pt) <= width or len(ct) < 6:
        return pts_text
    probe = ct[:6]
    for i in range(len(pt) - len(probe)):
        if pt[i:i + len(probe)] == probe:
            return ' '.join(pt[i:i + width])
    # sin coincidencia exacta: se busca la mejor ventana por solapamiento léxico
    head = set(ct[:60])
    best, bi = -1, None
    for i in range(0, max(1, len(pt) - 60), 20):
        ov = len(head & set(pt[i:i + 60]))
        if ov > best:
            best, bi = ov, i
    return ' '.join(pt[bi:bi + width]) if bi is not None and best >= 12 else pts_text


def main():
    dry = '--dry' in sys.argv
    do_all = '--all' in sys.argv
    n = int(sys.argv[sys.argv.index('--n') + 1]) if '--n' in sys.argv else 30
    only = set(sys.argv[sys.argv.index('--only') + 1].split(',')) if '--only' in sys.argv else None

    conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    SEC = {}
    for nip, stem, pref in SECCIONES:
        u = build_cst_units(stem)
        i2, r2 = excel_entries(pref)
        mk = pts_markers(cur, nip)
        SEC[nip] = {'stem': stem, 'units': u, 'ind': i2, 'rng': r2, 'marks': mk,
                    'colofones': colofones(cur, nip),
                    'by': {m['num']: m for m in mk}, 'cst': align(i2, u)}
        SEC[nip]['pts'] = align_pts(i2, mk, SEC[nip]['cst'])
    ind = [e for nip, _s, _p in SECCIONES for e in SEC[nip]['ind']]
    rng = [e for nip, _s, _p in SECCIONES for e in SEC[nip]['rng']]
    units = [u for nip, _s, _p in SECCIONES for u in SEC[nip]['units']]
    marks = [m for nip, _s, _p in SECCIONES for m in SEC[nip]['marks']]
    cst = {k: v for nip, _s, _p in SECCIONES for k, v in SEC[nip]['cst'].items()}
    pts = {k: v for nip, _s, _p in SECCIONES for k, v in SEC[nip]['pts'].items()}
    STEM = {e['num']: SEC[nip]['stem'] for nip, _s, _p in SECCIONES for e in SEC[nip]['ind']}

    print('=' * 92)
    print('AN III — PAÑCAKA y CHAKKA: pipeline VRI + estructura por colofones (anguttara-vol-III-info.txt)')
    print('=' * 92)
    # Los arranques de vagga se derivan de los COLOFONES, no de la aritmética `10k+1`: en el
    # Chakka los vaggas son de 10, 12, 12, 11 y 11 suttas —lo que Hardy mismo documenta («from one
    # to ten **or eleven**, if there are more Suttas than ten in a Vagga»)— y suponerlos uniformes
    # descuadraba 8 de sus 12.
    ok16 = tot16 = 0
    for nip, _s, _p in SECCIONES:
        starts = vagga_starts(SEC[nip]['marks'], SEC[nip]['colofones'])
        for k, (_v, _nm, pg) in enumerate(HARDY[nip]):
            tot16 += 1
            if k < len(starts) and starts[k] == pg:
                ok16 += 1
    print('PTS (BD libro 19): ' + ' | '.join(
        f"{nip} {len(SEC[nip]['marks'])} marc. (1..{max(SEC[nip]['by'])})"
        for nip, _s, _p in SECCIONES))
    print(f'Excel: {len(ind)} individuales + {len(rng)} de rango | CST: {len(units)} suttas')
    print(f'estructura vs índice de Hardy: {ok16}/{tot16} páginas de arranque de vagga exactas')

    tasks, no_cst, no_pts, name_ok, ord_ok, ord_n, page_ok = [], [], [], 0, 0, 0, 0
    for k, e in enumerate(ind):
        if only is not None and e['num'] not in only:
            continue
        u = cst.get(e['num'])
        if not u:
            no_cst.append(e); continue
        if an_names.score(u['title'], e['name']) >= 60:
            name_ok += 1
        oo = an_names.ordinal_ok(u['title'], e['name'])
        if oo is not None:
            ord_n += 1; ord_ok += bool(oo)
        m = pts.get(e['num'])
        if not m:
            no_pts.append(e); continue
        if isinstance(e['page'], int) and abs(m['page'] - e['page']) <= 1:
            page_ok += 1
        tasks.append((e, m, u))

    tot = max(1, len(ind))
    print(f'\nalineadas con sutta del CST: {len(ind) - len(no_cst)}/{len(ind)}')
    print(f'  cruzada — nombre PTS ≡ título CST: {name_ok}/{tot} ({100 * name_ok / tot:.0f}%)')
    print(f'  cruzada — ordinal (dígito PTS ≡ paṭhama/dutiya del CST): {ord_ok}/{ord_n}')
    print(f'  cruzada — página del marcador ≡ página del Excel (±1): {page_ok}/{tot} '
          f'({100 * page_ok / tot:.0f}%)  | sin marcador: {len(no_pts)}')
    if dry:
        print('\n(dry-run, nada validado ni escrito)')
        return

    run = tasks if (do_all or only) else tasks[:n]
    print(f'\nValidando {len(run)} (Gemini en vivo)...', flush=True)
    rows = []
    for k, (e, m, u) in enumerate(run, 1):
        cstx = ' '.join(sh.tokens(u['text'])[:350])
        ptsx = pts_window(m['text'], cstx)
        res = validate_pair(ptsx, cstx, e['name'], u['title'], 'AN', concordant=True)
        rows.append({'num': e['num'], 'name': e['name'], 'cst_title': u['title'],
                     'vri_ref': f"{STEM[e['num']]}:{min(u['pn'])}" if u['pn'] else None,
                     'pts_num': m['num'], 'pts_page': m['page'], 'pts_line': m['line'],
                     'legacy': e['legacy'], **res})
        if k % 25 == 0:
            print(f'  ...{k}/{len(run)}', flush=True)
    prev = json.load(open(OUT)) if os.path.exists(OUT) else []
    merged = {r['num']: r for r in prev}
    merged.update({r['num']: r for r in rows})
    json.dump(list(merged.values()), open(OUT, 'w'), ensure_ascii=False, indent=1)
    print(f"\nEstado: {dict(Counter(r['estado'] for r in rows))} | "
          f"Validation: {dict(Counter(r['validation'] for r in rows))}")
    print(f'Resultados → {OUT} ({len(merged)} en total; nada escrito al Excel).')


if __name__ == '__main__':
    main()
