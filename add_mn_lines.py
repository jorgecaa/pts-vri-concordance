#!/usr/bin/env python3
"""
MN line numbers — use sutta number marker pattern (centered number like "82.")
as the sutta start marker, following MN volume conventions.
"""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def find_sutta_line(book_no, page_no, sutta_num_str):
    """Find PTS line number where sutta begins. Returns (line, method)."""
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
               (book_no, page_no))
    r = cur.fetchone()
    if not r: return None, 'missing'
    
    head = r['head'] or ''
    text = r['unitext'] or ''
    lines = text.split('\n')
    ns = sutta_num_str.strip()
    
    # 1. Is sutta header in page HEAD? → starts at first text line
    if f'({ns})' in head or f'({ns}.)' in head:
        for i, line in enumerate(lines):
            if line.strip() and len(line.strip()) > 3:
                return i + 1, 'page_top'
        return 1, 'page_top'
    
    # 2. Search body for CENTERED sutta number: "82." alone on a line
    # Pattern: line consists of just a number + period, possibly with whitespace
    for i, line in enumerate(lines):
        stripped = line.strip()
        # Match: "82." or " 82. " — just the number and period
        if re.match(rf'^{re.escape(ns)}\.?$', stripped):
            return i + 1, f'num_marker'
    
    # 3. Number followed by text? "82. Evam me sutam..."
    for i, line in enumerate(lines):
        stripped = line.strip()
        if re.match(rf'^{re.escape(ns)}\.\s+', stripped):
            return i + 1, 'num+text'
    
    # 4. Opening formula as fallback
    for i, line in enumerate(lines):
        if re.search(r'evam\s+me\s+suta[mṃ]', line, re.IGNORECASE):
            return i + 1, 'evam_me'
    
    return None, 'not_found'

MN_BOOKS = {1:9, 2:10, 3:11}

wb = load_workbook(XL)
ws = wb['Complete Canon']

cols = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=1, column=c).value
    if h: cols[h] = c

results = {'page_top':0, 'num_marker':0, 'num+text':0, 'evam_me':0, 'not_found':0}

print('MN line numbers — using sutta number markers')
print('═'*60)

for row_idx in range(2, ws.max_row+1):
    if ws.cell(row=row_idx, column=cols['Nikaya']).value != 'MN': continue
    
    sn = str(ws.cell(row=row_idx, column=cols['Sutta #']).value or '')
    page = ws.cell(row=row_idx, column=cols['PTS Page']).value
    roman = str(ws.cell(row=row_idx, column=cols['PTS Roman']).value or '').strip()
    vol = str(ws.cell(row=row_idx, column=cols['PTS Vol']).value or '').strip()
    ref = str(ws.cell(row=row_idx, column=cols['PTS Ref']).value or '')
    
    rm = {'i':1,'ii':2,'iii':3}
    book_no = MN_BOOKS.get(rm.get(roman.lower(), 1))
    if not book_no or not page: continue
    
    line_num, method = find_sutta_line(book_no, page, sn)
    results[method] = results.get(method, 0) + 1
    
    if line_num:
        # Format: "M i 7" or "M i 7,24" (only add comma if line > 1)
        if line_num == 1:
            new_ref = re.sub(r',\d+$', '', ref)  # remove any existing comma
        else:
            new_ref = re.sub(r',\d+$', '', ref)  # remove old
            new_ref = f'{new_ref},{line_num}'     # add new
        
        if new_ref != ref:
            ws.cell(row=row_idx, column=cols['PTS Ref']).value = new_ref
    
    if method in ('num_marker', 'num+text', 'evam_me', 'not_found'):
        marker = {'num_marker':'N°', 'num+text':'N°+', 'evam_me':'evaṃ', 'not_found':'✗'}.get(method, method)
        ln_str = str(line_num) if line_num else '?'
        print(f'  MN {sn:>3s} p.{page:>3d} L{ln_str:>3s}  [{marker}]')

wb.save(XL)
print(f"\n  page_top:  {results['page_top']}  |  num_marker: {results['num_marker']}")
print(f"  num+text:  {results['num+text']}  |  evam_me:    {results['evam_me']}")
print(f"  not_found: {results['not_found']}")
print(f'  Total: {sum(results.values())}/152')
print(f'  Saved: {XL}')
conn.close()
