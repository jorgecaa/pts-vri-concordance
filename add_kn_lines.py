#!/usr/bin/env python3
"""
KN Line Numbers v5 — final fixes:
- Ja: extract roman from ref for correct book_no, preserve existing lines
- Nidd: add 'Nidd' to vol mapping
- Don't overwrite entries that already have line numbers
"""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def find_verse_starts(book_no, page_no):
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r or not r['unitext']: return []
    lines = r['unitext'].split('\n')
    ends = [i for i, line in enumerate(lines) if re.search(r'[║|]\s*\d+\s*[║|]', line)]
    starts = []
    for i, line in enumerate(lines):
        s = line.strip()
        if not s: continue
        if re.match(r'^[A-ZĀĪŪṄÑṬḌṆḶ\s\-\.\,\;\:║|]+$', s) and len(s) > 3: continue
        if re.match(r'^[\d\s\.\,\;\:\-║|]+$', s): continue
        if len(s) >= 8: starts.append(i+1); break
    for ep in ends:
        for j in range(ep+1, min(ep+6, len(lines))):
            if lines[j].strip() and len(lines[j].strip()) > 5: starts.append(j+1); break
    return sorted(set(starts))

def find_markers(book_no, page_no):
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r or not r['unitext']: return []
    return [i+1 for i, line in enumerate(r['unitext'].split('\n')) if re.match(r'^\d+\.\s+\S', line.strip())]

def find_evam_me(book_no, page_no):
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r or not r['unitext']: return []
    return [i+1 for i, line in enumerate(r['unitext'].split('\n')) if re.search(r'[Ee]va[mṃ]\s+me\s+suta[mṃ]', line)]

def find_first_content(book_no, page_no):
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r or not r['unitext']: return [1]
    for i, line in enumerate(r['unitext'].split('\n')):
        s = line.strip()
        if not s: continue
        if re.match(r'^[A-ZĀĪŪṄÑṬḌṆḶ\s\-\.\,\;\:║|]+$', s) and len(s) > 3: continue
        if re.match(r'^[\d\s\.\,\;\:\-║|]+$', s): continue
        if len(s) >= 6: return [i+1]
    return [1]

BOOK_DB = {
    'Khp':22,'Kh':22,'Dhp':23,'Dh':23,'Ud':24,'It':25,'Sn':26,
    'Vv':27,'Pv':28,'Th':29,'Th & Th':29,'Thi':29,
    'Ja':30,'Ja I':30,'Ja II':31,'Ja III':32,'Ja IV':33,'Ja V':34,'Ja VI':35,
    'Nidd':36,'Nidd I':36,'Nidd II':37,
    'Patis I':38,'Patis II':39,'Paṭis I':38,'Paṭis II':39,'Ap':40,'Bv':41,'Cp':42,
}
PIPELINE = {
    'Th':'verse','Th & Th':'verse','Thi':'verse',
    'Ap':'number','Ja':'ja','Ja I':'ja','Ja II':'ja','Ja III':'ja','Ja IV':'ja','Ja V':'ja','Ja VI':'ja',
    'Dhp':'number','Dh':'number','Khp':'number','Kh':'number',
    'It':'number','Sn':'number','Ud':'evam_me',
    'Vv':'page_top','Pv':'page_top',
    'Nidd':'page_top','Nidd I':'page_top','Nidd II':'page_top',
    'Patis I':'page_top','Patis II':'page_top','Paṭis I':'page_top','Paṭis II':'page_top','Bv':'page_top','Cp':'page_top',
}
JA_ROMAN_MAP = {'i':30,'ii':31,'iii':32,'iv':33,'v':34,'vi':35}

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}

entries_by_page = defaultdict(list)
for ri in range(2, ws.max_row+1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'KN': continue
    vol = str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip()
    snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    if not page: continue
    if (not snum or snum == 'None') and name in ('Therapadana','Theriapadana','Jataka','Niddesa','Mahaniddesa','Culaniddesa','Patisambhidamagga','Nettipakaraṇa','Petakopadesa','Milindapa ñha'):
        continue
    # Skip entries that already have line numbers
    has_line = ',' in ref
    entries_by_page[(vol, page)].append({
        'ri': ri, 'snum': snum, 'name': name, 'ref': ref,
        'has_line': has_line,
    })

print('KN Line Numbers v5')
print('=' * 70)
stats = defaultdict(lambda: {'found':0,'miss':0,'skip':0})
fixed = 0

for (vol, page), entries in sorted(entries_by_page.items()):
    # Resolve pipeline and book_no
    ptype = PIPELINE.get(vol, 'skip')
    
    if ptype == 'ja':
        # For Ja, determine book_no from the roman in the ref
        book_no = BOOK_DB.get(vol, 30)
        # Try to refine from ref
        for e in entries:
            m = re.match(r'Ja\s+([ivxlcdm]+)\s+\d+', e['ref'], re.IGNORECASE)
            if m:
                book_no = JA_ROMAN_MAP.get(m.group(1).lower(), book_no)
                break
    else:
        book_no = BOOK_DB.get(vol)
    
    if not book_no or ptype == 'skip':
        for e in entries: stats[vol]['miss'] += 1
        continue
    
    # Get only entries that need line numbers
    need_lines = [e for e in entries if not e['has_line']]
    skip_count = sum(1 for e in entries if e['has_line'])
    stats[vol]['skip'] += skip_count
    
    if not need_lines:
        stats[vol]['found'] += skip_count
        continue
    
    # Get markers
    if ptype == 'verse':
        markers = find_verse_starts(book_no, page)
    elif ptype in ('number', 'ja'):
        markers = find_markers(book_no, page)
        if not markers: markers = find_first_content(book_no, page)
    elif ptype == 'evam_me':
        markers = find_evam_me(book_no, page)
        if not markers: markers = find_markers(book_no, page)
        if not markers: markers = find_first_content(book_no, page)
    elif ptype == 'page_top':
        markers = find_first_content(book_no, page)
        markers = sorted(set(markers + find_markers(book_no, page)))
    
    # Assign markers sequentially to entries that need lines
    for ei, e in enumerate(need_lines):
        if ei < len(markers) and markers[ei]:
            line = markers[ei]
            stats[vol]['found'] += 1
            new_ref = re.sub(r',\d+$', '', e['ref'])
            if line > 1: new_ref = f'{new_ref},{line}'
            if new_ref != e['ref']:
                ws.cell(row=e['ri'], column=cols['PTS Ref']).value = new_ref
                fixed += 1
        else:
            stats[vol]['miss'] += 1
    
    stats[vol]['found'] += skip_count

wb.save(XL)

print()
grand_f = grand_m = 0
for vol in sorted(stats.keys()):
    s = stats[vol]; t = s['found']+s['miss']
    if t == 0: continue
    f = s['found']; pct = 100*f/max(t,1)
    bar = '█'*int(pct/5)+'░'*(20-int(pct/5))
    print(f'  {vol:>12s}: {f:>4d}/{t:<4d} ({pct:5.1f}%) {bar}')
    grand_f += f; grand_m += s['miss']

print(f'\n  {"TOTAL":>12s}: {grand_f}/{grand_f+grand_m} ({100*grand_f/max(grand_f+grand_m,1):.1f}%)')
print(f'  Refs fixed: {fixed}')
print(f'  Saved: {XL}')
conn.close()
