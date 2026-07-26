#!/usr/bin/env python3
"""
check_integrity — invariantes del proyecto: caza la **corrupción silenciosa** que el validador
no puede ver.

**Por qué existe.** El validador coteja un par PTS↔CST y dice si concuerdan. No puede detectar que
el par que recibe sea el equivocado, ni que el texto que se le entrega esté corrompido: si los dos
lados son coherentes *entre sí*, aprueba. Todos los defectos graves de este proyecto han vivido en
ese punto ciego, y ninguno lo destapó el validador:

| defecto real | cómo se veía | lo habría cazado |
|---|---|---|
| el tokenizador partía `n' atthi` y descartaba la `n` suelta, así que **«no hay» pasaba a «hay»** | texto perfectamente legible, sentido invertido | `tokens_no_pierden_palabras` |
| el importador de AN partió el `Sutta #` y dejó el resto en el nombre (166 filas) | nombres empezando por `-5 [AN…` | `nombre_sin_residuo_de_parseo` |
| una fila de AN llevaba una referencia del Saṃyutta (`S i 68,4`) | página imposible, 68 en un nipāta que abre en 101 | `prefijo_de_ref_coincide_con_nikaya` |
| 36 claves `Sutta #` duplicadas impedían emparejar | ninguna señal | `clave_sutta_no_duplicada` |
| asignaciones no inyectivas: dos filas sobre un marcador, una validada contra el locus de su vecina | ninguna señal | `audit_injectivity.py` (aparte) |

Uso:
    python3 check_integrity.py              # todas las invariantes
    python3 check_integrity.py --quick      # salta las que leen la BD entera
    python3 check_integrity.py -v           # lista cada incidencia

Devuelve 0 si todo pasa, 1 si alguna invariante falla. **Correr antes y después de cualquier
cambio que toque el texto, el Excel o un alineador.**
"""
import re
import sqlite3
import sys
from collections import Counter, defaultdict

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
DB = 'src/data/tipitaka.sqlite'

PREFIJO = {'DN': 'D', 'MN': 'M', 'SN': 'S', 'AN': 'A'}
ESTADOS = {'CONFIRMADO', 'PENDIENTE'}
CONFIRMA = {'VALIDADOR', 'VALIDADOR_HUMANO', 'PTS_CROSSREF_SN',
            'HELMER_APPROVED', 'HELMER_PTS_TRUNCATED', 'HELMER_FIXED'}

_checks = []


def check(nombre, porque, nivel='fallo'):
    """Registra una invariante. `porque` explica el incidente real que la motiva.

    `nivel='aviso'` para las que tienen excepciones legítimas conocidas: se listan pero no hacen
    fallar la suite. Una comprobación que grita en falso acaba ignorándose, y entonces no sirve
    para nada.
    """
    def deco(fn):
        fn.nombre, fn.porque, fn.nivel = nombre, porque, nivel
        _checks.append(fn)
        return fn
    return deco


def _excel():
    from openpyxl import load_workbook
    ws = load_workbook(XLSX, read_only=True)['Complete Canon']
    it = ws.iter_rows(values_only=True)
    hdr = [c for c in next(it)]
    ci = {n: i for i, n in enumerate(hdr)}
    return [dict(zip(hdr, r)) for r in it], ci


# ── A. preparación del texto: el punto ciego más peligroso ───────────────────
@check('tokens_no_pierden_palabras',
       'El tokenizador partía «n\' atthi» y descartaba la «n» como ruido de un carácter: '
       '«no hay acción» llegaba al validador como «hay acción». Sentido INVERTIDO, texto legible, '
       '12.561 casos en DN/MN/AN. Es el arquetipo de la corrupción que nadie ve.')
def inv_tokens(verbose=False, quick=False):
    import sutta_hash as sh
    casos = [("n' atthi kammaṃ", 'natthi'), ("c' eva", 'ceva'), ("ten' upasaṅkami", 'tenupasaṅkami'),
             ("kālass' eva", 'kālasseva'), ("pan' etaṃ", 'panetaṃ'), ("diṭṭh' eva", 'diṭṭheva')]
    mal = []
    for src, esperado in casos:
        got = sh.tokens(src)
        if esperado not in got:
            mal.append(f'{src!r} → {got} (falta {esperado!r})')
    # ninguna palabra del original puede DESAPARECER: cada token de salida debe venir de una
    # secuencia contigua del original, y el nº de caracteres alfabéticos no puede caer
    for src, _e in casos:
        a = len(re.sub(r'[^a-zāīūṃṅñṭḍṇḷ]', '', src.lower()))
        b = len(''.join(sh.tokens(src)))
        if b < a:
            mal.append(f'{src!r}: se pierden {a - b} caracteres al tokenizar')
    return mal


@check('negacion_preservada',
       'Corolario del anterior, comprobado sobre el texto REAL de la BD: si una página contiene '
       'una negación elidida, sus tokens deben seguir conteniéndola.')
def inv_negacion(verbose=False, quick=False):
    if quick:
        return []
    import sutta_hash as sh
    db = sqlite3.connect(DB); db.row_factory = sqlite3.Row
    mal = []
    q = db.execute("SELECT book_no,page_no,unitext FROM pages WHERE edition='mula' "
                   "AND unitext LIKE \"%n' atthi%\" LIMIT 40")
    for r in q:
        n_src = len(re.findall(r"n'\s*atthi", r['unitext']))
        n_tok = ' '.join(sh.tokens(r['unitext'])).count('natthi')
        if n_tok < n_src:
            mal.append(f"libro {r['book_no']} p{r['page_no']}: {n_src} «n' atthi» → sólo "
                       f"{n_tok} «natthi» tras tokenizar")
    return mal


# ── B. invariantes del Excel maestro ─────────────────────────────────────────
@check('prefijo_de_ref_coincide_con_nikaya',
       'Una fila de AN llevaba «S i 68,4» — una referencia del Saṃyutta. La p68 es imposible en '
       'un nipāta que abre en la 101, pero nada lo señalaba.')
def inv_prefijo(verbose=False, quick=False):
    rows, _ = _excel()
    mal = []
    for r in rows:
        nk, ref = str(r.get('Nikaya') or ''), str(r.get('PTS Ref') or '').strip()
        if nk not in PREFIJO or not ref:
            continue
        m = re.match(r'([A-Z][a-z]?)\s', ref)
        if m and m.group(1) != PREFIJO[nk]:
            mal.append(f"{nk} {r.get('Sutta #')} «{r.get('Sutta Name')}» ref={ref!r}")
    return mal


@check('pagina_coincide_con_la_ref',
       'La página de la columna y la de la referencia deben ser la misma; si divergen, una de las '
       'dos se editó sin la otra.')
def inv_pagina(verbose=False, quick=False):
    rows, _ = _excel()
    mal = []
    for r in rows:
        ref, pg = str(r.get('PTS Ref') or ''), r.get('PTS Page')
        m = re.match(r'[A-Z][a-z]?\s+[ivx]+\s+(\d+)', ref)
        if m and isinstance(pg, int) and int(m.group(1)) != pg:
            mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')} página={pg} vs ref={ref!r}")
    return mal


@check('nombre_sin_residuo_de_parseo',
       'El importador de AN partió el «Sutta #» por el sitio equivocado y dejó la cola en el '
       'nombre: 166 filas empezaban por «-5 [AN 1.1.1-5]: …» y 1126 por «: ». Un nombre que '
       'empieza por «-», «:» o que va vacío es residuo de parseo, no un nombre.')
def inv_nombre(verbose=False, quick=False):
    rows, _ = _excel()
    mal = []
    for r in rows:
        nm = str(r.get('Sutta Name') or '').strip()
        if not nm:
            mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')}: nombre VACÍO")
        elif nm[0] in '-:':
            mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')}: nombre empieza por {nm[0]!r} → {nm[:40]!r}")
    return mal


@check('sin_mojibake',
       'Seis filas de AN traían «á¹hāna» donde debe leerse «Ṭhāna»: a la Ṭ (U+1E6C) se le perdió '
       'el último byte en la fuente. Un nombre así no casa con nada y el fallo se atribuye al '
       'alineador en vez de al dato.')
def inv_mojibake(verbose=False, quick=False):
    rows, _ = _excel()
    # El segundo carácter debe estar en el rango de un byte de continuación leído como latin-1
    # (U+0080-U+00BF). Con un rango más amplio se capturaba el español normal: en «Página», la «á»
    # va seguida de «g», y una comprobación que grita en falso acaba ignorándose.
    moji = re.compile('[\u00c3\u00c2\u00e1\u00e2][\u0080-\u00bf]')
    mal = []
    for r in rows:
        for col in ('Sutta Name', 'Section', 'Detail'):
            v = str(r.get(col) or '')
            if moji.search(v):
                mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')} [{col}] {v[:38]!r}")
    return mal


@check('clave_sutta_no_duplicada',
       'Tras el defecto del importador quedaron 36 claves «Sutta #» duplicadas dentro de un mismo '
       'volumen, y con ellas es imposible emparejar por clave.')
def inv_clave(verbose=False, quick=False):
    rows, _ = _excel()
    # KN no numera sus suttas en esta columna: una clave vacía no es una clave duplicada
    c = Counter((str(r.get('Nikaya')), str(r.get('PTS Roman') or '').strip().lower(),
                 str(r.get('Sutta #'))) for r in rows
                if str(r.get('Sutta #') or '').strip() not in ('', 'None'))
    return [f'{k[0]} {k[1]} #{k[2]} × {v}' for k, v in sorted(c.items()) if v > 1]


@check('sutta_num_coincide_con_raw_id',
       '`Raw ID` conserva la cadena original de la fuente. Divergir de ella es legítimo SÓLO si '
       'hubo una re-identificación deliberada y documentada (S ii: el «Sutta #» de SN 17 iba '
       'desplazado +7 y se corrigió con `reid_sn2.py`). Cualquier otra divergencia es una edición '
       'fuera de la fuente.', nivel='aviso')
def inv_rawid(verbose=False, quick=False):
    rows, _ = _excel()
    mal = []
    for r in rows:
        raw, num = str(r.get('Raw ID') or ''), str(r.get('Sutta #'))
        m = re.match(r'^[A-Z]{2}\s+([\d.\-]+?)(?=\s|\[|:|$)', raw)
        if m and m.group(1) != num:
            mal.append(f"{r.get('Nikaya')} #{num} vs Raw ID «{m.group(1)}» ({raw[:44]!r})")
    return mal


@check('estado_coherente_con_validation',
       '`Estado` es el resumen binario de `Validation`. Si se editan por separado dejan de decir '
       'lo mismo, y `Estado` es lo que se publica.')
def inv_estado(verbose=False, quick=False):
    rows, _ = _excel()
    mal = []
    for r in rows:
        est, val = str(r.get('Estado') or ''), str(r.get('Validation') or '')
        if est and est not in ESTADOS:
            mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')}: Estado inválido {est!r}")
            continue
        esperado = 'CONFIRMADO' if val in CONFIRMA else 'PENDIENTE'
        if est and est != esperado:
            mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')}: Validation={val!r} pero Estado={est!r}")
    return mal


_CAPS = {}
_PNS = {}


def _pn_existe(stem, pn):
    """¿Existe ese paranum en el fichero VRI?"""
    if stem not in _PNS:
        try:
            import an_peyyala as ap
            _PNS[stem] = {x['pn'] for x in ap.cst_unidades(stem)}
        except Exception:
            _PNS[stem] = None
    return _PNS[stem] is None or pn in _PNS[stem]


def _cap_existe(stem, cap):
    """¿Tiene el fichero VRI ese capítulo (`<div rend="chapter">`)?"""
    import xml.etree.ElementTree as ET
    if stem not in _CAPS:
        try:
            root = ET.parse(f'/tmp/tipitaka-xml/romn/{stem}.mul.xml').getroot()
            _CAPS[stem] = sum(1 for d in root.iter('div')
                              if d.find('head') is not None
                              and d.find('head').get('rend') == 'chapter')
        except Exception:
            _CAPS[stem] = None
    n = _CAPS[stem]
    return n is None or 1 <= cap <= n


@check('vri_ref_bien_formada_y_existente',
       'La `VRI Ref` es la clave canónica del lado CST (regla (5)). Si apunta a un paranum que no '
       'existe en el XML, la fila está anclada a nada.')
def inv_vri(verbose=False, quick=False):
    if quick:
        return []
    import xml.etree.ElementTree as ET
    rows, _ = _excel()
    cache, mal = {}, []
    for r in rows:
        v = str(r.get('VRI Ref') or '').strip()
        if not v:
            continue
        # ⚠️ FORMA NUEVA para KN: `<fichero>:c<capítulo>[.<item>]`. En SN y AN el `n` del CST es un
        # paranum **corrido por fichero** y basta con él; en KN **reinicia dentro de cada
        # capítulo** —el `s0501m` tiene un `n=1` en el capítulo 2, otro en el 4, otro en el 5…— y
        # además hay capítulos sin `n` ninguno (Saraṇattaya, Dvattiṃsākāra). Un `s0501m:1` no
        # identificaría nada. La `c` marca que la referencia es al capítulo.
        mc = re.match(r'^([a-z0-9]+):c(\d+)(?:\.(\d+))?(?:-c?(\d+)(?:\.(\d+))?)?$', v)
        if mc:
            if not _cap_existe(mc.group(1), int(mc.group(2))):
                mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')}: el capítulo {mc.group(2)} no "
                           f"existe en {mc.group(1)}")
            continue
        # ⚠️ FORMA `<fichero>:<paranum>.<ítem>` — cuando el Excel numera más fino que el CST y la
        # subdivisión va **dentro** de un paranum: el peyyāla que `s0302m:73` mete entero y numera
        # `(2)`…`(12)` en su propio texto, o los frutos de la Aṭṭhikasaññā dentro de `s0305m:238`.
        # Es la tercera extensión de la clave, hermana de las dos de KN, y por el mismo motivo: la
        # unidad numerada no siempre es el paranum.
        mi = re.match(r'^([a-z0-9]+):(\d+)\.(\d+)$', v)
        if mi:
            if not _pn_existe(mi.group(1), int(mi.group(2))):
                mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')}: el paranum {mi.group(2)} no "
                           f"existe en {mi.group(1)}")
            continue
        m = re.match(r'^([a-z0-9]+):(\d+)(?:-(\d+))?$', v)
        if not m:
            mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')}: VRI Ref mal formada {v!r}")
            continue
        stem = m.group(1)
        if stem not in cache:
            try:
                root = ET.parse(f'/tmp/tipitaka-xml/romn/{stem}.mul.xml').getroot()
                s = set()
                for p in root.iter('p'):
                    mm = re.match(r'(\d+)(?:-(\d+))?$', p.get('n') or '')
                    if mm:
                        s |= set(range(int(mm.group(1)), int(mm.group(2) or mm.group(1)) + 1))
                cache[stem] = s
            except Exception:
                cache[stem] = None
        pns = cache[stem]
        if pns is None:
            continue                                    # XML no disponible: no es fallo del dato
        if int(m.group(2)) not in pns:
            mal.append(f"{r.get('Nikaya')} {r.get('Sutta #')}: paranum {m.group(2)} no existe en {stem}")
    return mal


@check('confirmado_exige_procedencia',
       'Regla del proyecto: sin el validador, nada es CONFIRMADO. Una fila CONFIRMADO cuya '
       '`Validation` no esté en la lista de procedencias válidas es una firma sin respaldo.')
def inv_procedencia(verbose=False, quick=False):
    rows, _ = _excel()
    return [f"{r.get('Nikaya')} {r.get('Sutta #')}: CONFIRMADO con Validation={r.get('Validation')!r}"
            for r in rows
            if str(r.get('Estado') or '') == 'CONFIRMADO'
            and str(r.get('Validation') or '') not in CONFIRMA]


@check('paginas_monotonas_por_volumen',
       'Dentro de un volumen las páginas deben crecer con el orden de las filas. Un salto hacia '
       'atrás delata una fila mal colocada o una referencia de otro nikāya (así se vio la fila de '
       'AN que llevaba «S i 68,4»).')
def inv_monotonia(verbose=False, quick=False):
    rows, _ = _excel()
    por = defaultdict(list)
    for r in rows:
        pg = r.get('PTS Page')
        if not isinstance(pg, int):
            continue
        nk = str(r.get('Nikaya'))
        num = str(r.get('Sutta #') or '').strip()
        # KN reúne OBRAS distintas, cada una con su propia paginación y sin volumen romano: se
        # agrupa por obra (el primer componente del «Sutta #»), o la monotonía no significa nada.
        # Las filas de KN sin número quedan fuera: no hay con qué agruparlas, y KN aún no se ha
        # procesado — comprobarlas sólo generaría ruido.
        if nk == 'KN' and not num:
            continue
        vol = str(r.get('PTS Roman') or '').strip().lower()
        clave = (nk, vol, num.split('.')[0] if nk == 'KN' else '')
        por[clave].append((num, pg))
    mal = []
    for k, v in sorted(por.items()):
        for i in range(len(v) - 1):
            if v[i + 1][1] < v[i][1] - 1:               # −1 de tolerancia: filas que comparten página
                mal.append(f'{k[0]} {k[1]}: #{v[i][0]} p{v[i][1]} → #{v[i+1][0]} p{v[i+1][1]}')
    return mal


def main():
    verbose = '-v' in sys.argv or '--verbose' in sys.argv
    quick = '--quick' in sys.argv
    print('=' * 88)
    print('INTEGRIDAD — invariantes que cazan lo que el validador no puede ver')
    print('=' * 88)
    fallos = 0
    for fn in _checks:
        try:
            mal = fn(verbose=verbose, quick=quick)
        except Exception as e:                          # una invariante rota no debe ocultar el resto
            print(f'  ⚠️  {fn.nombre}: ERROR al ejecutar — {e}')
            fallos += 1
            continue
        if mal:
            es_fallo = getattr(fn, 'nivel', 'fallo') == 'fallo'
            fallos += 1 if es_fallo else 0
            print(f'\n  {"✗" if es_fallo else "!"} {fn.nombre}: {len(mal)} incidencia(s)'
                  + ('' if es_fallo else '  [aviso, no hace fallar]'))
            print(f'      ({fn.porque})')
            for x in (mal if verbose else mal[:6]):
                print(f'      · {x}')
            if not verbose and len(mal) > 6:
                print(f'      … y {len(mal) - 6} más (usa -v)')
        else:
            print(f'  ✓ {fn.nombre}')
    print('\n' + '=' * 88)
    if fallos:
        print(f'{fallos} invariante(s) con incidencias. Revisar antes de dar por bueno el dato.')
    else:
        print('Todas las invariantes pasan.')
    print('Nota: la inyectividad de las asignaciones se comprueba aparte, con audit_injectivity.py')
    return 1 if fallos else 0


if __name__ == '__main__':
    sys.exit(main())
