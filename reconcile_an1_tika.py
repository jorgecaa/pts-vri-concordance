#!/usr/bin/env python3
"""reconcile_an1_tika.py — vuelca `validador_an1_tika.json` al Excel maestro (AN I, Tika-nipāta).

Regla del proyecto: CONFIRMADO = concordancia ∧ Gemini APPROVE → `Validation=VALIDADOR`.
Gemini REJECT → **no se auto-firma**: queda como `REVISAR`/PENDIENTE para arbitraje humano.

Escribe además la **`VRI Ref`** (`s0402m2:N`), que es la clave canónica del lado CST (regla (5)):
es la primera vez que una fila de AN la recibe, y viene del alineador, no de `massive.tsv` —que en
AN no sirve de puente (auditado por nombre, sólo 66 de 1738 filas concordaban).

Los valores legados de estas filas (`OK+RTE`, `RTE_ONLY`, `OK`, `UNVERIFIED`) vienen del pase RTE
(BUDSIR), **fuera de alcance** del proyecto: no se heredan, se sobrescriben con el veredicto del
validador.

⚠️ Las **9 filas de rango** del Tika (`3.2-9`, `3.163-182`…) no entran aquí: son redundantes con
filas individuales que ya existen y están a arbitraje, como en su día `12.74` (S ii) y `22.148`
(S iii).

Uso:
    python3 reconcile_an1_tika.py --dry
    python3 reconcile_an1_tika.py --write [--humano firmas.json]
"""
import datetime
import json
import shutil
import sys
from collections import Counter

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
RES = 'validador_an1_tika.json'


def main():
    dry = '--write' not in sys.argv
    humano = {}
    if '--humano' in sys.argv:
        humano = json.load(open(sys.argv[sys.argv.index('--humano') + 1]))
    res = {r['num']: r for r in json.load(open(RES))}

    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}
    vcol, ecol, vri = ci['Validation'] + 1, ci['Estado'] + 1, ci['VRI Ref'] + 1

    plan, writes, pend = Counter(), [], []
    for row in ws.iter_rows(min_row=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'AN' or str(v[ci['PTS Roman']] or '').strip().lower() != 'i':
            continue
        num = str(v[ci['Sutta #']])
        if not num.startswith('3.'):
            continue
        if '-' in num:
            plan['fila de rango (a arbitraje, no se toca)'] += 1
            continue
        r = res.get(num)
        if not r:
            plan['sin resultado'] += 1
            continue
        if num in humano:
            val = humano[num]
            est = 'PENDIENTE' if val in ('REVISAR', 'REF_ERROR_DPR') else 'CONFIRMADO'
            plan[f'humano:{val}'] += 1
        elif r['gemini'] == 'APPROVE':
            val, est = 'VALIDADOR', 'CONFIRMADO'
            plan['VALIDADOR(auto)'] += 1
        else:
            val, est = 'REVISAR', 'PENDIENTE'
            plan['arbitraje(gemini REJECT)'] += 1
            pend.append(r)
        writes.append((row[0].row, val, est, r.get('vri_ref')))

    print('Plan de escritura:', dict(plan))
    print('Filas a escribir:', len(writes))
    if pend:
        print(f'\nA arbitraje ({len(pend)}):')
        for r in pend:
            print(f"  AN {r['num']:>7} «{r['name'][:20]:20}» PTS nº{r['pts_num']} "
                  f"p{r['pts_page']},{r['pts_line']} | {str(r['reason'])[:66]}")
    if dry:
        print('\n(dry-run, nada escrito)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-an1tika.xlsx'
    shutil.copy(XLSX, bak)
    print('backup →', bak)
    for ridx, val, est, vr in writes:
        ws.cell(row=ridx, column=vcol).value = val
        ws.cell(row=ridx, column=ecol).value = est
        if vr:
            ws.cell(row=ridx, column=vri).value = vr
    wb.save(XLSX)
    print(f'Escritas {len(writes)} filas en {XLSX}.')


if __name__ == '__main__':
    main()
