#!/usr/bin/env python3
"""SN IV Helmer pilot: first 20 entries, rigorous marker matching + DeepSeek."""
import hashlib, json, os, re, sqlite3
from pathlib import Path
import sutta_hash as sh

CACHE_FILE=Path('helmer_ptscst_cache.json'); SN4_BOOK=15

conn=sqlite3.connect('src/data/tipitaka.sqlite'); conn.row_factory=sqlite3.Row; cur=conn.cursor()

# Build marker map for book 15
marker_map={}
cur.execute('SELECT page_no, unitext FROM pages WHERE book_no=? AND edition="mula"',(SN4_BOOK,))
for r in cur.fetchall():
    for i,line in enumerate(r['unitext'].split('\n')):
        s=line.strip()
        m=re.match(r'^(\d+)\.?\s*\((\d+)\)\s+(\S.*)',s)
        if m:
            marker_map.setdefault(r['page_no'],[]).append((i+1,int(m.group(1)),int(m.group(2)),m.group(3)[:60]))

# Load first 20 SN IV entries
from openpyxl import load_workbook
wb=load_workbook('PTS_Reference_Complete_Canon.xlsx'); ws=wb['Complete Canon']
cols={ws.cell(row=1,column=c).value:c for c in range(1,ws.max_column+1)}
entries=[]
for ri in range(2,ws.max_row+1):
    if str(ws.cell(row=ri,column=cols['Nikaya']).value or '')!='SN': continue
    snum=str(ws.cell(row=ri,column=cols['Sutta #']).value or '')
    if not snum or snum=='None': continue
    parts=snum.split('.')
    if len(parts)<2: continue
    if int(parts[0])!=35: continue  # only SN 35 for pilot
    if str(ws.cell(row=ri,column=cols['PTS Roman']).value or '').strip().lower()!='iv': continue
    page=ws.cell(row=ri,column=cols['PTS Page']).value
    ref=str(ws.cell(row=ri,column=cols['PTS Ref']).value or '')
    name=str(ws.cell(row=ri,column=cols['Sutta Name']).value or '')
    entries.append({'num':snum,'ref':ref,'name':name,'page':page,'sn_inner':int(parts[1])})

entries.sort(key=lambda e:e['sn_inner'])
entries=entries[:20]
print('SN IV pilot: {} entries (SN 35.1 - 35.20)'.format(len(entries)))

# For each entry, show: stated page, markers on that page, match result
print('\n--- Marker Analysis ---')
print('{:>8s} {:>15s} {:>4s} {:>4s} {:>12s} {}'.format('Sutta','Ref','Page','vpos','Match','Details'))
print('-'*85)

tasks=[]
for e in entries:
    page=e['page']; vpos=e['sn_inner']
    markers=marker_map.get(page,[])
    marker_str=','.join('{}({})'.format(g,v) for _,g,v,n in markers[:4])
    
    # Find matching marker
    match=None
    for line,gid,mvpos,name in markers:
        if mvpos==vpos or gid==vpos:
            match=(line,gid,name); break
    
    if match:
        line,gid,name=match
        print('{:>8s} {:>15s} {:>4d} {:>4d} {:>12s} L{} gid={} {}'.format(
            e['num'],e['ref'],page,vpos,'✓ MATCH',line,gid,name[:30]))
        
        # Extract text
        cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',(SN4_BOOK,page))
        r=cur.fetchone()
        lines=r['unitext'].split('\n')
        start=line-1; end=len(lines)
        for j in range(start+1,len(lines)):
            if re.match(r'^\d+\.?\s*\(\d+\)',lines[j].strip()): end=j; break
        text=' '.join(lines[start:end])
        tasks.append((e,' '.join(sh.tokens(text)[:300]),line,gid,match))
    else:
        print('{:>8s} {:>15s} {:>4d} {:>4d} {:>12s} markers on page: {}'.format(
            e['num'],e['ref'],page,vpos,'✗ NO MATCH',marker_str))

print('\n--- DeepSeek Validation ({} with markers) ---'.format(len(tasks)))

if tasks:
    segs=sh.load_cha(['s4m.xml'],'h4n')
    cache=json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}
    from openai import OpenAI
    cli=OpenAI(api_key=os.environ['DEEPSEEK_API_KEY'],base_url='https://api.deepseek.com')
    SYS='You are Helmer Smith. Are these two Pali passages the SAME sutta? Reply ONLY JSON: {"verdict":"APPROVE"|"REJECT","reason":"<=15 words"}'
    
    from collections import Counter; vc=Counter()
    for i,(e,pts,line,gid,(_,_,name)) in enumerate(tasks):
        cst_text=' '.join(sh.tokens(segs[i]['text'])[:300])
        cst_title=segs[i].get('title','')
        
        key=hashlib.md5((pts[:200]+cst_text[:200]).encode()).hexdigest()
        if key in cache: hv=cache[key]
        else:
            try:
                r=cli.chat.completions.create(model='deepseek-v4-flash',temperature=0,
                    response_format={'type':'json_object'},
                    messages=[{'role':'system','content':SYS},
                              {'role':'user','content':'PTS (marker {}):\n{}\n\nCST:\n{}'.format(gid,pts[:800],cst_text[:800])}])
                hv=json.loads(r.choices[0].message.content)
                cache[key]=hv; CACHE_FILE.write_text(json.dumps(cache,ensure_ascii=False,indent=2))
            except Exception as ex: hv={'verdict':'ERROR','reason':str(ex)[:60]}
        
        v=hv.get('verdict','?'); vc[v]+=1
        sym='✓' if v=='APPROVE' else '✗' if v=='REJECT' else '?'
        print('{} SN {:>8s} {:>15s} | {:8s} | CST: {} | {}'.format(
            sym,e['num'],e['ref'],v,cst_title[:30],hv.get('reason','')[:50]))
        print('   PTS: {:.80s}...'.format(pts))
        print('   CST: {:.80s}...'.format(cst_text))
    
    print('\n'+'='*85)
    acc=vc.get('APPROVE',0)+vc.get('REJECT',0)
    print('APPROVE: {} | REJECT: {} | Precision: {:.0f}%'.format(
        vc.get('APPROVE',0),vc.get('REJECT',0),100*vc.get('APPROVE',0)/acc if acc>0 else 0))

conn.close()
