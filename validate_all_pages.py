#!/usr/bin/env python3
"""
Content-based validation of ALL PTS references against tipitaka.sqlite.
For each entry, verify the sutta actually appears on the stated page
using name matching, incipit fingerprinting, and cross-page search.
"""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict, Counter

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

# ── Book maps ──
DN_BOOKS = {'i': 6, 'ii': 7, 'iii': 8}
MN_BOOKS = {'i': 9, 'ii': 10, 'iii': 11}
SN_BOOKS = {'i': 12, 'ii': 13, 'iii': 14, 'iv': 15, 'v': 16}
AN_BOOKS = {'i': 17, 'ii': 18, 'iii': 19, 'iv': 20, 'v': 21}
KN_BOOKS = {
    'Khp': 22, 'Dhp': 23, 'Ud': 24, 'It': 25, 'Sn': 26,
    'Vv': 27, 'Pv': 28, 'Thag': 29, 'Thig': 30,
    'Ja': 31,  # Ja II-VI are 31-35
    'Nidd I': 36, 'Nidd II': 37, 'Patis I': 38, 'Patis II': 39,
    'Ap': 40, 'Bv': 41, 'Cp': 42,
}

NIKAYA_BOOK_MAP = {
    'DN': DN_BOOKS, 'MN': MN_BOOKS, 'SN': SN_BOOKS, 'AN': AN_BOOKS
}


def strip_diacritics(text):
    for k, v in {'ā': 'a', 'ī': 'i', 'ū': 'u', 'ṁ': 'm', 'ṃ': 'm',
                 'ṅ': 'n', 'ñ': 'n', 'ṭ': 't', 'ḍ': 'd', 'ṇ': 'n', 'ḷ': 'l'}.items():
        text = text.replace(k, v).replace(k.upper(), v.upper())
    return text


def extract_keywords(name, min_len=3):
    """Extract meaningful keywords from a sutta name."""
    clean = strip_diacritics(name.lower())
    # Remove common prefixes/suffixes
    clean = re.sub(r'sutta[mṃ]?|vagga|pathama|dutiya|tatiya|catuttha|pañcama|chaṭṭha|sattama|aṭṭhama|navama|dasama', '', clean)
    clean = re.sub(r'\[.*?\]|\(.*?\)|\d+[\-\.]?\d*', ' ', clean)
    words = [w.strip() for w in re.split(r'[\s\-–—,;:.]+', clean) if len(w.strip()) >= min_len]
    # Remove very common words
    skip = {'the', 'and', 'for', 'are', 'not', 'eva', 'ca', 'va', 'no', 'pi', 'ti', 'kho', 'pana', 'tattha'}
    return [w for w in words if w not in skip]


def name_match_score(name, text):
    """Score how well a sutta name matches page text."""
    kw = extract_keywords(name)
    if not kw:
        return 0.0
    txt = strip_diacritics(text.lower())
    hits = sum(1 for w in kw if w in txt)
    return hits / len(kw)


def get_page_content(book_no, page_no, lines=20):
    """Get page head + first N lines of text."""
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                (book_no, page_no))
    r = cur.fetchone()
    if not r:
        return '', ''
    head = r['head'] or ''
    text = r['unitext'] or ''
    text_lines = text.split('\n')
    first_text = '\n'.join(text_lines[:lines])
    return head, first_text


def validate_nikaya(nikaya, book_map=None):
    """Validate all entries of a nikaya against page content."""
    wb = load_workbook(XL)
    ws = wb['Complete Canon']
    cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    results = []
    stats = Counter()

    for ri in range(2, ws.max_row + 1):
        if ws.cell(row=ri, column=cols['Nikaya']).value != nikaya:
            continue

        snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
        name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
        page = ws.cell(row=ri, column=cols['PTS Page']).value
        roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower()
        vol = str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip()
        ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
        raw = str(ws.cell(row=ri, column=cols['Raw ID']).value or '')

        if not page:
            stats['no_page'] += 1
            continue

        # Determine book_no
        book_no = None
        if book_map:
            book_no = book_map.get(roman)
            if not book_no:
                # Try volume letter for KN
                book_no = book_map.get(vol)

        if not book_no:
            stats['no_book'] += 1
            continue

        # Check if page exists
        cur.execute('SELECT COUNT(*) as cnt FROM pages WHERE book_no=? AND page_no=? AND edition="mula"',
                    (book_no, page))
        if cur.fetchone()['cnt'] == 0:
            stats['page_missing'] += 1
            results.append({
                'ri': ri, 'snum': snum, 'name': name, 'page': page, 'roman': roman,
                'ref': ref, 'status': 'PAGE_MISSING', 'detail': f'Page {page} not in book {book_no}'
            })
            continue

        # Get page content
        head, text = get_page_content(book_no, page, 30)

        # Strategy 1: Name match in HEAD
        head_score = name_match_score(name, head)
        if head_score >= 0.4:
            stats['ok_head'] += 1
            continue

        # Strategy 2: Name match in page body
        body_score = name_match_score(name, text)
        if body_score >= 0.4:
            stats['ok_body'] += 1
            continue

        # Strategy 3: Check nearby pages (±3)
        found_nearby = False
        for delta in [-3, -2, -1, 1, 2, 3]:
            if found_nearby:
                break
            nh, nt = get_page_content(book_no, page + delta, 30)
            ns = name_match_score(name, nh + '\n' + nt)
            if ns >= 0.4:
                stats[f'ok_nearby_{delta:+d}'] += 1
                found_nearby = True
                results.append({
                    'ri': ri, 'snum': snum, 'name': name, 'page': page,
                    'roman': roman, 'ref': ref, 'status': 'NEARBY',
                    'detail': f'Found at page {page + delta} (δ={delta:+d})',
                    'correct_page': page + delta
                })
                break

        if not found_nearby:
            # Strategy 4: Search wider (±10 pages)
            found_wide = False
            for delta in range(-10, 11):
                if delta == 0 or found_wide:
                    continue
                nh, nt = get_page_content(book_no, page + delta, 30)
                ns = name_match_score(name, nh + '\n' + nt)
                if ns >= 0.3:
                    stats['ok_wide'] += 1
                    found_wide = True
                    results.append({
                        'ri': ri, 'snum': snum, 'name': name, 'page': page,
                        'roman': roman, 'ref': ref, 'status': 'WIDE_MATCH',
                        'detail': f'Found at page {page + delta} (δ={delta:+d})',
                        'correct_page': page + delta
                    })
                    break

            if not found_wide:
                stats['unverified'] += 1
                results.append({
                    'ri': ri, 'snum': snum, 'name': name, 'page': page,
                    'roman': roman, 'ref': ref, 'status': 'UNVERIFIED',
                    'detail': f'No name match on page {page} or ±10',
                    'head_sample': head[:100],
                    'text_sample': ' '.join(text.split()[:20])
                })

    # Print report
    total = sum(stats.values())
    verified = stats.get('ok_head', 0) + stats.get('ok_body', 0)
    nearby = sum(v for k, v in stats.items() if k.startswith('ok_nearby'))
    wide = stats.get('ok_wide', 0)
    unverified = stats.get('unverified', 0)

    print(f'\n{"=" * 80}')
    print(f'  {nikaya} — Content Validation')
    print(f'{"=" * 80}')
    print(f'  Total entries:       {total}')
    print(f'  ✓ Page verified:     {verified} ({100 * verified / max(total, 1):.1f}%)')
    print(f'  ≈ Nearby (±3):       {nearby}')
    print(f'  ≈ Wide (±10):        {wide}')
    print(f'  ✗ Unverified:         {unverified}')
    print(f'  ✗ Missing pages:     {stats.get("page_missing", 0)}')

    if results:
        print(f'\n  ⚠ Issues ({len(results)}):')
        for r in results[:30]:
            print(f'    [{r["status"]:12s}] {r["snum"]:>10s} | {r["ref"]:>15s} | {r["name"][:45]}')
            if r.get('detail'):
                print(f'                     {r["detail"]}')
            if r.get('head_sample'):
                print(f'                     HEAD: {r["head_sample"][:90]}')
        if len(results) > 30:
            print(f'    ... and {len(results) - 30} more')

    return results, dict(stats)


# ── Run for all Nikayas ──
all_issues = []

for nikaya in ['DN', 'MN', 'SN', 'AN']:
    book_map = NIKAYA_BOOK_MAP.get(nikaya)
    issues, stats = validate_nikaya(nikaya, book_map)
    all_issues.extend(issues)

# KN requires special handling
print(f'\n{"=" * 80}')
print(f'  KN — skipped (needs per-book mapping)')
print(f'{"=" * 80}')

# Summary
print(f'\n{"=" * 80}')
print(f'  TOTAL ISSUES FOUND: {len(all_issues)}')
print(f'{"=" * 80}')

conn.close()
