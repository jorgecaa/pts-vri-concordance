#!/usr/bin/env python3
"""
Validate Saṃyutta Nikāya (1,806 entries) against tipitaka.sqlite.
"""

import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict

DB_PATH = '/home/jorge/Code/squashfs-root/src/data/tipitaka.sqlite'
EXCEL_PATH = '/home/jorge/Code/squashfs-root/PTS_Reference_Full_Canon_CORRECTED.xlsx'

SN_BOOK_MAP = {1: 12, 2: 13, 3: 14, 4: 15, 5: 16}
SN_RANGES = {12: (1, 240), 13: (1, 286), 14: (1, 279), 15: (1, 403), 16: (1, 478)}

def strip_diacritics(text):
    for k, v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        text = text.replace(k, v).replace(k.upper(), v.upper())
    return text

def sutta_words(name, min_len=3):
    clean = strip_diacritics(name.lower())
    clean = re.sub(r'sutta(m|nta)?', '', clean)
    clean = re.sub(r'\[.*?\]|\(.*?\)', '', clean)
    clean = re.sub(r'^\d+[\.\-\s]*', '', clean)
    return [w for w in re.split(r'[\s\-–—,;:.]+', clean) if len(w.strip()) >= min_len]

def match_in_text(name, text):
    sw = sutta_words(name)
    if not sw: return 0.0
    txt = strip_diacritics(text.lower())
    return sum(1 for w in sw if w in txt) / len(sw)

def load_excel():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['PTS Reference']
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] != 'SN': continue
        entries.append({
            'row_num': row[0], 'sutta_num': str(row[2] or ''),
            'sutta_name': (row[3] or '').strip(), 'sutta_raw': str(row[11] or '').strip(),
            'pts_vol': row[5], 'pts_roman': row[6], 'pts_page': row[7], 'pts_full': row[8]
        })
    return entries

def roman_to_int(r):
    return {'i':1,'ii':2,'iii':3,'iv':4,'v':5,'vi':6}.get(str(r).strip().lower(), 0)

def validate():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    entries = load_excel()
    print(f"Loaded {len(entries)} SN entries\n")
    
    # TOC
    cur.execute("SELECT book_no, page_no, section, title FROM contents WHERE book_no BETWEEN 12 AND 16")
    toc_index = {(r['book_no'], r['page_no']): (r['section'], r['title']) for r in cur.fetchall()}
    
    # Pages
    cur.execute("SELECT book_no, page_no, head, substr(unitext, 1, 500) as txt FROM pages WHERE book_no BETWEEN 12 AND 16 AND edition='mula'")
    pages_index = {(r['book_no'], r['page_no']): (r['head'] or '', r['txt'] or '') for r in cur.fetchall()}
    
    print(f"DB: {len(toc_index)} TOC, {len(pages_index)} pages\n")
    
    results = []
    stats = defaultdict(int)
    
    for e in entries:
        vol = roman_to_int(e['pts_roman'])
        book_no = SN_BOOK_MAP.get(vol)
        page = e['pts_page']
        
        r = {**e, 'book_no': book_no, 'db_page': page}
        
        if not book_no:
            r['status'] = 'ERROR'; results.append(r); stats['ERROR'] += 1; continue
        
        db_key = (book_no, page)
        if db_key not in pages_index:
            r['status'] = 'MISSING'; r['detail'] = f'p.{page} not in book {book_no}'; results.append(r); stats['MISSING'] += 1; continue
        
        head, text = pages_index[db_key]
        r['head'] = head[:100]; r['first_words'] = ' '.join(text.split()[:12])[:130]
        
        # Strategy 1: TOC
        toc = toc_index.get(db_key)
        if toc:
            section, title = toc
            r['toc_section'] = section; r['toc_title'] = title
            ratio = match_in_text(e['sutta_name'], title)
            if ratio >= 0.3:
                r['status'] = 'OK_TOC'; r['detail'] = title; stats['OK_TOC'] += 1
            else:
                # Check head/content too
                if match_in_text(e['sutta_name'], head + ' ' + text) >= 0.25:
                    r['status'] = 'OK_TOC'; r['detail'] = title; stats['OK_TOC'] += 1
                else:
                    r['status'] = 'TOC_NEAR'; r['detail'] = f'TOC: {title}'; stats['TOC_NEAR'] += 1
            results.append(r); continue
        
        # Strategy 2: HEAD
        if match_in_text(e['sutta_name'], head) >= 0.3:
            r['status'] = 'OK_HEAD'; r['detail'] = head[:70]; stats['OK_HEAD'] += 1
            results.append(r); continue
        
        # Strategy 3: content
        if match_in_text(e['sutta_name'], text) >= 0.25:
            r['status'] = 'OK_CONT'; stats['OK_CONT'] += 1
            results.append(r); continue
        
        # Strategy 4: nearby pages +-2
        found = False
        for delta in [-2, -1, 1, 2]:
            nearby = (book_no, page + delta)
            nh, nt = pages_index.get(nearby, ('', ''))
            if match_in_text(e['sutta_name'], nh + nt) >= 0.3:
                r['status'] = 'OK_NEAR'; r['detail'] = f'Match p.{page+delta}'; stats['OK_NEAR'] += 1
                found = True; break
        if found:
            results.append(r); continue
        
        r['status'] = 'UNVERIF'; r['detail'] = 'No name match'; stats['UNVERIF'] += 1
        results.append(r)
    
    conn.close()
    return results, stats

def print_report(results, stats):
    total = len(results)
    verified = total - stats.get('UNVERIF', 0) - stats.get('ERROR', 0) - stats.get('MISSING', 0)
    
    print(f"{'='*90}")
    print(f"  SAṂYUTTA NIKĀYA — Content Validation")
    print(f"  Database: tipitaka.sqlite (PTS edition, 'mula')")
    print(f"{'='*90}")
    print(f"  Total: {total}")
    print(f"  ✓ OK_TOC:  {stats.get('OK_TOC',0):>5}  |  OK_HEAD: {stats.get('OK_HEAD',0):>5}")
    print(f"  ≈ TOC_NEAR:{stats.get('TOC_NEAR',0):>5}  |  OK_NEAR: {stats.get('OK_NEAR',0):>5}  |  OK_CONT: {stats.get('OK_CONT',0):>5}")
    print(f"  ⚠ UNVERIF: {stats.get('UNVERIF',0):>5}  |  ERROR:   {stats.get('ERROR',0):>5}  |  MISSING: {stats.get('MISSING',0):>5}")
    print(f"  Verified: {verified}/{total} = {100*verified/total:.1f}%")
    
    # By volume
    by_vol = defaultdict(lambda: {'total': 0, 'ok': 0})
    for r in results:
        v = r['pts_roman']
        by_vol[v]['total'] += 1
        if r['status'] not in ('UNVERIF', 'ERROR', 'MISSING'):
            by_vol[v]['ok'] += 1
    
    print(f"\n  By volume:")
    for vol in ['i', 'ii', 'iii', 'iv', 'v']:
        d = by_vol[vol]
        print(f"    SN {vol}: {d['ok']}/{d['total']} verified ({100*d['ok']/max(1,d['total']):.0f}%)")
    
    print(f"{'─'*90}")
    
    issues = [r for r in results if r['status'] in ('UNVERIF', 'ERROR', 'MISSING')]
    if issues:
        print(f"\n  ⚠ Items needing review: {len(issues)}\n")
        # Show first 25 and last 5
        for r in issues[:25]:
            m = {'UNVERIF': '?', 'ERROR': '‼', 'MISSING': '✗'}.get(r['status'], '?')
            print(f"  {m} [{r['status']:7s}] {r['sutta_raw'][:55]:55s} {r['pts_full']:10s}")
            if r.get('toc_title'): print(f"       TOC: {r['toc_title']}")
            if r.get('head'): print(f"       Head: {r['head'][:90]}")
        if len(issues) > 30:
            print(f"\n  ... {len(issues)-25} more (showing last 5):\n")
            for r in issues[-5:]:
                print(f"  {m} [{r['status']:7s}] {r['sutta_raw'][:55]:55s} {r['pts_full']:10s}")
                if r.get('head'): print(f"       Head: {r['head'][:90]}")
    
    # Show volume I specially (no TOC)
    vol1 = [r for r in results if r['pts_roman'] == 'i']
    vol1_unv = [r for r in vol1 if r['status'] == 'UNVERIF']
    if vol1_unv:
        print(f"\n  SN I (no TOC, {len(vol1_unv)}/{len(vol1)} unverified) — spot checks:")
        for r in vol1[::50][:5]:
            m = '✓' if r['status'] not in ('UNVERIF',) else '?'
            head_short = (r.get('head','') or '')[:70]
            txt_short = (r.get('first_words','') or '')[:90]
            print(f"    {m} {r['sutta_raw'][:50]:50s} p.{r['db_page']:>4d}")
            print(f"       Head: {head_short}")
            print(f"       Text: {txt_short}")
    
    print(f"\n{'='*90}")
    print(f"  Combined valid: {verified}/{total} ({100*verified/total:.1f}%)")

if __name__ == '__main__':
    results, stats = validate()
    print_report(results, stats)
