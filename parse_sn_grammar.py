#!/usr/bin/env python3
"""SN — sequential position matching for remaining entries."""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict
import pyparsing as pp

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

# ── Grammar ──
integer = pp.pyparsing_common.integer
dot = pp.Suppress('.')
dash = pp.Suppress('-')
vagga = (integer + dash + integer) ^ integer
vagga_paren = pp.Suppress('(') + vagga('pos') + pp.Suppress(')')
sutta_id = (integer + dash + integer)('id') ^ integer('id')
name = pp.SkipTo(pp.LineEnd())('name')
name.setParseAction(lambda t: t[0].strip())

pat_dot = sutta_id + dot + vagga_paren + name
pat_nodot = sutta_id + vagga_paren + name
pat_section = pp.Suppress(pp.Literal('§')) + integer('id') + pp.Suppress('.') + name
pat_simple = sutta_id + ~pp.Literal('(') + name
all_pat = pat_dot | pat_nodot | pat_section | pat_simple

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

SN_MAP = {1:12, 2:13, 3:14, 4:15, 5:16}
VOL_LETTER = {1:'i',2:'ii',3:'iii',4:'iv',5:'v'}

def sd(t):
    for k,v in {'ā':'a','ī':'i','ū':'u','ṁ':'m','ṃ':'m','ṅ':'n','ñ':'n','ṭ':'t','ḍ':'d','ṇ':'n','ḷ':'l'}.items():
        t=t.replace(k,v).replace(k.upper(),v.upper())
    return t

# Parse all markers
print('Parsing markers...')
all_markers = {}
for vol_num in [1,2,3,4,5]:
    book = SN_MAP[vol_num]
    cur.execute('SELECT page_no, unitext FROM pages WHERE book_no=? AND edition="mula"', (book,))
    for r in cur.fetchall():
        text = r['unitext'] or ''
        markers = []
        try:
            for tokens, start, end in all_pat.scan_string(text):
                line_num = text[:start].count('\n') + 1
                rid = tokens.get('id')
                rname = tokens.get('name', '')
                if isinstance(rid, int): lo = hi = rid
                elif rid and len(rid) > 0: lo = rid[0]; hi = rid[-1]
                else: continue
                markers.append((line_num, lo, hi, rname))
        except: pass
        if markers:
            all_markers[(book, r['page_no'])] = markers

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1,column=c).value: c for c in range(1, ws.max_column+1)}

# For each page, collect Excel entries and PTS markers
# Match by ID first, then by sequential position
print('Matching by position...')
total_fixed = 0
stats = defaultdict(lambda: {'total':0, 'id':0, 'pos':0, 'name':0, 'none':0})

# Group Excel entries by (volume, page)
page_entries = defaultdict(list)
entry_data = {}
for ri in range(2, ws.max_row + 1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'SN': continue
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip()
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    if not page: continue
    vol_num = {'i':1,'ii':2,'iii':3,'iv':4,'v':5}.get(roman, 0)
    book = SN_MAP.get(vol_num)
    sutta_num = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    
    key = (book, page)
    page_entries[key].append({'ri': ri, 'num': sutta_num, 'name': name, 'ref': ref, 'roman': roman, 'matched': False})

# Process each page
for (book, page), entries in page_entries.items():
    markers = all_markers.get((book, page), [])
    
    # First pass: ID matching
    for e in entries:
        parts = e['num'].split('.')
        target_id = int(parts[1]) if len(parts) > 1 else None
        if target_id:
            for mline, lo, hi, mname in markers:
                if lo <= target_id <= hi:
                    e['matched'] = True
                    e['line'] = mline
                    e['method'] = 'id'
                    stats[e['roman']]['id'] += 1
                    break
    
    # Second pass: sequential position for remaining
    unmatched_entries = [e for e in entries if not e['matched']]
    if unmatched_entries and markers:
        # Only use markers that haven't been matched to other entries
        used_markers = set()
        for e in entries:
            if e.get('matched') and 'line' in e:
                used_markers.add(e['line'])
        
        available = [(l, lo, hi, n) for l, lo, hi, n in markers if l not in used_markers]
        
        if len(unmatched_entries) == len(available):
            # Perfect 1:1 — assign sequentially
            for i, e in enumerate(unmatched_entries):
                e['line'] = available[i][0]
                e['matched'] = True
                e['method'] = 'pos'
                stats[e['roman']]['pos'] += 1
        elif len(unmatched_entries) <= len(available):
            # Assign first N markers
            for i, e in enumerate(unmatched_entries):
                e['line'] = available[i][0]
                e['matched'] = True
                e['method'] = 'pos'
                stats[e['roman']]['pos'] += 1
    
    # Third pass: name matching for any still unmatched
    still_unmatched = [e for e in entries if not e['matched']]
    for e in still_unmatched:
        nw = [w for w in re.split(r'[\s\-,;.()]+', sd(e['name'].lower())) if len(w) >= 3]
        skip = {'sutta','suttam','vagga','pathama','dutiya','tatiya','catuttha','pancaka','adisu','tika','duka'}
        nw = [w for w in nw if w not in skip]
        if nw and markers:
            best_l, best_s = None, 0
            for mline, lo, hi, mname in markers:
                hits = sum(1 for w in nw if w in sd(mname.lower()))
                if hits > best_s: best_s = hits; best_l = mline
            if best_s >= max(1, len(nw)*0.4):
                e['line'] = best_l
                e['matched'] = True
                e['method'] = 'name'
                stats[e['roman']]['name'] += 1
    
    # Apply to Excel
    for e in entries:
        stats[e['roman']]['total'] += 1
        if not e.get('matched'):
            stats[e['roman']]['none'] += 1
            continue
        
        new_ref = 'S %s %d' % (e['roman'], page)
        if e['line'] > 1: new_ref += ',%d' % e['line']
        if new_ref != e['ref']:
            ws.cell(row=e['ri'], column=cols['PTS Ref']).value = new_ref
            total_fixed += 1

wb.save(XL)

# Report
print()
for r in ['i','ii','iii','iv','v']:
    s = stats[r]
    f = s['id'] + s['pos'] + s['name']
    print('SN %s: %d/%d (id=%d pos=%d name=%d miss=%d)' % (
        r.upper(), f, s['total'], s['id'], s['pos'], s['name'], s['none']))

total = sum(stats[r]['id']+stats[r]['pos']+stats[r]['name'] for r in stats)
print()
print('TOTAL: %d/1806 (%.1f%%)' % (total, 100*total/1806))
print('Fixed: %d' % total_fixed)
print('Saved: %s' % XL)
conn.close()
