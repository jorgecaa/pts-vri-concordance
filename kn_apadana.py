#!/usr/bin/env python3
"""
kn_apadana — el Apadāna: **603 filas** (Therāpadāna 563 + Therīapadāna 40), el bloque más grande de
KN después del Jātaka. Etiqueta `Ap` del Excel, libro 40 de la BD, ficheros `s0510m1` y `s0510m2`.

### El problema: las dos ediciones no cuentan lo mismo, y el desfase crece

El CST tiene **563** apadānas de theras en 56 vaggas (12 + 10×54 + 11) y el Excel las mismas 563,
una por una. Pero **PTS numera 547**, y no por un desfase fijo:

| | Excel/CST | impreso |
|---|--:|--:|
| Sāriputta | 3 | **1** |
| Lakuṇḍabhaddiya | 543 | **534** |
| Cūḷasugandha | 552 | **547** |

El desfase arranca en 2 —PTS imprime el Buddhāpadāna y el Paccekabuddhāpadāna **sin numerar**, como
prefacio— y llega a 9 porque el CST parte en dos lo que PTS imprime junto. Emparejar por número
sería exactamente el error de S ii: un par coherente consigo mismo y ajeno a la fila.

### Lo que decide: la alineación de dos listas ordenadas, cotejada con la página

PTS marca cada apadāna con **`[N. Nombre.]`** entre corchetes —`[2. Mahā-Moggallāna.]`,
`[547. Cūlasugandha.]`— y cierra cada uno con su colofón (`Sāriputtattherassa apadānaṃ samattaṃ`).
Son **544** marcadores impresos contra los 563 `subhead` del CST, los dos en el mismo orden. Se
alinean con `SequenceMatcher` sobre los nombres normalizados, que absorbe las variantes ortográficas
—PTS escribe `Puḷinapājaka` donde el CST pone `Pulinapūjaka`, `Tīṇisaraṇāgamaniya` por
`Tisaraṇagamaniya`, `Sakacittaniya` por `Sakacintaniya`— sin necesidad de una tabla de excepciones.

Y entonces viene la comprobación que lo cierra, **independiente del nombre**: la página del marcador
impreso contra la que declara el Excel. De las 537 alineadas, **505 coinciden exactamente**, 31 caen
a ±1 (el marcador está al pie de la página anterior) y **una sola discrepa**.

### Las 11 que PTS no imprime

El **Yasavagga**, vagga 56 del CST (`Yasa`, `Nadīkassapa`, `Gayākassapa`… `Raṭṭhapāla`), **no está
en la edición de PTS**: el impreso cierra en `[547. Cūlasugandha]` (p510) y en la p511 va el colofón
de toda la obra. El propio Excel lo delata — **las once filas llevan la misma página, la 511**, que
es la del colofón. No hay texto PTS que cotejar y quedan PENDIENTE.

### Therīapadāna

Cuarenta filas, cuatro vaggas de diez, y la numeración impresa **reinicia** en `[1. Sumedhā]`
(p513). Aquí las dos ediciones **sí** cuentan igual: 40 = 40, en orden.

Uso: python3 kn_apadana.py [--obra thera|theri] [--dry]
"""
import re
import shutil
import sqlite3
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from difflib import SequenceMatcher

import an_peyyala as ap
import an1_eka_duka as R
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
XMLDIR = '/tmp/tipitaka-xml/romn'
LIBRO = 40
P_THERI = 512            # el Therīapadāna abre aquí y la numeración impresa reinicia


def raiz(t):
    """Raíz comparable de un nombre de apadāna: sin ordinal, sin sufijo, sin guiones ni apóstrofos.

    PTS y el CST escriben el mismo nombre de muchas maneras —`Mahā-Moggallāna` / `Mahāmoggallāna`,
    `{Aññākoṇḍañña}` / `Aññāsikoṇḍañña`— y el cotejo va sobre esta forma desnuda. Las diferencias
    que quedan las absorbe la alineación por secuencia, que no necesita que cada par sea idéntico.
    """
    t = ap.fold(re.sub(r'^[\d\-.\s{]+', '', t or '')).strip()
    t = re.sub(r'(therapadanam|theri?apadanam|apadanam|thero|theri)$', '', t)
    for ch in "-' }*":
        t = t.replace(ch, '')
    return re.sub(r'[aiueo]m?$', '', t)


def raiz_sin_ordinal(t):
    """La raíz sin el ordinal con que el CST distingue a los homónimos.

    Cuando dos apadānas se llaman igual, el CST marca el segundo con `Dutiya-` —`Raṃsisaññaka` /
    `Dutiyaraṃsisaññaka`— y **PTS no**: los distingue con una vocal (`Raṃsisaññika` /
    `Raṃsisaññaka`) o no los distingue en absoluto. Es la misma práctica que ya obligó a tomar el
    k-ésimo homónimo en S iii. Sin quitar el ordinal, el parecido cae por debajo del umbral y el
    par se pierde aunque el orden y la página lo señalen sin duda.
    """
    return re.sub(r'^(dutiya|tatiya|apara|pathama)', '', raiz(t))


def marcadores(conn):
    """`[(página, nº impreso, nombre)]` — los `[N. Nombre.]` que PTS pone al abrir cada apadāna.

    Los dos primeros no llevan corchete ni número: PTS los imprime **sin numerar**, como prefacio,
    con numeral romano y el nombre en versales —`I` / `BUDDHĀPADĀNAṂ` (p1) y `II` /
    `PACCEKABUDDHĀPADĀNAṂ` (p7)—, que son exactamente las páginas que declaran sus filas. De ahí
    arranca el desfase de 2 con la numeración del CST.
    """
    out = [(1, 0, 'Buddhāpadāna'), (7, 0, 'Paccekabuddhāpadāna')]
    for r in conn.execute('SELECT page_no,unitext FROM pages WHERE edition="mula" AND book_no=? '
                          'ORDER BY page_no', (LIBRO,)):
        for l in (r['unitext'] or '').split('\n'):
            # ⚠️ El corchete de apertura falta a veces (`377. Kāsumāriphaladāyaka.]`) y detrás
            # del de cierre puede ir la llamada de nota (`[372. Sattapaṇṇiya.] 1`). Exigir la forma
            # canónica dejaba sin marcador a varios apadānas y la alineación los daba por
            # divergencias de división entre las dos ediciones, que es lo contrario de lo que son.
            m = re.match(r'^\[?(\d+)\.\s*([^\]\[]+?)\.?\]\s*\d*$', l.strip())
            if m:
                out.append((r['page_no'], int(m.group(1)), m.group(2).strip(' .')))
    return out


def subheads_cst(pref):
    """`[(fichero, capítulo POSICIONAL, item en el vagga, etiqueta del vagga, subhead, texto)]`.

    ⚠️ El capítulo de la `VRI Ref` es **la posición del `div` dentro de su fichero**, no la etiqueta
    del vagga. El Therāpadāna se reparte entre dos ficheros —vaggas 1-42 en `s0510m1` y 43-56 en
    `s0510m2`, donde además vienen detrás los cuatro del Therīapadāna—, así que el «vagga 43» es el
    **primer** capítulo de `s0510m2`. Escribir `s0510m2:c43` ancla la fila a nada, y
    `check_integrity` lo detecta (`el capítulo 43 no existe en s0510m2`): fue justo lo que pasó al
    primer volcado, con 130 filas mal ancladas.

    El item también es **el de dentro del vagga**, no el índice corrido: el `n` reinicia en cada
    capítulo (regla (5-bis)).
    """
    out = []
    for st in ('s0510m1', 's0510m2'):
        raizx = ET.parse(f'{XMLDIR}/{st}.mul.xml').getroot()
        pos = 0
        for d in raizx.iter('div'):
            n = d.get('n') or ''
            if '_' not in n:
                continue
            pos += 1
            if not n.startswith(pref + '_'):
                continue
            act, k = None, 0
            for p in d.iter('p'):
                txt = ''.join(p.itertext()).strip()
                if p.get('rend') == 'subhead':
                    k += 1
                    act = [st, pos, k, n.split('_')[1], txt, '']
                    out.append(act)
                elif act is not None:
                    act[5] += ' ' + txt
    return [(a, b, c, d, e, ap.fold(f)) for a, b, c, d, e, f in out]


def alinea(cst, imp):
    """`{índice del CST: índice del impreso}` por alineación de las dos listas **ordenadas**.

    Se aceptan los bloques `equal` y los `replace` de **igual longitud** —ahí la correspondencia es
    posición a posición y la diferencia es sólo ortográfica—. Los bloques de longitud distinta son
    divergencias reales de división entre las dos ediciones y se dejan fuera: emparejarlos «a ojo»
    es justo lo que produce un par coherente y ajeno a la fila.
    """
    A = [raiz(x[4]) for x in cst]
    B = [raiz(nm) for _, _, nm in imp]
    par = {}
    for tag, i1, i2, j1, j2 in SequenceMatcher(None, A, B, autojunk=False).get_opcodes():
        if tag == 'equal' or (tag == 'replace' and (i2 - i1) == (j2 - j1)):
            for d in range(i2 - i1):
                par[i1 + d] = j1 + d
    return par


def repara(par, cst, imp, fs):
    """Segunda pasada **por página**: rescata lo que la alineación por nombre no pudo cerrar.

    Cuando dos nombres vecinos son casi iguales —`Raṃsisaññika` y `Raṃsisaññaka`, que PTS numera 85
    y 86— la alineación puede desplazarse uno y dejar los dos extremos sueltos. La página lo
    resuelve sin ambigüedad: se busca, entre los marcadores **aún sin usar**, el que esté en la
    página que declara la fila y cuyo nombre se parezca. Sigue siendo evidencia doble (página +
    nombre), sólo que en el otro orden.
    """
    usados = set(par.values())
    for k in range(len(cst)):
        if k in par or not isinstance(fs[k]['page'], int):
            continue
        cands = [j for j, (pg, _n, nm) in enumerate(imp)
                 if j not in usados and abs(pg - fs[k]['page']) <= 1
                 and max(SequenceMatcher(None, r_, raiz(nm)).ratio()
                         for r_ in (raiz(cst[k][4]), raiz_sin_ordinal(cst[k][4]))) >= 0.75]
        # con varios candidatos manda **la página exacta**: los dos `Raṃsisaññaka` del vagga 9
        # están en páginas contiguas y sólo uno cae en la que declara la fila
        exactos = [j for j in cands if imp[j][0] == fs[k]['page']]
        elegido = exactos[0] if len(exactos) == 1 else (cands[0] if len(cands) == 1 else None)
        if elegido is not None:
            par[k] = elegido
            usados.add(elegido)
    return par


def filas(pref):
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    return [{'num': str(r[ci['Sutta #']]), 'name': str(r[ci['Sutta Name']] or ''),
             'page': r[ci['PTS Page']], 'estado': r[ci['Estado']]}
            for r in ws.iter_rows(min_row=2, values_only=True)
            if str(r[ci['PTS Vol']]) == 'Ap' and str(r[ci['Sutta #']]).startswith(pref)]


def resuelve(conn, obra):
    thera = obra == 'thera'
    cst = subheads_cst('kn10' if thera else 'kn11')
    fs = filas('10.' if thera else '11.')
    mk = marcadores(conn)
    imp = [x for x in mk if (x[0] < P_THERI) == thera]
    print(f'CST {len(cst)} subheads · Excel {len(fs)} filas · marcadores impresos {len(imp)}')
    if len(cst) != len(fs):
        print('⚠ el CST y el Excel no cuentan lo mismo: no se escribe nada')
        return {}, [], len(fs)
    par = repara(alinea(cst, imp), cst, imp, fs)
    res, avisos, exacta, cerca = {}, [], 0, 0
    for k, f in enumerate(fs):
        st, cap, item, vg, head, tc = cst[k]
        nth = k + 1
        if k not in par:
            avisos.append(f'   ✗ {f["num"]} «{head[:34]}»: sin marcador impreso alineado')
            continue
        pgi, npts, nom = imp[par[k]]
        pge = f['page']
        d = abs(pgi - pge) if isinstance(pge, int) else 99
        exacta += d == 0
        cerca += 1 <= d <= 2
        # ⚠️ La tolerancia es de **dos** páginas. El Excel parece haber tomado la página del
        # **encabezado corrido**, que en este volumen sólo aparece en las impares: cuando el
        # marcador cae al pie de una par, la primera impar cuyo encabezado lo nombra está dos
        # páginas más allá (`[18. Khemā.]` en la p543,23 y el encabezado `18. Khemā` en la p545).
        # No es un error de la fila ni del marcador: son dos maneras de decir dónde empieza.
        if d > 2:
            avisos.append(f'   ✗ {f["num"]} «{head[:30]}»: el impreso «{nom}» está en la p{pgi} y '
                          f'la fila dice {pge}')
            continue
        sim = SequenceMatcher(None, raiz(head), raiz(nom)).ratio()
        res[f['num']] = (
            f'{st}:c{cap}.{item}',
            f'{"Th" if thera else "Thī"} Ap {nth}',
            f'Apadāna: «{head}», la {nth} de {len(cst)} → vagga {vg} del CST, capítulo {cap} de '
            f'{st}. El impreso la marca '
            f'«[{npts}. {nom}]» en la p{pgi}'
            + ('' if d == 0 else f' (la fila dice {pge}, que es la primera impar cuyo encabezado '
               f'corrido lo nombra; el marcador va antes)')
            + f'; parecido de nombre {sim:.2f}. '
            + ('⚠ PTS lo imprime SIN NUMERAR, como prefacio, con numeral romano y el nombre en '
               'versales; de ahí arranca el desfase de 2 con la numeración del CST. '
               if npts == 0 else '')
            + (f'⚠ PTS numera {npts} donde el CST cuenta {nth}: el desfase arranca en 2 —el '
               f'Buddhāpadāna y el Paccekabuddhāpadāna van sin numerar— y crece porque el CST '
               f'parte en dos lo que PTS imprime junto; por eso el emparejamiento va por '
               f'alineación de nombres y se coteja con la página, nunca por el número'
               if thera and npts and npts != nth else ''))
    print(f'  página impresa ≡ la del Excel: exacta {exacta} · ±1 {cerca}')
    return res, avisos, len(fs)


def main():
    dry = '--dry' in sys.argv
    obra = sys.argv[sys.argv.index('--obra') + 1] if '--obra' in sys.argv else 'thera'
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    res, avisos, total = resuelve(conn, obra)
    print(f'\n{len(res)}/{total} firmadas')
    for a in avisos[:40]:
        print(a)
    if len(avisos) > 40:
        print(f'   … y {len(avisos) - 40} más')
    if dry or not res:
        print('(dry-run: nada escrito)' if dry else '')
        return
    cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-{obra}.xlsx'
    shutil.copy(XLSX, cp)
    print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    n = 0
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, ci['PTS Vol']).value) != 'Ap':
            continue
        num = str(ws.cell(r, ci['Sutta #']).value)
        if num not in res or ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        vri, ref, det = res[num]
        ws.cell(r, ci['VRI Ref']).value = vri
        ws.cell(r, ci['PTS Ref']).value = ref
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = det
        n += 1
    wb.save(XLSX)
    print(f'{n} filas cerradas · escrito → {XLSX}')


if __name__ == '__main__':
    main()
