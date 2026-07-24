#!/usr/bin/env python3
"""Add line numbers to DN — fixed: skip roman-numeral titles, section numbers."""
import sqlite3, re
from openpyxl import load_workbook

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

DN_MAP = {1:6, 2:7, 3:8}
ROMAN = {'i':1,'ii':2,'iii':3}

def is_skip(s):
    s = s.strip()
    if not s: return True
    low = s.lower()
    if low.startswith('namo'): return True
    if s.startswith('['): return True
    if re.match(r'^\-{3,}$', s): return True
    if re.match(r'^\d+\.?\d*\.?\s*$', s): return True
    # Roman numeral title: "i. Brahmajala Sutta.]" or "xiv. Mahapadana-Suttanta.]"
    if re.match(r'^[ivxlc]+\.\s+.+\.\]', s, re.IGNORECASE): return True
    # Section markers like "1.1." alone
    if re.match(r'^\d+\.\d+\.\s*$', s): return True
    # Footnote references like "1" alone
    if re.match(r'^\d+$', s) and len(s) <= 2: return True
    return False

def find_text(book, page):
    cur.execute('SELECT unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book, page))
    r = cur.fetchone()
    if not r: return None, ''
    for i, line in enumerate((r['unitext'] or '').split('\n')):
        if not is_skip(line):
            return i+1, re.sub(r'\s+',' ',line.strip())[:120]
    return None, ''

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {}
for c in range(1, ws.max_column+1):
    h = ws.cell(row=1, column=c).value
    if h: cols[h] = c

print('DN — Line numbers (corrected)')
print('='*70)
updates = 0

for ri in range(2, ws.max_row+1):
    if ws.cell(row=ri, column=cols['Nikaya']).value != 'DN': continue
    
    sn = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    page = ws.cell(row=ri, column=cols['PTS Page']).value
    roman = str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip()
    ref = str(ws.cell(row=ri, column=cols['PTS Ref']).value or '')
    name = str(ws.cell(row=ri, column=cols['Sutta Name']).value or '')
    
    book = DN_MAP.get(ROMAN.get(roman.lower(),1))
    if not book or not page: continue
    
    line, text = find_text(book, page)
    if not line: continue
    
    new_ref = f'D {roman} {page}'
    if line > 1: new_ref += f',{line}'
    
    if new_ref != ref:
        ws.cell(row=ri, column=cols['PTS Ref']).value = new_ref
        updates += 1
    
    print(f'  DN {sn:>2s}  {new_ref:16s}  {name[:35]:35s}  {text[:80]}')

wb.save(XL)
print(f'\n  Updated: {updates}')
print(f'  Saved: {XL}')
conn.close()
