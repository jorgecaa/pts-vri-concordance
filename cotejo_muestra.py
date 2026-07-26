#!/usr/bin/env python3
"""
cotejo_muestra — saca a la vista, **lado a lado**, el texto PTS y el texto CST de una fila concreta
del Excel, para que un humano arbitre.

No decide nada: no escribe en el Excel, no puntúa, no aprueba. Resuelve la `VRI Ref` contra el XML,
resuelve la página contra la BD, y lo pone en dos columnas con la cobertura al pie. Es la
herramienta del paso que el proyecto llama **revisión humana**, y sirve sobre todo para lo que el
validador **no** puede ver: que el par sea el equivocado.

Por eso el control importa tanto como el par: se imprime también la cobertura del texto PTS de la
fila contra el sutta **anterior y siguiente** del CST. Si el propio no gana con holgura, la fila es
sospechosa aunque su cobertura sea alta — es exactamente el caso que destapó el desplazamiento +1
del Petavatthu.

Uso:
    python3 cotejo_muestra.py --fila «<Sutta Name, Sutta # o PTS Ref>» [--ancho 92]
"""
import re
import sqlite3
import sys
import xml.etree.ElementTree as ET

from difflib import SequenceMatcher

import an_peyyala as ap
import an1_eka_duka as R
from kn_apadana import junta_corchetes
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
XMLDIR = '/tmp/tipitaka-xml/romn'
# etiqueta `PTS Vol` del Excel → libro de la BD
# ⚠️ En `Bv` (Buddhavaṃsa + Cariyāpiṭaka) la página del Excel es de **otra composición PTS** y
# apunta a un capítulo distinto en la BD: pedirle a la p68 el `Dhātubhājanīya` devuelve el
# `Siddhatthabuddhavaṃsa`. Para esas filas la ventana no sale de la página sino de **la posición del
# capítulo en la BD**, que el `Detail` ya trae escrita («En la BD el texto está en el libro 41,
# pp. 102-102»). La tabla no debe obligar a echar cuentas.
POR_CAPITULO = {'Bv', 'Cp'}
# siglas del CPD (ver `normaliza_siglas_cpd.py`)
LIBRO = {'Khp': 22, 'Dhp': 23, 'Ud': 24, 'It': 25, 'Sn': 26, 'Vv': 27, 'Pv': 28, 'Th': 29,
         'Thī': 29, 'Nidd I': 36, 'Nidd II': 37, 'Paṭis': 38, 'Th-ap': 40, 'Thī-ap': 40,
         'Bv': 41, 'Cp': 42, 'Ja': 30}


def filas():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb['Complete Canon']
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr)}
    return [{k: r[ci[k]] for k in hdr} for r in ws.iter_rows(min_row=2, values_only=True)]


def texto_cst(vri):
    """Texto del CST al que apunta una `VRI Ref`, en cualquiera de sus tres formas."""
    m = re.match(r'^([a-z0-9]+):c(\d+)(?:\.(\d+))?$', str(vri))
    if m:                                   # KN: `<fichero>:c<capítulo>[.<item>]`
        stem, cap, item = m.group(1), int(m.group(2)), m.group(3)
        raiz = ET.parse(f'{XMLDIR}/{stem}.mul.xml').getroot()
        divs = [d for d in raiz.iter('div') if '_' in (d.get('n') or '')]
        if cap > len(divs):
            return None, f'el capítulo {cap} no existe en {stem}'
        d = divs[cap - 1]
        if item is None:
            return ' '.join(''.join(p.itertext()) for p in d.iter('p')), f'{stem} div {cap}'
        act, k, out = None, 0, []
        for p in d.iter('p'):
            if p.get('rend') in ('subhead', 'title'):
                k += 1
                act = (k == int(item))
                if act:
                    out.append(''.join(p.itertext()).strip())
                continue
            if act:
                out.append(''.join(p.itertext()))
        return ' '.join(out), f'{stem} div {cap}, unidad {item}'
    m = re.match(r'^([a-z0-9]+):(\d+)(?:-(\d+))?$', str(vri))
    if m:                                   # SN/AN/KN corrido: `<fichero>:<a>[-<b>]`
        stem, a = m.group(1), int(m.group(2))
        b = int(m.group(3) or a)
        u = [x for x in ap.cst_unidades(stem) if a <= x['pn'] <= b]
        return ' '.join(x['texto'] for x in u), f'{stem} paranums {a}-{b}'
    return None, f'VRI Ref no reconocida: {vri!r}'


def texto_pts(conn, libro, a, b):
    return ' '.join((x['unitext'] or '') for x in conn.execute(
        'SELECT unitext FROM pages WHERE edition="mula" AND book_no=? AND page_no BETWEEN ? AND ?',
        (libro, a, b)))


def columnas(izq, der, ancho):
    """Dos textos en dos columnas, envueltos, con separador."""
    import textwrap
    w = (ancho - 3) // 2
    a = textwrap.wrap(re.sub(r'\s+', ' ', izq).strip(), w) or ['—']
    b = textwrap.wrap(re.sub(r'\s+', ' ', der).strip(), w) or ['—']
    n = max(len(a), len(b))
    a += [''] * (n - len(a))
    b += [''] * (n - len(b))
    return '\n'.join(f'{x:<{w}} │ {y}' for x, y in zip(a, b))


def ventana_del_detail(f):
    """`(libro, primera, última)` leídos del propio `Detail`, o `None`.

    Es el caso de `Bv`/`Cp`: la fila ya sabe dónde está su texto en la BD porque el alineador lo
    dejó escrito. Se lee de ahí en vez de recalcularlo — y si algún día el `Detail` y la BD dejaran
    de concordar, se vería aquí.
    """
    m = re.search(r'En la BD el texto está en el libro (\d+), pp\. (\d+)-(\d+)', str(f['Detail'] or ''))
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None


# ── Gramáticas de marcador, UNA POR OBRA (regla (6)) ────────────────────────────────────────
#
# ⚠️ El recorte genérico —«busca el nombre y corta»— falló tres veces seguidas, y cada fallo lo
# destapó el anterior: cortaba a mitad de verso, luego cortaba por el **colofón** (que en el
# Vimānavatthu lleva el mismo nombre que el encabezado), luego el número no casaba porque `ap.fold`
# borra los dígitos. La lección es la de siempre en este proyecto: **no hay marcador universal**.
# Cada obra tiene el suyo y hay que preguntárselo a ella.
#
# Estas expresiones son **las mismas** que usan los alineadores; `verifica_gramaticas()` comprueba
# que siguen dando el mismo número de marcadores que el módulo de origen, para que no se separen en
# silencio.
MARCADORES = {
    # `[329. Pupphacchattiya.]` — de kn_apadana.marcadores
    'Th-ap': (r'^\[?\s*\d+\.\s*([^\]\[]+?)\.?\]\s*\d*$', 'kn_apadana'),
    # `12 Dutiyapatibbatāvimānavatthu` — abre numerado; el colofón (`Dutiyapatibbatāvimānaṃ`) no
    'Vv': (r'^\s*\d+\s+([A-Za-zĀāĪīŪūṂṃṆṇṬṭḌḍÑñṄṅḶḷ\'\- ]+vimānavatthu)\s*\d*$', 'kn_vimanavatthu'),
    # `5 Soṇapaṇḍitacariyaṃ` — de kn_buddhavamsa.subtitulos_cp
    'Cp': (r'^\s*\d+\s+([A-Za-zĀāĪīŪūṂṃṆṇṬṭḌḍÑñṄṅḶḷ\'\- ]+cariya[ṃm])\s*\d*$', 'kn_buddhavamsa'),
}


def recorta(tp, propio, siguiente, vol):
    """Recorta el texto PTS al tramo de la fila usando **la gramática de marcador de su obra**.

    Devuelve `(texto, nota)`. La página es una resolución demasiado gruesa —en el Apadāna, el
    Vimānavatthu o el Cariyāpiṭaka dos unidades cortas comparten página y el control marca al vecino
    como ganador aunque el par sea bueno—, así que se corta por el **encabezado impreso**: desde el
    de la fila hasta el de la siguiente.

    Si la obra no tiene gramática registrada, o su encabezado no se lee en la ventana, **no se
    corta** y se dice. Más vale la ventana entera que un corte inventado: un recorte mal puesto no
    baja la cobertura, la **falsea**.
    """
    pat, _origen = MARCADORES.get(vol, (None, None))
    if not pat:
        return tp, f'no ({vol} no tiene gramática de marcador registrada)'
    # ⚠️ Se parte por `[\r\n]`, no por `\n`. El texto de la BD usa **`\r` suelto** como separador
    # —15.529 veces, aproximadamente uno por página— y ahí donde el impreso pone dos rótulos
    # seguidos, partir sólo por `\n` los fusiona en una línea: `Paṭhamapatibbatāvimānaṃ4\r    12
    # Dutiyapatibbatāvimānavatthu`. El encabezado deja de leerse y el recorte no salta. Comprobado
    # que los alineadores no pierden marcadores por esto (mismos recuentos con `\n` y con
    # `[\r\n]`), pero aquí sí importaba.
    lineas = junta_corchetes(re.split(r'[\r\n]+', tp))
    cabs = [(j, m.group(1)) for j, l in enumerate(lineas)
            if (m := re.match(pat, l.strip()))]
    if not cabs:
        return tp, 'no (ningún encabezado en la ventana)'

    def elige(nombre, desde):
        r = raiz_nombre(nombre)
        if len(r) < 5:
            return None
        cand = [(SequenceMatcher(None, r, raiz_nombre(nm)).ratio(), j)
                for j, nm in cabs if j >= desde]
        mejor = max(cand, default=(0, None))
        return mejor[1] if mejor[0] >= 0.7 else None

    a = elige(propio, 0)
    if a is None:
        return tp, 'no (el encabezado de esta unidad no se lee en la ventana)'
    b = elige(siguiente, a + 1) if siguiente else None
    return '\n'.join(lineas[a:b]), f'sí, por el encabezado impreso de la línea {a + 1}'


def raiz_nombre(t):
    """Raíz comparable de un nombre, venga del Excel o del impreso."""
    t = re.sub(r'^\s*\(KN[^)]*\)\s*', '', str(t or '').replace('\u00ad', ''))
    t = re.sub(r'^\s*(Th[īi]? Ap|Tha Ap|Cp|Bv|Vv|Pv)\s+[\d.]+\s*', '', t)
    t = ap.fold(re.sub(r'^[\d.\s\[]+', '', t))
    t = re.sub(r'(vimanavathu|petavathu|cariyam|cariya|apadanam|thera|theri)\b.*$', '', t)
    return re.sub(r'[^a-z]', '', t)


def verifica_gramaticas(conn):
    """¿Siguen las gramáticas de aquí dando lo mismo que el módulo del que salieron?

    Si un alineador afina su expresión y ésta se queda atrás, el cotejo empieza a no recortar sin
    que nadie se entere. Esto lo canta.
    """
    import kn_apadana
    n_mod = len([x for x in kn_apadana.marcadores(conn) if x[1]])
    pat = MARCADORES['Th-ap'][0]
    n_aqui = sum(1 for r in conn.execute(
        'SELECT unitext FROM pages WHERE edition="mula" AND book_no=40')
        for l in junta_corchetes((r['unitext'] or '').split('\n')) if re.match(pat, l.strip()))
    return [f'Th-ap: el módulo lee {n_mod} y esta gramática {n_aqui}'] if n_mod != n_aqui \
        else []


def coteja(conn, f, fs, ancho=92, lineas=8):
    vri = f['VRI Ref']
    tc, origen = texto_cst(vri)
    vol = str(f['PTS Vol'])
    libro = LIBRO.get(vol)
    pg = f['PTS Page']
    hermanas = [g for g in fs if str(g['PTS Vol']) == vol and isinstance(g['PTS Page'], int)]
    k = hermanas.index(f) if f in hermanas else -1
    sig = hermanas[k + 1]['PTS Page'] if 0 <= k < len(hermanas) - 1 else pg + 3
    cap = ventana_del_detail(f) if vol in POR_CAPITULO else None
    if cap:
        libro, a, b = cap
        nota = f'por CAPÍTULO (la p{pg} del Excel es de otra composición y no sirve)'
    else:
        a, b = max(1, pg - 1), max(pg, sig)
        nota = 'por página'
    tp = texto_pts(conn, libro, a, b) if libro and pg else ''
    sig_nom = hermanas[k + 1]['Sutta Name'] if 0 <= k < len(hermanas) - 1 else None
    tp, nota_rec = recorta(tp, f['Sutta Name'], sig_nom, vol)
    print('═' * ancho)
    print(f"  {f['Sutta Name']}")
    print(f"  PTS Ref «{f['PTS Ref']}»   ·   VRI Ref «{vri}» → {origen}")
    print(f"  {vol} p{pg} · BD libro {libro}, pp. {a}-{b} — resuelto {nota}"
          f"   ·   {f['Validation']} / {f['Estado']}")
    print(f'  recorte: {nota_rec}')
    print('─' * ancho)
    print(f"{'PTS (impreso, BD)':<{(ancho - 3) // 2}} │ CST (VRI)")
    print('─' * ancho)
    if tc is None:
        print(f'  ⚠ {origen}')
        return
    print('\n'.join(columnas(tp, tc, ancho).split('\n')[:lineas]))
    print('─' * ancho)
    cov = R.cobertura(ap.fold(tp), ap.fold(tc))
    ctl = []
    for d in (-1, 1):
        if 0 <= k + d < len(hermanas) and hermanas[k + d]['VRI Ref']:
            t2, _ = texto_cst(hermanas[k + d]['VRI Ref'])
            if t2:
                ctl.append(f'{"anterior" if d < 0 else "siguiente"} {R.cobertura(ap.fold(tp), ap.fold(t2)):.2f}')
    marca = '✓' if not ctl or cov >= max(float(x.split()[-1]) for x in ctl) else '⚠ EL VECINO GANA'
    print(f'  cobertura propia {cov:.2f}   ·   control: {" · ".join(ctl) or "—"}   {marca}')


def main():
    ancho = int(sys.argv[sys.argv.index('--ancho') + 1]) if '--ancho' in sys.argv else 92
    claves = sys.argv[sys.argv.index('--fila') + 1:] if '--fila' in sys.argv else []
    fs = filas()
    conn = sqlite3.connect(R.DB)
    conn.row_factory = sqlite3.Row
    for clave in claves:
        hit = [f for f in fs if clave in str(f['Sutta Name']) or clave == str(f['Sutta #'])
               or clave == str(f['PTS Ref'])]
        if not hit:
            print(f'⚠ sin fila para {clave!r}')
            continue
        coteja(conn, hit[0], fs, ancho)


if __name__ == '__main__':
    main()
