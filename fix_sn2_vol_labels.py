#!/usr/bin/env python3
"""fix_sn2_vol_labels.py — repara 7 filas de SN 22 etiquetadas «S ii» en vez de «S iii».

SN 22 (Khandhasaṃyutta) abre el volumen III — Feer lo dice explícitamente en la introducción de
S ii: «The first Samyutta of the third section (Khandha) will accordingly be numbered XXII».
Aun así, `PTS_Reference_Complete_Canon.xlsx` trae **SN 22.153–159** con `PTS Roman='ii'`.

Evidencia de que el volumen es lo único erróneo (la PÁGINA ya era correcta):

- `massive.tsv` da `cst_p_page = 3.0183 … 3.0187` → **vol 3**, páginas 183–187, las mismas que
  el Excel; y los títulos CST casan uno a uno con el `Sutta Name` de las filas.
- En la BD, S ii 183 (libro 13) habla de *kappā* (eones → Anamataggasaṃyutta), mientras S iii 183
  (libro 14) trata de rūpaṃ/attā/loko, que es Khandha.
- El Excel ya tiene SN 22.1–152 en «S iii»; 152 + 7 = **159**, el total canónico de SN 22.

Corrige `PTS Roman` (`ii`→`iii`) y `PTS Ref` (`S ii N`→`S iii N`); no toca `PTS Page`.

Uso: python3 fix_sn2_vol_labels.py [--write]
"""
import datetime, re, shutil, sys

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
TARGET = [f'22.{k}' for k in range(153, 160)]


def main():
    write = '--write' in sys.argv
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}

    fixes = []
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'SN':
            continue
        if str(v[ci['Sutta #']]) not in TARGET:
            continue
        if str(v[ci['PTS Roman']] or '').strip().lower() != 'ii':
            continue
        old_ref = str(v[ci['PTS Ref']] or '')
        new_ref = re.sub(r'^S\s+ii\b', 'S iii', old_ref)
        fixes.append((ridx, str(v[ci['Sutta #']]), old_ref, new_ref))

    print(f'Filas a corregir: {len(fixes)}')
    for _r, num, o, n in fixes:
        print(f'  SN {num:>7}: PTS Roman ii→iii | {o!r} → {n!r}')
    if not write:
        print('\n(dry-run, nada escrito — usa --write)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-sn2vol.xlsx'
    shutil.copy(XLSX, bak)
    print('backup →', bak)
    for ridx, _num, _o, new_ref in fixes:
        ws.cell(row=ridx, column=ci['PTS Roman'] + 1).value = 'iii'
        ws.cell(row=ridx, column=ci['PTS Ref'] + 1).value = new_ref
    wb.save(XLSX)
    print(f'Corregidas {len(fixes)} filas en {XLSX}.')


if __name__ == '__main__':
    main()
