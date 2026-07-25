#!/usr/bin/env python3
"""reconcile_an3.py — vuelca `validador_an3.json` al Excel maestro (AN III: Pañcaka y Chakka).

Regla del proyecto: CONFIRMADO = concordancia ∧ Gemini APPROVE → `Validation=VALIDADOR`.
Gemini REJECT → **no se auto-firma**: queda como `REVISAR`/PENDIENTE para arbitraje humano.

Escribe además la **`VRI Ref`**, la clave canónica del lado CST (regla (5)). Ojo: aquí el fichero
VRI **depende de la sección** —`s0403m1` para el Pañcaka y `s0403m2` para el Chakka—, así que la
referencia la trae cada registro del JSON, no se compone aquí.

Los valores legados (`OK+RTE`, `RTE_ONLY`, `OK`, `UNVERIFIED`) vienen del pase RTE (BUDSIR),
**fuera de alcance**: no se heredan, se sobrescriben con el veredicto del validador.

⚠️ Quedan **fuera** las 33 filas de rango y las **45 sin alinear** —la cola del peyyāla del
Pañcaka (`5.258`-`5.302`), donde el CST no nombra individualmente los suttas—. Ninguna se toca.

Uso:
    python3 reconcile_an3.py --dry
    python3 reconcile_an3.py --write [--humano firmas.json]
"""
import datetime
import json
import shutil
import sys
from collections import Counter

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
RES = 'validador_an3.json'


def main():
    dry = '--write' not in sys.argv
    humano = {}
    if '--humano' in sys.argv:
        humano = json.load(open(sys.argv[sys.argv.index('--humano') + 1]))
    res = {r['num']: r for r in json.load(open(RES))}

    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}
    vcol, ecol, vri = ci['Validation'] + 1, ci['Estado'] + 1, ci['VRI Ref'] + 1

    plan, writes, pend = Counter(), [], []
    for row in ws.iter_rows(min_row=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'AN' or str(v[ci['PTS Roman']] or '').strip().lower() != 'iii':
            continue
        num = str(v[ci['Sutta #']])
        if not (num.startswith('5.') or num.startswith('6.')):
            continue
        if '-' in num:
            plan['fila de rango (a arbitraje, no se toca)'] += 1
            continue
        r = res.get(num)
        if not r:
            plan['sin resultado'] += 1
            continue
        if num in humano:
            val = humano[num]
            est = 'PENDIENTE' if val in ('REVISAR', 'REF_ERROR_DPR') else 'CONFIRMADO'
            plan[f'humano:{val}'] += 1
        elif r['gemini'] == 'APPROVE':
            val, est = 'VALIDADOR', 'CONFIRMADO'
            plan['VALIDADOR(auto)'] += 1
        else:
            val, est = 'REVISAR', 'PENDIENTE'
            plan['arbitraje(gemini REJECT)'] += 1
            pend.append(r)
        writes.append((row[0].row, val, est, r.get('vri_ref')))

    print('Plan de escritura:', dict(plan))
    print('Filas a escribir:', len(writes))
    if pend:
        print(f'\nA arbitraje ({len(pend)}):')
        for r in pend:
            print(f"  AN {r['num']:>7} «{r['name'][:20]:20}» PTS nº{r['pts_num']} "
                  f"p{r['pts_page']},{r['pts_line']} | {str(r['reason'])[:66]}")
    if dry:
        print('\n(dry-run, nada escrito)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-an3.xlsx'
    shutil.copy(XLSX, bak)
    print('backup →', bak)
    for ridx, val, est, vr in writes:
        ws.cell(row=ridx, column=vcol).value = val
        ws.cell(row=ridx, column=ecol).value = est
        if vr:
            ws.cell(row=ridx, column=vri).value = vr
    wb.save(XLSX)
    print(f'Escritas {len(writes)} filas en {XLSX}.')


if __name__ == '__main__':
    main()
