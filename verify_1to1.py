#!/usr/bin/env python3
"""
1:1 Verification v3 — Fixed DN title detection (brackets in PTS text).
Page-by-page sequential matching for all Nikayas.
"""
import sqlite3, re
from openpyxl import load_workbook
from collections import defaultdict, Counter

DB = 'src/data/tipitaka.sqlite'
XL = 'PTS_Reference_Complete_Canon.xlsx'

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def norm(s):
    for a,b in [('ā','a'),('ī','i'),('ū','u'),('ṅ','n'),('ñ','n'),
                ('ṭ','t'),('ḍ','d'),('ṇ','n'),('ḷ','l'),('ṃ','m'),('ṁ','m')]:
        s = s.replace(a,b).replace(a.upper(),b.upper())
    return s

def get_page(book_no, page_no):
    cur.execute('SELECT head, unitext FROM pages WHERE book_no=? AND page_no=? AND edition="mula"', (book_no, page_no))
    r = cur.fetchone()
    if not r: return None, None
    return r['head'] or '', r['unitext'] or ''

wb = load_workbook(XL)
ws = wb['Complete Canon']
cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

entries = []
for ri in range(2, ws.max_row + 1):
    nik = str(ws.cell(row=ri, column=cols['Nikaya']).value or '')
    snum = str(ws.cell(row=ri, column=cols['Sutta #']).value or '')
    if not snum or snum == 'None': continue
    entries.append({
        'ri': ri, 'nik': nik, 'num': snum,
        'name': str(ws.cell(row=ri, column=cols['Sutta Name']).value or ''),
        'roman': str(ws.cell(row=ri, column=cols['PTS Roman']).value or '').strip().lower(),
        'vol': str(ws.cell(row=ri, column=cols['PTS Vol']).value or '').strip(),
        'page': ws.cell(row=ri, column=cols['PTS Page']).value,
        'ref': str(ws.cell(row=ri, column=cols['PTS Ref']).value or ''),
    })

def get_book_no(e):
    nik, roman, vol = e['nik'], e['roman'], e['vol']
    if nik == 'DN': return {'i':6,'ii':7,'iii':8}.get(roman)
    if nik == 'MN': return {'i':9,'ii':10,'iii':11}.get(roman)
    if nik == 'SN': return {'i':12,'ii':13,'iii':14,'iv':15,'v':16}.get(roman)
    if nik == 'AN': return {'i':17,'ii':18,'iii':19,'iv':20,'v':21}.get(roman)
    if nik == 'KN':
        m = {'Khp':22,'Kh':22,'Dhp':23,'Dh':23,'Ud':24,'It':25,'Sn':26,
             'Vv':27,'Pv':28,'Th':29,'Th & Th':29,'Thi':29,
             'Ja':30,'Ja I':30,'Ja II':31,'Ja III':32,'Ja IV':33,'Ja V':34,'Ja VI':35,
             'Nidd':36,'Nidd I':36,'Nidd II':37,
             'Patis I':38,'Patis II':39,'Ap':40,'Bv':41,'Cp':42}
        if vol in m: return m[vol]
        for k,v in m.items():
            if vol.startswith(k) or k.startswith(vol): return v
    return None

# ── Per-Nikaya marker finders ──

def find_markers_dn(text):
    """DN: Roman numeral titles like '[ii. Samanna-Phala Sutta.]' or 'xvi. Maha-Parinibbana'."""
    lines = text.split('\n')
    markers = []
    for i, line in enumerate(lines):
        s = line.strip()
        # Strip brackets common in PTS DN format
        s2 = s.lstrip('[').rstrip(']').strip()
        if re.match(r'^[ivxlc]+\.?\s+\S', s2.lower()):
            markers.append((i+1, 'title', s[:80]))
        elif re.match(r'^[ivxlc]+\.?\s+\S', s.lower()):
            markers.append((i+1, 'title', s[:80]))
        # Also catch "1. Evam me sutam" as content start
        elif re.search(r'[Ee]va[mM].*suta[mM]', s):
            markers.append((i+1, 'evam', s[:80]))
        # Catch "1.1. Evam me sutam" (some DN suttas)
        elif re.match(r'^\d+\.\d*\.?\s+[Ee]va[mM]', s):
            markers.append((i+1, 'evam_num', s[:80]))
    return markers

def find_markers_mn(text):
    lines = text.split('\n')
    markers = []
    for i, line in enumerate(lines):
        s = line.strip()
        if re.match(r'^\d+\.?\s*$', s) and len(s) <= 5:
            markers.append((i+1, 'num', s))
        elif re.match(r'^\d+\.\s+\S', s):
            markers.append((i+1, 'num_txt', s[:80]))
    return markers

def find_markers_sn(text):
    lines = text.split('\n')
    markers = []
    for i, line in enumerate(lines):
        s = line.strip()
        if re.match(r'^\d+\.?\s*\(\d+', s):
            markers.append((i+1, 'sn_id', s[:80]))
        elif re.match(r'^§\s+\d+', s):
            markers.append((i+1, 'section', s[:80]))
        elif re.match(r'^\d+\.\s+\S', s) and len(s) > 5:
            markers.append((i+1, 'num', s[:80]))
    return markers

def find_markers_an(text):
    lines = text.split('\n')
    markers = []
    for i, line in enumerate(lines):
        s = line.strip()
        if re.match(r'^[IVX]+\.?\s*$', s) and len(s) <= 8:
            markers.append((i+1, 'vagga', s))
        elif re.match(r'^\d+\.\s+\S', s):
            markers.append((i+1, 'num', s[:80]))
    return markers

def find_markers_kn(text, vol):
    lines = text.split('\n')
    markers = []
    if vol in ('Th', 'Th & Th', 'Thi'):
        for i, line in enumerate(lines):
            if re.search(r'[|]\s*\d+\s*[|]', line):
                markers.append((i+1, 'verse_end', line.strip()[:40]))
        return markers
    for i, line in enumerate(lines):
        s = line.strip()
        if re.match(r'^\d+\.\s+\S', s):
            markers.append((i+1, 'num', s[:80]))
        elif re.search(r'[Ee]va[mM].*suta[mM]', s):
            markers.append((i+1, 'evam', s[:80]))
    return markers

# ── Group entries by page ──
page_entries = defaultdict(list)
for e in entries:
    book_no = get_book_no(e)
    if not book_no: continue
    page_entries[(e['nik'], book_no, e['page'])].append(e)

print('1:1 Verification v3')
print('=' * 70)

results = []
stats = defaultdict(Counter)

for (nik, book_no, page), page_ents in sorted(page_entries.items()):
    head, text = get_page(book_no, page)
    if text is None:
        for e in page_ents:
            stats[nik]['PAGE_MISSING'] += 1
            results.append((e, 'PAGE_MISSING', None, None))
        continue
    
    if nik == 'DN': markers = find_markers_dn(text)
    elif nik == 'MN': markers = find_markers_mn(text)
    elif nik == 'SN': markers = find_markers_sn(text)
    elif nik == 'AN': markers = find_markers_an(text)
    elif nik == 'KN': markers = find_markers_kn(text, page_ents[0]['vol'] if page_ents else '')
    else: markers = []
    
    # Filter to content markers (exclude vagga-only for sequential matching)
    sutta_markers = [(l, t, c) for l, t, c in markers if t != 'vagga']
    
    for ei, e in enumerate(page_ents):
        if ei < len(sutta_markers):
            line, mtype, content = sutta_markers[ei]
            stats[nik][mtype] += 1
            results.append((e, mtype, page, line))
        else:
            # Fallback: keyword in page body
            kw = re.findall(r'[a-z]{5,}', norm(e['name'].lower()))
            body = norm(text).lower()
            hits = sum(1 for w in kw if w in body)
            if kw and hits >= max(1, len(kw)*0.4):
                stats[nik]['keyword'] += 1
                results.append((e, 'keyword', page, 1))
            else:
                stats[nik]['UNVERIFIED'] += 1
                results.append((e, 'UNVERIFIED', page, None))

# ── Report ──
def is_verified(status):
    return status not in ('UNVERIFIED', 'PAGE_MISSING', 'NO_BOOK')

for nik in ['DN', 'MN', 'SN', 'AN', 'KN']:
    s = stats[nik]
    total = sum(s.values())
    ver = sum(v for k,v in s.items() if is_verified(k))
    unv = s.get('UNVERIFIED', 0) + s.get('PAGE_MISSING', 0)
    print(f'\n{nik}: {total} entries')
    print(f'  Verified:   {ver:>5d} ({100*ver/max(total,1):.1f}%)')
    print(f'  Unverified: {unv:>5d}')
    methods = Counter({k:v for k,v in s.items() if is_verified(k)})
    for method, count in methods.most_common(4):
        print(f'    - {method}: {count}')

total_all = sum(sum(s.values()) for s in stats.values())
ver_all = sum(sum(v for k,v in s.items() if is_verified(k)) for s in stats.values())
unv_all = sum(s.get('UNVERIFIED',0)+s.get('PAGE_MISSING',0) for s in stats.values())
print(f'\n{"="*70}')
print(f'TOTAL: {total_all} | VERIFIED: {ver_all} ({100*ver_all/total_all:.1f}%) | UNVERIFIED: {unv_all}')
conn.close()
