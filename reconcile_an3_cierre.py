#!/usr/bin/env python3
"""
reconcile_an3_cierre — cierra lo cerrable de **A iii (Pañcaka + Chakka)**: 86 de las 88 pendientes.

A diferencia de A ii, A iv y A v, aquí **no se cierra el volumen entero**, y el motivo es del dato,
no de la herramienta: en el *sikkhāpada-peyyāla* PTS **no enumera todos los miembros que el CST sí
numera**. Su lista pasa de `sāmaṇerā` a `upāsikā` sin nombrar `upāsaka` ni `sāmaṇerī`, de modo que
para esas filas no hay término que buscar en el impreso. Se quedan PENDIENTE, con el motivo escrito.

### Cómo se decide qué se cierra

La regla es asimétrica **a propósito**, y conviene ver por qué:

- **el lado CST no necesita prueba léxica**: su ancla es el **paranum**, que es la clave canónica
  (regla (5)) y aquí coincide con el `Sutta #`;
- **el lado PTS sí**, porque es el que no tiene clave: el impreso no numera estos suttas. Se exige
  que el término distintivo aparezca **literalmente y en orden** en el texto y que caiga en la
  página que declara el Excel (±1).

Con eso, `validador_an_peyyala.analiza` sitúa **71 de las 76** filas del Pañcaka y **6 de 6** del
Chakka. El propio análisis se vuelve a correr aquí (no se copian listas de números a mano): lo que
se cierra es lo que la prueba sostiene en el momento de escribir.

### Las 2 que se quedan fuera, y por qué son irreductibles

| fila | por qué |
|---|---|
| `5.291` `Upāsaka` | **PTS no lo imprime**: su lista salta de `sāmaṇerā` a `upāsikā` |
| `5.300` `Āruddhaka` | el nombre **no aparece en ninguna página de A iii**: queda dentro del `…pe…` de la lista de sectarios |

Las dos son el mismo caso —PTS elide ese miembro— y **no hay prueba que darles por este camino**:
sin término en el impreso no hay nada que situar. Cerrarlas exigiría otra clase de evidencia (el
uddāna del vagga, o el cotejo contra el impreso en papel).

En la primera pasada quedaban tres más, y las tres eran **defectos míos o del Excel**, no del dato:

- `5.270`/`5.271` — el **sandhi de `apara-`**: `apara` + `anāgāmi` se escribe `aparānāgāmi`, así que
  quitar el prefijo dejaba `nāgāmi…`, que no existe en ningún texto. Hay que **restituir la vocal**.
  Y la serie monótona tiene que correrse sobre **todas** las filas del tramo, no sólo las
  pendientes: las ya cerradas son sus anclas, y sin ellas no hay de dónde agarrarse.
- `5.282` — no era divergencia de grafía entre ediciones sino un **error del Excel**: PTS escribe
  `sāṭiyaggāhāpako` (A iii 275) y el CST `sāṭiyagāhāpako`; sólo el Excel decía `Sāṭika-`.

⚠️ Y por el camino salió un defecto de fondo en `an_peyyala.cst_unidades`: el CST sólo pone el
atributo `n` en el **primer** `<p>` de cada sutta, y los de continuación van sin él. Descartarlos
tiraba el **74 % del texto** de `s0403m1` — justo la parte que enumera los miembros de los grupos
elididos. Corregido: se agregan a la unidad a la que pertenecen.

### Dos ajustes de dato

- **`5.1103-1152` → `VRI Ref` recortada a `1103-1151`**: el Excel numera la cola del Pañcaka hasta
  1152 y el CST **acaba en 1151**. Es el mismo desajuste que dejó fuera dos filas del Ekādasaka en
  A v, pero aquí sólo sobra el último paranum, así que la fila se ancla al rango real en vez de
  quedarse sin clave.
- Los nombres `Mada`/`Pamāda` ya se corrigieron aparte (`fix_kilesa_an3.py`).

### Fuera de los tramos peyyāla (6 filas)

`5.14-15`, `5.21-22`, `5.205-206` y `5.241-244` tienen **todos** sus miembros ya CONFIRMADO.
`5.187` rechazaba con cobertura **0.0** porque el CST imprime el sutta entero elidido
(`pañcime yathāsanthatikā…pe… sattamaṃ`): falso negativo puro. `6.82` es par correcto (A iii 433).

Uso: python3 reconcile_an3_cierre.py [--dry]
"""
import shutil
import sqlite3
import sys
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook

import validador_an_peyyala as VP

XLSX = 'PTS_Reference_Complete_Canon.xlsx'

FUERA = {          # filas del régimen numerado, con su motivo
    '5.14-15': 'rango: los dos miembros ya CONFIRMADO',
    '5.21-22': 'rango: los dos miembros ya CONFIRMADO',
    '5.205-206': 'rango: los dos miembros ya CONFIRMADO',
    '5.241-244': 'rango: los cuatro miembros ya CONFIRMADO',
    '5.187': 'el CST imprime el sutta entero elidido («pañcime yathāsanthatikā…pe… sattamaṃ»): '
             'cobertura 0.0 y REJECT por elisión, no por desalineación. Marcador 187, A iii 219',
    '6.82': 'par correcto (marcador 82, A iii 433); el REJECT venía del sutta gemelo',
}
VRI_CORREGIDA = {'5.1103-1152': 's0403m1:1103-1151'}
# Sólo dos filas quedan sin cerrar, y las dos por la misma razón: **PTS elide ese miembro de la
# lista**. Las otras tres de la primera pasada se recuperaron al arreglar tres defectos propios:
#   · `5.270`/`5.271` — el sandhi de `apara-`: `apara`+`anāgāmi` se escribe `aparānāgāmi`, así que
#     quitar el prefijo deja `nāgāmi…`, que no existe; hay que restituir la vocal (`an_peyyala`).
#     Y la serie tiene que correrse sobre TODAS las filas del tramo, no sólo las pendientes: sin
#     las ya cerradas como anclas, la monotonía no tiene de dónde agarrarse.
#   · `5.282` — no era grafía divergente entre ediciones sino un **error del Excel**: PTS escribe
#     `sāṭiyaggāhāpako` (A iii 275) y el CST `sāṭiyagāhāpako`; sólo el Excel decía `Sāṭika-`.
#     Corregido en `fix_kilesa_an3.py`.
NO_CERRAR = {
    '5.291': 'PTS NO imprime este miembro: su lista salta de «sāmaṇerā» a «upāsikā»',
    '5.300': 'el nombre «Āruddhaka» no aparece en NINGUNA página de A iii: PTS lo deja dentro del '
             '«…pe…» de la lista de sectarios',
}


def main():
    dry = '--dry' in sys.argv
    conn = sqlite3.connect(VP.DB)
    conn.row_factory = sqlite3.Row

    # se vuelve a correr el análisis: se cierra lo que la prueba sostiene AHORA, no una lista fija
    probadas = {}
    for region in VP.REGIONES:
        if region[0] != 'iii':
            continue
        clave = (region[0], region[1], region[3], region[4], region[5], region[6])
        # TODAS las filas del tramo, no sólo las pendientes: las ya cerradas son las anclas de la
        # serie monótona. Con sólo las pendientes, `5.270` y `5.271` se quedaban sin situar.
        filas = [f for f in VP.excel_pendientes(solo_pendientes=False)
                 if VP.region_de(f) == clave]
        if not filas:
            continue
        filas, _fallos, _T, _u = VP.analiza(conn, clave, filas)
        for f in filas:
            if not f['pendiente'] or f['num'] in NO_CERRAR:
                continue
            if f['pos_pts'] is not None and f['pagina_ok']:
                probadas[f['num']] = (f['vri'],
                                      f'peyyāla: «{f["sonda_pts"]}» literal y en orden en PTS '
                                      f'(A iii {f["locus"][0]},{f["locus"][1]}); ancla CST = paranum')
    print(f'prueba mecánica sostiene {len(probadas)} filas de los tramos peyyāla de A iii')

    if not dry:
        cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-an3cierre.xlsx'
        shutil.copy(XLSX, cp)
        print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    filas = [r for r in range(2, ws.max_row + 1)
             if ws.cell(r, ci['Nikaya']).value == 'AN'
             and str(ws.cell(r, ci['PTS Roman']).value or '').strip() == 'iii']
    hechas, pend_expl = Counter(), []
    for r in filas:
        num = str(ws.cell(r, ci['Sutta #']).value)
        if ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        if num in NO_CERRAR:
            ws.cell(r, ci['Detail']).value = NO_CERRAR[num]
            pend_expl.append(num)
            continue
        if num in probadas:
            vri, motivo = probadas[num]
        elif num in FUERA:
            lo, _, hi = num.split('.')[1].partition('-')
            stem = 's0403m1' if num.startswith('5.') else 's0403m2'
            vri, motivo = f'{stem}:{lo}' + (f'-{hi}' if hi else ''), FUERA[num]
        else:
            continue
        ws.cell(r, ci['VRI Ref']).value = VRI_CORREGIDA.get(num, vri)
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = motivo
        hechas['cerradas'] += 1
    pend = sum(1 for r in filas if ws.cell(r, ci['Estado']).value != 'CONFIRMADO')
    print(f"{hechas['cerradas']} filas cerradas · {len(pend_expl)} PENDIENTE con motivo: {pend_expl}")
    print(f'A iii: {len(filas) - pend}/{len(filas)} CONFIRMADO')
    if dry:
        print('(dry-run: nada escrito)')
        return
    wb.save(XLSX)
    print(f'escrito → {XLSX}')


if __name__ == '__main__':
    main()
