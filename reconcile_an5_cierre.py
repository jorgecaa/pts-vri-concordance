#!/usr/bin/env python3
"""
reconcile_an5_cierre — cierra **A v (Dasaka + Ekādasaka)**: las 30 filas que quedaban pendientes.

Mismo criterio que en A ii (`reconcile_an2_cierre.py`): donde las dos ediciones eliden no hay texto
que cotejar y el REJECT del LLM es un falso negativo, así que lo que firma cada fila es **prueba
mecánica o la lectura del impreso**. Cero llamadas a API. Cuatro vías, una por clase de fila:

**(1) Ruta LÉXICA — 11 filas del Ekādasaka (`11.22-1151`).** El `sāmañña-vagga` del Ekādasaka va
entero en **un solo bloque** de PTS (A v 359, el símil del gopālaka) y en **un solo `<p n="22-29">`**
del CST, y las dos ediciones enumeran las ocho *anupassanā* en el mismo orden —`aniccānupassī,
dukkhānupassī, anattānupassī, khayānupassī, vayānupassī, virāgānupassī, nirodhānupassī,
paṭinissaggānupassī`—. Las ocho se sitúan literalmente y en orden en los dos textos, con la página
del Excel. Las tres de rango (`11.30-501`, `11.502-981`, `11.982-1151`) se firman por **aritmética**:
el rango del Excel es la unión exacta de los `<p>` de rango del CST.

**(2) Ruta NUMÉRICA — 10 filas del Sāmañña-vagga del Dasaka (`10.221-236`).** Aquí no hay término
léxico que buscar: los nombres son **ordinales puros** (`Paṭhama`, `Dutiya`, `Chaṭṭhamādi`…). Pero
hay **dos claves independientes**: el `Sutta #` es el paranum del VRI y la `DPR Ref` es el nº
corrido de PTS (`10.221` → `AN 10.210` = el marcador `CCX.` de A v 303). Las diez cuadran, y el
marcador cae en la página que declara el Excel — **10/10**. Los sufijos `a`/`b` de la `DPR Ref`
describen además la asimetría: PTS numera 7 suttas donde el CST tiene 16, y `10.214a`/`10.214b`
dicen que su nº 214 cubre el `225` y el grupo `226-228`.

**(3) Marcador corregido — las 3 `REVISAR` (`10.36`, `10.37`, `10.38`).** El alineador las emparejó
con el marcador equivocado y por eso Gemini rechazó. Leído el impreso, el desplazamiento es claro:

| fila | CST | marcador que le puso el alineador | el que le toca |
|---|---|---|---|
| `10.36 Sāmaṇera` | `sāmaṇero upaṭṭhāpetabbo` | 35 (*Saṅghabheda*) | **34**, A v 73 |
| `10.37 Saṅghabheda` | `saṅghabhedo…` | 36 (*Saṅghasāmaggī*) | **35**, A v 73 |
| `10.38 Saṅghasāmaggi` | `saṅghasāmaggī…` | 38 | **36**, A v 74 |

**Las páginas del Excel ya eran correctas**; lo que estaba mal era el marcador, y de ahí el rechazo.

**(4) Filas de rango — 5.** `10.115-116`, `10.217-219` y `11.4-5` tienen **todos** sus miembros ya
CONFIRMADO individualmente. `10.156-166` casa exacto con el `subhead` de rango del CST
(`2-12. Bhajitabbādisuttāni`, `<p n="156-166">`) y sus once marcadores están en A v 248, la página
que declara el Excel. `10.199-210` casa con los `<p>` `199-210` del CST bajo
`Nasevitabbādisuttāni`, y su marcador `CXCIX` abre en A v 281, también la del Excel.

### La cola del Ekādasaka: por qué el Excel cuenta 1151 y el CST 671

En la primera pasada firmé `11.502-981` y `11.982-1151` con su `VRI Ref` aritmética y
`check_integrity` saltó: **el paranum 982 no existe en `s0404m4`**. Mirando el impreso se entiende
la aritmética entera, y no es un error de nadie:

- **PTS imprime el símil del gopālaka DOS veces**: `abhabbo` (incapaz, A v 359) y `bhabbo` (capaz,
  A v 360 §§4-6), cada uno con los mismos 60 objetos × 8 *anupassanā*.
- **El CST sólo numera la mitad negativa**: su `<p n="22-29">` dice `abhabbo` diez veces y `bhabbo`
  ninguna, y sus paranums del *sāmañña-vagga* son `22-501` = **480**.
- El Excel cuenta `22-981` = **960** = 480 × 2: numera **las dos mitades**.

De ahí que a partir de ahí vaya +480, y las cuentas cuadran al dígito:

| fila del Excel | unidades | contraparte del CST | unidades |
|---|--:|---|--:|
| `11.982-1151` (*Rāgādi peyyāla*) | **170** | `502` + `503-511` + `512-671` | **170** ✓ |
| `11.502-981` (*Aniccānupassanādi*) | 480 | — *el CST no numera la mitad positiva* | — |

Así que **`11.982-1151` se cierra** con la `VRI Ref` corregida a `s0404m4:502-671` (mismo recuento,
mismo contenido, misma página). Y **`11.502-981` se queda PENDIENTE**, pero ya no por
desconocimiento: es una fila que PTS imprime (A v 360 §§4-6) y que **el CST no numera**. Anclarla a
`22-501` sería darle la clave de su mitad gemela; borrarla sería negar lo que el impreso tiene
delante. Es decisión editorial de Jorge, del mismo tipo que la de `12.74` en S ii.

A v queda en **244/245**.

Uso: python3 reconcile_an5_cierre.py [--dry]
"""
import shutil
import sys
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
STEM = {'10': 's0404m3', '11': 's0404m4'}

CIERRE = {
    # (1) ruta léxica — Ekādasaka
    **{f'11.{n}': 'peyyāla: la anupassanā aparece literal y en orden en PTS (A v 359) y en el CST'
       for n in range(22, 30)},
    '11.30-501': 'rango = unión exacta de los `<p>` del CST (30-69 … 454-501); locus en A v 359',
    # `11.982-1151` SÍ se cierra, con la `VRI Ref` corregida: son los **170** paranums del
    # rāgapeyyāla del CST (`502` + `503-511` + `512-671`), exactamente los 170 que cuenta la fila.
    '11.982-1151': 'rāgapeyyāla = `502`+`503-511`+`512-671` del CST: 170 paranums, los mismos 170 '
                   'que cuenta la fila; texto en A v 360-361 tras la raya',
    # ⚠️ `11.502-981` NO se cierra: el CST no numera la mitad positiva. Ver la nota de abajo.
    # (2) ruta numérica — Sāmañña-vagga del Dasaka
    '10.221': 'ruta numérica: paranum 221 + DPR 10.210 = marcador CCX en A v 303',
    '10.222': 'ruta numérica: paranum 222 + DPR 10.211 = marcador 211 en A v 304',
    '10.223': 'ruta numérica: paranum 223 + DPR 10.212 = marcador 212 en A v 305',
    '10.224': 'ruta numérica: paranum 224 + DPR 10.213 = marcador 213 en A v 306',
    '10.225': 'ruta numérica: paranum 225 + DPR 10.214a = marcador 214 en A v 308',
    '10.226-228': 'ruta numérica: paranums 226-228 + DPR 10.214b = marcador 214 en A v 308',
    '10.229': 'ruta numérica: paranum 229 + DPR 10.215a = marcador 215 en A v 308',
    '10.230-232': 'ruta numérica: paranums 230-232 + DPR 10.215b = marcador 215 en A v 308',
    '10.233': 'ruta numérica: paranum 233 + DPR 10.216a = marcador 216 en A v 309',
    '10.234-236': 'ruta numérica: paranums 234-236 + DPR 10.216b = marcador 216 en A v 309',
    '10.237-746': 'rāgādi peyyāla; término literal en las dos ediciones, A v 309',
    # (3) marcador corregido
    '10.36': 'marcador 34 (A v 73), no el 35: el 35 es el Saṅghabheda. El REJECT venía del par malo',
    '10.37': 'marcador 35 (A v 73), no el 36',
    '10.38': 'marcador 36 (A v 74), no el 38',
    # (4) rangos
    '10.115-116': 'rango: los dos miembros ya CONFIRMADO',
    '10.156-166': 'rango = `subhead` «2-12. Bhajitabbādisuttāni», `<p n="156-166">`; 11 marcadores en A v 248',
    '10.199-210': 'rango = `<p>` 199-210 del CST («Nasevitabbādisuttāni»); marcador CXCIX en A v 281',
    '10.217-219': 'rango: los tres miembros ya CONFIRMADO',
    '11.4-5': 'rango: los dos miembros ya CONFIRMADO',
}


def main():
    dry = '--dry' in sys.argv
    if not dry:
        cp = f'{XLSX}.backup-{datetime.now():%Y%m%d-%H%M%S}-an5cierre.xlsx'
        shutil.copy(XLSX, cp)
        print(f'backup → {cp}')
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    ci = {c.value: i + 1 for i, c in enumerate(ws[1])}
    filas = [r for r in range(2, ws.max_row + 1)
             if ws.cell(r, ci['Nikaya']).value == 'AN'
             and str(ws.cell(r, ci['PTS Roman']).value or '').strip() == 'v']
    hechas, faltan = Counter(), []
    for r in filas:
        num = str(ws.cell(r, ci['Sutta #']).value)
        if ws.cell(r, ci['Estado']).value == 'CONFIRMADO':
            continue
        motivo = CIERRE.get(num)
        if not motivo:
            faltan.append(num)
            continue
        nip = num.split('.')[0]
        lo, _, hi = num.split('.')[1].partition('-')
        vri = f'{STEM[nip]}:{lo}' + (f'-{hi}' if hi else '')
        # el rāgapeyyāla del Ekādasaka lleva OTRA numeración en el Excel: 982-1151 son los
        # paranums 502-671 del CST (mismo recuento, 170)
        ws.cell(r, ci['VRI Ref']).value = ('s0404m4:502-671' if num == '11.982-1151' else vri)
        ws.cell(r, ci['Validation']).value = 'VALIDADOR_HUMANO'
        ws.cell(r, ci['Estado']).value = 'CONFIRMADO'
        ws.cell(r, ci['Detail']).value = motivo
        hechas['cerradas'] += 1
    pend = sum(1 for r in filas if ws.cell(r, ci['Estado']).value != 'CONFIRMADO')
    print(f"{hechas['cerradas']} filas cerradas · A v: {len(filas) - pend}/{len(filas)} CONFIRMADO")
    if faltan:
        print(f'  sin regla: {faltan}')
    if dry:
        print('(dry-run: nada escrito)')
        return
    wb.save(XLSX)
    print(f'escrito → {XLSX}')


if __name__ == '__main__':
    main()
