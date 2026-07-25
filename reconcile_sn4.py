#!/usr/bin/env python3
"""reconcile_sn4.py — vuelca `validador_sn4.json` al Excel maestro (SN IV, S iv).

Regla del proyecto: CONFIRMADO = concordancia (VRI) ∧ Gemini APPROVE → `Validation=VALIDADOR`.
Gemini REJECT → NO se auto-firma: queda para arbitraje humano (`--humano firmas.json`).

En SN IV el lado CST **no se resuelve por `Sutta #`** —en SN 35 esa clave es la cuenta reducida de
Feer (207) y no la del CST— sino por **posición**: los bloques de subhead del XML VRI y las filas
del Excel se corresponden 1:1 en los 10 saṃyuttas (344=344), con 344/344 de acuerdo por nombre. El
lado PTS lo fija `assign_volume` (inyectivo y monótono, hueco 0) sobre los marcadores de la BD,
con las capacidades medidas en el texto. Detalle en `STATUS.md` § Fase 3.

Los 94 `HELMER_APPROVED` y los 12 `HELMER_REJECT` que traía el volumen son del pase DeepSeek
retirado: **no se heredan**, se sobrescriben con el veredicto del validador.

Uso:
    python3 reconcile_sn4.py --dry
    python3 reconcile_sn4.py --write [--humano firmas.json]
"""
import json, re, sys, shutil, datetime
from collections import Counter
from openpyxl import load_workbook

XLSX = 'PTS_Reference_Complete_Canon.xlsx'
RES = 'validador_sn4.json'


def sn4_rows(ws, ci):
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = [c.value for c in row]
        if str(v[ci['Nikaya']] or '') != 'SN' or str(v[ci['PTS Roman']] or '').strip().lower() != 'iv':
            continue
        yield ridx, str(v[ci['Sutta #']]), v


def main():
    dry = '--write' not in sys.argv
    humano = {}
    if '--humano' in sys.argv:
        humano = json.load(open(sys.argv[sys.argv.index('--humano') + 1]))
    res = {r['num']: r for r in json.load(open(RES))}
    wb = load_workbook(XLSX)
    ws = wb['Complete Canon']
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}
    vcol, ecol = ci['Validation'] + 1, ci['Estado'] + 1

    plan, writes, pend = Counter(), [], []
    for ridx, num, v in sn4_rows(ws, ci):
        if num in humano:
            val = humano[num]
            est = 'PENDIENTE' if val in ('REVISAR', 'REF_ERROR_DPR') else 'CONFIRMADO'
            plan[f'humano:{val}'] += 1
            writes.append((ridx, val, est))
            continue
        cur_val = str(ws.cell(row=ridx, column=vcol).value or '')
        if cur_val in ('VALIDADOR_HUMANO', 'PTS_CROSSREF_SN'):
            plan['protegido(humano/crossref)'] += 1
            continue
        r = res.get(num)
        if not r:
            plan['sin_resultado'] += 1
            continue
        if r['gemini'] == 'APPROVE':
            plan['VALIDADOR(auto)'] += 1
            writes.append((ridx, 'VALIDADOR', 'CONFIRMADO'))
        else:
            # El REJECT se ESCRIBE como PENDIENTE/REVISAR: si sólo se listara, las filas con
            # `HELMER_APPROVED` del pase DeepSeek retirado seguirían figurando como CONFIRMADO
            # justo donde el validador vigente discrepa — que es la herencia que hay que evitar.
            plan['arbitraje(gemini REJECT)'] += 1
            writes.append((ridx, 'REVISAR', 'PENDIENTE'))
            pend.append(r)

    print('Plan de escritura:', dict(plan))
    print('Filas a escribir:', len(writes))
    if pend:
        print(f'\nA arbitraje ({len(pend)}) — alineación segura, discrepa el cotejo de texto:')
        for r in pend:
            print(f'  SN {r["num"]:>7} PTS "{(r.get("pts_name") or "—")[:20]:20}" vs CST '
                  f'"{r["cst_title"][:26]:26}" p{r.get("pts_page")} | {str(r["reason"])[:64]}')
    if dry:
        print('\n(dry-run, nada escrito)')
        return
    bak = f'{XLSX}.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}-sn4.xlsx'
    shutil.copy(XLSX, bak)
    print('backup →', bak)
    for ridx, val, est in writes:
        ws.cell(row=ridx, column=vcol).value = val
        ws.cell(row=ridx, column=ecol).value = est
    wb.save(XLSX)
    print(f'Escritas {len(writes)} filas en {XLSX}.')


if __name__ == '__main__':
    main()
